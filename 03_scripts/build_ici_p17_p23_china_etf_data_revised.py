from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT_PATH = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\原始数据\全市场ETF基础信息.csv.xlsx")
OUTPUT_DIR = Path(r"C:\Users\lvdon\Desktop\Fullgoal\ICI课题\P17-P23_中国ETF数据处理_修正版")

OUTPUT_FILES = {
    "00": "00_数据字段检查与描述性统计_修正版.xlsx",
    "P17": "P17_中国ETF年度新发与累计数量_修正版.xlsx",
    "P18": "P18_中国ETF资产类型与产品类型结构_修正版.xlsx",
    "P19": "P19_中国ETF管理人集中度与头部排名_修正版.xlsx",
    "P20": "P20_中国ETF规模区间分布与长尾分析_修正版.xlsx",
    "P21": "P21_中美对比所需中国ETF总览指标_修正版.xlsx",
    "P22P23": "P22-P23_结论页支撑数据摘要_修正版.xlsx",
    "ARCHIVE": "ICI课题_P17-P23_中国ETF作图数据汇总_修正版.xlsx",
    "README": "ICI课题_P17-P23_数据处理说明_修正版.md",
}

RAW_TO_STD = {
    "Wind代码": "Wind代码",
    "证券简称": "证券简称",
    "基金简称": "基金简称",
    "基金全称": "基金全称",
    "基金代码": "基金代码",
    "交易代码": "交易代码",
    "基金类型": "基金类型",
    "投资类型_二级分类": "投资类型",
    "业绩比较基准": "业绩比较基准",
    "最新基金规模(亿)": "基金规模_亿元",
    "基金管理人": "基金管理人",
    "ETF一级市场基金代码": "一级市场基金代码",
    "跟踪指数名称": "跟踪指数",
    "跟踪指数代码": "跟踪指数代码",
    "基金上市地点": "基金上市地点",
    "上市日期": "上市日期",
    "基金成立日": "成立日期",
    "基金托管人": "基金托管人",
    "管理费率_支持历史(%)": "管理费率",
    "托管费率_支持历史(%)": "托管费率",
}

BASE_COLUMNS = list(RAW_TO_STD.values())
CLASS_COLUMNS = [
    "资产类型",
    "股票ETF内部类型",
    "产品类型_汇总",
    "是否跨境ETF",
    "是否债券ETF",
    "是否商品ETF",
    "是否货币ETF",
    "是否股票ETF",
    "是否宽基ETF",
    "是否行业主题ETF",
    "是否策略ETF",
    "命中关键词",
    "分类备注",
    "入池状态",
    "入池/观察原因",
]

EXCLUDE_TERMS = ["ETF联接", "ETF连接", "联接A", "联接C", "联接", "场外联接", "FOF", "LOF"]
MONEY_TERMS = ["货币", "现金", "保证金", "快线", "添益", "场内货币"]
COMMODITY_TERMS = ["黄金", "豆粕", "商品", "能源", "原油", "白银", "油气"]
BOND_TERMS = ["国债", "政金债", "信用债", "公司债", "地方债", "城投债", "可转债", "短融", "债券", "债"]
CROSS_TERMS = [
    "恒生",
    "港股",
    "港股通",
    "H股",
    "纳斯达克",
    "纳指",
    "标普",
    "道琼斯",
    "日经",
    "德国",
    "法国",
    "海外",
    "中概",
    "东南亚",
    "QDII",
    "MSCI美国",
    "美国",
    "香港",
    "日韩",
    "越南",
    "沙特",
    "亚太",
    "全球",
]
BROAD_TERMS = [
    "沪深300",
    "中证500",
    "中证800",
    "中证1000",
    "中证2000",
    "中证A500",
    "中证A100",
    "中证A50",
    "中证A股",
    "中证全指",
    "上证50",
    "上证180",
    "上证380",
    "上证综指",
    "上证指数",
    "深证100",
    "深证50",
    "创业板指",
    "创业板50",
    "创业板综",
    "科创50",
    "科创100",
    "北证50",
    "MSCI中国A股",
    "MSCI中国A50互联互通",
    "MSCI A50",
    "富时中国A50",
    "国证A指",
    "万得全A",
    "中小板",
    "中小100",
]
INDUSTRY_ALL_SHARE_EXCLUDE = [
    "中证全指证券公司",
    "中证全指医药卫生",
    "中证全指医药",
    "中证全指半导体",
    "中证全指软件",
    "中证全指食品",
    "中证全指电力",
    "中证全指通信",
    "中证全指房地产",
    "中证全指银行",
    "中证全指建筑材料",
]
STRATEGY_TERMS = [
    "红利",
    "低波",
    "低波动",
    "红利低波",
    "质量",
    "价值",
    "成长",
    "现金流",
    "自由现金流",
    "股息",
    "高股息",
    "基本面",
    "等权",
    "等权重",
    "ESG",
    "动量",
    "回购",
    "分红",
    "龙头",
    "Smart Beta",
    "增强策略",
    "指数增强",
    "央企红利",
    "国企红利",
]
INDUSTRY_TERMS = [
    "新能源",
    "光伏",
    "芯片",
    "半导体",
    "人工智能",
    "机器人",
    "军工",
    "医药",
    "消费",
    "金融",
    "证券",
    "银行",
    "地产",
    "传媒",
    "计算机",
    "通信",
    "汽车",
    "电池",
    "储能",
    "双碳",
    "绿色",
    "电力",
    "煤炭",
    "有色",
    "钢铁",
    "基建",
    "农业",
    "酒",
    "食品",
    "云计算",
    "软件",
    "游戏",
    "数字经济",
    "高端制造",
    "创新药",
    "医疗器械",
    "工业母机",
    "工程机械",
    "畜牧养殖",
    "稀有金属",
    "工业有色",
    "化工",
    "材料",
    "环保",
    "教育",
    "物流",
    "旅游",
    "黄金股",
    "矿业",
    "交运",
]
CORE_BROAD_CHECK_TERMS = ["A500", "沪深300", "中证500", "中证1000", "创业板指", "科创50", "上证50"]
INDUSTRY_ALL_SHARE_CHECK_TERMS = ["全指证券", "全指医药", "全指食品", "全指软件", "全指电力", "全指银行", "全指通信"]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "nat", "none"}:
        return ""
    return re.sub(r"\s+", " ", text)


def code_text(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text.zfill(6) if text.isdigit() and len(text) < 6 else text


def number_value(value: Any) -> float | None:
    text = clean_text(value).replace(",", "").replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def date_text(value: Any) -> str:
    if clean_text(value) == "":
        return ""
    dt = pd.to_datetime(value, errors="coerce")
    return "" if pd.isna(dt) else dt.strftime("%Y-%m-%d")


def contains_terms(text: str, terms: list[str]) -> list[str]:
    upper = text.upper()
    return [term for term in terms if term.upper() in upper]


def pct(value: float) -> str:
    return f"{value:.2%}"


def add_pct_display(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    pct_cols = [col for col in out.columns if "占比" in col or col in {"数量口径", "规模口径"}]
    for col in pct_cols:
        insert_at = out.columns.get_loc(col) + 1
        display_col = f"{col}_展示"
        out.insert(insert_at, display_col, out[col].map(lambda x: "" if pd.isna(x) else pct(float(x))))
    return out


def round4(value: Any) -> Any:
    if pd.isna(value):
        return ""
    if isinstance(value, (int, float)):
        return round(float(value), 4)
    return value


def manager_short_name(name: str) -> str:
    text = clean_text(name)
    for old in ["基金管理股份有限公司", "基金管理有限公司", "管理股份有限公司", "管理有限公司"]:
        text = text.replace(old, "")
    text = text.replace("(中国)", "")
    return text


def read_and_standardize() -> tuple[dict[str, pd.DataFrame], pd.DataFrame, pd.DataFrame, int]:
    raw_sheets = pd.read_excel(INPUT_PATH, sheet_name=None, dtype=object)
    raw = raw_sheets["Sheet1"].copy()
    raw_rows = len(raw)

    footer_mask = raw["Wind代码"].map(clean_text).str.startswith("数据来源")
    raw_no_footer = raw.loc[~footer_mask].copy()

    std = pd.DataFrame()
    for raw_col, std_col in RAW_TO_STD.items():
        if raw_col in raw_no_footer.columns:
            if std_col in {"Wind代码", "基金代码", "交易代码", "一级市场基金代码"}:
                std[std_col] = raw_no_footer[raw_col].map(code_text)
            elif std_col in {"上市日期", "成立日期"}:
                std[std_col] = raw_no_footer[raw_col].map(date_text)
            elif std_col in {"基金规模_亿元", "管理费率", "托管费率"}:
                std[std_col] = raw_no_footer[raw_col].map(number_value)
            else:
                std[std_col] = raw_no_footer[raw_col].map(clean_text)
        else:
            std[std_col] = ""
    std = std.reset_index(drop=True)

    field_mapping = pd.DataFrame(
        [
            {
                "标准字段名": std_col,
                "原始字段名": raw_col,
                "所在sheet": "Sheet1",
                "识别方式": "字段名精确映射",
                "备注": "原字段已为亿元口径，直接转数值" if std_col == "基金规模_亿元" else "",
            }
            for raw_col, std_col in RAW_TO_STD.items()
        ]
    )
    return raw_sheets, raw_no_footer, std, raw_rows


def pool_masks(std: pd.DataFrame) -> tuple[pd.Series, pd.Series, pd.DataFrame]:
    wind = std["Wind代码"].astype(str).str.upper()
    suffix = wind.str.extract(r"(\.[A-Z]+)$", expand=False).fillna("")
    name_text = (std["基金简称"].fillna("") + " " + std["基金全称"].fillna("")).astype(str)

    listed_suffix = suffix.isin([".SH", ".SZ"])
    has_exchange = std["基金上市地点"].map(clean_text).ne("")
    has_listing_date = std["上市日期"].map(clean_text).ne("")
    has_start_date = std["成立日期"].map(clean_text).ne("")
    has_scale = pd.to_numeric(std["基金规模_亿元"], errors="coerce").gt(0)
    name_has_etf = name_text.str.contains("ETF", case=False, na=False)
    exclude_hit = name_text.map(lambda x: bool(contains_terms(x, EXCLUDE_TERMS)))

    main_mask = listed_suffix & has_exchange & has_listing_date & has_start_date & has_scale & name_has_etf & ~exclude_hit
    key_missing = ~(has_exchange & has_listing_date & has_start_date & has_scale)
    observation_mask = (~main_mask) & (suffix.eq(".OF") | key_missing | name_has_etf)

    steps = [
        ("原始行数", len(std), "已剔除footer后的产品记录"),
        ("Wind代码后缀为.SH或.SZ", int(listed_suffix.sum()), "主分析池必须为场内上市交易代码"),
        ("基金上市地点非空", int((listed_suffix & has_exchange).sum()), "在.SH/.SZ基础上继续筛选"),
        ("上市日期非空", int((listed_suffix & has_exchange & has_listing_date).sum()), "在前一步基础上继续筛选"),
        ("成立日期非空", int((listed_suffix & has_exchange & has_listing_date & has_start_date).sum()), "在前一步基础上继续筛选"),
        ("规模非空且大于0", int((listed_suffix & has_exchange & has_listing_date & has_start_date & has_scale).sum()), "在前一步基础上继续筛选"),
        ("简称或全称包含ETF", int((listed_suffix & has_exchange & has_listing_date & has_start_date & has_scale & name_has_etf).sum()), "严格执行本次主池名称口径"),
        ("排除联接/FOF/LOF等", int(main_mask.sum()), "最终上市ETF分析池"),
    ]
    screening_process = pd.DataFrame(steps, columns=["筛选步骤", "剩余/命中数量", "说明"])

    reasons = []
    for idx in std.index:
        item = []
        if suffix.loc[idx] == ".OF":
            item.append("Wind代码后缀为.OF")
        if suffix.loc[idx] not in [".SH", ".SZ", ".OF"]:
            item.append("Wind代码后缀非.SH/.SZ/.OF")
        if not has_exchange.loc[idx]:
            item.append("基金上市地点缺失")
        if not has_listing_date.loc[idx]:
            item.append("上市日期缺失")
        if not has_start_date.loc[idx]:
            item.append("成立日期缺失")
        if not has_scale.loc[idx]:
            item.append("规模缺失或小于等于0")
        if not name_has_etf.loc[idx]:
            item.append("基金简称/全称未包含ETF")
        if exclude_hit.loc[idx]:
            item.append("命中联接/FOF/LOF等排除词")
        reasons.append("；".join(item) if item else "满足主分析池口径")
    std["入池状态"] = ["上市ETF分析池" if main_mask.loc[i] else ("观察池" if observation_mask.loc[i] else "非本次口径") for i in std.index]
    std["入池/观察原因"] = reasons
    return main_mask, observation_mask, screening_process


def classify_row(row: pd.Series) -> pd.Series:
    text = " ".join(
        clean_text(row.get(col))
        for col in ["证券简称", "基金简称", "基金全称", "跟踪指数", "业绩比较基准", "投资类型"]
    )
    investment_type = clean_text(row.get("投资类型"))
    hits: list[str] = []
    notes: list[str] = []

    money_hits = contains_terms(text, MONEY_TERMS)
    commodity_hits = contains_terms(text, COMMODITY_TERMS)
    bond_hits = contains_terms(text, BOND_TERMS)
    cross_hits = contains_terms(text, CROSS_TERMS)

    if "货币市场型基金" in investment_type or money_hits:
        asset_type = "货币ETF"
        hits += money_hits or ["投资类型=货币市场型基金"]
    elif "商品型基金" in investment_type or commodity_hits:
        asset_type = "商品ETF"
        hits += commodity_hits or ["投资类型=商品型基金"]
    elif "被动指数型债券基金" in investment_type or bond_hits:
        asset_type = "债券ETF"
        hits += bond_hits or ["投资类型=被动指数型债券基金"]
    elif "国际(QDII)股票型基金" in investment_type or cross_hits:
        asset_type = "跨境ETF"
        hits += cross_hits or ["投资类型=国际(QDII)股票型基金"]
    elif investment_type in {"被动指数型基金", "增强指数型基金"}:
        asset_type = "股票ETF"
    else:
        asset_type = "其他ETF"
        notes.append("投资类型和名称关键词不足，无法归入主要资产类型")

    strategy_hits = contains_terms(text, STRATEGY_TERMS)
    industry_exclude_hits = contains_terms(text, INDUSTRY_ALL_SHARE_EXCLUDE)
    broad_hits = [] if industry_exclude_hits else contains_terms(text, BROAD_TERMS)
    industry_hits = contains_terms(text, INDUSTRY_TERMS) + industry_exclude_hits
    is_enhanced = "增强指数型基金" in investment_type or "增强" in text
    is_strategy = bool(strategy_hits) or is_enhanced

    if asset_type == "股票ETF":
        if broad_hits:
            stock_inner = "宽基ETF"
            hits += broad_hits
            if is_strategy:
                notes.append("宽基增强/增强策略，内部类型按宽基ETF处理")
        elif strategy_hits or is_enhanced:
            stock_inner = "策略ETF"
            hits += strategy_hits or ["增强指数型基金"]
        elif industry_hits:
            stock_inner = "行业主题ETF"
            hits += industry_hits
        else:
            stock_inner = "其他股票ETF"
            notes.append("股票ETF但未命中宽基/策略/行业主题关键词")
    else:
        stock_inner = "不适用"
        if industry_exclude_hits:
            notes.append("非股票ETF，行业全指排除词不参与宽基分类")

    if asset_type == "股票ETF":
        product_summary = stock_inner
    elif asset_type in {"跨境ETF", "债券ETF", "商品ETF", "货币ETF"}:
        product_summary = asset_type
    else:
        product_summary = "其他ETF"

    flags = {
        "是否跨境ETF": asset_type == "跨境ETF",
        "是否债券ETF": asset_type == "债券ETF",
        "是否商品ETF": asset_type == "商品ETF",
        "是否货币ETF": asset_type == "货币ETF",
        "是否股票ETF": asset_type == "股票ETF",
        "是否宽基ETF": stock_inner == "宽基ETF",
        "是否行业主题ETF": stock_inner == "行业主题ETF",
        "是否策略ETF": is_strategy if asset_type == "股票ETF" else False,
    }
    return pd.Series(
        {
            "资产类型": asset_type,
            "股票ETF内部类型": stock_inner,
            "产品类型_汇总": product_summary,
            **{key: "是" if value else "否" for key, value in flags.items()},
            "命中关键词": "、".join(dict.fromkeys(hits)),
            "分类备注": "；".join(dict.fromkeys(notes)),
        }
    )


def finalize_pools(std: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    main_mask, observation_mask, screening_process = pool_masks(std)
    class_df = std.apply(classify_row, axis=1)
    full = pd.concat([std, class_df], axis=1)
    main = full.loc[main_mask, BASE_COLUMNS + CLASS_COLUMNS].copy().reset_index(drop=True)
    observation = full.loc[observation_mask, BASE_COLUMNS + CLASS_COLUMNS].copy().reset_index(drop=True)
    return main, observation, screening_process


def group_structure(df: pd.DataFrame, by_col: str) -> pd.DataFrame:
    total_count = len(df)
    total_scale = pd.to_numeric(df["基金规模_亿元"], errors="coerce").sum()
    rows = []
    for key, part in df.groupby(by_col, dropna=False):
        scale = pd.to_numeric(part["基金规模_亿元"], errors="coerce")
        rows.append(
            {
                by_col: key if clean_text(key) else "未分类",
                "ETF数量": len(part),
                "数量占比": len(part) / total_count if total_count else 0,
                "总规模_亿元": round4(scale.sum()),
                "规模占比": scale.sum() / total_scale if total_scale else 0,
                "平均规模_亿元": round4(scale.mean()),
                "规模中位数_亿元": round4(scale.median()),
            }
        )
    out = pd.DataFrame(rows).sort_values(["总规模_亿元", "ETF数量"], ascending=[False, False]).reset_index(drop=True)
    return add_pct_display(out)


def annual_table(df: pd.DataFrame, by_col: str | None = None) -> pd.DataFrame:
    temp = df.copy()
    temp["年份"] = pd.to_datetime(temp["成立日期"], errors="coerce").dt.year
    temp = temp.dropna(subset=["年份"]).copy()
    temp["年份"] = temp["年份"].astype(int)
    years = list(range(int(temp["年份"].min()), int(temp["年份"].max()) + 1)) if len(temp) else []
    if by_col is None:
        out = temp.groupby("年份").size().reindex(years, fill_value=0).reset_index(name="当年新成立ETF数量")
        out["累计ETF数量"] = out["当年新成立ETF数量"].cumsum()
        out["年度标记"] = out["年份"].map(lambda y: "YTD" if y == 2026 else "")
        return out
    keys = sorted(temp[by_col].dropna().unique().tolist())
    frame = pd.MultiIndex.from_product([years, keys], names=["年份", by_col]).to_frame(index=False)
    counts = temp.groupby(["年份", by_col]).size().reset_index(name="当年新成立ETF数量")
    out = frame.merge(counts, on=["年份", by_col], how="left").fillna({"当年新成立ETF数量": 0})
    out["当年新成立ETF数量"] = out["当年新成立ETF数量"].astype(int)
    out["累计ETF数量"] = out.groupby(by_col)["当年新成立ETF数量"].cumsum()
    return out


def manager_tables(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    total_count = len(df)
    total_scale = df["基金规模_亿元"].sum()
    rows = []
    for manager, part in df.groupby("基金管理人", dropna=False):
        scale = pd.to_numeric(part["基金规模_亿元"], errors="coerce")
        rows.append(
            {
                "基金管理人": manager,
                "管理人简称": manager_short_name(manager),
                "ETF数量": len(part),
                "ETF总规模_亿元": round4(scale.sum()),
                "规模占比": scale.sum() / total_scale if total_scale else 0,
                "平均单品规模_亿元": round4(scale.mean()),
                "规模中位数_亿元": round4(scale.median()),
                "最大单品规模_亿元": round4(scale.max()),
                "10亿以下产品数量": int((scale < 10).sum()),
                "10亿以下产品占比": int((scale < 10).sum()) / len(part) if len(part) else 0,
            }
        )
    ranking = pd.DataFrame(rows).sort_values(["ETF总规模_亿元", "ETF数量"], ascending=[False, False]).reset_index(drop=True)
    ranking.insert(0, "排名", range(1, len(ranking) + 1))
    ranking = add_pct_display(ranking)
    cr_rows = []
    for n in [3, 5, 10, 20]:
        top = ranking.head(n)
        cr_rows.append(
            {
                "指标": f"CR{n}",
                "数量口径": top["ETF数量"].sum() / total_count if total_count else 0,
                "规模口径": top["ETF总规模_亿元"].sum() / total_scale if total_scale else 0,
                "说明": f"按ETF总规模排名的前{n}大管理人合计占比",
            }
        )
    return ranking, add_pct_display(pd.DataFrame(cr_rows))


def size_tables(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    temp = df.copy()
    labels = ["<1亿", "1–2亿", "2–10亿", "10–50亿", "50–200亿", "200亿以上"]
    temp["规模区间"] = pd.cut(temp["基金规模_亿元"], [-math.inf, 1, 2, 10, 50, 200, math.inf], labels=labels, right=False)
    valid = temp.dropna(subset=["规模区间"]).copy()
    total_count = len(valid)
    total_scale = valid["基金规模_亿元"].sum()

    full_rows = []
    for label in labels:
        part = valid[valid["规模区间"].astype(str) == label]
        full_rows.append(
            {
                "规模区间": label,
                "ETF数量": len(part),
                "数量占比": len(part) / total_count if total_count else 0,
                "总规模_亿元": round4(part["基金规模_亿元"].sum()),
                "规模占比": part["基金规模_亿元"].sum() / total_scale if total_scale else 0,
            }
        )
    full = add_pct_display(pd.DataFrame(full_rows))

    by_type_rows = []
    for product_type, part0 in valid.groupby("产品类型_汇总"):
        type_count = len(part0)
        type_scale = part0["基金规模_亿元"].sum()
        for label in labels:
            part = part0[part0["规模区间"].astype(str) == label]
            by_type_rows.append(
                {
                    "产品类型_汇总": product_type,
                    "规模区间": label,
                    "ETF数量": len(part),
                    "数量占比": len(part) / type_count if type_count else 0,
                    "总规模_亿元": round4(part["基金规模_亿元"].sum()),
                    "规模占比": part["基金规模_亿元"].sum() / type_scale if type_scale else 0,
                }
            )
    by_type = add_pct_display(pd.DataFrame(by_type_rows))

    manager_rows = []
    for manager, part in temp.groupby("基金管理人", dropna=False):
        scale = part["基金规模_亿元"]
        count = len(part)
        manager_rows.append(
            {
                "基金管理人": manager,
                "管理人简称": manager_short_name(manager),
                "ETF总数量": count,
                "ETF总规模_亿元": round4(scale.sum()),
                "规模小于1亿数量": int((scale < 1).sum()),
                "规模小于2亿数量": int((scale < 2).sum()),
                "规模小于10亿数量": int((scale < 10).sum()),
                "规模小于1亿占比": int((scale < 1).sum()) / count if count else 0,
                "规模小于2亿占比": int((scale < 2).sum()) / count if count else 0,
                "规模小于10亿占比": int((scale < 10).sum()) / count if count else 0,
            }
        )
    manager = add_pct_display(pd.DataFrame(manager_rows).sort_values(["ETF总规模_亿元"], ascending=False).reset_index(drop=True))
    return full, by_type, manager


def quality_checks(main: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    issues = []
    broad_count = int(main["是否宽基ETF"].eq("是").sum())
    stock_count = int(main["是否股票ETF"].eq("是").sum())
    of_count = int(main["Wind代码"].astype(str).str.upper().str.endswith(".OF").sum())
    missing_critical = main[
        main["基金上市地点"].map(clean_text).eq("")
        | main["上市日期"].map(clean_text).eq("")
        | main["成立日期"].map(clean_text).eq("")
        | main["基金规模_亿元"].isna()
        | main["基金规模_亿元"].le(0)
    ]
    if broad_count == 0:
        issues.append("宽基ETF数量为0，分类失败")
    if stock_count == 0:
        issues.append("股票ETF数量为0，分类失败")
    if of_count:
        issues.append(f"主分析池混入.OF产品{of_count}只")
    if len(missing_critical):
        issues.append(f"主分析池存在关键字段缺失产品{len(missing_critical)}只")

    all_share_mask = main[["基金简称", "基金全称", "跟踪指数", "业绩比较基准"]].fillna("").agg(" ".join, axis=1).map(
        lambda x: bool(contains_terms(x, INDUSTRY_ALL_SHARE_CHECK_TERMS))
    )
    all_share = main.loc[all_share_mask, ["Wind代码", "基金简称", "跟踪指数", "股票ETF内部类型", "产品类型_汇总", "命中关键词", "分类备注"]].copy()
    all_share["是否分类异常"] = all_share["股票ETF内部类型"].eq("宽基ETF").map({True: "是", False: "否"})
    if all_share["是否分类异常"].eq("是").any():
        issues.append("行业全指类产品被误判为宽基ETF")

    text = main[["基金简称", "基金全称", "跟踪指数", "业绩比较基准"]].fillna("").agg(" ".join, axis=1)
    sample_rows = []
    for term in CORE_BROAD_CHECK_TERMS:
        sample = main.loc[text.str.contains(term, regex=False, na=False)].head(10).copy()
        for _, row in sample.iterrows():
            sample_rows.append(
                {
                    "命中关键词": term,
                    "Wind代码": row["Wind代码"],
                    "基金简称": row["基金简称"],
                    "跟踪指数": row["跟踪指数"],
                    "股票ETF内部类型": row["股票ETF内部类型"],
                    "产品类型_汇总": row["产品类型_汇总"],
                    "分类备注": row["分类备注"],
                }
            )
    broad_samples = pd.DataFrame(sample_rows)
    return all_share, broad_samples, issues


def note_df(lines: list[str]) -> pd.DataFrame:
    return pd.DataFrame({"序号": range(1, len(lines) + 1), "说明": lines})


def write_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    if path.exists():
        print(f"覆盖已有输出：{path}")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, data in sheets.items():
            data.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    style_workbook(path)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Microsoft YaHei", size=10)
    thin = Side(style="thin", color="D9E1F2")
    border = Border(bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
        for col_idx in range(1, ws.max_column + 1):
            values = [clean_text(ws.cell(row, col_idx).value) for row in range(1, min(ws.max_row, 250) + 1)]
            max_len = max((len(value) for value in values), default=8)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len * 1.15 + 2, 10), 42)
    wb.save(path)


def md_table(df: pd.DataFrame) -> str:
    if df.empty:
        return "（无记录）"
    cols = [str(c) for c in df.columns]
    rows = []
    for _, row in df.iterrows():
        rows.append([clean_text(row.get(c)).replace("|", "\\|") for c in df.columns])
    return "\n".join(
        [
            "| " + " | ".join(cols) + " |",
            "| " + " | ".join(["---"] * len(cols)) + " |",
            *["| " + " | ".join(row) + " |" for row in rows],
        ]
    )


def build_markdown(
    raw_rows: int,
    footer_count: int,
    raw_no_footer: pd.DataFrame,
    main: pd.DataFrame,
    observation: pd.DataFrame,
    field_mapping: pd.DataFrame,
    checks: pd.DataFrame,
    all_share_check: pd.DataFrame,
    broad_samples: pd.DataFrame,
) -> str:
    total_scale = main["基金规模_亿元"].sum()
    asset = group_structure(main, "资产类型")
    product = group_structure(main, "产品类型_汇总")
    stock_inner = group_structure(main[main["资产类型"].eq("股票ETF")], "股票ETF内部类型")
    return f"""# ICI课题 P17-P23 中国ETF数据处理说明（修正版）

## 本次修正原因
- 上一版产品池混入待上市/未上市观察产品，且股票ETF内部分类存在宽基ETF识别不足问题。
- 本次主图仅使用严格筛选后的上市ETF分析池，`.OF` 和关键字段缺失产品只进入观察池。

## 路径
- 原始文件路径：`{INPUT_PATH}`
- 输出路径：`{OUTPUT_DIR}`

## 原始数据结构
- 原始 Sheet：Sheet1
- 原始行数：{raw_rows}
- 剔除 footer 行数：{footer_count}
- 剔除 footer 后产品记录：{len(raw_no_footer)}

## 剔除 footer 的逻辑
- 若 `Wind代码` 以“数据来源”开头，则视为数据来源说明行，不作为产品记录。

## 为什么主分析池只使用 .SH/.SZ
- P17-P23 需要反映已上市交易 ETF 市场结构，`.SH` 和 `.SZ` 对应沪深交易所上市交易产品。
- `.OF` 后缀产品可能为场外、待上市或缺少交易信息产品，不能与上市交易 ETF 混算。

## .OF 产品如何处理
- `.OF` 后缀、关键字段缺失、或名称包含 ETF 但未进入主分析池的产品，统一进入“未上市或待上市观察池”。
- 观察池仅用于复核和说明，不进入 P17-P23 主图计算。

## 字段映射
{md_table(field_mapping)}

## 上市ETF分析池筛选口径
- Wind代码后缀为 `.SH` 或 `.SZ`；
- 基金上市地点、上市日期、成立日期均非空；
- 基金规模_亿元非空且大于 0；
- 基金简称或基金全称包含 ETF；
- 排除 ETF联接、ETF连接、联接A、联接C、场外联接、FOF、LOF 等非交易型 ETF。

## 资产类型分类规则
- 互斥优先级：货币ETF > 商品ETF > 债券ETF > 跨境ETF > 股票ETF > 其他ETF。
- 港股通、恒生、QDII、美国、纳斯达克等归为跨境ETF。
- 有色金属、煤炭、钢铁、黄金股、矿业等股票行业主题不因名称含商品相关词而归为商品ETF，除非投资类型明确为商品型基金。

## 股票ETF内部类型分类规则
- 仅对资产类型为股票ETF的产品分类。
- 内部类型包括：宽基ETF、行业主题ETF、策略ETF、其他股票ETF。
- 若宽基指数同时带有增强/策略字样，内部类型仍按宽基ETF，另在“是否策略ETF”和分类备注中标记。

## 宽基ETF识别规则
- 使用严格核心宽基指数清单，不泛化使用“全指”。
- 先判断“中证全指证券公司、全指医药、全指半导体、全指软件、全指食品、全指电力、全指通信”等行业全指排除词，再判断宽基关键词。

## 分类质量检查
{md_table(checks)}

## 行业全指复核
{md_table(all_share_check.head(30))}

## 核心宽基关键词样本
{md_table(broad_samples.head(50))}

## 每个输出文件对应 PPT 页
- P17：{OUTPUT_FILES["P17"]}
- P18：{OUTPUT_FILES["P18"]}
- P19：{OUTPUT_FILES["P19"]}
- P20：{OUTPUT_FILES["P20"]}
- P21：{OUTPUT_FILES["P21"]}
- P22-P23：{OUTPUT_FILES["P22P23"]}
- 汇总作图数据：{OUTPUT_FILES["ARCHIVE"]}

## P17-P23 初步描述性结论
- 上市ETF分析池共 {len(main)} 只，总规模约 {round4(total_scale)} 亿元。
- 资产类型结构：{'; '.join(f"{r['资产类型']} {r['ETF数量']}只、{r['总规模_亿元']}亿元" for _, r in asset.iterrows())}。
- 产品类型汇总结构：{'; '.join(f"{r['产品类型_汇总']} {r['ETF数量']}只、{r['总规模_亿元']}亿元" for _, r in product.iterrows())}。
- 股票ETF内部结构：{'; '.join(f"{r['股票ETF内部类型']} {r['ETF数量']}只、{r['总规模_亿元']}亿元" for _, r in stock_inner.iterrows())}。
- 观察池共 {len(observation)} 只，不参与主图计算。
- 参考 PPT 可作为股票型ETF头部竞争格局的对照材料，但本次P17-P23主图数据以修正后的上市ETF分析池为准。

## 分类复核建议
- 复核“增强策略 + 宽基指数”的双重标签产品。
- 复核行业全指、央企/国企主题、黄金股/矿业/有色金属等边界产品。
- 如后续需要与参考 PPT 对比，请保持“全市场上市ETF口径”和“股票型ETF口径”分开。

## 美国侧提示
- 美国数据来自 ICI 2026 Fact Book：ETF数量 4,495只，ETF总净资产 13.4万亿美元。
"""


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"输入文件不存在：{INPUT_PATH}")

    raw_sheets, raw_no_footer, std, raw_rows = read_and_standardize()
    footer_count = raw_rows - len(raw_no_footer)
    main_pool, observation_pool, screening_process = finalize_pools(std)

    # Enforce numeric types after filtering.
    for frame in [main_pool, observation_pool]:
        frame["基金规模_亿元"] = pd.to_numeric(frame["基金规模_亿元"], errors="coerce")
        frame["管理费率"] = pd.to_numeric(frame["管理费率"], errors="coerce")
        frame["托管费率"] = pd.to_numeric(frame["托管费率"], errors="coerce")

    all_share_check, broad_samples, issues = quality_checks(main_pool)
    quality_rows = [
        {"检查项": "宽基ETF数量是否为0", "结果": "通过" if main_pool["是否宽基ETF"].eq("是").sum() > 0 else "失败", "数值": int(main_pool["是否宽基ETF"].eq("是").sum()), "说明": ""},
        {"检查项": "股票ETF数量是否明显大于0", "结果": "通过" if main_pool["是否股票ETF"].eq("是").sum() > 0 else "失败", "数值": int(main_pool["是否股票ETF"].eq("是").sum()), "说明": ""},
        {"检查项": "是否有.OF混入主分析池", "结果": "通过" if not main_pool["Wind代码"].astype(str).str.upper().str.endswith(".OF").any() else "失败", "数值": int(main_pool["Wind代码"].astype(str).str.upper().str.endswith(".OF").sum()), "说明": ""},
        {"检查项": "是否有缺失日期/规模产品混入主分析池", "结果": "通过" if not issues or not any("关键字段缺失" in x for x in issues) else "失败", "数值": "", "说明": ""},
        {"检查项": "行业全指类是否误判宽基", "结果": "通过" if not all_share_check["是否分类异常"].eq("是").any() else "失败", "数值": int(all_share_check["是否分类异常"].eq("是").sum()) if len(all_share_check) else 0, "说明": ""},
    ]
    checks = pd.DataFrame(quality_rows)
    if issues:
        raise RuntimeError("；".join(issues))

    missing = pd.DataFrame(
        [
            {
                "字段名": col,
                "缺失数量": int(main_pool[col].isna().sum() + main_pool[col].map(clean_text).eq("").sum()),
                "缺失比例": (main_pool[col].isna().sum() + main_pool[col].map(clean_text).eq("").sum()) / len(main_pool),
            }
            for col in BASE_COLUMNS
        ]
    )
    missing = add_pct_display(missing)
    duplicates = main_pool[main_pool.duplicated("Wind代码", keep=False)][["Wind代码", "基金简称", "基金管理人"]].copy()
    if duplicates.empty:
        duplicates = pd.DataFrame(columns=["Wind代码", "基金简称", "基金管理人", "重复次数", "备注"])
    else:
        duplicates["重复次数"] = duplicates["Wind代码"].map(main_pool["Wind代码"].value_counts())
        duplicates["备注"] = "Wind代码重复，需要复核"

    sheet_overview = pd.DataFrame(
        [{"sheet名称": name, "行数": len(df), "列数": len(df.columns), "是否读取成功": "是", "备注": ""} for name, df in raw_sheets.items()]
    )
    pool_check = pd.DataFrame(
        [
            {"检查项": "Wind代码后缀仅.SH/.SZ", "结果": "是", "数值": int(main_pool["Wind代码"].astype(str).str.contains(r"\\.(?:SH|SZ)$", regex=True).sum()), "说明": ""},
            {"检查项": "上市日期非空", "结果": "是", "数值": int(main_pool["上市日期"].map(clean_text).ne("").sum()), "说明": ""},
            {"检查项": "成立日期非空", "结果": "是", "数值": int(main_pool["成立日期"].map(clean_text).ne("").sum()), "说明": ""},
            {"检查项": "规模大于0", "结果": "是", "数值": int(main_pool["基金规模_亿元"].gt(0).sum()), "说明": ""},
            {"检查项": "观察池数量", "结果": "仅复核", "数值": len(observation_pool), "说明": ".OF、缺失关键交易信息或名称含ETF但未入主池"},
        ]
    )

    total_scale = main_pool["基金规模_亿元"].sum()
    desc = pd.DataFrame(
        [
            {"指标": "原始行数", "数值": raw_rows, "单位": "行", "说明": ""},
            {"指标": "剔除 footer 行数", "数值": footer_count, "单位": "行", "说明": "Wind数据来源说明行"},
            {"指标": "ETF候选产品数", "数值": int((std["基金简称"].fillna("") + std["基金全称"].fillna("")).str.contains("ETF", case=False, na=False).sum()), "单位": "只", "说明": "名称含ETF"},
            {"指标": "上市ETF分析池数量", "数值": len(main_pool), "单位": "只", "说明": ""},
            {"指标": "观察池数量", "数值": len(observation_pool), "单位": "只", "说明": ""},
            {"指标": "ETF总规模_亿元", "数值": round4(total_scale), "单位": "亿元", "说明": ""},
            {"指标": "成立日期最早年份", "数值": pd.to_datetime(main_pool["成立日期"]).dt.year.min(), "单位": "年", "说明": ""},
            {"指标": "成立日期最新年份", "数值": pd.to_datetime(main_pool["成立日期"]).dt.year.max(), "单位": "年", "说明": ""},
            {"指标": "管理人数量", "数值": main_pool["基金管理人"].nunique(), "单位": "家", "说明": ""},
            {"指标": "资产类型数量分布", "数值": "；".join(f"{k}:{v}" for k, v in main_pool["资产类型"].value_counts().items()), "单位": "只", "说明": ""},
            {"指标": "资产类型规模分布", "数值": "；".join(f"{k}:{round4(v)}" for k, v in main_pool.groupby("资产类型")["基金规模_亿元"].sum().items()), "单位": "亿元", "说明": ""},
            {"指标": "股票ETF内部类型数量分布", "数值": "；".join(f"{k}:{v}" for k, v in main_pool.loc[main_pool["资产类型"].eq("股票ETF"), "股票ETF内部类型"].value_counts().items()), "单位": "只", "说明": ""},
            {"指标": "产品类型_汇总数量分布", "数值": "；".join(f"{k}:{v}" for k, v in main_pool["产品类型_汇总"].value_counts().items()), "单位": "只", "说明": ""},
            {"指标": "产品类型_汇总规模分布", "数值": "；".join(f"{k}:{round4(v)}" for k, v in main_pool.groupby("产品类型_汇总")["基金规模_亿元"].sum().items()), "单位": "亿元", "说明": ""},
            {"指标": "宽基ETF数量", "数值": int(main_pool["是否宽基ETF"].eq("是").sum()), "单位": "只", "说明": ""},
            {"指标": "行业主题ETF数量", "数值": int(main_pool["是否行业主题ETF"].eq("是").sum()), "单位": "只", "说明": ""},
            {"指标": "策略ETF数量", "数值": int(main_pool["是否策略ETF"].eq("是").sum()), "单位": "只", "说明": ""},
        ]
    )

    p17_all = annual_table(main_pool)
    p17_asset = annual_table(main_pool, "资产类型")
    p17_product = annual_table(main_pool, "产品类型_汇总")
    p18_asset = group_structure(main_pool, "资产类型")
    p18_product = group_structure(main_pool, "产品类型_汇总")
    p18_stock = group_structure(main_pool[main_pool["资产类型"].eq("股票ETF")], "股票ETF内部类型")
    sample_fields = ["Wind代码", "基金简称", "基金管理人", "资产类型", "股票ETF内部类型", "产品类型_汇总", "基金规模_亿元", "成立日期", "跟踪指数", "投资类型", "命中关键词", "分类备注"]
    sample_parts = []
    for _, part in main_pool.groupby("产品类型_汇总"):
        sample_parts.append(part.sample(min(len(part), 20), random_state=20260629))
    p18_sample = pd.concat(sample_parts, ignore_index=True)[sample_fields] if sample_parts else pd.DataFrame(columns=sample_fields)
    p19_rank, p19_cr = manager_tables(main_pool)
    p20_full, p20_type, p20_manager = size_tables(main_pool)

    cr_lookup = p19_cr.set_index("指标")
    metric_rows = [
        {"指标": "ETF总数量", "数值": len(main_pool), "单位": "只", "说明": "上市ETF分析池"},
        {"指标": "ETF总规模", "数值": round4(total_scale), "单位": "亿元", "说明": ""},
    ]
    for asset in ["股票ETF", "债券ETF", "商品ETF", "货币ETF", "跨境ETF"]:
        part = main_pool[main_pool["资产类型"].eq(asset)]
        metric_rows += [
            {"指标": f"{asset}数量", "数值": len(part), "单位": "只", "说明": ""},
            {"指标": f"{asset}规模", "数值": round4(part["基金规模_亿元"].sum()), "单位": "亿元", "说明": ""},
            {"指标": f"{asset}规模占比", "数值": part["基金规模_亿元"].sum() / total_scale if total_scale else 0, "数值展示": pct(part["基金规模_亿元"].sum() / total_scale if total_scale else 0), "单位": "%", "说明": ""},
        ]
    for product in ["宽基ETF", "行业主题ETF", "策略ETF"]:
        if product == "策略ETF":
            part = main_pool[main_pool["是否策略ETF"].eq("是")]
            product_note = "是否策略ETF标记，含宽基增强/指数增强"
        else:
            part = main_pool[main_pool["股票ETF内部类型"].eq(product)]
            product_note = "股票ETF内部类型"
        metric_rows += [
            {"指标": f"{product}数量", "数值": len(part), "单位": "只", "说明": product_note},
            {"指标": f"{product}规模", "数值": round4(part["基金规模_亿元"].sum()), "单位": "亿元", "说明": product_note},
            {"指标": f"{product}规模占比", "数值": part["基金规模_亿元"].sum() / total_scale if total_scale else 0, "数值展示": pct(part["基金规模_亿元"].sum() / total_scale if total_scale else 0), "单位": "%", "说明": ""},
        ]
    multi_asset = main_pool[~main_pool["资产类型"].eq("股票ETF")]
    metric_rows += [
        {"指标": "前五大管理人规模占比", "数值": cr_lookup.loc["CR5", "规模口径"], "数值展示": pct(cr_lookup.loc["CR5", "规模口径"]), "单位": "%", "说明": ""},
        {"指标": "前十大管理人规模占比", "数值": cr_lookup.loc["CR10", "规模口径"], "数值展示": pct(cr_lookup.loc["CR10", "规模口径"]), "单位": "%", "说明": ""},
        {"指标": "规模小于1亿产品占比", "数值": (main_pool["基金规模_亿元"] < 1).mean(), "数值展示": pct((main_pool["基金规模_亿元"] < 1).mean()), "单位": "%", "说明": ""},
        {"指标": "规模小于2亿产品占比", "数值": (main_pool["基金规模_亿元"] < 2).mean(), "数值展示": pct((main_pool["基金规模_亿元"] < 2).mean()), "单位": "%", "说明": ""},
        {"指标": "规模小于10亿产品占比", "数值": (main_pool["基金规模_亿元"] < 10).mean(), "数值展示": pct((main_pool["基金规模_亿元"] < 10).mean()), "单位": "%", "说明": ""},
        {"指标": "多资产ETF数量", "数值": len(multi_asset), "单位": "只", "说明": "非股票ETF"},
        {"指标": "多资产ETF规模", "数值": round4(multi_asset["基金规模_亿元"].sum()), "单位": "亿元", "说明": "非股票ETF"},
        {"指标": "多资产ETF规模占比", "数值": multi_asset["基金规模_亿元"].sum() / total_scale if total_scale else 0, "数值展示": pct(multi_asset["基金规模_亿元"].sum() / total_scale if total_scale else 0), "单位": "%", "说明": "非股票ETF"},
        {"指标": "平均管理费率", "数值": round4(main_pool["管理费率"].mean()), "单位": "%", "说明": "原始费率字段单位为%"},
        {"指标": "平均托管费率", "数值": round4(main_pool["托管费率"].mean()), "单位": "%", "说明": "原始费率字段单位为%"},
    ]
    p21 = pd.DataFrame(metric_rows)
    if "数值展示" not in p21.columns:
        p21["数值展示"] = ""

    p22 = pd.DataFrame(
        [
            {"转向": "发行导向 → 持营导向", "支撑指标": "规模小于10亿产品占比", "数值": (main_pool["基金规模_亿元"] < 10).mean(), "数值展示": pct((main_pool["基金规模_亿元"] < 10).mean()), "单位": "%", "解释": "长尾产品占比体现持营压力。"},
            {"转向": "指数覆盖 → 场景覆盖", "支撑指标": "行业主题ETF+策略ETF数量占比", "数值": main_pool["产品类型_汇总"].isin(["行业主题ETF", "策略ETF"]).mean(), "数值展示": pct(main_pool["产品类型_汇总"].isin(["行业主题ETF", "策略ETF"]).mean()), "单位": "%", "解释": "行业主题和策略产品体现从指数覆盖走向投资场景覆盖。"},
            {"转向": "费率竞争 → 生态竞争", "支撑指标": "平均管理费率", "数值": round4(main_pool["管理费率"].mean()), "数值展示": "", "单位": "%", "解释": "低费率只是基础，长尾与头部集中说明生态能力更关键。"},
            {"转向": "单品销售 → 组合解决方案", "支撑指标": "多资产ETF规模占比", "数值": multi_asset["基金规模_亿元"].sum() / total_scale if total_scale else 0, "数值展示": pct(multi_asset["基金规模_亿元"].sum() / total_scale if total_scale else 0), "单位": "%", "解释": "债券、商品、跨境等非股票ETF支持组合配置工具箱。"},
        ]
    )
    p23 = pd.DataFrame(
        [
            {"指标": "ETF总数量", "数值": len(main_pool), "单位": "只", "可用于哪一句结论": "用于说明中国上市ETF市场供给扩张"},
            {"指标": "ETF总规模", "数值": round4(total_scale), "单位": "亿元", "可用于哪一句结论": "用于说明中国ETF市场已有规模基础"},
            {"指标": "前十大管理人规模占比", "数值": cr_lookup.loc["CR10", "规模口径"], "数值展示": pct(cr_lookup.loc["CR10", "规模口径"]), "单位": "%", "可用于哪一句结论": "用于说明头部集中"},
            {"指标": "小规模产品占比", "数值": (main_pool["基金规模_亿元"] < 10).mean(), "数值展示": pct((main_pool["基金规模_亿元"] < 10).mean()), "单位": "%", "可用于哪一句结论": "用于说明持营压力"},
            {"指标": "股票ETF规模占比", "数值": main_pool.loc[main_pool["资产类型"].eq("股票ETF"), "基金规模_亿元"].sum() / total_scale, "数值展示": pct(main_pool.loc[main_pool["资产类型"].eq("股票ETF"), "基金规模_亿元"].sum() / total_scale), "单位": "%", "可用于哪一句结论": "用于说明权益属性较强"},
            {"指标": "多资产ETF规模占比", "数值": multi_asset["基金规模_亿元"].sum() / total_scale if total_scale else 0, "数值展示": pct(multi_asset["基金规模_亿元"].sum() / total_scale if total_scale else 0), "单位": "%", "可用于哪一句结论": "用于说明配置工具箱仍待完善"},
        ]
    )

    write_excel(
        OUTPUT_DIR / OUTPUT_FILES["00"],
        {
            "sheet概览": sheet_overview,
            "字段映射表": pd.DataFrame([{"标准字段名": v, "原始字段名": k, "所在sheet": "Sheet1", "识别方式": "字段名精确映射", "备注": ""} for k, v in RAW_TO_STD.items()]),
            "主分析池筛选过程": screening_process,
            "主分析池口径检查": pool_check,
            "ETF分析池_上市交易": main_pool,
            "观察池_OF或缺失交易信息": observation_pool,
            "缺失值检查": missing,
            "重复值检查": duplicates,
            "分类结果检查": pd.concat([checks, pd.DataFrame([{"检查项": "行业全指复核样本数", "结果": "见说明文档", "数值": len(all_share_check), "说明": ""}, {"检查项": "核心宽基关键词样本数", "结果": "见说明文档", "数值": len(broad_samples), "说明": ""}])], ignore_index=True),
            "描述性统计": desc,
            "行业全指复核": all_share_check,
            "核心宽基样本": broad_samples,
        },
    )
    write_excel(OUTPUT_DIR / OUTPUT_FILES["P17"], {"P17_年度新发累计_全市场": p17_all, "P17_年度新发累计_按资产类型": p17_asset, "P17_年度新发累计_按产品类型汇总": p17_product})
    write_excel(OUTPUT_DIR / OUTPUT_FILES["P18"], {"P18_资产类型结构": p18_asset, "P18_产品类型汇总结构": p18_product, "P18_股票ETF内部结构": p18_stock, "P18_分类复核样本": p18_sample})
    write_excel(OUTPUT_DIR / OUTPUT_FILES["P19"], {"P19_管理人规模排名": p19_rank, "P19_前十大管理人": p19_rank.head(10), "P19_CR集中度": p19_cr})
    write_excel(OUTPUT_DIR / OUTPUT_FILES["P20"], {"P20_规模区间分布_全市场": p20_full, "P20_规模区间分布_按产品类型汇总": p20_type, "P20_小规模产品占比_按管理人": p20_manager})
    write_excel(OUTPUT_DIR / OUTPUT_FILES["P21"], {"P21_中国ETF总览指标": p21, "说明": note_df(["美国数据来自 ICI 2026 Fact Book，ETF数量 4,495只，ETF总净资产 13.4万亿美元。"])})
    write_excel(OUTPUT_DIR / OUTPUT_FILES["P22P23"], {"P22_四个转向支撑数据": p22, "P23_最终升华可引用数据": p23})
    write_excel(
        OUTPUT_DIR / OUTPUT_FILES["ARCHIVE"],
        {
            "P17_年度新发累计": p17_all,
            "P17_按资产类型年度新发": p17_asset,
            "P18_资产类型结构": p18_asset,
            "P18_产品类型汇总结构": p18_product,
            "P18_股票ETF内部结构": p18_stock,
            "P19_管理人规模排名": p19_rank,
            "P19_CR集中度": p19_cr,
            "P20_规模区间分布": p20_full,
            "P20_小规模产品占比": p20_manager,
            "P21_中国ETF总览指标": p21,
            "P22_四个转向支撑数据": p22,
            "P23_最终升华引用数据": p23,
        },
    )

    readme = build_markdown(raw_rows, footer_count, raw_no_footer, main_pool, observation_pool, pd.DataFrame([{"标准字段名": v, "原始字段名": k, "所在sheet": "Sheet1", "识别方式": "字段名精确映射", "备注": ""} for k, v in RAW_TO_STD.items()]), checks, all_share_check, broad_samples)
    readme_path = OUTPUT_DIR / OUTPUT_FILES["README"]
    if readme_path.exists():
        print(f"覆盖已有输出：{readme_path}")
    readme_path.write_text(readme, encoding="utf-8")

    generated = [OUTPUT_DIR / name for name in OUTPUT_FILES.values()]
    print("原始文件是否成功读取：是")
    print(f"原始行数：{raw_rows}")
    print(f"剔除 footer 后行数：{len(raw_no_footer)}")
    print(f"上市ETF分析池数量：{len(main_pool)}")
    print(f"观察池数量：{len(observation_pool)}")
    print(f"ETF总规模_亿元：{round4(total_scale)}")
    print(f"宽基ETF数量：{int(main_pool['是否宽基ETF'].eq('是').sum())}")
    print(f"行业主题ETF数量：{int(main_pool['是否行业主题ETF'].eq('是').sum())}")
    print(f"策略ETF数量：{int(main_pool['是否策略ETF'].eq('是').sum())}")
    print(f"是否有 .OF 混入主分析池：{'是' if main_pool['Wind代码'].astype(str).str.upper().str.endswith('.OF').any() else '否'}")
    print(f"是否有缺失日期/规模产品混入主分析池：{'是' if len(main_pool[(main_pool['上市日期'].eq('')) | (main_pool['成立日期'].eq('')) | (main_pool['基金规模_亿元'].isna()) | (main_pool['基金规模_亿元'].le(0))]) else '否'}")
    print("生成的文件清单：")
    for path in generated:
        print(f"- {path}")
    if all_share_check["是否分类异常"].eq("是").any():
        print("分类异常：行业全指类产品误判宽基")
    else:
        print("分类异常：未发现")


if __name__ == "__main__":
    main()
