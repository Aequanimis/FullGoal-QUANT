from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\全市场ETF基础信息_策略ETF池二次修正版.xlsx")
OUTPUT_DIR = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池")
OUTPUT_XLSX = OUTPUT_DIR / "广义策略ETF_Wind代码池.xlsx"
TXT_BROAD = OUTPUT_DIR / "广义策略ETF_223只_Wind代码.txt"
TXT_CORE = OUTPUT_DIR / "核心策略指数ETF_168只_Wind代码.txt"
TXT_ENHANCED = OUTPUT_DIR / "指数增强多因子ETF_55只_Wind代码.txt"

SOURCE_SHEET = "策略ETF_最终统计池"
SHEET_BROAD = "广义策略ETF_223只"
SHEET_CORE = "核心策略指数ETF_168只"
SHEET_ENHANCED = "指数增强多因子ETF_55只"
SHEET_STATS = "代码池统计"
SHEET_ERRORS = "异常代码检查"

OUTPUT_FIELDS = [
    "Wind代码",
    "基金代码",
    "交易代码",
    "证券简称",
    "基金简称",
    "基金全称",
    "基金管理人",
    "基金上市地点",
    "上市日期",
    "基金成立日",
    "跟踪指数代码",
    "跟踪指数名称",
    "最新基金规模(亿)",
    "一级策略大类",
    "二级策略类别",
    "市场范围_二次修正",
    "统计口径分类",
    "是否纳入核心策略ETF统计",
    "是否纳入广义策略ETF统计",
]

PRIMARY_ORDER = [
    "红利",
    "自由现金流",
    "质量",
    "价值",
    "成长",
    "低波",
    "ESG",
    "基本面策略",
    "等权/另类加权",
    "指数增强/多因子",
    "待核验",
]

SECONDARY_ORDER = [
    "普通红利",
    "红利低波",
    "港股通低波红利",
    "红利质量",
    "红利价值",
    "红利成长",
    "央企红利",
    "国企红利",
    "港股通红利",
    "股东回报",
    "央企股东回报",
    "自由现金流",
    "质量",
    "价值",
    "成长",
    "低波",
    "ESG",
    "基本面策略",
    "等权/另类加权",
    "指数增强/多因子",
    "待核验",
]

CLASS_ORDER = {
    "核心策略指数ETF": 0,
    "广义策略ETF_指数增强": 1,
    "观察池": 2,
    "排除": 3,
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=9, color="1F1F1F")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN_GRAY = Side(style="thin", color="D9E1F2")
BODY_BORDER = Border(bottom=THIN_GRAY)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    value = str(value).replace("\u3000", " ").replace("\u200b", "")
    value = re.sub(r"[\r\n\t]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def normalize_wind_code(value: Any) -> str:
    raw = clean_text(value).upper().replace("。", ".")
    match = re.fullmatch(r"(\d{6})\.(SH|SZ)", raw)
    return f"{match.group(1)}.{match.group(2)}" if match else raw


def validate_and_deduplicate(
    records: list[dict[str, Any]],
    pool_name: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    valid: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    seen: set[str] = set()

    for record in records:
        code = normalize_wind_code(record.get("Wind代码"))
        reason = ""
        if not code:
            reason = "Wind代码为空"
        elif not re.fullmatch(r"\d{6}\.(SH|SZ)", code):
            reason = "Wind代码不是6位数字加.SH/.SZ后缀"
        elif code in seen:
            reason = "重复Wind代码"

        if reason:
            errors.append(
                {
                    "代码池": pool_name,
                    "Wind代码原值": record.get("Wind代码"),
                    "Wind代码标准化": code,
                    "证券简称": record.get("证券简称"),
                    "统计口径分类": record.get("统计口径分类"),
                    "异常原因": reason,
                }
            )
            continue

        copied = dict(record)
        copied["Wind代码"] = code
        seen.add(code)
        valid.append(copied)

    return valid, errors


def sort_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    primary_rank = {value: index for index, value in enumerate(PRIMARY_ORDER)}
    secondary_rank = {value: index for index, value in enumerate(SECONDARY_ORDER)}
    return sorted(
        records,
        key=lambda r: (
            CLASS_ORDER.get(clean_text(r.get("统计口径分类")), 999),
            primary_rank.get(clean_text(r.get("一级策略大类")), 999),
            secondary_rank.get(clean_text(r.get("二级策略类别")), 999),
            -(float(r["最新基金规模(亿)"]) if isinstance(r.get("最新基金规模(亿)"), (int, float)) else -1),
            clean_text(r.get("Wind代码")),
        ),
    )


def write_records(ws, headers: list[str], records: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for record in records:
        ws.append([record.get(header) for header in headers])


def style_sheet(ws, headers: list[str]) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BODY_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    sample_limit = min(ws.max_row, 300)
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, sample_limit + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 50))
        if header in {"基金全称", "跟踪指数名称", "说明", "异常原因"}:
            width = min(max(max_len * 1.2, 18), 40)
        else:
            width = min(max(max_len * 1.15, 11), 27)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for header in ["上市日期", "基金成立日"]:
        if header in headers:
            col = headers.index(header) + 1
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col).value, (date, datetime)):
                    ws.cell(row_idx, col).number_format = "yyyy-mm-dd"

    for col_idx, header in enumerate(headers, 1):
        if "规模" in header:
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "#,##0.0000"


def pool_size(records: list[dict[str, Any]]) -> float:
    return round(
        sum(
            float(record["最新基金规模(亿)"])
            for record in records
            if isinstance(record.get("最新基金规模(亿)"), (int, float))
        ),
        4,
    )


def build_stats_rows(
    broad: list[dict[str, Any]],
    core: list[dict[str, Any]],
    enhanced: list[dict[str, Any]],
    error_count: int,
) -> list[list[Any]]:
    rows: list[list[Any]] = []

    def add(section: str, item: str, count: Any, size: Any = None, note: str = "") -> None:
        rows.append([section, item, count, size, note])

    add("总体统计", "广义策略ETF", len(broad), pool_size(broad), "核心策略指数ETF + 指数增强/多因子")
    add("总体统计", "核心策略指数ETF", len(core), pool_size(core))
    add("总体统计", "指数增强多因子ETF", len(enhanced), pool_size(enhanced))
    add("数据质量", "异常代码记录", error_count, None, "正常情况下应为0")

    primary_count = Counter(clean_text(r.get("一级策略大类")) or "待核验" for r in broad)
    primary_size: defaultdict[str, float] = defaultdict(float)
    manager_count = Counter(clean_text(r.get("基金管理人")) or "待补充" for r in broad)
    manager_size: defaultdict[str, float] = defaultdict(float)

    for record in broad:
        size = record.get("最新基金规模(亿)")
        if isinstance(size, (int, float)):
            primary_size[clean_text(record.get("一级策略大类")) or "待核验"] += float(size)
            manager_size[clean_text(record.get("基金管理人")) or "待补充"] += float(size)

    primary_rank = {value: index for index, value in enumerate(PRIMARY_ORDER)}
    for primary in sorted(primary_count, key=lambda x: (primary_rank.get(x, 999), x)):
        add(
            "各一级策略大类",
            primary,
            primary_count[primary],
            round(primary_size.get(primary, 0), 4),
            "按广义策略ETF统计",
        )

    for manager, count in manager_count.most_common():
        add(
            "各基金管理人",
            manager,
            count,
            round(manager_size.get(manager, 0), 4),
            "按广义策略ETF统计",
        )

    return rows


def write_txt(path: Path, records: list[dict[str, Any]]) -> None:
    content = "\n".join(clean_text(record.get("Wind代码")) for record in records)
    if content:
        content += "\n"
    # UTF-8 with BOM is convenient for Windows Notepad and Wind copy/paste.
    path.write_text(content, encoding="utf-8-sig", newline="\n")


def main() -> None:
    if not INPUT.exists():
        raise FileNotFoundError(f"输入文件不存在：{INPUT}")

    source_wb = load_workbook(INPUT, read_only=True, data_only=True)
    if SOURCE_SHEET not in source_wb.sheetnames:
        raise KeyError(f"输入文件缺少sheet：{SOURCE_SHEET}")

    source_ws = source_wb[SOURCE_SHEET]
    headers = [cell.value for cell in next(source_ws.iter_rows())]
    missing_fields = [field for field in OUTPUT_FIELDS if field not in headers]
    if missing_fields:
        raise KeyError(f"源sheet缺少输出字段：{missing_fields}")

    records = []
    for values in source_ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and clean_text(value) for value in values):
            continue
        records.append(dict(zip(headers, values)))

    broad_source = [r for r in records if clean_text(r.get("是否纳入广义策略ETF统计")) == "是"]
    core_source = [
        r
        for r in broad_source
        if clean_text(r.get("统计口径分类")) == "核心策略指数ETF"
    ]
    enhanced_source = [
        r
        for r in broad_source
        if clean_text(r.get("统计口径分类")) == "广义策略ETF_指数增强"
    ]

    broad, broad_errors = validate_and_deduplicate(broad_source, "广义策略ETF")
    core, core_errors = validate_and_deduplicate(core_source, "核心策略指数ETF")
    enhanced, enhanced_errors = validate_and_deduplicate(enhanced_source, "指数增强多因子ETF")
    all_errors = broad_errors + core_errors + enhanced_errors

    broad = sort_records(broad)
    core = sort_records(core)
    enhanced = sort_records(enhanced)

    expected = {
        "广义策略ETF": (len(broad), 223),
        "核心策略指数ETF": (len(core), 168),
        "指数增强多因子ETF": (len(enhanced), 55),
    }
    mismatch_notes = [
        {
            "代码池": pool,
            "Wind代码原值": None,
            "Wind代码标准化": None,
            "证券简称": None,
            "统计口径分类": None,
            "异常原因": f"数量与预期不一致：实际{actual}，预期{target}",
        }
        for pool, (actual, target) in expected.items()
        if actual != target
    ]
    all_errors.extend(mismatch_notes)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    broad_ws = wb.create_sheet(SHEET_BROAD)
    core_ws = wb.create_sheet(SHEET_CORE)
    enhanced_ws = wb.create_sheet(SHEET_ENHANCED)
    stats_ws = wb.create_sheet(SHEET_STATS)
    errors_ws = wb.create_sheet(SHEET_ERRORS)

    write_records(broad_ws, OUTPUT_FIELDS, broad)
    write_records(core_ws, OUTPUT_FIELDS, core)
    write_records(enhanced_ws, OUTPUT_FIELDS, enhanced)

    stats_headers = ["统计类别", "项目", "产品数量", "最新规模合计(亿)", "说明"]
    stats_ws.append(stats_headers)
    for row in build_stats_rows(broad, core, enhanced, len(all_errors)):
        stats_ws.append(row)

    error_headers = [
        "代码池",
        "Wind代码原值",
        "Wind代码标准化",
        "证券简称",
        "统计口径分类",
        "异常原因",
    ]
    write_records(errors_ws, error_headers, all_errors)

    for ws, sheet_headers in [
        (broad_ws, OUTPUT_FIELDS),
        (core_ws, OUTPUT_FIELDS),
        (enhanced_ws, OUTPUT_FIELDS),
        (stats_ws, stats_headers),
        (errors_ws, error_headers),
    ]:
        style_sheet(ws, sheet_headers)

    previous_section = None
    for row_idx in range(2, stats_ws.max_row + 1):
        section = stats_ws.cell(row_idx, 1).value
        if section != previous_section:
            for col_idx in range(1, len(stats_headers) + 1):
                stats_ws.cell(row_idx, col_idx).fill = SECTION_FILL
                stats_ws.cell(row_idx, col_idx).font = Font(
                    name="微软雅黑", size=9, bold=True, color="1F1F1F"
                )
        previous_section = section
        stats_ws.cell(row_idx, 3).number_format = "#,##0"
        stats_ws.cell(row_idx, 4).number_format = "#,##0.0000"

    wb.save(OUTPUT_XLSX)
    write_txt(TXT_BROAD, broad)
    write_txt(TXT_CORE, core)
    write_txt(TXT_ENHANCED, enhanced)

    print(f"Excel输出路径：{OUTPUT_XLSX}")
    print(f"广义策略ETF代码池数量：{len(broad)}")
    print(f"核心策略指数ETF代码池数量：{len(core)}")
    print(f"指数增强多因子ETF代码池数量：{len(enhanced)}")
    print(f"异常代码数量：{len(all_errors)}")
    print(f"TXT输出路径：{OUTPUT_DIR}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"处理失败：{exc}", file=sys.stderr)
        raise
