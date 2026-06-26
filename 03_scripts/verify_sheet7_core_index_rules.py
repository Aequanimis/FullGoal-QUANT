from pathlib import Path

from openpyxl import load_workbook


path = (
    Path.home()
    / "Desktop"
    / "Fullgoal"
    / "课题研究"
    / "处理后数据"
    / "wind代码池"
    / "sheet7"
    / "核心策略ETF_跟踪指数规则_验收修正版.xlsx"
)

workbook = load_workbook(path, read_only=True, data_only=True)
print("FILE_SIZE", path.stat().st_size)
print("SHEETS", workbook.sheetnames)
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    print(sheet_name, sheet.max_row, sheet.max_column)

sheet1 = workbook["核心ETF_指数映射_标准版"]
headers1 = [cell.value for cell in next(sheet1.iter_rows(max_row=1))]
rows1 = list(sheet1.iter_rows(min_row=2, values_only=True))
etf_idx = headers1.index("ETF_Code")
idx_idx = headers1.index("Index_Code")
print(
    "ETF_ROWS",
    len(rows1),
    "ETF_UNIQUE",
    len({row[etf_idx] for row in rows1}),
    "INDEX_UNIQUE",
    len({row[idx_idx] for row in rows1}),
)

sheet3 = workbook["指数规则明细_标准版"]
headers3 = [cell.value for cell in next(sheet3.iter_rows(max_row=1))]
complete_idx = headers3.index("是否规则已补全")
complete_values = [row[complete_idx] for row in sheet3.iter_rows(min_row=2, values_only=True)]
print(
    "RULE_COMPLETE",
    complete_values.count("是"),
    "RULE_PENDING",
    len(complete_values) - complete_values.count("是"),
)

has_wind_source_row = False
for sheet in workbook.worksheets:
    for row in sheet.iter_rows(values_only=True):
        if any(isinstance(value, str) and "数据来源：Wind" in value for value in row):
            has_wind_source_row = True
            break
print("HAS_WIND_SOURCE_ROW", has_wind_source_row)
