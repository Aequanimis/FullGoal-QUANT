from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT_PATH = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\原始数据\全市场ETF基础信息.csv.xlsx")
OUTPUT_DIR = Path(r"C:\Users\lvdon\Desktop\Fullgoal\ICI课题\P17-P23_中国ETF数据处理")

OUTPUT_FILES = {
    "00": "00_数据字段检查与描述性统计.xlsx",
    "P17": "P17_中国ETF年度新发与累计数量.xlsx",
    "P18": "P18_中国ETF资产类型与产品类型结构.xlsx",
    "P19": "P19_中国ETF管理人集中度与头部排名.xlsx",
    "P20": "P20_中国ETF规模区间分布与长尾分析.xlsx",
    "P21": "P21_中美对比所需中国ETF总览指标.xlsx",
    "P22P23": "P22-P23_结论页支撑数据摘要.xlsx",
    "ARCHIVE": "ICI课题_P17-P23_中国ETF作图数据汇总.xlsx",
    "README": "ICI课题_P17-P23_数据处理说明.md",
}

STANDARD_FIELDS = [
    "基金代码",
    "基金简称",
    "基金全称",
    "基金管理人",
    "基金公司",
    "成立日期",
    "上市日期",
    "基金类型",
    "投资类型",
    "资产类型",
    "产品类型",
    "跟踪指数",
    "跟踪指数代码",
    "基金规模_亿元",
    "最新规模_亿元",
    "发行规模_亿元",
    "管理费率",
    "托管费率",
    "是否ETF",
    "是否股票ETF",
    "是否跨境ETF",
    "是否策略ETF",
    "是否行业主题ETF",
    "是否宽基ETF",
    "命中关键词",
    "ETF筛选依据",
    "ETF排除依据",
]

FIELD_MAP = {
    "基金代码": "基金代码",
    "基金简称": "基金简称",
    "基金全称": "基金全称",
    "基金管理人": "基金管理人",
    "基金公司": "基金管理人",
    "成立日期": "基金成立日",
    "上市日期": "上市日期",
    "基金类型": "基金类型",
    "投资类型": "投资类型_二级分类",
    "跟踪指数": "跟踪指数名称",
    "跟踪指数代码": "跟踪指数代码",
    "基金规模_亿元": "最新基金规模(亿)",
    "最新规模_亿元": "最新基金规模(亿)",
    "管理费率": "管理费率_支持历史(%)",
    "托管费率": "托管费率_支持历史(%)",
}

BROAD_KEYWORDS = [
    "沪深300",
    "中证500",
    "中证1000",
    "中证2000",
    "创业板",
    "科创50",
    "科创100",
    "上证50",
    "上证180",
    "深证100",
    "A50",
    "A100",
    "A500",
    "MSCI中国A股",
    "全指",
    "综指",
    "中证800",
    "中证A500",
]
INDUSTRY_KEYWORDS = [
    "新能源",
    "光伏",
    "芯片",
    "半导体",
    "人工智能",
    "机器人",
    "军工",
    "医药",
    "消费",
    "金融",
    "证券",
    "银行",
    "地产",
    "传媒",
    "计算机",
    "通信",
    "汽车",
    "电池",
    "储能",
    "双碳",
    "绿色",
    "电力",
    "煤炭",
    "有色",
    "钢铁",
    "基建",
    "农业",
    "酒",
    "食品",
    "云计算",
    "软件",
    "游戏",
    "数字经济",
    "创新药",
    "生物",
    "稀土",
    "家电",
]
STRATEGY_KEYWORDS = [
    "红利",
    "低波",
    "质量",
    "价值",
    "成长",
    "动量",
    "现金流",
    "基本面",
    "股息",
    "ESG",
    "Smart Beta",
    "增强策略",
    "增强",
    "等权",
    "龙头",
    "央企",
    "国企",
    "分红",
    "回购",
    "智选",
]
CROSS_BORDER_KEYWORDS = [
    "纳斯达克",
    "纳指",
    "标普",
    "恒生",
    "港股",
    "日经",
    "德国",
    "法国",
    "海外",
    "QDII",
    "中概",
    "东南亚",
    "恒指",
    "香港",
    "美国",
]
BOND_KEYWORDS = ["债", "国债", "信用债", "可转债", "政金债", "公司债"]
COMMODITY_KEYWORDS = ["黄金", "有色", "商品", "豆粕", "原油", "油气", "能源化工"]
MONEY_KEYWORDS = ["货币", "现金", "保证金", "银华日利"]
ETF_EXCLUDE_KEYWORDS = ["ETF联接", "ETF连接", "联接A", "联接C", "联接", "场外联接", "FOF", "LOF", "指数增强非ETF"]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "nat", "none"}:
        return ""
    return re.sub(r"\s+", " ", text)


def code_text(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text.zfill(6) if text.isdigit() and len(text) < 6 else text


def to_number(value: Any) -> float | None:
    text = clean_text(value).replace(",", "").replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def to_date_str(value: Any) -> str:
    if value is None or clean_text(value) == "":
        return ""
    dt = pd.to_datetime(value, errors="coerce")
    if pd.isna(dt):
        return ""
    return dt.strftime("%Y-%m-%d")


def contains_any(text: str, keywords: list[str]) -> list[str]:
    text_upper = text.upper()
    hits = []
    for keyword in keywords:
        if keyword.upper() in text_upper:
            hits.append(keyword)
    return hits


def safe_ratio(numerator: float, denominator: float) -> float:
    return numerator / denominator if denominator else 0.0


def round4(value: Any) -> Any:
    if pd.isna(value):
        return ""
    if isinstance(value, (int, float)):
        return round(float(value), 4)
    return value


def choose_main_sheet(raw_sheets: dict[str, pd.DataFrame]) -> tuple[str, pd.DataFrame]:
    sheet_name = max(raw_sheets, key=lambda name: len(raw_sheets[name]))
    df = raw_sheets[sheet_name].copy()
    df = df.dropna(how="all")
    return sheet_name, df


def scale_to_yi(value: Any, source_field: str) -> float | None:
    number = to_number(value)
    if number is None:
        return None
    if "万元" in source_field:
        return number / 10000
    if "元" in source_field and "万元" not in source_field and "亿" not in source_field:
        return number / 100000000
    return number


def classify_etf(row: dict[str, Any]) -> tuple[str, str, str]:
    text = " ".join(
        clean_text(row.get(field))
        for field in ["基金简称", "基金全称", "基金类型", "投资类型", "跟踪指数"]
    )
    hits: list[str] = []

    cross_hits = contains_any(text, CROSS_BORDER_KEYWORDS)
    bond_hits = contains_any(text, BOND_KEYWORDS)
    commodity_hits = contains_any(text, COMMODITY_KEYWORDS)
    money_hits = contains_any(text, MONEY_KEYWORDS)
    broad_hits = contains_any(text, BROAD_KEYWORDS)
    industry_hits = contains_any(text, INDUSTRY_KEYWORDS)
    strategy_hits = contains_any(text, STRATEGY_KEYWORDS)

    investment_type = clean_text(row.get("投资类型"))
    if "债券" in investment_type and "债" not in bond_hits:
        bond_hits.append("投资类型_债券")
    if "货币" in investment_type and "货币" not in money_hits:
        money_hits.append("投资类型_货币")
    if "商品" in investment_type and "商品" not in commodity_hits:
        commodity_hits.append("投资类型_商品")
    if "QDII" in investment_type.upper() and "QDII" not in cross_hits:
        cross_hits.append("投资类型_QDII")

    if cross_hits:
        asset_type = "跨境ETF"
        hits.extend(cross_hits)
    elif bond_hits:
        asset_type = "债券ETF"
        hits.extend(bond_hits)
    elif commodity_hits:
        asset_type = "商品ETF"
        hits.extend(commodity_hits)
    elif money_hits:
        asset_type = "货币ETF"
        hits.extend(money_hits)
    elif text:
        asset_type = "股票ETF"
    else:
        asset_type = "其他ETF"

    if asset_type in {"跨境ETF", "债券ETF", "商品ETF", "货币ETF"}:
        product_type = asset_type
    elif strategy_hits:
        product_type = "策略ETF"
        hits.extend(strategy_hits)
    elif industry_hits:
        product_type = "行业主题ETF"
        hits.extend(industry_hits)
    elif broad_hits:
        product_type = "宽基ETF"
        hits.extend(broad_hits)
    elif asset_type == "股票ETF":
        product_type = "宽基ETF" if broad_hits else "其他ETF"
        hits.extend(broad_hits)
    else:
        product_type = "其他ETF"

    return asset_type, product_type, "、".join(dict.fromkeys(hits))


def standardize_dataframe(df: pd.DataFrame, source_sheet: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    std = pd.DataFrame(index=df.index)
    mapping_rows = []
    for std_field in STANDARD_FIELDS:
        source_field = FIELD_MAP.get(std_field, "")
        if source_field and source_field in df.columns:
            if std_field in {"基金代码"}:
                std[std_field] = df[source_field].map(code_text)
            elif std_field in {"成立日期", "上市日期"}:
                std[std_field] = df[source_field].map(to_date_str)
            elif std_field in {"基金规模_亿元", "最新规模_亿元"}:
                std[std_field] = df[source_field].map(lambda v: scale_to_yi(v, source_field))
            elif std_field in {"管理费率", "托管费率"}:
                std[std_field] = df[source_field].map(to_number)
            else:
                std[std_field] = df[source_field].map(clean_text)
            mapping_rows.append(
                {
                    "标准字段名": std_field,
                    "原始字段名": source_field,
                    "所在sheet": source_sheet,
                    "识别方式": "字段名精确/语义映射",
                    "备注": "基金公司由基金管理人同步，便于公司口径统计" if std_field == "基金公司" else "",
                }
            )
        else:
            std[std_field] = ""
            mapping_rows.append(
                {
                    "标准字段名": std_field,
                    "原始字段名": "",
                    "所在sheet": source_sheet,
                    "识别方式": "原始数据未提供/后续派生",
                    "备注": "派生字段" if std_field in {"资产类型", "产品类型", "是否ETF", "命中关键词"} else "原始数据未提供",
                }
            )

    std["基金简称"] = std["基金简称"].where(std["基金简称"] != "", df.get("证券简称", "").map(clean_text))
    if "基金代码" in df.columns:
        std["基金代码"] = df["基金代码"].map(code_text)
    elif "交易代码" in df.columns:
        std["基金代码"] = df["交易代码"].map(code_text)
    elif "Wind代码" in df.columns:
        std["基金代码"] = df["Wind代码"].map(lambda v: code_text(clean_text(v).split(".")[0]))

    screening_basis = []
    exclude_basis = []
    is_etf = []
    for idx, row in std.iterrows():
        combined = " ".join(clean_text(row.get(field)) for field in ["基金简称", "基金全称", "基金类型", "投资类型"])
        wind_code = clean_text(df.loc[idx, "Wind代码"]) if "Wind代码" in df.columns else ""
        listed_code = bool(re.fullmatch(r"\d{6}\.(SH|SZ)", wind_code.upper()))
        include_by_name = "ETF" in combined.upper()
        include_by_source = listed_code and clean_text(row.get("投资类型")) != ""
        include = include_by_name or include_by_source
        excluded = contains_any(combined, ETF_EXCLUDE_KEYWORDS)
        is_etf.append("是" if include and not excluded else "否")
        if include_by_name:
            screening_basis.append("名称/类型字段包含ETF")
        elif include_by_source:
            screening_basis.append("原始文件为全市场ETF基础信息，且Wind代码为场内交易代码")
        else:
            screening_basis.append("名称/类型字段未直接包含ETF，且缺少场内交易代码依据")
        exclude_basis.append("、".join(excluded))
    std["是否ETF"] = is_etf
    std["ETF筛选依据"] = screening_basis
    std["ETF排除依据"] = exclude_basis

    classifications = std.apply(lambda row: classify_etf(row.to_dict()), axis=1)
    std["资产类型"] = [item[0] for item in classifications]
    std["产品类型"] = [item[1] for item in classifications]
    std["命中关键词"] = [item[2] for item in classifications]
    std["是否股票ETF"] = std["资产类型"].eq("股票ETF").map({True: "是", False: "否"})
    std["是否跨境ETF"] = std["资产类型"].eq("跨境ETF").map({True: "是", False: "否"})
    std["是否策略ETF"] = std["产品类型"].eq("策略ETF").map({True: "是", False: "否"})
    std["是否行业主题ETF"] = std["产品类型"].eq("行业主题ETF").map({True: "是", False: "否"})
    std["是否宽基ETF"] = std["产品类型"].eq("宽基ETF").map({True: "是", False: "否"})

    std = std[std["基金代码"].astype(str).str.len() > 0].copy()
    return std.reset_index(drop=True), pd.DataFrame(mapping_rows)


def build_anomaly_tables(etf_pool: pd.DataFrame) -> dict[str, pd.DataFrame]:
    missing = pd.DataFrame(
        [
            {
                "字段名": col,
                "缺失数量": int(etf_pool[col].isna().sum() + (etf_pool[col].astype(str).str.strip() == "").sum()),
                "缺失比例": safe_ratio(
                    int(etf_pool[col].isna().sum() + (etf_pool[col].astype(str).str.strip() == "").sum()),
                    len(etf_pool),
                ),
                "备注": "",
            }
            for col in STANDARD_FIELDS
            if col in etf_pool.columns
        ]
    )
    dup_counts = etf_pool["基金代码"].value_counts()
    duplicates = etf_pool[etf_pool["基金代码"].isin(dup_counts[dup_counts > 1].index)][["基金代码", "基金简称"]].copy()
    if duplicates.empty:
        duplicates = pd.DataFrame(columns=["基金代码", "重复次数", "基金简称", "备注"])
    else:
        duplicates["重复次数"] = duplicates["基金代码"].map(dup_counts)
        duplicates["备注"] = "基金代码重复，需要人工复核"
        duplicates = duplicates[["基金代码", "重复次数", "基金简称", "备注"]]

    date_bad = etf_pool[etf_pool["成立日期"].astype(str).str.strip() == ""][["基金代码", "基金简称", "成立日期"]].copy()
    date_bad["异常原因"] = "成立日期缺失或无法解析，未进入P17年度统计"

    scale_bad = etf_pool[
        etf_pool["基金规模_亿元"].isna() | (pd.to_numeric(etf_pool["基金规模_亿元"], errors="coerce") <= 0)
    ][["基金代码", "基金简称", "基金规模_亿元"]].copy()
    scale_bad["异常原因"] = "规模缺失、无法解析或小于等于0"

    generic = pd.concat(
        [
            date_bad.rename(columns={"成立日期": "异常值", "异常原因": "异常原因"})[["基金代码", "基金简称", "异常值", "异常原因"]],
            scale_bad.rename(columns={"基金规模_亿元": "异常值", "异常原因": "异常原因"})[["基金代码", "基金简称", "异常值", "异常原因"]],
        ],
        ignore_index=True,
    )
    return {
        "missing": missing,
        "duplicates": duplicates,
        "date_bad": date_bad,
        "scale_bad": scale_bad,
        "generic": generic,
    }


def group_structure(df: pd.DataFrame, by_col: str) -> pd.DataFrame:
    total_count = len(df)
    total_scale = pd.to_numeric(df["基金规模_亿元"], errors="coerce").sum()
    rows = []
    for key, part in df.groupby(by_col, dropna=False):
        scale = pd.to_numeric(part["基金规模_亿元"], errors="coerce")
        rows.append(
            {
                by_col: key or "未分类",
                "ETF数量": len(part),
                "数量占比": safe_ratio(len(part), total_count),
                "总规模_亿元": round4(scale.sum()),
                "规模占比": safe_ratio(scale.sum(), total_scale),
                "平均规模_亿元": round4(scale.mean()),
                "规模中位数_亿元": round4(scale.median()),
            }
        )
    return pd.DataFrame(rows).sort_values(["总规模_亿元", "ETF数量"], ascending=[False, False])


def annual_table(df: pd.DataFrame, by_col: str | None = None) -> pd.DataFrame:
    temp = df.copy()
    temp["年份"] = pd.to_datetime(temp["成立日期"], errors="coerce").dt.year
    temp = temp.dropna(subset=["年份"])
    temp["年份"] = temp["年份"].astype(int)
    years = list(range(int(temp["年份"].min()), int(temp["年份"].max()) + 1)) if len(temp) else []
    if by_col is None:
        yearly = temp.groupby("年份").size().reindex(years, fill_value=0).reset_index(name="当年新成立ETF数量")
        yearly["累计ETF数量"] = yearly["当年新成立ETF数量"].cumsum()
        return yearly
    keys = sorted(temp[by_col].dropna().unique().tolist())
    frame = pd.MultiIndex.from_product([years, keys], names=["年份", by_col]).to_frame(index=False)
    counts = temp.groupby(["年份", by_col]).size().reset_index(name="当年新成立ETF数量")
    result = frame.merge(counts, on=["年份", by_col], how="left").fillna({"当年新成立ETF数量": 0})
    result["当年新成立ETF数量"] = result["当年新成立ETF数量"].astype(int)
    result["累计ETF数量"] = result.groupby(by_col)["当年新成立ETF数量"].cumsum()
    return result


def manager_tables(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    total_count = len(df)
    total_scale = pd.to_numeric(df["基金规模_亿元"], errors="coerce").sum()
    rows = []
    for manager, part in df.groupby("基金管理人", dropna=False):
        scale = pd.to_numeric(part["基金规模_亿元"], errors="coerce")
        rows.append(
            {
                "基金管理人": manager or "未提供",
                "ETF数量": len(part),
                "ETF总规模_亿元": round4(scale.sum()),
                "规模占比": safe_ratio(scale.sum(), total_scale),
                "平均单品规模_亿元": round4(scale.mean()),
                "规模中位数_亿元": round4(scale.median()),
                "最大单品规模_亿元": round4(scale.max()),
            }
        )
    ranking = pd.DataFrame(rows).sort_values(["ETF总规模_亿元", "ETF数量"], ascending=[False, False]).reset_index(drop=True)
    ranking.insert(0, "排名", range(1, len(ranking) + 1))

    cr_rows = []
    for n in [3, 5, 10, 20]:
        top = ranking.head(n)
        cr_rows.append(
            {
                "指标": f"CR{n}",
                "数量口径": safe_ratio(top["ETF数量"].sum(), total_count),
                "规模口径": safe_ratio(top["ETF总规模_亿元"].sum(), total_scale),
                "说明": f"按ETF总规模排名的前{n}大管理人合计占比",
            }
        )
    cr = pd.DataFrame(cr_rows)
    top20_managers = ranking.head(20)["基金管理人"].tolist()
    top_detail = df[df["基金管理人"].isin(top20_managers)][
        ["基金管理人", "基金代码", "基金简称", "资产类型", "产品类型", "基金规模_亿元", "成立日期", "跟踪指数"]
    ].sort_values(["基金管理人", "基金规模_亿元"], ascending=[True, False])
    return ranking, cr, top_detail


def size_interval_tables(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    temp = df.copy()
    temp["基金规模_亿元"] = pd.to_numeric(temp["基金规模_亿元"], errors="coerce")
    labels = ["<1亿", "1–2亿", "2–10亿", "10–50亿", "50–200亿", "200亿以上"]
    temp["规模区间"] = pd.cut(temp["基金规模_亿元"], [-math.inf, 1, 2, 10, 50, 200, math.inf], labels=labels, right=False)
    valid = temp.dropna(subset=["规模区间"]).copy()
    total_count = len(valid)
    total_scale = valid["基金规模_亿元"].sum()

    full_rows = []
    for label in labels:
        part = valid[valid["规模区间"].astype(str) == label]
        full_rows.append(
            {
                "规模区间": label,
                "ETF数量": len(part),
                "数量占比": safe_ratio(len(part), total_count),
                "总规模_亿元": round4(part["基金规模_亿元"].sum()),
                "规模占比": safe_ratio(part["基金规模_亿元"].sum(), total_scale),
            }
        )
    full = pd.DataFrame(full_rows)

    product_rows = []
    for product_type, part0 in valid.groupby("产品类型"):
        type_count = len(part0)
        type_scale = part0["基金规模_亿元"].sum()
        for label in labels:
            part = part0[part0["规模区间"].astype(str) == label]
            product_rows.append(
                {
                    "产品类型": product_type,
                    "规模区间": label,
                    "ETF数量": len(part),
                    "数量占比": safe_ratio(len(part), type_count),
                    "总规模_亿元": round4(part["基金规模_亿元"].sum()),
                    "规模占比": safe_ratio(part["基金规模_亿元"].sum(), type_scale),
                }
            )
    by_product = pd.DataFrame(product_rows)

    manager_rows = []
    for manager, part in temp.groupby("基金管理人", dropna=False):
        scale = part["基金规模_亿元"]
        total = len(part)
        manager_rows.append(
            {
                "基金管理人": manager or "未提供",
                "ETF总数量": total,
                "ETF总规模_亿元": round4(scale.sum()),
                "规模小于1亿数量": int((scale < 1).sum()),
                "规模小于2亿数量": int((scale < 2).sum()),
                "规模小于10亿数量": int((scale < 10).sum()),
                "规模小于1亿占比": safe_ratio(int((scale < 1).sum()), total),
                "规模小于2亿占比": safe_ratio(int((scale < 2).sum()), total),
                "规模小于10亿占比": safe_ratio(int((scale < 10).sum()), total),
            }
        )
    by_manager = pd.DataFrame(manager_rows).sort_values(["规模小于10亿占比", "ETF总数量"], ascending=[False, False])
    small_detail = valid[valid["基金规模_亿元"] < 10][
        ["基金代码", "基金简称", "基金管理人", "基金规模_亿元", "规模区间", "资产类型", "产品类型", "成立日期", "跟踪指数"]
    ].sort_values("基金规模_亿元")
    return full, by_product, by_manager, small_detail


def note_df(lines: list[str]) -> pd.DataFrame:
    return pd.DataFrame({"序号": range(1, len(lines) + 1), "说明": lines})


def dataframe_to_markdown(df: pd.DataFrame) -> str:
    """Small dependency-free Markdown table writer for the processing note."""
    if df.empty:
        return "（无记录）"
    columns = [str(col) for col in df.columns]
    rows = []
    for _, row in df.iterrows():
        rows.append([clean_text(row.get(col)).replace("|", "\\|") for col in df.columns])
    header = "| " + " | ".join(columns) + " |"
    separator = "| " + " | ".join(["---"] * len(columns)) + " |"
    body = ["| " + " | ".join(row) + " |" for row in rows]
    return "\n".join([header, separator, *body])


def write_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    if path.exists():
        print(f"覆盖已有输出：{path}")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, data in sheets.items():
            data.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    style_workbook(path)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Microsoft YaHei", size=10)
    thin = Side(style="thin", color="D9E1F2")
    border = Border(bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            values = [clean_text(ws.cell(row, col_idx).value) for row in range(1, min(ws.max_row, 300) + 1)]
            width = min(max(max((len(v) for v in values), default=8) * 1.25 + 2, 10), 38)
            ws.column_dimensions[letter].width = width
    wb.save(path)


def build_markdown(
    raw_sheets: dict[str, pd.DataFrame],
    field_mapping: pd.DataFrame,
    etf_pool: pd.DataFrame,
    anomalies: dict[str, pd.DataFrame],
    p21_overview: pd.DataFrame,
) -> str:
    total = len(etf_pool)
    total_scale = pd.to_numeric(etf_pool["基金规模_亿元"], errors="coerce").sum()
    top_asset = group_structure(etf_pool, "资产类型").head(3)
    top_product = group_structure(etf_pool, "产品类型").head(3)
    small_lt10 = safe_ratio((pd.to_numeric(etf_pool["基金规模_亿元"], errors="coerce") < 10).sum(), total)
    mapping_md = dataframe_to_markdown(field_mapping)
    sheet_lines = "\n".join(f"- {name}: {len(df)} 行，{len(df.columns)} 列" for name, df in raw_sheets.items())
    file_lines = "\n".join(f"- P{k}: {v}" for k, v in OUTPUT_FILES.items() if k.startswith("P"))
    asset_text = "；".join(f"{r['资产类型']} {r['ETF数量']}只、规模{round4(r['总规模_亿元'])}亿元" for _, r in top_asset.iterrows())
    product_text = "；".join(f"{r['产品类型']} {r['ETF数量']}只、规模{round4(r['总规模_亿元'])}亿元" for _, r in top_product.iterrows())
    missing_key = anomalies["missing"].sort_values("缺失数量", ascending=False).head(8)
    missing_text = "；".join(f"{r['字段名']}缺失{r['缺失数量']}" for _, r in missing_key.iterrows())
    return f"""# ICI课题 P17-P23 中国ETF数据处理说明

## 1. 本次任务目的
- 基于本地全市场 ETF 基础信息，生成 P17-P23 可直接作图的数据底表。
- 所有输出均为新文件，未修改原始文件。

## 2. 输入与输出路径
- 输入文件：`{INPUT_PATH}`
- 输出文件夹：`{OUTPUT_DIR}`

## 3. 原始数据 sheet
{sheet_lines}

## 4. 字段标准化映射
{mapping_md}

## 5. ETF筛选口径
- 优先使用基金简称、基金全称、基金类型、投资类型是否包含 ETF 判断。
- 排除 ETF联接、ETF连接、联接A、联接C、场外联接、FOF、LOF、指数增强非ETF等非ETF产品。
- 本次标准化后 ETF 产品池数量：{total} 只。

## 6. ETF分类规则
- 资产类型按跨境ETF、债券ETF、商品ETF、货币ETF、股票ETF、其他ETF识别。
- 产品类型按跨境ETF > 债券ETF > 商品ETF > 货币ETF > 策略ETF > 行业主题ETF > 宽基ETF > 其他ETF 的优先级识别。
- 每只产品保留“命中关键词”，便于后续人工复核。

## 7. 规模单位换算逻辑
- 原始字段为 `最新基金规模(亿)`，已明确为“亿”口径，因此直接转换为数值型 `基金规模_亿元`。
- 若后续原始字段出现“万元”或“元”，脚本已预留自动换算为亿元的逻辑。

## 8. 缺失值与异常处理
- 缺失字段不编造，保留空列或写入异常/缺失检查。
- 成立日期缺失或无法解析的产品不进入 P17 年度统计。
- 规模缺失、无法解析或小于等于 0 的产品进入规模异常检查。
- 主要缺失概况：{missing_text}。

## 9. 输出文件对应 PPT 页
{file_lines}
- 汇总文件：{OUTPUT_FILES["ARCHIVE"]}

## 10. 建议后续人工复核事项
- “有色”“能源”等关键词可能在商品ETF与行业主题ETF之间存在边界，需要人工复核。
- 策略ETF与宽基/行业主题ETF的重叠产品，建议结合指数编制规则二次确认。
- 若需要费率竞争分析的正式结论，建议补充历史费率和同类产品费率变动数据。

## 11. P17-P23 初步描述性结论
- 中国 ETF 产品池共 {total} 只，总规模约 {round4(total_scale)} 亿元，已具备较大的供给基础。
- 资产类型结构中，规模靠前类别为：{asset_text}。
- 产品类型结构中，规模靠前类别为：{product_text}。
- 小于 10 亿元产品占比约 {round4(small_lt10)}，长尾产品运营压力需要关注。
- 管理人规模集中度、长尾占比和产品类型分化共同支持“从产品扩张走向生态竞争”的判断。
- 美国侧数据不在本地表中计算，汇报时可单独引用 ICI 2026 Fact Book：ETF 数量 4,495 只，ETF 总净资产 13.4 万亿美元。
"""


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"输入文件不存在：{INPUT_PATH}")

    raw_sheets = pd.read_excel(INPUT_PATH, sheet_name=None, dtype=object)
    main_sheet, raw_df = choose_main_sheet(raw_sheets)
    etf_all, field_mapping = standardize_dataframe(raw_df, main_sheet)
    etf_pool = etf_all[etf_all["是否ETF"].eq("是")].copy().reset_index(drop=True)
    etf_pool["基金规模_亿元"] = pd.to_numeric(etf_pool["基金规模_亿元"], errors="coerce")
    etf_pool["最新规模_亿元"] = pd.to_numeric(etf_pool["最新规模_亿元"], errors="coerce")

    anomalies = build_anomaly_tables(etf_pool)

    sheet_overview = pd.DataFrame(
        [
            {"sheet名称": name, "行数": len(df), "列数": len(df.columns), "是否读取成功": "是", "备注": ""}
            for name, df in raw_sheets.items()
        ]
    )
    desc_rows = [
        {"指标": "ETF总数量", "数值": len(etf_pool), "单位": "只", "说明": ""},
        {"指标": "ETF总规模_亿元", "数值": round4(etf_pool["基金规模_亿元"].sum()), "单位": "亿元", "说明": ""},
        {"指标": "平均规模_亿元", "数值": round4(etf_pool["基金规模_亿元"].mean()), "单位": "亿元", "说明": ""},
        {"指标": "规模中位数_亿元", "数值": round4(etf_pool["基金规模_亿元"].median()), "单位": "亿元", "说明": ""},
        {"指标": "最大规模_亿元", "数值": round4(etf_pool["基金规模_亿元"].max()), "单位": "亿元", "说明": ""},
        {"指标": "最小规模_亿元", "数值": round4(etf_pool["基金规模_亿元"].min()), "单位": "亿元", "说明": ""},
        {
            "指标": "成立日期最早年份",
            "数值": pd.to_datetime(etf_pool["成立日期"], errors="coerce").dt.year.min(),
            "单位": "年",
            "说明": "",
        },
        {
            "指标": "成立日期最新年份",
            "数值": pd.to_datetime(etf_pool["成立日期"], errors="coerce").dt.year.max(),
            "单位": "年",
            "说明": "",
        },
        {"指标": "管理人数量", "数值": etf_pool["基金管理人"].nunique(), "单位": "家", "说明": ""},
        {
            "指标": "资产类型数量分布",
            "数值": "；".join(f"{k}:{v}" for k, v in etf_pool["资产类型"].value_counts().items()),
            "单位": "只",
            "说明": "",
        },
        {
            "指标": "产品类型数量分布",
            "数值": "；".join(f"{k}:{v}" for k, v in etf_pool["产品类型"].value_counts().items()),
            "单位": "只",
            "说明": "",
        },
    ]
    desc = pd.DataFrame(desc_rows)

    p17_all = annual_table(etf_pool)
    p17_asset = annual_table(etf_pool, "资产类型")
    p17_product = annual_table(etf_pool, "产品类型")

    p18_asset = group_structure(etf_pool, "资产类型")
    p18_product = group_structure(etf_pool, "产品类型")
    p18_detail = etf_pool[
        ["基金代码", "基金简称", "基金管理人", "资产类型", "产品类型", "基金规模_亿元", "成立日期", "跟踪指数", "命中关键词"]
    ].copy()

    p19_rank, p19_cr, p19_top_detail = manager_tables(etf_pool)
    p20_full, p20_product, p20_manager, p20_small_detail = size_interval_tables(etf_pool)

    total_scale = etf_pool["基金规模_亿元"].sum()
    cr_lookup = p19_cr.set_index("指标")
    overview_rows = []
    for asset_type in ["股票ETF", "债券ETF", "商品ETF", "货币ETF", "跨境ETF"]:
        part = etf_pool[etf_pool["资产类型"].eq(asset_type)]
        overview_rows.append({"指标": f"{asset_type}数量", "数值": len(part), "单位": "只", "说明": "按资产类型分类"})
        overview_rows.append({"指标": f"{asset_type}规模", "数值": round4(part["基金规模_亿元"].sum()), "单位": "亿元", "说明": "按资产类型分类"})
    for product_type in ["宽基ETF", "行业主题ETF", "策略ETF"]:
        part = etf_pool[etf_pool["产品类型"].eq(product_type)]
        overview_rows.append({"指标": f"{product_type}数量", "数值": len(part), "单位": "只", "说明": "按产品类型分类"})
        overview_rows.append({"指标": f"{product_type}规模", "数值": round4(part["基金规模_亿元"].sum()), "单位": "亿元", "说明": "按产品类型分类"})
    overview_rows = [
        {"指标": "ETF总数量", "数值": len(etf_pool), "单位": "只", "说明": "ETF产品池数量"},
        {"指标": "ETF总规模", "数值": round4(total_scale), "单位": "亿元", "说明": "基金规模_亿元合计"},
        *overview_rows,
        {"指标": "前五大管理人规模占比", "数值": round4(cr_lookup.loc["CR5", "规模口径"]), "单位": "%", "说明": "按管理人规模排名"},
        {"指标": "前十大管理人规模占比", "数值": round4(cr_lookup.loc["CR10", "规模口径"]), "单位": "%", "说明": "按管理人规模排名"},
        {
            "指标": "规模小于1亿产品占比",
            "数值": round4(safe_ratio((etf_pool["基金规模_亿元"] < 1).sum(), len(etf_pool))),
            "单位": "%",
            "说明": "按产品数量",
        },
        {
            "指标": "规模小于2亿产品占比",
            "数值": round4(safe_ratio((etf_pool["基金规模_亿元"] < 2).sum(), len(etf_pool))),
            "单位": "%",
            "说明": "按产品数量",
        },
        {
            "指标": "规模小于10亿产品占比",
            "数值": round4(safe_ratio((etf_pool["基金规模_亿元"] < 10).sum(), len(etf_pool))),
            "单位": "%",
            "说明": "按产品数量",
        },
    ]
    p21_overview = pd.DataFrame(overview_rows)
    p21_compare = pd.DataFrame(
        [
            {"维度": "市场阶段", "中国ETF市场数据": f"{len(etf_pool)}只，约{round4(total_scale)}亿元", "备注": "本地数据口径"},
            {"维度": "ETF数量", "中国ETF市场数据": len(etf_pool), "备注": "美国侧：ICI 2026 Fact Book 4,495只"},
            {"维度": "ETF总规模", "中国ETF市场数据": round4(total_scale), "备注": "单位：亿元；美国侧：13.4万亿美元"},
            {"维度": "资产结构", "中国ETF市场数据": p18_asset.iloc[0]["资产类型"] if len(p18_asset) else "", "备注": "以规模最大资产类型描述"},
            {"维度": "管理人集中度", "中国ETF市场数据": round4(cr_lookup.loc["CR10", "规模口径"]), "备注": "前十大管理人规模占比"},
            {
                "维度": "小规模产品占比",
                "中国ETF市场数据": round4(safe_ratio((etf_pool["基金规模_亿元"] < 10).sum(), len(etf_pool))),
                "备注": "规模小于10亿元产品数量占比",
            },
            {"维度": "主导产品类型", "中国ETF市场数据": p18_product.iloc[0]["产品类型"] if len(p18_product) else "", "备注": "以规模最大产品类型描述"},
            {"维度": "后续竞争重点", "中国ETF市场数据": "持营能力、场景化配置、生态服务能力", "备注": "基于长尾与集中度指标"},
        ]
    )

    multi_asset = etf_pool[etf_pool["资产类型"].isin(["跨境ETF", "债券ETF", "商品ETF", "货币ETF"])]
    fee_available = etf_pool["管理费率"].notna().any()
    avg_fee = pd.to_numeric(etf_pool["管理费率"], errors="coerce").mean()
    p22 = pd.DataFrame(
        [
            {
                "转向": "发行导向 → 持营导向",
                "支撑指标": "规模小于10亿产品占比",
                "数值": round4(safe_ratio((etf_pool["基金规模_亿元"] < 10).sum(), len(etf_pool))),
                "单位": "%",
                "解释": "长尾产品占比较高时，后续重点从发行转向持续运营。",
            },
            {
                "转向": "指数覆盖 → 场景覆盖",
                "支撑指标": "行业主题+策略ETF数量占比",
                "数值": round4(safe_ratio(etf_pool["产品类型"].isin(["行业主题ETF", "策略ETF"]).sum(), len(etf_pool))),
                "单位": "%",
                "解释": "产品从单纯指数覆盖走向投资场景覆盖。",
            },
            {
                "转向": "费率竞争 → 生态竞争",
                "支撑指标": "平均管理费率" if fee_available else "原始数据缺少费率字段",
                "数值": round4(avg_fee) if fee_available else "",
                "单位": "%" if fee_available else "",
                "解释": "低费率只是基础，长尾与头部集中说明生态能力更关键。",
            },
            {
                "转向": "单品销售 → 组合解决方案",
                "支撑指标": "多资产ETF规模占比",
                "数值": round4(safe_ratio(multi_asset["基金规模_亿元"].sum(), total_scale)),
                "单位": "%",
                "解释": "跨境/债券/商品/货币ETF为组合配置工具箱提供支撑。",
            },
        ]
    )
    p23 = pd.DataFrame(
        [
            {"指标": "ETF总数量", "数值": len(etf_pool), "单位": "只", "可用于哪一句结论": "用于说明中国市场供给扩张"},
            {"指标": "ETF总规模", "数值": round4(total_scale), "单位": "亿元", "可用于哪一句结论": "用于说明中国市场已有基础"},
            {
                "指标": "前十大管理人规模占比",
                "数值": round4(cr_lookup.loc["CR10", "规模口径"]),
                "单位": "%",
                "可用于哪一句结论": "用于说明头部集中",
            },
            {
                "指标": "小规模产品占比",
                "数值": round4(safe_ratio((etf_pool["基金规模_亿元"] < 10).sum(), len(etf_pool))),
                "单位": "%",
                "可用于哪一句结论": "用于说明持营压力",
            },
            {
                "指标": "股票ETF规模占比",
                "数值": round4(safe_ratio(etf_pool.loc[etf_pool["资产类型"].eq("股票ETF"), "基金规模_亿元"].sum(), total_scale)),
                "单位": "%",
                "可用于哪一句结论": "用于说明权益属性较强",
            },
            {
                "指标": "多资产ETF规模占比",
                "数值": round4(safe_ratio(multi_asset["基金规模_亿元"].sum(), total_scale)),
                "单位": "%",
                "可用于哪一句结论": "用于说明配置工具箱仍待完善",
            },
        ]
    )

    common_note = note_df(
        [
            "所有规模字段统一为亿元口径。",
            "所有日期字段统一为YYYY-MM-DD格式。",
            "基金代码按文本保存，保留前导0。",
            "分类结果由关键词规则自动生成，建议结合指数规则人工复核。",
        ]
    )
    p17_note = note_df(["年份来自成立日期；成立日期缺失产品不进入年度统计。", "柱状图建议使用当年新成立ETF数量，折线图使用累计ETF数量。"])
    p18_note = note_df(["结构表按资产类型/产品类型汇总ETF数量与规模。", "规模占比以可解析基金规模_亿元合计为分母。"])
    p19_note = note_df(["管理人排名按ETF总规模_亿元降序。", "CR指标按规模排名前N大管理人计算。"])
    p20_note = note_df(["规模区间按基金规模_亿元分箱。", "小规模占比按管理人旗下产品数量计算。"])
    p21_note = note_df(["美国侧数据不在本地文件中计算，仅在说明中提示 ICI 2026 Fact Book 口径。"])
    p22_note = note_df(["结论页摘要仅使用本地可支撑指标；不编造外部数据。"])

    write_excel(
        OUTPUT_DIR / OUTPUT_FILES["00"],
        {
            "00_sheet概览": sheet_overview,
            "01_字段映射表": field_mapping,
            "02_缺失值检查": anomalies["missing"],
            "03_重复值检查": anomalies["duplicates"],
            "04_日期异常检查": anomalies["date_bad"],
            "05_规模异常检查": anomalies["scale_bad"],
            "06_描述性统计": desc,
            "ETF产品池_清洗底表": etf_pool,
            "口径说明": common_note,
        },
    )
    write_excel(
        OUTPUT_DIR / OUTPUT_FILES["P17"],
        {
            "P17_年度新发累计_全市场": p17_all,
            "P17_年度新发累计_按资产类型": p17_asset,
            "P17_年度新发累计_按产品类型": p17_product,
            "P17_作图建议": note_df(["图1：柱状图=当年新成立ETF数量，折线图=累计ETF数量。", "图2：可选堆积柱状图=按资产类型拆分年度新发数量。"]),
            "口径说明": p17_note,
            "异常缺失检查": anomalies["date_bad"],
        },
    )
    write_excel(
        OUTPUT_DIR / OUTPUT_FILES["P18"],
        {
            "P18_资产类型结构": p18_asset,
            "P18_产品类型结构": p18_product,
            "P18_资产类型明细": p18_detail,
            "P18_作图建议": note_df(["左图：按资产类型拆分规模，可用环形图或横向堆积条形图。", "右图：按产品类型拆分数量，可用柱状图或横向条形图。"]),
            "口径说明": p18_note,
            "异常缺失检查": anomalies["scale_bad"],
        },
    )
    write_excel(
        OUTPUT_DIR / OUTPUT_FILES["P19"],
        {
            "P19_管理人规模排名": p19_rank,
            "P19_前十大管理人": p19_rank.head(10),
            "P19_CR集中度": p19_cr,
            "P19_头部管理人明细": p19_top_detail,
            "P19_作图建议": note_df(["左图：前十大ETF管理人规模排名柱状图。", "右图：CR5/CR10/CR20集中度柱状图或指标卡。"]),
            "口径说明": p19_note,
            "异常缺失检查": anomalies["scale_bad"],
        },
    )
    write_excel(
        OUTPUT_DIR / OUTPUT_FILES["P20"],
        {
            "P20_规模区间分布_全市场": p20_full,
            "P20_规模区间分布_按产品类型": p20_product,
            "P20_小规模产品占比_按管理人": p20_manager,
            "P20_小规模产品明细": p20_small_detail,
            "P20_作图建议": note_df(["图1：全市场ETF规模区间分布柱状图。", "图2：主要管理人小规模产品占比对比图。", "可重点关注 <1亿、<2亿、<10亿三个口径。"]),
            "口径说明": p20_note,
            "异常缺失检查": anomalies["scale_bad"],
        },
    )
    write_excel(
        OUTPUT_DIR / OUTPUT_FILES["P21"],
        {
            "P21_中国ETF总览指标": p21_overview,
            "P21_中美对比表_中国侧": p21_compare,
            "口径说明": p21_note,
            "异常缺失检查": anomalies["generic"],
        },
    )
    write_excel(
        OUTPUT_DIR / OUTPUT_FILES["P22P23"],
        {
            "P22_四个转向支撑数据": p22,
            "P23_最终升华可引用数据": p23,
            "口径说明": p22_note,
            "异常缺失检查": anomalies["generic"],
        },
    )
    write_excel(
        OUTPUT_DIR / OUTPUT_FILES["ARCHIVE"],
        {
            "P17_年度新发累计": p17_all,
            "P17_按资产类型年度新发": p17_asset,
            "P18_资产类型结构": p18_asset,
            "P18_产品类型结构": p18_product,
            "P19_管理人规模排名": p19_rank,
            "P19_CR集中度": p19_cr,
            "P20_规模区间分布": p20_full,
            "P20_小规模产品占比": p20_manager,
            "P21_中国ETF总览指标": p21_overview,
            "P22_四个转向支撑数据": p22,
            "P23_最终升华引用数据": p23,
        },
    )

    readme = build_markdown(raw_sheets, field_mapping, etf_pool, anomalies, p21_overview)
    readme_path = OUTPUT_DIR / OUTPUT_FILES["README"]
    if readme_path.exists():
        print(f"覆盖已有输出：{readme_path}")
    readme_path.write_text(readme, encoding="utf-8")

    generated = [OUTPUT_DIR / name for name in OUTPUT_FILES.values()]
    print("输入文件是否成功读取：是")
    print(f"输出文件夹路径：{OUTPUT_DIR}")
    print("生成文件：")
    for path in generated:
        print(f"- {path.name}")
    print(f"ETF产品池数量：{len(etf_pool)}")
    print(f"ETF总规模：{round4(total_scale)} 亿元")
    print("P17-P23每页对应的数据文件名：")
    for key in ["P17", "P18", "P19", "P20", "P21", "P22P23"]:
        print(f"- {key}: {OUTPUT_FILES[key]}")


if __name__ == "__main__":
    main()
