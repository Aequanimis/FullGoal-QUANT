from __future__ import annotations

import hashlib
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据")
ARCHIVE = BASE / "阶段性归档_境内广义策略ETF数据底座"
README_DIR = ARCHIVE / "99_归档说明"
README_MD = README_DIR / "阶段性归档说明.md"
INDEX_XLSX = README_DIR / "阶段性归档索引表.xlsx"

MODULES = [
    {
        "序号": 1,
        "模块名称": "产品池与分类口径",
        "源文件": BASE / "全市场ETF基础信息_策略ETF池二次修正版.xlsx",
        "目标目录": ARCHIVE / "01_产品池与分类口径",
        "目标文件名": "01_广义策略ETF_产品池与分类口径_主文件.xlsx",
        "文件定位": "定义223只广义策略ETF、168只核心策略指数ETF、55只指数增强/多因子ETF，并保留策略分类、市场范围和核心/广义统计标记。",
        "正式分析使用sheet": "策略ETF_最终统计池",
        "是否核心产出": "是",
        "注意事项": "产品筛选、分类匹配和后续数据关联均以Wind代码及本文件分类字段为准。",
    },
    {
        "序号": 2,
        "模块名称": "Wind口径 vs 自建口径差异",
        "源文件": BASE / "Wind策略类ETF口径_vs_自建策略ETF口径_差异分析.xlsx",
        "目标目录": ARCHIVE / "02_Wind口径_vs_自建口径差异",
        "目标文件名": "02_Wind口径_vs_自建策略ETF口径_差异分析.xlsx",
        "文件定位": "解释Wind策略类ETF、自建核心策略ETF和自建广义策略ETF数量不一致的原因，并形成研究口径建议。",
        "正式分析使用sheet": "口径对比汇总；策略分类差异分析；研究范围建议",
        "是否核心产出": "是",
        "注意事项": "用于口径解释与交叉验证，不替代产品池主文件。",
    },
    {
        "序号": 3,
        "模块名称": "月度规模份额",
        "源文件": BASE / "广义策略ETF_月度规模份额表_上市后分析版.xlsx",
        "目标目录": ARCHIVE / "03_月度规模份额",
        "目标文件名": "03_广义策略ETF_月度规模份额_上市后分析版.xlsx",
        "文件定位": "用于研究月度规模、份额、规模增长、策略结构变化、发行节奏、基金公司布局和产品规模分布。",
        "正式分析使用sheet": "月度规模份额_上市后分析版；各上市后月度汇总sheet",
        "是否核心产出": "是",
        "注意事项": "使用上市后有效数据；不要把上市前记录纳入总量统计。",
    },
    {
        "序号": 4,
        "模块名称": "月度交易流动性",
        "源文件": BASE / "wind代码池" / "广义策略ETF月度流动性数据_合并验收清洗版_v2.xlsx",
        "目标目录": ARCHIVE / "04_月度交易流动性",
        "目标文件名": "04_广义策略ETF_月度交易流动性_合并验收清洗版_v2.xlsx",
        "文件定位": "用于分析月度成交额、日均成交额、成交量、换手率、折溢价率、月振幅和交易活跃度。",
        "正式分析使用sheet": "流动性数据_分析可用版",
        "是否核心产出": "是",
        "注意事项": "补充数据的固定收盘价不用于收益风险分析；收益风险应使用后续单独导出的净值或收益率数据。",
    },
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=9, color="1F1F1F")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN_GRAY = Side(style="thin", color="D9E1F2")
BODY_BORDER = Border(bottom=THIN_GRAY)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_records(ws, headers: list[str], records: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for record in records:
        ws.append([record.get(header) for header in headers])


def style_sheet(ws, headers: list[str]) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BODY_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
    sample_limit = min(ws.max_row, 200)
    wide = {
        "归档后文件路径",
        "原始文件路径",
        "文件定位",
        "正式分析使用sheet",
        "注意事项",
        "关键字段",
        "用途",
        "备注",
    }
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, sample_limit + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 80))
        width = min(max(max_len * 1.1, 12), 65 if header in wide else 32)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def make_markdown(archive_time: str, module_rows: list[dict[str, Any]]) -> str:
    table_lines = [
        "| 序号 | 文件模块 | 归档后文件路径 | 原始来源路径 | 文件定位 | 正式分析建议使用的 sheet | 注意事项 |",
        "|---:|---|---|---|---|---|---|",
    ]
    for row in module_rows:
        table_lines.append(
            "| {序号} | {模块名称} | `{归档后文件路径}` | `{原始文件路径}` | {文件定位} | {正式分析使用sheet} | {注意事项} |".format(
                **row
            )
        )

    return f"""# 境内广义策略ETF数据底座阶段性归档说明

## 1. 归档时间

{archive_time}

## 2. 当前研究阶段说明

当前阶段已经完成境内广义策略 ETF 的产品池构建、Wind 口径与自建口径差异分析、月度规模份额数据清洗，以及月度交易流动性数据合并验收。当前数据底座可以支持境内广义策略 ETF 的产品数量、规模结构、发行节奏、策略分类、基金公司布局和交易流动性等总量分析。

## 3. 当前产品池口径

- 广义策略 ETF：223 只；
- 核心策略指数 ETF：168 只；
- 指数增强/多因子 ETF：55 只；
- 主分析建议使用核心策略指数 ETF 口径；
- 广义策略 ETF 口径用于补充观察指数增强/多因子 ETF。

## 4. 四个主文件说明

{chr(10).join(table_lines)}

## 5. 正式分析建议

- 产品池和分类口径以“01_广义策略ETF_产品池与分类口径_主文件.xlsx”为准；
- 规模份额分析以“03_广义策略ETF_月度规模份额_上市后分析版.xlsx”为准；
- 流动性分析以“04_广义策略ETF_月度交易流动性_合并验收清洗版_v2.xlsx”中的“流动性数据_分析可用版”为准；
- Wind口径差异解释以“02_Wind口径_vs_自建策略ETF口径_差异分析.xlsx”为准。

## 6. 已完成数据模块

- 产品池代码表；
- 策略分类口径；
- Wind口径与自建口径差异；
- 月度规模份额；
- 月度交易流动性。

## 7. 后续待补充数据模块

- 收益风险表现表；
- 持有人结构表；
- 跟踪指数表现估值表；
- 跟踪指数规则表。

## 8. 重要注意事项

- 月度流动性数据可以用于成交额、成交量、换手率、折溢价率、月振幅分析；
- 月度流动性数据中的补充收盘价不建议用于收益风险分析；
- 收益风险分析应以后续单独导出的净值、复权净值或收益率数据为准；
- 规模份额分析应使用上市后有效数据，避免将上市前空值纳入统计；
- 后续所有新增数据都应通过 Wind 代码与产品池主文件匹配分类字段。

## 9. 归档完整性

本次归档采用复制方式，不删除或覆盖原始文件。文件复制状态、文件大小及 SHA256 校验值记录在“阶段性归档索引表.xlsx”的“归档文件索引”sheet中。
"""


def main() -> None:
    archive_time = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %z")
    subdirs = [
        ARCHIVE / "01_产品池与分类口径",
        ARCHIVE / "02_Wind口径_vs_自建口径差异",
        ARCHIVE / "03_月度规模份额",
        ARCHIVE / "04_月度交易流动性",
        README_DIR,
    ]
    for directory in subdirs:
        directory.mkdir(parents=True, exist_ok=True)

    index_rows = []
    missing_paths = []
    for module in MODULES:
        source: Path = module["源文件"]
        target = module["目标目录"] / module["目标文件名"]
        exists = source.exists()
        copied = False
        error = ""
        if exists:
            try:
                shutil.copy2(source, target)
                copied = target.exists() and target.stat().st_size > 0
            except Exception as exc:
                error = str(exc)
        else:
            missing_paths.append(str(source))
            error = "源文件缺失"

        index_rows.append(
            {
                "序号": module["序号"],
                "模块名称": module["模块名称"],
                "归档后文件名": module["目标文件名"],
                "归档后文件路径": str(target),
                "原始文件路径": str(source),
                "文件定位": module["文件定位"],
                "正式分析使用sheet": module["正式分析使用sheet"],
                "是否核心产出": module["是否核心产出"],
                "注意事项": module["注意事项"],
                "源文件是否存在": "是" if exists else "否",
                "是否成功复制": "是" if copied else "否",
                "归档文件大小_字节": target.stat().st_size if copied else None,
                "SHA256": sha256(target) if copied else None,
                "校验备注": error,
            }
        )

    README_MD.write_text(
        make_markdown(archive_time, index_rows),
        encoding="utf-8-sig",
        newline="\n",
    )

    completion_rows = [
        {
            "数据模块": "产品池与分类口径",
            "当前状态": "已完成",
            "对应文件": MODULES[0]["目标文件名"],
            "主要用途": "定义产品范围、核心/广义口径和策略分类字段",
            "后续是否还需补充": "否",
            "备注": "后续新增数据须按Wind代码关联本文件",
        },
        {
            "数据模块": "Wind口径与自建口径差异",
            "当前状态": "已完成",
            "对应文件": MODULES[1]["目标文件名"],
            "主要用途": "解释研究范围口径差异并提供交叉验证",
            "后续是否还需补充": "视产品池更新",
            "备注": "产品池大幅更新时建议重跑",
        },
        {
            "数据模块": "月度规模份额",
            "当前状态": "已完成",
            "对应文件": MODULES[2]["目标文件名"],
            "主要用途": "规模、份额、发行节奏与结构分析",
            "后续是否还需补充": "需定期更新",
            "备注": "使用上市后分析口径",
        },
        {
            "数据模块": "月度交易流动性",
            "当前状态": "已完成",
            "对应文件": MODULES[3]["目标文件名"],
            "主要用途": "成交额、成交量、换手率、折溢价率与振幅分析",
            "后续是否还需补充": "需定期更新",
            "备注": "正式使用流动性数据_分析可用版",
        },
        {
            "数据模块": "收益风险表现",
            "当前状态": "待收集",
            "对应文件": "广义策略ETF_收益风险表现表.xlsx",
            "主要用途": "收益、波动、回撤与风险调整收益分析",
            "后续是否还需补充": "是",
            "备注": "应使用净值、复权净值或收益率数据",
        },
        {
            "数据模块": "持有人结构",
            "当前状态": "待收集",
            "对应文件": "广义策略ETF_持有人结构表.xlsx",
            "主要用途": "机构/个人持有结构和集中度分析",
            "后续是否还需补充": "是",
            "备注": "建议按半年报/年报频率",
        },
        {
            "数据模块": "跟踪指数表现估值",
            "当前状态": "待收集",
            "对应文件": "核心策略ETF_跟踪指数表现估值表.xlsx",
            "主要用途": "指数收益、估值和策略特征比较",
            "后续是否还需补充": "是",
            "备注": "优先覆盖核心策略指数ETF",
        },
        {
            "数据模块": "跟踪指数规则",
            "当前状态": "待收集",
            "对应文件": "核心策略ETF_跟踪指数规则表.xlsx",
            "主要用途": "指数选样、加权、调样和策略逻辑分析",
            "后续是否还需补充": "是",
            "备注": "以指数公司官方编制方案为准",
        },
    ]

    pool_rows = [
        {"指标": "广义策略ETF数量", "内容": 223},
        {"指标": "核心策略指数ETF数量", "内容": 168},
        {"指标": "指数增强/多因子ETF数量", "内容": 55},
        {"指标": "主分析口径", "内容": "核心策略指数ETF"},
        {"指标": "补充分析口径", "内容": "广义策略ETF"},
        {
            "指标": "分类字段",
            "内容": "统计口径分类、一级策略大类、二级策略类别、市场范围_二次修正、是否纳入核心策略ETF统计、是否纳入广义策略ETF统计",
        },
    ]

    future_rows = [
        {
            "后续数据模块": "收益风险表现",
            "建议文件名": "广义策略ETF_收益风险表现表.xlsx",
            "数据范围": "223只广义策略ETF，核心168只重点分析",
            "建议频率": "月度",
            "关键字段": "复权净值、月收益率、年化收益、年化波动、最大回撤、夏普比率、跟踪误差",
            "用途": "收益风险表现、策略有效性和代表产品比较",
        },
        {
            "后续数据模块": "持有人结构",
            "建议文件名": "广义策略ETF_持有人结构表.xlsx",
            "数据范围": "223只广义策略ETF",
            "建议频率": "半年/年度",
            "关键字段": "机构持有比例、个人持有比例、前十大持有人、持有人户数、集中度",
            "用途": "投资者结构、机构化程度与产品稳定性分析",
        },
        {
            "后续数据模块": "跟踪指数表现估值",
            "建议文件名": "核心策略ETF_跟踪指数表现估值表.xlsx",
            "数据范围": "核心策略ETF对应指数，指数代码去重",
            "建议频率": "月度",
            "关键字段": "指数收益率、波动率、最大回撤、PE、PB、股息率、成分股数量",
            "用途": "策略指数表现、估值和市场环境适应性比较",
        },
        {
            "后续数据模块": "跟踪指数规则",
            "建议文件名": "核心策略ETF_跟踪指数规则表.xlsx",
            "数据范围": "核心策略ETF对应指数，指数代码去重",
            "建议频率": "规则变更时更新",
            "关键字段": "样本空间、选样因子、选样方法、成分数量、加权方式、调样频率、权重限制",
            "用途": "解释策略逻辑、分类依据及代表指数差异",
        },
    ]

    wb = Workbook()
    wb.remove(wb.active)
    index_ws = wb.create_sheet("归档文件索引")
    completion_ws = wb.create_sheet("当前数据完成情况")
    pool_ws = wb.create_sheet("产品池口径摘要")
    future_ws = wb.create_sheet("后续数据收集清单")

    index_headers = [
        "序号",
        "模块名称",
        "归档后文件名",
        "归档后文件路径",
        "原始文件路径",
        "文件定位",
        "正式分析使用sheet",
        "是否核心产出",
        "注意事项",
        "源文件是否存在",
        "是否成功复制",
        "归档文件大小_字节",
        "SHA256",
        "校验备注",
    ]
    write_records(index_ws, index_headers, index_rows)
    completion_headers = [
        "数据模块",
        "当前状态",
        "对应文件",
        "主要用途",
        "后续是否还需补充",
        "备注",
    ]
    write_records(completion_ws, completion_headers, completion_rows)
    pool_headers = ["指标", "内容"]
    write_records(pool_ws, pool_headers, pool_rows)
    future_headers = [
        "后续数据模块",
        "建议文件名",
        "数据范围",
        "建议频率",
        "关键字段",
        "用途",
    ]
    write_records(future_ws, future_headers, future_rows)

    for ws, headers in [
        (index_ws, index_headers),
        (completion_ws, completion_headers),
        (pool_ws, pool_headers),
        (future_ws, future_headers),
    ]:
        style_sheet(ws, headers)

    INDEX_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(INDEX_XLSX)

    print(f"1. 阶段性归档目录路径：{ARCHIVE}")
    for row in index_rows:
        print(f"2. 源文件存在[{row['模块名称']}]：{row['源文件是否存在']}")
    for row in index_rows:
        print(f"3. 成功复制[{row['模块名称']}]：{row['是否成功复制']}")
    print(f"4. Markdown说明文件路径：{README_MD}")
    print(f"5. Excel索引表路径：{INDEX_XLSX}")
    print("6. 当前完成的数据模块：产品池代码表、策略分类口径、Wind口径与自建口径差异、月度规模份额、月度交易流动性")
    print("7. 后续待补充的数据模块：收益风险表现表、持有人结构表、跟踪指数表现估值表、跟踪指数规则表")
    if missing_paths:
        for path in missing_paths:
            print(f"8. 缺失路径：{path}")
    else:
        print("8. 缺失文件：无")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"归档失败：{exc}", file=sys.stderr)
        raise
