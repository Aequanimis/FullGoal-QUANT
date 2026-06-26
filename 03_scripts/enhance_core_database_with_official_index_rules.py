from __future__ import annotations

import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = (
    Path.home()
    / "Desktop"
    / "Fullgoal"
    / "课题研究"
    / "处理后数据"
    / "阶段性归档_境内广义策略ETF数据底座"
)
CORE_SOURCE = ROOT / "境内广义策略ETF_核心数据库.xlsx"
OFFICIAL_SOURCE = ROOT / "82个核心策略指数_官网来源全量核验表.xlsx"
CORE_OUTPUT = ROOT / "境内广义策略ETF_核心数据库_指数规则官网核验增强版.xlsx"
DIFF_OUTPUT = ROOT / "指数规则_官网核验字段级差异表.xlsx"

CORE_RULE_SHEET = "10_指数规则明细"
MAPPING_SHEET = "09_ETF_指数映射"
QUALITY_SHEET = "12_数据质量总览"
OFFICIAL_MASTER_SHEET = "01_82指数官网核验总表"
OFFICIAL_EXTRACTED_SHEET = "02_已抽取字段可修正"
OFFICIAL_REVIEW_SHEET = "03_待逐字段复核清单"

SENTINELS = {"", "待官网核验", "待逐字段复核", "未抽取", "未明确", "wind暂无", "待补充", "n/a", "nan", "-", "--"}
KEY_FIELDS = ["单只成分权重上限", "行业权重限制", "调样频率"]
TRACKED_FIELDS = [
    "指数编制方案链接",
    "信息来源",
    "单只成分权重上限",
    "行业权重限制",
    "调样频率",
    "指数简介",
    "规则完整性状态",
    "是否已核验",
    "是否仍需官网核验",
    "仍缺失字段说明",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
HEADER_BORDER = Border(bottom=Side(style="medium", color="1F4E78"))


def normalize_code(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\u3000", " ").upper()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_valid_official_value(value: Any) -> bool:
    return clean_text(value).lower() not in SENTINELS


def is_missing_or_pending(value: Any) -> bool:
    return clean_text(value).lower() in SENTINELS


def load_table(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"未找到sheet：{path.name} / {sheet_name}")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header_raw = next(rows, None)
    if header_raw is None:
        return [], []
    headers = [clean_text(value) for value in header_raw]
    records: list[dict[str, Any]] = []
    for row in rows:
        if all(value is None or clean_text(value) == "" for value in row):
            continue
        records.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return headers, records


def last_nonempty_row(ws) -> int:
    last = 1
    for row in ws.iter_rows():
        if any(cell.value is not None and clean_text(cell.value) != "" for cell in row):
            last = row[0].row
    return last


def nonempty_row_count(ws) -> int:
    return max(0, last_nonempty_row(ws) - 1)


def set_column_widths(ws, headers: list[str], rows: list[list[Any]], max_width: int = 48) -> None:
    widths = [min(max(len(header) + 2, 10), max_width) for header in headers]
    for row in rows:
        for i, value in enumerate(row):
            if i >= len(widths):
                continue
            widths[i] = max(widths[i], min(max(len(clean_text(value)) + 2, 10), max_width))
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def style_table_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = BODY_ALIGNMENT
            if cell.column == 1:
                cell.number_format = "@"
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    set_column_widths(ws, headers, rows)


def replace_sheet_with_rows(wb, sheet_name: str, headers: list[str], rows: list[list[Any]], index: int | None = None):
    if sheet_name in wb.sheetnames:
        old = wb[sheet_name]
        target_index = wb.index(old) if index is None else index
        wb.remove(old)
    else:
        target_index = index if index is not None else len(wb.sheetnames)
    ws = wb.create_sheet(sheet_name, target_index)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_table_sheet(ws, headers, rows)
    return ws


def core_table_records(ws) -> tuple[list[str], list[dict[str, Any]]]:
    headers = [clean_text(cell.value) for cell in ws[1]]
    records: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if all(value is None or clean_text(value) == "" for value in values):
            continue
        records.append({headers[i]: values[i] if i < len(values) else None for i in range(len(headers))})
    return headers, records


def distinct_join(values: list[str]) -> str:
    seen: list[str] = []
    for value in values:
        text = clean_text(value)
        if text and text not in seen:
            seen.append(text)
    return "；".join(seen)


def main() -> None:
    if not CORE_SOURCE.exists():
        raise FileNotFoundError(f"核心数据库文件不存在：{CORE_SOURCE}")
    if not OFFICIAL_SOURCE.exists():
        raise FileNotFoundError(f"官网核验文件不存在：{OFFICIAL_SOURCE}")

    official_headers, official_records = load_table(OFFICIAL_SOURCE, OFFICIAL_MASTER_SHEET)
    extracted_headers, extracted_records = load_table(OFFICIAL_SOURCE, OFFICIAL_EXTRACTED_SHEET)
    review_headers, review_records = load_table(OFFICIAL_SOURCE, OFFICIAL_REVIEW_SHEET)
    official_by_code = {
        normalize_code(record.get("指数代码")): record
        for record in official_records
        if normalize_code(record.get("指数代码"))
    }
    extracted_by_code = {
        normalize_code(record.get("指数代码")): record
        for record in extracted_records
        if normalize_code(record.get("指数代码"))
    }
    review_by_code = {
        normalize_code(record.get("指数代码")): record
        for record in review_records
        if normalize_code(record.get("指数代码"))
    }

    # Copy first so no source file can be modified.
    shutil.copy2(CORE_SOURCE, CORE_OUTPUT)
    wb = load_workbook(CORE_OUTPUT)
    required_sheets = [CORE_RULE_SHEET, MAPPING_SHEET, QUALITY_SHEET]
    missing_sheets = [name for name in required_sheets if name not in wb.sheetnames]
    if missing_sheets:
        raise KeyError(f"核心数据库缺少必需sheet：{missing_sheets}")

    # Snapshot every source sheet row-count before the intended updates.
    original_row_counts = {name: nonempty_row_count(wb[name]) for name in wb.sheetnames}
    rule_sheet_index = wb.index(wb[CORE_RULE_SHEET])
    rule_headers, core_rules = core_table_records(wb[CORE_RULE_SHEET])
    if "指数代码" not in rule_headers:
        raise KeyError("10_指数规则明细缺少“指数代码”字段")
    if len(core_rules) != 82:
        raise ValueError(f"10_指数规则明细应为82行，实际为{len(core_rules)}行")
    core_codes = {normalize_code(record.get("指数代码")) for record in core_rules}
    if len(core_codes) != len(core_rules):
        raise ValueError("10_指数规则明细存在重复指数代码")

    # Verify all mapping-sheet tracked indexes exist in 10 before applying updates.
    mapping_headers, mapping_records = core_table_records(wb[MAPPING_SHEET])
    mapped_index_codes = {normalize_code(record.get("跟踪指数代码")) for record in mapping_records if normalize_code(record.get("跟踪指数代码"))}
    missing_mapped_indexes_before = sorted(mapped_index_codes - core_codes)
    if missing_mapped_indexes_before:
        raise ValueError(f"09_ETF_指数映射有指数未出现在10_指数规则明细：{missing_mapped_indexes_before}")

    diff_headers = [
        "指数代码", "指数名称", "字段名", "原表字段值", "官网核验字段值", "最终字段值", "是否发生更新", "更新类型",
        "官网来源URL", "官网来源域名", "来源类型", "字段抽取状态", "建议处理动作", "备注",
    ]
    diff_rows: list[list[Any]] = []
    index_summary_rows: list[list[Any]] = []
    review_output_rows: list[list[Any]] = []
    updated_records: list[dict[str, Any]] = []

    auto_updated_index_codes: set[str] = set()
    auto_updated_field_count = 0
    source_only_index_codes: set[str] = set()
    key_update_summary: list[str] = []

    field_to_official = {
        "单只成分权重上限": "官网核验_单只成分权重上限",
        "行业权重限制": "官网核验_行业权重限制",
        "调样频率": "官网核验_调样频率",
    }

    for original in core_rules:
        record = dict(original)
        code = normalize_code(record.get("指数代码"))
        name = record.get("指数名称")
        official = official_by_code.get(code, {})
        extracted = extracted_by_code.get(code, {})
        url = official.get("官网来源URL") or extracted.get("官网来源URL")
        domain = official.get("官网来源域名")
        source_type = official.get("官网来源类型")
        extract_status = official.get("字段抽取状态")
        action = extracted.get("建议处理动作") or official.get("建议处理动作")
        source_points = extracted.get("来源要点") or official.get("来源要点") or official.get("核验结果摘要")

        originals = {field: original.get(field) for field in TRACKED_FIELDS}
        official_values: dict[str, Any] = {
            "指数编制方案链接": url,
            "信息来源": f"官网核验：{domain}；来源类型：{source_type}" if is_valid_official_value(domain) else None,
            "单只成分权重上限": extracted.get("官网核验_单只成分权重上限"),
            "行业权重限制": extracted.get("官网核验_行业权重限制"),
            "调样频率": extracted.get("官网核验_调样频率"),
            "指数简介": extracted.get("官网核验_样本/规则"),
        }

        changed_key_fields: list[str] = []
        changed_source_fields: list[str] = []
        # Source link: a non-empty official URL always takes precedence.
        if is_valid_official_value(url):
            if clean_text(record.get("指数编制方案链接")) != clean_text(url):
                record["指数编制方案链接"] = url
                changed_source_fields.append("指数编制方案链接")
        # Information source: only replace when the official domain is explicit.
        if is_valid_official_value(domain):
            official_source_text = f"官网核验：{domain}；来源类型：{source_type}"
            if clean_text(record.get("信息来源")) != official_source_text:
                record["信息来源"] = official_source_text
                changed_source_fields.append("信息来源")

        # Only explicit values from the extracted-fields table can overwrite key rule fields.
        for field, official_field in field_to_official.items():
            candidate = extracted.get(official_field)
            if is_valid_official_value(candidate) and clean_text(record.get(field)) != clean_text(candidate):
                record[field] = candidate
                changed_key_fields.append(field)

        # The sample/rule text is supplementary: it fills an empty/pending index intro only.
        intro_candidate = extracted.get("官网核验_样本/规则")
        if is_missing_or_pending(record.get("指数简介")) and is_valid_official_value(intro_candidate):
            record["指数简介"] = intro_candidate
            changed_source_fields.append("指数简介")

        key_missing = [field for field in KEY_FIELDS if is_missing_or_pending(record.get(field))]
        has_url = is_valid_official_value(record.get("指数编制方案链接"))
        if has_url and not key_missing:
            record["规则完整性状态"] = "官网来源已补充，关键字段较完整"
        elif has_url:
            record["规则完整性状态"] = "官网来源已定位，部分字段仍需复核"
        else:
            record["规则完整性状态"] = "官网来源待补充"

        if code in extracted_by_code and changed_key_fields:
            record["是否已核验"] = "是-关键字段已官网核验"
        elif has_url:
            record["是否已核验"] = "部分-官网来源已定位"
        else:
            record["是否已核验"] = "否"

        record["是否仍需官网核验"] = "是" if key_missing else "否"
        if not key_missing:
            record["仍缺失字段说明"] = "关键规则字段已补充或已核验"
        elif has_url:
            record["仍缺失字段说明"] = f"官网来源已定位，仍需逐字段复核：{'、'.join(key_missing)}"
        else:
            record["仍缺失字段说明"] = f"仍需复核：{'、'.join(key_missing)}"

        if changed_key_fields:
            auto_updated_index_codes.add(code)
            auto_updated_field_count += len(changed_key_fields)
            key_update_summary.append(f"{code}：{'、'.join(changed_key_fields)}")
        elif has_url and ("指数编制方案链接" in changed_source_fields or "信息来源" in changed_source_fields):
            source_only_index_codes.add(code)

        # Field-level trace: exactly one row per tracked field per core index.
        for field in TRACKED_FIELDS:
            original_value = originals.get(field)
            final_value = record.get(field)
            official_value = official_values.get(field)
            changed = clean_text(original_value) != clean_text(final_value)
            if field in KEY_FIELDS and is_valid_official_value(official_value):
                update_type = "官网明确字段覆盖" if changed else "原值保留"
            elif field in {"指数编制方案链接", "信息来源"} and is_valid_official_value(official_value):
                update_type = "官网来源补充" if changed else "原值保留"
            elif field == "指数简介" and changed:
                update_type = "官网明确字段覆盖"
            elif field in {"规则完整性状态", "是否已核验", "是否仍需官网核验", "仍缺失字段说明"}:
                update_type = "状态字段重算" if changed else "原值保留"
            elif not official:
                update_type = "待人工复核"
            else:
                update_type = "原值保留"
            note_parts = [clean_text(source_points)]
            if field in KEY_FIELDS and not is_valid_official_value(official_value) and code in review_by_code:
                note_parts.append("官网来源已定位，但该关键字段未明确抽取")
            diff_rows.append([
                code, name, field, original_value, official_value, final_value,
                "是" if changed else "否", update_type, url, domain, source_type, extract_status, action,
                distinct_join(note_parts),
            ])

        updated_records.append(record)

    # Sort the updated rule sheet by publisher and index code as requested.
    updated_records.sort(key=lambda row: (clean_text(row.get("指数发布机构")), normalize_code(row.get("指数代码"))))
    rule_rows = [[record.get(header) for header in rule_headers] for record in updated_records]
    replace_sheet_with_rows(wb, CORE_RULE_SHEET, rule_headers, rule_rows, rule_sheet_index)

    # Rebuild 12_数据质量总览 by appending the requested official-verification metrics only.
    quality_ws = wb[QUALITY_SHEET]
    quality_headers = [clean_text(cell.value) for cell in quality_ws[1]]
    if quality_headers[:4] != ["模块", "指标", "数值", "状态/说明"]:
        raise ValueError("12_数据质量总览表头与预期不一致")
    old_quality_last = last_nonempty_row(quality_ws)
    matched_codes = core_codes & set(official_by_code)
    official_unmatched_codes = sorted(set(official_by_code) - core_codes)
    core_unmatched_codes = sorted(core_codes - set(official_by_code))
    still_review_codes = {normalize_code(row.get("指数代码")) for row in updated_records if row.get("是否仍需官网核验") == "是"}
    not_review_codes = {normalize_code(row.get("指数代码")) for row in updated_records if row.get("是否仍需官网核验") == "否"}
    source_only_count = len(source_only_index_codes - auto_updated_index_codes)
    quality_append_rows = [
        ["指数规则官网核验增强", "官网核验表指数数量", len(official_by_code), "01_82指数官网核验总表"],
        ["指数规则官网核验增强", "指数规则成功匹配数量", len(matched_codes), "按指数代码匹配"],
        ["指数规则官网核验增强", "未匹配指数数量", len(official_unmatched_codes) + len(core_unmatched_codes), f"官网表未匹配：{'、'.join(official_unmatched_codes) or '无'}；规则表未匹配：{'、'.join(core_unmatched_codes) or '无'}"],
        ["指数规则官网核验增强", "关键字段发生更新的指数数量", len(auto_updated_index_codes), "单只成分权重上限、行业权重限制、调样频率"],
        ["指数规则官网核验增强", "关键字段发生更新的字段数量", auto_updated_field_count, ""],
        ["指数规则官网核验增强", "仅补充官网来源的指数数量", source_only_count, "未覆盖关键字段"],
        ["指数规则官网核验增强", "更新后仍需官网核验指数数量", len(still_review_codes), ""],
        ["指数规则官网核验增强", "更新后无需官网核验指数数量", len(not_review_codes), ""],
        ["指数规则官网核验增强", "是否生成差异表", "是", "13_指数规则官网核验差异及独立差异表文件"],
        ["指数规则官网核验增强", "是否生成更新日志", "是", "14_指数规则更新日志"],
    ]
    for row in quality_append_rows:
        quality_ws.append(row)
    for row in quality_ws.iter_rows(min_row=old_quality_last + 1, max_row=old_quality_last + len(quality_append_rows)):
        for cell in row:
            cell.alignment = BODY_ALIGNMENT
    quality_ws.auto_filter.ref = f"A1:D{last_nonempty_row(quality_ws)}"
    quality_ws.freeze_panes = "A2"

    # New sheet 13: field-level difference output inside the enhanced database.
    diff_rows.sort(key=lambda row: (0 if row[6] == "是" else 1, normalize_code(row[0]), clean_text(row[2])))
    replace_sheet_with_rows(wb, "13_指数规则官网核验差异", diff_headers, diff_rows)

    # New sheet 14: update log with all requested operating statistics.
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_headers = ["项目", "结果/内容"]
    log_rows = [
        ["本次处理时间", now_text],
        ["核心数据库文件路径", str(CORE_SOURCE)],
        ["官网核验文件路径", str(OFFICIAL_SOURCE)],
        ["输出核心数据库路径", str(CORE_OUTPUT)],
        ["核心数据库指数规则数量", len(updated_records)],
        ["官网核验表指数数量", len(official_by_code)],
        ["成功匹配指数数量", len(matched_codes)],
        ["未匹配指数数量", len(official_unmatched_codes) + len(core_unmatched_codes)],
        ["自动更新关键字段的指数数量", len(auto_updated_index_codes)],
        ["自动更新关键字段的字段数", auto_updated_field_count],
        ["仅补充官网来源但未覆盖字段的指数数量", source_only_count],
        ["更新后是否仍需官网核验=是的指数数量", len(still_review_codes)],
        ["更新后是否仍需官网核验=否的指数数量", len(not_review_codes)],
        ["未匹配指数清单", f"官网核验表未匹配：{'、'.join(official_unmatched_codes) or '无'}；规则表未匹配：{'、'.join(core_unmatched_codes) or '无'}"],
        ["关键字段更新清单摘要", "；".join(key_update_summary) or "无"],
        ["处理说明", "本次只更新指数规则sheet；其他产品、规模、流动性、收益风险、持有人结构、指数表现sheet均保持不变；自动覆盖仅限官网核验表中明确抽取的字段；仅有官网来源但未抽取字段的指数保留原规则字段，并标记为部分核验。"],
    ]
    replace_sheet_with_rows(wb, "14_指数规则更新日志", log_headers, log_rows)

    # Independent difference workbook: details, one-row-per-index summary, and review list.
    for code in official_unmatched_codes:
        official = official_by_code[code]
        review_output_rows.append([
            "官网核验表未匹配核心规则", code, official.get("指数名称"), official.get("官网来源URL"), official.get("官网来源域名"),
            official.get("字段抽取状态"), official.get("建议处理动作"), "官网核验表中的指数未能在核心数据库10_指数规则明细中按指数代码找到"],
        )
    for code in core_unmatched_codes:
        core = next(record for record in updated_records if normalize_code(record.get("指数代码")) == code)
        review_output_rows.append([
            "核心规则未匹配官网核验表", code, core.get("指数名称"), None, None, None, None, "核心数据库中的指数未在官网核验总表中找到"],
        )
    for record in updated_records:
        code = normalize_code(record.get("指数代码"))
        official = official_by_code.get(code, {})
        if is_valid_official_value(record.get("指数编制方案链接")) and record.get("是否仍需官网核验") == "是":
            review_output_rows.append([
                "官网来源已定位但关键字段仍缺失", code, record.get("指数名称"), record.get("指数编制方案链接"), official.get("官网来源域名"),
                official.get("字段抽取状态"), official.get("建议处理动作"), record.get("仍缺失字段说明")],
            )
        if code in review_by_code:
            review = review_by_code[code]
            review_output_rows.append([
                "官网核验表建议继续人工复核", code, record.get("指数名称"), review.get("官网来源URL"), None,
                None, review.get("建议下一步"), "来自03_待逐字段复核清单"],
            )
    # De-duplicate exact review rows and sort by issue then code.
    seen_review = set()
    review_unique: list[list[Any]] = []
    for row in review_output_rows:
        marker = tuple(clean_text(value) for value in row)
        if marker not in seen_review:
            review_unique.append(row)
            seen_review.add(marker)
    review_unique.sort(key=lambda row: (clean_text(row[0]), normalize_code(row[1])))

    summary_headers = [
        "指数代码", "指数名称", "官网来源URL", "官网来源域名", "来源类型", "字段抽取状态", "建议处理动作",
        "关键字段更新数", "已更新字段", "仅补充官网来源但未覆盖字段", "是否仍需官网核验", "规则完整性状态", "是否已核验", "仍缺失字段说明", "处理备注",
    ]
    index_updates: dict[str, list[str]] = defaultdict(list)
    for row in diff_rows:
        if row[6] == "是":
            index_updates[row[0]].append(row[2])
    for record in updated_records:
        code = normalize_code(record.get("指数代码"))
        official = official_by_code.get(code, {})
        index_summary_rows.append([
            code, record.get("指数名称"), record.get("指数编制方案链接"), official.get("官网来源域名"), official.get("官网来源类型"),
            official.get("字段抽取状态"), official.get("建议处理动作"), len(index_updates.get(code, [])),
            "；".join(index_updates.get(code, [])) or "无", "是" if code in (source_only_index_codes - auto_updated_index_codes) else "否",
            record.get("是否仍需官网核验"), record.get("规则完整性状态"), record.get("是否已核验"), record.get("仍缺失字段说明"),
            official.get("核验结果摘要") or official.get("来源要点"),
        ])
    index_summary_rows.sort(key=lambda row: normalize_code(row[0]))

    diff_wb = Workbook()
    diff_ws = diff_wb.active
    diff_ws.title = "字段级差异明细"
    diff_ws.append(diff_headers)
    for row in diff_rows:
        diff_ws.append(row)
    style_table_sheet(diff_ws, diff_headers, diff_rows)
    summary_ws = diff_wb.create_sheet("指数级更新摘要")
    summary_ws.append(summary_headers)
    for row in index_summary_rows:
        summary_ws.append(row)
    style_table_sheet(summary_ws, summary_headers, index_summary_rows)
    review_headers = ["问题类型", "指数代码", "指数名称", "官网来源URL", "官网来源域名", "字段抽取状态", "建议处理动作", "说明"]
    review_ws = diff_wb.create_sheet("未匹配与待人工复核")
    review_ws.append(review_headers)
    for row in review_unique:
        review_ws.append(row)
    style_table_sheet(review_ws, review_headers, review_unique)
    diff_wb.save(DIFF_OUTPUT)

    # Final hard validations before saving the enhanced core database.
    if len(updated_records) != 82:
        raise ValueError("更新后10_指数规则明细不是82行")
    updated_codes = {normalize_code(record.get("指数代码")) for record in updated_records}
    if len(updated_codes) != 82:
        raise ValueError("更新后10_指数规则明细存在重复指数代码")
    if mapped_index_codes - updated_codes:
        raise ValueError("更新后09_ETF_指数映射存在未匹配指数规则")
    protected_sheets = [
        "01_产品主表", "02_月度规模份额", "03_月度交易流动性", "04_收益风险表现", "05_持有人结构面板",
        "06_持有人结构最新截面", "07_核心指数月度表现估值", "08_核心指数最新收益", "09_ETF_指数映射", "11_口径差异摘要",
    ]
    changed_protected = [
        name for name in protected_sheets if nonempty_row_count(wb[name]) != original_row_counts[name]
    ]
    if changed_protected:
        raise ValueError(f"不应改变的sheet行数发生变化：{changed_protected}")
    if len(wb.sheetnames) != 15:
        raise ValueError(f"输出核心数据库sheet数量应为15，实际为{len(wb.sheetnames)}")
    for ws in wb.worksheets:
        if ws.merged_cells.ranges:
            raise ValueError(f"不允许存在合并单元格：{ws.title}")
        if not clean_text(ws["A1"].value):
            raise ValueError(f"存在空白首行或空表头：{ws.title}")
    wb.save(CORE_OUTPUT)

    print(f"输出核心数据库路径：{CORE_OUTPUT}")
    print(f"输出独立差异表路径：{DIFF_OUTPUT}")
    print(f"核心数据库指数规则行数：{len(updated_records)}")
    print(f"官网核验表指数数量：{len(official_by_code)}")
    print(f"成功匹配指数数量：{len(matched_codes)}")
    print(f"未匹配指数数量：{len(official_unmatched_codes) + len(core_unmatched_codes)}")
    print(f"关键字段发生更新的指数数量：{len(auto_updated_index_codes)}")
    print(f"关键字段发生更新的字段数量：{auto_updated_field_count}")
    print(f"仅补充官网来源的指数数量：{source_only_count}")
    print(f"更新后仍需官网核验指数数量：{len(still_review_codes)}")
    print(f"更新后无需官网核验指数数量：{len(not_review_codes)}")
    print("是否通过所有校验：是")


if __name__ == "__main__":
    main()
