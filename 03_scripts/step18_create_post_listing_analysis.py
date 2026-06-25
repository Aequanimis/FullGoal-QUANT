from __future__ import annotations

import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\广义策略ETF_月度规模份额表_分析版.xlsx")
OUTPUT = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\广义策略ETF_月度规模份额表_上市后分析版.xlsx")
SOURCE_SHEET = "月度规模份额_分析版"

POST_SHEET = "月度规模份额_上市后分析版"
PRE_SHEET = "上市前记录检查"
TOTAL_SHEET = "月度汇总_广义与核心_上市后"
PRIMARY_SHEET = "月度汇总_按一级策略_上市后"
SECONDARY_SHEET = "月度汇总_按二级策略_上市后"
MANAGER_SHEET = "月度汇总_按基金公司_上市后"
QUALITY_SHEET = "上市后质量检查"

PRE_FIELDS = [
    "日期",
    "Wind代码",
    "证券简称",
    "基金简称",
    "基金管理人",
    "基金成立日",
    "上市日期",
    "基金规模_亿元",
    "基金份额_亿份",
    "一级策略大类",
    "二级策略类别",
    "统计口径分类",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=9, color="1F1F1F")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN_GRAY = Side(style="thin", color="D9E1F2")
BODY_BORDER = Border(bottom=THIN_GRAY)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def as_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if value is None or clean_text(value) == "":
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(clean_text(value), fmt)
        except ValueError:
            continue
    return None


def write_records(ws, headers: list[str], records: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for record in records:
        ws.append([record.get(header) for header in headers])


def style_sheet(ws, headers: list[str]) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BODY_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    sample_limit = min(ws.max_row, 300)
    wide_fields = {"基金全称", "跟踪指数名称", "数据备注", "说明"}
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, sample_limit + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 55))
        width = min(max(max_len * 1.15, 11), 44 if header in wide_fields else 28)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for col_idx, header in enumerate(headers, 1):
        if "日期" in header or header in {"基金成立日", "上市日期"}:
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (datetime, date)):
                    ws.cell(row_idx, col_idx).number_format = "yyyy-mm-dd"
        if "规模" in header:
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "#,##0.0000"
        if "亿份" in header:
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "#,##0.0000"
        elif header == "基金份额_份":
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "#,##0"


def sum_numeric(records: list[dict[str, Any]], field: str) -> float:
    return round(
        sum(float(record[field]) for record in records if isinstance(record.get(field), (int, float))),
        4,
    )


def scale_count(records: list[dict[str, Any]]) -> int:
    return len({record["Wind代码"] for record in records if record.get("基金规模_亿元") is not None})


def main() -> None:
    if not INPUT.exists():
        raise FileNotFoundError(f"输入文件不存在：{INPUT}")

    input_wb = load_workbook(INPUT, read_only=True, data_only=True)
    if SOURCE_SHEET not in input_wb.sheetnames:
        raise KeyError(f"输入文件缺少sheet：{SOURCE_SHEET}")
    source_ws = input_wb[SOURCE_SHEET]
    headers = [cell.value for cell in next(source_ws.iter_rows())]
    required = {
        "日期",
        "Wind代码",
        "上市日期",
        "基金管理人",
        "一级策略大类",
        "二级策略类别",
        "统计口径分类",
        "是否纳入核心策略ETF统计",
        "是否纳入广义策略ETF统计",
        "基金规模_亿元",
        "基金份额_亿份",
    }
    missing = required - set(headers)
    if missing:
        raise KeyError(f"源sheet缺少字段：{sorted(missing)}")

    original_records: list[dict[str, Any]] = []
    post_records: list[dict[str, Any]] = []
    pre_records: list[dict[str, Any]] = []
    missing_listing_date = 0

    for values in source_ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and clean_text(value) for value in values):
            continue
        record = dict(zip(headers, values))
        month = as_datetime(record.get("日期"))
        listing_date = as_datetime(record.get("上市日期"))
        record["日期"] = month
        record["上市日期"] = listing_date
        original_records.append(record)

        if listing_date is None:
            missing_listing_date += 1
            record["是否上市后月份"] = "否"
            pre_records.append(record)
        elif month is not None and month >= listing_date:
            record["是否上市后月份"] = "是"
            post_records.append(record)
        else:
            record["是否上市后月份"] = "否"
            pre_records.append(record)

    post_records.sort(key=lambda r: (r.get("日期") or datetime.max, clean_text(r.get("Wind代码"))))
    pre_records.sort(key=lambda r: (r.get("日期") or datetime.max, clean_text(r.get("Wind代码"))))

    dates = [record["日期"] for record in post_records if record.get("日期") is not None]
    latest_date = max(dates) if dates else None

    monthly: defaultdict[datetime, list[dict[str, Any]]] = defaultdict(list)
    for record in post_records:
        if record.get("日期") is not None:
            monthly[record["日期"]].append(record)

    total_rows: list[dict[str, Any]] = []
    primary_rows: list[dict[str, Any]] = []
    secondary_rows: list[dict[str, Any]] = []
    manager_rows: list[dict[str, Any]] = []

    for month in sorted(monthly):
        records = monthly[month]
        broad = [r for r in records if clean_text(r.get("是否纳入广义策略ETF统计")) == "是"]
        core = [r for r in broad if clean_text(r.get("是否纳入核心策略ETF统计")) == "是"]
        enhanced = [
            r for r in broad if clean_text(r.get("统计口径分类")) == "广义策略ETF_指数增强"
        ]
        total_rows.append(
            {
                "日期": month,
                "广义策略ETF总规模_亿元": sum_numeric(broad, "基金规模_亿元"),
                "核心策略指数ETF总规模_亿元": sum_numeric(core, "基金规模_亿元"),
                "指数增强多因子ETF总规模_亿元": sum_numeric(enhanced, "基金规模_亿元"),
                "广义策略ETF总份额_亿份": sum_numeric(broad, "基金份额_亿份"),
                "核心策略指数ETF总份额_亿份": sum_numeric(core, "基金份额_亿份"),
                "指数增强多因子ETF总份额_亿份": sum_numeric(enhanced, "基金份额_亿份"),
                "广义策略ETF有规模数据产品数": scale_count(broad),
                "核心策略指数ETF有规模数据产品数": scale_count(core),
                "指数增强多因子ETF有规模数据产品数": scale_count(enhanced),
            }
        )

        by_primary: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
        by_secondary: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
        by_manager: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
        for record in broad:
            primary = clean_text(record.get("一级策略大类")) or "待补充"
            secondary = clean_text(record.get("二级策略类别")) or "待补充"
            manager = clean_text(record.get("基金管理人")) or "待补充"
            by_primary[primary].append(record)
            by_secondary[(primary, secondary)].append(record)
            by_manager[manager].append(record)

        for primary, items in sorted(by_primary.items()):
            primary_rows.append(
                {
                    "日期": month,
                    "一级策略大类": primary,
                    "有规模数据产品数量": scale_count(items),
                    "总规模_亿元": sum_numeric(items, "基金规模_亿元"),
                    "总份额_亿份": sum_numeric(items, "基金份额_亿份"),
                }
            )
        for (primary, secondary), items in sorted(by_secondary.items()):
            secondary_rows.append(
                {
                    "日期": month,
                    "一级策略大类": primary,
                    "二级策略类别": secondary,
                    "有规模数据产品数量": scale_count(items),
                    "总规模_亿元": sum_numeric(items, "基金规模_亿元"),
                    "总份额_亿份": sum_numeric(items, "基金份额_亿份"),
                }
            )
        for manager, items in sorted(by_manager.items()):
            manager_rows.append(
                {
                    "日期": month,
                    "基金管理人": manager,
                    "有规模数据产品数量": scale_count(items),
                    "总规模_亿元": sum_numeric(items, "基金规模_亿元"),
                    "总份额_亿份": sum_numeric(items, "基金份额_亿份"),
                }
            )

    pre_codes = {clean_text(record.get("Wind代码")) for record in pre_records}
    pre_scale = sum_numeric(pre_records, "基金规模_亿元")
    pre_at_latest = [record for record in pre_records if record.get("日期") == latest_date]
    pre_latest_codes = {clean_text(record.get("Wind代码")) for record in pre_at_latest}
    pre_latest_scale = sum_numeric(pre_at_latest, "基金规模_亿元")

    duplicate_groups = defaultdict(int)
    invalid_codes = set()
    for record in post_records:
        duplicate_groups[(record.get("Wind代码"), record.get("日期"))] += 1
        code = clean_text(record.get("Wind代码")).upper()
        if not re.fullmatch(r"\d{6}\.(SH|SZ)", code):
            invalid_codes.add(code or "空白")
    duplicate_count = sum(1 for count in duplicate_groups.values() if count > 1)

    latest_summary = total_rows[-1] if total_rows else {}
    latest_records = monthly.get(latest_date, []) if latest_date else []
    latest_broad = [r for r in latest_records if clean_text(r.get("是否纳入广义策略ETF统计")) == "是"]
    latest_core = [r for r in latest_broad if clean_text(r.get("是否纳入核心策略ETF统计")) == "是"]
    latest_enhanced = [
        r for r in latest_broad if clean_text(r.get("统计口径分类")) == "广义策略ETF_指数增强"
    ]

    quality_rows = [
        {"统计类别": "行数", "指标": "原分析版行数", "数值": len(original_records)},
        {"统计类别": "行数", "指标": "被剔除的上市前记录数量", "数值": len(pre_records)},
        {"统计类别": "行数", "指标": "上市后分析版有效行数", "数值": len(post_records)},
        {"统计类别": "覆盖", "指标": "覆盖Wind代码数量", "数值": len({r["Wind代码"] for r in post_records})},
        {"统计类别": "日期", "指标": "最新日期", "数值": latest_date},
        {"统计类别": "最新月", "指标": "最新日期广义策略ETF产品数", "数值": scale_count(latest_broad)},
        {"统计类别": "最新月", "指标": "最新日期核心策略指数ETF产品数", "数值": scale_count(latest_core)},
        {"统计类别": "最新月", "指标": "最新日期指数增强多因子ETF产品数", "数值": scale_count(latest_enhanced)},
        {"统计类别": "最新月", "指标": "最新日期广义策略ETF总规模", "数值": latest_summary.get("广义策略ETF总规模_亿元"), "说明": "单位：亿元"},
        {"统计类别": "最新月", "指标": "最新日期核心策略指数ETF总规模", "数值": latest_summary.get("核心策略指数ETF总规模_亿元"), "说明": "单位：亿元"},
        {"统计类别": "最新月", "指标": "最新日期指数增强多因子ETF总规模", "数值": latest_summary.get("指数增强多因子ETF总规模_亿元"), "说明": "单位：亿元"},
        {"统计类别": "质量", "指标": "是否存在重复Wind代码+日期", "数值": "是" if duplicate_count else "否", "说明": f"重复键数量：{duplicate_count}"},
        {"统计类别": "质量", "指标": "是否存在非.SH/.SZ代码", "数值": "是" if invalid_codes else "否", "说明": "；".join(sorted(invalid_codes))},
        {"统计类别": "缺失", "指标": "基金规模_亿元缺失数量", "数值": sum(r.get("基金规模_亿元") is None for r in post_records)},
        {"统计类别": "缺失", "指标": "基金份额_亿份缺失数量", "数值": sum(r.get("基金份额_亿份") is None for r in post_records)},
        {"统计类别": "上市日期", "指标": "上市日期缺失记录数量", "数值": missing_listing_date},
    ]

    wb = Workbook()
    wb.remove(wb.active)
    post_ws = wb.create_sheet(POST_SHEET)
    pre_ws = wb.create_sheet(PRE_SHEET)
    total_ws = wb.create_sheet(TOTAL_SHEET)
    primary_ws = wb.create_sheet(PRIMARY_SHEET)
    secondary_ws = wb.create_sheet(SECONDARY_SHEET)
    manager_ws = wb.create_sheet(MANAGER_SHEET)
    quality_ws = wb.create_sheet(QUALITY_SHEET)

    post_headers = headers + ["是否上市后月份"]
    write_records(post_ws, post_headers, post_records)
    write_records(pre_ws, PRE_FIELDS, pre_records)
    # Append compact statistics under the pre-listing detail table.
    pre_stats_start = pre_ws.max_row + 2
    pre_ws.cell(pre_stats_start, 1, "统计指标")
    pre_ws.cell(pre_stats_start, 2, "数值")
    pre_stats = [
        ("上市前记录数量", len(pre_records)),
        ("涉及Wind代码数量", len(pre_codes)),
        ("上市前记录规模合计_亿元", pre_scale),
        ("最新日期中被剔除的产品数量", len(pre_latest_codes)),
        ("最新日期中被剔除的规模_亿元", pre_latest_scale),
    ]
    for offset, (label, value) in enumerate(pre_stats, 1):
        pre_ws.cell(pre_stats_start + offset, 1, label)
        pre_ws.cell(pre_stats_start + offset, 2, value)

    total_headers = [
        "日期",
        "广义策略ETF总规模_亿元",
        "核心策略指数ETF总规模_亿元",
        "指数增强多因子ETF总规模_亿元",
        "广义策略ETF总份额_亿份",
        "核心策略指数ETF总份额_亿份",
        "指数增强多因子ETF总份额_亿份",
        "广义策略ETF有规模数据产品数",
        "核心策略指数ETF有规模数据产品数",
        "指数增强多因子ETF有规模数据产品数",
    ]
    write_records(total_ws, total_headers, total_rows)
    primary_headers = ["日期", "一级策略大类", "有规模数据产品数量", "总规模_亿元", "总份额_亿份"]
    write_records(primary_ws, primary_headers, primary_rows)
    secondary_headers = ["日期", "一级策略大类", "二级策略类别", "有规模数据产品数量", "总规模_亿元", "总份额_亿份"]
    write_records(secondary_ws, secondary_headers, secondary_rows)
    manager_headers = ["日期", "基金管理人", "有规模数据产品数量", "总规模_亿元", "总份额_亿份"]
    write_records(manager_ws, manager_headers, manager_rows)
    quality_headers = ["统计类别", "指标", "数值", "说明"]
    write_records(quality_ws, quality_headers, quality_rows)

    for ws, sheet_headers in [
        (post_ws, post_headers),
        (pre_ws, PRE_FIELDS),
        (total_ws, total_headers),
        (primary_ws, primary_headers),
        (secondary_ws, secondary_headers),
        (manager_ws, manager_headers),
        (quality_ws, quality_headers),
    ]:
        style_sheet(ws, sheet_headers)

    # Style pre-listing statistics block separately.
    for cell in pre_ws[pre_stats_start]:
        if cell.column <= 2:
            cell.fill = SECTION_FILL
            cell.font = Font(name="微软雅黑", size=9, bold=True)
    for row_idx in range(pre_stats_start + 1, pre_stats_start + 1 + len(pre_stats)):
        pre_ws.cell(row_idx, 1).font = BODY_FONT
        pre_ws.cell(row_idx, 2).font = BODY_FONT
        if "规模" in clean_text(pre_ws.cell(row_idx, 1).value):
            pre_ws.cell(row_idx, 2).number_format = "#,##0.0000"

    previous = None
    for row_idx in range(2, quality_ws.max_row + 1):
        section = quality_ws.cell(row_idx, 1).value
        if section != previous:
            for col_idx in range(1, quality_ws.max_column + 1):
                quality_ws.cell(row_idx, col_idx).fill = SECTION_FILL
                quality_ws.cell(row_idx, col_idx).font = Font(name="微软雅黑", size=9, bold=True)
        previous = section
        if quality_ws.cell(row_idx, 2).value == "最新日期":
            quality_ws.cell(row_idx, 3).number_format = "yyyy-mm-dd"
        if "总规模" in clean_text(quality_ws.cell(row_idx, 2).value):
            quality_ws.cell(row_idx, 3).number_format = "#,##0.0000"

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    latest_product_count = scale_count(latest_broad)
    latest_scale = latest_summary.get("广义策略ETF总规模_亿元")
    print(f"原分析版行数：{len(original_records)}")
    print(f"剔除上市前记录数量：{len(pre_records)}")
    print(f"上市后有效行数：{len(post_records)}")
    print(f"最新日期：{latest_date.strftime('%Y-%m-%d') if latest_date else '无'}")
    print(f"最新日期广义策略ETF产品数：{latest_product_count}")
    print(f"最新日期广义策略ETF总规模：{latest_scale:.4f}亿元" if latest_scale is not None else "最新日期广义策略ETF总规模：无")
    print(f"输出文件路径：{OUTPUT}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"处理失败：{exc}", file=sys.stderr)
        raise
