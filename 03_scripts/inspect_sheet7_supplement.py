from pathlib import Path

from openpyxl import load_workbook


base = Path.home() / "Desktop" / "Fullgoal" / "课题研究" / "处理后数据" / "wind代码池" / "sheet7"
paths = [
    base / "核心策略ETF_跟踪指数规则_验收修正版.xlsx",
    base / "33只核心策略指数编制规则信息补充表.xlsx",
]

for path in paths:
    print(f"\nFILE {path.name}")
    wb = load_workbook(path, read_only=True, data_only=True)
    print("Sheets:", wb.sheetnames)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(" Sheet:", sheet_name, "max", ws.max_row, ws.max_column)
        for row in ws.iter_rows(max_row=4, values_only=True):
            print("  ", tuple(row[:25]))
        break
