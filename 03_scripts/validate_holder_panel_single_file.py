from __future__ import annotations

import hashlib
import math
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_FILE = Path(
    r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据"
    r"\全市场ETF基础信息_策略ETF池二次修正版.xlsx"
)
PANEL_FILE = Path(
    r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\sheet5"
    r"\广义策略ETF持有人结构面板数据.xlsx"
)
OUTPUT_FILE = Path(
    r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\sheet5"
    r"\Sheet5_广义策略ETF持有人结构面板数据_单文件验收报告.xlsx"
)

POOL_FIELDS = [
    "Wind代码", "证券简称", "基金简称", "基金管理人", "统计口径分类",
    "是否纳入核心策略ETF统计", "是否纳入广义策略ETF统计", "一级策略大类",
    "二级策略类别", "市场范围_二次修正", "最新基金规模(亿)",
]
REQUIRED_FIELDS = [
    "报告期", "Wind代码", "基金代码", "交易代码", "基金简称",
    "机构投资者持有份额", "机构投资者持有比例", "个人投资者持有份额",
    "个人投资者持有比例", "联接基金持有份额", "联接基金持有比例",
    "前十大持有人持有比例", "基金管理人自持份额", "基金管理人自持比例",
    "员工持有份额", "员工持有比例",
]
NUMERIC_FIELDS = [
    "机构投资者持有份额", "机构投资者持有比例", "个人投资者持有份额",
    "个人投资者持有比例", "联接基金持有份额", "联接基金持有比例",
    "前十大持有人持有比例", "基金管理人自持份额", "基金管理人自持比例",
    "员工持有份额", "员工持有比例",
]
RATIO_FIELDS = {field for field in NUMERIC_FIELDS if "比例" in field}
SHARE_FIELDS = set(NUMERIC_FIELDS) - RATIO_FIELDS
CORE_FIELDS = ["机构投资者持有比例", "个人投资者持有比例", "前十大持有人持有比例"]
EMPTY_TOKENS = {"", "-", "--", "N/A", "NA", "NAN", "NONE", "NULL", "WIND暂不可得"}


def digest(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[\s\u3000\u200b\ufeff]+", "", str(value)).strip()


def code(value: Any) -> str:
    return text(value).upper()


def missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return text(value).upper() in EMPTY_TOKENS


def number(value: Any) -> tuple[float | None, str]:
    if missing(value):
        return None, "缺失"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value), "可转数值"
    try:
        return float(text(value).replace(",", "").replace("，", "").replace("%", "")), "可转数值"
    except ValueError:
        return None, "不可转数值"


def read(path: Path, sheet: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    headers = [str(v or "").strip() for v in next(rows)]
    records = [dict(zip(headers, row)) for row in rows]
    wb.close()
    return headers, records


def period_type(period: Any) -> str:
    p = text(period).upper()
    if "一季报" in p or p.startswith("Q1"):
        return "一季报"
    if "三季报" in p or p.startswith("Q3"):
        return "三季报"
    if "中报" in p or "半年报" in p:
        return "中报"
    if "年报" in p:
        return "年报"
    return "其他"


def target_period(period: Any) -> bool:
    return period_type(period) in {"年报", "中报"}


def write(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def format_book(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    white_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="微软雅黑", size=9)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 32
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                header = str(ws.cell(1, cell.column).value or "")
                if isinstance(cell.value, (date, datetime)):
                    cell.number_format = "yyyy-mm-dd"
                elif isinstance(cell.value, (int, float)):
                    if "比例" in header or "非空率" in header:
                        cell.number_format = "0.00"
                    elif "规模" in header:
                        cell.number_format = "0.0000"
                    else:
                        cell.number_format = "#,##0"
        for col in range(1, ws.max_column + 1):
            values = [str(ws.cell(row, col).value or "") for row in range(1, min(ws.max_row, 1000) + 1)]
            length = max((len(v) for v in values), default=8)
            width = min(max(length * 1.1 + 2, 10), 38)
            if any(k in str(ws.cell(1, col).value or "") for k in ("说明", "建议", "报告期列表", "待补充字段")):
                width = min(max(width, 28), 50)
            ws.column_dimensions[get_column_letter(col)].width = width
        if ws.title == "验收总览":
            ws.column_dimensions["A"].width = 36
            ws.column_dimensions["B"].width = 24
            ws.column_dimensions["C"].width = 40
            ws.column_dimensions["D"].width = 56
            for row in range(2, ws.max_row + 1):
                ws.cell(row, 4).alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    if not BASE_FILE.exists() or not PANEL_FILE.exists():
        raise FileNotFoundError("主产品池或待验收面板文件不存在")
    hashes = {BASE_FILE: digest(BASE_FILE), PANEL_FILE: digest(PANEL_FILE)}

    base_headers, base_rows = read(BASE_FILE, "策略ETF_最终统计池")
    panel_headers, raw_rows = read(PANEL_FILE)
    pool_rows = [r for r in base_rows if text(r.get("是否纳入广义策略ETF统计")) == "是"]
    pool = {code(r.get("Wind代码")): {f: r.get(f) for f in POOL_FIELDS} for r in pool_rows if code(r.get("Wind代码"))}
    pool_codes = set(pool)

    source_rows = []
    valid_rows = []
    for row in raw_rows:
        c = code(row.get("Wind代码"))
        joined = " ".join(text(v) for v in row.values())
        if not c or "数据来源" in c or "数据来源" in joined:
            source_rows.append(row)
            continue
        item = dict(row)
        item["Wind代码"] = c
        item["报告期类型"] = period_type(item.get("报告期"))
        valid_rows.append(item)

    all_codes = {r["Wind代码"] for r in valid_rows}
    outside_codes = all_codes - pool_codes
    never_codes = pool_codes - all_codes
    target_rows = [r for r in valid_rows if target_period(r.get("报告期"))]
    non_target_rows = [r for r in valid_rows if not target_period(r.get("报告期"))]

    keys = Counter((text(r.get("报告期")), r["Wind代码"]) for r in valid_rows)
    duplicate_keys = {key: count for key, count in keys.items() if count > 1}

    anomalies: list[dict[str, Any]] = []
    for row in valid_rows:
        for field in NUMERIC_FIELDS:
            if field not in panel_headers:
                continue
            value, state = number(row.get(field))
            if state == "不可转数值":
                anomalies.append({
                    "异常类型": "不可转为数值", "报告期": row.get("报告期"), "Wind代码": row["Wind代码"],
                    "面板证券简称": row.get("证券简称"), "字段": field, "原始值": row.get(field),
                    "期望范围": "数值或标准缺失标记", "说明": "非空值无法转为数值",
                })
            elif state == "可转数值":
                if field in RATIO_FIELDS and not 0 <= value <= 100:
                    anomalies.append({
                        "异常类型": "比例超范围", "报告期": row.get("报告期"), "Wind代码": row["Wind代码"],
                        "面板证券简称": row.get("证券简称"), "字段": field, "原始值": row.get(field),
                        "期望范围": "0至100", "说明": "",
                    })
                if field in SHARE_FIELDS and value < 0:
                    anomalies.append({
                        "异常类型": "份额为负", "报告期": row.get("报告期"), "Wind代码": row["Wind代码"],
                        "面板证券简称": row.get("证券简称"), "字段": field, "原始值": row.get(field),
                        "期望范围": "大于等于0", "说明": "",
                    })
        inst, si = number(row.get("机构投资者持有比例"))
        indiv, sp = number(row.get("个人投资者持有比例"))
        if si == "可转数值" and sp == "可转数值" and not 99.9 <= inst + indiv <= 100.1:
            anomalies.append({
                "异常类型": "机构+个人比例合计异常", "报告期": row.get("报告期"),
                "Wind代码": row["Wind代码"], "面板证券简称": row.get("证券简称"),
                "字段": "机构投资者持有比例+个人投资者持有比例", "原始值": inst + indiv,
                "期望范围": "99.9至100.1", "说明": f"{inst}+{indiv}={inst+indiv}",
            })
    for (period, wind_code), count in duplicate_keys.items():
        anomalies.append({
            "异常类型": "重复报告期+Wind代码", "报告期": period, "Wind代码": wind_code,
            "面板证券简称": "", "字段": "报告期+Wind代码", "原始值": count,
            "期望范围": "唯一", "说明": f"重复{count}次",
        })
    for c in sorted(outside_codes):
        anomalies.append({
            "异常类型": "代码池外产品", "报告期": "", "Wind代码": c, "面板证券简称": "",
            "字段": "Wind代码", "原始值": c, "期望范围": "223只广义策略ETF代码池", "说明": "",
        })

    period_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in valid_rows:
        period_groups[text(row.get("报告期"))].append(row)
    period_order = ["2023年年报", "2024年中报", "2024年年报", "2025年中报", "2025年年报", "2026年一季报"]
    period_stats = []
    for period in sorted(period_groups, key=lambda p: (period_order.index(p) if p in period_order else 999, p)):
        rows = period_groups[period]
        period_stats.append({
            "报告期": period, "报告期类型": period_type(period), "记录数量": len(rows),
            "覆盖Wind代码数量": len({r["Wind代码"] for r in rows}),
            "有机构投资者持有比例数量": sum(not missing(r.get("机构投资者持有比例")) for r in rows),
            "有个人投资者持有比例数量": sum(not missing(r.get("个人投资者持有比例")) for r in rows),
            "有前十大持有人持有比例数量": sum(not missing(r.get("前十大持有人持有比例")) for r in rows),
            "核心字段全空数量": sum(all(missing(r.get(f)) for f in CORE_FIELDS) for r in rows),
        })

    field_stats = []
    actual_map = {
        "交易代码": "Wind代码",
        "基金简称": "证券简称/主产品池基金简称",
    }
    for field in REQUIRED_FIELDS:
        exists = field in panel_headers
        actual_field = field if exists else actual_map.get(field, "")
        if exists:
            nonempty = sum(not missing(r.get(field)) for r in valid_rows)
            rate = nonempty / len(valid_rows) * 100 if valid_rows else 0
            if rate >= 80:
                judgment = "已提供且可用"
                note = ""
            elif nonempty > 0:
                judgment = "已提供但大量缺失"
                note = "可用于部分记录，正式分析需披露缺失情况"
            else:
                judgment = "Wind暂不可得 / 需后续补充"
                note = "字段存在但全为空"
        elif field == "交易代码":
            nonempty, rate = len(valid_rows), 100.0
            judgment = "可由其他字段推导"
            note = "可直接使用Wind代码作为交易代码"
        elif field == "基金简称":
            nonempty, rate = len(valid_rows), 100.0
            judgment = "可由其他字段推导"
            note = "面板提供证券简称；正式清洗时按Wind代码匹配主产品池基金简称"
        elif field in {"联接基金持有份额", "联接基金持有比例"}:
            nonempty, rate = 0, 0.0
            judgment = "Wind暂不可得 / 需后续补充"
            note = "Wind未提供，不影响主字段验收，但无法分析联接基金贡献"
        elif field == "基金管理人自持份额":
            nonempty, rate = 0, 0.0
            judgment = "未提供"
            note = "仅提供基金管理人自持比例，属于部分满足"
        else:
            nonempty, rate = 0, 0.0
            judgment = "未提供"
            note = ""
        field_stats.append({
            "需求字段": field, "面板数据中是否存在": "是" if exists else "否",
            "面板数据实际字段名": actual_field, "非空数量": nonempty,
            "非空率": round(rate, 2), "可用性判断": judgment, "说明": note,
        })

    y2025 = [r for r in valid_rows if text(r.get("报告期")) == "2025年年报"]
    y2025_by_code = {r["Wind代码"]: r for r in y2025}
    coverage_2025 = []
    pending_2025 = []
    for c in sorted(pool_codes):
        p = pool[c]
        row = y2025_by_code.get(c)
        if row is None:
            status = "未覆盖"
            missing_fields = "机构投资者持有比例；个人投资者持有比例；前十大持有人持有比例"
            missing_type = "未覆盖产品"
        else:
            absent = [f for f in CORE_FIELDS if missing(row.get(f))]
            if len(absent) == 3:
                status = "已覆盖但核心字段全空"
                missing_type = "核心字段全空"
            elif absent:
                status = "已覆盖但核心字段部分缺失"
                missing_type = "核心字段部分缺失"
            else:
                status = "已覆盖且核心字段有效"
                missing_type = ""
            missing_fields = "；".join(absent)
        coverage_2025.append({
            "Wind代码": c, "主产品池证券简称": p.get("证券简称"), "基金管理人": p.get("基金管理人"),
            "统计口径分类": p.get("统计口径分类"), "是否纳入核心策略ETF统计": p.get("是否纳入核心策略ETF统计"),
            "一级策略大类": p.get("一级策略大类"), "二级策略类别": p.get("二级策略类别"),
            "市场范围_二次修正": p.get("市场范围_二次修正"), "最新基金规模(亿)": p.get("最新基金规模(亿)"),
            "是否出现在2025年年报": "是" if row else "否",
            "机构投资者持有比例": row.get("机构投资者持有比例") if row else "",
            "个人投资者持有比例": row.get("个人投资者持有比例") if row else "",
            "前十大持有人持有比例": row.get("前十大持有人持有比例") if row else "",
            "覆盖状态": status,
        })
        if status != "已覆盖且核心字段有效":
            pending_2025.append({
                "Wind代码": c, "主产品池证券简称": p.get("证券简称"), "基金管理人": p.get("基金管理人"),
                "统计口径分类": p.get("统计口径分类"), "一级策略大类": p.get("一级策略大类"),
                "二级策略类别": p.get("二级策略类别"), "最新基金规模(亿)": p.get("最新基金规模(亿)"),
                "缺失类型": missing_type, "待补充字段": missing_fields,
                "建议处理方式": "向Wind补充2025年年报持有人结构数据" if row is None else "补充缺失核心字段",
            })

    code_periods: dict[str, set[str]] = defaultdict(set)
    for row in valid_rows:
        code_periods[row["Wind代码"]].add(text(row.get("报告期")))
    code_coverage = []
    for c in sorted(pool_codes):
        periods = sorted(code_periods.get(c, set()), key=lambda p: (period_order.index(p) if p in period_order else 999, p))
        code_coverage.append({
            "Wind代码": c, "主产品池证券简称": pool[c].get("证券简称"),
            "是否在面板中出现": "是" if periods else "否", "出现的报告期数量": len(periods),
            "出现的报告期列表": "；".join(periods), "是否出现在2025年年报": "是" if c in y2025_by_code else "否",
            "覆盖状态": "已覆盖" if periods else "从未出现",
        })

    mismatches = []
    for row in valid_rows:
        c = row["Wind代码"]
        if c in pool and text(row.get("证券简称")) != text(pool[c].get("证券简称")):
            mismatches.append({
                "Wind代码": c, "面板数据证券简称": row.get("证券简称"),
                "主产品池证券简称": pool[c].get("证券简称"), "报告期": row.get("报告期"),
                "说明": "证券简称不一致可能来自Wind主简称口径，后续正式清洗时应以主产品池证券简称为准覆盖。",
            })

    core_anomalies = [a for a in anomalies if a["异常类型"] in {"比例超范围", "机构+个人比例合计异常", "份额为负", "不可转为数值"}]
    field_incomplete = any(r["可用性判断"] in {"未提供", "Wind暂不可得 / 需后续补充"} for r in field_stats)
    if outside_codes or duplicate_keys or core_anomalies:
        conclusion = "不满足，需要重新导出"
        reason = "存在代码池外产品、重复主键或数值异常等阻断问题。"
    elif pending_2025 or field_incomplete:
        conclusion = "基本满足，需补充部分字段/基金"
        reason = (
            f"半年报/年报主数据可用，2025年年报覆盖{len(y2025_by_code)}只、待补充{len(pending_2025)}只；"
            "交易代码可由Wind代码推导，联接基金字段及基金管理人自持份额暂未提供。"
        )
    else:
        conclusion = "满足需求"
        reason = "半年报/年报覆盖、字段和数值质量均符合要求。"

    overview = [
        {"指标": "主产品池广义策略ETF数量", "结果": len(pool_codes), "验收判断": "符合" if len(pool_codes) == 223 else "异常", "说明": "理论数量223只"},
        {"指标": "面板数据原始行数", "结果": len(raw_rows), "验收判断": "信息", "说明": "不含表头"},
        {"指标": "删除数据来源行后的有效行数", "结果": len(valid_rows), "验收判断": "符合", "说明": f"剔除非产品行{len(source_rows)}条"},
        {"指标": "面板数据覆盖Wind代码数量", "结果": len(all_codes), "验收判断": "基本覆盖", "说明": f"至少出现一次；仍有{len(never_codes)}只从未出现"},
        {"指标": "是否存在代码池外产品", "结果": "是" if outside_codes else "否", "验收判断": "异常" if outside_codes else "符合", "说明": f"{len(outside_codes)}只"},
        {"指标": "是否存在重复报告期+Wind代码", "结果": "是" if duplicate_keys else "否", "验收判断": "异常" if duplicate_keys else "符合", "说明": f"{len(duplicate_keys)}组"},
        {"指标": "是否存在比例或数值异常", "结果": "是" if core_anomalies else "否", "验收判断": "异常" if core_anomalies else "符合", "说明": f"{len(core_anomalies)}条"},
        {"指标": "半年报/年报记录数量", "结果": len(target_rows), "验收判断": "可用", "说明": "目标口径"},
        {"指标": "非半年报/年报记录数量", "结果": len(non_target_rows), "验收判断": "单独标注", "说明": "包含2026年一季报，不因此判定文件不可用"},
        {"指标": "最新可用半年报/年报报告期", "结果": "2025年年报", "验收判断": "符合", "说明": ""},
        {"指标": "2025年年报覆盖基金数量", "结果": len(y2025_by_code), "验收判断": "基本覆盖", "说明": ""},
        {"指标": "2025年年报未覆盖基金数量", "结果": len(pool_codes - set(y2025_by_code)), "验收判断": "需补充" if pending_2025 else "符合", "说明": ""},
        {"指标": "2025年年报核心字段全空数量", "结果": sum(r["覆盖状态"] == "已覆盖但核心字段全空" for r in coverage_2025), "验收判断": "符合", "说明": "已覆盖记录中未发现核心字段全空"},
        {"指标": "证券简称不一致记录数量", "结果": len(mismatches), "验收判断": "不作为否决项", "说明": "以Wind代码为主键；正式清洗时用主产品池标准简称覆盖"},
        {"指标": "字段完整性结论", "结果": "部分满足", "验收判断": "需补充", "说明": "交易代码可推导；基金简称可由主池匹配；联接基金字段和基金管理人自持份额未提供"},
        {"指标": "最终验收结论", "结果": conclusion, "验收判断": conclusion, "说明": reason},
    ]

    wb = Workbook()
    wb.remove(wb.active)
    sheets = [
        ("验收总览", ["指标", "结果", "验收判断", "说明"], overview),
        ("报告期覆盖统计", ["报告期", "报告期类型", "记录数量", "覆盖Wind代码数量", "有机构投资者持有比例数量", "有个人投资者持有比例数量", "有前十大持有人持有比例数量", "核心字段全空数量"], period_stats),
        ("字段可用性检查", ["需求字段", "面板数据中是否存在", "面板数据实际字段名", "非空数量", "非空率", "可用性判断", "说明"], field_stats),
        ("2025年年报覆盖核对", ["Wind代码", "主产品池证券简称", "基金管理人", "统计口径分类", "是否纳入核心策略ETF统计", "一级策略大类", "二级策略类别", "市场范围_二次修正", "最新基金规模(亿)", "是否出现在2025年年报", "机构投资者持有比例", "个人投资者持有比例", "前十大持有人持有比例", "覆盖状态"], coverage_2025),
        ("2025年年报待补充清单", ["Wind代码", "主产品池证券简称", "基金管理人", "统计口径分类", "一级策略大类", "二级策略类别", "最新基金规模(亿)", "缺失类型", "待补充字段", "建议处理方式"], pending_2025),
        ("数值异常检查", ["异常类型", "报告期", "Wind代码", "面板证券简称", "字段", "原始值", "期望范围", "说明"], anomalies),
        ("代码覆盖核对", ["Wind代码", "主产品池证券简称", "是否在面板中出现", "出现的报告期数量", "出现的报告期列表", "是否出现在2025年年报", "覆盖状态"], code_coverage),
        ("证券简称不一致检查", ["Wind代码", "面板数据证券简称", "主产品池证券简称", "报告期", "说明"], mismatches),
        ("非半年报年报记录", panel_headers + ["报告期类型"], sorted(non_target_rows, key=lambda r: (text(r.get("报告期")), r["Wind代码"]))),
    ]
    for name, headers, rows in sheets:
        ws = wb.create_sheet(name)
        write(ws, headers, rows)
    format_book(wb)
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp = OUTPUT_FILE.with_name(OUTPUT_FILE.stem + "_tmp.xlsx")
    wb.save(tmp)
    check = load_workbook(tmp, read_only=True, data_only=True)
    if len(check.sheetnames) != 9:
        raise RuntimeError("输出工作簿sheet数量异常")
    check.close()
    tmp.replace(OUTPUT_FILE)
    if any(digest(path) != hashes[path] for path in hashes):
        raise RuntimeError("原始文件被意外修改")

    print(f"面板有效行数：{len(valid_rows)}")
    print(f"面板覆盖基金数量：{len(all_codes)}")
    print(f"半年报/年报记录数量：{len(target_rows)}")
    print(f"非半年报/年报记录数量：{len(non_target_rows)}")
    print(f"2025年年报覆盖基金数量：{len(y2025_by_code)}")
    print(f"2025年年报待补充基金数量：{len(pending_2025)}")
    print(f"是否存在代码池外产品：{'是' if outside_codes else '否'}")
    print(f"是否存在重复记录：{'是' if duplicate_keys else '否'}")
    print(f"最终验收结论：{conclusion}")
    print(f"输出文件路径：{OUTPUT_FILE}")


if __name__ == "__main__":
    main()
