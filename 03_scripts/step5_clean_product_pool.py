"""Step 5：清洗境内策略 ETF 产品池并同步主 Excel。"""

from __future__ import annotations

import csv
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
INPUT_CSV = PROJECT_ROOT / "01_processed_data" / "product_pool" / "境内策略ETF产品池_公开搜索初版.csv"
MAIN_EXCEL = PROJECT_ROOT / "02_outputs" / "excel" / "境内策略ETF产品梳理_初版.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "01_processed_data" / "product_pool"
CLEAN_CSV = OUTPUT_DIR / "境内策略ETF产品池_清洗版.csv"
CLEAN_EXCEL = OUTPUT_DIR / "境内策略ETF产品池_清洗版.xlsx"
ANOMALY_CSV = OUTPUT_DIR / "境内策略ETF产品池_异常行.csv"
LOG_PATH = OUTPUT_DIR / "step5_cleaning_log.md"

INPUT_FIELDS = [
    "策略大类",
    "策略细分",
    "基金代码",
    "产品名称",
    "基金公司",
    "上市交易所",
    "跟踪指数或策略线索",
    "纳入级别",
    "待校验事项",
    "信息来源备注",
]

OUTPUT_FIELDS = [
    "策略大类",
    "策略细分",
    "基金代码",
    "产品名称",
    "基金公司",
    "上市交易所",
    "跟踪指数",
    "指数公司",
    "纳入级别",
    "是否主线",
    "是否补充观察",
    "是否待校验",
    "待校验事项",
    "是否代表产品初筛",
    "代表产品初筛理由",
    "信息来源备注",
    "清洗备注",
]

ANOMALY_FIELDS = ["原始行号", "异常原因", "原始内容", *INPUT_FIELDS]

STRATEGY_NORMALIZATION = {"高股息": "红利", "ESG/可持续": "ESG"}
LEVEL_NORMALIZATION = {
    "主线重点": "主线",
    "主线": "主线",
    "补充观察": "补充观察",
    "跨境补充": "跨境补充",
    "单列观察": "单列观察",
    "仅作线索": "仅作线索",
}
REPRESENTATIVE_KEYWORDS = [
    "上证红利ETF",
    "中证红利低波动ETF",
    "自由现金流ETF",
    "富国中证价值ETF",
    "基本面50ETF",
]


def strip_field(value: str | None) -> str:
    return "" if value is None else value.strip(" \t\r\n")


def join_unique(parts: Iterable[str]) -> str:
    result: list[str] = []
    for part in parts:
        if part and part not in result:
            result.append(part)
    return "；".join(result)


def read_input() -> tuple[int, list[dict[str, str]], list[dict[str, str]]]:
    if not INPUT_CSV.exists():
        raise FileNotFoundError(f"缺少输入文件：{INPUT_CSV}")

    lines = INPUT_CSV.read_text(encoding="utf-8-sig").splitlines()
    if not lines:
        raise ValueError("输入 CSV 为空")

    header = next(csv.reader([lines[0]]))
    if header != INPUT_FIELDS:
        raise ValueError(f"输入 CSV 字段不匹配：{header}")

    valid_rows: list[dict[str, str]] = []
    anomalies: list[dict[str, str]] = []
    original_row_count = 0

    for line_number, raw_line in enumerate(lines[1:], start=2):
        original_row_count += 1
        parsed = next(csv.reader([raw_line]))

        if not parsed or all(not strip_field(value) for value in parsed):
            continue

        cleaned_values = [strip_field(value) for value in parsed]
        anomaly_reason = ""
        if len(cleaned_values) != len(INPUT_FIELDS):
            anomaly_reason = f"字段数量异常：应为10个，实际为{len(cleaned_values)}个"

        padded = (cleaned_values + [""] * len(INPUT_FIELDS))[: len(INPUT_FIELDS)]
        record = dict(zip(INPUT_FIELDS, padded))

        # Step 4 为保证 CSV 列数统一，曾将字段不足的原始行补空；缺少产品名称时仍应隔离。
        if not anomaly_reason and not record["产品名称"]:
            anomaly_reason = "缺少产品名称，无法判断为完整产品记录"

        if anomaly_reason:
            anomalies.append(
                {
                    "原始行号": str(line_number),
                    "异常原因": anomaly_reason,
                    "原始内容": raw_line,
                    **record,
                }
            )
        else:
            valid_rows.append(record)

    return original_row_count, valid_rows, anomalies


def infer_index_company(index_name: str) -> str:
    rules = [
        ("中证", "中证"),
        ("国证", "国证"),
        ("上证", "上证"),
        ("深证", "深证"),
        ("标普", "标普"),
        ("富时", "富时罗素"),
    ]
    for keyword, company in rules:
        if keyword in index_name:
            return company
    return "待补充"


def pending_reasons(record: dict[str, str], tracking_index: str) -> list[str]:
    reasons: list[str] = []
    if record["基金代码"] == "待校验":
        reasons.append("基金代码待校验")
    if record["基金公司"] == "待校验":
        reasons.append("基金公司待校验")
    if record["上市交易所"] == "待校验":
        reasons.append("上市交易所待校验")
    if any(keyword in tracking_index for keyword in ("相关指数", "待校验", "策略线索")):
        reasons.append("跟踪指数全称待校验")
    if "待校验" in record["产品名称"]:
        reasons.append("产品名称待校验")
    return reasons


def clean_record(record: dict[str, str]) -> dict[str, str]:
    original_strategy = record["策略大类"]
    strategy = STRATEGY_NORMALIZATION.get(original_strategy, original_strategy)
    tracking_index = record["跟踪指数或策略线索"]
    index_company = infer_index_company(tracking_index)

    original_level = record["纳入级别"]
    level = LEVEL_NORMALIZATION.get(original_level, original_level or "待判断")
    is_mainline = "是" if level == "主线" else "否"
    is_supplemental = (
        "是" if level in {"补充观察", "跨境补充", "单列观察", "仅作线索"} else "否"
    )

    generated_pending_reasons = pending_reasons(record, tracking_index)
    is_pending = "是" if record["待校验事项"] or generated_pending_reasons else "否"
    pending_item = record["待校验事项"]
    if is_pending == "是" and not pending_item:
        pending_item = join_unique(generated_pending_reasons)

    representative_reasons: list[str] = []
    if original_level == "主线重点":
        representative_reasons.append("原始标记为主线重点")
    if any(keyword in record["产品名称"] for keyword in REPRESENTATIVE_KEYWORDS):
        representative_reasons.append("产品名称属于典型策略方向")
    is_representative = "是" if representative_reasons else "否"

    cleaning_notes: list[str] = []
    if strategy != original_strategy:
        cleaning_notes.append(f"策略大类由{original_strategy}标准化为{strategy}")
    if level != original_level:
        cleaning_notes.append(f"纳入级别由{original_level or '空值'}标准化为{level}")
    if is_pending == "是":
        cleaning_notes.append("标记为待校验")
    if index_company == "待补充":
        cleaning_notes.append("指数公司待补充")
    else:
        cleaning_notes.append("指数公司由指数名称初步推断")
    if not cleaning_notes:
        cleaning_notes.append("正常保留")

    return {
        "策略大类": strategy,
        "策略细分": record["策略细分"],
        "基金代码": record["基金代码"],
        "产品名称": record["产品名称"],
        "基金公司": record["基金公司"],
        "上市交易所": record["上市交易所"],
        "跟踪指数": tracking_index,
        "指数公司": index_company,
        "纳入级别": level,
        "是否主线": is_mainline,
        "是否补充观察": is_supplemental,
        "是否待校验": is_pending,
        "待校验事项": pending_item,
        "是否代表产品初筛": is_representative,
        "代表产品初筛理由": join_unique(representative_reasons),
        "信息来源备注": record["信息来源备注"],
        "清洗备注": join_unique(cleaning_notes),
    }


def completeness_score(record: dict[str, str]) -> tuple[int, int]:
    nonempty = sum(bool(record[field]) for field in OUTPUT_FIELDS)
    text_length = sum(len(record[field]) for field in OUTPUT_FIELDS)
    return nonempty, text_length


def deduplicate(
    rows: list[dict[str, str]],
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    selected: dict[tuple[str, ...], tuple[int, dict[str, str]]] = {}
    removed: list[dict[str, str]] = []

    for position, record in enumerate(rows):
        if record["基金代码"] != "待校验":
            key = ("基金代码", record["基金代码"])
        else:
            key = (
                "组合键",
                record["产品名称"],
                record["基金公司"],
                record["跟踪指数"],
            )

        if key not in selected:
            selected[key] = (position, record)
            continue

        prior_position, prior_record = selected[key]
        if completeness_score(record) > completeness_score(prior_record):
            removed.append({**prior_record, "去重键": " | ".join(key), "删除原因": "重复记录，保留信息更完整的后出现记录"})
            selected[key] = (prior_position, record)
        else:
            removed.append({**record, "去重键": " | ".join(key), "删除原因": "重复记录，保留信息更完整或更早出现的记录"})

    kept = [record for _, record in sorted(selected.values(), key=lambda item: item[0])]
    return kept, removed


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fields, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def display_width(value: object) -> int:
    text = "" if value is None else str(value)
    return sum(
        2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1
        for char in text
    )


def apply_sheet_format(worksheet, text_columns: set[str] | None = None) -> None:
    text_columns = text_columns or set()
    max_row = max(worksheet.max_row, 1)
    max_column = max(worksheet.max_column, 1)
    last_column = get_column_letter(max_column)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{last_column}{max_row}"
    worksheet.sheet_view.zoomScale = 90

    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    header_font = Font(name="Microsoft YaHei", size=10, bold=True)
    body_font = Font(name="Microsoft YaHei", size=10)
    bottom_border = Border(bottom=Side(style="thin", color="A6A6A6"))

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bottom_border
    worksheet.row_dimensions[1].height = 26

    headers = [cell.value for cell in worksheet[1]]
    for row in worksheet.iter_rows(min_row=2, max_row=max_row, max_col=max_column):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_index, header in enumerate(headers, start=1):
        values = [worksheet.cell(row, column_index).value for row in range(1, max_row + 1)]
        measured = max(display_width(value) for value in values) + 3
        cap = 44 if header in {
            "跟踪指数",
            "待校验事项",
            "代表产品初筛理由",
            "信息来源备注",
            "清洗备注",
            "异常原因",
            "原始内容",
        } else 26
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(measured, 12), cap)

        if header in text_columns:
            for row in range(2, max_row + 1):
                worksheet.cell(row, column_index).number_format = "@"


def append_dict_rows(worksheet, fields: list[str], rows: list[dict[str, str]]) -> None:
    worksheet.append(fields)
    for record in rows:
        worksheet.append([record.get(field, "") for field in fields])


def build_statistics(rows: list[dict[str, str]]) -> list[dict[str, str | int]]:
    strategy_counts = Counter(record["策略大类"] for record in rows)
    level_counts = Counter(record["纳入级别"] for record in rows)
    cross_counts = Counter((record["策略大类"], record["纳入级别"]) for record in rows)

    statistics: list[dict[str, str | int]] = []
    for strategy, count in sorted(strategy_counts.items()):
        statistics.append({"统计类型": "按策略大类", "策略大类": strategy, "纳入级别": "", "产品数量": count})
    for level, count in sorted(level_counts.items()):
        statistics.append({"统计类型": "按纳入级别", "策略大类": "", "纳入级别": level, "产品数量": count})
    for (strategy, level), count in sorted(cross_counts.items()):
        statistics.append({"统计类型": "策略大类+纳入级别", "策略大类": strategy, "纳入级别": level, "产品数量": count})
    statistics.append(
        {
            "统计类型": "汇总指标",
            "策略大类": "待校验产品",
            "纳入级别": "",
            "产品数量": sum(record["是否待校验"] == "是" for record in rows),
        }
    )
    statistics.append(
        {
            "统计类型": "汇总指标",
            "策略大类": "代表产品初筛",
            "纳入级别": "",
            "产品数量": sum(record["是否代表产品初筛"] == "是" for record in rows),
        }
    )
    return statistics


def write_clean_excel(rows: list[dict[str, str]], anomalies: list[dict[str, str]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    clean_sheet = workbook.create_sheet("产品池_清洗版")
    append_dict_rows(clean_sheet, OUTPUT_FIELDS, rows)
    apply_sheet_format(clean_sheet, {"基金代码"})

    pending_sheet = workbook.create_sheet("待校验清单")
    pending_rows = [record for record in rows if record["是否待校验"] == "是"]
    append_dict_rows(pending_sheet, OUTPUT_FIELDS, pending_rows)
    apply_sheet_format(pending_sheet, {"基金代码"})

    anomaly_sheet = workbook.create_sheet("异常行隔离")
    append_dict_rows(anomaly_sheet, ANOMALY_FIELDS, anomalies)
    apply_sheet_format(anomaly_sheet, {"基金代码"})

    statistics_sheet = workbook.create_sheet("分类统计")
    statistics_fields = ["统计类型", "策略大类", "纳入级别", "产品数量"]
    append_dict_rows(statistics_sheet, statistics_fields, build_statistics(rows))
    apply_sheet_format(statistics_sheet)

    workbook.save(CLEAN_EXCEL)


def sync_main_excel(rows: list[dict[str, str]]) -> None:
    if not MAIN_EXCEL.exists():
        raise FileNotFoundError(f"缺少主 Excel：{MAIN_EXCEL}")

    workbook = load_workbook(MAIN_EXCEL)
    if "产品总表" not in workbook.sheetnames:
        raise ValueError("主 Excel 中不存在 Sheet【产品总表】")
    worksheet = workbook["产品总表"]
    headers = [cell.value for cell in worksheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}

    field_map = {
        "策略大类": "策略大类",
        "策略细分": "策略细分",
        "产品名称": "产品名称",
        "基金代码": "基金代码",
        "基金公司": "基金公司",
        "上市交易所": "上市交易所",
        "跟踪指数": "跟踪指数",
        "指数公司": "指数公司",
        "是否代表产品初筛": "是否代表产品",
        "代表产品初筛理由": "代表性理由",
        "信息来源备注": "信息来源",
    }
    missing = (set(field_map.values()) | {"备注"}) - set(headers)
    if missing:
        raise ValueError(f"主 Excel 产品总表缺少字段：{sorted(missing)}")

    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)

    for row_number, record in enumerate(rows, start=2):
        for clean_field, main_field in field_map.items():
            worksheet.cell(row_number, columns[main_field], record[clean_field])

        note_parts: list[str] = []
        if record["待校验事项"]:
            note_parts.append(f"待校验事项：{record['待校验事项']}")
        if record["清洗备注"]:
            note_parts.append(f"清洗备注：{record['清洗备注']}")
        worksheet.cell(row_number, columns["备注"], "\n".join(note_parts))

    last_row = len(rows) + 1
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    for row in worksheet.iter_rows(min_row=2, max_row=last_row, max_col=len(headers)):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    code_column = columns["基金代码"]
    for row_number in range(2, last_row + 1):
        worksheet.cell(row_number, code_column).number_format = "@"

    for column_index, header in enumerate(headers, start=1):
        values = [worksheet.cell(row, column_index).value for row in range(1, last_row + 1)]
        measured = max(display_width(value) for value in values) + 3
        cap = 44 if header in {"跟踪指数", "代表性理由", "备注", "信息来源"} else 26
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(measured, 12), cap)

    workbook.save(MAIN_EXCEL)


def validate_outputs(rows: list[dict[str, str]], anomalies: list[dict[str, str]]) -> None:
    if CLEAN_CSV.read_bytes()[:3] != b"\xef\xbb\xbf":
        raise ValueError("清洗版 CSV 不是 UTF-8 with BOM")
    if ANOMALY_CSV.read_bytes()[:3] != b"\xef\xbb\xbf":
        raise ValueError("异常行 CSV 不是 UTF-8 with BOM")

    with CLEAN_CSV.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reloaded = list(csv.DictReader(csv_file))
    if reloaded != rows:
        raise ValueError("清洗版 CSV 回读校验失败")

    clean_workbook = load_workbook(CLEAN_EXCEL)
    expected_sheets = ["产品池_清洗版", "待校验清单", "异常行隔离", "分类统计"]
    if clean_workbook.sheetnames != expected_sheets:
        raise ValueError(f"清洗版 Excel Sheet 不匹配：{clean_workbook.sheetnames}")
    if clean_workbook["产品池_清洗版"].max_row != len(rows) + 1:
        raise ValueError("清洗版 Excel 产品行数不匹配")
    if clean_workbook["待校验清单"].max_row != sum(r["是否待校验"] == "是" for r in rows) + 1:
        raise ValueError("待校验清单行数不匹配")
    if clean_workbook["异常行隔离"].max_row != len(anomalies) + 1:
        raise ValueError("异常行隔离 Sheet 行数不匹配")

    for sheet in clean_workbook.worksheets:
        if sheet.freeze_panes != "A2" or not sheet.auto_filter.ref:
            raise ValueError(f"{sheet.title} 的冻结窗格或筛选未生效")

    main_workbook = load_workbook(MAIN_EXCEL)
    main_sheet = main_workbook["产品总表"]
    if main_sheet.max_row != len(rows) + 1:
        raise ValueError("主 Excel 同步行数不匹配")
    if main_sheet.freeze_panes != "A2" or not main_sheet.auto_filter.ref:
        raise ValueError("主 Excel 冻结窗格或筛选未生效")


def write_log(
    original_count: int,
    valid_count: int,
    anomalies: list[dict[str, str]],
    before_dedup_count: int,
    rows: list[dict[str, str]],
    duplicates: list[dict[str, str]],
) -> None:
    level_counts = Counter(record["纳入级别"] for record in rows)
    strategy_counts = Counter(record["策略大类"] for record in rows)
    pending_count = sum(record["是否待校验"] == "是" for record in rows)
    representative_count = sum(record["是否代表产品初筛"] == "是" for record in rows)

    strategy_lines = "\n".join(
        f"- {strategy}：{count}" for strategy, count in sorted(strategy_counts.items())
    )
    anomaly_lines = "无"
    if anomalies:
        anomaly_lines = "\n".join(
            f"- 第 {record['原始行号']} 行：{record['异常原因']}；原始内容：`{record['原始内容']}`"
            for record in anomalies
        )
    duplicate_lines = "无"
    if duplicates:
        duplicate_lines = "\n".join(
            f"- {record['产品名称']}（{record['基金代码']}）：{record['删除原因']}；去重键：{record['去重键']}"
            for record in duplicates
        )

    log = f"""# Step 5 产品池清洗日志

1. 输入文件路径：`{INPUT_CSV}`
2. 输出文件路径：
   - 清洗版 CSV：`{CLEAN_CSV}`
   - 清洗版 Excel：`{CLEAN_EXCEL}`
   - 异常行 CSV：`{ANOMALY_CSV}`
   - 同步主 Excel：`{MAIN_EXCEL}`
3. 原始读取行数：{original_count}
4. 有效产品行数：{valid_count}
5. 异常行数量：{len(anomalies)}
6. 去重前产品数量：{before_dedup_count}
7. 去重后产品数量：{len(rows)}
8. 删除重复行数量：{len(duplicates)}
9. 主线产品数量：{level_counts['主线']}
10. 补充观察产品数量：{level_counts['补充观察']}
11. 单列观察产品数量：{level_counts['单列观察']}
12. 跨境补充产品数量：{level_counts['跨境补充']}
13. 仅作线索产品数量：{level_counts['仅作线索']}
14. 待校验产品数量：{pending_count}
15. 代表产品初筛数量：{representative_count}
16. 各策略大类产品数量：
{strategy_lines}
17. 被隔离的异常行原始内容：
{anomaly_lines}
18. 被删除的重复行信息：
{duplicate_lines}
19. 是否成功同步写入主 Excel：是
"""
    LOG_PATH.write_text(log, encoding="utf-8")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    original_count, valid_input_rows, anomalies = read_input()
    cleaned_before_dedup = [clean_record(record) for record in valid_input_rows]
    cleaned_rows, duplicates = deduplicate(cleaned_before_dedup)

    write_csv(CLEAN_CSV, OUTPUT_FIELDS, cleaned_rows)
    write_csv(ANOMALY_CSV, ANOMALY_FIELDS, anomalies)
    write_clean_excel(cleaned_rows, anomalies)
    sync_main_excel(cleaned_rows)
    validate_outputs(cleaned_rows, anomalies)
    write_log(
        original_count,
        len(valid_input_rows),
        anomalies,
        len(cleaned_before_dedup),
        cleaned_rows,
        duplicates,
    )

    print(f"original={original_count}")
    print(f"valid={len(valid_input_rows)}")
    print(f"anomalies={len(anomalies)}")
    print(f"before_dedup={len(cleaned_before_dedup)}")
    print(f"after_dedup={len(cleaned_rows)}")
    print(f"duplicates={len(duplicates)}")
    print(f"pending={sum(r['是否待校验'] == '是' for r in cleaned_rows)}")
    print(f"representative={sum(r['是否代表产品初筛'] == '是' for r in cleaned_rows)}")
    print("strategy_counts=" + repr(dict(sorted(Counter(r['策略大类'] for r in cleaned_rows).items()))))


if __name__ == "__main__":
    main()
