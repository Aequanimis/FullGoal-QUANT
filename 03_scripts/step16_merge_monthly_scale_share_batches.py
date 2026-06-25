from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT_DIR = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池")
POOL_FILE = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\全市场ETF基础信息_策略ETF池二次修正版.xlsx")
POOL_SHEET = "策略ETF_最终统计池"
OUTPUT = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\广义策略ETF_月度规模份额表_五批合并清洗版.xlsx")
FILE_PATTERN = "*广义策略ETF_月度规模份额表*.xlsx"

MAIN_SHEET = "月度规模份额_清洗版"
COVERAGE_SHEET = "代码覆盖检查"
QUALITY_SHEET = "数据质量检查"
DUPLICATE_SHEET = "重复记录检查"
MONTHLY_TOTAL_SHEET = "月度汇总_广义与核心"
MONTHLY_STRATEGY_SHEET = "月度汇总_按一级策略"
MONTHLY_MANAGER_SHEET = "月度汇总_按基金公司"
UNIT_SHEET = "规模单位检查"

MAIN_FIELDS = [
    "日期",
    "Wind代码",
    "证券简称",
    "基金代码",
    "交易代码",
    "基金简称",
    "基金全称",
    "基金管理人",
    "基金上市地点",
    "上市日期",
    "基金成立日",
    "跟踪指数代码",
    "跟踪指数名称",
    "一级策略大类",
    "二级策略类别",
    "市场范围_二次修正",
    "统计口径分类",
    "是否纳入核心策略ETF统计",
    "是否纳入广义策略ETF统计",
    "基金规模_原始",
    "基金规模_原始单位",
    "基金规模_亿元",
    "基金份额_份",
    "基金份额_亿份",
    "单位净值",
    "复权单位净值",
    "累计净值",
    "单位净值币种",
    "最新基金规模(亿)",
    "来源批次",
    "来源文件",
    "数据备注",
]

POOL_FIELDS = [
    "基金代码",
    "交易代码",
    "基金简称",
    "基金全称",
    "基金管理人",
    "基金上市地点",
    "上市日期",
    "基金成立日",
    "跟踪指数代码",
    "跟踪指数名称",
    "最新基金规模(亿)",
    "一级策略大类",
    "二级策略类别",
    "市场范围_二次修正",
    "统计口径分类",
    "是否纳入核心策略ETF统计",
    "是否纳入广义策略ETF统计",
]

BATCH_ORDER = {"第一批": 1, "第二批": 2, "第三批": 3, "第四批": 4, "第五批": 5}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=9, color="1F1F1F")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN_GRAY = Side(style="thin", color="D9E1F2")
BODY_BORDER = Border(bottom=THIN_GRAY)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    value = str(value).replace("\u3000", " ").replace("\u200b", "")
    value = re.sub(r"[\r\n\t]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def normalized_header(value: Any) -> str:
    return re.sub(r"\s+", "", clean_text(value))


def normalize_code(value: Any) -> str:
    raw = clean_text(value).upper().replace("。", ".")
    match = re.fullmatch(r"(\d{6})\.(SH|SZ)", raw)
    return f"{match.group(1)}.{match.group(2)}" if match else raw


def to_number(value: Any) -> float | int | None:
    if value is None or clean_text(value) in {"", "--", "N/A", "nan"}:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    raw = clean_text(value).replace(",", "").replace("%", "")
    try:
        number = float(raw)
        return int(number) if number.is_integer() else number
    except ValueError:
        return None


def to_date(value: Any) -> datetime | None:
    if value is None or clean_text(value) == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    raw = clean_text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def find_field(headers: list[str], field_type: str) -> tuple[int | None, str | None]:
    for idx, header in enumerate(headers):
        h = normalized_header(header)
        if field_type == "wind_code" and h == "Wind代码":
            return idx, header
        if field_type == "security_name" and h == "证券简称":
            return idx, header
        if field_type == "date" and h == "日期":
            return idx, header
        if field_type == "scale" and "基金规模" in h:
            return idx, header
        if field_type == "share" and "基金份额" in h and "份" in h:
            return idx, header
        if field_type == "adjusted_nav" and "复权单位净值" in h:
            return idx, header
        if field_type == "cumulative_nav" and ("累计单位净值" in h or "累计净值" in h):
            return idx, header
        if (
            field_type == "unit_nav"
            and "单位净值" in h
            and "复权" not in h
            and "累计" not in h
            and "币种" not in h
        ):
            return idx, header
        if field_type == "currency" and h == "单位净值币种":
            return idx, header
    return None, None


def scale_unit_from_header(header: str | None) -> str:
    h = normalized_header(header)
    if "基金规模(亿)" in h or "基金规模（亿）" in h:
        return "亿"
    if "基金规模(百万)" in h or "基金规模（百万）" in h:
        return "百万"
    return "待确认"


def extract_batch(filename: str) -> str:
    for batch in BATCH_ORDER:
        if batch in filename:
            return batch
    match = re.search(r"第[一二三四五六七八九十]+批", filename)
    return match.group(0) if match else "批次待识别"


def comparable_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        return round(value, 10)
    return value


def raw_signature(record: dict[str, Any]) -> tuple[Any, ...]:
    fields = [
        "日期",
        "Wind代码",
        "证券简称",
        "基金规模_原始",
        "基金规模_原始单位",
        "基金规模_亿元",
        "基金份额_份",
        "基金份额_亿份",
        "单位净值",
        "复权单位净值",
        "累计净值",
        "单位净值币种",
    ]
    return tuple(comparable_value(record.get(field)) for field in fields)


def value_signature(record: dict[str, Any]) -> tuple[Any, ...]:
    fields = [
        "基金规模_亿元",
        "基金份额_份",
        "单位净值",
        "复权单位净值",
        "累计净值",
        "单位净值币种",
    ]
    return tuple(comparable_value(record.get(field)) for field in fields)


def write_records(ws, headers: list[str], records: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for record in records:
        ws.append([record.get(header) for header in headers])


def style_sheet(ws, headers: list[str]) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BODY_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
    sample_limit = min(ws.max_row, 300)
    wide_fields = {
        "基金全称",
        "跟踪指数名称",
        "数据备注",
        "说明",
        "异常/冲突类型",
        "原始规模字段名",
    }
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, sample_limit + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 55))
        width = min(max(max_len * 1.15, 11), 44 if header in wide_fields else 28)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for col_idx, header in enumerate(headers, 1):
        if "日期" in header or header in {"上市日期", "基金成立日", "日期最小值", "日期最大值"}:
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (datetime, date)):
                    ws.cell(row_idx, col_idx).number_format = "yyyy-mm-dd"
        if "规模" in header:
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "#,##0.0000"
        if "亿份" in header:
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "#,##0.0000"
        elif header == "基金份额_份":
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "#,##0"


def load_product_pool() -> tuple[dict[str, dict[str, Any]], list[str]]:
    wb = load_workbook(POOL_FILE, read_only=True, data_only=True)
    if POOL_SHEET not in wb.sheetnames:
        raise KeyError(f"产品池文件缺少sheet：{POOL_SHEET}")
    ws = wb[POOL_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows())]
    missing = [field for field in ["Wind代码", "是否纳入广义策略ETF统计", *POOL_FIELDS] if field not in headers]
    if missing:
        raise KeyError(f"产品池sheet缺少字段：{missing}")
    pool: dict[str, dict[str, Any]] = {}
    duplicate_codes: list[str] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, values))
        if clean_text(record.get("是否纳入广义策略ETF统计")) != "是":
            continue
        code = normalize_code(record.get("Wind代码"))
        if code in pool:
            duplicate_codes.append(code)
        else:
            pool[code] = record
    return pool, duplicate_codes


def main() -> None:
    files = [
        path
        for path in INPUT_DIR.glob(FILE_PATTERN)
        if path.resolve() != OUTPUT.resolve()
    ]
    files.sort(key=lambda p: (BATCH_ORDER.get(extract_batch(p.name), 999), p.name))
    if not files:
        raise FileNotFoundError(f"未在输入文件夹找到匹配文件：{FILE_PATTERN}")

    pool, pool_duplicate_codes = load_product_pool()
    pool_codes = set(pool)

    stats = Counter()
    batch_stats: dict[str, dict[str, Any]] = {}
    unit_rows: list[dict[str, Any]] = []
    parsed_records: list[dict[str, Any]] = []
    parse_notes: list[str] = []

    for path in files:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [clean_text(cell.value) for cell in next(ws.iter_rows())]
        detected = {
            field: find_field(headers, field)
            for field in [
                "wind_code",
                "security_name",
                "date",
                "scale",
                "share",
                "unit_nav",
                "adjusted_nav",
                "cumulative_nav",
                "currency",
            ]
        }
        required_missing = [
            field for field in ["wind_code", "date"] if detected[field][0] is None
        ]
        if required_missing:
            raise KeyError(f"{path.name} 缺少必要字段：{required_missing}")

        batch = extract_batch(path.name)
        scale_header = detected["scale"][1]
        unit = scale_unit_from_header(scale_header)
        batch_valid_codes: set[str] = set()
        batch_valid_rows = 0
        sample_raw = None
        sample_converted = None

        for values in ws.iter_rows(min_row=2, values_only=True):
            stats["原始合并行数"] += 1
            if not any(value is not None and clean_text(value) for value in values):
                stats["删除全空行数"] += 1
                continue

            code_idx = detected["wind_code"][0]
            raw_code = values[code_idx] if code_idx is not None else None
            code = normalize_code(raw_code)
            if clean_text(raw_code).startswith("数据来源：Wind"):
                stats["删除Wind来源行数"] += 1
                continue
            if not code:
                stats["删除Wind代码为空行数"] += 1
                continue
            if not re.fullmatch(r"\d{6}\.(SH|SZ)", code):
                stats["删除非SH_SZ代码行数"] += 1
                continue

            def value_for(field: str) -> Any:
                idx = detected[field][0]
                return values[idx] if idx is not None else None

            raw_scale = to_number(value_for("scale"))
            if unit == "亿":
                scale_billion = raw_scale
            elif unit == "百万":
                scale_billion = raw_scale / 100 if raw_scale is not None else None
            else:
                scale_billion = None
            raw_share = to_number(value_for("share"))
            share_billion = raw_share / 100000000 if raw_share is not None else None
            parsed_date = to_date(value_for("date"))

            notes: list[str] = []
            if unit == "待确认":
                notes.append("规模单位待确认")
            if value_for("date") not in (None, "") and parsed_date is None:
                notes.append("日期无法解析")
                stats["日期无法解析行数"] += 1

            record = {
                "日期": parsed_date,
                "Wind代码": code,
                "证券简称": clean_text(value_for("security_name")) or None,
                "基金规模_原始": raw_scale,
                "基金规模_原始单位": unit,
                "基金规模_亿元": scale_billion,
                "基金份额_份": raw_share,
                "基金份额_亿份": share_billion,
                "单位净值": to_number(value_for("unit_nav")),
                "复权单位净值": to_number(value_for("adjusted_nav")),
                "累计净值": to_number(value_for("cumulative_nav")),
                "单位净值币种": clean_text(value_for("currency")) or None,
                "来源批次": batch,
                "来源文件": path.name,
                "数据备注": "；".join(notes),
            }

            product = pool.get(code)
            if product:
                for field in POOL_FIELDS:
                    record[field] = product.get(field)
            else:
                for field in POOL_FIELDS:
                    record[field] = None
                record["数据备注"] = "；".join(filter(None, [record["数据备注"], "未匹配到广义策略ETF产品池"]))

            parsed_records.append(record)
            batch_valid_codes.add(code)
            batch_valid_rows += 1
            if sample_raw is None and raw_scale is not None:
                sample_raw = raw_scale
                sample_converted = scale_billion

        batch_stats[path.name] = {
            "来源批次": batch,
            "Wind代码数量": len(batch_valid_codes),
            "有效数据行数": batch_valid_rows,
            "代码集合": batch_valid_codes,
        }
        unit_rows.append(
            {
                "来源文件": path.name,
                "原始规模字段名": scale_header,
                "识别单位": unit,
                "样例数值": sample_raw,
                "转换后样例数值": sample_converted,
                "是否需要人工确认": "是" if unit == "待确认" else "否",
            }
        )
        if unit == "待确认":
            parse_notes.append(f"{path.name} 的规模字段单位待确认：{scale_header}")

    # Remove fully identical standardized rows, ignoring source batch/file.
    exact_seen: set[tuple[Any, ...]] = set()
    unique_records: list[dict[str, Any]] = []
    for record in parsed_records:
        signature = raw_signature(record)
        if signature in exact_seen:
            stats["删除完全重复行数"] += 1
            continue
        exact_seen.add(signature)
        unique_records.append(record)

    # Isolate conflicting Wind-code/date groups so the cleaned main table stays unique.
    groups: defaultdict[tuple[str, Any], list[dict[str, Any]]] = defaultdict(list)
    no_date_records: list[dict[str, Any]] = []
    for record in unique_records:
        if record["日期"] is None:
            no_date_records.append(record)
        else:
            groups[(record["Wind代码"], record["日期"])].append(record)

    main_records: list[dict[str, Any]] = []
    duplicate_rows: list[dict[str, Any]] = []
    for key, group in groups.items():
        if len(group) == 1:
            main_records.append(group[0])
            continue
        signatures = {value_signature(record) for record in group}
        conflict_type = "数值冲突" if len(signatures) > 1 else "重复记录_数值一致"
        for record in group:
            duplicate = dict(record)
            duplicate["异常/冲突类型"] = conflict_type
            duplicate_rows.append(duplicate)
        if conflict_type == "重复记录_数值一致":
            main_records.append(group[0])
            stats["合并同键同值重复记录数"] += len(group) - 1
        else:
            stats["隔离数值冲突记录数"] += len(group)
            stats["数值冲突键数量"] += 1
    main_records.extend(no_date_records)
    main_records.sort(
        key=lambda r: (
            r["日期"] if r["日期"] is not None else datetime.max,
            r["Wind代码"],
        )
    )

    actual_codes = {record["Wind代码"] for record in main_records}
    all_valid_codes = {record["Wind代码"] for record in unique_records}
    missing_codes = sorted(pool_codes - all_valid_codes)
    extra_codes = sorted(all_valid_codes - pool_codes)
    unmatched_codes = sorted(
        {record["Wind代码"] for record in unique_records if record["Wind代码"] not in pool}
    )

    valid_dates = [record["日期"] for record in main_records if record["日期"] is not None]
    min_date = min(valid_dates) if valid_dates else None
    max_date = max(valid_dates) if valid_dates else None

    # Monthly totals.
    monthly_groups: defaultdict[datetime, list[dict[str, Any]]] = defaultdict(list)
    for record in main_records:
        if record["日期"] is not None:
            monthly_groups[record["日期"]].append(record)

    monthly_total_rows: list[dict[str, Any]] = []
    monthly_strategy_rows: list[dict[str, Any]] = []
    monthly_manager_rows: list[dict[str, Any]] = []

    for month in sorted(monthly_groups):
        records = monthly_groups[month]
        broad_records = [
            r for r in records if clean_text(r.get("是否纳入广义策略ETF统计")) == "是"
        ]
        core_records = [
            r for r in broad_records if clean_text(r.get("是否纳入核心策略ETF统计")) == "是"
        ]
        enhanced_records = [
            r
            for r in broad_records
            if clean_text(r.get("统计口径分类")) == "广义策略ETF_指数增强"
        ]

        def sum_field(items: list[dict[str, Any]], field: str) -> float:
            return round(
                sum(float(r[field]) for r in items if isinstance(r.get(field), (int, float))),
                4,
            )

        def count_with_scale(items: list[dict[str, Any]]) -> int:
            return len({r["Wind代码"] for r in items if r.get("基金规模_亿元") is not None})

        monthly_total_rows.append(
            {
                "日期": month,
                "广义策略ETF总规模_亿元": sum_field(broad_records, "基金规模_亿元"),
                "核心策略指数ETF总规模_亿元": sum_field(core_records, "基金规模_亿元"),
                "指数增强多因子ETF总规模_亿元": sum_field(enhanced_records, "基金规模_亿元"),
                "广义策略ETF总份额_亿份": sum_field(broad_records, "基金份额_亿份"),
                "核心策略指数ETF总份额_亿份": sum_field(core_records, "基金份额_亿份"),
                "指数增强多因子ETF总份额_亿份": sum_field(enhanced_records, "基金份额_亿份"),
                "广义策略ETF有规模数据产品数": count_with_scale(broad_records),
                "核心策略指数ETF有规模数据产品数": count_with_scale(core_records),
                "指数增强多因子ETF有规模数据产品数": count_with_scale(enhanced_records),
            }
        )

        by_strategy: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
        by_manager: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
        for record in broad_records:
            by_strategy[clean_text(record.get("一级策略大类")) or "待补充"].append(record)
            by_manager[clean_text(record.get("基金管理人")) or "待补充"].append(record)

        for strategy, items in sorted(by_strategy.items()):
            monthly_strategy_rows.append(
                {
                    "日期": month,
                    "一级策略大类": strategy,
                    "产品数量": len({r["Wind代码"] for r in items}),
                    "有规模数据产品数量": count_with_scale(items),
                    "总规模_亿元": sum_field(items, "基金规模_亿元"),
                    "总份额_亿份": sum_field(items, "基金份额_亿份"),
                }
            )
        for manager, items in sorted(by_manager.items()):
            monthly_manager_rows.append(
                {
                    "日期": month,
                    "基金管理人": manager,
                    "产品数量": len({r["Wind代码"] for r in items}),
                    "有规模数据产品数量": count_with_scale(items),
                    "总规模_亿元": sum_field(items, "基金规模_亿元"),
                    "总份额_亿份": sum_field(items, "基金份额_亿份"),
                }
            )

    # Coverage sheet.
    coverage_rows: list[dict[str, Any]] = [
        {"检查类别": "总体", "项目": "产品池应有Wind代码数量", "数量": len(pool_codes), "说明": "是否纳入广义策略ETF统计=是"},
        {"检查类别": "总体", "项目": "五批合并实际覆盖Wind代码数量", "数量": len(all_valid_codes)},
        {"检查类别": "总体", "项目": "缺失Wind代码数量", "数量": len(missing_codes)},
        {"检查类别": "总体", "项目": "多余Wind代码数量", "数量": len(extra_codes)},
        {"检查类别": "总体", "项目": "代码是否全部以.SH或.SZ结尾", "说明": "是" if all(re.fullmatch(r"\d{6}\.(SH|SZ)", c) for c in all_valid_codes) else "否"},
        {"检查类别": "总体", "项目": "未匹配产品池分类字段代码数量", "数量": len(unmatched_codes)},
    ]
    for code in missing_codes:
        product = pool.get(code, {})
        coverage_rows.append(
            {
                "检查类别": "缺失代码清单",
                "项目": "缺失",
                "Wind代码": code,
                "证券简称": product.get("证券简称"),
                "说明": product.get("基金简称"),
            }
        )
    for code in extra_codes:
        coverage_rows.append({"检查类别": "多余代码清单", "项目": "多余", "Wind代码": code})
    for filename, info in batch_stats.items():
        coverage_rows.append(
            {
                "检查类别": "批次覆盖",
                "项目": info["来源批次"],
                "数量": info["Wind代码数量"],
                "有效数据行数": info["有效数据行数"],
                "说明": filename,
            }
        )
    for code in unmatched_codes:
        coverage_rows.append(
            {"检查类别": "未匹配产品池分类", "项目": "未匹配", "Wind代码": code}
        )

    quality_rows: list[dict[str, Any]] = []

    def add_quality(category: str, item: str, value: Any, note: str = "") -> None:
        quality_rows.append({"统计类别": category, "指标": item, "数值": value, "说明": note})

    add_quality("总体", "读取文件数量", len(files))
    add_quality("总体", "原始合并行数", stats["原始合并行数"])
    add_quality("清洗删除", "删除“数据来源：Wind”行数", stats["删除Wind来源行数"])
    add_quality("清洗删除", "删除Wind代码为空行数", stats["删除Wind代码为空行数"])
    add_quality("清洗删除", "删除全空行数", stats["删除全空行数"])
    add_quality("清洗删除", "删除非.SH/.SZ代码行数", stats["删除非SH_SZ代码行数"])
    add_quality("清洗删除", "删除完全重复行数", stats["删除完全重复行数"])
    add_quality("重复处理", "合并同键同值重复记录数", stats["合并同键同值重复记录数"])
    add_quality("重复处理", "数值冲突键数量", stats["数值冲突键数量"])
    add_quality("重复处理", "隔离数值冲突记录数", stats["隔离数值冲突记录数"])
    add_quality("结果", "清洗后有效数据行数", len(main_records))
    add_quality("结果", "清洗后有效Wind代码数量", len(actual_codes))
    add_quality("结果", "日期最小值", min_date)
    add_quality("结果", "日期最大值", max_date)
    for field in [
        "基金规模_原始",
        "基金规模_亿元",
        "基金份额_份",
        "单位净值",
        "复权单位净值",
        "累计净值",
    ]:
        add_quality("缺失值", f"{field}缺失数量", sum(r.get(field) is None for r in main_records))
    unit_counts = Counter(r.get("基金规模_原始单位") or "待确认" for r in main_records)
    for unit in ["亿", "百万", "待确认"]:
        add_quality("规模原始单位分布", unit, unit_counts.get(unit, 0))
    for month, records in sorted(monthly_groups.items()):
        add_quality(
            "每月覆盖ETF数量",
            month.strftime("%Y-%m-%d"),
            len({record["Wind代码"] for record in records}),
        )
    code_counts = Counter(record["Wind代码"] for record in main_records)
    for code in sorted(code_counts):
        add_quality("每个Wind代码月度记录数量", code, code_counts[code])
    if pool_duplicate_codes:
        add_quality("产品池问题", "产品池重复Wind代码数量", len(pool_duplicate_codes), "；".join(pool_duplicate_codes))
    for note in parse_notes:
        add_quality("字段识别提示", "提示", None, note)

    wb = Workbook()
    wb.remove(wb.active)
    main_ws = wb.create_sheet(MAIN_SHEET)
    coverage_ws = wb.create_sheet(COVERAGE_SHEET)
    quality_ws = wb.create_sheet(QUALITY_SHEET)
    duplicate_ws = wb.create_sheet(DUPLICATE_SHEET)
    monthly_total_ws = wb.create_sheet(MONTHLY_TOTAL_SHEET)
    monthly_strategy_ws = wb.create_sheet(MONTHLY_STRATEGY_SHEET)
    monthly_manager_ws = wb.create_sheet(MONTHLY_MANAGER_SHEET)
    unit_ws = wb.create_sheet(UNIT_SHEET)

    write_records(main_ws, MAIN_FIELDS, main_records)

    coverage_headers = ["检查类别", "项目", "数量", "有效数据行数", "Wind代码", "证券简称", "说明"]
    write_records(coverage_ws, coverage_headers, coverage_rows)
    quality_headers = ["统计类别", "指标", "数值", "说明"]
    write_records(quality_ws, quality_headers, quality_rows)

    duplicate_headers = MAIN_FIELDS + ["异常/冲突类型"]
    if duplicate_rows:
        write_records(duplicate_ws, duplicate_headers, duplicate_rows)
    else:
        duplicate_ws.append(duplicate_headers)
        duplicate_ws.append(["无重复 Wind代码-日期记录"] + [None] * (len(duplicate_headers) - 1))

    monthly_total_headers = [
        "日期",
        "广义策略ETF总规模_亿元",
        "核心策略指数ETF总规模_亿元",
        "指数增强多因子ETF总规模_亿元",
        "广义策略ETF总份额_亿份",
        "核心策略指数ETF总份额_亿份",
        "指数增强多因子ETF总份额_亿份",
        "广义策略ETF有规模数据产品数",
        "核心策略指数ETF有规模数据产品数",
        "指数增强多因子ETF有规模数据产品数",
    ]
    write_records(monthly_total_ws, monthly_total_headers, monthly_total_rows)
    monthly_strategy_headers = ["日期", "一级策略大类", "产品数量", "有规模数据产品数量", "总规模_亿元", "总份额_亿份"]
    write_records(monthly_strategy_ws, monthly_strategy_headers, monthly_strategy_rows)
    monthly_manager_headers = ["日期", "基金管理人", "产品数量", "有规模数据产品数量", "总规模_亿元", "总份额_亿份"]
    write_records(monthly_manager_ws, monthly_manager_headers, monthly_manager_rows)
    unit_headers = ["来源文件", "原始规模字段名", "识别单位", "样例数值", "转换后样例数值", "是否需要人工确认"]
    write_records(unit_ws, unit_headers, unit_rows)

    for ws, headers in [
        (main_ws, MAIN_FIELDS),
        (coverage_ws, coverage_headers),
        (quality_ws, quality_headers),
        (duplicate_ws, duplicate_headers),
        (monthly_total_ws, monthly_total_headers),
        (monthly_strategy_ws, monthly_strategy_headers),
        (monthly_manager_ws, monthly_manager_headers),
        (unit_ws, unit_headers),
    ]:
        style_sheet(ws, headers)

    for ws in [coverage_ws, quality_ws]:
        previous = None
        for row_idx in range(2, ws.max_row + 1):
            section = ws.cell(row_idx, 1).value
            if section != previous:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row_idx, col_idx).fill = SECTION_FILL
                    ws.cell(row_idx, col_idx).font = Font(name="微软雅黑", size=9, bold=True)
            previous = section

    # Quality dates live in the generic “数值” column, so format them explicitly.
    for row_idx in range(2, quality_ws.max_row + 1):
        if quality_ws.cell(row_idx, 2).value in {"日期最小值", "日期最大值"}:
            quality_ws.cell(row_idx, 3).number_format = "yyyy-mm-dd"

    if duplicate_rows:
        conflict_col = duplicate_headers.index("异常/冲突类型") + 1
        for row_idx in range(2, duplicate_ws.max_row + 1):
            if duplicate_ws.cell(row_idx, conflict_col).value == "数值冲突":
                for col_idx in range(1, duplicate_ws.max_column + 1):
                    duplicate_ws.cell(row_idx, col_idx).fill = WARNING_FILL

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    print(f"读取文件数量：{len(files)}")
    for row in unit_rows:
        print(f"{row['来源文件']}：{row['原始规模字段名']} -> {row['识别单位']}")
    print(f"清洗后有效数据行数：{len(main_records)}")
    print(f"覆盖Wind代码数量：{len(all_valid_codes)}")
    print(f"缺失代码数量：{len(missing_codes)}")
    print(f"多余代码数量：{len(extra_codes)}")
    print(
        f"日期范围：{min_date.strftime('%Y-%m-%d') if min_date else '无'} 至 "
        f"{max_date.strftime('%Y-%m-%d') if max_date else '无'}"
    )
    print(f"是否存在重复Wind代码+日期：{'是' if duplicate_rows else '否'}")
    print(f"输出文件路径：{OUTPUT}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"处理失败：{exc}", file=sys.stderr)
        raise
