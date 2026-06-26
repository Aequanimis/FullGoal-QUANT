from pathlib import Path

from openpyxl import load_workbook


path = (
    Path.home()
    / "Desktop"
    / "Fullgoal"
    / "课题研究"
    / "处理后数据"
    / "wind代码池"
    / "sheet7"
    / "核心策略ETF_跟踪指数规则_官网补充核验版.xlsx"
)

wb = load_workbook(path, read_only=True, data_only=True)
print("FILE_SIZE", path.stat().st_size)
print("SHEETS", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(name, ws.max_row, ws.max_column)

ws1 = wb["核心ETF_指数映射_标准版"]
h1 = [c.value for c in next(ws1.iter_rows(max_row=1))]
r1 = list(ws1.iter_rows(min_row=2, values_only=True))
print("SHEET1_ROWS", len(r1), "ETF_UNIQUE", len({r[h1.index('ETF_Code')] for r in r1}))

ws3 = wb["指数规则明细_官网补充版"]
h3 = [c.value for c in next(ws3.iter_rows(max_row=1))]
r3 = list(ws3.iter_rows(min_row=2, values_only=True))
print("SHEET3_ROWS", len(r3), "INDEX_UNIQUE", len({r[h3.index('Index_Code')] for r in r3}))
print("CORE_COMPLETE", sum(1 for r in r3 if r[h3.index("规则完整性状态")] == "核心规则已补全"))
print("BASIC_VERIFY", sum(1 for r in r3 if isinstance(r[h3.index("规则完整性状态")], str) and "基本补全" in r[h3.index("规则完整性状态")]))
print("PARTIAL", sum(1 for r in r3 if isinstance(r[h3.index("规则完整性状态")], str) and "部分补全" in r[h3.index("规则完整性状态")]))
print("NEED_VERIFY", sum(1 for r in r3 if r[h3.index("是否仍需官网核验")] in ("是", "建议核验")))

ws4 = wb["ETF_指数规则合并表_官网补充版"]
h4 = [c.value for c in next(ws4.iter_rows(max_row=1))]
r4 = list(ws4.iter_rows(min_row=2, values_only=True))
map1 = {r[h1.index("ETF_Code")]: r[h1.index("Index_Code")] for r in r1}
mismatch = 0
for r in r4:
    if map1.get(r[h4.index("ETF_Code")]) != r[h4.index("Index_Code")]:
        mismatch += 1
print("SHEET4_ROWS", len(r4), "MISMATCH", mismatch)

old_tokens = {"Wind暂无": 0, "待补充": 0, "待官网补充": 0}
blank_count = 0
for name in ["指数规则明细_官网补充版", "ETF_指数规则合并表_官网补充版"]:
    ws = wb[name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for value in row:
            text = "" if value is None else str(value).strip()
            if text == "":
                blank_count += 1
            for token in old_tokens:
                if token in text:
                    old_tokens[token] += 1
print("BLANK_COUNT_RULE_OUTPUTS", blank_count)
print("OLD_TOKENS", old_tokens)
