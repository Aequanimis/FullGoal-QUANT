from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]

INPUT_PRODUCT_POOL = ROOT / "01_processed_data" / "classification" / "境内策略ETF产品池_代表产品筛选版.xlsx"
INPUT_STRATEGY_TABLE = ROOT / "01_processed_data" / "classification" / "策略分类表.xlsx"
INPUT_STATISTICS = ROOT / "01_processed_data" / "statistics" / "策略ETF统计分析表.xlsx"
MAIN_EXCEL = ROOT / "02_outputs" / "excel" / "境内策略ETF产品梳理_初版.xlsx"

OUTPUT_PRODUCT_POOL = ROOT / "01_processed_data" / "classification" / "境内策略ETF产品池_mentor分类调整版.xlsx"
OUTPUT_STRATEGY_TABLE = ROOT / "01_processed_data" / "classification" / "策略分类表_mentor分类调整版.xlsx"
OUTPUT_STATISTICS = ROOT / "01_processed_data" / "statistics" / "策略ETF统计分析表_mentor分类调整版.xlsx"
OUTPUT_MD = ROOT / "02_outputs" / "summary" / "mentor分类口径调整说明.md"
OUTPUT_LOG = ROOT / "01_processed_data" / "classification" / "step11_mentor_classification_adjustment_log.md"

NEW_PRODUCT_FIELDS = [
    "一级策略大类",
    "二级策略类别",
    "市场范围",
    "分类口径说明",
    "是否红利加策略",
]

NEW_STRATEGY_HEADERS = [
    "一级策略大类",
    "二级策略类别",
    "核心因子",
    "投资逻辑",
    "代表指数",
    "代表ETF",
    "市场范围",
    "适合市场环境",
    "主要风险",
    "境内发展成熟度",
    "备注",
]

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)


def s(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").replace("\t", " ").strip()


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def read_sheet(path: Path, preferred_sheet: str) -> Tuple[List[str], List[Dict[str, str]], str]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    wb = load_workbook(path, data_only=False)
    ws = wb[preferred_sheet] if preferred_sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [s(cell.value) for cell in ws[1]]
    rows: List[Dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        values = [s(v) for v in row]
        if not any(values):
            continue
        rows.append(dict(zip(headers, values)))
    actual_sheet = ws.title
    wb.close()
    return headers, rows, actual_sheet


def first_nonempty(row: Dict[str, str], names: Iterable[str]) -> str:
    for name in names:
        value = s(row.get(name))
        if value:
            return value
    return ""


def contains_any(text: str, keywords: Iterable[str]) -> bool:
    return any(k in text for k in keywords)


def combined_text(row: Dict[str, str]) -> str:
    fields = [
        "建议策略大类",
        "策略大类",
        "建议策略细分",
        "策略细分",
        "产品名称",
        "跟踪指数",
        "跟踪指数或策略线索",
    ]
    return " ".join(s(row.get(f)) for f in fields)


def product_classification(row: Dict[str, str]) -> Dict[str, str]:
    major = first_nonempty(row, ["建议策略大类", "策略大类"])
    minor = first_nonempty(row, ["建议策略细分", "策略细分"])
    product_name = s(row.get("产品名称"))
    index_name = s(row.get("跟踪指数") or row.get("跟踪指数或策略线索"))
    text = combined_text(row)

    result = {
        "一级策略大类": "",
        "二级策略类别": "",
        "市场范围": "",
        "分类口径说明": "",
        "是否红利加策略": "否",
    }

    def set_result(primary: str, secondary: str, market: str, is_dividend: str, note: str) -> Dict[str, str]:
        result["一级策略大类"] = primary
        result["二级策略类别"] = secondary
        result["市场范围"] = market
        result["是否红利加策略"] = is_dividend
        result["分类口径说明"] = note
        return result

    if "沪深港红利" in text:
        return set_result(
            "红利",
            "沪深港红利",
            "AH跨市场",
            "是",
            "沪深港红利本质是红利策略，市场范围为 AH 跨市场。",
        )

    if major == "港股通红利/低波" or contains_any(text, ["港股通红利", "港股通低波红利", "标普港股通低波红利"]):
        secondary = "港股通低波红利" if "低波" in text else "港股通红利"
        return set_result(
            "红利",
            secondary,
            "港股通/H",
            "是",
            "港股通红利本质仍是红利策略，只是底层市场为港股通/H，因此归入红利一级大类，市场范围单独标注。",
        )

    if major == "央国企红利/股东回报" or contains_any(
        text, ["央企股东回报", "央企红利", "中央企业红利", "国企红利", "国有企业红利"]
    ):
        if "央企股东回报" in text:
            secondary = "央企股东回报"
        elif contains_any(text, ["央企红利", "中央企业红利"]):
            secondary = "央企红利"
        elif contains_any(text, ["国企红利", "国有企业红利"]):
            secondary = "国企红利"
        else:
            secondary = "央国企红利/股东回报"
        return set_result(
            "红利",
            secondary,
            "A股",
            "是",
            "央国企红利/股东回报兼具主题属性和红利/股东回报策略属性，按 mentor 建议归入红利一级大类下的二级类别。",
        )

    if major == "红利低波" or contains_any(text, ["红利低波", "低波红利"]):
        return set_result(
            "红利",
            "红利低波",
            "A股",
            "是",
            "红利低波属于“红利+低波”的红利加策略，调整为红利一级大类下的二级类别。",
        )

    if major == "红利质量" or "红利质量" in text:
        return set_result(
            "红利",
            "红利质量",
            "A股",
            "是",
            "红利质量属于“红利+质量”的红利加策略，调整为红利一级大类下的二级类别。",
        )

    if major == "红利" or major == "高股息" or contains_any(text, ["高股息", "股息率"]):
        return set_result(
            "红利",
            "普通红利",
            "A股",
            "是",
            "按 mentor 建议，红利为一级策略大类，普通红利作为二级类别。",
        )

    if major == "自由现金流" or "自由现金流" in text:
        if contains_any(text, ["国证自由现金流"]):
            secondary = "国证自由现金流"
        elif contains_any(text, ["中证全指自由现金流"]):
            secondary = "中证全指自由现金流"
        elif contains_any(text, ["中证800自由现金流"]):
            secondary = "中证800自由现金流"
        elif contains_any(text, ["沪深300自由现金流"]):
            secondary = "沪深300自由现金流"
        else:
            secondary = "自由现金流"
        return set_result(
            "自由现金流",
            secondary,
            "A股",
            "否",
            "自由现金流作为一级策略大类保留。",
        )

    if major == "价值":
        return set_result("价值", minor or "价值", "A股", "否", "价值作为一级策略大类保留。")

    if major == "成长":
        return set_result("成长", minor or "成长", "A股", "否", "成长作为一级策略大类保留。")

    if major == "低波":
        return set_result("低波", minor or "低波", "A股", "否", "低波作为一级策略大类保留。")

    if major == "质量":
        return set_result("质量", minor or "质量", "A股", "否", "质量作为一级策略大类保留。")

    if major == "基本面策略":
        return set_result("基本面策略", "基本面加权", "A股", "否", "基本面策略作为一级策略大类保留。")

    if major in {"指数增强/多因子", "多因子/增强"} or contains_any(major, ["增强", "多因子"]):
        return set_result(
            "指数增强/多因子",
            "指数增强",
            "A股",
            "否",
            "指数增强 ETF 不等同于纯 Smart Beta 策略 ETF，继续单列观察。",
        )

    if major == "ESG" or "ESG" in text or "可持续发展" in text:
        if "可持续发展" in text:
            secondary = "可持续发展"
        elif "央企ESG" in text:
            secondary = "央企ESG"
        else:
            secondary = "ESG"
        return set_result("ESG", secondary, "A股", "否", "ESG 作为一级策略大类保留。")

    fallback_primary = major or first_nonempty(row, ["原策略大类"]) or "其他待补充"
    fallback_secondary = minor or first_nonempty(row, ["原策略细分"]) or "待校验"
    return set_result(
        fallback_primary,
        fallback_secondary,
        "待校验",
        "否",
        "暂按原分类保留，后续人工校验。",
    )


def add_or_update_fields(headers: List[str], rows: List[Dict[str, str]]) -> Tuple[List[str], List[Dict[str, str]]]:
    out_headers = list(headers)
    for field in NEW_PRODUCT_FIELDS:
        if field not in out_headers:
            out_headers.append(field)
    out_rows: List[Dict[str, str]] = []
    for row in rows:
        new_row = dict(row)
        new_row.update(product_classification(row))
        out_rows.append(new_row)
    return out_headers, out_rows


def format_sheet(ws, code_header: str = "基金代码") -> None:
    if ws.max_row >= 1:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    code_col = None
    for cell in ws[1]:
        if s(cell.value) == code_header:
            code_col = cell.column
            break

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            value = s(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if code_col and col_idx == code_col:
                cell.number_format = "@"
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)


def write_rows(ws, headers: List[str], rows: List[Dict[str, str]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    format_sheet(ws)


def write_table(ws, headers: List[str], matrix: List[List]) -> None:
    ws.append(headers)
    for row in matrix:
        ws.append(row)
    format_sheet(ws)


def save_product_workbook(headers: List[str], rows: List[Dict[str, str]]) -> None:
    ensure_parent(OUTPUT_PRODUCT_POOL)
    wb = Workbook()
    ws = wb.active
    ws.title = "产品池_mentor分类调整版"
    write_rows(ws, headers, rows)

    primary_counts = Counter(r["一级策略大类"] for r in rows)
    red_rows = [r for r in rows if r["一级策略大类"] == "红利"]

    ws2 = wb.create_sheet("红利一级检查")
    red_headers = ["基金代码", "产品名称", "一级策略大类", "二级策略类别", "市场范围", "是否红利加策略", "分类口径说明"]
    write_rows(ws2, red_headers, red_rows)

    ws3 = wb.create_sheet("一级策略统计")
    write_table(ws3, ["一级策略大类", "产品数量"], [[k, v] for k, v in primary_counts.most_common()])

    wb.save(OUTPUT_PRODUCT_POOL)
    wb.close()


def read_strategy_rows() -> Tuple[List[Dict[str, str]], List[str]]:
    headers, rows, _ = read_sheet(INPUT_STRATEGY_TABLE, "策略分类表")
    return rows, headers


def build_strategy_table(original_rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    by_type = {s(r.get("策略类型")): r for r in original_rows}

    def from_source(primary: str, secondary: str, source_type: str, market: str = "A股", extra_note: str = "") -> Dict[str, str]:
        src = by_type.get(source_type, {})
        note = s(src.get("备注"))
        if extra_note:
            note = f"{note}；{extra_note}" if note else extra_note
        return {
            "一级策略大类": primary,
            "二级策略类别": secondary,
            "核心因子": s(src.get("核心因子")),
            "投资逻辑": s(src.get("投资逻辑")),
            "代表指数": s(src.get("代表指数")),
            "代表ETF": s(src.get("代表ETF")),
            "市场范围": market,
            "适合市场环境": s(src.get("适合市场环境")),
            "主要风险": s(src.get("主要风险")),
            "境内发展成熟度": s(src.get("境内发展成熟度")),
            "备注": note,
        }

    rows = [
        from_source("红利", "普通红利", "红利", "A股", "按 mentor 口径，普通红利作为红利一级大类下的二级类别。"),
        from_source("红利", "红利低波", "红利低波", "A股", "红利低波不是被删除，而是作为红利一级大类下的二级策略保留。"),
        from_source("红利", "红利质量", "红利质量", "A股", "红利质量不是被删除，而是作为红利一级大类下的二级策略保留。"),
        from_source(
            "红利",
            "央企红利/国企红利/股东回报",
            "央国企红利/股东回报",
            "A股",
            "央国企红利/股东回报归入红利一级大类，主题属性在二级类别和备注中表达。",
        ),
        from_source(
            "红利",
            "港股通红利/低波",
            "港股通红利/低波",
            "港股通/H",
            "港股通红利/低波归入红利一级大类，市场范围单独标注。",
        ),
        from_source("自由现金流", "自由现金流", "自由现金流"),
        from_source("质量", "质量", "质量"),
        from_source("价值", "价值", "价值"),
        from_source("成长", "成长", "成长"),
        from_source("低波", "低波", "低波"),
        from_source("基本面策略", "基本面加权", "基本面策略"),
        from_source("指数增强/多因子", "指数增强", "指数增强/多因子"),
        from_source("ESG", "ESG", "ESG"),
        from_source("等权/另类加权", "等权/另类加权", "等权/另类加权"),
        from_source("动量", "动量", "动量"),
    ]
    return rows


def save_strategy_workbook(rows: List[Dict[str, str]], output_path: Path) -> None:
    ensure_parent(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "策略分类表_mentor调整版"
    write_rows(ws, NEW_STRATEGY_HEADERS, rows)

    ws2 = wb.create_sheet("红利二级类别")
    red_rows = [r for r in rows if r["一级策略大类"] == "红利"]
    write_rows(ws2, NEW_STRATEGY_HEADERS, red_rows)

    wb.save(output_path)
    wb.close()


def save_statistics_workbook(rows: List[Dict[str, str]]) -> Dict[str, Counter]:
    ensure_parent(OUTPUT_STATISTICS)
    wb = Workbook()

    primary_counts = Counter(r["一级策略大类"] for r in rows)
    cross_counts = Counter((r["一级策略大类"], r["二级策略类别"]) for r in rows)
    market_counts = Counter(r["市场范围"] for r in rows)
    dividend_counts = Counter(r["是否红利加策略"] for r in rows)
    red_rows = [r for r in rows if r["一级策略大类"] == "红利"]

    ws = wb.active
    ws.title = "一级策略大类统计"
    write_table(ws, ["一级策略大类", "产品数量"], [[k, v] for k, v in primary_counts.most_common()])

    ws = wb.create_sheet("一级策略_二级类别交叉")
    cross_matrix = [[k[0], k[1], v] for k, v in sorted(cross_counts.items(), key=lambda item: (item[0][0], item[0][1]))]
    write_table(ws, ["一级策略大类", "二级策略类别", "产品数量"], cross_matrix)

    ws = wb.create_sheet("市场范围统计")
    write_table(ws, ["市场范围", "产品数量"], [[k, v] for k, v in market_counts.most_common()])

    ws = wb.create_sheet("红利加策略统计")
    write_table(ws, ["统计项", "数量"], [["是否红利加策略=是", dividend_counts.get("是", 0)], ["是否红利加策略=否", dividend_counts.get("否", 0)]])
    start_row = ws.max_row + 3
    red_headers = ["基金代码", "产品名称", "一级策略大类", "二级策略类别", "市场范围", "是否红利加策略"]
    for col, header in enumerate(red_headers, start=1):
        ws.cell(start_row, col, header)
    for r_idx, row in enumerate(red_rows, start=start_row + 1):
        values = [row.get(h, "") for h in red_headers]
        for c_idx, value in enumerate(values, start=1):
            ws.cell(r_idx, c_idx, value)
    format_sheet(ws)

    ws = wb.create_sheet("调整说明")
    notes = [
        ["说明项", "内容"],
        ["调整原因", "mentor 建议先分一级策略大类，再按二级类别和市场范围细分。"],
        ["一级策略", "一级策略看核心策略逻辑，例如红利、自由现金流、价值、成长、低波等。"],
        ["二级类别", "二级类别承接红利低波、红利质量、央企红利、港股通红利等细分方向。"],
        ["市场范围", "市场范围单独标注 A股、港股通/H、AH跨市场。"],
        ["红利策略处理", "红利低波和红利质量不是被删除，而是作为红利一级大类下的二级策略保留。"],
    ]
    for row in notes:
        ws.append(row)
    format_sheet(ws)

    wb.save(OUTPUT_STATISTICS)
    wb.close()

    return {
        "primary": primary_counts,
        "cross": cross_counts,
        "market": market_counts,
        "dividend": dividend_counts,
    }


def write_markdown(stats: Dict[str, Counter]) -> None:
    ensure_parent(OUTPUT_MD)
    primary_lines = "\n".join(f"- {k}：{v} 只" for k, v in stats["primary"].most_common())
    red_secondary = Counter({k[1]: v for k, v in stats["cross"].items() if k[0] == "红利"})
    red_lines = "\n".join(f"- {k}：{v} 只" for k, v in red_secondary.most_common())
    content = f"""# Mentor 分类口径调整说明

## 1. 本次调整原因

mentor 建议先分一级策略大类，再按二级类别和市场范围细分，不要把港股通红利/低波等市场范围因素和策略类型放在同一级。

## 2. 本次调整原则

- 一级策略看核心策略逻辑；
- 二级类别看红利低波、红利质量、央企红利、港股通红利等细分；
- 市场范围单独标注 A股、港股通/H、AH跨市场。

## 3. 红利大类调整

红利低波、红利质量、央国企红利/股东回报、港股通红利/低波都归入一级策略“红利”下面。

它们没有被删除，而是作为红利大类下的二级类别保留。

## 4. 调整后的示例

- 中证红利 ETF：一级策略大类=红利；二级策略类别=普通红利；市场范围=A股。
- 中证红利低波动 ETF：一级策略大类=红利；二级策略类别=红利低波；市场范围=A股。
- 中证红利质量 ETF：一级策略大类=红利；二级策略类别=红利质量；市场范围=A股。
- 港股通低波红利 ETF：一级策略大类=红利；二级策略类别=港股通低波红利；市场范围=港股通/H。
- 沪深港红利 ETF：一级策略大类=红利；二级策略类别=沪深港红利；市场范围=AH跨市场。

## 5. 后续影响

后续汇报时，一级分类会更清晰；红利可以作为最大的策略母类；红利低波、红利质量、央企红利、港股通红利等作为红利+延伸方向展开；这样也更方便后续和海外红利、Dividend Growth、High Dividend、Low Vol High Dividend 等策略做对比。

## 6. 一级策略大类统计

{primary_lines}

## 7. 红利一级大类下各二级类别数量

{red_lines}
"""
    OUTPUT_MD.write_text(content, encoding="utf-8")


def product_key(row: Dict[str, str]) -> Tuple[str, ...]:
    code = s(row.get("基金代码"))
    if code and code != "待校验":
        return ("code", code)
    return ("combo", s(row.get("产品名称")), s(row.get("基金公司")), s(row.get("跟踪指数")))


def sync_main_excel(product_rows: List[Dict[str, str]], strategy_rows: List[Dict[str, str]]) -> bool:
    if not MAIN_EXCEL.exists():
        return False

    wb = load_workbook(MAIN_EXCEL)

    if "产品总表" in wb.sheetnames:
        ws = wb["产品总表"]
        headers = [s(cell.value) for cell in ws[1]]
        for field in NEW_PRODUCT_FIELDS:
            if field not in headers:
                headers.append(field)
                ws.cell(1, len(headers), field)

        header_to_col = {h: i + 1 for i, h in enumerate(headers)}
        lookup = {product_key(r): r for r in product_rows}

        for row_idx in range(2, ws.max_row + 1):
            current = {h: s(ws.cell(row_idx, c).value) for h, c in header_to_col.items() if c <= ws.max_column}
            match = lookup.get(product_key(current))
            if match is None:
                # Main Excel follows 产品名称 + 基金代码; if 跟踪指数 is blank in old rows, try code-only again.
                code = s(current.get("基金代码"))
                if code and code != "待校验":
                    match = lookup.get(("code", code))
            if match:
                for field in NEW_PRODUCT_FIELDS:
                    ws.cell(row_idx, header_to_col[field], match.get(field, ""))
        format_sheet(ws)

    sheet_name = "策略分类表_mentor调整版"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    write_rows(ws, NEW_STRATEGY_HEADERS, strategy_rows)

    wb.save(MAIN_EXCEL)
    wb.close()
    return True


def write_log(
    product_sheet: str,
    product_row_count: int,
    adjusted_count: int,
    stats: Dict[str, Counter],
    success_flags: Dict[str, bool],
    missing_inputs: List[str],
) -> None:
    ensure_parent(OUTPUT_LOG)
    primary_counts = stats["primary"]
    cross = stats["cross"]
    market = stats["market"]
    red_secondary_lines = "\n".join(
        f"- {secondary}：{count} 只"
        for (primary, secondary), count in sorted(cross.items(), key=lambda item: (item[0][0], item[0][1]))
        if primary == "红利"
    )
    primary_lines = "\n".join(f"- {k}：{v} 只" for k, v in primary_counts.most_common())

    content = f"""# Step 11 Mentor 分类口径调整日志

## 1. 输入文件路径

- 产品池：{INPUT_PRODUCT_POOL}
- 策略分类表：{INPUT_STRATEGY_TABLE}
- 统计分析表：{INPUT_STATISTICS}
- 主 Excel：{MAIN_EXCEL}

## 2. 输出文件路径

- 调整后产品池 Excel：{OUTPUT_PRODUCT_POOL}
- 调整后策略分类表 Excel：{OUTPUT_STRATEGY_TABLE}
- 调整后统计表 Excel：{OUTPUT_STATISTICS}
- Mentor 分类调整说明 Markdown：{OUTPUT_MD}
- 调整日志：{OUTPUT_LOG}

## 3. 输入文件缺失情况

{chr(10).join(f"- {p}" for p in missing_inputs) if missing_inputs else "- 无"}

## 4. 产品池读取情况

- 读取 sheet：{product_sheet}
- 产品池读取行数：{product_row_count}
- 成功新增一级/二级分类字段的行数：{adjusted_count}

## 5. 一级策略大类统计

{primary_lines}

## 6. 红利相关统计

- 红利一级大类产品数量：{primary_counts.get("红利", 0)}
- 红利低波二级类别产品数量：{cross.get(("红利", "红利低波"), 0)}
- 红利质量二级类别产品数量：{cross.get(("红利", "红利质量"), 0)}
- 港股通/H 市场范围产品数量：{market.get("港股通/H", 0)}

### 红利一级大类下各二级类别数量

{red_secondary_lines}

## 7. 生成状态

- 是否成功生成 mentor 调整版产品池：{"是" if success_flags.get("product_pool") else "否"}
- 是否成功生成 mentor 调整版策略分类表：{"是" if success_flags.get("strategy_table") else "否"}
- 是否成功生成 mentor 调整版统计表：{"是" if success_flags.get("statistics") else "否"}
- 是否成功同步主 Excel：{"是" if success_flags.get("main_excel") else "否"}
- 是否成功生成 Markdown 说明：{"是" if success_flags.get("markdown") else "否"}
"""
    OUTPUT_LOG.write_text(content, encoding="utf-8")


def main() -> None:
    missing_inputs = [str(p) for p in [INPUT_PRODUCT_POOL, INPUT_STRATEGY_TABLE, INPUT_STATISTICS, MAIN_EXCEL] if not p.exists()]

    product_headers: List[str] = []
    product_rows: List[Dict[str, str]] = []
    product_sheet = ""
    if INPUT_PRODUCT_POOL.exists():
        product_headers, product_rows, product_sheet = read_sheet(INPUT_PRODUCT_POOL, "产品池_代表产品筛选版")

    strategy_rows_original: List[Dict[str, str]] = []
    if INPUT_STRATEGY_TABLE.exists():
        strategy_rows_original, _ = read_strategy_rows()

    adjusted_headers, adjusted_rows = add_or_update_fields(product_headers, product_rows)
    strategy_rows = build_strategy_table(strategy_rows_original)

    success_flags = defaultdict(bool)

    save_product_workbook(adjusted_headers, adjusted_rows)
    success_flags["product_pool"] = OUTPUT_PRODUCT_POOL.exists()

    save_strategy_workbook(strategy_rows, OUTPUT_STRATEGY_TABLE)
    success_flags["strategy_table"] = OUTPUT_STRATEGY_TABLE.exists()

    stats = save_statistics_workbook(adjusted_rows)
    success_flags["statistics"] = OUTPUT_STATISTICS.exists()

    write_markdown(stats)
    success_flags["markdown"] = OUTPUT_MD.exists()

    success_flags["main_excel"] = sync_main_excel(adjusted_rows, strategy_rows)

    write_log(
        product_sheet=product_sheet,
        product_row_count=len(product_rows),
        adjusted_count=len(adjusted_rows),
        stats=stats,
        success_flags=success_flags,
        missing_inputs=missing_inputs,
    )

    print("STEP11_DONE")
    print(f"PRODUCT_ROWS={len(product_rows)}")
    print(f"ADJUSTED_ROWS={len(adjusted_rows)}")
    print("PRIMARY_COUNTS=" + ";".join(f"{k}:{v}" for k, v in stats["primary"].most_common()))
    red_secondary = Counter({k[1]: v for k, v in stats["cross"].items() if k[0] == "红利"})
    print("RED_SECONDARY_COUNTS=" + ";".join(f"{k}:{v}" for k, v in red_secondary.most_common()))
    print(f"MAIN_SYNC={'YES' if success_flags['main_excel'] else 'NO'}")


if __name__ == "__main__":
    main()
