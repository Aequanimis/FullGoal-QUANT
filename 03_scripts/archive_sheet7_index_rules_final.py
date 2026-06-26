from __future__ import annotations

import hashlib
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path.home() / "Desktop" / "Fullgoal" / "课题研究" / "处理后数据"
ARCHIVE_DIR = BASE / "阶段性归档_境内广义策略ETF数据底座"
SOURCE_FILE = BASE / "wind代码池" / "sheet7" / "核心策略ETF_跟踪指数规则_官网补充核验最终版.xlsx"
TARGET_FILE = ARCHIVE_DIR / "08_核心策略ETF_跟踪指数规则_官网补充核验最终版.xlsx"
INDEX_FILE = ARCHIVE_DIR / "阶段性归档索引表.xlsx"
README_FILE = ARCHIVE_DIR / "阶段性归档说明.md"


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def auto_width_and_header(ws) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    ws.freeze_panes = "A2"
    if ws.max_row and ws.max_column:
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 80)


def update_index() -> bool:
    if not INDEX_FILE.exists():
        return False

    wb = load_workbook(INDEX_FILE)
    ws = wb["归档文件索引"] if "归档文件索引" in wb.sheetnames else wb.worksheets[0]
    headers = [cell.value for cell in ws[1]]

    row_data_by_possible_header = {
        "序号": "08",
        "模块名称": "核心策略ETF跟踪指数规则",
        "文件名称": TARGET_FILE.name,
        "归档后文件名": TARGET_FILE.name,
        "文件类型": "Excel",
        "归档后文件路径": str(TARGET_FILE),
        "原始文件路径": str(SOURCE_FILE),
        "文件定位": "核心策略指数ETF跟踪指数编制规则终版，用于梳理82个核心跟踪指数的样本空间、选样指标、选样方法、加权方式、调样频率及官网核验状态",
        "文件用途": "用于分析核心策略ETF跟踪指数规则、策略因子逻辑、指数编制方法和后续指数规则表建设",
        "正式分析使用sheet": "指数规则明细_官网补充版；ETF_指数规则合并表_官网补充版；剩余待官网核验清单",
        "是否核心产出": "是",
        "注意事项": "ETF-指数映射保持168只ETF、82个核心指数、错配0；仍标记为待官网核验的字段后续可继续人工核验",
        "数据范围": "168只核心策略指数ETF，82个唯一跟踪指数",
        "数据状态": "已完成官网补充核验最终版",
        "覆盖情况": "168只ETF、82个核心指数；ETF-指数错配0；Sheet3/Sheet4空值0",
        "原始路径": str(SOURCE_FILE),
        "归档路径": str(TARGET_FILE),
        "归档时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "备注": "本文件为Sheet7跟踪指数规则正式归档版，可用于后续规则解读和策略分类分析",
    }

    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, ws.max_column + 1)]
        joined = " ".join(str(v) for v in values if v is not None)
        first = str(values[0]).strip() if values and values[0] is not None else ""
        if first == "08" or TARGET_FILE.name in joined or "跟踪指数规则" in joined:
            target_row = row_idx
            break

    if target_row is None:
        target_row = ws.max_row + 1

    for col_idx, header in enumerate(headers, start=1):
        if header in row_data_by_possible_header:
            ws.cell(target_row, col_idx).value = row_data_by_possible_header[header]

    # If the existing index is the older four-column/flexible format, ensure at least key cells are populated.
    if not any(h in headers for h in ["文件名称", "归档后文件名", "模块名称"]):
        ws.cell(target_row, 1).value = "08"
        ws.cell(target_row, 2).value = TARGET_FILE.name
        ws.cell(target_row, 3).value = str(TARGET_FILE)
        ws.cell(target_row, 4).value = str(SOURCE_FILE)

    # Update completion sheet if present.
    if "当前数据完成情况" in wb.sheetnames:
        ws2 = wb["当前数据完成情况"]
        headers2 = [cell.value for cell in ws2[1]]
        target = None
        for row_idx in range(2, ws2.max_row + 1):
            row_text = " ".join(str(ws2.cell(row_idx, col_idx).value) for col_idx in range(1, ws2.max_column + 1) if ws2.cell(row_idx, col_idx).value is not None)
            if "跟踪指数规则" in row_text or "指数规则" in row_text:
                target = row_idx
                break
        if target is None:
            target = ws2.max_row + 1
        values = {
            "数据模块": "跟踪指数规则",
            "当前状态": "已完成",
            "对应文件": TARGET_FILE.name,
            "主要用途": "核心策略ETF跟踪指数规则、选样方法、加权方式与核验状态分析",
            "后续是否还需补充": "部分字段可继续官网人工核验",
            "备注": "168只核心策略指数ETF，82个唯一跟踪指数，ETF-指数错配0",
        }
        for col_idx, header in enumerate(headers2, start=1):
            if header in values:
                ws2.cell(target, col_idx).value = values[header]

    # Remove the module from follow-up collection list if it exists.
    if "后续数据收集清单" in wb.sheetnames:
        ws4 = wb["后续数据收集清单"]
        rows_to_delete = []
        for row_idx in range(2, ws4.max_row + 1):
            row_text = " ".join(str(ws4.cell(row_idx, col_idx).value) for col_idx in range(1, ws4.max_column + 1) if ws4.cell(row_idx, col_idx).value is not None)
            if "跟踪指数规则" in row_text or "核心策略ETF_跟踪指数规则表" in row_text:
                rows_to_delete.append(row_idx)
        for row_idx in reversed(rows_to_delete):
            ws4.delete_rows(row_idx)

    for sheet in wb.worksheets:
        auto_width_and_header(sheet)
    wb.save(INDEX_FILE)
    return True


def update_readme() -> bool:
    if not README_FILE.exists():
        return False

    text = README_FILE.read_text(encoding="utf-8")
    section = f"""

## 核心策略ETF跟踪指数规则归档

08_核心策略ETF_跟踪指数规则_官网补充核验最终版.xlsx 已完成归档。本文件基于核心策略ETF产品池中的168只核心策略指数ETF和82个唯一跟踪指数，保留标准ETF-指数映射，并整合Wind补充与官网补充核验后的指数规则字段。当前版本保持ETF-指数错配数量为0，Sheet 3与Sheet 4已无空值，缺失或未确认字段统一标记为“待官网核验”。本文件可用于后续分析核心策略ETF跟踪指数的样本空间、选样指标、选样方法、加权方式、调样频率、策略因子逻辑及剩余官网核验事项。
"""
    if "08_核心策略ETF_跟踪指数规则_官网补充核验最终版.xlsx" not in text:
        text = text.rstrip() + section + "\n"

    replacements = {
        "- 跟踪指数规则：待收集": "- 跟踪指数规则：已完成，覆盖168只核心策略指数ETF和82个唯一跟踪指数。",
        "- 跟踪指数规则表。": "- 跟踪指数规则表：已完成。",
        "- 跟踪指数规则：待补充": "- 跟踪指数规则：已完成，覆盖168只核心策略指数ETF和82个唯一跟踪指数。",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)

    README_FILE.write_text(text, encoding="utf-8")
    return True


def main() -> None:
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(SOURCE_FILE)

    backup_path = None
    if TARGET_FILE.exists():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = TARGET_FILE.with_name(f"{TARGET_FILE.stem}_旧版备份_{stamp}{TARGET_FILE.suffix}")
        shutil.copy2(TARGET_FILE, backup_path)

    shutil.copy2(SOURCE_FILE, TARGET_FILE)
    source_hash = sha256(SOURCE_FILE)
    target_hash = sha256(TARGET_FILE)
    copied = source_hash == target_hash and TARGET_FILE.stat().st_size > 0

    index_updated = update_index()
    readme_updated = update_readme()

    print(f"源文件存在: 是")
    print(f"复制成功: {'是' if copied else '否'}")
    print(f"是否备份旧版文件: {'是' if backup_path else '否'}")
    if backup_path:
        print(f"旧版备份路径: {backup_path}")
    print(f"阶段性归档索引表更新: {'是' if index_updated else '否'}")
    print(f"阶段性归档说明更新: {'是' if readme_updated else '否'}")
    print(f"归档文件路径: {TARGET_FILE}")
    print(f"文件大小: {TARGET_FILE.stat().st_size}")
    print(f"SHA256一致: {'是' if source_hash == target_hash else '否'}")


if __name__ == "__main__":
    main()
