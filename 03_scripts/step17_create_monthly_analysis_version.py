from __future__ import annotations

import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\广义策略ETF_月度规模份额表_五批合并清洗版.xlsx")
OUTPUT = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\广义策略ETF_月度规模份额表_分析版.xlsx")
SOURCE_SHEET = "月度规模份额_清洗版"

ANALYSIS_SHEET = "月度规模份额_分析版"
TOTAL_SHEET = "月度汇总_广义与核心"
PRIMARY_SHEET = "月度汇总_按一级策略"
SECONDARY_SHEET = "月度汇总_按二级策略"
MANAGER_SHEET = "月度汇总_按基金公司"
QUALITY_SHEET = "分析版质量检查"

KEY_FIELDS = ["基金规模_亿元", "基金份额_亿份", "单位净值", "复权单位净值", "累计净值"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=9, color="1F1F1F")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN_GRAY = Side(style="thin", color="D9E1F2")
BODY_BORDER = Border(bottom=THIN_GRAY)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    value = str(value).replace("\u3000", " ").replace("\u200b", "")
    return re.sub(r"\s+", " ", value).strip()


def write_records(ws, headers: list[str], records: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for record in records:
        ws.append([record.get(header) for header in headers])


def style_sheet(ws, headers: list[str]) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BODY_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    sample_limit = min(ws.max_row, 300)
    wide_fields = {"基金全称", "跟踪指数名称", "数据备注", "说明"}
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, sample_limit + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 55))
        width = min(max(max_len * 1.15, 11), 44 if header in wide_fields else 28)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for col_idx, header in enumerate(headers, 1):
        if "日期" in header or header in {"上市日期", "基金成立日"}:
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (datetime, date)):
                    ws.cell(row_idx, col_idx).number_format = "yyyy-mm-dd"
        if "规模" in header:
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "#,##0.0000"
        if "亿份" in header:
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "#,##0.0000"
        elif header == "基金份额_份":
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "#,##0"


def sum_numeric(records: list[dict[str, Any]], field: str) -> float:
    return round(
        sum(float(record[field]) for record in records if isinstance(record.get(field), (int, float))),
        4,
    )


def scale_count(records: list[dict[str, Any]]) -> int:
    return len({record["Wind代码"] for record in records if record.get("基金规模_亿元") is not None})


def main() -> None:
    if not INPUT.exists():
        raise FileNotFoundError(f"输入文件不存在：{INPUT}")

    source_wb = load_workbook(INPUT, read_only=True, data_only=True)
    if SOURCE_SHEET not in source_wb.sheetnames:
        raise KeyError(f"输入文件缺少sheet：{SOURCE_SHEET}")
    source_ws = source_wb[SOURCE_SHEET]
    headers = [cell.value for cell in next(source_ws.iter_rows())]
    required = {
        "日期",
        "Wind代码",
        "基金管理人",
        "一级策略大类",
        "二级策略类别",
        "统计口径分类",
        "是否纳入核心策略ETF统计",
        "是否纳入广义策略ETF统计",
        *KEY_FIELDS,
    }
    missing = required - set(headers)
    if missing:
        raise KeyError(f"源sheet缺少字段：{sorted(missing)}")

    original_records: list[dict[str, Any]] = []
    analysis_records: list[dict[str, Any]] = []
    deleted_count = 0
    for values in source_ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and clean_text(value) for value in values):
            continue
        record = dict(zip(headers, values))
        original_records.append(record)
        if all(record.get(field) is None for field in KEY_FIELDS):
            deleted_count += 1
        else:
            analysis_records.append(record)

    analysis_records.sort(
        key=lambda record: (
            record.get("日期") if isinstance(record.get("日期"), datetime) else datetime.max,
            clean_text(record.get("Wind代码")),
        )
    )

    duplicate_keys = defaultdict(int)
    invalid_codes = set()
    for record in analysis_records:
        duplicate_keys[(record.get("Wind代码"), record.get("日期"))] += 1
        code = clean_text(record.get("Wind代码")).upper()
        if not re.fullmatch(r"\d{6}\.(SH|SZ)", code):
            invalid_codes.add(code or "空白")
    duplicate_count = sum(1 for count in duplicate_keys.values() if count > 1)

    dates = [record["日期"] for record in analysis_records if isinstance(record.get("日期"), datetime)]
    min_date = min(dates) if dates else None
    max_date = max(dates) if dates else None

    monthly: defaultdict[datetime, list[dict[str, Any]]] = defaultdict(list)
    for record in analysis_records:
        if isinstance(record.get("日期"), datetime):
            monthly[record["日期"]].append(record)

    total_rows: list[dict[str, Any]] = []
    primary_rows: list[dict[str, Any]] = []
    secondary_rows: list[dict[str, Any]] = []
    manager_rows: list[dict[str, Any]] = []

    for month in sorted(monthly):
        records = monthly[month]
        broad = [r for r in records if clean_text(r.get("是否纳入广义策略ETF统计")) == "是"]
        core = [r for r in broad if clean_text(r.get("是否纳入核心策略ETF统计")) == "是"]
        enhanced = [
            r for r in broad if clean_text(r.get("统计口径分类")) == "广义策略ETF_指数增强"
        ]
        total_rows.append(
            {
                "日期": month,
                "广义策略ETF总规模_亿元": sum_numeric(broad, "基金规模_亿元"),
                "核心策略指数ETF总规模_亿元": sum_numeric(core, "基金规模_亿元"),
                "指数增强多因子ETF总规模_亿元": sum_numeric(enhanced, "基金规模_亿元"),
                "广义策略ETF总份额_亿份": sum_numeric(broad, "基金份额_亿份"),
                "核心策略指数ETF总份额_亿份": sum_numeric(core, "基金份额_亿份"),
                "指数增强多因子ETF总份额_亿份": sum_numeric(enhanced, "基金份额_亿份"),
                "广义策略ETF有规模数据产品数": scale_count(broad),
                "核心策略指数ETF有规模数据产品数": scale_count(core),
                "指数增强多因子ETF有规模数据产品数": scale_count(enhanced),
            }
        )

        primary_groups: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
        secondary_groups: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
        manager_groups: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
        for record in broad:
            primary = clean_text(record.get("一级策略大类")) or "待补充"
            secondary = clean_text(record.get("二级策略类别")) or "待补充"
            manager = clean_text(record.get("基金管理人")) or "待补充"
            primary_groups[primary].append(record)
            secondary_groups[(primary, secondary)].append(record)
            manager_groups[manager].append(record)

        for primary, items in sorted(primary_groups.items()):
            primary_rows.append(
                {
                    "日期": month,
                    "一级策略大类": primary,
                    "有规模数据产品数量": scale_count(items),
                    "总规模_亿元": sum_numeric(items, "基金规模_亿元"),
                    "总份额_亿份": sum_numeric(items, "基金份额_亿份"),
                }
            )
        for (primary, secondary), items in sorted(secondary_groups.items()):
            secondary_rows.append(
                {
                    "日期": month,
                    "一级策略大类": primary,
                    "二级策略类别": secondary,
                    "有规模数据产品数量": scale_count(items),
                    "总规模_亿元": sum_numeric(items, "基金规模_亿元"),
                    "总份额_亿份": sum_numeric(items, "基金份额_亿份"),
                }
            )
        for manager, items in sorted(manager_groups.items()):
            manager_rows.append(
                {
                    "日期": month,
                    "基金管理人": manager,
                    "有规模数据产品数量": scale_count(items),
                    "总规模_亿元": sum_numeric(items, "基金规模_亿元"),
                    "总份额_亿份": sum_numeric(items, "基金份额_亿份"),
                }
            )

    all_codes = {clean_text(record.get("Wind代码")) for record in analysis_records}
    broad_codes = {
        clean_text(record.get("Wind代码"))
        for record in analysis_records
        if clean_text(record.get("是否纳入广义策略ETF统计")) == "是"
    }
    core_codes = {
        clean_text(record.get("Wind代码"))
        for record in analysis_records
        if clean_text(record.get("是否纳入核心策略ETF统计")) == "是"
    }
    enhanced_codes = {
        clean_text(record.get("Wind代码"))
        for record in analysis_records
        if clean_text(record.get("统计口径分类")) == "广义策略ETF_指数增强"
    }

    latest_row = next((row for row in reversed(total_rows) if row["日期"] == max_date), None)
    quality_rows = [
        {"统计类别": "行数", "指标": "原始清洗版行数", "数值": len(original_records)},
        {"统计类别": "行数", "指标": "删除关键字段全空行数", "数值": deleted_count},
        {"统计类别": "行数", "指标": "分析版有效行数", "数值": len(analysis_records)},
        {"统计类别": "覆盖", "指标": "覆盖Wind代码数量", "数值": len(all_codes)},
        {"统计类别": "覆盖", "指标": "覆盖广义策略ETF数量", "数值": len(broad_codes)},
        {"统计类别": "覆盖", "指标": "覆盖核心策略指数ETF数量", "数值": len(core_codes)},
        {"统计类别": "覆盖", "指标": "覆盖指数增强/多因子ETF数量", "数值": len(enhanced_codes)},
        {"统计类别": "日期", "指标": "日期最小值", "数值": min_date},
        {"统计类别": "日期", "指标": "日期最大值", "数值": max_date},
        {"统计类别": "日期", "指标": "最新日期", "数值": max_date},
        {
            "统计类别": "最新月",
            "指标": "最新日期广义策略ETF总规模",
            "数值": latest_row["广义策略ETF总规模_亿元"] if latest_row else None,
            "说明": "单位：亿元",
        },
        {
            "统计类别": "最新月",
            "指标": "最新日期核心策略指数ETF总规模",
            "数值": latest_row["核心策略指数ETF总规模_亿元"] if latest_row else None,
            "说明": "单位：亿元",
        },
        {
            "统计类别": "最新月",
            "指标": "最新日期指数增强/多因子ETF总规模",
            "数值": latest_row["指数增强多因子ETF总规模_亿元"] if latest_row else None,
            "说明": "单位：亿元",
        },
        {
            "统计类别": "质量",
            "指标": "是否存在重复Wind代码+日期",
            "数值": "是" if duplicate_count else "否",
            "说明": f"重复键数量：{duplicate_count}",
        },
        {
            "统计类别": "质量",
            "指标": "是否存在非.SH/.SZ代码",
            "数值": "是" if invalid_codes else "否",
            "说明": "；".join(sorted(invalid_codes)),
        },
    ]

    wb = Workbook()
    wb.remove(wb.active)
    analysis_ws = wb.create_sheet(ANALYSIS_SHEET)
    total_ws = wb.create_sheet(TOTAL_SHEET)
    primary_ws = wb.create_sheet(PRIMARY_SHEET)
    secondary_ws = wb.create_sheet(SECONDARY_SHEET)
    manager_ws = wb.create_sheet(MANAGER_SHEET)
    quality_ws = wb.create_sheet(QUALITY_SHEET)

    write_records(analysis_ws, headers, analysis_records)
    total_headers = [
        "日期",
        "广义策略ETF总规模_亿元",
        "核心策略指数ETF总规模_亿元",
        "指数增强多因子ETF总规模_亿元",
        "广义策略ETF总份额_亿份",
        "核心策略指数ETF总份额_亿份",
        "指数增强多因子ETF总份额_亿份",
        "广义策略ETF有规模数据产品数",
        "核心策略指数ETF有规模数据产品数",
        "指数增强多因子ETF有规模数据产品数",
    ]
    write_records(total_ws, total_headers, total_rows)
    primary_headers = ["日期", "一级策略大类", "有规模数据产品数量", "总规模_亿元", "总份额_亿份"]
    write_records(primary_ws, primary_headers, primary_rows)
    secondary_headers = ["日期", "一级策略大类", "二级策略类别", "有规模数据产品数量", "总规模_亿元", "总份额_亿份"]
    write_records(secondary_ws, secondary_headers, secondary_rows)
    manager_headers = ["日期", "基金管理人", "有规模数据产品数量", "总规模_亿元", "总份额_亿份"]
    write_records(manager_ws, manager_headers, manager_rows)
    quality_headers = ["统计类别", "指标", "数值", "说明"]
    write_records(quality_ws, quality_headers, quality_rows)

    for ws, sheet_headers in [
        (analysis_ws, headers),
        (total_ws, total_headers),
        (primary_ws, primary_headers),
        (secondary_ws, secondary_headers),
        (manager_ws, manager_headers),
        (quality_ws, quality_headers),
    ]:
        style_sheet(ws, sheet_headers)

    previous = None
    for row_idx in range(2, quality_ws.max_row + 1):
        section = quality_ws.cell(row_idx, 1).value
        if section != previous:
            for col_idx in range(1, quality_ws.max_column + 1):
                quality_ws.cell(row_idx, col_idx).fill = SECTION_FILL
                quality_ws.cell(row_idx, col_idx).font = Font(name="微软雅黑", size=9, bold=True)
        previous = section
        if quality_ws.cell(row_idx, 2).value in {"日期最小值", "日期最大值", "最新日期"}:
            quality_ws.cell(row_idx, 3).number_format = "yyyy-mm-dd"
        if "总规模" in clean_text(quality_ws.cell(row_idx, 2).value):
            quality_ws.cell(row_idx, 3).number_format = "#,##0.0000"

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    latest_scale = latest_row["广义策略ETF总规模_亿元"] if latest_row else None
    print(f"删除空白行数量：{deleted_count}")
    print(f"分析版有效行数：{len(analysis_records)}")
    print(f"覆盖Wind代码数量：{len(all_codes)}")
    print(
        f"日期范围：{min_date.strftime('%Y-%m-%d') if min_date else '无'} 至 "
        f"{max_date.strftime('%Y-%m-%d') if max_date else '无'}"
    )
    print(f"最新月广义策略ETF总规模：{latest_scale:.4f}亿元" if latest_scale is not None else "最新月广义策略ETF总规模：无")
    print(f"输出文件路径：{OUTPUT}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"处理失败：{exc}", file=sys.stderr)
        raise
