from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT_PATH = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\原始数据\全市场ETF基础信息.csv.xlsx")
OUTPUT_PATH = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\全市场ETF基础信息_清洗与策略筛选.xlsx")

SOURCE_SHEET = "原始数据_备份"
CLEAN_SHEET = "ETF全市场_清洗版"
CANDIDATE_SHEET = "策略ETF_候选池"
FORMAL_SHEET = "策略ETF_正式池_待复核"
EXCLUDED_SHEET = "排除产品清单"
QUALITY_SHEET = "数据质量检查"

NUMERIC_FIELDS = {
    "最新基金规模(亿)",
    "最新基金份额(份)",
    "托管费率_支持历史(%)",
    "管理费率_支持历史(%)",
}

DATE_FIELDS = {
    "上市日期",
    "基金成立日",
    "托管费率_支持历史时间",
    "管理费率_支持历史时间",
}

SEARCH_FIELDS = [
    "证券简称",
    "基金简称",
    "基金全称",
    "跟踪指数名称",
    "业绩比较基准",
    "投资类型_二级分类",
]

# Long phrases are placed first so that the evidence is easier to read.
STRONG_KEYWORDS = [
    "港股通低波红利",
    "红利低波",
    "低波红利",
    "高股息低波",
    "自由现金流",
    "社会责任",
    "股东回报",
    "增强策略",
    "等权重",
    "低波动",
    "高股息",
    "高分红",
    "多因子",
    "基本面",
    "现金流",
    "红利",
    "股息",
    "低波",
    "质量",
    "价值",
    "成长",
    "动量",
    "增强",
    "ESG",
    "治理",
    "回购",
    "等权",
]

WEAK_KEYWORDS = ["策略", "智选", "优选", "央企", "国企"]

RESTRICTED_TYPES = {"货币市场型基金", "商品型基金", "被动指数型债券基金"}

OBVIOUS_STATE_THEME_TERMS = [
    "央企创新",
    "央企现代能源",
    "央企科技引领",
    "国企改革",
    "国企数字经济",
]

FORMAL_FIELDS = [
    "Wind代码",
    "证券简称",
    "基金简称",
    "基金全称",
    "基金代码",
    "交易代码",
    "基金管理人",
    "基金托管人",
    "基金上市地点",
    "上市日期",
    "基金成立日",
    "投资类型_二级分类",
    "业绩比较基准",
    "跟踪指数代码",
    "跟踪指数名称",
    "最新基金规模(亿)",
    "最新基金份额(份)",
    "管理费率_支持历史(%)",
    "托管费率_支持历史(%)",
    "是否上市ETF",
    "是否待核验",
    "命中关键词",
    "命中字段",
    "策略识别强度",
    "是否纳入策略ETF统计",
    "一级策略大类",
    "二级策略类别",
    "市场范围",
    "是否红利+策略",
    "分类依据",
    "备注",
]

DERIVED_CLEAN_FIELDS = [
    "Wind代码后缀",
    "交易所_标准化",
    "是否上市ETF",
    "是否有跟踪指数",
    "是否有规模数据",
    "是否待核验",
]

CANDIDATE_FIELDS = ["命中关键词", "命中字段", "策略识别强度", "是否纳入策略ETF统计"]

PRIMARY_ORDER = [
    "红利",
    "自由现金流",
    "质量",
    "价值",
    "成长",
    "低波",
    "指数增强/多因子",
    "ESG",
    "基本面策略",
    "等权/另类加权",
    "待核验",
]

SECONDARY_ORDER = [
    "普通红利",
    "红利低波",
    "港股通低波红利",
    "红利质量",
    "红利价值",
    "红利成长",
    "央企红利",
    "国企红利",
    "港股通红利",
    "股东回报",
    "央企股东回报",
    "自由现金流",
    "质量",
    "价值",
    "成长",
    "低波",
    "指数增强/多因子",
    "ESG",
    "基本面策略",
    "等权/另类加权",
    "待核验",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=9, color="1F1F1F")
LIGHT_BLUE_FILL = PatternFill("solid", fgColor="D9EAF7")
LIGHT_GREEN_FILL = PatternFill("solid", fgColor="E2F0D9")
LIGHT_YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
LIGHT_RED_FILL = PatternFill("solid", fgColor="FCE4D6")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN_GRAY = Side(style="thin", color="D9E1F2")
BODY_BORDER = Border(bottom=THIN_GRAY)


def clean_text(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = re.sub(r"[\r\n\t]+", " ", value)
        value = re.sub(r"\s+", " ", value).strip()
        return value or None
    return value


def to_number(value: Any) -> float | int | None:
    value = clean_text(value)
    if value in (None, "", "--", "N/A", "nan"):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    text = str(value).replace(",", "").replace("%", "").strip()
    try:
        number = float(text)
        return int(number) if number.is_integer() else number
    except (TypeError, ValueError):
        return None


def to_date(value: Any) -> datetime | None:
    value = clean_text(value)
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def suffix_from_wind_code(value: Any) -> str:
    text = str(value or "").strip().upper()
    if "." in text:
        return text.rsplit(".", 1)[-1]
    match = re.search(r"(SH|SZ|OF)$", text)
    return match.group(1) if match else ""


def has_text(value: Any) -> bool:
    return clean_text(value) not in (None, "")


def searchable_text(record: dict[str, Any], fields: list[str] | None = None) -> str:
    fields = fields or SEARCH_FIELDS
    return " ".join(str(record.get(field) or "") for field in fields)


def match_keywords(record: dict[str, Any]) -> tuple[list[str], list[str], dict[str, list[str]]]:
    hits_by_field: dict[str, list[str]] = {}
    strong_hits: list[str] = []
    weak_hits: list[str] = []

    for field in SEARCH_FIELDS:
        text = str(record.get(field) or "")
        text_upper = text.upper()
        field_hits: list[str] = []
        for keyword in STRONG_KEYWORDS:
            matched = keyword.upper() in text_upper if keyword == "ESG" else keyword in text
            if matched:
                field_hits.append(keyword)
                if keyword not in strong_hits:
                    strong_hits.append(keyword)
        for keyword in WEAK_KEYWORDS:
            if keyword in text:
                field_hits.append(keyword)
                if keyword not in weak_hits:
                    weak_hits.append(keyword)
        if field_hits:
            hits_by_field[field] = list(dict.fromkeys(field_hits))

    return strong_hits, weak_hits, hits_by_field


def identify_candidate(record: dict[str, Any]) -> dict[str, Any] | None:
    strong_hits, weak_hits, hits_by_field = match_keywords(record)
    investment_type = str(record.get("投资类型_二级分类") or "")
    enhanced_type = investment_type == "增强指数型基金"

    if enhanced_type:
        if "增强指数型基金" not in strong_hits:
            strong_hits.append("增强指数型基金")
        hits_by_field.setdefault("投资类型_二级分类", []).append("增强指数型基金")

    if not strong_hits and not weak_hits and not enhanced_type:
        return None

    combined = searchable_text(record)
    restricted = investment_type in RESTRICTED_TYPES
    weak_only = not strong_hits and not enhanced_type
    obvious_state_theme = weak_only and any(term in combined for term in OBVIOUS_STATE_THEME_TERMS)

    if restricted and not strong_hits:
        return None

    if obvious_state_theme:
        strength = "待核验"
        include_status = "否"
    elif restricted:
        strength = "待核验"
        include_status = "待核验"
    elif record.get("是否上市ETF") != "是":
        strength = "待核验"
        include_status = "待核验"
    elif weak_only:
        strength = "弱"
        include_status = "待核验"
    else:
        strength = "强"
        include_status = "是"

    keyword_hits = strong_hits + [item for item in weak_hits if item not in strong_hits]
    hit_fields = []
    for field, keywords in hits_by_field.items():
        hit_fields.append(f"{field}({','.join(dict.fromkeys(keywords))})")

    return {
        "命中关键词": "、".join(keyword_hits),
        "命中字段": "；".join(hit_fields),
        "策略识别强度": strength,
        "是否纳入策略ETF统计": include_status,
    }


def contains_any(text: str, terms: list[str] | tuple[str, ...]) -> bool:
    return any(term in text for term in terms)


def classify_strategy(record: dict[str, Any]) -> dict[str, str]:
    text = searchable_text(
        record,
        ["证券简称", "基金简称", "基金全称", "跟踪指数名称", "业绩比较基准"],
    )
    text_upper = text.upper()
    investment_type = str(record.get("投资类型_二级分类") or "")
    hits = str(record.get("命中关键词") or "")

    has_dividend = contains_any(text, ["红利", "高股息", "股息", "高分红"])
    has_shareholder_return = contains_any(text, ["股东回报", "回购"])
    has_low_vol = contains_any(text, ["低波", "低波动", "最小波动"])
    has_quality = "质量" in text
    has_value = "价值" in text
    has_growth = "成长" in text
    has_cashflow = contains_any(text, ["自由现金流", "现金流"])
    has_enhanced = contains_any(text, ["增强策略", "增强", "多因子"]) or investment_type == "增强指数型基金"
    has_esg = "ESG" in text_upper or contains_any(text, ["社会责任", "治理"])
    has_fundamental = "基本面" in text
    has_equal_weight = contains_any(text, ["等权", "等权重"])

    if has_dividend or has_shareholder_return:
        primary = "红利"
    elif has_cashflow:
        primary = "自由现金流"
    elif has_quality:
        primary = "质量"
    elif has_value:
        primary = "价值"
    elif has_growth:
        primary = "成长"
    elif has_low_vol:
        primary = "低波"
    elif has_enhanced:
        primary = "指数增强/多因子"
    elif has_esg:
        primary = "ESG"
    elif has_fundamental:
        primary = "基本面策略"
    elif has_equal_weight:
        primary = "等权/另类加权"
    else:
        primary = "待核验"

    if has_dividend or has_shareholder_return:
        if "港股通" in text and has_low_vol:
            secondary = "港股通低波红利"
        elif has_dividend and has_low_vol:
            secondary = "红利低波"
        elif has_dividend and has_quality:
            secondary = "红利质量"
        elif has_dividend and has_value:
            secondary = "红利价值"
        elif has_dividend and has_growth:
            secondary = "红利成长"
        elif "央企" in text and has_shareholder_return:
            secondary = "央企股东回报"
        elif has_shareholder_return:
            secondary = "股东回报"
        elif "港股通" in text:
            secondary = "港股通红利"
        elif contains_any(text, ["央企", "中央企业"]):
            secondary = "央企红利"
        elif contains_any(text, ["国企", "国有企业"]):
            secondary = "国企红利"
        else:
            secondary = "普通红利"
    elif has_cashflow:
        secondary = "自由现金流"
    elif has_quality:
        secondary = "质量"
    elif has_value:
        secondary = "价值"
    elif has_growth:
        secondary = "成长"
    elif has_low_vol:
        secondary = "低波"
    elif has_enhanced:
        secondary = "指数增强/多因子"
    elif has_esg:
        secondary = "ESG"
    elif has_fundamental:
        secondary = "基本面策略"
    elif has_equal_weight:
        secondary = "等权/另类加权"
    else:
        secondary = "待核验"

    if "港股通" in text:
        market_scope = "港股通"
    elif contains_any(text, ["恒生", "标普港股通"]) or "MSCI" in text_upper:
        market_scope = "跨境/港股"
    elif contains_any(text, ["央企", "中央企业"]):
        market_scope = "央企"
    elif contains_any(text, ["国企", "国有企业"]):
        market_scope = "国企"
    elif contains_any(
        text,
        [
            "中证A500",
            "沪深300",
            "中证500",
            "中证800",
            "中证1000",
            "中证2000",
            "创业板",
            "科创板",
            "科创50",
        ],
    ):
        market_scope = "A股"
    else:
        market_scope = "待核验"

    dividend_plus = "是" if primary == "红利" and secondary != "普通红利" else "否"
    basis = f"命中关键词：{hits or '无明确关键词'}；结合产品名称、跟踪指数及投资类型按规则初步分类"

    notes: list[str] = []
    if record.get("是否纳入策略ETF统计") == "待核验":
        notes.append("需人工复核是否具有明确策略选样或加权规则")
    elif record.get("是否纳入策略ETF统计") == "否":
        notes.append("弱关键词对应主题属性，暂不纳入统计")
    if record.get("是否待核验") == "是":
        notes.append("基础信息存在待核验项")
    if "动量" in hits and primary == "待核验":
        notes.append("命中动量，但当前一级策略口径未单列")
    if investment_type in RESTRICTED_TYPES:
        notes.append(f"{investment_type}原则上不纳入本次权益策略ETF统计")

    return {
        "一级策略大类": primary,
        "二级策略类别": secondary,
        "市场范围": market_scope,
        "是否红利+策略": dividend_plus,
        "分类依据": basis,
        "备注": "；".join(notes),
    }


def exclusion_reason(record: dict[str, Any]) -> str:
    investment_type = str(record.get("投资类型_二级分类") or "")
    text = searchable_text(
        record,
        ["证券简称", "基金简称", "基金全称", "跟踪指数名称", "业绩比较基准"],
    )
    text_upper = text.upper()

    if investment_type == "货币市场型基金":
        return "货币 ETF"
    if investment_type == "商品型基金":
        return "商品 ETF"
    if investment_type == "被动指数型债券基金":
        return "债券 ETF"
    if record.get("是否上市ETF") != "是":
        return "未上市 / 待核验"

    broad_terms = [
        "沪深300",
        "中证500",
        "中证1000",
        "中证2000",
        "中证A500",
        "上证50",
        "科创50",
        "创业板指",
        "创业板50",
        "深证100",
        "中证800",
        "中证全指",
    ]
    if investment_type == "国际(QDII)股票型基金":
        cross_border_broad = [
            "标普500",
            "纳斯达克100",
            "纳指100",
            "日经225",
            "恒生指数",
            "恒生科技",
            "德国DAX",
            "法国CAC",
            "MSCI",
        ]
        if contains_any(text, cross_border_broad) or "MSCI" in text_upper:
            return "跨境宽基 ETF"

    if contains_any(text, broad_terms):
        return "宽基 ETF"

    industry_terms = [
        "证券公司",
        "证券行业",
        "银行",
        "芯片",
        "半导体",
        "医药",
        "医疗",
        "创新药",
        "新能源",
        "食品",
        "畜牧",
        "养殖",
        "化工",
        "电力",
        "军工",
        "煤炭",
        "钢铁",
        "有色金属",
        "房地产",
        "传媒",
        "通信",
        "计算机",
        "软件",
        "汽车",
        "机械",
        "建材",
        "家电",
        "消费",
        "白酒",
        "农业",
        "光伏",
        "电池",
        "储能",
        "装备产业",
    ]
    if contains_any(text, industry_terms):
        return "行业主题 ETF"

    concept_terms = [
        "人工智能",
        "机器人",
        "数字经济",
        "云计算",
        "大数据",
        "物联网",
        "低空经济",
        "卫星",
        "航天",
        "绿色低碳",
        "碳中和",
        "国企改革",
        "央企创新",
        "央企科技",
    ]
    if contains_any(text, concept_terms):
        return "主题概念 ETF"

    if investment_type == "国际(QDII)股票型基金":
        return "跨境宽基 ETF"
    if investment_type == "被动指数型基金":
        return "指数规则不足，暂不纳入"
    return "其他"


def clone_cell_style(source_cell, target_cell) -> None:
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.protection:
        target_cell.protection = copy(source_cell.protection)


def style_table_sheet(ws, headers: list[str], row_count: int, add_filter: bool = True) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    max_col = len(headers)
    if add_filter and row_count >= 1 and max_col:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max(row_count, 1)}"

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34

    sample_limit = min(row_count, 300)
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, sample_limit + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 50))
        if header in {"基金全称", "业绩比较基准", "跟踪指数名称", "命中字段", "分类依据", "备注"}:
            width = min(max(max_len * 1.25, 18), 40)
        else:
            width = min(max(max_len * 1.15, 11), 26)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    if row_count >= 2:
        body_range = ws.iter_rows(min_row=2, max_row=row_count, min_col=1, max_col=max_col)
        for row in body_range:
            for cell in row:
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                cell.border = BODY_BORDER

    for header in DATE_FIELDS:
        if header in headers:
            col = headers.index(header) + 1
            for row in range(2, row_count + 1):
                ws.cell(row, col).number_format = "yyyy-mm-dd"

    numeric_formats = {
        "最新基金规模(亿)": "#,##0.0000",
        "最新基金份额(份)": "#,##0",
        "管理费率_支持历史(%)": "0.0000",
        "托管费率_支持历史(%)": "0.0000",
    }
    for header, fmt in numeric_formats.items():
        if header in headers:
            col = headers.index(header) + 1
            for row in range(2, row_count + 1):
                ws.cell(row, col).number_format = fmt


def add_status_formatting(ws, headers: list[str], row_count: int) -> None:
    if row_count < 2:
        return
    for header in ["是否纳入策略ETF统计", "是否待核验", "是否上市ETF"]:
        if header not in headers:
            continue
        col = headers.index(header) + 1
        letter = get_column_letter(col)
        target = f"{letter}2:{letter}{row_count}"
        ws.conditional_formatting.add(
            target,
            FormulaRule(formula=[f'{letter}2="是"'], fill=LIGHT_GREEN_FILL),
        )
        ws.conditional_formatting.add(
            target,
            FormulaRule(formula=[f'OR({letter}2="待核验",{letter}2="否/待核验")'], fill=LIGHT_YELLOW_FILL),
        )
        ws.conditional_formatting.add(
            target,
            FormulaRule(formula=[f'{letter}2="否"'], fill=LIGHT_RED_FILL),
        )


def write_records(ws, headers: list[str], records: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for record in records:
        ws.append([record.get(header) for header in headers])


def build_quality_rows(
    raw_total_rows: int,
    clean_records: list[dict[str, Any]],
    candidate_records: list[dict[str, Any]],
    formal_records: list[dict[str, Any]],
) -> list[list[Any]]:
    listed_yes = sum(r.get("是否上市ETF") == "是" for r in clean_records)
    suffix_counts = Counter(r.get("Wind代码后缀") or "空白" for r in clean_records)
    included = [r for r in formal_records if r.get("是否纳入策略ETF统计") == "是"]
    pending = [r for r in formal_records if r.get("是否纳入策略ETF统计") == "待核验"]

    rows: list[list[Any]] = []

    def add(section: str, metric: str, value: Any, note: str = "") -> None:
        rows.append([section, metric, value, note])

    add("总体检查", "原始总行数", raw_total_rows, "不含表头，包含 Wind 来源行及可能的空白行")
    add("总体检查", "删除 Wind 来源行后的有效产品数", len(clean_records))
    add("总体检查", "是否上市ETF=是 的数量", listed_yes)
    add("总体检查", "是否上市ETF=否/待核验 的数量", len(clean_records) - listed_yes)
    add("总体检查", "Wind代码后缀为 SH 的数量", suffix_counts.get("SH", 0))
    add("总体检查", "Wind代码后缀为 SZ 的数量", suffix_counts.get("SZ", 0))
    add("总体检查", "Wind代码后缀为 OF 的数量", suffix_counts.get("OF", 0))
    add("总体检查", "最新基金规模缺失数量", sum(r.get("最新基金规模(亿)") is None for r in clean_records))
    add("总体检查", "跟踪指数名称缺失数量", sum(not has_text(r.get("跟踪指数名称")) for r in clean_records))
    add("总体检查", "上市日期缺失数量", sum(r.get("上市日期") is None for r in clean_records))
    add("策略识别", "策略ETF候选池数量", len(candidate_records))
    add("策略识别", "策略ETF正式池中“是否纳入策略ETF统计=是”的数量", len(included))
    add("策略识别", "“待核验”数量", len(pending), "指是否纳入策略ETF统计=待核验")
    add(
        "策略识别",
        "基础信息是否待核验=是 的候选数量",
        sum(r.get("是否待核验") == "是" for r in formal_records),
    )

    primary_counts = Counter(r.get("一级策略大类") or "待核验" for r in included)
    primary_sizes: defaultdict[str, float] = defaultdict(float)
    secondary_counts = Counter(r.get("二级策略类别") or "待核验" for r in included)
    manager_counts = Counter(r.get("基金管理人") or "待补充" for r in included)
    manager_sizes: defaultdict[str, float] = defaultdict(float)

    for record in included:
        size = record.get("最新基金规模(亿)")
        if isinstance(size, (int, float)):
            primary_sizes[record.get("一级策略大类") or "待核验"] += float(size)
            manager_sizes[record.get("基金管理人") or "待补充"] += float(size)

    for primary in sorted(primary_counts, key=lambda x: (PRIMARY_ORDER.index(x) if x in PRIMARY_ORDER else 999, x)):
        add("一级策略大类产品数量", primary, primary_counts[primary], "仅统计是否纳入策略ETF统计=是")
    for primary in sorted(primary_sizes, key=lambda x: (PRIMARY_ORDER.index(x) if x in PRIMARY_ORDER else 999, x)):
        add("一级策略大类规模合计", primary, round(primary_sizes[primary], 4), "单位：亿元；忽略规模缺失值")
    for secondary in sorted(
        secondary_counts,
        key=lambda x: (SECONDARY_ORDER.index(x) if x in SECONDARY_ORDER else 999, x),
    ):
        add("二级策略类别产品数量", secondary, secondary_counts[secondary], "仅统计是否纳入策略ETF统计=是")
    for manager, count in manager_counts.most_common():
        add("基金管理人策略ETF数量", manager, count, "仅统计是否纳入策略ETF统计=是")
    for manager, size in sorted(manager_sizes.items(), key=lambda item: (-item[1], item[0])):
        add("基金管理人策略ETF规模合计", manager, round(size, 4), "单位：亿元；忽略规模缺失值")

    return rows


def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"输入文件不存在：{INPUT_PATH}")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    # Load the source workbook and reuse it as the output workbook so the first
    # sheet remains a faithful in-workbook backup of the Wind export.
    wb = load_workbook(INPUT_PATH)
    source_ws = wb[wb.sheetnames[0]]
    source_ws.title = SOURCE_SHEET

    # “原始总行数”按数据行口径统计，不含表头，但包含末尾 Wind 来源行。
    raw_total_rows = max(source_ws.max_row - 1, 0)
    raw_headers = [clean_text(cell.value) for cell in source_ws[1]]
    if len(set(raw_headers)) != len(raw_headers):
        raise ValueError("原始表头存在重复字段，无法安全处理。")

    clean_records: list[dict[str, Any]] = []
    for values in source_ws.iter_rows(min_row=2, values_only=True):
        if not any(has_text(value) for value in values):
            continue
        first_value = clean_text(values[0]) if values else None
        if isinstance(first_value, str) and first_value.startswith("数据来源：Wind"):
            continue

        record: dict[str, Any] = {}
        for header, value in zip(raw_headers, values):
            if header in NUMERIC_FIELDS:
                record[header] = to_number(value)
            elif header in DATE_FIELDS:
                record[header] = to_date(value)
            else:
                record[header] = clean_text(value)

        suffix = suffix_from_wind_code(record.get("Wind代码"))
        exchange = {"SH": "上交所", "SZ": "深交所", "OF": "场外或待核验"}.get(suffix, "待核验")
        listed = (
            suffix in {"SH", "SZ"}
            and has_text(record.get("基金上市地点"))
            and record.get("上市日期") is not None
        )
        pending = (
            suffix == "OF"
            or record.get("上市日期") is None
            or not has_text(record.get("基金上市地点"))
            or not has_text(record.get("跟踪指数名称"))
        )

        record.update(
            {
                "Wind代码后缀": suffix,
                "交易所_标准化": exchange,
                "是否上市ETF": "是" if listed else "否/待核验",
                "是否有跟踪指数": "是" if has_text(record.get("跟踪指数名称")) else "否",
                "是否有规模数据": "是" if record.get("最新基金规模(亿)") is not None else "否",
                "是否待核验": "是" if pending else "否",
            }
        )
        clean_records.append(record)

    candidate_records: list[dict[str, Any]] = []
    excluded_records: list[dict[str, Any]] = []

    for record in clean_records:
        candidate_fields = identify_candidate(record)
        if candidate_fields is None:
            excluded = dict(record)
            excluded["排除原因"] = exclusion_reason(record)
            excluded_records.append(excluded)
            continue
        candidate = dict(record)
        candidate.update(candidate_fields)
        candidate_records.append(candidate)

    formal_records: list[dict[str, Any]] = []
    for candidate in candidate_records:
        formal = dict(candidate)
        formal.update(classify_strategy(formal))
        formal_records.append(formal)

    primary_rank = {value: idx for idx, value in enumerate(PRIMARY_ORDER)}
    secondary_rank = {value: idx for idx, value in enumerate(SECONDARY_ORDER)}
    formal_records.sort(
        key=lambda r: (
            primary_rank.get(r.get("一级策略大类"), 999),
            secondary_rank.get(r.get("二级策略类别"), 999),
            -(float(r.get("最新基金规模(亿)")) if isinstance(r.get("最新基金规模(亿)"), (int, float)) else -1),
            str(r.get("Wind代码") or ""),
        )
    )

    clean_headers = raw_headers + DERIVED_CLEAN_FIELDS
    candidate_headers = clean_headers + CANDIDATE_FIELDS
    excluded_headers = clean_headers + ["排除原因"]

    # Rebuild all generated sheets to make reruns idempotent.
    for sheet_name in [CLEAN_SHEET, CANDIDATE_SHEET, FORMAL_SHEET, EXCLUDED_SHEET, QUALITY_SHEET]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    clean_ws = wb.create_sheet(CLEAN_SHEET)
    candidate_ws = wb.create_sheet(CANDIDATE_SHEET)
    formal_ws = wb.create_sheet(FORMAL_SHEET)
    excluded_ws = wb.create_sheet(EXCLUDED_SHEET)
    quality_ws = wb.create_sheet(QUALITY_SHEET)

    write_records(clean_ws, clean_headers, clean_records)
    write_records(candidate_ws, candidate_headers, candidate_records)
    write_records(formal_ws, FORMAL_FIELDS, formal_records)
    write_records(excluded_ws, excluded_headers, excluded_records)

    quality_headers = ["统计类别", "指标", "数量/规模", "说明"]
    quality_rows = build_quality_rows(raw_total_rows, clean_records, candidate_records, formal_records)
    quality_ws.append(quality_headers)
    for row in quality_rows:
        quality_ws.append(row)

    # The backup values stay unchanged. Only presentation settings requested for
    # every sheet are added.
    style_table_sheet(source_ws, raw_headers, source_ws.max_row)
    style_table_sheet(clean_ws, clean_headers, clean_ws.max_row)
    style_table_sheet(candidate_ws, candidate_headers, candidate_ws.max_row)
    style_table_sheet(formal_ws, FORMAL_FIELDS, formal_ws.max_row)
    style_table_sheet(excluded_ws, excluded_headers, excluded_ws.max_row)
    style_table_sheet(quality_ws, quality_headers, quality_ws.max_row)

    for ws, headers in [
        (clean_ws, clean_headers),
        (candidate_ws, candidate_headers),
        (formal_ws, FORMAL_FIELDS),
        (excluded_ws, excluded_headers),
    ]:
        add_status_formatting(ws, headers, ws.max_row)

    # Make the quality sheet easier to scan by visually grouping sections.
    previous_section = None
    for row_idx in range(2, quality_ws.max_row + 1):
        section = quality_ws.cell(row_idx, 1).value
        if section != previous_section:
            for col_idx in range(1, 5):
                quality_ws.cell(row_idx, col_idx).fill = SECTION_FILL
                quality_ws.cell(row_idx, col_idx).font = Font(
                    name="微软雅黑", size=9, bold=True, color="1F1F1F"
                )
        previous_section = section
        if isinstance(quality_ws.cell(row_idx, 3).value, float):
            quality_ws.cell(row_idx, 3).number_format = "#,##0.0000"
        else:
            quality_ws.cell(row_idx, 3).number_format = "#,##0"

    # Format source date/numeric fields without changing source values.
    for header in DATE_FIELDS:
        if header in raw_headers:
            col = raw_headers.index(header) + 1
            for row in range(2, source_ws.max_row + 1):
                if isinstance(source_ws.cell(row, col).value, (datetime, date)):
                    source_ws.cell(row, col).number_format = "yyyy-mm-dd"

    # Preserve the source workbook's original tab as the first sheet.
    wb._sheets = [
        wb[SOURCE_SHEET],
        wb[CLEAN_SHEET],
        wb[CANDIDATE_SHEET],
        wb[FORMAL_SHEET],
        wb[EXCLUDED_SHEET],
        wb[QUALITY_SHEET],
    ]

    # Remove any stale calc cache behavior and ask Excel to recalculate on open.
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except AttributeError:
        pass

    wb.save(OUTPUT_PATH)

    listed_count = sum(r.get("是否上市ETF") == "是" for r in clean_records)
    included_count = sum(r.get("是否纳入策略ETF统计") == "是" for r in formal_records)
    pending_count = sum(r.get("是否纳入策略ETF统计") == "待核验" for r in formal_records)

    print(f"1. 原始有效产品数量：{len(clean_records)}")
    print(f"2. 上市 ETF 数量：{listed_count}")
    print(f"3. 策略 ETF 候选数量：{len(candidate_records)}")
    print(f"4. 纳入统计的策略 ETF 数量：{included_count}")
    print(f"5. 待核验数量：{pending_count}")
    print(f"6. 输出文件路径：{OUTPUT_PATH}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"处理失败：{exc}", file=sys.stderr)
        raise
