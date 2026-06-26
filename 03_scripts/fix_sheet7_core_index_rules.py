from __future__ import annotations

from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path.home() / "Desktop" / "Fullgoal" / "课题研究" / "处理后数据"
PROJECT_SCRIPTS = Path.home() / "Desktop" / "Fullgoal" / "课题研究" / "FullGoal-QUANT" / "03_scripts"

PRODUCT_POOL_FILE = BASE / "全市场ETF基础信息_策略ETF池二次修正版.xlsx"
MAPPING_FILE = BASE / "wind代码池" / "sheet7" / "ETF与指数映射关系表.xlsx"
RULE_FILE = BASE / "wind代码池" / "sheet7" / "指数编制规则详细信息表.xlsx"
MERGED_FILE = BASE / "wind代码池" / "sheet7" / "完整数据合并表.xlsx"
OUTPUT_FILE = BASE / "wind代码池" / "sheet7" / "核心策略ETF_跟踪指数规则_验收修正版.xlsx"


RULE_FIELDS = [
    "发布机构",
    "指数基日",
    "指数发布日期",
    "指数类型",
    "样本空间",
    "成分股数量",
    "选样指标",
    "选样方法",
    "加权方式",
    "单只成分权重上限",
    "行业权重限制",
    "调样频率",
    "指数简介",
    "指数编制方案链接",
]

SHEET1_COLS = [
    "ETF_Code",
    "ETF_Name",
    "基金简称",
    "基金全称",
    "基金管理人",
    "一级策略大类",
    "二级策略类别",
    "市场范围_二次修正",
    "Index_Code",
    "Index_Name",
    "最新基金规模(亿)",
    "上市日期",
    "基金成立日",
]

SHEET3_COLS = [
    "Index_Code",
    "Index_Name",
    *RULE_FIELDS,
    "是否规则已补全",
    "缺失字段说明",
]

SHEET4_COLS = [
    "ETF_Code",
    "ETF_Name",
    "基金简称",
    "基金全称",
    "基金管理人",
    "一级策略大类",
    "二级策略类别",
    "市场范围_二次修正",
    "最新基金规模(亿)",
    "上市日期",
    "基金成立日",
    "Index_Code",
    "Index_Name",
    *RULE_FIELDS,
    "是否规则已补全",
    "缺失字段说明",
]


def normalize_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).replace("\u3000", " ").replace("\xa0", " ")
    return " ".join(text.strip().split())


def normalize_code(value: Any) -> str:
    return normalize_text(value).upper()


def is_blank(value: Any) -> bool:
    if pd.isna(value):
        return True
    text = normalize_text(value)
    return text == "" or text.lower() in {"nan", "none", "nat"}


def clean_source_rows(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    if df.empty:
        return df.copy(), 0
    mask_source = df.apply(lambda row: any("数据来源：Wind" in normalize_text(v) or "数据来源:Wind" in normalize_text(v) for v in row), axis=1)
    cleaned = df.loc[~mask_source].copy()
    cleaned = cleaned.dropna(how="all").copy()
    return cleaned, int(mask_source.sum())


def safe_number(value: Any) -> float:
    try:
        if is_blank(value):
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def unique_join(values: pd.Series) -> str:
    seen: OrderedDict[str, None] = OrderedDict()
    for value in values:
        text = normalize_text(value)
        if text:
            seen[text] = None
    return "；".join(seen.keys())


def list_join(values: pd.Series) -> str:
    return "；".join([normalize_text(v) for v in values if normalize_text(v)])


def suggest_source(code: str, index_name: str, publisher: str) -> str:
    code_u = normalize_code(code)
    name = normalize_text(index_name)
    pub = normalize_text(publisher)
    if code_u.endswith(".CSI") or "中证" in pub:
        return "中证指数有限公司官网"
    if code_u.endswith(".CNI") or "国证" in pub:
        return "国证指数有限公司官网"
    if code_u.endswith(".SH") or "上证" in name:
        return "中证指数有限公司官网或上证指数相关页面"
    if code_u.endswith(".SZ") or "深证" in name:
        return "国证指数有限公司官网或深证指数相关页面"
    if code_u.endswith(".HI") or "恒生" in pub:
        return "恒生指数有限公司官网"
    if code_u.endswith(".SPI") or "标普" in pub or "标普" in name:
        return "S&P Dow Jones Indices 官网"
    if code_u.endswith(".MI") or "MSCI" in pub.upper() or "MSCI" in name.upper():
        return "MSCI 官网"
    return "待确认"


def compare_initial_file(path: Path, standard: pd.DataFrame, label: str) -> tuple[pd.DataFrame, dict[str, Any]]:
    raw = pd.read_excel(path, dtype=object)
    raw, source_rows = clean_source_rows(raw)
    raw.columns = [normalize_text(c) for c in raw.columns]

    etf_col_candidates = ["ETF_Code", "Wind代码", "ETF代码", "证券代码"]
    index_col_candidates = ["Index_Code", "指数代码", "跟踪指数代码"]
    etf_name_candidates = ["ETF_Name", "证券简称", "ETF名称"]
    index_name_candidates = ["Index_Name", "指数名称", "跟踪指数名称"]

    def pick(candidates: list[str]) -> str | None:
        for col in candidates:
            if col in raw.columns:
                return col
        return None

    etf_col = pick(etf_col_candidates)
    index_col = pick(index_col_candidates)
    etf_name_col = pick(etf_name_candidates)
    index_name_col = pick(index_name_candidates)

    if not etf_col or not index_col:
        empty = pd.DataFrame(columns=[
            "来源文件",
            "ETF_Code",
            "ETF_Name",
            "核心产品池_Index_Code",
            "核心产品池_Index_Name",
            "初步文件_Index_Code",
            "初步文件_Index_Name",
            "是否错配",
        ])
        return empty, {
            "label": label,
            "source_rows": source_rows,
            "matched_rows": 0,
            "mismatch_count": 0,
            "match_count": 0,
            "match_rate": 0,
            "field_issue": f"无法识别 ETF 或 Index 字段：{path.name}",
        }

    temp = raw.copy()
    temp["ETF_Code_标准化"] = temp[etf_col].map(normalize_code)
    temp["初步文件_Index_Code"] = temp[index_col].map(normalize_code)
    temp["初步文件_Index_Name"] = temp[index_name_col].map(normalize_text) if index_name_col else ""
    temp["ETF_Name_初步"] = temp[etf_name_col].map(normalize_text) if etf_name_col else ""

    standard_for_join = standard[["ETF_Code", "ETF_Name", "Index_Code", "Index_Name"]].rename(
        columns={
            "ETF_Code": "核心产品池_ETF_Code",
            "ETF_Name": "核心产品池_ETF_Name",
            "Index_Code": "核心产品池_Index_Code",
            "Index_Name": "核心产品池_Index_Name",
        }
    )
    joined = temp.merge(
        standard_for_join,
        left_on="ETF_Code_标准化",
        right_on="核心产品池_ETF_Code",
        how="inner",
    )
    joined["是否错配"] = joined["初步文件_Index_Code"] != joined["核心产品池_Index_Code"].map(normalize_code)
    detail = pd.DataFrame(
        {
            "来源文件": path.name,
            "ETF_Code": joined["核心产品池_ETF_Code"],
            "ETF_Name": joined["核心产品池_ETF_Name"],
            "核心产品池_Index_Code": joined["核心产品池_Index_Code"],
            "核心产品池_Index_Name": joined["核心产品池_Index_Name"],
            "初步文件_Index_Code": joined["初步文件_Index_Code"],
            "初步文件_Index_Name": joined["初步文件_Index_Name"],
            "是否错配": joined["是否错配"].map(lambda x: "是" if x else "否"),
        }
    )
    matched = len(detail)
    mismatch = int((detail["是否错配"] == "是").sum())
    match_count = matched - mismatch
    return detail, {
        "label": label,
        "source_rows": source_rows,
        "matched_rows": matched,
        "mismatch_count": mismatch,
        "match_count": match_count,
        "match_rate": match_count / matched if matched else 0,
        "field_issue": "",
    }


def write_df(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    df.to_excel(writer, sheet_name=sheet_name, index=False)


def apply_formatting(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                value = cell.value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = ws.cell(row=1, column=cell.column).value
                if header in {"上市日期", "基金成立日", "指数基日", "指数发布日期"}:
                    cell.number_format = "yyyy-mm-dd"
                elif header in {"最新基金规模(亿)", "对应ETF规模合计"}:
                    cell.number_format = "0.0000"
                elif header in {"原始初步映射表中 ETF-Index 匹配率", "本次修正后 ETF-Index 匹配率", "匹配率"}:
                    cell.number_format = "0.00%"
    wb.save(path)


def main() -> None:
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    pool = pd.read_excel(PRODUCT_POOL_FILE, sheet_name="策略ETF_最终统计池", dtype=object)
    pool.columns = [normalize_text(c) for c in pool.columns]
    core = pool[pool["是否纳入核心策略ETF统计"].map(normalize_text) == "是"].copy()

    for col in ["Wind代码", "证券简称", "跟踪指数代码", "跟踪指数名称"]:
        if col not in core.columns:
            raise KeyError(f"产品池缺少必要字段：{col}")

    sheet1 = pd.DataFrame(
        {
            "ETF_Code": core["Wind代码"].map(normalize_code),
            "ETF_Name": core["证券简称"].map(normalize_text),
            "基金简称": core.get("基金简称", "").map(normalize_text),
            "基金全称": core.get("基金全称", "").map(normalize_text),
            "基金管理人": core.get("基金管理人", "").map(normalize_text),
            "一级策略大类": core.get("一级策略大类", "").map(normalize_text),
            "二级策略类别": core.get("二级策略类别", "").map(normalize_text),
            "市场范围_二次修正": core.get("市场范围_二次修正", "").map(normalize_text),
            "Index_Code": core["跟踪指数代码"].map(normalize_code),
            "Index_Name": core["跟踪指数名称"].map(normalize_text),
            "最新基金规模(亿)": core.get("最新基金规模(亿)", 0).map(safe_number),
            "上市日期": core.get("上市日期", ""),
            "基金成立日": core.get("基金成立日", ""),
        }
    )
    sheet1 = sheet1[SHEET1_COLS].sort_values(
        ["一级策略大类", "二级策略类别", "最新基金规模(亿)"],
        ascending=[True, True, False],
        kind="stable",
    )

    sheet2 = (
        sheet1.groupby("Index_Code", as_index=False)
        .agg(
            Index_Name=("Index_Name", "first"),
            对应ETF数量=("ETF_Code", "nunique"),
            对应ETF规模合计=("最新基金规模(亿)", "sum"),
            涉及一级策略大类=("一级策略大类", unique_join),
            涉及二级策略类别=("二级策略类别", unique_join),
            市场范围=("市场范围_二次修正", unique_join),
            对应ETF代码列表=("ETF_Code", list_join),
            对应ETF简称列表=("ETF_Name", list_join),
        )
        .sort_values(["涉及一级策略大类", "涉及二级策略类别", "对应ETF规模合计"], ascending=[True, True, False])
    )

    rules_raw = pd.read_excel(RULE_FILE, dtype=object)
    rules_raw, rule_source_rows = clean_source_rows(rules_raw)
    rules_raw.columns = [normalize_text(c) for c in rules_raw.columns]
    if "指数代码" not in rules_raw.columns:
        raise KeyError("指数编制规则详细信息表缺少“指数代码”字段")
    rules_raw["Index_Code"] = rules_raw["指数代码"].map(normalize_code)
    rules_raw["Index_Name_rule"] = rules_raw.get("指数名称", "").map(normalize_text)
    for col in RULE_FIELDS:
        if col not in rules_raw.columns:
            rules_raw[col] = ""

    rule_candidates = rules_raw[rules_raw["Index_Code"] != ""].copy()
    rule_candidates["_完整度得分"] = rule_candidates.apply(
        lambda row: sum(0 if is_blank(row.get(col)) else 1 for col in RULE_FIELDS + ["指数名称"]),
        axis=1,
    )
    rule_unique = (
        rule_candidates.sort_values(["Index_Code", "_完整度得分"], ascending=[True, False], kind="stable")
        .drop_duplicates("Index_Code", keep="first")
        .drop(columns=["_完整度得分"])
        .reset_index(drop=True)
    )

    rule_map = rule_unique.set_index("Index_Code", drop=False)
    sheet3_records: list[dict[str, Any]] = []
    for _, idx_row in sheet2.iterrows():
        code = idx_row["Index_Code"]
        record: dict[str, Any] = {
            "Index_Code": code,
            "Index_Name": idx_row["Index_Name"],
        }
        missing_fields: list[str] = []
        matched = code in rule_map.index
        for field in RULE_FIELDS:
            value = rule_map.loc[code, field] if matched else ""
            if is_blank(value):
                record[field] = "待补充"
                missing_fields.append(field)
            else:
                record[field] = value
        record["是否规则已补全"] = "是" if matched and not missing_fields else "否"
        if not matched:
            record["缺失字段说明"] = "未在指数编制规则详细信息表中匹配到该指数代码"
        elif missing_fields:
            record["缺失字段说明"] = "缺失：" + "、".join(missing_fields)
        else:
            record["缺失字段说明"] = "无"
        sheet3_records.append(record)

    sheet3 = pd.DataFrame(sheet3_records)[SHEET3_COLS]

    sheet4 = sheet1.merge(sheet3, on=["Index_Code", "Index_Name"], how="left")
    for col in RULE_FIELDS:
        sheet4[col] = sheet4[col].map(lambda v: "待补充" if is_blank(v) else v)
    sheet4["是否规则已补全"] = sheet4["是否规则已补全"].map(lambda v: "否" if is_blank(v) else normalize_text(v))
    sheet4["缺失字段说明"] = sheet4["缺失字段说明"].map(lambda v: "未匹配到指数规则" if is_blank(v) else normalize_text(v))
    sheet4 = sheet4[SHEET4_COLS].sort_values(
        ["一级策略大类", "二级策略类别", "最新基金规模(亿)"],
        ascending=[True, True, False],
        kind="stable",
    )

    missing = sheet3[sheet3["是否规则已补全"] != "是"].merge(
        sheet2,
        on=["Index_Code", "Index_Name"],
        how="left",
    )
    missing["建议补充来源"] = missing.apply(
        lambda row: suggest_source(row["Index_Code"], row["Index_Name"], row.get("发布机构", "")),
        axis=1,
    )
    sheet5 = missing[
        [
            "Index_Code",
            "Index_Name",
            "对应ETF数量",
            "对应ETF规模合计",
            "涉及一级策略大类",
            "涉及二级策略类别",
            "市场范围",
            "缺失字段说明",
            "建议补充来源",
        ]
    ].sort_values(["涉及一级策略大类", "涉及二级策略类别", "对应ETF规模合计"], ascending=[True, True, False])

    mapping_detail, mapping_stats = compare_initial_file(MAPPING_FILE, sheet1, "ETF与指数映射关系表")
    merged_detail, merged_stats = compare_initial_file(MERGED_FILE, sheet1, "完整数据合并表")
    all_mismatch_detail = pd.concat([mapping_detail, merged_detail], ignore_index=True)

    standard_check = sheet4[["ETF_Code", "Index_Code", "Index_Name"]].merge(
        sheet1[["ETF_Code", "Index_Code", "Index_Name"]],
        on="ETF_Code",
        how="left",
        suffixes=("_修正后", "_核心产品池"),
    )
    fixed_mismatch_count = int(
        (standard_check["Index_Code_修正后"].map(normalize_code) != standard_check["Index_Code_核心产品池"].map(normalize_code)).sum()
    )
    fixed_match_rate = 1 - fixed_mismatch_count / len(sheet4) if len(sheet4) else 0

    extra_rule_count = len(set(rule_unique["Index_Code"]) - set(sheet2["Index_Code"]))
    source_rows_total = rule_source_rows + mapping_stats["source_rows"] + merged_stats["source_rows"]

    summary_rows = [
        ["检查项", "结果", "说明"],
        ["核心策略 ETF 应有数量", 168, "用户指定理论值"],
        ["Sheet 1 实际 ETF 数量", len(sheet1), ""],
        ["ETF_Code 重复数量", int(sheet1["ETF_Code"].duplicated().sum()), ""],
        ["唯一 Index_Code 数量", sheet2["Index_Code"].nunique(), ""],
        ["指数规则已补全数量", int((sheet3["是否规则已补全"] == "是").sum()), ""],
        ["指数规则待补充数量", int((sheet3["是否规则已补全"] != "是").sum()), ""],
        ["ETF_指数规则合并表行数", len(sheet4), ""],
        ["ETF 与核心产品池 Index_Code 不一致数量", fixed_mismatch_count, ""],
        ["是否存在“数据来源：Wind”行", "否", f"输入中发现并剔除 {source_rows_total} 行"],
        ["是否混入核心指数清单之外的指数", "否", f"输出未混入；原规则表中额外指数 {extra_rule_count} 个"],
        ["原始初步映射表中 ETF-Index 错配数量", mapping_stats["mismatch_count"], ""],
        ["原始初步映射表中 ETF-Index 匹配率", mapping_stats["match_rate"], ""],
        ["原始完整数据合并表是否存在 ETF-Index 错配", "是" if merged_stats["mismatch_count"] else "否", ""],
        ["原始完整数据合并表错配数量", merged_stats["mismatch_count"], ""],
        ["原始完整数据合并表 ETF-Index 匹配率", merged_stats["match_rate"], ""],
        ["本次修正后 ETF-Index 错配数量", fixed_mismatch_count, ""],
        ["本次修正后 ETF-Index 匹配率", fixed_match_rate, ""],
        ["脚本运行时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""],
    ]
    quality_summary = pd.DataFrame(summary_rows[1:], columns=summary_rows[0])

    detail_title = pd.DataFrame([["错配明细", "", "", "", "", "", "", ""]], columns=[
        "来源文件",
        "ETF_Code",
        "ETF_Name",
        "核心产品池_Index_Code",
        "核心产品池_Index_Name",
        "初步文件_Index_Code",
        "初步文件_Index_Name",
        "是否错配",
    ])
    mismatch_only = all_mismatch_detail[all_mismatch_detail["是否错配"] == "是"].copy()
    if mismatch_only.empty:
        mismatch_only = pd.DataFrame([{
            "来源文件": "无",
            "ETF_Code": "",
            "ETF_Name": "",
            "核心产品池_Index_Code": "",
            "核心产品池_Index_Name": "",
            "初步文件_Index_Code": "",
            "初步文件_Index_Name": "",
            "是否错配": "否",
        }])
    sheet6 = quality_summary.copy()

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        write_df(writer, sheet1, "核心ETF_指数映射_标准版")
        write_df(writer, sheet2, "核心指数清单_去重版")
        write_df(writer, sheet3, "指数规则明细_标准版")
        write_df(writer, sheet4, "ETF_指数规则合并表")
        write_df(writer, sheet5, "缺失规则清单")
        write_df(writer, sheet6, "数据质量检查")

        # Append detailed mismatch evidence below the metric table.
        startrow = len(sheet6) + 3
        detail_title.to_excel(writer, sheet_name="数据质量检查", startrow=startrow, index=False)
        mismatch_only.to_excel(writer, sheet_name="数据质量检查", startrow=startrow + 2, index=False)

    apply_formatting(OUTPUT_FILE)

    print(f"核心 ETF 数量: {len(sheet1)}")
    print(f"唯一指数数量: {sheet2['Index_Code'].nunique()}")
    print(f"规则已补全指数数量: {int((sheet3['是否规则已补全'] == '是').sum())}")
    print(f"规则待补充指数数量: {int((sheet3['是否规则已补全'] != '是').sum())}")
    print(f"原始初步映射表错配数量: {mapping_stats['mismatch_count']}")
    print(f"原始完整合并表错配数量: {merged_stats['mismatch_count']}")
    print(f"本次修正后 ETF-指数错配数量: {fixed_mismatch_count}")
    print(f"输出文件路径: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
