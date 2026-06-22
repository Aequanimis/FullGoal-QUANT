"""Step 9：生成策略分类框架并同步主 Excel。"""

from __future__ import annotations

import csv
import io
import unicodedata
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
ATTACHMENT = Path(
    r"C:\Users\lvdon\.codex\attachments\dafe8830-acc7-4021-901b-7ce706a59f4a\pasted-text.txt"
)
REPRESENTATIVE_EXCEL = PROJECT_ROOT / "01_processed_data" / "classification" / "境内策略ETF产品池_代表产品筛选版.xlsx"
INDEX_RULES_EXCEL = PROJECT_ROOT / "01_processed_data" / "index_rules" / "代表指数规则整理.xlsx"
MAIN_EXCEL = PROJECT_ROOT / "02_outputs" / "excel" / "境内策略ETF产品梳理_初版.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "01_processed_data" / "classification"
OUTPUT_CSV = OUTPUT_DIR / "策略分类表.csv"
OUTPUT_EXCEL = OUTPUT_DIR / "策略分类表.xlsx"
REVIEW_MD = OUTPUT_DIR / "step9_strategy_classification.md"
LOG_MD = OUTPUT_DIR / "step9_strategy_classification_log.md"

FIELDS = [
    "策略类型",
    "核心因子",
    "投资逻辑",
    "代表指数",
    "代表ETF",
    "适合市场环境",
    "主要风险",
    "境内发展成熟度",
    "备注",
]

MAINLINE_TYPES = [
    "红利",
    "红利低波",
    "自由现金流",
    "红利质量",
    "质量",
    "价值",
    "成长",
    "低波",
    "基本面策略",
]

SUPPLEMENTAL_TYPES = [
    "指数增强/多因子",
    "央国企红利/股东回报",
    "ESG",
    "港股通红利/低波",
    "等权/另类加权",
    "动量",
]

PENDING_TYPES = ["等权/另类加权", "动量"]

STEP8_FIX = {
    "Step8指数规则核验状态": "部分核验",
    "Step8指数规则简述": "富国中证价值ETF对应中证国信价值指数，价值策略规则已部分核验，仍需后续补齐加权方式、调样频率和权重限制。",
    "Step8信息来源URL": "https://oss-ch.csindex.com.cn/static/html/csindex/public/uploads/indices/detail/files/zh_CN/20231208180210-931052_Index_Methodology_cn.pdf",
    "Step8待补充事项": "Step8 回写时因跟踪指数名称不一致未匹配，已按基金代码 512040 补写；后续仍需二次核验完整指数规则。",
}


def clean(value: object) -> str:
    return "" if value is None else str(value).strip(" \t\r\n")


def verify_inputs() -> None:
    missing = [
        path for path in (REPRESENTATIVE_EXCEL, INDEX_RULES_EXCEL, MAIN_EXCEL) if not path.exists()
    ]
    if missing:
        raise FileNotFoundError("缺少输入文件：" + "；".join(str(path) for path in missing))

    representative_book = load_workbook(REPRESENTATIVE_EXCEL, read_only=True)
    if "产品池_代表产品筛选版" not in representative_book.sheetnames:
        raise ValueError("代表产品筛选版缺少 Sheet【产品池_代表产品筛选版】")
    rules_book = load_workbook(INDEX_RULES_EXCEL, read_only=True)
    if "代表指数规则表" not in rules_book.sheetnames:
        raise ValueError("指数规则整理缺少 Sheet【代表指数规则表】")


def read_data_block() -> list[dict[str, str]]:
    text = ATTACHMENT.read_text(encoding="utf-8-sig")
    header = "\t".join(FIELDS)
    start = text.find(header)
    if start < 0:
        raise ValueError("附件中未找到策略分类 TSV 表头")
    end = text.find("\n\n五、输出 Excel 要求", start)
    if end < 0:
        raise ValueError("附件中未找到策略分类 TSV 结束位置")
    block = text[start:end].strip("\r\n")
    reader = csv.DictReader(io.StringIO(block), delimiter="\t")
    if reader.fieldnames != FIELDS:
        raise ValueError(f"策略分类 TSV 字段不匹配：{reader.fieldnames}")
    rows = [{field: clean(record.get(field, "")) for field in FIELDS} for record in reader]
    if [row["策略类型"] for row in rows] != [*MAINLINE_TYPES, *SUPPLEMENTAL_TYPES]:
        raise ValueError("DATA_BLOCK 策略类型或顺序与任务要求不一致")
    return rows


def write_csv(rows: list[dict[str, str]]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def display_width(value: object) -> int:
    text = clean(value)
    return sum(
        2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1
        for char in text
    )


def append_rows(worksheet, fields: list[str], rows: list[dict[str, object]]) -> None:
    worksheet.append(fields)
    for record in rows:
        worksheet.append([record.get(field, "") for field in fields])


def format_sheet(worksheet) -> None:
    max_row = max(worksheet.max_row, 1)
    max_column = max(worksheet.max_column, 1)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_column)}{max_row}"
    worksheet.sheet_view.zoomScale = 90

    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    header_font = Font(name="Microsoft YaHei", size=10, bold=True)
    body_font = Font(name="Microsoft YaHei", size=10)
    border = Border(bottom=Side(style="thin", color="A6A6A6"))
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    worksheet.row_dimensions[1].height = 28

    headers = [clean(cell.value) for cell in worksheet[1]]
    for row in worksheet.iter_rows(min_row=2, max_row=max_row, max_col=max_column):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_index, header in enumerate(headers, start=1):
        values = [worksheet.cell(row, column_index).value for row in range(1, max_row + 1)]
        measured = max(display_width(value) for value in values) + 3
        cap = 52 if header in set(FIELDS) - {"策略类型", "境内发展成熟度"} else 26
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(measured, 12), cap)
        if header == "代表ETF":
            for row in range(2, max_row + 1):
                worksheet.cell(row, column_index).number_format = "@"


def build_statistics(rows: list[dict[str, str]]) -> list[dict[str, object]]:
    maturity_high = sum("成熟度高" in row["境内发展成熟度"] for row in rows)
    rapid_growth = sum("快速发展" in row["境内发展成熟度"] for row in rows)
    maturity_pending = sum("待补充" in row["境内发展成熟度"] for row in rows)
    return [
        {"统计项目": "策略类型数量", "数量": len(rows)},
        {"统计项目": "主线策略数量", "数量": len(MAINLINE_TYPES)},
        {"统计项目": "补充观察策略数量", "数量": len(SUPPLEMENTAL_TYPES) - len(PENDING_TYPES)},
        {"统计项目": "待补充策略数量", "数量": len(PENDING_TYPES)},
        {"统计项目": "包含成熟度高的策略数量", "数量": maturity_high},
        {"统计项目": "包含快速发展的策略数量", "数量": rapid_growth},
        {"统计项目": "包含待补充的策略数量", "数量": maturity_pending},
    ]


def write_excel(rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    by_type = {row["策略类型"]: row for row in rows}

    for sheet_name, selected in [
        ("策略分类表", rows),
        ("主线策略", [by_type[name] for name in MAINLINE_TYPES]),
        ("补充观察与待补充", [by_type[name] for name in SUPPLEMENTAL_TYPES]),
    ]:
        worksheet = workbook.create_sheet(sheet_name)
        append_rows(worksheet, FIELDS, selected)
        format_sheet(worksheet)

    stats_sheet = workbook.create_sheet("分类统计")
    append_rows(stats_sheet, ["统计项目", "数量"], build_statistics(rows))
    format_sheet(stats_sheet)
    workbook.save(OUTPUT_EXCEL)


def sync_main_excel(rows: list[dict[str, str]]) -> None:
    workbook = load_workbook(MAIN_EXCEL)
    if "策略分类表" not in workbook.sheetnames:
        worksheet = workbook.create_sheet("策略分类表")
        worksheet.append(FIELDS)
    else:
        worksheet = workbook["策略分类表"]

    headers = [clean(cell.value) for cell in worksheet[1]]
    if not any(headers):
        worksheet.delete_rows(1, worksheet.max_row)
        worksheet.append(FIELDS)
        headers = FIELDS.copy()
    missing = set(FIELDS) - set(headers)
    if missing:
        raise ValueError(f"主 Excel 策略分类表缺少字段：{sorted(missing)}")
    columns = {header: index + 1 for index, header in enumerate(headers)}

    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)
    for row_number, record in enumerate(rows, start=2):
        for field in FIELDS:
            worksheet.cell(row_number, columns[field], record[field])

    last_row = len(rows) + 1
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    for row in worksheet.iter_rows(min_row=2, max_row=last_row, max_col=len(headers)):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if "代表ETF" in columns:
        for row in range(2, last_row + 1):
            worksheet.cell(row, columns["代表ETF"]).number_format = "@"
    for column_index, header in enumerate(headers, start=1):
        values = [worksheet.cell(row, column_index).value for row in range(1, last_row + 1)]
        measured = max(display_width(value) for value in values) + 3
        cap = 52 if header in set(FIELDS) - {"策略类型", "境内发展成熟度"} else 26
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(measured, 12), cap)
    workbook.save(MAIN_EXCEL)


def fix_512040_step8() -> tuple[bool, str]:
    workbook = load_workbook(REPRESENTATIVE_EXCEL)
    worksheet = workbook["产品池_代表产品筛选版"]
    headers = [clean(cell.value) for cell in worksheet[1]]
    required = {"基金代码", "产品名称", *STEP8_FIX.keys()}
    missing = required - set(headers)
    if missing:
        return False, "未修复：代表产品筛选版缺少字段：" + "、".join(sorted(missing))
    columns = {header: index + 1 for index, header in enumerate(headers)}

    matching_rows = []
    for row in range(2, worksheet.max_row + 1):
        code = clean(worksheet.cell(row, columns["基金代码"]).value)
        name = clean(worksheet.cell(row, columns["产品名称"]).value)
        if code == "512040" or "富国中证价值ETF" in name:
            matching_rows.append(row)
    if not matching_rows:
        return False, "未修复：未找到基金代码 512040 或富国中证价值ETF"
    if len(matching_rows) > 1:
        return False, f"未修复：匹配到 {len(matching_rows)} 行，无法唯一确定"

    row = matching_rows[0]
    current_status = clean(worksheet.cell(row, columns["Step8指数规则核验状态"]).value)
    current_summary = clean(worksheet.cell(row, columns["Step8指数规则简述"]).value)
    current_url = clean(worksheet.cell(row, columns["Step8信息来源URL"]).value)
    needs_fix = current_status in {"", "未匹配"} or not current_summary or not current_url
    if not needs_fix:
        return True, "无需修改：512040 Step8 指数规则字段已完整"

    for field, value in STEP8_FIX.items():
        cell = worksheet.cell(row, columns[field], value)
        cell.font = Font(name="Microsoft YaHei", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if field == "Step8信息来源URL":
            cell.number_format = "@"
    workbook.save(REPRESENTATIVE_EXCEL)
    return True, "已按基金代码 512040 补写中证国信价值指数的部分核验信息"


def write_review() -> None:
    content = """# Step 9 策略分类说明

## 1. Step 9 目标

将产品池、代表产品和指数规则进一步归纳为统一的境内策略 ETF 分类框架，为后续统计分析、汇报和 PPT 制作提供直接输入。

## 2. 策略分类表的用途

分类表集中呈现每类策略的核心因子、投资逻辑、代表指数、代表 ETF、适用环境、主要风险及境内发展成熟度，可用于比较不同策略的逻辑边界与发展阶段。

本轮 Step 9 将前面整理的产品池、代表产品和指数规则上升为策略分类框架。当前主线策略包括红利、红利低波、自由现金流、红利质量、质量、价值、成长、低波和基本面策略；指数增强/多因子、央国企红利/股东回报、ESG 和港股通红利/低波暂作为单列或补充观察；等权/另类加权和动量属于研报口径下的策略类 ETF 重要补充方向，但当前产品池尚未系统纳入，后续需通过 Wind / Choice 继续补齐。本表后续可直接用于 PPT 中的“境内策略 ETF 分类框架”页面。

## 3. 主线策略分类

- 红利、红利低波、自由现金流、红利质量；
- 质量、价值、成长、低波；
- 基本面策略。

## 4. 补充观察分类

- 指数增强/多因子；
- 央国企红利/股东回报；
- ESG；
- 港股通红利/低波。

## 5. 当前产品池仍需补充的分类

- 等权/另类加权；
- 动量；
- 纯质量 ETF 也需要在现有红利质量产品之外继续补充。

## 6. Step 10 统计分析注意事项

- 区分产品数量与代表指数数量，避免同一指数对应多个 ETF 时重复统计。
- 将主线、补充观察、跨境补充和待补充类别分开展示。
- 规模、成交额、费率、成立日期和上市日期仍需 Wind / Choice 补齐后再做市场发展判断。
- 对红利质量、指数增强和央国企股东回报等边界类别保留分类说明。
"""
    REVIEW_MD.write_text(content, encoding="utf-8")


def write_log(rows: list[dict[str, str]], fix_attempted: bool, fix_result: str) -> None:
    content = f"""# Step 9 策略分类处理日志

1. 输入文件路径：
   - 代表产品筛选版：`{REPRESENTATIVE_EXCEL}`
   - 指数规则整理：`{INDEX_RULES_EXCEL}`
   - 主 Excel：`{MAIN_EXCEL}`
2. 输出文件路径：
   - 策略分类 CSV：`{OUTPUT_CSV}`
   - 策略分类 Excel：`{OUTPUT_EXCEL}`
   - Step 9 说明：`{REVIEW_MD}`
   - 同步主 Excel：`{MAIN_EXCEL}`
3. DATA_BLOCK 策略类型数量：{len(rows)}
4. 主线策略数量：{len(MAINLINE_TYPES)}
5. 补充观察与待补充策略数量：{len(SUPPLEMENTAL_TYPES)}
6. 是否成功生成策略分类表 Excel：是
7. 是否成功同步更新主 Excel：是
8. 是否尝试修复 512040 Step8 回写问题：{'是' if fix_attempted else '否'}
9. 512040 Step8 回写修复结果：{fix_result}
"""
    LOG_MD.write_text(content, encoding="utf-8")


def validate(rows: list[dict[str, str]]) -> None:
    if OUTPUT_CSV.read_bytes()[:3] != b"\xef\xbb\xbf":
        raise ValueError("策略分类 CSV 不是 UTF-8 with BOM")
    with OUTPUT_CSV.open("r", encoding="utf-8-sig", newline="") as csv_file:
        loaded = list(csv.DictReader(csv_file))
    if loaded != rows:
        raise ValueError("策略分类 CSV 回读不一致")

    workbook = load_workbook(OUTPUT_EXCEL)
    expected = ["策略分类表", "主线策略", "补充观察与待补充", "分类统计"]
    if workbook.sheetnames != expected:
        raise ValueError(f"策略分类 Excel Sheet 不匹配：{workbook.sheetnames}")
    expected_rows = {
        "策略分类表": len(rows),
        "主线策略": len(MAINLINE_TYPES),
        "补充观察与待补充": len(SUPPLEMENTAL_TYPES),
    }
    for sheet_name, count in expected_rows.items():
        if workbook[sheet_name].max_row != count + 1:
            raise ValueError(f"{sheet_name} 行数不匹配")
    for worksheet in workbook.worksheets:
        if worksheet.freeze_panes != "A2" or not worksheet.auto_filter.ref:
            raise ValueError(f"{worksheet.title} 的冻结窗格或筛选未生效")

    main_workbook = load_workbook(MAIN_EXCEL)
    main_sheet = main_workbook["策略分类表"]
    if main_sheet.max_row != len(rows) + 1:
        raise ValueError("主 Excel 策略分类表行数不匹配")

    representative_book = load_workbook(REPRESENTATIVE_EXCEL)
    worksheet = representative_book["产品池_代表产品筛选版"]
    headers = [clean(cell.value) for cell in worksheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}
    for row in range(2, worksheet.max_row + 1):
        if clean(worksheet.cell(row, columns["基金代码"]).value) == "512040":
            if clean(worksheet.cell(row, columns["Step8指数规则核验状态"]).value) != "部分核验":
                raise ValueError("512040 Step8 回写修复校验失败")
            break


def main() -> None:
    verify_inputs()
    rows = read_data_block()
    write_csv(rows)
    write_excel(rows)
    sync_main_excel(rows)
    fix_attempted = True
    fix_success, fix_result = fix_512040_step8()
    write_review()
    validate(rows)
    write_log(rows, fix_attempted, fix_result)

    print(f"strategies={len(rows)}")
    print(f"mainline={len(MAINLINE_TYPES)}")
    print(f"supplemental_and_pending={len(SUPPLEMENTAL_TYPES)}")
    print(f"fix_512040_success={fix_success}")
    print(f"fix_512040_result={fix_result}")


if __name__ == "__main__":
    main()
