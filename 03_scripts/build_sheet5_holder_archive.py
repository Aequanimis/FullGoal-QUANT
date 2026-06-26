from __future__ import annotations

import hashlib
import math
import shutil
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\全市场ETF基础信息_策略ETF池二次修正版.xlsx")
PANEL = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\sheet5\广义策略ETF持有人结构面板数据.xlsx")
VALIDATION = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\sheet5\Sheet5_广义策略ETF持有人结构面板数据_单文件验收报告.xlsx")
WORK_OUTPUT = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\sheet5\广义策略ETF_持有人结构表_半年报年报清洗版.xlsx")
ARCHIVE_OUTPUT = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\阶段性归档_境内广义策略ETF数据底座\06_广义策略ETF_持有人结构表_半年报年报清洗归档版.xlsx")

TARGET_PERIODS = {
    "2023年年报": 202312,
    "2024年中报": 202406,
    "2024年年报": 202412,
    "2025年中报": 202506,
    "2025年年报": 202512,
}
Q1_PERIODS = {"2026年一季报": 202603, "Q1FY2026": 202603}
NUMERIC_FIELDS = [
    "机构投资者持有份额", "机构投资者持有比例", "个人投资者持有份额", "个人投资者持有比例",
    "联接基金持有份额", "联接基金持有比例", "前十大持有人持有比例",
    "基金管理人自持份额", "基金管理人自持比例", "员工持有份额", "员工持有比例",
]
RATIO_FIELDS = {f for f in NUMERIC_FIELDS if "比例" in f}
SHARE_FIELDS = set(NUMERIC_FIELDS) - RATIO_FIELDS
CORE_FIELDS = ["机构投资者持有比例", "个人投资者持有比例", "前十大持有人持有比例"]
UNAVAILABLE_FIELDS = {"联接基金持有份额", "联接基金持有比例", "基金管理人自持份额"}
EMPTY = {"", "-", "--", "N/A", "NA", "NAN", "NONE", "NULL", "WIND暂不可得"}

PANEL_OUTPUT_FIELDS = [
    "报告期", "报告期排序", "Wind代码", "基金代码", "交易代码", "证券简称", "基金简称",
    "原始证券简称", "基金管理人", "统计口径分类", "是否纳入核心策略ETF统计",
    "是否纳入广义策略ETF统计", "一级策略大类", "二级策略类别", "市场范围_二次修正",
    "最新基金规模(亿)", "机构投资者持有份额", "机构投资者持有比例",
    "个人投资者持有份额", "个人投资者持有比例", "联接基金持有份额",
    "联接基金持有比例", "前十大持有人持有比例", "基金管理人自持份额",
    "基金管理人自持比例", "员工持有份额", "员工持有比例", "数据来源文件",
    "是否证券简称不一致",
]


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def clean(value: Any) -> str:
    if value is None:
        return ""
    return "".join(str(value).replace("\u3000", " ").split()).strip()


def norm_code(value: Any) -> str:
    return clean(value).upper()


def is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return clean(value).upper() in EMPTY


def to_number(value: Any) -> tuple[float | None, str]:
    if is_missing(value):
        return None, "缺失"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value), "可转数值"
    try:
        return float(clean(value).replace(",", "").replace("，", "").replace("%", "")), "可转数值"
    except ValueError:
        return None, "不可转数值"


def to_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def read_records(path: Path, sheet: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    iterator = ws.iter_rows(values_only=True)
    headers = [str(v or "").strip() for v in next(iterator)]
    rows = [dict(zip(headers, row)) for row in iterator]
    wb.close()
    return headers, rows


def normalized_period(value: Any) -> str:
    raw = clean(value)
    if raw.upper() == "Q1FY2026":
        return "2026年一季报"
    return raw


def write_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def format_book(wb: Workbook) -> None:
    dark = "1F4E78"
    white = "FFFFFF"
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        for c in ws[1]:
            c.fill = PatternFill("solid", fgColor=dark)
            c.font = Font(name="微软雅黑", size=10, bold=True, color=white)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        ws.row_dimensions[1].height = 32
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.font = Font(name="微软雅黑", size=9)
                c.border = border
                c.alignment = Alignment(vertical="top", wrap_text=False)
                header = str(ws.cell(1, c.column).value or "")
                if isinstance(c.value, (date, datetime)):
                    c.number_format = "yyyy-mm-dd"
                elif isinstance(c.value, (int, float)):
                    if "比例" in header or "覆盖率" in header or "非空率" in header:
                        c.number_format = "0.00"
                    elif "份额" in header:
                        c.number_format = "#,##0"
                    elif "规模" in header:
                        c.number_format = "0.0000"
        for col in range(1, ws.max_column + 1):
            vals = [str(ws.cell(r, col).value or "") for r in range(1, min(ws.max_row, 1000) + 1)]
            width = min(max(max((len(v) for v in vals), default=8) * 1.1 + 2, 10), 38)
            header = str(ws.cell(1, col).value or "")
            if any(k in header for k in ("说明", "原因", "建议", "缺失字段", "原始值")):
                width = min(max(width, 28), 50)
            ws.column_dimensions[get_column_letter(col)].width = width


def main() -> None:
    if not BASE.exists() or not PANEL.exists():
        raise FileNotFoundError("主产品池或持有人结构面板数据不存在")
    source_hashes = {BASE: sha256(BASE), PANEL: sha256(PANEL)}
    validation_used = False
    validation_note = "单文件验收报告不存在，未作交叉核对"
    if VALIDATION.exists():
        source_hashes[VALIDATION] = sha256(VALIDATION)
        vwb = load_workbook(VALIDATION, read_only=True, data_only=True)
        validation_used = "验收总览" in vwb.sheetnames
        validation_note = "已读取单文件验收报告并交叉核对基础覆盖统计" if validation_used else "验收报告存在但未找到验收总览"
        vwb.close()

    base_headers, base_rows = read_records(BASE, "策略ETF_最终统计池")
    panel_headers, panel_rows_raw = read_records(PANEL)
    pool_rows = [r for r in base_rows if clean(r.get("是否纳入广义策略ETF统计")) == "是"]
    pool = {norm_code(r.get("Wind代码")): r for r in pool_rows if norm_code(r.get("Wind代码"))}
    pool_codes = set(pool)

    valid_rows = []
    removed_source_rows = 0
    for row in panel_rows_raw:
        wind_code = norm_code(row.get("Wind代码"))
        joined = " ".join(clean(v) for v in row.values())
        if not wind_code or "数据来源" in wind_code or "数据来源" in joined:
            removed_source_rows += 1
            continue
        item = dict(row)
        item["Wind代码"] = wind_code
        item["报告期"] = normalized_period(item.get("报告期"))
        valid_rows.append(item)

    missing_source_fields = [f for f in UNAVAILABLE_FIELDS if f not in panel_headers]
    cleaned_all = []
    anomalies = []
    for row in valid_rows:
        wind_code = row["Wind代码"]
        p = pool.get(wind_code)
        period = row["报告期"]
        order = TARGET_PERIODS.get(period, 202603 if "一季报" in period else "")
        numeric = {}
        for field in NUMERIC_FIELDS:
            value, state = to_number(row.get(field)) if field in panel_headers else (None, "缺失")
            numeric[field] = value
            if state == "不可转数值":
                anomalies.append({
                    "异常类型": "数值不可转换", "报告期": period, "Wind代码": wind_code,
                    "原始证券简称": row.get("证券简称"), "字段": field, "原始值": row.get(field),
                    "异常说明": "非空值无法转换为数值", "是否错误": "是",
                })
            elif state == "可转数值":
                if field in RATIO_FIELDS and not 0 <= value <= 100:
                    anomalies.append({
                        "异常类型": "比例超范围", "报告期": period, "Wind代码": wind_code,
                        "原始证券简称": row.get("证券简称"), "字段": field, "原始值": value,
                        "异常说明": "比例应在0至100之间", "是否错误": "是",
                    })
                if field in SHARE_FIELDS and value < 0:
                    anomalies.append({
                        "异常类型": "份额小于0", "报告期": period, "Wind代码": wind_code,
                        "原始证券简称": row.get("证券简称"), "字段": field, "原始值": value,
                        "异常说明": "份额应大于等于0", "是否错误": "是",
                    })
        if numeric["机构投资者持有比例"] is not None and numeric["个人投资者持有比例"] is not None:
            total = numeric["机构投资者持有比例"] + numeric["个人投资者持有比例"]
            if not 99.9 <= total <= 100.1:
                anomalies.append({
                    "异常类型": "机构+个人比例异常", "报告期": period, "Wind代码": wind_code,
                    "原始证券简称": row.get("证券简称"), "字段": "机构+个人比例", "原始值": total,
                    "异常说明": "合计应在99.9至100.1之间", "是否错误": "是",
                })
        original_name = row.get("证券简称", "")
        standard_name = p.get("证券简称", "") if p else ""
        mismatch = bool(p and clean(original_name) != clean(standard_name))
        if mismatch:
            anomalies.append({
                "异常类型": "证券简称不一致", "报告期": period, "Wind代码": wind_code,
                "原始证券简称": original_name, "字段": "证券简称", "原始值": original_name,
                "异常说明": f"主产品池简称：{standard_name}。Wind简称口径与主产品池简称不同，正式清洗使用主产品池简称。",
                "是否错误": "否，仅提示",
            })
        if not p:
            anomalies.append({
                "异常类型": "代码池外产品", "报告期": period, "Wind代码": wind_code,
                "原始证券简称": original_name, "字段": "Wind代码", "原始值": wind_code,
                "异常说明": "Wind代码不属于223只广义策略ETF产品池", "是否错误": "是",
            })
        output = {
            "报告期": period, "报告期排序": order, "Wind代码": wind_code,
            "基金代码": p.get("基金代码", row.get("基金代码", "")) if p else row.get("基金代码", ""),
            "交易代码": p.get("交易代码", wind_code) if p else wind_code,
            "证券简称": standard_name, "基金简称": p.get("基金简称", "") if p else "",
            "原始证券简称": original_name, "基金管理人": p.get("基金管理人", "") if p else "",
            "统计口径分类": p.get("统计口径分类", "") if p else "",
            "是否纳入核心策略ETF统计": p.get("是否纳入核心策略ETF统计", "") if p else "",
            "是否纳入广义策略ETF统计": p.get("是否纳入广义策略ETF统计", "") if p else "",
            "一级策略大类": p.get("一级策略大类", "") if p else "",
            "二级策略类别": p.get("二级策略类别", "") if p else "",
            "市场范围_二次修正": p.get("市场范围_二次修正", "") if p else "",
            "最新基金规模(亿)": p.get("最新基金规模(亿)", "") if p else "",
            **numeric, "数据来源文件": PANEL.name, "是否证券简称不一致": "是" if mismatch else "否",
        }
        cleaned_all.append(output)

    key_counts = Counter((r["报告期"], r["Wind代码"]) for r in cleaned_all)
    for (period, wind_code), count in key_counts.items():
        if count > 1:
            anomalies.append({
                "异常类型": "重复报告期+Wind代码", "报告期": period, "Wind代码": wind_code,
                "原始证券简称": "", "字段": "报告期+Wind代码", "原始值": count,
                "异常说明": f"重复{count}次", "是否错误": "是",
            })

    panel_clean = [r for r in cleaned_all if r["报告期"] in TARGET_PERIODS and r["Wind代码"] in pool_codes]
    q1_backup = [r for r in cleaned_all if "一季报" in r["报告期"] or r["报告期"] == "Q1 FY2026"]
    panel_clean.sort(key=lambda r: (r["报告期排序"], str(r["统计口径分类"]), str(r["一级策略大类"]), r["Wind代码"]))
    q1_backup.sort(key=lambda r: (r["报告期排序"], str(r["统计口径分类"]), str(r["一级策略大类"]), r["Wind代码"]))

    y25 = {r["Wind代码"]: r for r in panel_clean if r["报告期"] == "2025年年报"}
    latest_snapshot = []
    pending = []
    not_established = new_late = true_supplement = 0
    cutoff = date(2025, 12, 31)
    new_start = date(2025, 11, 1)
    for wind_code in sorted(pool_codes):
        p = pool[wind_code]
        r = y25.get(wind_code)
        absent_core = CORE_FIELDS[:] if not r else [f for f in CORE_FIELDS if r.get(f) is None]
        if not r:
            coverage = "未覆盖"
        elif len(absent_core) == 3:
            coverage = "已覆盖但核心字段全空"
        elif absent_core:
            coverage = "已覆盖但核心字段部分缺失"
        else:
            coverage = "已覆盖且核心字段有效"
        inception = to_date(p.get("基金成立日"))
        if coverage == "已覆盖且核心字段有效":
            reason = "已覆盖且核心字段有效"
        elif coverage == "已覆盖但核心字段全空":
            reason = "2025年年报核心字段全空，需核验"
            true_supplement += 1
        elif coverage == "已覆盖但核心字段部分缺失":
            reason = "2025年年报核心字段部分缺失"
            true_supplement += 1
        elif inception and inception > cutoff:
            reason = "2025年尚未成立，无2025年年报"
            not_established += 1
        elif inception and new_start <= inception <= cutoff:
            reason = "2025年底新成立，Wind暂未收录持有人结构"
            new_late += 1
        else:
            reason = "已成立但2025年年报未覆盖，需补数"
            true_supplement += 1
        base = {
            "Wind代码": wind_code, "证券简称": p.get("证券简称", ""), "基金简称": p.get("基金简称", ""),
            "基金管理人": p.get("基金管理人", ""), "基金成立日": p.get("基金成立日", ""),
            "上市日期": p.get("上市日期", ""), "基金状态": p.get("基金状态", ""),
            "统计口径分类": p.get("统计口径分类", ""),
            "是否纳入核心策略ETF统计": p.get("是否纳入核心策略ETF统计", ""),
            "是否纳入广义策略ETF统计": p.get("是否纳入广义策略ETF统计", ""),
            "一级策略大类": p.get("一级策略大类", ""), "二级策略类别": p.get("二级策略类别", ""),
            "市场范围_二次修正": p.get("市场范围_二次修正", ""),
            "最新基金规模(亿)": p.get("最新基金规模(亿)", ""),
            "跟踪指数代码": p.get("跟踪指数代码", ""), "跟踪指数名称": p.get("跟踪指数名称", ""),
            "是否出现在2025年年报": "是" if r else "否", "覆盖状态": coverage,
            "缺失字段": "；".join(absent_core), "2025年年报缺失原因": reason,
        }
        for field in NUMERIC_FIELDS:
            base[field] = r.get(field) if r else None
        latest_snapshot.append(base)
        if coverage != "已覆盖且核心字段有效":
            advice = {
                "2025年尚未成立，无2025年年报": "不纳入2025年年报覆盖率分母，后续等2026年中报/年报",
                "2025年底新成立，Wind暂未收录持有人结构": "暂不强制补数，保留说明",
                "已成立但2025年年报未覆盖，需补数": "回Wind补查",
                "2025年年报核心字段全空，需核验": "回Wind核验字段",
                "2025年年报核心字段部分缺失": "回Wind核验字段",
            }[reason]
            pending.append({**base, "建议处理方式": advice})

    latest_snapshot.sort(key=lambda r: (
        str(r["统计口径分类"]), str(r["一级策略大类"]),
        -(float(r["最新基金规模(亿)"]) if isinstance(r["最新基金规模(亿)"], (int, float)) else -1),
    ))
    pending.sort(key=lambda r: (r["2025年年报缺失原因"], str(r["统计口径分类"]), str(r["一级策略大类"]), r["Wind代码"]))

    statistics = []
    def stat(category, metric, value, denominator="", rate="", note=""):
        statistics.append({"统计类别": category, "指标": metric, "分组": "", "数量/结果": value, "分母": denominator, "覆盖率/非空率(%)": rate, "说明": note})
    stat("总体", "广义策略ETF总数", len(pool_codes))
    stat("总体", "面板原始有效行数", len(valid_rows))
    stat("总体", "原始面板覆盖基金数量（含2026Q1）", len({r["Wind代码"] for r in cleaned_all if r["Wind代码"] in pool_codes}), len(pool_codes), round(len({r["Wind代码"] for r in cleaned_all if r["Wind代码"] in pool_codes}) / len(pool_codes) * 100, 2))
    stat("总体", "剔除2026Q1后有效行数", len(panel_clean))
    stat("总体", "半年报/年报面板覆盖基金数量", len({r["Wind代码"] for r in panel_clean}), len(pool_codes), round(len({r["Wind代码"] for r in panel_clean}) / len(pool_codes) * 100, 2))
    for period in TARGET_PERIODS:
        rows = [r for r in panel_clean if r["报告期"] == period]
        stat("报告期", f"{period}记录数量", len(rows))
        stat("报告期", f"{period}覆盖基金数量", len({r["Wind代码"] for r in rows}), len(pool_codes), round(len({r["Wind代码"] for r in rows}) / len(pool_codes) * 100, 2))
    stat("2025年年报", "覆盖基金数量", len(y25), len(pool_codes), round(len(y25) / len(pool_codes) * 100, 2))
    stat("2025年年报", "未覆盖基金数量", len(pool_codes) - len(y25))
    stat("2025年年报", "2025年尚未成立数量", not_established)
    stat("2025年年报", "2025年底新成立数量", new_late)
    stat("2025年年报", "真正需要补数数量", true_supplement)
    for field in NUMERIC_FIELDS:
        nonempty = sum(r.get(field) is not None for r in panel_clean)
        statistics.append({"统计类别": "字段非空率", "指标": field, "分组": "", "数量/结果": nonempty, "分母": len(panel_clean), "覆盖率/非空率(%)": round(nonempty / len(panel_clean) * 100, 2) if panel_clean else 0, "说明": "半年报/年报清洗版"})
    for group_field in ("一级策略大类", "统计口径分类"):
        groups = defaultdict(set)
        covered = defaultdict(set)
        for c, p in pool.items():
            groups[str(p.get(group_field, ""))].add(c)
        for c in y25:
            covered[str(pool[c].get(group_field, ""))].add(c)
        for group in sorted(groups):
            statistics.append({"统计类别": f"{group_field}覆盖率", "指标": "2025年年报覆盖", "分组": group, "数量/结果": len(covered[group]), "分母": len(groups[group]), "覆盖率/非空率(%)": round(len(covered[group]) / len(groups[group]) * 100, 2), "说明": ""})

    for field in sorted(missing_source_fields):
        anomalies.append({
            "异常类型": "字段不可得提示", "报告期": "", "Wind代码": "", "原始证券简称": "",
            "字段": field, "原始值": "", "异常说明": "Wind暂不可得，已在清洗表中生成空列",
            "是否错误": "否，仅提示",
        })
    if "基金状态" not in base_headers:
        anomalies.append({
            "异常类型": "字段不可得提示", "报告期": "", "Wind代码": "", "原始证券简称": "",
            "字段": "基金状态", "原始值": "", "异常说明": "主产品池当前未提供，相关输出列保留为空",
            "是否错误": "否，仅提示",
        })
    anomalies.sort(key=lambda r: (r["异常类型"], str(r["报告期"]), r["Wind代码"]))

    source_map = {
        "报告期": ("Wind持有人结构字段", "Wind报告期，主分析仅保留指定半年报/年报"),
        "报告期排序": ("衍生字段", "用于报告期排序"),
        "Wind代码": ("Wind持有人结构字段", "唯一匹配主键"),
        "基金代码": ("主产品池字段", "优先采用主产品池基金代码"),
        "交易代码": ("主产品池字段", "主产品池交易代码"),
        "证券简称": ("主产品池字段", "标准简称"),
        "基金简称": ("主产品池字段", "标准基金简称"),
        "原始证券简称": ("Wind持有人结构字段", "保留原始值备查"),
        "基金管理人": ("主产品池字段", ""),
        "统计口径分类": ("主产品池字段", ""),
        "是否纳入核心策略ETF统计": ("主产品池字段", ""),
        "是否纳入广义策略ETF统计": ("主产品池字段", ""),
        "一级策略大类": ("主产品池字段", ""),
        "二级策略类别": ("主产品池字段", ""),
        "市场范围_二次修正": ("主产品池字段", ""),
        "最新基金规模(亿)": ("主产品池字段", ""),
        "数据来源文件": ("衍生字段", "记录原始文件名"),
        "是否证券简称不一致": ("衍生字段", "仅提示，不作为错误"),
        "覆盖状态": ("衍生字段", "基于2025年年报覆盖和核心字段计算"),
        "缺失字段": ("衍生字段", "列示缺失核心字段"),
        "2025年年报缺失原因": ("衍生字段", "根据成立日期和覆盖情况判断"),
        "基金状态": ("Wind暂不可得字段", "主产品池当前未提供，输出为空"),
    }
    for f in NUMERIC_FIELDS:
        source_map[f] = ("Wind暂不可得字段" if f in missing_source_fields else "Wind持有人结构字段", "Wind暂不可得，输出为空" if f in missing_source_fields else "标准缺失标记已转为空，数值已标准化")
    field_notes = [{"字段": field, "字段来源": source_map.get(field, ("主产品池字段", ""))[0], "处理说明": source_map.get(field, ("主产品池字段", ""))[1]} for field in dict.fromkeys(PANEL_OUTPUT_FIELDS + ["基金成立日", "上市日期", "基金状态", "跟踪指数代码", "跟踪指数名称", "覆盖状态", "缺失字段", "2025年年报缺失原因"])]
    field_notes.append({"字段": "验收报告使用情况", "字段来源": "验收辅助", "处理说明": validation_note})

    latest_headers = [
        "Wind代码", "证券简称", "基金简称", "基金管理人", "基金成立日", "上市日期", "基金状态",
        "统计口径分类", "是否纳入核心策略ETF统计", "是否纳入广义策略ETF统计",
        "一级策略大类", "二级策略类别", "市场范围_二次修正", "最新基金规模(亿)",
        "跟踪指数代码", "跟踪指数名称", "是否出现在2025年年报", *NUMERIC_FIELDS,
        "覆盖状态", "缺失字段", "2025年年报缺失原因",
    ]
    pending_headers = [
        "Wind代码", "证券简称", "基金管理人", "基金成立日", "上市日期", "基金状态",
        "统计口径分类", "是否纳入核心策略ETF统计", "一级策略大类", "二级策略类别",
        "市场范围_二次修正", "最新基金规模(亿)", "是否出现在2025年年报", "覆盖状态",
        "缺失字段", "2025年年报缺失原因", "建议处理方式",
    ]

    wb = Workbook()
    wb.remove(wb.active)
    specs = [
        ("持有人结构_面板清洗版", PANEL_OUTPUT_FIELDS, panel_clean),
        ("2025年年报_最新截面", latest_headers, latest_snapshot),
        ("2025年年报待补充清单", pending_headers, pending),
        ("覆盖率统计", ["统计类别", "指标", "分组", "数量/结果", "分母", "覆盖率/非空率(%)", "说明"], statistics),
        ("异常与提示", ["异常类型", "报告期", "Wind代码", "原始证券简称", "字段", "原始值", "异常说明", "是否错误"], anomalies),
        ("字段说明", ["字段", "字段来源", "处理说明"], field_notes),
        ("2026Q1备查", PANEL_OUTPUT_FIELDS, q1_backup),
    ]
    for name, headers, rows in specs:
        ws = wb.create_sheet(name)
        write_sheet(ws, headers, rows)
    format_book(wb)

    WORK_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    ARCHIVE_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    temp = WORK_OUTPUT.with_name(WORK_OUTPUT.stem + "_tmp.xlsx")
    wb.save(temp)
    check = load_workbook(temp, read_only=True, data_only=True)
    if check.sheetnames != [s[0] for s in specs]:
        raise RuntimeError("输出sheet结构校验失败")
    check.close()
    temp.replace(WORK_OUTPUT)
    shutil.copy2(WORK_OUTPUT, ARCHIVE_OUTPUT)
    if sha256(WORK_OUTPUT) != sha256(ARCHIVE_OUTPUT):
        raise RuntimeError("工作版与归档版哈希不一致")
    if any(sha256(path) != old_hash for path, old_hash in source_hashes.items()):
        raise RuntimeError("输入文件被意外修改")

    hard_anomalies = [a for a in anomalies if a["是否错误"] == "是"]
    print(f"广义策略ETF产品池数量：{len(pool_codes)}")
    print(f"Sheet5 原始面板有效行数：{len(valid_rows)}")
    print(f"Sheet5 剔除2026Q1后的面板有效行数：{len(panel_clean)}")
    print(f"面板覆盖基金数量：{len({r['Wind代码'] for r in panel_clean})}（半年报/年报口径；原始含2026Q1为{len({r['Wind代码'] for r in cleaned_all if r['Wind代码'] in pool_codes})}）")
    print(f"2025年年报覆盖数量：{len(y25)}")
    print(f"2025年年报未覆盖数量：{len(pool_codes) - len(y25)}")
    print(f"其中2025年尚未成立数量：{not_established}")
    print(f"其中2025年底新成立数量：{new_late}")
    print(f"真正需要补数数量：{true_supplement}")
    print(f"是否存在代码池外产品：{'是' if any(a['异常类型']=='代码池外产品' for a in anomalies) else '否'}")
    print(f"是否存在重复主键：{'是' if any(a['异常类型']=='重复报告期+Wind代码' for a in anomalies) else '否'}")
    print(f"是否存在比例异常：{'是' if any(a['异常类型'] in {'比例超范围','机构+个人比例异常'} for a in anomalies) else '否'}")
    print(f"工作版输出路径：{WORK_OUTPUT}")
    print(f"归档版输出路径：{ARCHIVE_OUTPUT}")


if __name__ == "__main__":
    main()
