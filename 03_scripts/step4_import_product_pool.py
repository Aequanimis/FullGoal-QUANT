"""Step 4：保存附件中的 DATA_BLOCK，生成 CSV，并写入产品总表。"""

from __future__ import annotations

import csv
import io
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
REQUEST_FILE = Path(
    r"C:\Users\lvdon\.codex\attachments\ced4726f-5c2c-4a26-9f26-58123b35650b\pasted-text.txt"
)
MARKDOWN_PATH = (
    PROJECT_ROOT
    / "00_raw_data"
    / "fund_company_pages"
    / "公开搜索_策略ETF产品清单_20260622.md"
)
CSV_PATH = (
    PROJECT_ROOT
    / "01_processed_data"
    / "product_pool"
    / "境内策略ETF产品池_公开搜索初版.csv"
)
LOG_PATH = (
    PROJECT_ROOT
    / "01_processed_data"
    / "product_pool"
    / "step4_product_pool_log.md"
)
EXCEL_PATH = (
    PROJECT_ROOT
    / "02_outputs"
    / "excel"
    / "境内策略ETF产品梳理_初版.xlsx"
)

EXPECTED_HEADERS = [
    "策略大类",
    "策略细分",
    "基金代码",
    "产品名称",
    "基金公司",
    "上市交易所",
    "跟踪指数或策略线索",
    "纳入级别",
    "待校验事项",
    "信息来源备注",
]

EXCEL_FIELD_MAP = {
    "策略大类": "策略大类",
    "策略细分": "策略细分",
    "基金代码": "基金代码",
    "产品名称": "产品名称",
    "基金公司": "基金公司",
    "上市交易所": "上市交易所",
    "跟踪指数或策略线索": "跟踪指数",
    "信息来源备注": "信息来源",
}


def extract_data_block(request_text: str) -> str:
    start_marker = "DATA_BLOCK 如下："
    end_marker = "完成后请告诉我："
    if start_marker not in request_text or end_marker not in request_text:
        raise ValueError("附件中未找到完整 DATA_BLOCK 边界")
    return request_text.split(start_marker, 1)[1].split(end_marker, 1)[0].strip("\r\n")


def parse_rows(data_block: str) -> tuple[list[str], list[list[str]], list[tuple[int, int]]]:
    parsed = list(csv.reader(io.StringIO(data_block)))
    if not parsed or parsed[0] != EXPECTED_HEADERS:
        raise ValueError("DATA_BLOCK 表头与要求不一致")

    rows: list[list[str]] = []
    malformed: list[tuple[int, int]] = []
    for source_line, row in enumerate(parsed[1:], start=2):
        if len(row) != len(EXPECTED_HEADERS):
            malformed.append((source_line, len(row)))
        if len(row) < len(EXPECTED_HEADERS):
            row = row + [""] * (len(EXPECTED_HEADERS) - len(row))
        elif len(row) > len(EXPECTED_HEADERS):
            raise ValueError(f"DATA_BLOCK 第 {source_line} 行超过 10 个字段，无法无损写入")
        rows.append(row)
    return parsed[0], rows, malformed


def display_width(value: object) -> int:
    text = "" if value is None else str(value)
    return sum(
        2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1
        for char in text
    )


def save_markdown(data_block: str) -> None:
    MARKDOWN_PATH.parent.mkdir(parents=True, exist_ok=True)
    content = "# 公开搜索_策略ETF产品清单_20260622\n\n" + data_block + "\n"
    MARKDOWN_PATH.write_text(content, encoding="utf-8")


def save_csv(headers: list[str], rows: list[list[str]]) -> None:
    CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.writer(csv_file, lineterminator="\n")
        writer.writerow(headers)
        writer.writerows(rows)


def build_excel_rows(rows: list[list[str]]) -> list[dict[str, str]]:
    return [dict(zip(EXPECTED_HEADERS, row)) for row in rows]


def write_excel(rows: list[list[str]]) -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel 模板不存在：{EXCEL_PATH}")

    workbook = load_workbook(EXCEL_PATH)
    if "产品总表" not in workbook.sheetnames:
        raise ValueError("Excel 中不存在 Sheet【产品总表】")

    worksheet = workbook["产品总表"]
    excel_headers = [cell.value for cell in worksheet[1]]
    header_columns = {header: index + 1 for index, header in enumerate(excel_headers)}

    required_excel_headers = set(EXCEL_FIELD_MAP.values()) | {"备注"}
    missing_headers = required_excel_headers - set(header_columns)
    if missing_headers:
        raise ValueError(f"Excel 产品总表缺少字段：{sorted(missing_headers)}")

    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)

    records = build_excel_rows(rows)
    for excel_row, record in enumerate(records, start=2):
        for csv_field, excel_field in EXCEL_FIELD_MAP.items():
            worksheet.cell(excel_row, header_columns[excel_field], record[csv_field])

        note_parts = [f"纳入级别：{record['纳入级别']}"]
        if record["待校验事项"]:
            note_parts.append(f"待校验事项：{record['待校验事项']}")
        worksheet.cell(excel_row, header_columns["备注"], "\n".join(note_parts))

    last_row = len(rows) + 1
    last_column = len(excel_headers)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(last_column)}{last_row}"

    for row in worksheet.iter_rows(min_row=2, max_row=last_row, max_col=last_column):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_index, header in enumerate(excel_headers, start=1):
        values = [header]
        values.extend(
            worksheet.cell(row_index, column_index).value
            for row_index in range(2, last_row + 1)
        )
        measured = max(display_width(value) for value in values) + 3
        cap = 42 if header in {"跟踪指数", "代表性理由", "备注", "信息来源"} else 24
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(measured, 12), cap
        )

    workbook.save(EXCEL_PATH)


def validate(headers: list[str], rows: list[list[str]]) -> None:
    if MARKDOWN_PATH.read_text(encoding="utf-8").splitlines()[0] != (
        "# 公开搜索_策略ETF产品清单_20260622"
    ):
        raise ValueError("Markdown 标题校验失败")
    if CSV_PATH.read_bytes()[:3] != b"\xef\xbb\xbf":
        raise ValueError("CSV 不是 UTF-8 with BOM 编码")
    with CSV_PATH.open("r", encoding="utf-8-sig", newline="") as csv_file:
        csv_rows = list(csv.reader(csv_file))
    if csv_rows != [headers, *rows]:
        raise ValueError("CSV 内容回读校验失败")

    workbook = load_workbook(EXCEL_PATH, read_only=False)
    worksheet = workbook["产品总表"]
    if worksheet.max_row != len(rows) + 1:
        raise ValueError("Excel 写入行数校验失败")
    if worksheet.freeze_panes != "A2" or not worksheet.auto_filter.ref:
        raise ValueError("Excel 冻结窗格或筛选校验失败")


def make_log(rows: list[list[str]], malformed: list[tuple[int, int]]) -> None:
    records = build_excel_rows(rows)
    mainline_count = sum(record["纳入级别"].startswith("主线") for record in records)
    supplemental_count = sum(record["纳入级别"] == "补充观察" for record in records)
    pending_count = sum(
        any(value.strip() == "待校验" for value in record.values()) for record in records
    )

    malformed_text = "无"
    if malformed:
        malformed_text = "；".join(
            f"DATA_BLOCK 第 {line_number} 行仅有 {field_count} 个字段"
            for line_number, field_count in malformed
        )

    log = f"""# Step 4 产品池处理日志

1. Markdown 文件是否成功创建：是
2. CSV 文件是否成功创建：是（UTF-8 with BOM）
3. Excel 是否成功写入：是
4. 总共写入多少只产品：{len(rows)}
5. 主线产品多少只：{mainline_count}（包含“主线”和“主线重点”）
6. 补充观察产品多少只：{supplemental_count}（按“纳入级别=补充观察”统计）
7. 有多少只产品标记为“待校验”：{pending_count}（按任一字段值严格等于“待校验”统计）

## 数据完整性备注

- 未删除任何 DATA_BLOCK 数据行。
- 格式异常原始行：{malformed_text}；该行已原样保留，并在 CSV 中以空字段补齐至 10 列。
"""
    LOG_PATH.write_text(log, encoding="utf-8")


def main() -> None:
    request_text = REQUEST_FILE.read_text(encoding="utf-8-sig")
    data_block = extract_data_block(request_text)
    headers, rows, malformed = parse_rows(data_block)

    save_markdown(data_block)
    save_csv(headers, rows)
    write_excel(rows)
    validate(headers, rows)
    make_log(rows, malformed)

    records = build_excel_rows(rows)
    print(f"total={len(rows)}")
    print(f"mainline={sum(r['纳入级别'].startswith('主线') for r in records)}")
    print(f"supplemental={sum(r['纳入级别'] == '补充观察' for r in records)}")
    print(
        "pending="
        + str(sum(any(v.strip() == "待校验" for v in r.values()) for r in records))
    )
    print(f"malformed={malformed}")


if __name__ == "__main__":
    main()
