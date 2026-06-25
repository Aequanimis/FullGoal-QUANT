from __future__ import annotations

import hashlib
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE = Path(
    r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池"
    r"\广义策略ETF_收益风险表现表_合并验收清洗版.xlsx"
)
ARCHIVE_DIR = Path(
    r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据"
    r"\阶段性归档_境内广义策略ETF数据底座"
)
TARGET_NAME = "05_广义策略ETF_收益风险表现_合并验收清洗版.xlsx"
TARGET = ARCHIVE_DIR / TARGET_NAME
INDEX_FILE = ARCHIVE_DIR / "阶段性归档索引表.xlsx"
README_FILE = ARCHIVE_DIR / "阶段性归档说明.md"

PURPOSE = (
    "用于分析广义策略ETF的收益率、年化波动率、最大回撤、Sharpe、"
    "跟踪误差、信息比率和相对基准表现"
)
DATA_SCOPE = "223只广义策略ETF，其中核心策略指数ETF 168只，指数增强/多因子ETF 55只"
COVERAGE = "223只全部覆盖，缺失代码0，多余代码0，重复代码0"
NOTE = "本文件为收益风险表现表正式清洗版，可用于后续描述性统计和收益风险分析"
ARCHIVE_PARAGRAPH = (
    "05_广义策略ETF_收益风险表现_合并验收清洗版.xlsx 已完成归档。"
    "本表覆盖223只广义策略ETF，其中核心策略指数ETF 168只、指数增强/多因子ETF 55只。"
    "数据已完成分批合并、字段标准化、缺失代码检查、重复代码检查和产品池匹配验收。"
    "验收结果显示：223只全部覆盖，缺失代码0，多余代码0，重复代码0，产品名称匹配一致。"
    "该表可用于后续分析不同策略类型的收益率、年化波动率、最大回撤、Sharpe、跟踪误差、"
    "信息比率和超越基准收益表现。"
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def header_map(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None and str(cell.value).strip()
    }


def ensure_headers(ws, names: list[str]) -> dict[str, int]:
    headers = header_map(ws)
    style_source = ws.cell(1, max(1, ws.max_column))
    for name in names:
        if name not in headers:
            col = ws.max_column + 1
            cell = ws.cell(1, col, name)
            if style_source.has_style:
                cell._style = copy(style_source._style)
            else:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9E1F2")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            headers[name] = col
    return headers


def find_row(ws, column: int, terms: tuple[str, ...]) -> int | None:
    for row in range(2, ws.max_row + 1):
        text = str(ws.cell(row, column).value or "")
        if any(term in text for term in terms):
            return row
    return None


def auto_width(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        values = [str(ws.cell(row, col_idx).value or "") for row in range(1, ws.max_row + 1)]
        max_len = min(max((len(value) for value in values), default=0) + 2, 80)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, max_len)


def update_index(now_text: str, target_size: int, target_hash: str, backup: Path | None) -> None:
    if not INDEX_FILE.exists():
        raise FileNotFoundError(f"归档索引表不存在：{INDEX_FILE}")

    wb = load_workbook(INDEX_FILE)
    ws = wb["归档文件索引"]
    requested = ["文件类型", "文件用途", "数据范围", "数据状态", "覆盖情况", "归档时间"]
    headers = ensure_headers(ws, requested)

    module_col = headers.get("模块名称")
    filename_col = headers.get("归档后文件名")
    row = None
    if module_col:
        row = find_row(ws, module_col, ("收益风险",))
    if row is None and filename_col:
        row = find_row(ws, filename_col, ("收益风险", TARGET_NAME))
    if row is None:
        row = ws.max_row + 1

    values = {
        "序号": "05",
        "模块名称": "收益风险表现",
        "归档后文件名": TARGET_NAME,
        "归档后文件路径": str(TARGET),
        "原始文件路径": str(SOURCE),
        "文件定位": PURPOSE,
        "正式分析使用sheet": "收益风险_合并清洗版",
        "是否核心产出": "是",
        "注意事项": NOTE,
        "源文件是否存在": "是",
        "是否成功复制": "是",
        "归档文件大小_字节": target_size,
        "SHA256": target_hash,
        "校验备注": f"旧版已备份：{backup}" if backup else "源文件与归档文件SHA256一致",
        "文件类型": "Excel",
        "文件用途": PURPOSE,
        "数据范围": DATA_SCOPE,
        "数据状态": "已完成合并、清洗、验收",
        "覆盖情况": COVERAGE,
        "归档时间": now_text,
    }
    for name, value in values.items():
        if name in headers:
            ws.cell(row, headers[name], value)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    auto_width(ws)

    if "当前数据完成情况" in wb.sheetnames:
        status_ws = wb["当前数据完成情况"]
        status_headers = header_map(status_ws)
        module_idx = status_headers.get("数据模块")
        status_row = find_row(status_ws, module_idx, ("收益风险表现",)) if module_idx else None
        if status_row is None:
            status_row = status_ws.max_row + 1
        status_values = {
            "数据模块": "收益风险表现",
            "当前状态": "已完成",
            "对应文件": TARGET_NAME,
            "主要用途": PURPOSE,
            "后续是否还需补充": "需定期更新",
            "备注": COVERAGE,
        }
        for name, value in status_values.items():
            if name in status_headers:
                status_ws.cell(status_row, status_headers[name], value)
        status_ws.freeze_panes = "A2"
        status_ws.auto_filter.ref = (
            f"A1:{get_column_letter(status_ws.max_column)}{status_ws.max_row}"
        )
        auto_width(status_ws)

    if "后续数据收集清单" in wb.sheetnames:
        pending_ws = wb["后续数据收集清单"]
        pending_headers = header_map(pending_ws)
        module_idx = pending_headers.get("后续数据模块")
        if module_idx:
            for row_idx in range(pending_ws.max_row, 1, -1):
                value = str(pending_ws.cell(row_idx, module_idx).value or "")
                if "收益风险表现" in value:
                    pending_ws.delete_rows(row_idx)
        pending_ws.freeze_panes = "A2"
        pending_ws.auto_filter.ref = (
            f"A1:{get_column_letter(pending_ws.max_column)}{pending_ws.max_row}"
        )
        auto_width(pending_ws)

    for sheet in wb.worksheets:
        sheet.freeze_panes = sheet.freeze_panes or "A2"
        for cell in sheet[1]:
            cell.font = copy(cell.font)
            cell.font = Font(
                name=cell.font.name,
                size=cell.font.size,
                bold=True,
                italic=cell.font.italic,
                color=cell.font.color,
            )

    wb.save(INDEX_FILE)


def update_markdown() -> None:
    if not README_FILE.exists():
        raise FileNotFoundError(f"归档说明不存在：{README_FILE}")

    text = README_FILE.read_text(encoding="utf-8-sig")
    text = text.replace("## 四个主文件说明", "## 五个主文件说明")
    text = text.replace("四个主文件说明", "五个主文件说明")

    table_row = (
        f"| 5 | 收益风险表现 | `{TARGET}` | `{SOURCE}` | {PURPOSE} | "
        f"收益风险_合并清洗版 | {NOTE} |"
    )
    if table_row not in text:
        lines = text.splitlines()
        table_indices = [i for i, line in enumerate(lines) if line.startswith("| 4 |")]
        if table_indices:
            lines.insert(table_indices[-1] + 1, table_row)
            text = "\n".join(lines)

    analysis_bullet = (
        f'- 收益风险分析以“{TARGET_NAME}”中的“收益风险_合并清洗版”为准；'
    )
    if analysis_bullet not in text:
        marker = "- Wind口径差异解释"
        position = text.find(marker)
        if position >= 0:
            line_end = text.find("\n", position)
            line_end = len(text) if line_end < 0 else line_end
            text = text[:line_end] + "\n" + analysis_bullet + text[line_end:]

    completed_bullet = "- 收益风险表现数据：已完成，覆盖223只广义策略ETF。"
    if completed_bullet not in text:
        heading = "## 已完成数据模块"
        position = text.find(heading)
        if position >= 0:
            next_heading = text.find("\n## ", position + len(heading))
            insert_at = len(text) if next_heading < 0 else next_heading
            text = text[:insert_at].rstrip() + "\n" + completed_bullet + "\n\n" + text[insert_at:].lstrip()

    text = text.replace("- 收益风险表现表；\n", "")
    text = text.replace("- 收益风险表现表\n", "")

    section_heading = "## 收益风险表现数据归档"
    if section_heading in text:
        start = text.index(section_heading)
        next_heading = text.find("\n## ", start + len(section_heading))
        if next_heading < 0:
            text = text[:start].rstrip()
        else:
            text = text[:start].rstrip() + "\n\n" + text[next_heading + 1 :].lstrip()
    text = text.rstrip() + f"\n\n{section_heading}\n\n{ARCHIVE_PARAGRAPH}\n"
    README_FILE.write_text(text, encoding="utf-8-sig")


def main() -> None:
    now = datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    now_text = now.strftime("%Y-%m-%d %H:%M:%S")

    source_found = SOURCE.exists()
    if not source_found:
        print("是否找到源文件：否")
        raise FileNotFoundError(SOURCE)

    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    source_hash_before = sha256(SOURCE)

    backup = None
    if TARGET.exists():
        backup = TARGET.with_name(f"{TARGET.stem}_旧版备份_{timestamp}{TARGET.suffix}")
        shutil.copy2(TARGET, backup)

    shutil.copy2(SOURCE, TARGET)
    copied = TARGET.exists() and TARGET.stat().st_size > 0
    source_hash_after = sha256(SOURCE)
    target_hash = sha256(TARGET)
    if source_hash_before != source_hash_after:
        raise RuntimeError("源文件在归档过程中发生变化")
    if source_hash_after != target_hash:
        raise RuntimeError("复制后的文件哈希与源文件不一致")

    update_index(now_text, TARGET.stat().st_size, target_hash, backup)
    update_markdown()

    print(f"是否找到源文件：{'是' if source_found else '否'}")
    print(f"是否成功复制：{'是' if copied else '否'}")
    print(f"是否备份旧版文件：{'是，' + str(backup) if backup else '否（目标文件原先不存在）'}")
    print(f"是否成功更新阶段性归档索引表：{'是' if INDEX_FILE.exists() else '否'}")
    print(f"是否成功更新阶段性归档说明：{'是' if README_FILE.exists() else '否'}")
    print(f"最新归档文件路径：{TARGET}")
    print(f"文件大小：{TARGET.stat().st_size} 字节")
    print("处理完成：仅归档并更新了收益风险表现文件。")


if __name__ == "__main__":
    main()
