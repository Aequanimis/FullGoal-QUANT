from __future__ import annotations

import hashlib
import math
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_FILE = Path(
    r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据"
    r"\全市场ETF基础信息_策略ETF池二次修正版.xlsx"
)
PANEL_FILE = Path(
    r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\sheet5"
    r"\广义策略ETF持有人结构面板数据.xlsx"
)
LATEST_FILE = Path(
    r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\sheet5"
    r"\广义策略ETF持有人结构数据（最新报告期Q1 FY2026）.xlsx"
)
MISSING_FILE = Path(
    r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\sheet5"
    r"\缺失基金清单.xlsx"
)
OUTPUT_FILE = Path(
    r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\sheet5"
    r"\Sheet5_持有人结构数据_验收报告.xlsx"
)

POOL_FIELDS = [
    "Wind代码",
    "证券简称",
    "基金简称",
    "基金管理人",
    "一级策略大类",
    "二级策略类别",
    "统计口径分类",
    "是否纳入核心策略ETF统计",
    "是否纳入广义策略ETF统计",
    "市场范围_二次修正",
    "最新基金规模(亿)",
]

PANEL_REQUIRED_FIELDS = [
    "报告期",
    "Wind代码",
    "基金代码",
    "证券简称",
    "机构投资者持有份额",
    "机构投资者持有比例",
    "个人投资者持有份额",
    "个人投资者持有比例",
    "前十大持有人持有比例",
    "基金管理人自持比例",
    "员工持有份额",
    "员工持有比例",
]

NUMERIC_FIELDS = [
    "机构投资者持有份额",
    "机构投资者持有比例",
    "个人投资者持有份额",
    "个人投资者持有比例",
    "前十大持有人持有比例",
    "基金管理人自持比例",
    "员工持有份额",
    "员工持有比例",
    "联接基金持有份额",
    "联接基金持有比例",
    "基金管理人自持份额",
]

RATIO_FIELDS = {
    "机构投资者持有比例",
    "个人投资者持有比例",
    "前十大持有人持有比例",
    "基金管理人自持比例",
    "员工持有比例",
    "联接基金持有比例",
}
SHARE_FIELDS = set(NUMERIC_FIELDS) - RATIO_FIELDS
EMPTY_TOKENS = {"", "-", "--", "N/A", "NA", "NAN", "NONE", "NULL", "WIND暂不可得"}
EXPECTED_PERIODS = {
    "2023年年报",
    "2024年中报",
    "2024年年报",
    "2025年中报",
    "2025年年报",
    "2026年一季报",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[\s\u3000\u200b\ufeff]+", "", str(value)).strip()


def normalize_code(value: Any) -> str:
    return clean_text(value).upper()


def normalize_name(value: Any) -> str:
    return clean_text(value)


def is_missing_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return clean_text(value).upper() in EMPTY_TOKENS


def to_number(value: Any) -> tuple[float | None, str]:
    if is_missing_value(value):
        return None, "缺失"
    if isinstance(value, bool):
        return float(value), "可转数值"
    if isinstance(value, (int, float)):
        return float(value), "可转数值"
    text = clean_text(value).replace(",", "").replace("，", "").replace("%", "")
    try:
        return float(text), "可转数值"
    except (TypeError, ValueError):
        return None, "不可转数值"


def read_records(path: Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
    iterator = ws.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(iterator)]
    records = [dict(zip(headers, row)) for row in iterator]
    wb.close()
    return headers, records


def is_non_product_row(record: dict[str, Any]) -> bool:
    code = clean_text(record.get("Wind代码"))
    joined = " ".join(clean_text(value) for value in record.values())
    return not code or "数据来源" in code or "数据来源：Wind" in joined or "数据来源:Wind" in joined


def add_anomaly(
    anomalies: list[dict[str, Any]],
    anomaly_type: str,
    source: str,
    record: dict[str, Any] | None = None,
    field: str = "",
    original: Any = "",
    expected: Any = "",
    description: str = "",
    blocking: str = "是",
) -> None:
    record = record or {}
    anomalies.append(
        {
            "异常类型": anomaly_type,
            "来源文件": source,
            "报告期": record.get("报告期", ""),
            "Wind代码": normalize_code(record.get("Wind代码")),
            "证券简称": record.get("证券简称", ""),
            "字段": field,
            "原始值": original,
            "相关值或期望值": expected,
            "异常说明": description,
            "是否阻断通过": blocking,
        }
    )


def analyse_numeric_dataset(
    dataset_name: str,
    records: list[dict[str, Any]],
    headers: list[str],
    anomalies: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for field in NUMERIC_FIELDS:
        present = field in headers
        nonempty = convertible = nonconvertible = abnormal = 0
        if present:
            for record in records:
                raw = record.get(field)
                if not is_missing_value(raw):
                    nonempty += 1
                number, state = to_number(raw)
                if state == "可转数值":
                    convertible += 1
                    bad = False
                    if field in RATIO_FIELDS and not (0 <= number <= 100):
                        bad = True
                        reason = "比例字段应在0至100之间"
                    elif field in SHARE_FIELDS and number < 0:
                        bad = True
                        reason = "份额字段应大于等于0"
                    if bad:
                        abnormal += 1
                        add_anomaly(
                            anomalies,
                            "数值超范围",
                            dataset_name,
                            record,
                            field,
                            raw,
                            "比例0-100；份额>=0",
                            reason,
                        )
                elif state == "不可转数值":
                    nonconvertible += 1
                    abnormal += 1
                    add_anomaly(
                        anomalies,
                        "数值不可转换",
                        dataset_name,
                        record,
                        field,
                        raw,
                        "数值或标准缺失标记",
                        "非缺失值无法转换为数值",
                    )
        results.append(
            {
                "数据集": dataset_name,
                "字段": field,
                "字段是否存在": "是" if present else "否",
                "总记录数": len(records) if present else 0,
                "非空数量": nonempty,
                "缺失数量": (len(records) - nonempty) if present else 0,
                "可转数值数量": convertible,
                "不可转数值数量": nonconvertible,
                "异常数量": abnormal,
                "说明": "" if present else "该源文件未提供此字段",
            }
        )

    for record in records:
        inst, inst_state = to_number(record.get("机构投资者持有比例"))
        indiv, indiv_state = to_number(record.get("个人投资者持有比例"))
        if inst_state == "可转数值" and indiv_state == "可转数值":
            total = inst + indiv
            if not 99.9 <= total <= 100.1:
                add_anomaly(
                    anomalies,
                    "机构与个人比例合计异常",
                    dataset_name,
                    record,
                    "机构投资者持有比例+个人投资者持有比例",
                    total,
                    "99.9至100.1",
                    f"机构比例{inst:g}+个人比例{indiv:g}={total:g}",
                )
    return results


def write_sheet(ws, headers: list[str], records: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for record in records:
        ws.append([record.get(header, "") for header in headers])


def format_workbook(wb: Workbook) -> None:
    navy = "1F4E78"
    light_blue = "D9EAF7"
    pale_red = "FCE4D6"
    pale_green = "E2F0D9"
    pale_yellow = "FFF2CC"
    white = "FFFFFF"
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 1)}"
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(name="微软雅黑", size=10, bold=True, color=white)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="微软雅黑", size=9)
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                cell.border = border
                if isinstance(cell.value, (datetime, date)):
                    cell.number_format = "yyyy-mm-dd"

        for col_idx in range(1, ws.max_column + 1):
            values = [str(ws.cell(row_idx, col_idx).value or "") for row_idx in range(1, min(ws.max_row, 1000) + 1)]
            max_len = max((len(value) for value in values), default=8)
            width = min(max(max_len * 1.15 + 2, 10), 38)
            header = str(ws.cell(1, col_idx).value or "")
            if "路径" in header or "说明" in header or "原因" in header:
                width = min(max(width, 28), 48)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        if ws.title == "验收总览":
            ws.column_dimensions["A"].width = 34
            ws.column_dimensions["B"].width = 24
            ws.column_dimensions["C"].width = 32
            ws.column_dimensions["D"].width = 16
            ws.column_dimensions["E"].width = 58
            for row_idx in range(2, ws.max_row + 1):
                status = str(ws.cell(row_idx, 4).value or "")
                fill = pale_green if status in {"通过", "符合", "无"} else pale_red if status in {"不通过", "异常"} else pale_yellow
                ws.cell(row_idx, 4).fill = PatternFill("solid", fgColor=fill)
                ws.cell(row_idx, 4).font = Font(name="微软雅黑", size=9, bold=True)
                ws.cell(row_idx, 5).alignment = Alignment(vertical="top", wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = str(ws.cell(1, cell.column).value or "")
                if any(token in header for token in ("比例", "规模", "份额", "数量", "行数")) and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.0000"

        if ws.title == "异常记录明细" and ws.max_row > 1:
            block_col = next(
                (cell.column for cell in ws[1] if cell.value == "是否阻断通过"),
                None,
            )
            if block_col:
                letter = get_column_letter(block_col)
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    FormulaRule(
                        formula=[f'${letter}2="是"'],
                        fill=PatternFill("solid", fgColor=pale_red),
                    ),
                )
        if ws.title == "证券简称不一致检查" and ws.max_row > 1:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, ws.max_column).fill = PatternFill("solid", fgColor=pale_yellow)


def main() -> None:
    inputs = [BASE_FILE, PANEL_FILE, LATEST_FILE, MISSING_FILE]
    missing_inputs = [path for path in inputs if not path.exists()]
    if missing_inputs:
        raise FileNotFoundError("缺少输入文件：" + "；".join(map(str, missing_inputs)))

    source_hashes = {path: sha256(path) for path in inputs}
    anomalies: list[dict[str, Any]] = []

    base_headers, base_records = read_records(BASE_FILE, "策略ETF_最终统计池")
    missing_base_fields = [field for field in POOL_FIELDS if field not in base_headers]
    if missing_base_fields:
        raise ValueError(f"主产品池缺少字段：{missing_base_fields}")
    pool_records = [
        record
        for record in base_records
        if clean_text(record.get("是否纳入广义策略ETF统计")) == "是"
    ]
    pool: dict[str, dict[str, Any]] = {}
    for record in pool_records:
        code = normalize_code(record.get("Wind代码"))
        if not code:
            add_anomaly(
                anomalies,
                "主产品池Wind代码为空",
                BASE_FILE.name,
                record,
                "Wind代码",
                record.get("Wind代码"),
                ".SH或.SZ代码",
                "广义策略ETF产品池存在空代码",
            )
            continue
        if code in pool:
            add_anomaly(
                anomalies,
                "主产品池重复Wind代码",
                BASE_FILE.name,
                record,
                "Wind代码",
                code,
                "唯一",
                "主产品池Wind代码重复",
            )
        pool[code] = {field: record.get(field) for field in POOL_FIELDS}
    pool_codes = set(pool)

    panel_headers, panel_raw = read_records(PANEL_FILE)
    latest_headers, latest_raw = read_records(LATEST_FILE)
    missing_headers, missing_raw = read_records(MISSING_FILE)

    datasets = [
        (PANEL_FILE.name, panel_raw),
        (LATEST_FILE.name, latest_raw),
        (MISSING_FILE.name, missing_raw),
    ]
    non_product_rows: list[tuple[str, dict[str, Any]]] = []
    cleaned: dict[str, list[dict[str, Any]]] = {}
    for source_name, records in datasets:
        valid = []
        for record in records:
            if is_non_product_row(record):
                non_product_rows.append((source_name, record))
            else:
                record = dict(record)
                record["Wind代码"] = normalize_code(record.get("Wind代码"))
                valid.append(record)
        cleaned[source_name] = valid

    panel_records = cleaned[PANEL_FILE.name]
    latest_records = cleaned[LATEST_FILE.name]
    original_missing_records = cleaned[MISSING_FILE.name]

    panel_missing_fields = [field for field in PANEL_REQUIRED_FIELDS if field not in panel_headers]
    for field in panel_missing_fields:
        add_anomaly(
            anomalies,
            "面板缺少必需字段",
            PANEL_FILE.name,
            field=field,
            expected="字段必须存在",
            description="面板数据结构不完整",
        )

    outside_records: list[tuple[str, dict[str, Any]]] = []
    for source_name, records in [
        (PANEL_FILE.name, panel_records),
        (LATEST_FILE.name, latest_records),
        (MISSING_FILE.name, original_missing_records),
    ]:
        for record in records:
            if record["Wind代码"] not in pool_codes:
                outside_records.append((source_name, record))
                add_anomaly(
                    anomalies,
                    "代码池外产品",
                    source_name,
                    record,
                    "Wind代码",
                    record["Wind代码"],
                    "223只广义策略ETF代码池",
                    "Wind代码不属于主产品池",
                )

    latest_code_counts = Counter(record["Wind代码"] for record in latest_records)
    for code, count in latest_code_counts.items():
        if count > 1:
            for record in [r for r in latest_records if r["Wind代码"] == code]:
                add_anomaly(
                    anomalies,
                    "最新报告期重复Wind代码",
                    LATEST_FILE.name,
                    record,
                    "Wind代码",
                    code,
                    "每个代码一条记录",
                    f"重复{count}次",
                )

    missing_code_counts = Counter(record["Wind代码"] for record in original_missing_records)
    for code, count in missing_code_counts.items():
        if count > 1:
            for record in [r for r in original_missing_records if r["Wind代码"] == code]:
                add_anomaly(
                    anomalies,
                    "缺失清单重复Wind代码",
                    MISSING_FILE.name,
                    record,
                    "Wind代码",
                    code,
                    "每个代码一条记录",
                    f"重复{count}次",
                )

    panel_key_counts = Counter(
        (clean_text(record.get("报告期")), record["Wind代码"]) for record in panel_records
    )
    for key, count in panel_key_counts.items():
        if count > 1:
            for record in [
                r
                for r in panel_records
                if (clean_text(r.get("报告期")), r["Wind代码"]) == key
            ]:
                add_anomaly(
                    anomalies,
                    "面板重复键",
                    PANEL_FILE.name,
                    record,
                    "报告期+Wind代码",
                    f"{key[0]}|{key[1]}",
                    "唯一",
                    f"重复{count}次",
                )

    latest_empty_rows: list[dict[str, Any]] = []
    latest_data_records: list[dict[str, Any]] = []
    for record in latest_records:
        empty = all(
            is_missing_value(record.get(field))
            for field in (
                "机构投资者持有比例",
                "个人投资者持有比例",
                "前十大持有人持有比例",
            )
        )
        (latest_empty_rows if empty else latest_data_records).append(record)

    latest_data_codes = {record["Wind代码"] for record in latest_data_records}
    latest_empty_codes = {record["Wind代码"] for record in latest_empty_rows}
    original_missing_by_code = {record["Wind代码"]: record for record in original_missing_records}
    corrected_missing_codes = set(original_missing_by_code) | latest_empty_codes
    corrected_missing_codes -= latest_data_codes
    combined_coverage = latest_data_codes | corrected_missing_codes
    missing_from_combined = pool_codes - combined_coverage
    extra_in_combined = combined_coverage - pool_codes
    overlap_data_missing = latest_data_codes & corrected_missing_codes

    for code in sorted(missing_from_combined):
        add_anomaly(
            anomalies,
            "最新数据与缺失清单均未覆盖",
            "覆盖核对",
            {"Wind代码": code, "证券简称": pool[code].get("证券简称")},
            "Wind代码",
            code,
            "最新有数据或修正后缺失清单",
            "主产品池代码未被完整覆盖",
        )
    for code in sorted(overlap_data_missing):
        add_anomaly(
            anomalies,
            "最新有数据与缺失清单重叠",
            "覆盖核对",
            {"Wind代码": code, "证券简称": pool.get(code, {}).get("证券简称")},
            "Wind代码",
            code,
            "两者互斥",
            "同一代码同时被标记为有数据和缺失",
        )

    name_mismatches: list[dict[str, Any]] = []
    for source_name, records in [
        (LATEST_FILE.name, latest_records),
        (PANEL_FILE.name, panel_records),
    ]:
        for record in records:
            code = record["Wind代码"]
            if code not in pool:
                continue
            holder_name = normalize_name(record.get("证券简称"))
            pool_name = normalize_name(pool[code].get("证券简称"))
            if holder_name != pool_name:
                mismatch = {
                    "Wind代码": code,
                    "持有人数据证券简称": record.get("证券简称", ""),
                    "主产品池证券简称": pool[code].get("证券简称", ""),
                    "来源文件": source_name,
                    "报告期": record.get("报告期", ""),
                    "是否需要人工确认": "是",
                }
                name_mismatches.append(mismatch)
                add_anomaly(
                    anomalies,
                    "证券简称不一致",
                    source_name,
                    record,
                    "证券简称",
                    record.get("证券简称"),
                    pool[code].get("证券简称"),
                    "Wind代码一致但证券简称不一致，需确认是否存在字段或排序错位",
                    "否（需人工确认）",
                )

    numeric_stats = []
    numeric_stats.extend(
        analyse_numeric_dataset(LATEST_FILE.name, latest_records, latest_headers, anomalies)
    )
    numeric_stats.extend(
        analyse_numeric_dataset(PANEL_FILE.name, panel_records, panel_headers, anomalies)
    )

    panel_period_counts = Counter(clean_text(record.get("报告期")) for record in panel_records)
    missing_periods = EXPECTED_PERIODS - set(panel_period_counts)
    for period in sorted(missing_periods):
        add_anomaly(
            anomalies,
            "预期报告期缺失",
            PANEL_FILE.name,
            {"报告期": period},
            "报告期",
            period,
            "应存在",
            "面板未覆盖指定报告期",
        )

    panel_period_summary = []
    for period in sorted(panel_period_counts):
        period_records = [r for r in panel_records if clean_text(r.get("报告期")) == period]
        with_ratios = sum(
            1
            for r in period_records
            if not is_missing_value(r.get("机构投资者持有比例"))
            or not is_missing_value(r.get("个人投资者持有比例"))
        )
        with_shares = sum(
            1
            for r in period_records
            if not is_missing_value(r.get("机构投资者持有份额"))
            or not is_missing_value(r.get("个人投资者持有份额"))
        )
        panel_period_summary.append(
            {
                "报告期": period,
                "记录数量": len(period_records),
                "有机构/个人比例的记录数量": with_ratios,
                "有持有份额的记录数量": with_shares,
                "覆盖Wind代码数量": len({r["Wind代码"] for r in period_records}),
            }
        )

    corrected_missing_rows: list[dict[str, Any]] = []
    for code in sorted(corrected_missing_codes):
        source_record = original_missing_by_code.get(code, {})
        pool_record = pool.get(code, {})
        corrected_missing_rows.append(
            {
                "Wind代码": code,
                "主产品池证券简称": pool_record.get("证券简称", ""),
                "缺失清单证券简称": source_record.get("证券简称", ""),
                "是否来自原缺失清单": "是" if code in original_missing_by_code else "否",
                "是否来自最新报告期空数据行": "是" if code in latest_empty_codes else "否",
                "基金成立日": source_record.get("基金成立日", ""),
                "上市日期": source_record.get("上市日期", ""),
                "基金状态": source_record.get("基金状态", ""),
                "最新基金规模(亿)": pool_record.get("最新基金规模(亿)", source_record.get("最新基金规模(亿)", "")),
                "缺失原因或备注": source_record.get("缺失原因或备注", "")
                or "最新报告期关键持有人比例字段均为空",
                "一级策略大类": pool_record.get("一级策略大类", ""),
                "二级策略类别": pool_record.get("二级策略类别", ""),
            }
        )

    latest_empty_output: list[dict[str, Any]] = []
    for record in latest_empty_rows:
        code = record["Wind代码"]
        output = dict(record)
        output.update(
            {
                "主产品池证券简称": pool.get(code, {}).get("证券简称", ""),
                "处理建议": "并入修正后缺失基金清单",
            }
        )
        latest_empty_output.append(output)

    panel_codes = {record["Wind代码"] for record in panel_records}
    coverage_rows = []
    for code in sorted(pool_codes):
        in_data = code in latest_data_codes
        in_empty = code in latest_empty_codes
        in_original_missing = code in original_missing_by_code
        if in_data:
            status = "最新报告期有数据"
        elif code in corrected_missing_codes:
            status = "修正后缺失清单覆盖"
        else:
            status = "未覆盖"
        coverage_rows.append(
            {
                "Wind代码": code,
                "主产品池证券简称": pool[code].get("证券简称", ""),
                "是否在最新报告期有数据": "是" if in_data else "否",
                "是否最新报告期空数据": "是" if in_empty else "否",
                "是否在缺失基金清单": "是" if in_original_missing else "否",
                "是否在面板数据中出现": "是" if code in panel_codes else "否",
                "覆盖状态": status,
            }
        )

    panel_duplicate_count = sum(count - 1 for count in panel_key_counts.values() if count > 1)
    latest_duplicate_count = sum(count - 1 for count in latest_code_counts.values() if count > 1)
    missing_duplicate_count = sum(count - 1 for count in missing_code_counts.values() if count > 1)
    duplicate_exists = panel_duplicate_count + latest_duplicate_count + missing_duplicate_count > 0
    code_outside_exists = bool(outside_records or extra_in_combined)
    numeric_blocking_anomalies = sum(
        1
        for item in anomalies
        if item["异常类型"] in {"数值超范围", "数值不可转换", "机构与个人比例合计异常"}
    )
    name_mismatch_codes = {
        row["Wind代码"] for row in name_mismatches
    }
    full_coverage = (
        len(combined_coverage) == len(pool_codes)
        and not missing_from_combined
        and not extra_in_combined
        and not overlap_data_missing
    )

    hard_fail = any(
        [
            len(pool_codes) != 223,
            not full_coverage,
            code_outside_exists,
            duplicate_exists,
            bool(panel_missing_fields),
            numeric_blocking_anomalies > 0,
        ]
    )
    if hard_fail:
        conclusion = "不通过"
        conclusion_note = "存在覆盖、代码池、重复键、字段结构或数值质量方面的阻断性问题。"
    elif name_mismatch_codes:
        conclusion = "需要人工确认"
        conclusion_note = (
            f"代码覆盖和数值规则验收通过，但发现{name_mismatch_codes.__len__()}只基金证券简称不一致。"
            "不一致数量较多，可能存在证券简称字段错位或排序错位。若确认Wind代码对应的数据无误，"
            "后续整合时可用主产品池简称覆盖；若无法确认，则不建议整合。"
        )
    else:
        conclusion = "通过"
        conclusion_note = "覆盖、主键、数值及简称一致性均符合验收标准。"

    overview = [
        {"指标": "主产品池广义策略ETF数量", "实际结果": len(pool_codes), "验收标准": "223", "状态": "符合" if len(pool_codes) == 223 else "异常", "说明": ""},
        {"指标": "最新报告期原始数据行数（不含表头）", "实际结果": len(latest_raw), "验收标准": "记录原始行数", "状态": "信息", "说明": f"其中非产品行{len(latest_raw)-len(latest_records)}条"},
        {"指标": "最新报告期有效Wind代码行数", "实际结果": len(latest_records), "验收标准": "剔除空代码和Wind来源行", "状态": "符合", "说明": ""},
        {"指标": "最新报告期真正有数据数量", "实际结果": len(latest_data_codes), "验收标准": "141", "状态": "符合" if len(latest_data_codes) == 141 else "异常", "说明": "关键比例字段并非全部为空"},
        {"指标": "最新报告期空数据行数量", "实际结果": len(latest_empty_rows), "验收标准": "并入缺失清单", "状态": "需修正" if latest_empty_rows else "无", "说明": "机构比例、个人比例、前十大持有人比例均为空"},
        {"指标": "原缺失基金清单有效数量", "实际结果": len(original_missing_by_code), "验收标准": "剔除Wind来源行", "状态": "符合", "说明": ""},
        {"指标": "修正后缺失基金数量", "实际结果": len(corrected_missing_codes), "验收标准": "82", "状态": "符合" if len(corrected_missing_codes) == 82 else "异常", "说明": "原缺失清单与最新空数据代码取并集，并剔除真正有数据代码"},
        {"指标": "面板数据有效行数", "实际结果": len(panel_records), "验收标准": "剔除非产品行", "状态": "符合", "说明": ""},
        {"指标": "面板数据覆盖Wind代码数量", "实际结果": len(panel_codes), "验收标准": "信息项", "状态": "信息", "说明": f"主产品池另有{len(pool_codes-panel_codes)}只未在历史面板出现"},
        {"指标": "是否完整覆盖223只", "实际结果": "是" if full_coverage else "否", "验收标准": "最新有数据基金+修正后缺失基金完整且互斥", "状态": "符合" if full_coverage else "异常", "说明": f"未覆盖{len(missing_from_combined)}只；额外{len(extra_in_combined)}只；重叠{len(overlap_data_missing)}只"},
        {"指标": "是否存在代码池外产品", "实际结果": "是" if code_outside_exists else "否", "验收标准": "否", "状态": "异常" if code_outside_exists else "无", "说明": f"代码池外记录{len(outside_records)}条"},
        {"指标": "是否存在重复记录", "实际结果": "是" if duplicate_exists else "否", "验收标准": "否", "状态": "异常" if duplicate_exists else "无", "说明": f"面板重复键{panel_duplicate_count}条；最新重复代码{latest_duplicate_count}条；缺失清单重复代码{missing_duplicate_count}条"},
        {"指标": "是否存在证券简称不一致", "实际结果": "是" if name_mismatch_codes else "否", "验收标准": "不一致需明确标记", "状态": "需人工确认" if name_mismatch_codes else "无", "说明": f"不一致记录{name_mismatches.__len__()}条，涉及{name_mismatch_codes.__len__()}只代码"},
        {"指标": "数值字段阻断性异常数量", "实际结果": numeric_blocking_anomalies, "验收标准": "0", "状态": "符合" if numeric_blocking_anomalies == 0 else "异常", "说明": "含不可转数值、比例超范围、份额为负及机构+个人比例合计异常"},
        {"指标": "面板指定报告期覆盖", "实际结果": "完整" if not missing_periods else "缺失：" + "、".join(sorted(missing_periods)), "验收标准": "覆盖2023年年报至2026年一季报的六个指定报告期", "状态": "符合" if not missing_periods else "异常", "说明": ""},
        {"指标": "非产品行数量", "实际结果": len(non_product_rows), "验收标准": "验收统计时剔除", "状态": "符合", "说明": "原始文件未修改"},
        {"指标": "最终验收结论", "实际结果": conclusion, "验收标准": "通过/不通过/需要人工确认", "状态": conclusion, "说明": conclusion_note},
    ]

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("验收总览")
    write_sheet(ws, ["指标", "实际结果", "验收标准", "状态", "说明"], overview)

    ws = wb.create_sheet("代码覆盖核对")
    write_sheet(
        ws,
        [
            "Wind代码",
            "主产品池证券简称",
            "是否在最新报告期有数据",
            "是否最新报告期空数据",
            "是否在缺失基金清单",
            "是否在面板数据中出现",
            "覆盖状态",
        ],
        coverage_rows,
    )

    ws = wb.create_sheet("最新报告期空数据行")
    empty_headers = latest_headers + ["主产品池证券简称", "处理建议"]
    write_sheet(ws, empty_headers, latest_empty_output)

    ws = wb.create_sheet("修正后缺失基金清单")
    missing_output_headers = [
        "Wind代码",
        "主产品池证券简称",
        "缺失清单证券简称",
        "是否来自原缺失清单",
        "是否来自最新报告期空数据行",
        "基金成立日",
        "上市日期",
        "基金状态",
        "最新基金规模(亿)",
        "缺失原因或备注",
        "一级策略大类",
        "二级策略类别",
    ]
    write_sheet(ws, missing_output_headers, corrected_missing_rows)

    ws = wb.create_sheet("证券简称不一致检查")
    mismatch_headers = [
        "Wind代码",
        "持有人数据证券简称",
        "主产品池证券简称",
        "来源文件",
        "报告期",
        "是否需要人工确认",
    ]
    write_sheet(
        ws,
        mismatch_headers,
        sorted(name_mismatches, key=lambda r: (r["Wind代码"], r["来源文件"], str(r["报告期"]))),
    )

    ws = wb.create_sheet("面板报告期覆盖统计")
    period_headers = [
        "报告期",
        "记录数量",
        "有机构/个人比例的记录数量",
        "有持有份额的记录数量",
        "覆盖Wind代码数量",
    ]
    write_sheet(ws, period_headers, panel_period_summary)

    ws = wb.create_sheet("数值字段完整性检查")
    numeric_headers = [
        "数据集",
        "字段",
        "字段是否存在",
        "总记录数",
        "非空数量",
        "缺失数量",
        "可转数值数量",
        "不可转数值数量",
        "异常数量",
        "说明",
    ]
    write_sheet(ws, numeric_headers, numeric_stats)

    ws = wb.create_sheet("异常记录明细")
    anomaly_headers = [
        "异常类型",
        "来源文件",
        "报告期",
        "Wind代码",
        "证券简称",
        "字段",
        "原始值",
        "相关值或期望值",
        "异常说明",
        "是否阻断通过",
    ]
    write_sheet(
        ws,
        anomaly_headers,
        sorted(
            anomalies,
            key=lambda r: (
                str(r["异常类型"]),
                str(r["Wind代码"]),
                str(r["报告期"]),
                str(r["来源文件"]),
            ),
        ),
    )

    format_workbook(wb)
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    temp_file = OUTPUT_FILE.with_name(OUTPUT_FILE.stem + "_tmp.xlsx")
    wb.save(temp_file)

    # 导出后可打开性与结构验证
    check_wb = load_workbook(temp_file, read_only=True, data_only=True)
    expected_sheets = [
        "验收总览",
        "代码覆盖核对",
        "最新报告期空数据行",
        "修正后缺失基金清单",
        "证券简称不一致检查",
        "面板报告期覆盖统计",
        "数值字段完整性检查",
        "异常记录明细",
    ]
    if check_wb.sheetnames != expected_sheets:
        raise RuntimeError(f"输出sheet结构异常：{check_wb.sheetnames}")
    check_wb.close()
    temp_file.replace(OUTPUT_FILE)

    changed_sources = [str(path) for path in inputs if sha256(path) != source_hashes[path]]
    if changed_sources:
        raise RuntimeError("原始文件被意外修改：" + "；".join(changed_sources))

    print(f"主产品池广义策略ETF数量：{len(pool_codes)}")
    print(f"最新报告期有效Wind代码行数：{len(latest_records)}")
    print(f"最新报告期真正有数据数量：{len(latest_data_codes)}")
    print(f"最新报告期空数据行数量：{len(latest_empty_rows)}")
    print(f"原缺失基金清单有效数量：{len(original_missing_by_code)}")
    print(f"修正后缺失基金数量：{len(corrected_missing_codes)}")
    print(f"面板有效行数：{len(panel_records)}")
    print(f"面板覆盖Wind代码数量：{len(panel_codes)}")
    print(f"证券简称不一致记录数：{len(name_mismatches)}")
    print(f"证券简称不一致Wind代码数：{len(name_mismatch_codes)}")
    print(f"数值字段阻断性异常数量：{numeric_blocking_anomalies}")
    print(f"最终验收结论：{conclusion}")
    print(f"输出文件路径：{OUTPUT_FILE}")


if __name__ == "__main__":
    main()
