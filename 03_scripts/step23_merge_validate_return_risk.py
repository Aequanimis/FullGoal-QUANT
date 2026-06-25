from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT_DIR = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池")
POOL_FILE = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\全市场ETF基础信息_策略ETF池二次修正版.xlsx")
POOL_SHEET = "策略ETF_最终统计池"
OUTPUT = INPUT_DIR / "广义策略ETF_收益风险表现表_合并验收清洗版.xlsx"

MAIN_SHEET = "收益风险_合并清洗版"
SUP_RAW_SHEET = "补充文件_原始宽表"
DUP_SHEET = "重复代码检查"
MISSING_SHEET = "缺失代码清单"
EXTRA_SHEET = "多余代码清单"
FIELD_MISSING_SHEET = "字段缺失统计"
SUMMARY_SHEET = "验收汇总"
METHOD_SHEET = "口径说明"

POOL_FIELDS = [
    "Wind代码",
    "证券简称",
    "基金简称",
    "基金全称",
    "基金代码",
    "交易代码",
    "基金管理人",
    "基金托管人",
    "基金上市地点",
    "上市日期",
    "基金成立日",
    "投资类型_二级分类",
    "跟踪指数代码",
    "跟踪指数名称",
    "最新基金规模(亿)",
    "一级策略大类",
    "二级策略类别",
    "市场范围_二次修正",
    "统计口径分类",
    "是否纳入核心策略ETF统计",
    "是否纳入广义策略ETF统计",
]

STANDARD_RISK_FIELDS = [
    "证券简称_Wind",
    "复权单位净值增长率(%)",
    "年化波动率(%)",
    "最大回撤(%)",
    "SHARPE",
    "跟踪误差(%)",
    "信息比率",
    "区间净值超越基准收益率(%)",
    "今年以来收益率(%)",
    "近1月收益率(%)",
    "近3月收益率(%)",
    "近6月收益率(%)",
    "近1年收益率(%)",
    "近3年收益率(%)",
    "近5年收益率(%)",
    "成立以来收益率(%)",
    "成立以来收益率起始时间",
    "成立以来收益率截止时间",
    "基金成立日_Wind",
    "数据来源文件",
]

NUMERIC_FIELDS = [
    "复权单位净值增长率(%)",
    "年化波动率(%)",
    "最大回撤(%)",
    "SHARPE",
    "跟踪误差(%)",
    "信息比率",
    "区间净值超越基准收益率(%)",
    "今年以来收益率(%)",
    "近1月收益率(%)",
    "近3月收益率(%)",
    "近6月收益率(%)",
    "近1年收益率(%)",
    "近3年收益率(%)",
    "近5年收益率(%)",
    "成立以来收益率(%)",
]

DATE_FIELDS = [
    "成立以来收益率起始时间",
    "成立以来收益率截止时间",
    "基金成立日_Wind",
]

MAPPED_FIELDS = {
    "复权单位净值增长率(%)": ["复权单位净值增长率(%)", "成立以来复权单位净值增长率(%)"],
    "年化波动率(%)": ["年化波动率(%)", "成立以来年化波动率(%)"],
    "最大回撤(%)": ["最大回撤(%)", "成立以来最大回撤(%)"],
    "SHARPE": ["SHARPE", "成立以来SHARPE"],
    "跟踪误差(%)": ["跟踪误差(%)", "成立以来跟踪误差(%)"],
    "信息比率": ["信息比率", "成立以来信息比率"],
    "区间净值超越基准收益率(%)": [
        "区间净值超越基准收益率(%)",
        "成立以来净值超越基准收益率(%)",
    ],
    "今年以来收益率(%)": ["今年以来收益率(%)", "今年以来复权单位净值增长率(%)"],
    "近1月收益率(%)": ["近1月收益率(%)", "近1月复权单位净值增长率(%)"],
    "近3月收益率(%)": ["近3月收益率(%)", "近3月复权单位净值增长率(%)"],
    "近6月收益率(%)": ["近6月收益率(%)", "近6月复权单位净值增长率(%)"],
    "近1年收益率(%)": ["近1年收益率(%)", "近1年复权单位净值增长率(%)"],
    "近3年收益率(%)": ["近3年收益率(%)", "近3年复权单位净值增长率(%)"],
    "近5年收益率(%)": ["近5年收益率(%)", "近5年复权单位净值增长率(%)"],
    "成立以来收益率(%)": ["成立以来收益率(%)", "成立以来复权单位净值增长率(%)"],
    "成立以来收益率起始时间": [
        "成立以来收益率起始时间",
        "成立以来复权单位净值增长率起始时间",
    ],
    "成立以来收益率截止时间": [
        "成立以来收益率截止时间",
        "成立以来复权单位净值增长率截止时间",
    ],
}

MISSING_CHECK_FIELDS = NUMERIC_FIELDS + DATE_FIELDS

PRIMARY_ORDER = [
    "红利",
    "自由现金流",
    "质量",
    "价值",
    "成长",
    "低波",
    "ESG",
    "基本面策略",
    "等权/另类加权",
    "指数增强/多因子",
    "待核验",
]

SECONDARY_ORDER = [
    "普通红利",
    "红利低波",
    "港股通低波红利",
    "红利质量",
    "红利价值",
    "红利成长",
    "央企红利",
    "国企红利",
    "港股通红利",
    "股东回报",
    "央企股东回报",
    "自由现金流",
    "质量",
    "价值",
    "成长",
    "低波",
    "ESG",
    "基本面策略",
    "等权/另类加权",
    "指数增强/多因子",
    "待核验",
]

CLASS_ORDER = {"核心策略指数ETF": 0, "广义策略ETF_指数增强": 1}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=9, color="1F1F1F")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN_GRAY = Side(style="thin", color="D9E1F2")
BODY_BORDER = Border(bottom=THIN_GRAY)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    value = str(value).replace("\u3000", " ").replace("\u200b", "")
    return re.sub(r"\s+", " ", value).strip()


def normalize_code(value: Any) -> str:
    raw = clean_text(value).upper().replace("。", ".")
    match = re.search(r"(\d{6})\s*[.]?\s*(SH|SZ)$", raw)
    return f"{match.group(1)}.{match.group(2)}" if match else raw


def to_number(value: Any) -> float | int | None:
    if value is None or clean_text(value) in {"", "--", "N/A", "nan"}:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    try:
        number = float(clean_text(value).replace(",", "").replace("%", ""))
        return int(number) if number.is_integer() else number
    except ValueError:
        return None


def to_date(value: Any) -> datetime | None:
    if value is None or not clean_text(value):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    raw = clean_text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def first_nonempty(raw: dict[str, Any], candidates: list[str]) -> Any:
    for field in candidates:
        value = raw.get(field)
        if value is not None and clean_text(value) != "":
            return value
    return None


def file_rank(path: Path) -> int:
    if "补充" in path.name:
        return 6
    match = re.search(r"第(\d+)批", path.name)
    return int(match.group(1)) if match else 0


def load_pool() -> dict[str, dict[str, Any]]:
    wb = load_workbook(POOL_FILE, read_only=True, data_only=True)
    ws = wb[POOL_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows())]
    missing = [field for field in POOL_FIELDS if field not in headers]
    if missing:
        raise KeyError(f"产品池缺少字段：{missing}")
    pool = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, values))
        if clean_text(record.get("是否纳入广义策略ETF统计")) != "是":
            continue
        code = normalize_code(record.get("Wind代码"))
        copied = {field: record.get(field) for field in POOL_FIELDS}
        copied["Wind代码"] = code
        copied["上市日期"] = to_date(copied.get("上市日期"))
        copied["基金成立日"] = to_date(copied.get("基金成立日"))
        pool[code] = copied
    return pool


def parse_files() -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    int,
]:
    files = [
        path
        for path in INPUT_DIR.glob("*收益风险表现表*.xlsx")
        if path.resolve() != OUTPUT.resolve() and "合并验收清洗版" not in path.name
    ]
    files.sort(key=lambda p: (file_rank(p), p.name))
    parsed = []
    supplement_raw = []
    file_info = []
    source_order = {path.name: idx for idx, path in enumerate(files)}

    for path in files:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [clean_text(cell.value) for cell in next(ws.iter_rows())]
        valid_count = 0
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not any(value is not None and clean_text(value) for value in values):
                continue
            raw = dict(zip(headers, values))
            code = normalize_code(raw.get("Wind代码"))
            security_name = clean_text(raw.get("证券简称"))
            if (
                not code
                or code.startswith("数据来源")
                or security_name.startswith("数据来源")
                or not re.fullmatch(r"\d{6}\.(SH|SZ)", code)
            ):
                continue
            valid_count += 1
            record = {
                "Wind代码": code,
                "证券简称_Wind": security_name or None,
                "基金成立日_Wind": to_date(raw.get("基金成立日")),
                "数据来源文件": path.name,
                "_文件顺序": source_order[path.name],
            }
            for field, candidates in MAPPED_FIELDS.items():
                value = first_nonempty(raw, candidates)
                if field in NUMERIC_FIELDS:
                    value = to_number(value)
                elif field in DATE_FIELDS:
                    value = to_date(value)
                record[field] = value
            parsed.append(record)

            if "补充" in path.name:
                raw_copy = {header: raw.get(header) for header in headers}
                raw_copy["数据来源文件"] = path.name
                supplement_raw.append(raw_copy)

        file_info.append(
            {
                "文件名": path.name,
                "sheet": ws.title,
                "原始字段数": len(headers),
                "有效代码行数": valid_count,
            }
        )
    return parsed, supplement_raw, file_info, len(files)


def choose_unique(
    records: list[dict[str, Any]],
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    groups: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        groups[record["Wind代码"]].append(record)
    unique = {}
    duplicate_rows = []
    completeness_fields = [
        field for field in STANDARD_RISK_FIELDS if field not in {"证券简称_Wind", "数据来源文件"}
    ]
    for code, items in groups.items():
        ranked = sorted(
            items,
            key=lambda r: (
                sum(r.get(field) is not None and clean_text(r.get(field)) != "" for field in completeness_fields),
                r["_文件顺序"],
            ),
            reverse=True,
        )
        unique[code] = ranked[0]
        if len(items) > 1:
            files = "；".join(sorted({r["数据来源文件"] for r in items}))
            duplicate_rows.append(
                {
                    "Wind代码": code,
                    "出现次数": len(items),
                    "出现文件": files,
                    "保留文件": ranked[0]["数据来源文件"],
                    "处理规则": "优先保留非空字段更多的记录；相同时保留文件排序更靠后的批次",
                }
            )
    return unique, duplicate_rows


def normalize_name(value: Any) -> str:
    return re.sub(r"[\s（）()]+", "", clean_text(value)).upper()


def write_records(ws, headers: list[str], records: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for record in records:
        ws.append([record.get(header) for header in headers])


def style_sheet(ws, headers: list[str]) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BODY_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
    sample_limit = min(ws.max_row, 300)
    wide = {
        "基金全称",
        "跟踪指数名称",
        "缺失字段列表",
        "备注",
        "说明",
        "口径内容",
        "数据来源文件",
        "出现文件",
        "处理规则",
    }
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, sample_limit + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 70))
        width = min(max(max_len * 1.1, 11), 55 if header in wide else 30)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for col_idx, header in enumerate(headers, 1):
        if "日期" in header or "时间" in header or header in {"上市日期", "基金成立日", "基金成立日_Wind"}:
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (datetime, date)):
                    ws.cell(row_idx, col_idx).number_format = "yyyy-mm-dd"
        if header in NUMERIC_FIELDS or "规模" in header:
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "0.0000" if header in NUMERIC_FIELDS else "#,##0.0000"


def main() -> None:
    pool = load_pool()
    parsed, supplement_raw, file_info, file_count = parse_files()
    unique, duplicate_rows = choose_unique(parsed)
    data_codes = set(unique)
    pool_codes = set(pool)
    missing_codes = sorted(pool_codes - data_codes)
    extra_codes = sorted(data_codes - pool_codes)

    main_rows = []
    for code, product in pool.items():
        risk = unique.get(code)
        row = {field: product.get(field) for field in POOL_FIELDS}
        if risk:
            for field in STANDARD_RISK_FIELDS:
                row[field] = risk.get(field)
        else:
            for field in STANDARD_RISK_FIELDS:
                row[field] = None
        missing_fields = [field for field in MISSING_CHECK_FIELDS if row.get(field) is None]
        row["是否匹配收益风险数据"] = "是" if risk else "否"
        row["缺失字段数量"] = len(missing_fields)
        row["缺失字段列表"] = "；".join(missing_fields)
        if risk:
            name_match = normalize_name(row.get("基金简称")) == normalize_name(row.get("证券简称_Wind"))
            row["名称是否一致"] = "是" if name_match else "否"
            notes = []
            if not name_match:
                notes.append("产品池基金简称与Wind证券简称不完全一致")
            if any(field in missing_fields for field in ["近1年收益率(%)", "近3年收益率(%)", "近5年收益率(%)"]):
                notes.append("长期区间收益缺失可能因成立时间较短")
            row["备注"] = "；".join(notes)
        else:
            row["名称是否一致"] = None
            row["备注"] = "产品池中有代码，但收益风险批次未覆盖"
        main_rows.append(row)

    class_rank = CLASS_ORDER
    primary_rank = {value: idx for idx, value in enumerate(PRIMARY_ORDER)}
    secondary_rank = {value: idx for idx, value in enumerate(SECONDARY_ORDER)}
    main_rows.sort(
        key=lambda r: (
            class_rank.get(clean_text(r.get("统计口径分类")), 999),
            primary_rank.get(clean_text(r.get("一级策略大类")), 999),
            secondary_rank.get(clean_text(r.get("二级策略类别")), 999),
            -(float(r["最新基金规模(亿)"]) if isinstance(r.get("最新基金规模(亿)"), (int, float)) else -1),
            r["Wind代码"],
        )
    )

    missing_rows = [
        {
            "Wind代码": code,
            "证券简称": pool[code].get("证券简称"),
            "基金简称": pool[code].get("基金简称"),
            "统计口径分类": pool[code].get("统计口径分类"),
            "一级策略大类": pool[code].get("一级策略大类"),
            "二级策略类别": pool[code].get("二级策略类别"),
            "市场范围_二次修正": pool[code].get("市场范围_二次修正"),
            "跟踪指数代码": pool[code].get("跟踪指数代码"),
            "跟踪指数名称": pool[code].get("跟踪指数名称"),
            "缺失原因": "产品池中有代码，但所有收益风险批次均未出现",
        }
        for code in missing_codes
    ]
    extra_rows = [
        {
            "Wind代码": code,
            "证券简称_Wind": unique[code].get("证券简称_Wind"),
            "数据来源文件": unique[code].get("数据来源文件"),
            "处理建议": "不纳入223只广义策略ETF主表，人工核验产品范围",
        }
        for code in extra_codes
    ]

    field_missing_rows = []
    matched_rows = [row for row in main_rows if row["是否匹配收益风险数据"] == "是"]
    for field in NUMERIC_FIELDS + DATE_FIELDS:
        nonempty = sum(row.get(field) is not None for row in matched_rows)
        missing = len(matched_rows) - nonempty
        field_missing_rows.append(
            {
                "字段": field,
                "总行数": len(matched_rows),
                "非空数量": nonempty,
                "缺失数量": missing,
                "缺失比例": missing / len(matched_rows) if matched_rows else None,
                "说明": (
                    "长期区间字段缺失可能由产品成立时间较短导致，不删除产品"
                    if field in {"近1年收益率(%)", "近3年收益率(%)", "近5年收益率(%)"}
                    else ""
                ),
            }
        )

    summary_rows = []

    def add_summary(category: str, item: str, value: Any, note: str = "") -> None:
        summary_rows.append({"统计类别": category, "指标": item, "数值": value, "说明": note})

    matched_codes = pool_codes & data_codes
    add_summary("总体", "产品池理论数量", len(pool))
    add_summary("总体", "收益风险批次原始有效代码数量", len(data_codes))
    add_summary("总体", "清洗后唯一有效代码数量", len(unique))
    add_summary("总体", "匹配到产品池的数量", len(matched_codes))
    add_summary("总体", "缺失代码数量", len(missing_codes))
    add_summary("总体", "多余代码数量", len(extra_codes))
    add_summary("总体", "重复代码数量", len(duplicate_rows))
    core_codes = {code for code, p in pool.items() if p.get("是否纳入核心策略ETF统计") == "是"}
    enhanced_codes = {code for code, p in pool.items() if p.get("统计口径分类") == "广义策略ETF_指数增强"}
    add_summary("口径覆盖", "核心策略指数ETF理论数量", len(core_codes))
    add_summary("口径覆盖", "核心策略指数ETF已匹配数量", len(core_codes & data_codes))
    add_summary("口径覆盖", "指数增强/多因子ETF理论数量", len(enhanced_codes))
    add_summary("口径覆盖", "指数增强/多因子ETF已匹配数量", len(enhanced_codes & data_codes))
    primary_counts = Counter(pool[code].get("一级策略大类") for code in matched_codes)
    class_counts = Counter(pool[code].get("统计口径分类") for code in matched_codes)
    for primary in PRIMARY_ORDER:
        if primary_counts.get(primary):
            add_summary("各一级策略大类匹配数量", primary, primary_counts[primary])
    for category, count in sorted(class_counts.items()):
        add_summary("各统计口径分类匹配数量", category, count)
    key_fields = [
        "年化波动率(%)",
        "最大回撤(%)",
        "SHARPE",
        "跟踪误差(%)",
        "信息比率",
        "今年以来收益率(%)",
        "近1月收益率(%)",
        "近3月收益率(%)",
        "近6月收益率(%)",
        "近1年收益率(%)",
        "近3年收益率(%)",
        "近5年收益率(%)",
        "成立以来收益率(%)",
    ]
    for field in key_fields:
        add_summary(
            "关键字段缺失数量",
            field,
            sum(row.get(field) is None for row in matched_rows),
        )
    for info in file_info:
        add_summary(
            "读取文件",
            info["文件名"],
            info["有效代码行数"],
            f"sheet={info['sheet']}；原始字段数={info['原始字段数']}",
        )

    method_rows = [
        {"项目": "理论产品池", "口径内容": "以“策略ETF_最终统计池”中“是否纳入广义策略ETF统计=是”的223只ETF为理论产品池。"},
        {"项目": "产品池构成", "口径内容": "广义策略ETF包括168只核心策略指数ETF和55只指数增强/多因子ETF。"},
        {"项目": "数据来源", "口径内容": "收益风险数据来自Wind分批导出文件。"},
        {"项目": "第1–4批", "口径内容": "第1–4批字段格式较统一。"},
        {"项目": "第5批", "口径内容": "第5批字段顺序不同，本次按字段名称识别和合并。"},
        {"项目": "补充文件", "口径内容": "补充文件字段更宽，核心字段已按优先级映射到主表，全部原始宽字段保留在“补充文件_原始宽表”。"},
        {"项目": "百分比口径", "口径内容": "百分比字段保留Wind原始百分数数值，不除以100，例如12.5表示12.5%。"},
        {"项目": "长期字段缺失", "口径内容": "近1年、近3年、近5年等字段缺失可能由产品成立时间较短导致，仅记录缺失，不删除产品。"},
        {"项目": "重复代码规则", "口径内容": "同一Wind代码优先保留非空标准字段更多的记录；相同时保留文件名排序更靠后的批次。"},
    ]

    wb = Workbook()
    wb.remove(wb.active)
    main_ws = wb.create_sheet(MAIN_SHEET)
    sup_ws = wb.create_sheet(SUP_RAW_SHEET)
    dup_ws = wb.create_sheet(DUP_SHEET)
    missing_ws = wb.create_sheet(MISSING_SHEET)
    extra_ws = wb.create_sheet(EXTRA_SHEET)
    field_ws = wb.create_sheet(FIELD_MISSING_SHEET)
    summary_ws = wb.create_sheet(SUMMARY_SHEET)
    method_ws = wb.create_sheet(METHOD_SHEET)

    main_headers = POOL_FIELDS + STANDARD_RISK_FIELDS + [
        "是否匹配收益风险数据",
        "缺失字段数量",
        "缺失字段列表",
        "名称是否一致",
        "备注",
    ]
    write_records(main_ws, main_headers, main_rows)

    if supplement_raw:
        sup_headers = list(supplement_raw[0].keys())
        write_records(sup_ws, sup_headers, supplement_raw)
    else:
        sup_headers = ["说明"]
        write_records(sup_ws, sup_headers, [{"说明": "未读取到补充文件有效记录"}])

    dup_headers = ["Wind代码", "出现次数", "出现文件", "保留文件", "处理规则"]
    if duplicate_rows:
        write_records(dup_ws, dup_headers, duplicate_rows)
    else:
        write_records(dup_ws, dup_headers, [{"Wind代码": "无重复代码"}])

    missing_headers = [
        "Wind代码", "证券简称", "基金简称", "统计口径分类", "一级策略大类",
        "二级策略类别", "市场范围_二次修正", "跟踪指数代码", "跟踪指数名称", "缺失原因",
    ]
    if missing_rows:
        write_records(missing_ws, missing_headers, missing_rows)
    else:
        write_records(missing_ws, missing_headers, [{"Wind代码": "无缺失，广义策略 ETF 223 只全部覆盖"}])

    extra_headers = ["Wind代码", "证券简称_Wind", "数据来源文件", "处理建议"]
    if extra_rows:
        write_records(extra_ws, extra_headers, extra_rows)
    else:
        write_records(extra_ws, extra_headers, [{"Wind代码": "无多余有效 ETF 代码"}])

    field_headers = ["字段", "总行数", "非空数量", "缺失数量", "缺失比例", "说明"]
    write_records(field_ws, field_headers, field_missing_rows)
    summary_headers = ["统计类别", "指标", "数值", "说明"]
    write_records(summary_ws, summary_headers, summary_rows)
    method_headers = ["项目", "口径内容"]
    write_records(method_ws, method_headers, method_rows)

    for ws, headers in [
        (main_ws, main_headers),
        (sup_ws, sup_headers),
        (dup_ws, dup_headers),
        (missing_ws, missing_headers),
        (extra_ws, extra_headers),
        (field_ws, field_headers),
        (summary_ws, summary_headers),
        (method_ws, method_headers),
    ]:
        style_sheet(ws, headers)

    ratio_col = field_headers.index("缺失比例") + 1
    for row_idx in range(2, field_ws.max_row + 1):
        field_ws.cell(row_idx, ratio_col).number_format = "0.00%"

    previous = None
    for row_idx in range(2, summary_ws.max_row + 1):
        section = summary_ws.cell(row_idx, 1).value
        if section != previous:
            for col_idx in range(1, summary_ws.max_column + 1):
                summary_ws.cell(row_idx, col_idx).fill = SECTION_FILL
                summary_ws.cell(row_idx, col_idx).font = Font(name="微软雅黑", size=9, bold=True)
        previous = section

    method_ws.column_dimensions["A"].width = 24
    method_ws.column_dimensions["B"].width = 110
    for row_idx in range(2, method_ws.max_row + 1):
        method_ws.cell(row_idx, 2).alignment = Alignment(wrap_text=True, vertical="top")
        method_ws.row_dimensions[row_idx].height = 45

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    print(f"读取的批次文件数量：{file_count}")
    print(f"产品池理论数量：{len(pool)}")
    print(f"清洗后唯一有效代码数量：{len(unique)}")
    print(f"缺失代码数量：{len(missing_codes)}")
    print(f"多余代码数量：{len(extra_codes)}")
    print(f"重复代码数量：{len(duplicate_rows)}")
    print(f"核心策略指数ETF匹配数量：{len(core_codes & data_codes)}")
    print(f"指数增强/多因子ETF匹配数量：{len(enhanced_codes & data_codes)}")
    print(f"输出文件路径：{OUTPUT}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"处理失败：{exc}", file=sys.stderr)
        raise
