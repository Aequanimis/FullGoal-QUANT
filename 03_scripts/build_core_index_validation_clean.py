from __future__ import annotations

import hashlib
import math
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


POOL_FILE = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\全市场ETF基础信息_策略ETF池二次修正版.xlsx")
RAW_FILE = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\sheet6\核心策略指数ETF跟踪指数全量数据.xlsx")
SUP_FILE = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\sheet6\7个补充指数全量数据.xlsx")
OUTPUT_FILE = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\sheet6\核心策略指数ETF跟踪指数_验收清洗版.xlsx")

SUPPLEMENT_CODES = {
    "888888.FI", "CSPSADRP.CI", "932305.CSI", "987016.CNI",
    "931243.CSI", "930782.CSI", "931476.CSI",
}
VALUATION_FIELDS = ["PE_TTM", "PB_LF", "股息率_TTM"]
MONTHLY_FIELDS = [
    "日期", "指数代码", "指数名称", "指数发布机构", "收盘点位", "月度收益率",
    "PE_TTM", "PB_LF", "股息率_TTM", "成分股数量", "数据可得性说明",
]
LATEST_FIELDS = [
    "统计日期", "指数代码", "指数名称", "今年以来收益率", "近1年收益率",
    "近3年收益率", "近5年收益率", "PE_TTM", "PB_LF", "股息率_TTM",
]


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def txt(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[\s\u3000\u200b\ufeff]+", "", str(value)).strip()


def code(value: Any) -> str:
    return txt(value).upper()


def number(value: Any) -> float | None:
    if value is None or txt(value).upper() in {"", "-", "--", "N/A", "NA", "NAN", "NONE", "NULL"}:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        result = float(value)
        return None if math.isnan(result) else result
    try:
        return float(txt(value).replace(",", "").replace("，", "").replace("%", ""))
    except ValueError:
        return None


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d %H:%M:%S", "%Y%m%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def read(path: Path, sheet: str) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    iterator = ws.iter_rows(values_only=True)
    headers = [str(v or "").strip() for v in next(iterator)]
    rows = [dict(zip(headers, row)) for row in iterator]
    wb.close()
    return headers, rows


def unique_join(values: list[Any]) -> str:
    result = []
    for value in values:
        v = str(value or "").strip()
        if v and v not in result:
            result.append(v)
    return "；".join(result)


def completeness(row: dict[str, Any], fields: list[str]) -> int:
    return sum(row.get(field) not in (None, "") for field in fields)


def consecutive_month(previous: date, current: date) -> bool:
    return (current.year * 12 + current.month) - (previous.year * 12 + previous.month) == 1


def write_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def format_workbook(wb: Workbook) -> None:
    dark = "1F4E78"
    white = "FFFFFF"
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=dark)
            cell.font = Font(name="微软雅黑", size=10, bold=True, color=white)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 32
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="微软雅黑", size=9)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                header = str(ws.cell(1, cell.column).value or "")
                if isinstance(cell.value, (date, datetime)):
                    cell.number_format = "yyyy-mm-dd"
                elif isinstance(cell.value, (int, float)):
                    if header == "月度收益率":
                        cell.number_format = "0.0000%"
                    elif "收益率" in header or "非空率" in header or "股息率" in header:
                        cell.number_format = "0.0000"
                    elif header in {"PE_TTM", "PB_LF"}:
                        cell.number_format = "0.0000"
                    elif "规模" in header:
                        cell.number_format = "0.0000"
                    elif "点位" in header:
                        cell.number_format = "0.0000"
        for col in range(1, ws.max_column + 1):
            values = [str(ws.cell(row, col).value or "") for row in range(1, min(ws.max_row, 1200) + 1)]
            width = min(max(max((len(v) for v in values), default=8) * 1.05 + 2, 10), 38)
            header = str(ws.cell(1, col).value or "")
            if any(k in header for k in ("备注", "说明", "列表", "明细", "问题")):
                width = min(max(width, 28), 55)
            ws.column_dimensions[get_column_letter(col)].width = width


def main() -> None:
    inputs = [POOL_FILE, RAW_FILE, SUP_FILE]
    missing_files = [str(path) for path in inputs if not path.exists()]
    if missing_files:
        raise FileNotFoundError("缺少输入文件：" + "；".join(missing_files))
    hashes = {path: sha256(path) for path in inputs}
    issues: list[str] = []

    _, pool_rows_all = read(POOL_FILE, "策略ETF_最终统计池")
    core_etfs = [r for r in pool_rows_all if txt(r.get("统计口径分类")) == "核心策略指数ETF"]
    index_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in core_etfs:
        normalized = code(row.get("跟踪指数代码"))
        if not normalized:
            issues.append(f"核心ETF {row.get('Wind代码')} 跟踪指数代码为空")
            continue
        index_groups[normalized].append(row)
    core_codes = set(index_groups)

    index_list = []
    index_meta: dict[str, dict[str, Any]] = {}
    for normalized in sorted(core_codes):
        rows = index_groups[normalized]
        sizes = [number(r.get("最新基金规模(亿)")) for r in rows]
        item = {
            "指数代码_标准化": normalized,
            "原始跟踪指数代码": unique_join([r.get("跟踪指数代码") for r in rows]),
            "跟踪指数名称": unique_join([r.get("跟踪指数名称") for r in rows]),
            "对应ETF数量": len(rows),
            "对应ETF规模合计": sum(v for v in sizes if v is not None),
            "一级策略大类": unique_join([r.get("一级策略大类") for r in rows]),
            "二级策略类别": unique_join([r.get("二级策略类别") for r in rows]),
            "市场范围_二次修正": unique_join([r.get("市场范围_二次修正") for r in rows]),
            "对应ETF代码列表": "；".join(str(r.get("Wind代码") or "") for r in rows),
            "对应ETF简称列表": "；".join(str(r.get("证券简称") or r.get("基金简称") or "") for r in rows),
        }
        index_list.append(item)
        index_meta[normalized] = item

    raw_sheets = {
        "Sheet1-指数清单核对表": read(RAW_FILE, "Sheet1-指数清单核对表")[1],
        "Sheet2-指数基础信息表": read(RAW_FILE, "Sheet2-指数基础信息表")[1],
        "Sheet3-策略指数月度表现估值表": read(RAW_FILE, "Sheet3-策略指数月度表现估值表")[1],
        "Sheet4-指数最新期间收益表": read(RAW_FILE, "Sheet4-指数最新期间收益表")[1],
        "Sheet5-对照基准指数月度表现估值表": read(RAW_FILE, "Sheet5-对照基准指数月度表现估值表")[1],
    }
    sup_sheets = {
        "指数基础信息": read(SUP_FILE, "指数基础信息")[1],
        "指数月度表现估值": read(SUP_FILE, "指数月度表现估值")[1],
        "指数最新期间收益": read(SUP_FILE, "指数最新期间收益")[1],
    }

    raw_basic = raw_sheets["Sheet2-指数基础信息表"]
    raw_monthly = raw_sheets["Sheet3-策略指数月度表现估值表"]
    raw_latest = raw_sheets["Sheet4-指数最新期间收益表"]
    raw_benchmark = raw_sheets["Sheet5-对照基准指数月度表现估值表"]
    sup_basic = sup_sheets["指数基础信息"]
    sup_monthly = sup_sheets["指数月度表现估值"]
    sup_latest = sup_sheets["指数最新期间收益"]

    def code_set(rows: list[dict[str, Any]]) -> set[str]:
        return {code(r.get("指数代码")) for r in rows if code(r.get("指数代码"))}

    # 覆盖数量按月度主数据口径统计；基础信息覆盖与其一致。
    # 不能用“任一sheet出现”口径，否则仅在最新收益表偶然出现的代码会虚增覆盖数量。
    raw_core_coverage = core_codes & code_set(raw_monthly)
    sup_core_coverage = core_codes & code_set(sup_monthly)

    basic_by_code: dict[str, dict[str, Any]] = {}
    for source_rank, source_name, rows in [
        (1, "原始文件", raw_basic),
        (2, "补充文件", sup_basic),
    ]:
        for row in rows:
            c = code(row.get("指数代码"))
            if not c:
                continue
            candidate = dict(row)
            candidate["_source_rank"] = source_rank
            candidate["_source_name"] = source_name
            if c not in basic_by_code or source_rank >= basic_by_code[c]["_source_rank"]:
                basic_by_code[c] = candidate

    zero_value_count = 0
    monthly_candidates: dict[tuple[str, date], dict[str, Any]] = {}
    for source_rank, source_name, rows in [
        (1, "原始文件", raw_monthly),
        (2, "补充文件", sup_monthly),
    ]:
        for sequence, row in enumerate(rows):
            c = code(row.get("指数代码"))
            dt = parse_date(row.get("日期"))
            if not c or not dt:
                continue
            candidate = {field: row.get(field) for field in MONTHLY_FIELDS}
            candidate["指数代码"] = c
            candidate["日期"] = dt
            candidate["收盘点位"] = number(row.get("收盘点位"))
            monthly_return = number(row.get("月度收益率"))
            if source_name == "补充文件" and monthly_return is not None:
                monthly_return /= 100.0
            candidate["月度收益率"] = monthly_return
            for field in VALUATION_FIELDS:
                value = number(row.get(field))
                if value == 0:
                    zero_value_count += 1
                    value = None
                candidate[field] = value
            candidate["成分股数量"] = number(row.get("成分股数量"))
            candidate["_source_rank"] = source_rank
            candidate["_source_name"] = source_name
            candidate["_sequence"] = sequence
            key = (c, dt)
            existing = monthly_candidates.get(key)
            if (
                existing is None
                or source_rank > existing["_source_rank"]
                or (
                    source_rank == existing["_source_rank"]
                    and completeness(candidate, MONTHLY_FIELDS) > completeness(existing, MONTHLY_FIELDS)
                )
            ):
                monthly_candidates[key] = candidate

    cleaned_monthly = []
    by_index: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for (c, _), row in monthly_candidates.items():
        if c in core_codes:
            by_index[c].append(row)
    calculated_returns = 0
    for c, rows in by_index.items():
        rows.sort(key=lambda r: r["日期"])
        previous = None
        for row in rows:
            if (
                row["月度收益率"] is None
                and previous is not None
                and previous["收盘点位"] is not None
                and row["收盘点位"] is not None
                and previous["收盘点位"] != 0
                and consecutive_month(previous["日期"], row["日期"])
            ):
                row["月度收益率"] = row["收盘点位"] / previous["收盘点位"] - 1
                note = str(row.get("数据可得性说明") or "")
                calc_note = "月度收益率由相邻月收盘点位计算补充"
                row["数据可得性说明"] = f"{note}；{calc_note}".strip("；")
                calculated_returns += 1
            meta = index_meta[c]
            cleaned_monthly.append({
                "日期": row["日期"], "指数代码": c,
                "指数名称": row.get("指数名称") or meta["跟踪指数名称"],
                "指数发布机构": row.get("指数发布机构") or basic_by_code.get(c, {}).get("指数发布机构", ""),
                "收盘点位": row["收盘点位"], "月度收益率": row["月度收益率"],
                "PE_TTM": row["PE_TTM"], "PB_LF": row["PB_LF"], "股息率_TTM": row["股息率_TTM"],
                "成分股数量": row["成分股数量"], "数据可得性说明": row.get("数据可得性说明", ""),
                "对应ETF数量": meta["对应ETF数量"], "对应ETF规模合计": meta["对应ETF规模合计"],
                "一级策略大类": meta["一级策略大类"], "二级策略类别": meta["二级策略类别"],
                "市场范围_二次修正": meta["市场范围_二次修正"],
            })
            previous = row
    cleaned_monthly.sort(key=lambda r: (r["指数代码"], r["日期"]))

    latest_candidates: dict[str, dict[str, Any]] = {}
    for source_rank, source_name, rows in [
        (1, "原始文件", raw_latest),
        (2, "补充文件", sup_latest),
    ]:
        for sequence, row in enumerate(rows):
            c = code(row.get("指数代码"))
            if not c or c not in core_codes:
                continue
            dt = parse_date(row.get("统计日期"))
            candidate = {field: row.get(field) for field in LATEST_FIELDS}
            candidate["指数代码"] = c
            candidate["统计日期"] = dt
            note_parts = []
            for field in ["今年以来收益率", "近1年收益率", "近3年收益率", "近5年收益率"]:
                candidate[field] = number(row.get(field))
            for field in VALUATION_FIELDS:
                value = number(row.get(field))
                if value == 0:
                    zero_value_count += 1
                    value = None
                    note_parts.append("估值字段0值已按缺失处理")
                candidate[field] = value
            candidate["备注"] = "；".join(dict.fromkeys(note_parts))
            candidate["_source_rank"] = source_rank
            candidate["_source_name"] = source_name
            candidate["_sequence"] = sequence
            current = latest_candidates.get(c)
            current_date = current.get("统计日期") if current else None
            if (
                current is None
                or (dt or date.min) > (current_date or date.min)
                or ((dt or date.min) == (current_date or date.min) and source_rank > current["_source_rank"])
            ):
                latest_candidates[c] = candidate

    cleaned_latest = []
    for c, row in latest_candidates.items():
        meta = index_meta[c]
        cleaned_latest.append({
            "统计日期": row["统计日期"], "指数代码": c,
            "指数名称": row.get("指数名称") or meta["跟踪指数名称"],
            "今年以来收益率": row["今年以来收益率"], "近1年收益率": row["近1年收益率"],
            "近3年收益率": row["近3年收益率"], "近5年收益率": row["近5年收益率"],
            "PE_TTM": row["PE_TTM"], "PB_LF": row["PB_LF"], "股息率_TTM": row["股息率_TTM"],
            "对应ETF数量": meta["对应ETF数量"], "对应ETF规模合计": meta["对应ETF规模合计"],
            "一级策略大类": meta["一级策略大类"], "二级策略类别": meta["二级策略类别"],
            "市场范围_二次修正": meta["市场范围_二次修正"], "备注": row["备注"],
        })
    cleaned_latest.sort(key=lambda r: r["指数代码"])

    benchmark_codes = code_set(raw_benchmark)
    benchmark_candidates: dict[tuple[str, date], dict[str, Any]] = {}
    for row in raw_benchmark:
        c = code(row.get("指数代码"))
        dt = parse_date(row.get("日期"))
        if not c or not dt:
            continue
        candidate = {field: row.get(field) for field in MONTHLY_FIELDS}
        candidate["指数代码"] = c
        candidate["日期"] = dt
        candidate["收盘点位"] = number(row.get("收盘点位"))
        candidate["月度收益率"] = number(row.get("月度收益率"))
        for field in VALUATION_FIELDS:
            value = number(row.get(field))
            if value == 0:
                zero_value_count += 1
                value = None
            candidate[field] = value
        candidate["成分股数量"] = number(row.get("成分股数量"))
        key = (c, dt)
        if key not in benchmark_candidates or completeness(candidate, MONTHLY_FIELDS) > completeness(benchmark_candidates[key], MONTHLY_FIELDS):
            benchmark_candidates[key] = candidate
    cleaned_benchmark = sorted(benchmark_candidates.values(), key=lambda r: (r["指数代码"], r["日期"]))

    all_raw_code_sources: dict[str, set[str]] = defaultdict(set)
    all_names: dict[str, list[str]] = defaultdict(list)
    all_institutions: dict[str, list[str]] = defaultdict(list)
    for sheet_name, rows in raw_sheets.items():
        for row in rows:
            c = code(row.get("指数代码"))
            if not c:
                continue
            all_raw_code_sources[c].add(sheet_name)
            if row.get("指数名称"):
                all_names[c].append(row.get("指数名称"))
            if row.get("指数发布机构"):
                all_institutions[c].append(row.get("指数发布机构"))
    extra_codes = set(all_raw_code_sources) - core_codes - benchmark_codes

    def extra_reason(name: str) -> str:
        industry_words = ("行业", "银行", "证券", "医药", "芯片", "科技", "互联网", "能源", "消费", "军工", "新能源", "食品", "传媒", "地产")
        broad_words = ("沪深300", "中证500", "中证1000", "上证50", "创业板", "恒生指数", "标普500", "全指")
        reasons = ["不在168只核心策略ETF跟踪指数清单中"]
        if any(word in name for word in industry_words):
            reasons.append("行业主题指数")
        elif any(word in name for word in broad_words):
            reasons.append("宽基或规模指数")
        else:
            reasons.append("观察池或非本次主分析指数")
        return "；".join(reasons)

    extra_rows = []
    for c in sorted(extra_codes):
        name = unique_join(all_names[c])
        extra_rows.append({
            "指数代码": c, "指数名称": name, "指数发布机构": unique_join(all_institutions[c]),
            "出现在哪些原始sheet": "；".join(sorted(all_raw_code_sources[c])),
            "不纳入原因": extra_reason(name),
        })

    monthly_by_code: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in cleaned_monthly:
        monthly_by_code[row["指数代码"]].append(row)
    latest_codes = {r["指数代码"] for r in cleaned_latest}
    basic_codes = core_codes & set(basic_by_code)
    coverage_rows = []
    for item in index_list:
        c = item["指数代码_标准化"]
        rows = monthly_by_code.get(c, [])
        total = len(rows)
        def nonempty_rate(field: str) -> float:
            return round(sum(r.get(field) is not None for r in rows) / total * 100, 2) if total else 0.0
        has_basic = c in basic_codes
        has_monthly = total > 0
        has_latest = c in latest_codes
        close_rate = nonempty_rate("收盘点位")
        pe_rate = nonempty_rate("PE_TTM")
        pb_rate = nonempty_rate("PB_LF")
        div_rate = nonempty_rate("股息率_TTM")
        remarks = []
        if not has_basic:
            remarks.append("缺少基础信息")
        if not has_latest:
            remarks.append("缺少最新期间收益")
        if has_monthly and min(pe_rate, pb_rate, div_rate) < 95:
            remarks.append("估值字段存在明显缺失")
        if not has_monthly:
            status = "代码待核验" if c not in (basic_codes | latest_codes) else "缺失"
        elif has_basic and has_latest and close_rate >= 95 and min(pe_rate, pb_rate, div_rate) >= 95:
            status = "完整覆盖"
        else:
            status = "部分覆盖"
        coverage_rows.append({
            "指数代码_标准化": c, "跟踪指数名称": item["跟踪指数名称"],
            "是否有基础信息": "是" if has_basic else "否",
            "是否有月度表现估值数据": "是" if has_monthly else "否",
            "是否有最新期间收益数据": "是" if has_latest else "否",
            "月度数据最早日期": min((r["日期"] for r in rows), default=None),
            "月度数据最晚日期": max((r["日期"] for r in rows), default=None),
            "月度数据行数": total, "收盘点位非空率": close_rate,
            "月度收益率非空率": nonempty_rate("月度收益率"),
            "PE_TTM非空率": pe_rate, "PB_LF非空率": pb_rate,
            "股息率_TTM非空率": div_rate, "数据状态": status,
            "备注": "；".join(remarks),
        })
    coverage_rows.sort(key=lambda r: r["指数代码_标准化"])

    status_counts = Counter(r["数据状态"] for r in coverage_rows)
    final_coverage = {
        c for c in core_codes
        if c in basic_codes or c in monthly_by_code or c in latest_codes
    }
    monthly_dates = [r["日期"] for r in cleaned_monthly]
    total_monthly = len(cleaned_monthly)
    def overall_rate(field: str) -> float:
        return round(sum(r.get(field) is not None for r in cleaned_monthly) / total_monthly * 100, 2) if total_monthly else 0.0
    missing_indices = [r for r in coverage_rows if r["数据状态"] in {"缺失", "代码待核验"}]

    quality_rows = []
    def quality(category: str, metric: str, value: Any = "", detail: str = "", note: str = ""):
        quality_rows.append({"类别": category, "指标": metric, "数值/结果": value, "指数代码": "", "指数名称": "", "明细/说明": detail, "备注": note})
    quality("总体统计", "核心策略ETF数量", len(core_etfs), "理论值168")
    quality("总体统计", "核心策略ETF唯一跟踪指数数量", len(core_codes), "理论值82")
    quality("覆盖统计", "原始指数文件中核心指数覆盖数量", len(raw_core_coverage))
    quality("覆盖统计", "补充指数文件中核心指数覆盖数量", len(sup_core_coverage))
    quality("覆盖统计", "合并后最终核心指数覆盖数量", len(final_coverage))
    quality("覆盖统计", "完整覆盖指数数量", status_counts["完整覆盖"])
    quality("覆盖统计", "部分覆盖指数数量", status_counts["部分覆盖"])
    quality("覆盖统计", "缺失指数数量", status_counts["缺失"])
    quality("覆盖统计", "代码待核验指数数量", status_counts["代码待核验"])
    quality("范围统计", "额外指数数量", len(extra_codes))
    quality("范围统计", "对照基准指数数量", len(benchmark_codes))
    quality("月度数据", "月度数据最早日期", min(monthly_dates) if monthly_dates else "")
    quality("月度数据", "月度数据最晚日期", max(monthly_dates) if monthly_dates else "")
    quality("月度数据", "月度数据总行数", total_monthly)
    quality("完整性", "收盘点位整体非空率", overall_rate("收盘点位"))
    quality("完整性", "月度收益率整体非空率", overall_rate("月度收益率"))
    quality("完整性", "PE_TTM整体非空率", overall_rate("PE_TTM"))
    quality("完整性", "PB_LF整体非空率", overall_rate("PB_LF"))
    quality("完整性", "股息率_TTM整体非空率", overall_rate("股息率_TTM"))
    quality("清洗处理", "0值估值字段处理数量", zero_value_count)
    quality("清洗处理", "月度收益率由收盘点位补算数量", calculated_returns)
    quality("清洗处理", "补充文件月度收益率单位转换", len([r for r in sup_monthly if number(r.get("月度收益率")) is not None]), "补充文件由百分数除以100，统一为小数收益率")
    for row in missing_indices:
        quality_rows.append({
            "类别": "缺失指数明细", "指标": row["数据状态"], "数值/结果": "",
            "指数代码": row["指数代码_标准化"], "指数名称": row["跟踪指数名称"],
            "明细/说明": row["备注"], "备注": "",
        })
    for row in extra_rows:
        quality_rows.append({
            "类别": "额外指数明细", "指标": row["不纳入原因"], "数值/结果": "",
            "指数代码": row["指数代码"], "指数名称": row["指数名称"],
            "明细/说明": row["出现在哪些原始sheet"], "备注": "",
        })
    if len(core_etfs) != 168:
        issues.append(f"核心策略ETF数量为{len(core_etfs)}，非理论值168")
    if len(core_codes) != 82:
        issues.append(f"唯一核心指数数量为{len(core_codes)}，非理论值82")
    if not SUPPLEMENT_CODES <= core_codes:
        issues.append("指定7个补充指数并非全部存在于产品池核心指数清单")
    if status_counts["缺失"] or status_counts["代码待核验"]:
        issues.append(f"仍有{status_counts['缺失'] + status_counts['代码待核验']}个指数缺失或待核验")
    missing_latest_rows = [r for r in coverage_rows if r["是否有最新期间收益数据"] == "否"]
    for row in missing_latest_rows:
        issues.append(
            f"{row['指数代码_标准化']}（{row['跟踪指数名称']}）缺少最新期间收益数据，"
            "月度数据仍可用于历史分析"
        )
    valuation_partial_count = sum("估值字段存在明显缺失" in str(r.get("备注") or "") for r in coverage_rows)
    if valuation_partial_count:
        issues.append(
            f"{valuation_partial_count}个指数的PE、PB或股息率历史非空率低于95%，"
            "未作前值、后值、均值或中位数填补"
        )
    for issue in issues:
        quality_rows.append({
            "类别": "需要人工核验的问题清单", "指标": "人工核验", "数值/结果": "",
            "指数代码": "", "指数名称": "", "明细/说明": issue, "备注": "",
        })
    if not issues:
        quality_rows.append({
            "类别": "需要人工核验的问题清单", "指标": "无阻断问题", "数值/结果": 0,
            "指数代码": "", "指数名称": "", "明细/说明": "未发现需要人工核验的阻断性问题", "备注": "",
        })

    wb = Workbook()
    wb.remove(wb.active)
    specs = [
        ("核心指数清单_来自产品池", [
            "指数代码_标准化", "原始跟踪指数代码", "跟踪指数名称", "对应ETF数量",
            "对应ETF规模合计", "一级策略大类", "二级策略类别", "市场范围_二次修正",
            "对应ETF代码列表", "对应ETF简称列表",
        ], index_list),
        ("指数数据覆盖验收表", [
            "指数代码_标准化", "跟踪指数名称", "是否有基础信息", "是否有月度表现估值数据",
            "是否有最新期间收益数据", "月度数据最早日期", "月度数据最晚日期", "月度数据行数",
            "收盘点位非空率", "月度收益率非空率", "PE_TTM非空率", "PB_LF非空率",
            "股息率_TTM非空率", "数据状态", "备注",
        ], coverage_rows),
        ("核心策略指数月度表现估值_清洗版", [
            "日期", "指数代码", "指数名称", "指数发布机构", "收盘点位", "月度收益率",
            "PE_TTM", "PB_LF", "股息率_TTM", "成分股数量", "数据可得性说明",
            "对应ETF数量", "对应ETF规模合计", "一级策略大类", "二级策略类别",
            "市场范围_二次修正",
        ], cleaned_monthly),
        ("核心策略指数最新期间收益_清洗版", [
            "统计日期", "指数代码", "指数名称", "今年以来收益率", "近1年收益率",
            "近3年收益率", "近5年收益率", "PE_TTM", "PB_LF", "股息率_TTM",
            "对应ETF数量", "对应ETF规模合计", "一级策略大类", "二级策略类别",
            "市场范围_二次修正", "备注",
        ], cleaned_latest),
        ("对照基准指数月度表现估值_清洗版", MONTHLY_FIELDS, cleaned_benchmark),
        ("额外指数_不纳入主分析", [
            "指数代码", "指数名称", "指数发布机构", "出现在哪些原始sheet", "不纳入原因",
        ], extra_rows),
        ("数据质量报告", [
            "类别", "指标", "数值/结果", "指数代码", "指数名称", "明细/说明", "备注",
        ], quality_rows),
    ]
    for name, headers, rows in specs:
        ws = wb.create_sheet(name)
        write_sheet(ws, headers, rows)
    format_workbook(wb)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    temp = OUTPUT_FILE.with_name(OUTPUT_FILE.stem + "_tmp.xlsx")
    wb.save(temp)
    check = load_workbook(temp, read_only=True, data_only=True)
    if check.sheetnames != [spec[0] for spec in specs]:
        raise RuntimeError("输出sheet结构校验失败")
    check.close()
    temp.replace(OUTPUT_FILE)
    changed = [str(path) for path, old_hash in hashes.items() if sha256(path) != old_hash]
    if changed:
        raise RuntimeError("输入文件被意外修改：" + "；".join(changed))

    print(f"核心策略ETF数量：{len(core_etfs)}")
    print(f"唯一核心指数数量：{len(core_codes)}")
    print(f"原始文件覆盖指数数量：{len(raw_core_coverage)}")
    print(f"补充文件覆盖指数数量：{len(sup_core_coverage)}")
    print(f"合并后覆盖指数数量：{len(final_coverage)}")
    print(f"缺失指数数量：{status_counts['缺失'] + status_counts['代码待核验']}")
    print(f"额外指数数量：{len(extra_codes)}")
    print(f"输出文件路径：{OUTPUT_FILE}")


if __name__ == "__main__":
    main()
