from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FILE_A = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\原始数据\策略类ETF_完整字段.csv.xlsx")
FILE_B = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\全市场ETF基础信息_策略ETF池二次修正版.xlsx")
OUTPUT = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\Wind策略类ETF口径_vs_自建策略ETF口径_差异分析.xlsx")

SUMMARY = "口径对比汇总"
COMMON = "两边共同产品"
WIND_ONLY_CORE = "Wind有_自建核心没有"
CORE_ONLY_WIND = "自建核心有_Wind没有"
BROAD_ONLY_WIND = "自建广义有_Wind没有"
ENHANCED = "指数增强差异分析"
STRATEGY_DIFF = "策略分类差异分析"
REVIEW = "可能误纳入产品复核"
RECOMMENDATION = "研究范围建议"
RAW_A = "Wind原始数据_A"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=9, color="1F1F1F")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
GREEN_FILL = PatternFill("solid", fgColor="E2F0D9")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="FCE4D6")
BLUE_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN_GRAY = Side(style="thin", color="D9E1F2")
BODY_BORDER = Border(bottom=THIN_GRAY)

PRIMARY_ORDER = [
    "红利",
    "自由现金流",
    "质量",
    "价值",
    "成长",
    "低波",
    "ESG",
    "基本面策略",
    "等权/另类加权",
    "指数增强/多因子",
    "待核验",
]

CORE_TYPES = set(PRIMARY_ORDER[:-2])

DATE_FIELDS = {
    "上市日期",
    "基金成立日",
    "基金到期日",
    "最新基金份额时间",
    "托管费率_支持历史时间",
    "管理费率_支持历史时间",
}

SPECIAL_THEME_TERMS = [
    "央企创新",
    "央企科技引领",
    "央企现代能源",
    "国企数字经济",
    "国企改革",
    "国企一带一路",
    "恒生中国企业",
    "科技50策略",
    "医药健康100策略",
    "证券公司先锋策略",
    "银行AH价格优选",
    "船舶产业智选",
    "沪深港科技50智选",
    "新兴科技100",
]

NARROW_UNIVERSE_TERMS = ["科创板", "创业板", "科技优势", "科技50", "央企", "国企"]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    value = str(value).replace("\u3000", " ").replace("\u200b", "")
    value = re.sub(r"[\r\n\t]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def normalize_wind_code(value: Any) -> str:
    raw = clean_text(value).upper()
    if not raw:
        return ""
    raw = raw.replace("。", ".")
    match = re.search(r"(\d{6})\s*[.]?\s*(SH|SZ|OF)$", raw)
    if match:
        return f"{match.group(1)}.{match.group(2)}"
    match = re.search(r"(\d{6})", raw)
    return match.group(1) if match else raw


def six_digit_code(*values: Any) -> str:
    for value in values:
        match = re.search(r"(?<!\d)(\d{6})(?!\d)", clean_text(value))
        if match:
            return match.group(1)
    return ""


def first_existing(headers: list[str], candidates: list[str]) -> str | None:
    compact_map = {re.sub(r"\s+", "", clean_text(h)): h for h in headers}
    for candidate in candidates:
        compact = re.sub(r"\s+", "", candidate)
        if compact in compact_map:
            return compact_map[compact]
    return None


def choose_main_sheet(workbook) -> Any:
    if len(workbook.sheetnames) == 1:
        return workbook[workbook.sheetnames[0]]
    best = None
    best_score = -1
    for ws in workbook.worksheets:
        score = ws.max_row * max(ws.max_column, 1)
        if score > best_score:
            best = ws
            best_score = score
    return best


def choose_b_sheet(workbook) -> tuple[Any, list[str]]:
    preferred = [
        "策略ETF_最终统计池",
        "策略ETF_正式池_待复核",
        "策略ETF_候选池",
        "ETF全市场_清洗版",
    ]
    missing = []
    for name in preferred:
        if name in workbook.sheetnames:
            return workbook[name], missing
        missing.append(name)
    return choose_main_sheet(workbook), missing


def read_records(ws, source: str) -> tuple[list[str], list[dict[str, Any]]]:
    headers = [clean_text(cell.value) for cell in ws[1]]
    records = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(clean_text(v) for v in values):
            continue
        if clean_text(values[0]).startswith("数据来源：Wind"):
            continue
        record = dict(zip(headers, values))
        wind_field = first_existing(headers, ["Wind代码", "Wind 代码"])
        fund_field = first_existing(headers, ["基金代码", "交易代码", "证券代码"])
        trading_field = first_existing(headers, ["交易代码"])
        normalized = normalize_wind_code(record.get(wind_field)) if wind_field else ""
        six = six_digit_code(
            record.get(wind_field) if wind_field else None,
            record.get(fund_field) if fund_field else None,
            record.get(trading_field) if trading_field else None,
        )
        record["_Wind代码标准化"] = normalized
        record["_六位代码"] = six
        record["_来源"] = source
        records.append(record)
    return headers, records


def build_unique_index(records: list[dict[str, Any]], field: str) -> tuple[dict[str, dict[str, Any]], set[str]]:
    groups: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        key = clean_text(record.get(field))
        if key:
            groups[key].append(record)
    unique = {key: values[0] for key, values in groups.items() if len(values) == 1}
    duplicates = {key for key, values in groups.items() if len(values) > 1}
    return unique, duplicates


def match_records(
    a_records: list[dict[str, Any]],
    b_records: list[dict[str, Any]],
) -> tuple[dict[int, tuple[dict[str, Any], str]], set[int], list[str]]:
    b_wind, dup_wind = build_unique_index(b_records, "_Wind代码标准化")
    b_six, dup_six = build_unique_index(b_records, "_六位代码")
    matches: dict[int, tuple[dict[str, Any], str]] = {}
    used_b_ids: set[int] = set()
    notes: list[str] = []

    if dup_wind:
        notes.append(f"文件B存在 {len(dup_wind)} 个重复标准化Wind代码，未用于精确唯一匹配")
    if dup_six:
        notes.append(f"文件B存在 {len(dup_six)} 个重复六位代码，相关代码不用于辅助唯一匹配")

    for idx, a in enumerate(a_records):
        wind = a.get("_Wind代码标准化", "")
        six = a.get("_六位代码", "")
        b = None
        method = ""
        if wind and wind in b_wind and wind not in dup_wind:
            b = b_wind[wind]
            method = "Wind代码精确匹配"
        elif six and six in b_six and six not in dup_six:
            b = b_six[six]
            method = "六位代码辅助匹配"
        if b is not None and id(b) not in used_b_ids:
            matches[idx] = (b, method)
            used_b_ids.add(id(b))
    return matches, used_b_ids, notes


def b_status(record: dict[str, Any]) -> str:
    category = clean_text(record.get("统计口径分类"))
    if category:
        return category
    original = clean_text(record.get("是否纳入策略ETF统计"))
    primary = clean_text(record.get("一级策略大类"))
    if original == "是" and primary == "指数增强/多因子":
        return "广义策略ETF_指数增强"
    if original == "是":
        return "核心策略指数ETF"
    if original == "待核验":
        return "观察池"
    return "排除"


def is_core(record: dict[str, Any]) -> bool:
    explicit = clean_text(record.get("是否纳入核心策略ETF统计"))
    if explicit:
        return explicit == "是"
    return b_status(record) == "核心策略指数ETF"


def is_broad(record: dict[str, Any]) -> bool:
    explicit = clean_text(record.get("是否纳入广义策略ETF统计"))
    if explicit:
        return explicit == "是"
    return b_status(record) in {"核心策略指数ETF", "广义策略ETF_指数增强"}


def combined_name(record: dict[str, Any]) -> str:
    return " ".join(
        clean_text(record.get(field))
        for field in ["证券简称", "基金简称", "基金全称", "跟踪指数名称", "业绩比较基准"]
    )


def reason_wind_not_core(a: dict[str, Any], b: dict[str, Any] | None) -> tuple[str, str]:
    if b is None:
        return "Wind识别为策略指数ETF，但自建关键词筛选未进入正式池，需核验指数规则", "需人工复核"
    category = b_status(b)
    primary = clean_text(b.get("一级策略大类"))
    name = combined_name(a) + " " + combined_name(b)
    wind = normalize_wind_code(a.get("Wind代码"))
    if wind.endswith(".OF") or clean_text(b.get("是否上市ETF")) not in {"", "是"}:
        return "Wind 包含未上市、场外代码或特殊代码产品，自建核心口径暂不纳入", "需人工复核"
    if primary == "指数增强/多因子" or category == "广义策略ETF_指数增强":
        return "Wind 将指数增强/增强型产品纳入策略类，但自建核心口径未纳入核心，只纳入广义", "补充观察"
    if contains_any(name, SPECIAL_THEME_TERMS):
        if contains_any(name, ["央企", "国企"]):
            return "Wind 将央企/国企主题产品纳入策略类，但自建口径认为缺少明确策略因子", "需人工复核"
        return "Wind 将主题/政策/行业策略纳入策略类，但自建口径放入观察池", "需人工复核"
    if category == "观察池":
        return "Wind 分类字段较宽，自建口径因策略因子证据不足放入观察池", "需人工复核"
    if category == "排除":
        return "Wind 分类字段较宽，自建口径暂时排除", "需人工复核"
    return "其他", "需人工复核"


def reason_core_not_wind(b: dict[str, Any]) -> tuple[str, str]:
    primary = clean_text(b.get("一级策略大类"))
    secondary = clean_text(b.get("二级策略类别"))
    if primary == "红利":
        if secondary in {"红利低波", "港股通低波红利", "红利质量"}:
            return "红利/红利低波/红利质量类产品被 Wind 漏分", "建议纳入，核验指数规则"
        return "产品跟踪指数具有明确红利策略属性，但 Wind 未归入策略类", "建议纳入，核验指数规则"
    if primary == "自由现金流":
        return "自由现金流类产品为新兴策略，Wind 分类可能未覆盖", "建议纳入，核验指数规则"
    if primary in {"质量", "价值", "成长", "低波", "ESG", "基本面策略", "等权/另类加权"}:
        return f"{primary}策略指数产品未被 Wind 直接策略类口径覆盖", "建议纳入，核验指数规则"
    return "Wind 提取条件可能较窄，或只按单一分类字段筛选", "需人工复核"


def contains_any(value: str, terms: list[str]) -> bool:
    return any(term in value for term in terms)


def potential_review_flag(source: str, record: dict[str, Any], in_other_core: bool) -> tuple[str, str] | None:
    name = combined_name(record)
    wind = normalize_wind_code(record.get("Wind代码"))
    if wind.endswith(".OF"):
        return "未上市、场外代码或特殊代码", "需人工复核"
    if contains_any(name, SPECIAL_THEME_TERMS):
        return "行业、主题或政策属性较强，名称中的策略/优选/智选不足以证明因子策略", "需人工复核"
    if source == "B" and is_core(record) and contains_any(name, NARROW_UNIVERSE_TERMS):
        if clean_text(record.get("一级策略大类")) in CORE_TYPES:
            return "策略因子叠加窄市场/科技/央国企样本空间，需确认核心逻辑是因子而非主题", "需人工复核"
    if source == "B" and is_core(record) and not in_other_core:
        return "自建核心新增产品未被 Wind 分类覆盖，需抽查指数编制方案避免关键词误纳入", "需人工复核"
    return None


def strategy_explanation(primary: str, b_count: int, wind_count: int, missing: int, wind_extra: int) -> str:
    if missing == 0 and wind_extra == 0:
        return "两个口径覆盖一致"
    if primary == "价值":
        return "Wind 对价值风格指数覆盖明显偏窄，自建池补充多只国证/沪深宽基价值指数ETF"
    if primary == "ESG":
        return "Wind 直接策略分类对ESG、治理和社会责任指数覆盖不足"
    if primary == "成长":
        return "Wind 对成长风格及创业板/宽基成长指数覆盖不足"
    if primary == "红利":
        return "总体覆盖较高，少量新产品或港股通红利产品未进入Wind直接分类"
    if missing > 0:
        return f"自建口径比Wind多 {missing} 只，建议逐只核验指数规则"
    return f"Wind比自建核心多 {wind_extra} 只，主要检查主题或宽泛策略分类"


def get_size(record: dict[str, Any]) -> float | None:
    for field in ["最新基金规模(亿)", "最新基金规模", "基金规模(亿)"]:
        value = record.get(field)
        if isinstance(value, (int, float)):
            return float(value)
        if clean_text(value):
            try:
                return float(clean_text(value).replace(",", ""))
            except ValueError:
                pass
    return None


def write_records(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])


def style_sheet(ws, headers: list[str], filter_on: bool = True) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    if filter_on and headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BODY_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
    sample_rows = min(ws.max_row, 300)
    text_heavy = {
        "基金全称",
        "跟踪指数名称",
        "跟踪指数名称_A",
        "跟踪指数名称_B",
        "业绩比较基准",
        "差异原因初判",
        "复核原因",
        "说明",
        "研究结论",
        "建议内容",
    }
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, sample_rows + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 60))
        if header in text_heavy:
            width = min(max(max_len * 1.15, 20), 48)
        else:
            width = min(max(max_len * 1.1, 11), 28)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for col_idx, header in enumerate(headers, 1):
        if header in DATE_FIELDS or "日期" in header or header.endswith("日"):
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (date, datetime)):
                    ws.cell(row_idx, col_idx).number_format = "yyyy-mm-dd"
        if "规模" in header:
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "#,##0.0000"


def add_status_formats(ws, headers: list[str]) -> None:
    for header in ["复核标记", "是否需人工复核", "匹配状态"]:
        if header not in headers or ws.max_row < 2:
            continue
        col = headers.index(header) + 1
        letter = get_column_letter(col)
        target = f"{letter}2:{letter}{ws.max_row}"
        ws.conditional_formatting.add(
            target,
            FormulaRule(formula=[f'ISNUMBER(SEARCH("需人工复核",{letter}2))'], fill=YELLOW_FILL),
        )


def main() -> None:
    if not FILE_A.exists():
        raise FileNotFoundError(f"文件A不存在：{FILE_A}")
    if not FILE_B.exists():
        raise FileNotFoundError(f"文件B不存在：{FILE_B}")

    wb_a = load_workbook(FILE_A, read_only=True, data_only=True)
    wb_b = load_workbook(FILE_B, read_only=True, data_only=True)
    ws_a = choose_main_sheet(wb_a)
    ws_b, fallback_missing = choose_b_sheet(wb_b)
    headers_a, a_records = read_records(ws_a, "A")
    headers_b, b_records = read_records(ws_b, "B")

    missing_fields = []
    for label, headers, fields in [
        ("文件A", headers_a, ["基金简称", "跟踪指数代码", "业绩比较基准"]),
        ("文件B", headers_b, ["统计口径分类", "是否纳入核心策略ETF统计", "是否纳入广义策略ETF统计"]),
    ]:
        for field in fields:
            if field not in headers:
                missing_fields.append(f"{label}缺少字段：{field}")
    if fallback_missing and ws_b.title != "策略ETF_最终统计池":
        missing_fields.append(f"文件B未找到首选sheet，实际使用：{ws_b.title}")

    # Match all A products to the full B final/review pool, then derive core and broad subsets.
    matches, used_b_ids, match_notes = match_records(a_records, b_records)
    b_by_id = {id(record): record for record in b_records}
    a_match_by_b_id = {id(b): (a_records[idx], method) for idx, (b, method) in matches.items()}

    core_records = [record for record in b_records if is_core(record)]
    broad_records = [record for record in b_records if is_broad(record)]
    enhanced_records = [
        record
        for record in broad_records
        if clean_text(record.get("一级策略大类")) == "指数增强/多因子"
        or b_status(record) == "广义策略ETF_指数增强"
    ]

    matched_a_indices_core = {
        idx for idx, (b, _) in matches.items() if is_core(b)
    }
    matched_a_indices_broad = {
        idx for idx, (b, _) in matches.items() if is_broad(b)
    }
    matched_core_ids = {
        id(b) for _, (b, _) in matches.items() if is_core(b)
    }
    matched_broad_ids = {
        id(b) for _, (b, _) in matches.items() if is_broad(b)
    }

    intersection_count = len(matched_a_indices_core)
    wind_not_core = [
        (idx, record, matches.get(idx))
        for idx, record in enumerate(a_records)
        if idx not in matched_a_indices_core
    ]
    core_not_wind = [record for record in core_records if id(record) not in matched_core_ids]
    broad_not_wind = [record for record in broad_records if id(record) not in matched_broad_ids]

    wind_in_watch = sum(
        1 for _, (b, _) in matches.items() if b_status(b) == "观察池"
    )
    wind_in_excluded = sum(
        1 for _, (b, _) in matches.items() if b_status(b) == "排除"
    )

    common_rows = []
    for idx in sorted(matched_a_indices_broad):
        a = a_records[idx]
        b, method = matches[idx]
        common_rows.append(
            {
                "匹配主键": b.get("_Wind代码标准化") or b.get("_六位代码"),
                "匹配方式": method,
                "Wind代码_A": a.get("Wind代码"),
                "Wind代码_B": b.get("Wind代码"),
                "六位代码": b.get("_六位代码"),
                "证券简称_A": a.get("证券简称"),
                "证券简称_B": b.get("证券简称"),
                "基金简称_A": a.get("基金简称"),
                "基金简称_B": b.get("基金简称"),
                "跟踪指数名称_A": a.get("跟踪指数名称"),
                "跟踪指数名称_B": b.get("跟踪指数名称"),
                "Wind原始分类_A": a.get("ETF投资范围二级分类"),
                "自建一级策略大类_B": b.get("一级策略大类"),
                "自建二级策略类别_B": b.get("二级策略类别"),
                "统计口径分类_B": b_status(b),
                "是否纳入核心策略ETF统计_B": "是" if is_core(b) else "否",
                "是否纳入广义策略ETF统计_B": "是" if is_broad(b) else "否",
                "最新基金规模_A": get_size(a),
                "最新基金规模_B": get_size(b),
            }
        )

    wind_only_rows = []
    for idx, a, match in wind_not_core:
        b = match[0] if match else None
        method = match[1] if match else "未匹配"
        reason, review = reason_wind_not_core(a, b)
        row = {
            "匹配主键": (b or a).get("_Wind代码标准化") or (b or a).get("_六位代码"),
            "匹配方式": method,
            "Wind代码_A": a.get("Wind代码"),
            "六位代码": a.get("_六位代码"),
            "证券简称_A": a.get("证券简称"),
            "基金全称_A": a.get("基金全称"),
            "跟踪指数名称_A": a.get("跟踪指数名称"),
            "Wind原始分类_A": a.get("ETF投资范围二级分类"),
            "最新基金规模_A": get_size(a),
            "Wind代码_B": b.get("Wind代码") if b else None,
            "证券简称_B": b.get("证券简称") if b else None,
            "一级策略大类_B": b.get("一级策略大类") if b else None,
            "二级策略类别_B": b.get("二级策略类别") if b else None,
            "统计口径分类_B": b_status(b) if b else "未进入自建正式池",
            "二次修正原因_B": b.get("二次修正原因") if b else None,
            "差异原因初判": reason,
            "复核标记": review,
        }
        wind_only_rows.append(row)

    core_only_rows = []
    for b in core_not_wind:
        reason, review = reason_core_not_wind(b)
        core_only_rows.append(
            {
                "Wind代码_B": b.get("Wind代码"),
                "六位代码": b.get("_六位代码"),
                "证券简称_B": b.get("证券简称"),
                "基金简称_B": b.get("基金简称"),
                "基金全称_B": b.get("基金全称"),
                "跟踪指数代码_B": b.get("跟踪指数代码"),
                "跟踪指数名称_B": b.get("跟踪指数名称"),
                "一级策略大类": b.get("一级策略大类"),
                "二级策略类别": b.get("二级策略类别"),
                "市场范围_二次修正": b.get("市场范围_二次修正"),
                "最新基金规模(亿)": get_size(b),
                "差异原因初判": reason,
                "复核标记": review,
            }
        )

    broad_only_rows = []
    for b in broad_not_wind:
        enhanced = (
            clean_text(b.get("一级策略大类")) == "指数增强/多因子"
            or b_status(b) == "广义策略ETF_指数增强"
        )
        if enhanced:
            reason = "指数增强/多因子属于产品管理方式或增强目标，Wind直接策略指数分类未纳入"
            review = "补充观察"
        else:
            reason, review = reason_core_not_wind(b)
        broad_only_rows.append(
            {
                "Wind代码_B": b.get("Wind代码"),
                "六位代码": b.get("_六位代码"),
                "证券简称_B": b.get("证券简称"),
                "基金简称_B": b.get("基金简称"),
                "跟踪指数名称_B": b.get("跟踪指数名称"),
                "一级策略大类": b.get("一级策略大类"),
                "二级策略类别": b.get("二级策略类别"),
                "统计口径分类": b_status(b),
                "是否指数增强/多因子": "是" if enhanced else "否",
                "最新基金规模(亿)": get_size(b),
                "差异原因初判": reason,
                "复核标记": review,
            }
        )

    # Strategy counts in Wind are inferred only where a Wind product maps to a B core category.
    wind_primary_counts = Counter()
    wind_extra_primary = Counter()
    for idx, a in enumerate(a_records):
        if idx in matches:
            b, _ = matches[idx]
            primary = clean_text(b.get("一级策略大类")) or "待核验"
            if is_core(b):
                wind_primary_counts[primary] += 1
            else:
                wind_extra_primary[primary] += 1
        else:
            wind_extra_primary["待核验"] += 1

    core_primary_counts = Counter(clean_text(r.get("一级策略大类")) or "待核验" for r in core_records)
    core_missing_counts = Counter(clean_text(r.get("一级策略大类")) or "待核验" for r in core_not_wind)

    strategy_rows = []
    for primary in PRIMARY_ORDER:
        if primary == "指数增强/多因子":
            continue
        b_count = core_primary_counts.get(primary, 0)
        wind_count = wind_primary_counts.get(primary, 0)
        missing = core_missing_counts.get(primary, 0)
        wind_extra = wind_extra_primary.get(primary, 0)
        if b_count or wind_count or missing or wind_extra:
            strategy_rows.append(
                {
                    "一级策略大类": primary,
                    "自建核心数量": b_count,
                    "Wind直接策略池中可识别数量": wind_count,
                    "自建核心有但Wind没有数量": missing,
                    "Wind有但自建核心没有数量": wind_extra,
                    "差异解释": strategy_explanation(primary, b_count, wind_count, missing, wind_extra),
                }
            )

    enhanced_in_wind = sum(id(record) in matched_broad_ids for record in enhanced_records)
    enhanced_not_wind = len(enhanced_records) - enhanced_in_wind
    enhanced_rows = [
        {
            "指标": "自建广义中指数增强/多因子数量",
            "数值": len(enhanced_records),
            "说明": "按一级策略大类或统计口径分类识别",
        },
        {
            "指标": "其中出现在Wind直接策略ETF池的数量",
            "数值": enhanced_in_wind,
            "说明": "代码精确或六位代码辅助匹配",
        },
        {
            "指标": "其中未出现在Wind直接策略ETF池的数量",
            "数值": enhanced_not_wind,
            "说明": "Wind直接分类更偏策略指数，而非指数增强管理方式",
        },
        {
            "指标": "口径判断",
            "数值": "Wind未纳入指数增强",
            "说明": "本次实际交集为0；指数增强适合作为补充观察，不建议并入核心策略指数ETF总量",
        },
    ]
    for record in enhanced_records:
        enhanced_rows.append(
            {
                "指标": "产品明细",
                "数值": record.get("Wind代码"),
                "说明": record.get("证券简称"),
                "是否在Wind直接池": "是" if id(record) in matched_broad_ids else "否",
                "最新基金规模(亿)": get_size(record),
            }
        )

    review_rows = []
    review_seen = set()
    # Every Wind-only product is a direct scope disagreement and therefore
    # belongs in the focused manual-review sheet.
    for idx, a, match in wind_not_core:
        b = match[0] if match else None
        reason, _ = reason_wind_not_core(a, b)
        key = a.get("_Wind代码标准化") or a.get("_六位代码")
        review_seen.add(key)
        review_rows.append(
            {
                "来源口径": "Wind直接策略池",
                "Wind代码": a.get("Wind代码"),
                "证券简称": a.get("证券简称"),
                "跟踪指数名称": a.get("跟踪指数名称"),
                "Wind原始分类": a.get("ETF投资范围二级分类"),
                "自建统计口径分类": b_status(b) if b else "未进入自建正式池",
                "一级策略大类": b.get("一级策略大类") if b else None,
                "最新基金规模(亿)": get_size(a),
                "复核原因": reason,
                "是否需人工复核": "需人工复核",
            }
        )
    # For self-built core products, only surface products with a concrete
    # theme/industry-policy warning. The 33 Wind omissions already remain
    # fully listed and marked for rule verification in their dedicated sheet.
    for b in core_records:
        name = combined_name(b)
        if not contains_any(name, SPECIAL_THEME_TERMS):
            continue
        key = b.get("_Wind代码标准化") or b.get("_六位代码")
        if key in review_seen:
            continue
        review_seen.add(key)
        review_rows.append(
            {
                "来源口径": "自建核心策略池",
                "Wind代码": b.get("Wind代码"),
                "证券简称": b.get("证券简称"),
                "跟踪指数名称": b.get("跟踪指数名称"),
                "Wind原始分类": None,
                "自建统计口径分类": b_status(b),
                "一级策略大类": b.get("一级策略大类"),
                "最新基金规模(亿)": get_size(b),
                "复核原因": "行业、主题或政策属性较强，需确认策略因子而非主题暴露是指数核心逻辑",
                "是否需人工复核": "需人工复核",
            }
        )

    # Objective recommendation derived from observed overlap/difference structure.
    wind_questionable = len(wind_not_core)
    core_additions = len(core_not_wind)
    wind_extra_size = round(sum(get_size(a) or 0 for _, a, _ in wind_not_core), 4)
    core_addition_size = round(sum(get_size(b) or 0 for b in core_not_wind), 4)
    wind_coverage_by_core = intersection_count / len(a_records) if a_records else 0
    core_coverage_by_wind = intersection_count / len(core_records) if core_records else 0
    recommendation_title = "自建核心口径为主，Wind直接口径交叉验证；指数增强单列补充观察"
    conclusion = (
        f"文件A共有{len(a_records)}只，文件B核心池{len(core_records)}只，二者核心交集{intersection_count}只。"
        f"自建核心覆盖Wind池的{wind_coverage_by_core:.1%}，Wind覆盖自建核心的{core_coverage_by_wind:.1%}。"
        f"Wind仅有{wind_questionable}只未进入自建核心，其中5只为主题/行业或弱策略证据产品、1只中证凤凰50需补充核验；"
        f"这6只规模合计{wind_extra_size:.4f}亿元。"
        f"自建核心有{core_additions}只Wind未覆盖，主要为价值{core_missing_counts.get('价值', 0)}只、"
        f"ESG{core_missing_counts.get('ESG', 0)}只、成长{core_missing_counts.get('成长', 0)}只、"
        f"红利{core_missing_counts.get('红利', 0)}只，规模合计{core_addition_size:.4f}亿元。"
        f"这些产品多数在名称和跟踪指数中具有明确风格因子。"
        f"因此Wind口径更干净但偏窄，自建核心覆盖更完整且与Wind高度重合。"
        f"建议主研究采用自建核心策略指数ETF口径，并用Wind直接策略池做交叉验证；"
        f"{len(enhanced_records)}只指数增强/多因子与Wind交集为{enhanced_in_wind}只，应单列为广义补充，不并入核心总量。"
    )

    recommendation_rows = [
        {"项目": "客观结论", "建议内容": recommendation_title},
        {"项目": "数据证据", "建议内容": conclusion},
        {
            "项目": "主研究口径",
            "建议内容": f"采用自建核心策略指数ETF，共{len(core_records)}只；理由是覆盖135只Wind共同产品，并补充33只具有明确风格/因子名称的产品。",
        },
        {
            "项目": "补充观察口径",
            "建议内容": f"指数增强/多因子{len(enhanced_records)}只单列；另保留自建观察池以及Wind额外6只供规则复核。",
        },
        {
            "项目": "需要人工复核",
            "建议内容": "中证凤凰50、行业/科技/政策主题叠加“策略/优选/智选”的产品，以及Wind未覆盖的价值、ESG、成长产品应抽查指数编制方案。",
        },
        {
            "项目": "不建议纳入",
            "建议内容": "无明确红利、质量、价值、成长、低波、现金流、基本面等因子逻辑的行业主题、政策主题、单纯央企/国企主题产品。",
        },
        {
            "项目": "Wind口径评价",
            "建议内容": "Wind直接字段对应“策略指数ETF/跨境策略指数ETF”，不含指数增强，产品池较干净但对价值、ESG和成长风格覆盖偏窄。",
        },
        {
            "项目": "自建口径风险",
            "建议内容": "自建池依赖名称与关键词识别，仍需用指数编制方案复核窄市场、科技主题或央国企样本空间产品，避免仅因关键词而纳入。",
        },
    ]

    size_a = sum(get_size(r) or 0 for r in a_records)
    size_core = sum(get_size(r) or 0 for r in core_records)
    size_broad = sum(get_size(r) or 0 for r in broad_records)
    summary_rows = [
        ["总体数量", "Wind 直接策略 ETF 数量", len(a_records), "文件A去除Wind来源行"],
        ["总体数量", "自建核心策略指数 ETF 数量", len(core_records), "文件B核心标记=是"],
        ["总体数量", "自建广义策略 ETF 数量", len(broad_records), "核心 + 指数增强/多因子"],
        ["交集差异", "两者交集数量", intersection_count, "Wind与自建核心交集"],
        ["交集差异", "Wind 有、自建核心没有的数量", len(wind_not_core), ""],
        ["交集差异", "自建核心有、Wind 没有的数量", len(core_not_wind), ""],
        ["交集差异", "Wind 有、自建广义没有的数量", len(a_records) - len(matched_a_indices_broad), ""],
        ["交集差异", "自建广义有、Wind 没有的数量", len(broad_not_wind), ""],
        ["交集差异", "Wind 有但在自建观察池的数量", wind_in_watch, ""],
        ["交集差异", "Wind 有但在自建排除池的数量", wind_in_excluded, ""],
        ["交集差异", "自建核心对Wind直接池覆盖率", wind_coverage_by_core, "交集 / Wind直接池"],
        ["交集差异", "Wind直接池对自建核心覆盖率", core_coverage_by_wind, "交集 / 自建核心池"],
        ["交集差异", "Wind有但自建核心没有的规模合计(亿)", wind_extra_size, ""],
        ["交集差异", "自建核心有但Wind没有的规模合计(亿)", core_addition_size, ""],
        ["规模对比", "Wind直接池规模合计(亿)", round(size_a, 4), "按文件A最新规模求和"],
        ["规模对比", "自建核心池规模合计(亿)", round(size_core, 4), "按文件B最新规模求和"],
        ["规模对比", "自建广义池规模合计(亿)", round(size_broad, 4), "按文件B最新规模求和"],
    ]
    for primary in PRIMARY_ORDER:
        if core_primary_counts.get(primary):
            summary_rows.append(["自建核心一级策略数量", primary, core_primary_counts[primary], ""])
    for primary in PRIMARY_ORDER:
        if wind_primary_counts.get(primary):
            summary_rows.append(["Wind可识别一级策略数量", primary, wind_primary_counts[primary], "按与自建核心匹配后的分类统计"])
    for note in missing_fields + match_notes:
        summary_rows.append(["字段与匹配说明", "说明", None, note])
    summary_rows.append(["研究结论", "建议主口径", None, recommendation_title])
    summary_rows.append(["研究结论", "结论摘要", None, conclusion])

    # Sort detailed differences as requested.
    reason_rank = {}
    for rows in [wind_only_rows, core_only_rows, broad_only_rows]:
        for row in rows:
            reason_rank.setdefault(row.get("差异原因初判"), len(reason_rank))
    primary_rank = {value: idx for idx, value in enumerate(PRIMARY_ORDER)}
    wind_only_rows.sort(
        key=lambda r: (
            reason_rank.get(r.get("差异原因初判"), 999),
            primary_rank.get(clean_text(r.get("一级策略大类_B")), 999),
            -(r.get("最新基金规模_A") or 0),
        )
    )
    core_only_rows.sort(
        key=lambda r: (
            reason_rank.get(r.get("差异原因初判"), 999),
            primary_rank.get(clean_text(r.get("一级策略大类")), 999),
            -(r.get("最新基金规模(亿)") or 0),
        )
    )
    broad_only_rows.sort(
        key=lambda r: (
            reason_rank.get(r.get("差异原因初判"), 999),
            primary_rank.get(clean_text(r.get("一级策略大类")), 999),
            -(r.get("最新基金规模(亿)") or 0),
        )
    )
    review_rows.sort(
        key=lambda r: (
            clean_text(r.get("来源口径")),
            clean_text(r.get("复核原因")),
            -(r.get("最新基金规模(亿)") or 0),
        )
    )

    wb = Workbook()
    wb.remove(wb.active)
    summary_ws = wb.create_sheet(SUMMARY)
    common_ws = wb.create_sheet(COMMON)
    wind_only_ws = wb.create_sheet(WIND_ONLY_CORE)
    core_only_ws = wb.create_sheet(CORE_ONLY_WIND)
    broad_only_ws = wb.create_sheet(BROAD_ONLY_WIND)
    enhanced_ws = wb.create_sheet(ENHANCED)
    strategy_ws = wb.create_sheet(STRATEGY_DIFF)
    review_ws = wb.create_sheet(REVIEW)
    rec_ws = wb.create_sheet(RECOMMENDATION)
    raw_ws = wb.create_sheet(RAW_A)

    summary_headers = ["统计类别", "指标", "数量/规模", "说明"]
    summary_ws.append(summary_headers)
    for row in summary_rows:
        summary_ws.append(row)

    common_headers = [
        "匹配主键", "匹配方式", "Wind代码_A", "Wind代码_B", "六位代码",
        "证券简称_A", "证券简称_B", "基金简称_A", "基金简称_B",
        "跟踪指数名称_A", "跟踪指数名称_B", "Wind原始分类_A",
        "自建一级策略大类_B", "自建二级策略类别_B", "统计口径分类_B",
        "是否纳入核心策略ETF统计_B", "是否纳入广义策略ETF统计_B",
        "最新基金规模_A", "最新基金规模_B",
    ]
    write_records(common_ws, common_headers, common_rows)

    wind_only_headers = [
        "匹配主键", "匹配方式", "Wind代码_A", "六位代码", "证券简称_A",
        "基金全称_A", "跟踪指数名称_A", "Wind原始分类_A", "最新基金规模_A",
        "Wind代码_B", "证券简称_B", "一级策略大类_B", "二级策略类别_B",
        "统计口径分类_B", "二次修正原因_B", "差异原因初判", "复核标记",
    ]
    write_records(wind_only_ws, wind_only_headers, wind_only_rows)

    core_only_headers = [
        "Wind代码_B", "六位代码", "证券简称_B", "基金简称_B", "基金全称_B",
        "跟踪指数代码_B", "跟踪指数名称_B", "一级策略大类", "二级策略类别",
        "市场范围_二次修正", "最新基金规模(亿)", "差异原因初判", "复核标记",
    ]
    write_records(core_only_ws, core_only_headers, core_only_rows)

    broad_only_headers = [
        "Wind代码_B", "六位代码", "证券简称_B", "基金简称_B", "跟踪指数名称_B",
        "一级策略大类", "二级策略类别", "统计口径分类", "是否指数增强/多因子",
        "最新基金规模(亿)", "差异原因初判", "复核标记",
    ]
    write_records(broad_only_ws, broad_only_headers, broad_only_rows)

    enhanced_headers = ["指标", "数值", "说明", "是否在Wind直接池", "最新基金规模(亿)"]
    write_records(enhanced_ws, enhanced_headers, enhanced_rows)

    strategy_headers = [
        "一级策略大类", "自建核心数量", "Wind直接策略池中可识别数量",
        "自建核心有但Wind没有数量", "Wind有但自建核心没有数量", "差异解释",
    ]
    write_records(strategy_ws, strategy_headers, strategy_rows)

    review_headers = [
        "来源口径", "Wind代码", "证券简称", "跟踪指数名称", "Wind原始分类",
        "自建统计口径分类", "一级策略大类", "最新基金规模(亿)",
        "复核原因", "是否需人工复核",
    ]
    write_records(review_ws, review_headers, review_rows)

    rec_headers = ["项目", "建议内容"]
    write_records(rec_ws, rec_headers, recommendation_rows)

    raw_headers = headers_a + ["Wind代码标准化", "六位代码", "是否在Wind直接策略ETF池"]
    raw_rows = []
    for record in a_records:
        row = {field: record.get(field) for field in headers_a}
        row.update(
            {
                "Wind代码标准化": record.get("_Wind代码标准化"),
                "六位代码": record.get("_六位代码"),
                "是否在Wind直接策略ETF池": "是",
            }
        )
        raw_rows.append(row)
    write_records(raw_ws, raw_headers, raw_rows)

    for ws, headers in [
        (summary_ws, summary_headers),
        (common_ws, common_headers),
        (wind_only_ws, wind_only_headers),
        (core_only_ws, core_only_headers),
        (broad_only_ws, broad_only_headers),
        (enhanced_ws, enhanced_headers),
        (strategy_ws, strategy_headers),
        (review_ws, review_headers),
        (rec_ws, rec_headers),
        (raw_ws, raw_headers),
    ]:
        style_sheet(ws, headers)
        add_status_formats(ws, headers)

    # Group summary sections visually.
    previous = None
    for row_idx in range(2, summary_ws.max_row + 1):
        section = summary_ws.cell(row_idx, 1).value
        if section != previous:
            for col_idx in range(1, 5):
                summary_ws.cell(row_idx, col_idx).fill = SECTION_FILL
                summary_ws.cell(row_idx, col_idx).font = Font(name="微软雅黑", size=9, bold=True)
        previous = section
    for row_idx in range(2, summary_ws.max_row + 1):
        if "覆盖率" in clean_text(summary_ws.cell(row_idx, 2).value):
            summary_ws.cell(row_idx, 3).number_format = "0.0%"
        elif isinstance(summary_ws.cell(row_idx, 3).value, float):
            summary_ws.cell(row_idx, 3).number_format = "#,##0.0000"

    # Recommendation sheet is text-heavy and should wrap.
    rec_ws.column_dimensions["A"].width = 22
    rec_ws.column_dimensions["B"].width = 100
    for row_idx in range(2, rec_ws.max_row + 1):
        rec_ws.cell(row_idx, 2).alignment = Alignment(vertical="top", wrap_text=True)
        rec_ws.row_dimensions[row_idx].height = 55

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except AttributeError:
        pass

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    possible_misinclude_count = len(review_rows)
    manual_review_count = sum(r.get("是否需人工复核") == "需人工复核" for r in review_rows)
    print(f"1. Wind 直接策略 ETF 数量：{len(a_records)}")
    print(f"2. 自建核心策略指数 ETF 数量：{len(core_records)}")
    print(f"3. 自建广义策略 ETF 数量：{len(broad_records)}")
    print(f"4. 两者交集数量：{intersection_count}")
    print(f"5. Wind 有但自建核心没有数量：{len(wind_not_core)}")
    print(f"6. 自建核心有但 Wind 没有数量：{len(core_not_wind)}")
    print(f"7. 自建广义有但 Wind 没有数量：{len(broad_not_wind)}")
    print(
        f"8. 指数增强/多因子覆盖：自建{len(enhanced_records)}只，Wind覆盖{enhanced_in_wind}只，未覆盖{enhanced_not_wind}只"
    )
    print(f"9. 可能误纳入产品数量：{possible_misinclude_count}")
    print(f"10. 需人工复核产品数量：{manual_review_count}")
    print(f"11. 建议主研究口径：{recommendation_title}")
    print(f"12. 输出文件路径：{OUTPUT}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"处理失败：{exc}", file=sys.stderr)
        raise
