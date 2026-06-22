"""创建境内策略 ETF 产品梳理 Excel 初版模板。"""

from pathlib import Path
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEETS = {
    "产品总表": [
        "策略大类",
        "策略细分",
        "产品名称",
        "基金代码",
        "基金公司",
        "上市交易所",
        "成立日期",
        "上市日期",
        "跟踪指数",
        "指数公司",
        "最新规模",
        "规模日期",
        "近一年规模变化",
        "日均成交额",
        "成交额统计区间",
        "管理费率",
        "托管费率",
        "是否代表产品",
        "代表性理由",
        "备注",
        "信息来源",
    ],
    "策略分类表": [
        "策略类型",
        "核心因子",
        "投资逻辑",
        "代表指数",
        "代表ETF",
        "适合市场环境",
        "主要风险",
        "境内发展成熟度",
        "备注",
    ],
    "指数规则表": [
        "策略类型",
        "指数名称",
        "指数代码",
        "指数公司",
        "样本空间",
        "选样因子",
        "选样方法",
        "成分股数量",
        "加权方式",
        "调样频率",
        "个股权重限制",
        "行业限制",
        "前十大成分股",
        "行业分布",
        "策略理解",
        "信息来源",
    ],
}


def display_width(text: str) -> int:
    """按中英文字符显示宽度估算 Excel 列宽。"""
    return sum(2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1 for char in text)


def style_worksheet(ws, headers: list[str]) -> None:
    ws.append(headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.sheet_view.zoomScale = 90

    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(bottom=Side(style="thin", color="A6A6A6"))

    for column_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=column_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border

        # 当前模板只有表头，因此按表头显示长度自动估算并留出筛选按钮空间。
        width = min(max(display_width(header) + 4, 12), 28)
        ws.column_dimensions[get_column_letter(column_index)].width = width

    ws.row_dimensions[1].height = 26


def create_template(output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, headers in SHEETS.items():
        worksheet = workbook.create_sheet(title=sheet_name)
        style_worksheet(worksheet, headers)

    workbook.save(output_path)


def validate_template(output_path: Path) -> None:
    workbook = load_workbook(output_path)
    if workbook.sheetnames != list(SHEETS):
        raise ValueError(f"工作表不匹配：{workbook.sheetnames}")

    for sheet_name, expected_headers in SHEETS.items():
        worksheet = workbook[sheet_name]
        actual_headers = [cell.value for cell in worksheet[1]]
        expected_filter = f"A1:{get_column_letter(len(expected_headers))}1"

        if actual_headers != expected_headers:
            raise ValueError(f"{sheet_name} 表头不匹配")
        if worksheet.max_row != 1:
            raise ValueError(f"{sheet_name} 除表头外存在数据")
        if worksheet.freeze_panes != "A2":
            raise ValueError(f"{sheet_name} 未冻结首行")
        if worksheet.auto_filter.ref != expected_filter:
            raise ValueError(f"{sheet_name} 筛选范围不正确")
        if not all(cell.font.bold for cell in worksheet[1]):
            raise ValueError(f"{sheet_name} 表头未全部加粗")
        if not all(cell.fill.fgColor.rgb == "00D9D9D9" for cell in worksheet[1]):
            raise ValueError(f"{sheet_name} 表头底色不正确")

        print(
            f"{sheet_name}: {len(actual_headers)} fields, "
            f"freeze={worksheet.freeze_panes}, filter={worksheet.auto_filter.ref}"
        )


if __name__ == "__main__":
    project_root = Path(__file__).resolve().parents[1]
    target = project_root / "02_outputs" / "excel" / "境内策略ETF产品梳理_初版.xlsx"
    create_template(target)
    validate_template(target)
    print(f"Excel template created: {target}")
