"""Step 6：复核策略分类、生成修正版并同步主 Excel。"""

from __future__ import annotations

import csv
import io
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
ATTACHMENT = Path(
    r"C:\Users\lvdon\.codex\attachments\dab15e6e-0e6f-415a-9bdf-b648f11fbcfa\pasted-text.txt"
)
INPUT_CSV = PROJECT_ROOT / "01_processed_data" / "product_pool" / "境内策略ETF产品池_清洗版.csv"
INPUT_EXCEL = PROJECT_ROOT / "01_processed_data" / "product_pool" / "境内策略ETF产品池_清洗版.xlsx"
MAIN_EXCEL = PROJECT_ROOT / "02_outputs" / "excel" / "境内策略ETF产品梳理_初版.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "01_processed_data" / "classification"
OUTPUT_CSV = OUTPUT_DIR / "境内策略ETF产品池_分类修正表.csv"
OUTPUT_EXCEL = OUTPUT_DIR / "境内策略ETF产品池_分类修正版.xlsx"
REVIEW_MD = OUTPUT_DIR / "step6_classification_review.md"
LOG_MD = OUTPUT_DIR / "step6_classification_log.md"

BASE_FIELDS = [
    "策略大类",
    "策略细分",
    "基金代码",
    "产品名称",
    "基金公司",
    "上市交易所",
    "跟踪指数",
    "指数公司",
    "纳入级别",
    "是否主线",
    "是否补充观察",
    "是否待校验",
    "待校验事项",
    "是否代表产品初筛",
    "代表产品初筛理由",
    "信息来源备注",
    "清洗备注",
]

REVIEW_FIELDS = [
    "原策略大类",
    "原策略细分",
    "原纳入级别",
    "原指数公司",
    "建议策略大类",
    "建议策略细分",
    "建议纳入级别",
    "建议指数公司",
    "是否需要修改",
    "修改类型",
    "修改原因",
    "后续是否需要查指数规则",
    "分类修正备注",
]

OUTPUT_FIELDS = [*BASE_FIELDS, *REVIEW_FIELDS]

CORRECTION_FIELDS = [
    "基金代码",
    "产品名称",
    "建议策略大类",
    "建议策略细分",
    "建议纳入级别",
    "建议指数公司",
    "修改类型",
    "修改原因",
    "后续是否需要查指数规则",
    "分类修正备注",
]


def strip_value(value: object) -> str:
    return "" if value is None else str(value).strip(" \t\r\n")


def product_key(record: dict[str, str]) -> tuple[str, str]:
    return record["基金代码"], record["产品名称"]


def read_product_pool() -> tuple[list[dict[str, str]], str]:
    if INPUT_CSV.exists():
        with INPUT_CSV.open("r", encoding="utf-8-sig", newline="") as csv_file:
            reader = csv.DictReader(csv_file)
            if reader.fieldnames != BASE_FIELDS:
                raise ValueError(f"清洗版 CSV 字段不匹配：{reader.fieldnames}")
            rows = [
                {field: strip_value(record.get(field, "")) for field in BASE_FIELDS}
                for record in reader
                if any(strip_value(value) for value in record.values())
            ]
        return rows, str(INPUT_CSV)

    if INPUT_EXCEL.exists():
        workbook = load_workbook(INPUT_EXCEL, data_only=False)
        if "产品池_清洗版" not in workbook.sheetnames:
            raise ValueError("清洗版 Excel 中不存在 Sheet【产品池_清洗版】")
        worksheet = workbook["产品池_清洗版"]
        headers = [strip_value(cell.value) for cell in worksheet[1]]
        if headers != BASE_FIELDS:
            raise ValueError(f"清洗版 Excel 字段不匹配：{headers}")
        rows = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            record = {field: strip_value(value) for field, value in zip(BASE_FIELDS, values)}
            if any(record.values()):
                rows.append(record)
        return rows, str(INPUT_EXCEL)

    raise FileNotFoundError(f"缺少输入文件：{INPUT_CSV}；备用文件也不存在：{INPUT_EXCEL}")


def read_corrections() -> list[dict[str, str]]:
    text = ATTACHMENT.read_text(encoding="utf-8-sig")
    header = "\t".join(CORRECTION_FIELDS)
    start = text.find(header)
    if start < 0:
        raise ValueError("附件中未找到分类修正 TSV 表头")
    end_marker = "\n\n五、分类修正版 Excel 要求"
    end = text.find(end_marker, start)
    if end < 0:
        raise ValueError("附件中未找到分类修正 TSV 结束位置")

    block = text[start:end].strip("\r\n")
    reader = csv.DictReader(io.StringIO(block), delimiter="\t")
    if reader.fieldnames != CORRECTION_FIELDS:
        raise ValueError(f"分类修正 TSV 字段不匹配：{reader.fieldnames}")
    return [
        {field: strip_value(record.get(field, "")) for field in CORRECTION_FIELDS}
        for record in reader
    ]


def apply_corrections(
    products: list[dict[str, str]], corrections: list[dict[str, str]]
) -> tuple[
    list[dict[str, str]],
    list[dict[str, str]],
    list[dict[str, str]],
    list[tuple[str, str]],
    list[tuple[str, str]],
]:
    product_counts = Counter(product_key(record) for record in products)
    correction_by_key = {product_key(record): record for record in corrections}
    correction_keys = set(correction_by_key)

    unmatched = [key for key in correction_keys if product_counts[key] == 0]
    ambiguous = [key for key in correction_keys if product_counts[key] > 1]
    usable_keys = correction_keys - set(unmatched) - set(ambiguous)

    reviewed_rows: list[dict[str, str]] = []
    matched_rows_by_key: dict[tuple[str, str], dict[str, str]] = {}
    for product in products:
        key = product_key(product)
        correction = correction_by_key.get(key) if key in usable_keys else None

        if correction:
            suggestions = {
                "建议策略大类": correction["建议策略大类"],
                "建议策略细分": correction["建议策略细分"],
                "建议纳入级别": correction["建议纳入级别"],
                "建议指数公司": correction["建议指数公司"],
                "是否需要修改": "是",
                "修改类型": correction["修改类型"],
                "修改原因": correction["修改原因"],
                "后续是否需要查指数规则": correction["后续是否需要查指数规则"],
                "分类修正备注": correction["分类修正备注"],
            }
        else:
            suggestions = {
                "建议策略大类": product["策略大类"],
                "建议策略细分": product["策略细分"],
                "建议纳入级别": product["纳入级别"],
                "建议指数公司": product["指数公司"],
                "是否需要修改": "否",
                "修改类型": "",
                "修改原因": "",
                "后续是否需要查指数规则": "是" if product["是否待校验"] == "是" else "否",
                "分类修正备注": "",
            }

        reviewed = {
            **product,
            "原策略大类": product["策略大类"],
            "原策略细分": product["策略细分"],
            "原纳入级别": product["纳入级别"],
            "原指数公司": product["指数公司"],
            **suggestions,
        }
        reviewed_rows.append(reviewed)
        if correction:
            matched_rows_by_key[key] = reviewed

    matched_rows = [
        matched_rows_by_key[product_key(correction)]
        for correction in corrections
        if product_key(correction) in matched_rows_by_key
    ]
    modification_rows = [row for row in reviewed_rows if row["是否需要修改"] == "是"]
    rule_check_rows = [
        row for row in reviewed_rows if row["后续是否需要查指数规则"] == "是"
    ]
    return reviewed_rows, matched_rows, modification_rows, rule_check_rows, unmatched, ambiguous


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=OUTPUT_FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def display_width(value: object) -> int:
    text = strip_value(value)
    return sum(
        2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1
        for char in text
    )


def append_rows(worksheet, fields: list[str], rows: list[dict[str, object]]) -> None:
    worksheet.append(fields)
    for record in rows:
        worksheet.append([record.get(field, "") for field in fields])


def format_sheet(worksheet) -> None:
    max_row = max(worksheet.max_row, 1)
    max_column = max(worksheet.max_column, 1)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_column)}{max_row}"
    worksheet.sheet_view.zoomScale = 85

    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    header_font = Font(name="Microsoft YaHei", size=10, bold=True)
    body_font = Font(name="Microsoft YaHei", size=10)
    border = Border(bottom=Side(style="thin", color="A6A6A6"))

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    worksheet.row_dimensions[1].height = 28

    headers = [cell.value for cell in worksheet[1]]
    for row in worksheet.iter_rows(min_row=2, max_row=max_row, max_col=max_column):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    long_text_headers = {
        "跟踪指数",
        "待校验事项",
        "信息来源备注",
        "清洗备注",
        "修改原因",
        "分类修正备注",
        "代表产品初筛理由",
    }
    for column_index, header in enumerate(headers, start=1):
        values = [worksheet.cell(row, column_index).value for row in range(1, max_row + 1)]
        measured = max(display_width(value) for value in values) + 3
        cap = 48 if header in long_text_headers else 28
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(measured, 12), cap
        )
        if header == "基金代码":
            for row in range(2, max_row + 1):
                worksheet.cell(row, column_index).number_format = "@"


def build_statistics(rows: list[dict[str, str]]) -> list[dict[str, object]]:
    strategy_counts = Counter(row["建议策略大类"] for row in rows)
    level_counts = Counter(row["建议纳入级别"] for row in rows)
    cross_counts = Counter((row["建议策略大类"], row["建议纳入级别"]) for row in rows)
    statistics: list[dict[str, object]] = []

    for strategy, count in sorted(strategy_counts.items()):
        statistics.append(
            {"统计类型": "按建议策略大类", "建议策略大类": strategy, "建议纳入级别": "", "产品数量": count}
        )
    for level, count in sorted(level_counts.items()):
        statistics.append(
            {"统计类型": "按建议纳入级别", "建议策略大类": "", "建议纳入级别": level, "产品数量": count}
        )
    for (strategy, level), count in sorted(cross_counts.items()):
        statistics.append(
            {"统计类型": "建议策略大类+建议纳入级别", "建议策略大类": strategy, "建议纳入级别": level, "产品数量": count}
        )

    summary_items = [
        ("需要修改产品", sum(row["是否需要修改"] == "是" for row in rows)),
        ("后续需要查指数规则", sum(row["后续是否需要查指数规则"] == "是" for row in rows)),
        ("主线", level_counts["主线"]),
        ("补充观察", level_counts["补充观察"]),
        ("跨境补充", level_counts["跨境补充"]),
        ("单列观察", level_counts["单列观察"]),
        ("仅作线索", level_counts["仅作线索"]),
    ]
    for label, count in summary_items:
        statistics.append(
            {"统计类型": "汇总指标", "建议策略大类": label, "建议纳入级别": "", "产品数量": count}
        )
    return statistics


def write_excel(
    all_rows: list[dict[str, str]],
    matched_rows: list[dict[str, str]],
    modification_rows: list[dict[str, str]],
    rule_check_rows: list[dict[str, str]],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    sheets = [
        ("产品池_分类修正版", all_rows),
        ("分类修正表", matched_rows),
        ("需要修改清单", modification_rows),
        ("后续查指数规则清单", rule_check_rows),
    ]
    for name, rows in sheets:
        worksheet = workbook.create_sheet(name)
        append_rows(worksheet, OUTPUT_FIELDS, rows)
        format_sheet(worksheet)

    stats_sheet = workbook.create_sheet("分类统计")
    stats_fields = ["统计类型", "建议策略大类", "建议纳入级别", "产品数量"]
    append_rows(stats_sheet, stats_fields, build_statistics(all_rows))
    format_sheet(stats_sheet)
    workbook.save(OUTPUT_EXCEL)


def sync_main_excel(rows: list[dict[str, str]]) -> tuple[list[tuple[str, str]], list[tuple[str, str]]]:
    if not MAIN_EXCEL.exists():
        raise FileNotFoundError(f"缺少主 Excel：{MAIN_EXCEL}")
    workbook = load_workbook(MAIN_EXCEL)
    if "产品总表" not in workbook.sheetnames:
        raise ValueError("主 Excel 中不存在 Sheet【产品总表】")

    worksheet = workbook["产品总表"]
    headers = [strip_value(cell.value) for cell in worksheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}
    required = {"策略大类", "策略细分", "基金代码", "产品名称", "指数公司", "备注"}
    missing_headers = required - set(headers)
    if missing_headers:
        raise ValueError(f"主 Excel 缺少字段：{sorted(missing_headers)}")

    main_rows_by_key: dict[tuple[str, str], list[int]] = defaultdict(list)
    for row_number in range(2, worksheet.max_row + 1):
        key = (
            strip_value(worksheet.cell(row_number, columns["基金代码"]).value),
            strip_value(worksheet.cell(row_number, columns["产品名称"]).value),
        )
        main_rows_by_key[key].append(row_number)

    missing_products: list[tuple[str, str]] = []
    ambiguous_products: list[tuple[str, str]] = []
    for record in rows:
        key = product_key(record)
        positions = main_rows_by_key.get(key, [])
        if not positions:
            missing_products.append(key)
            continue
        if len(positions) > 1:
            ambiguous_products.append(key)
            continue

        row_number = positions[0]
        worksheet.cell(row_number, columns["策略大类"], record["建议策略大类"])
        worksheet.cell(row_number, columns["策略细分"], record["建议策略细分"])
        worksheet.cell(row_number, columns["指数公司"], record["建议指数公司"])

        correction_note = record["分类修正备注"]
        if correction_note:
            note_cell = worksheet.cell(row_number, columns["备注"])
            existing_note = strip_value(note_cell.value)
            appended_note = f"Step 6 分类修正备注：{correction_note}"
            if appended_note not in existing_note:
                note_cell.value = existing_note + ("\n" if existing_note else "") + appended_note
            note_cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{worksheet.max_row}"
    workbook.save(MAIN_EXCEL)
    return missing_products, ambiguous_products


def write_review() -> None:
    content = """# Step 6 策略分类检查说明

## 1. Step 6 目标

基于 Step 5 清洗版产品池和给定分类修正表，对策略分类边界进行复核，形成可供 Step 7 代表产品筛选使用的分类修正版。全程不新增产品、不联网补充事实。

## 2. 本轮分类检查的主要原则

- 保留分类逻辑清晰、边界明确的原有类别。
- 对复合因子、指数增强、央国企主题与策略交叉、港股通跨境产品进行单列或归并。
- 分类修正与事实信息分离，不修改基金代码、产品名称、基金公司、上市交易所、跟踪指数和信息来源。
- 对仍需确认的边界问题统一标记为后续需要查指数规则。

## 3. 保留原分类的类别

红利、红利低波、自由现金流、低波、价值、成长、基本面策略等类别整体保持原分类。分类合理但指数公司尚需补充的产品，仅调整建议指数公司。

## 4. 本轮建议修正的类别

- 将红利质量产品从纯质量类调整为“红利质量”复合策略。
- 将指数增强产品调整为“指数增强/多因子”，维持单列观察。
- 将央企股东回报、央企红利和国企红利统一归入“央国企红利/股东回报”。
- 将港股通红利和港股通低波红利统一归入“港股通红利/低波”跨境补充池。

本轮主要对 Step 5 的机械分类结果进行策略逻辑复核。整体来看，红利、红利低波、自由现金流、低波、价值、成长、基本面策略等主线分类基本合理。需要修正的是几类边界产品：第一，红利质量类产品不宜简单归为纯质量，应作为“红利+质量”复合策略单列；第二，指数增强ETF虽然使用多因子模型，但不同于纯被动策略指数ETF，建议单列观察；第三，央企股东回报、央企红利、国企红利兼具主题属性和策略属性，建议统一归入“央国企红利/股东回报”补充观察池；第四，港股通红利及低波红利产品应统一作为跨境补充，不混入境内A股策略ETF主线。

## 5. 需要和 mentor 确认的问题

- 指数增强 ETF 是否纳入策略 ETF 主线，还是仅作为单列观察。
- 央企股东回报、央企红利和国企红利是否进入红利主线。
- 港股通策略产品是否纳入境内部分，还是仅用于后续境外对比。
- 红利质量复合策略在最终框架中单列，还是归入多因子类。

## 6. 后续 Step 7 代表产品筛选的建议衔接

优先在各建议策略大类中选择分类清晰、主线属性明确且指数规则可获得的产品；对指数增强、央国企红利/股东回报及港股通产品，待 mentor 明确研究边界后再决定是否进入代表产品名单。
"""
    REVIEW_MD.write_text(content, encoding="utf-8")


def format_key_list(keys: Iterable[tuple[str, str]]) -> str:
    items = list(keys)
    if not items:
        return "无"
    return "\n".join(f"- {code}｜{name}" for code, name in items)


def write_log(
    input_path: str,
    all_rows: list[dict[str, str]],
    corrections: list[dict[str, str]],
    matched_rows: list[dict[str, str]],
    unmatched: list[tuple[str, str]],
    ambiguous: list[tuple[str, str]],
    main_missing: list[tuple[str, str]],
    main_ambiguous: list[tuple[str, str]],
) -> None:
    strategy_counts = Counter(row["建议策略大类"] for row in all_rows)
    level_counts = Counter(row["建议纳入级别"] for row in all_rows)
    rule_count = sum(row["后续是否需要查指数规则"] == "是" for row in all_rows)
    strategy_lines = "\n".join(
        f"- {strategy}：{count}" for strategy, count in sorted(strategy_counts.items())
    )
    level_lines = "\n".join(
        f"- {level}：{count}" for level, count in sorted(level_counts.items())
    )

    content = f"""# Step 6 分类修正日志

1. 输入文件路径：`{input_path}`
2. 输出文件路径：
   - 分类修正表 CSV：`{OUTPUT_CSV}`
   - 分类修正版 Excel：`{OUTPUT_EXCEL}`
   - 分类检查说明：`{REVIEW_MD}`
   - 同步主 Excel：`{MAIN_EXCEL}`
3. 原始产品数量：{len(all_rows)}
4. DATA_BLOCK 修正产品数量：{len(corrections)}
5. 成功匹配并修正产品数量：{len(matched_rows)}
6. 未匹配到的 DATA_BLOCK 产品：
{format_key_list(unmatched)}
   - 无法唯一匹配的 DATA_BLOCK 产品：
{format_key_list(ambiguous)}
7. 是否成功生成分类修正版 Excel：是
8. 是否成功同步更新主 Excel：是
   - 主 Excel 中未找到的产品：
{format_key_list(main_missing)}
   - 主 Excel 中无法唯一匹配的产品：
{format_key_list(main_ambiguous)}
9. 按建议策略大类统计产品数量：
{strategy_lines}
10. 按建议纳入级别统计产品数量：
{level_lines}
11. 后续需要查指数规则产品数量：{rule_count}
"""
    LOG_MD.write_text(content, encoding="utf-8")


def validate(
    all_rows: list[dict[str, str]],
    matched_rows: list[dict[str, str]],
    modification_rows: list[dict[str, str]],
    rule_check_rows: list[dict[str, str]],
) -> None:
    if OUTPUT_CSV.read_bytes()[:3] != b"\xef\xbb\xbf":
        raise ValueError("分类修正 CSV 不是 UTF-8 with BOM")
    with OUTPUT_CSV.open("r", encoding="utf-8-sig", newline="") as csv_file:
        loaded = list(csv.DictReader(csv_file))
    if loaded != all_rows:
        raise ValueError("分类修正 CSV 回读内容不一致")

    workbook = load_workbook(OUTPUT_EXCEL)
    expected_sheets = [
        "产品池_分类修正版",
        "分类修正表",
        "需要修改清单",
        "后续查指数规则清单",
        "分类统计",
    ]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(f"分类修正版 Excel Sheet 不匹配：{workbook.sheetnames}")
    expected_counts = {
        "产品池_分类修正版": len(all_rows),
        "分类修正表": len(matched_rows),
        "需要修改清单": len(modification_rows),
        "后续查指数规则清单": len(rule_check_rows),
    }
    for sheet_name, count in expected_counts.items():
        if workbook[sheet_name].max_row != count + 1:
            raise ValueError(f"{sheet_name} 行数不匹配")
    for worksheet in workbook.worksheets:
        if worksheet.freeze_panes != "A2" or not worksheet.auto_filter.ref:
            raise ValueError(f"{worksheet.title} 的冻结窗格或筛选未生效")

    main_workbook = load_workbook(MAIN_EXCEL)
    main_sheet = main_workbook["产品总表"]
    headers = [strip_value(cell.value) for cell in main_sheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}
    main_map = {
        (
            strip_value(main_sheet.cell(row, columns["基金代码"]).value),
            strip_value(main_sheet.cell(row, columns["产品名称"]).value),
        ): row
        for row in range(2, main_sheet.max_row + 1)
    }
    for record in matched_rows:
        key = product_key(record)
        if key not in main_map:
            continue
        row = main_map[key]
        if strip_value(main_sheet.cell(row, columns["策略大类"]).value) != record["建议策略大类"]:
            raise ValueError(f"主 Excel 策略大类同步失败：{key}")
        if strip_value(main_sheet.cell(row, columns["指数公司"]).value) != record["建议指数公司"]:
            raise ValueError(f"主 Excel 指数公司同步失败：{key}")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    products, input_path = read_product_pool()
    corrections = read_corrections()
    (
        all_rows,
        matched_rows,
        modification_rows,
        rule_check_rows,
        unmatched,
        ambiguous,
    ) = apply_corrections(products, corrections)

    write_csv(OUTPUT_CSV, all_rows)
    write_excel(all_rows, matched_rows, modification_rows, rule_check_rows)
    write_review()
    main_missing, main_ambiguous = sync_main_excel(all_rows)
    validate(all_rows, matched_rows, modification_rows, rule_check_rows)
    write_log(
        input_path,
        all_rows,
        corrections,
        matched_rows,
        unmatched,
        ambiguous,
        main_missing,
        main_ambiguous,
    )

    print(f"products={len(all_rows)}")
    print(f"corrections={len(corrections)}")
    print(f"matched={len(matched_rows)}")
    print(f"unmatched={len(unmatched)}")
    print(f"ambiguous={len(ambiguous)}")
    print(f"rule_checks={len(rule_check_rows)}")
    print(f"main_missing={len(main_missing)}")
    print(f"main_ambiguous={len(main_ambiguous)}")
    print("strategy_counts=" + repr(dict(sorted(Counter(r['建议策略大类'] for r in all_rows).items()))))


if __name__ == "__main__":
    main()
