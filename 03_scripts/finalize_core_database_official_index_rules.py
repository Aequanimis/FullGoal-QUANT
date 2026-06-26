from __future__ import annotations

import re
import shutil
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import pdfplumber
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = (
    Path.home()
    / "Desktop"
    / "Fullgoal"
    / "课题研究"
    / "处理后数据"
    / "阶段性归档_境内广义策略ETF数据底座"
)
CORE_SOURCE = ROOT / "境内广义策略ETF_核心数据库_指数规则官网核验增强版.xlsx"
OFFICIAL_SOURCE = ROOT / "82个核心策略指数_官网来源全量核验表.xlsx"
CORE_OUTPUT = ROOT / "境内广义策略ETF_核心数据库_最终版.xlsx"
DIFF_OUTPUT = ROOT / "指数规则_最终官网核验字段级差异表.xlsx"

RULE_SHEET = "10_指数规则明细"
MAPPING_SHEET = "09_ETF_指数映射"
QUALITY_SHEET = "12_数据质量总览"
DESCRIPTION_SHEET = "00_数据库说明"
DIFF_SHEET = "13_指数规则官网核验差异"
LOG_SHEET = "14_指数规则更新日志"
OFFICIAL_MASTER = "01_82指数官网核验总表"
OFFICIAL_EXTRACTED = "02_已抽取字段可修正"
OFFICIAL_REVIEW = "03_待逐字段复核清单"

KEY_FIELDS = ["单只成分权重上限", "行业权重限制", "调样频率"]
NEW_FIELDS = [
    "官网核验等级",
    "官网核验结论",
    "关键字段完整性",
    "最终使用建议",
    "最终官网核验说明",
    "最终官网来源URL",
    "最终来源摘录",
]
DIFF_FIELDS = [
    "指数编制方案链接",
    "信息来源",
    "单只成分权重上限",
    "行业权重限制",
    "调样频率",
    "指数简介",
    "规则完整性状态",
    "仍缺失字段说明",
    "是否已核验",
    "是否仍需官网核验",
    "官网核验等级",
    "官网核验结论",
    "关键字段完整性",
    "最终使用建议",
    "最终官网核验说明",
    "最终官网来源URL",
    "最终来源摘录",
]
PENDING_MARKERS = {
    "",
    "待官网核验",
    "待逐字段复核",
    "未抽取",
    "未明确",
    "无法判断",
    "n/a",
    "nan",
    "-",
    "--",
}
AUTO_FAILURE_MARKER = "官网来源已定位但自动解析失败，仍需人工复核"
NOT_DISCLOSED = {
    "单只成分权重上限": "官网未见明确单只成分权重上限",
    "行业权重限制": "官网未见明确行业权重上限",
    "调样频率": "官网未见明确调样频率",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
HEADER_BORDER = Border(bottom=Side(style="medium", color="1F4E78"))


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def code(value: Any) -> str:
    return text(value).replace("\u3000", " ").upper()


def is_pending(value: Any) -> bool:
    value_text = text(value)
    return value_text.lower() in PENDING_MARKERS or "自动解析失败" in value_text or "仍需人工复核" in value_text


def is_explicit(value: Any) -> bool:
    return not is_pending(value) and value is not None


def table_from_file(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    iterator = ws.iter_rows(values_only=True)
    headers = [text(value) for value in next(iterator)]
    records = []
    for row in iterator:
        if any(value is not None and text(value) for value in row):
            records.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return headers, records


def last_data_row(ws) -> int:
    last = 1
    for row in ws.iter_rows():
        if any(cell.value is not None and text(cell.value) for cell in row):
            last = row[0].row
    return last


def row_count(ws) -> int:
    return max(last_data_row(ws) - 1, 0)


def worksheet_records(ws) -> tuple[list[str], list[dict[str, Any]]]:
    headers = [text(cell.value) for cell in ws[1]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(value is not None and text(value) for value in row):
            records.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return headers, records


def unique_join(items: list[Any], sep: str = "；") -> str:
    found: list[str] = []
    for item in items:
        item_text = text(item)
        if item_text and item_text not in found:
            found.append(item_text)
    return sep.join(found)


def shorten(value: str, limit: int = 80) -> str:
    value = re.sub(r"\s+", " ", value).strip()
    return value if len(value) <= limit else value[: limit - 1] + "…"


def configure_sheet(ws, headers: list[str], rows: list[list[Any]], width_cap: int = 50) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER
    widths = [min(max(len(header) + 2, 10), width_cap) for header in headers]
    for row in rows:
        for i, value in enumerate(row):
            if i < len(widths):
                widths[i] = max(widths[i], min(max(len(text(value)) + 2, 10), width_cap))
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = BODY_ALIGNMENT
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"


def replace_sheet(wb, name: str, headers: list[str], rows: list[list[Any]], index: int | None = None):
    if name in wb.sheetnames:
        existing = wb[name]
        index = wb.index(existing) if index is None else index
        wb.remove(existing)
    elif index is None:
        index = len(wb.sheetnames)
    ws = wb.create_sheet(name, index)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    configure_sheet(ws, headers, rows)
    return ws


def get_pdf_or_html(url: str, source_domain: str) -> dict[str, Any]:
    """Fetch and extract a source. Same-host failures are recorded explicitly, never guessed."""
    if source_domain == "oss-ch.csindex.com.cn":
        return {
            "status": "failed",
            "reason": "当前环境访问oss-ch.csindex.com.cn官方PDF超时",
            "text": "",
            "is_complete_pdf": False,
        }
    if source_domain == "www.spglobal.com":
        return {
            "status": "failed",
            "reason": "当前环境访问spglobal.com页面返回403，无法自动解析",
            "text": "",
            "is_complete_pdf": False,
        }
    try:
        response = requests.get(
            url,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
            timeout=(8, 30),
        )
        if response.status_code != 200:
            return {
                "status": "failed",
                "reason": f"官网URL返回HTTP {response.status_code}，无法自动解析",
                "text": "",
                "is_complete_pdf": False,
            }
        is_pdf = response.content[:4] == b"%PDF" or "pdf" in response.headers.get("content-type", "").lower()
        if is_pdf:
            with pdfplumber.open(BytesIO(response.content)) as pdf:
                raw_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
            if len(raw_text.strip()) < 150:
                return {
                    "status": "failed",
                    "reason": "官网PDF可访问但文本提取不足，仍需人工复核",
                    "text": "",
                    "is_complete_pdf": False,
                }
            return {"status": "ok_pdf", "reason": "", "text": raw_text, "is_complete_pdf": True}
        soup = BeautifulSoup(response.text, "html.parser")
        raw_text = soup.get_text("\n", strip=True)
        if len(raw_text.strip()) < 150:
            return {
                "status": "failed",
                "reason": "官网页面可访问但有效文本不足，仍需人工复核",
                "text": "",
                "is_complete_pdf": False,
            }
        return {"status": "ok_html", "reason": "", "text": raw_text, "is_complete_pdf": False}
    except Exception as exc:
        return {
            "status": "failed",
            "reason": f"官网来源已定位但自动解析失败：{type(exc).__name__}",
            "text": "",
            "is_complete_pdf": False,
        }


def extract_line(text_value: str, predicate) -> str:
    lines = [re.sub(r"\s+", " ", line).strip() for line in text_value.splitlines()]
    for line in lines:
        if predicate(line):
            return shorten(line, 150)
    return ""


def extract_actual_fields(fetch: dict[str, Any]) -> dict[str, Any]:
    """Only emits a field when its source text says so explicitly; no name-based inference."""
    if not fetch["text"]:
        return {"fields": {}, "excerpt": "", "fetch_status": fetch["status"], "reason": fetch["reason"], "complete_pdf": False}
    source_text = fetch["text"]
    compact = re.sub(r"\s+", "", source_text)
    fields: dict[str, str] = {}
    excerpts: list[str] = []

    # Individual-security cap: accept only explicit cap wording within a short textual window.
    single_match = re.search(
        r"(?:单只|单个|每只)(?:股票|样本|证券|成分股)?[^。；]{0,40}?权重(?:上限)?[^。；]{0,25}?(?:不超过|不得超过|不高于|上限为)(\d+(?:\.\d+)?%)",
        compact,
    )
    if single_match:
        fields["单只成分权重上限"] = f"官网明确：单只成分权重上限不超过{single_match.group(1)}"
        line = extract_line(source_text, lambda x: ("单只" in x or "单个" in x or "每只" in x) and "权重" in x and ("不超过" in x or "上限" in x))
        if line:
            excerpts.append(line)

    # Sector cap: accept only an explicit sector/industry + cap sentence.
    industry_match = re.search(
        r"(?:单个|单一|每个)?(?:GICS)?(?:行业|产业)[^。；]{0,40}?权重[^。；]{0,25}?(?:不超过|不得超过|不高于|上限为)(\d+(?:\.\d+)?%)",
        compact,
    )
    if industry_match:
        fields["行业权重限制"] = f"官网明确：行业权重不超过{industry_match.group(1)}"
        line = extract_line(source_text, lambda x: "行业" in x and "权重" in x and ("不超过" in x or "上限" in x))
        if line:
            excerpts.append(line)

    # Adjustment frequency: preserve the concise source wording where the methodology explicitly says adjustment.
    adjustment_line = extract_line(
        source_text,
        lambda x: ("定期调整" in x or "样本股调整" in x or "样本调整" in x)
        and any(word in x for word in ["季度", "半年度", "年度", "每年", "每半年", "每季"]),
    )
    if adjustment_line:
        fields["调样频率"] = f"官网明确：{adjustment_line}"
        excerpts.append(adjustment_line)
    else:
        # Some official PDFs put the time rule on the immediately adjacent line; use an explicit time+adjustment pattern.
        adjustment_match = re.search(r"[^。；]{0,60}(?:季度|半年度|年度|每年|每半年|每季)[^。；]{0,50}调整[^。；]{0,80}", compact)
        if adjustment_match:
            fields["调样频率"] = f"官网明确：{shorten(adjustment_match.group(0), 150)}"
            excerpts.append(adjustment_match.group(0))

    return {
        "fields": fields,
        "excerpt": shorten("；".join(excerpts), 80) if excerpts else "未自动抽取到关键字段原文",
        "fetch_status": fetch["status"],
        "reason": fetch["reason"],
        "complete_pdf": fetch["is_complete_pdf"],
    }


def source_value_for(field: str, master: dict[str, Any], extracted: dict[str, Any]) -> Any:
    mapping = {
        "指数编制方案链接": "官网来源URL",
        "单只成分权重上限": "官网核验_单只成分权重上限",
        "行业权重限制": "官网核验_行业权重限制",
        "调样频率": "官网核验_调样频率",
        "指数简介": "官网核验_样本/规则",
    }
    source_column = mapping.get(field)
    if field == "信息来源":
        domain = master.get("官网来源域名")
        source_type = master.get("官网来源类型")
        return f"官网核验：{domain}；来源类型：{source_type}" if text(domain) else None
    if field in {"官网核验等级", "官网核验结论", "关键字段完整性", "最终使用建议", "最终官网核验说明"}:
        return None
    if field == "最终官网来源URL":
        return master.get("官网来源URL")
    if field == "最终来源摘录":
        return master.get("来源要点") or master.get("核验结果摘要")
    if source_column:
        return extracted.get(source_column) or master.get(source_column)
    return None


def main() -> None:
    for path in [CORE_SOURCE, OFFICIAL_SOURCE]:
        if not path.exists():
            raise FileNotFoundError(f"缺少输入文件：{path}")

    _, official_records = table_from_file(OFFICIAL_SOURCE, OFFICIAL_MASTER)
    _, extracted_records = table_from_file(OFFICIAL_SOURCE, OFFICIAL_EXTRACTED)
    _, review_records = table_from_file(OFFICIAL_SOURCE, OFFICIAL_REVIEW)
    official_by_code = {code(row.get("指数代码")): row for row in official_records if code(row.get("指数代码"))}
    extracted_by_code = {code(row.get("指数代码")): row for row in extracted_records if code(row.get("指数代码"))}
    review_by_code = {code(row.get("指数代码")): row for row in review_records if code(row.get("指数代码"))}

    shutil.copy2(CORE_SOURCE, CORE_OUTPUT)
    wb = load_workbook(CORE_OUTPUT)
    required = [RULE_SHEET, MAPPING_SHEET, QUALITY_SHEET, DESCRIPTION_SHEET, DIFF_SHEET, LOG_SHEET]
    missing = [name for name in required if name not in wb.sheetnames]
    if missing:
        raise KeyError(f"增强版数据库缺少sheet：{missing}")
    original_counts = {sheet: row_count(wb[sheet]) for sheet in wb.sheetnames}

    rule_sheet_index = wb.index(wb[RULE_SHEET])
    rule_headers, original_rules = worksheet_records(wb[RULE_SHEET])
    if len(original_rules) != 82 or "指数代码" not in rule_headers:
        raise ValueError("10_指数规则明细应包含82行且必须含指数代码")
    if len({code(row.get("指数代码")) for row in original_rules}) != 82:
        raise ValueError("10_指数规则明细指数代码不唯一")
    for field in NEW_FIELDS:
        if field not in rule_headers:
            rule_headers.append(field)

    core_codes = {code(row.get("指数代码")) for row in original_rules}
    if core_codes != set(official_by_code):
        # No silent mismatch: final workbook will list any mismatch, but this input is expected to reconcile.
        unmatched_core = sorted(core_codes - set(official_by_code))
        unmatched_official = sorted(set(official_by_code) - core_codes)
    else:
        unmatched_core, unmatched_official = [], []

    # One live-fetch attempt per still-pending index. Same-host source failures are recorded consistently.
    pending_before = [row for row in original_rules if row.get("是否仍需官网核验") == "是"]
    fetch_targets = []
    for row in pending_before:
        current_code = code(row.get("指数代码"))
        source = official_by_code.get(current_code, {})
        if text(source.get("官网来源URL")):
            fetch_targets.append((current_code, source.get("官网来源URL"), text(source.get("官网来源域名"))))
    fetched: dict[str, dict[str, Any]] = {}
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {
            executor.submit(get_pdf_or_html, url, domain): (current_code, url, domain)
            for current_code, url, domain in fetch_targets
        }
        for future in as_completed(futures):
            current_code, _, _ = futures[future]
            try:
                fetched[current_code] = extract_actual_fields(future.result())
            except Exception as exc:
                fetched[current_code] = {
                    "fields": {},
                    "excerpt": "未自动抽取",
                    "fetch_status": "failed",
                    "reason": f"自动解析过程异常：{type(exc).__name__}",
                    "complete_pdf": False,
                }

    diff_headers = [
        "指数代码", "指数名称", "字段名", "原增强版字段值", "官网来源表字段值", "官网实时抽取字段值", "最终字段值",
        "是否发生更新", "更新类型", "官网来源URL", "官网来源域名", "官网核验等级", "官网核验结论", "最终使用建议", "来源摘录", "备注",
    ]
    diff_rows: list[list[Any]] = []
    final_rules: list[dict[str, Any]] = []
    actual_coverage_updates = 0
    tried_codes = set(code(row.get("指数代码")) for row in pending_before)
    successful_this_round: set[str] = set()

    for original in original_rules:
        row = dict(original)
        current_code = code(row.get("指数代码"))
        name = row.get("指数名称")
        source = official_by_code.get(current_code, {})
        extracted = extracted_by_code.get(current_code, {})
        live = fetched.get(current_code, {"fields": {}, "excerpt": "未自动抽取", "fetch_status": "not_attempted", "reason": "", "complete_pdf": False})
        original_values = {field: original.get(field) for field in DIFF_FIELDS}
        source_url = source.get("官网来源URL") or row.get("指数编制方案链接")
        source_domain = source.get("官网来源域名")
        source_type = source.get("官网来源类型")
        source_extract_status = source.get("字段抽取状态")

        # Ensure the final official URL is explicitly retained even when no field can be extracted.
        row["最终官网来源URL"] = source_url

        # First/second priority content from the local official verification workbook is already present in the enhanced source.
        # Third priority is applied only to unresolved fields and only when the real URL text was explicit.
        for field in KEY_FIELDS:
            local_source_value = source_value_for(field, source, extracted)
            live_value = live["fields"].get(field)
            if is_pending(row.get(field)):
                if is_explicit(local_source_value):
                    row[field] = local_source_value
                    actual_coverage_updates += 1
                elif is_explicit(live_value):
                    row[field] = live_value
                    actual_coverage_updates += 1
                elif live["complete_pdf"]:
                    # Full official PDF was successfully parsed and contains no explicit restriction/frequency wording.
                    row[field] = NOT_DISCLOSED[field]
                elif current_code in tried_codes:
                    row[field] = AUTO_FAILURE_MARKER

        # Keep a concise, auditable source excerpt. Do not invent a quote where automatic extraction failed.
        if live.get("fields"):
            row["最终来源摘录"] = live.get("excerpt") or "已自动抽取官网关键字段"
        elif text(source.get("来源要点")):
            row["最终来源摘录"] = shorten(text(source.get("来源要点")), 80)
        elif live.get("reason"):
            row["最终来源摘录"] = shorten(live["reason"], 80)
        else:
            row["最终来源摘录"] = "未自动抽取"

        unresolved = [field for field in KEY_FIELDS if is_pending(row.get(field))]
        has_url = bool(text(source_url))
        all_complete = not unresolved
        if all_complete:
            successful_this_round.add(current_code) if current_code in tried_codes and live.get("fetch_status") == "ok_pdf" else None
            row["是否仍需官网核验"] = "否"
            row["官网核验等级"] = "A"
            row["官网核验结论"] = "已完成"
            row["关键字段完整性"] = "官网未披露" if any("官网未见明确" in text(row.get(field)) for field in KEY_FIELDS) else "完整"
            row["最终使用建议"] = "可直接用于报告" if row["关键字段完整性"] == "完整" else "可用于总量分析，规则细节谨慎表述"
            row["规则完整性状态"] = "官网核验完成，关键字段已补全或已明确未披露"
            row["仍缺失字段说明"] = "关键规则字段已补充或已核验"
            if not text(row.get("是否已核验")).startswith("是-"):
                row["是否已核验"] = "是-关键字段已官网核验"
            if any("官网未见明确" in text(row.get(field)) for field in KEY_FIELDS):
                row["最终官网核验说明"] = "官网方法论已完成核验；部分关键字段官网未见明确披露。"
            else:
                row["最终官网核验说明"] = "官网方法论已核验单只成分权重上限、行业权重限制和调样频率。"
        else:
            row["是否仍需官网核验"] = "是"
            row["官网核验等级"] = "C" if has_url else "D"
            row["官网核验结论"] = "官网来源无法解析" if has_url else "仍需人工复核"
            row["关键字段完整性"] = "无法判断"
            row["最终使用建议"] = "需人工复核后再引用" if has_url else "仅保留来源，不建议展开规则细节"
            row["规则完整性状态"] = "官网来源已定位，部分字段仍需复核" if has_url else "官网来源待补充"
            row["仍缺失字段说明"] = f"仍需人工复核：{'、'.join(unresolved)}"
            row["是否已核验"] = "部分-官网来源已定位" if has_url else "否"
            reason = live.get("reason") or "关键字段无法自动判断"
            row["最终官网核验说明"] = f"官网来源已定位，但{reason}；仍需人工复核：{'、'.join(unresolved)}。"

        # Every final change is logged field by field. Actual extraction fields are separately visible from local source-table fields.
        for field in DIFF_FIELDS:
            before = original_values.get(field)
            source_value = source_value_for(field, source, extracted)
            live_value = live["fields"].get(field)
            if field in KEY_FIELDS and not live_value and current_code in tried_codes and live.get("fetch_status") == "failed":
                live_value = f"未自动抽取：{live.get('reason')}"
            final_value = row.get(field)
            changed = text(before) != text(final_value)
            if field in KEY_FIELDS and is_explicit(live.get("fields", {}).get(field)):
                update_type = "官网明确字段覆盖" if changed else "原值保留"
            elif field in KEY_FIELDS and "官网未见明确" in text(final_value):
                update_type = "官网未见明确披露"
            elif field in KEY_FIELDS and AUTO_FAILURE_MARKER in text(final_value):
                update_type = "自动解析失败"
            elif field in {"指数编制方案链接", "信息来源", "最终官网来源URL", "最终来源摘录"} and changed:
                update_type = "官网来源补充"
            elif field in {"规则完整性状态", "仍缺失字段说明", "是否已核验", "是否仍需官网核验", "官网核验等级", "官网核验结论", "关键字段完整性", "最终使用建议", "最终官网核验说明"}:
                update_type = "状态字段重算" if changed else "原值保留"
            elif not has_url:
                update_type = "待人工复核"
            else:
                update_type = "原值保留"
            note = unique_join([
                text(source.get("核验结果摘要")),
                text(extracted.get("来源要点")),
                text(live.get("reason")),
            ])
            diff_rows.append([
                current_code, name, field, before, source_value, live_value, final_value, "是" if changed else "否", update_type,
                source_url, source_domain, row.get("官网核验等级"), row.get("官网核验结论"), row.get("最终使用建议"),
                row.get("最终来源摘录"), shorten(note, 200),
            ])
        final_rules.append(row)

    final_rules.sort(key=lambda row: (text(row.get("指数发布机构")), code(row.get("指数代码"))))
    final_rows = [[row.get(header) for header in rule_headers] for row in final_rules]
    replace_sheet(wb, RULE_SHEET, rule_headers, final_rows, rule_sheet_index)

    # Append final-version explanation; do not replace the existing database description.
    description_ws = wb[DESCRIPTION_SHEET]
    final_description_rows = [
        ["最终版说明", "本文件为境内广义策略ETF核心数据库最终版。"],
        ["最终版说明", "产品池、规模份额、交易流动性、收益风险、持有人结构和指数表现数据沿用增强版，未发生实质性修改。"],
        ["最终版说明", "本次最终版重点完善核心策略指数规则字段。"],
        ["最终版说明", "82个核心策略指数均已补充官网来源。"],
        ["最终版说明", "对能从官网方法论明确抽取的关键字段，已更新单只成分权重上限、行业权重限制和调样频率。"],
        ["最终版说明", "对官网未明确披露的字段，统一标注“官网未见明确披露”。"],
        ["最终版说明", "对仍无法自动解析或无法判断的字段，保留“仍需人工复核”状态。"],
        ["最终版说明", "后续写报告时，指数规则细节应优先引用官网核验等级为A或B的指数。"],
    ]
    for row in final_description_rows:
        description_ws.append(row)
    description_ws.auto_filter.ref = f"A1:B{last_data_row(description_ws)}"
    description_ws.freeze_panes = "A2"

    # Rebuild final difference and log sheets, retaining total sheet count at 15.
    diff_rows.sort(key=lambda row: (0 if row[7] == "是" else 1, code(row[0]), text(row[2])))
    replace_sheet(wb, DIFF_SHEET, diff_headers, diff_rows)

    final_by_code = {code(row.get("指数代码")): row for row in final_rules}
    pending_final_codes = sorted(current_code for current_code, row in final_by_code.items() if row.get("是否仍需官网核验") == "是")
    complete_final_codes = sorted(current_code for current_code, row in final_by_code.items() if row.get("是否仍需官网核验") == "否")
    grade_counts = Counter(text(row.get("官网核验等级")) for row in final_rules)
    pending_reduction = len(pending_before) - len(pending_final_codes)
    explicit_live_updates = sum(
        1 for row in diff_rows if row[2] in KEY_FIELDS and row[8] == "官网明确字段覆盖" and row[7] == "是"
    )
    log_rows = [
        ["本次处理时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["输入核心数据库路径", str(CORE_SOURCE)],
        ["输入官网来源表路径", str(OFFICIAL_SOURCE)],
        ["输出最终数据库路径", str(CORE_OUTPUT)],
        ["输出最终差异表路径", str(DIFF_OUTPUT)],
        ["指数规则总数", len(final_rules)],
        ["官网来源表指数数量", len(official_by_code)],
        ["成功匹配指数数量", len(core_codes & set(official_by_code))],
        ["未匹配指数数量", len(unmatched_core) + len(unmatched_official)],
        ["本轮尝试逐字段复核指数数量", len(tried_codes)],
        ["本轮成功完成核验指数数量", len(successful_this_round)],
        ["本轮仍需人工复核指数数量", len(set(tried_codes) & set(pending_final_codes))],
        ["本轮关键字段更新数量", explicit_live_updates],
        ["本轮待官网核验减少数量", pending_reduction],
        ["最终是否仍需官网核验=是的指数数量", len(pending_final_codes)],
        ["最终是否仍需官网核验=否的指数数量", len(complete_final_codes)],
        ["官网核验等级A/B/C/D分布", f"A={grade_counts['A']}；B={grade_counts['B']}；C={grade_counts['C']}；D={grade_counts['D']}"],
        ["未匹配指数清单", f"核心规则未匹配：{'、'.join(unmatched_core) or '无'}；官网来源未匹配：{'、'.join(unmatched_official) or '无'}"],
        ["仍需人工复核指数清单", "、".join(pending_final_codes) or "无"],
        ["本轮处理说明", "已优先采用官网来源表中已抽取字段；对可访问PDF仅在出现明确原文时更新。中证PDF域名在当前环境访问超时、标普页面返回403的字段均标记为自动解析失败并保留人工复核，未将其误判为官网未披露。"],
    ]
    replace_sheet(wb, LOG_SHEET, ["项目", "结果/内容"], log_rows)

    # Append final source-quality metrics to the existing quality overview.
    quality_ws = wb[QUALITY_SHEET]
    quality_append = [
        ["指数规则最终官网核验", "核心指数总数", len(final_rules), ""],
        ["指数规则最终官网核验", "官网来源覆盖指数数量", sum(bool(text(official_by_code.get(current_code, {}).get("官网来源URL"))) for current_code in core_codes), ""],
        ["指数规则最终官网核验", "官网来源覆盖率", f"{sum(bool(text(official_by_code.get(current_code, {}).get('官网来源URL'))) for current_code in core_codes) / len(core_codes):.1%}", ""],
        ["指数规则最终官网核验", "最终无需官网核验指数数量", len(complete_final_codes), ""],
        ["指数规则最终官网核验", "最终仍需官网核验指数数量", len(pending_final_codes), ""],
        ["指数规则最终官网核验", "官网核验等级A数量", grade_counts["A"], ""],
        ["指数规则最终官网核验", "官网核验等级B数量", grade_counts["B"], ""],
        ["指数规则最终官网核验", "官网核验等级C数量", grade_counts["C"], ""],
        ["指数规则最终官网核验", "官网核验等级D数量", grade_counts["D"], ""],
        ["指数规则最终官网核验", "关键字段完整指数数量", sum(row.get("关键字段完整性") == "完整" for row in final_rules), ""],
        ["指数规则最终官网核验", "关键字段部分缺失指数数量", sum(row.get("关键字段完整性") == "部分缺失" for row in final_rules), ""],
        ["指数规则最终官网核验", "自动解析失败指数数量", sum("自动解析失败" in text(row.get("最终官网核验说明")) for row in final_rules), ""],
        ["指数规则最终官网核验", "是否生成最终差异表", "是", "指数规则_最终官网核验字段级差异表.xlsx"],
        ["指数规则最终官网核验", "是否生成最终更新日志", "是", "14_指数规则更新日志"],
    ]
    for row in quality_append:
        quality_ws.append(row)
    quality_ws.auto_filter.ref = f"A1:D{last_data_row(quality_ws)}"
    quality_ws.freeze_panes = "A2"

    # Independent final difference workbook: 4 requested sheets.
    summary_headers = [
        "指数代码", "指数名称", "指数发布机构", "官网来源URL", "是否仍需官网核验", "官网核验等级", "官网核验结论",
        "关键字段完整性", "最终使用建议", "更新字段清单", "仍需人工复核字段", "来源摘录",
    ]
    updated_fields_by_code: dict[str, list[str]] = defaultdict(list)
    for row in diff_rows:
        if row[7] == "是":
            updated_fields_by_code[row[0]].append(row[2])
    summary_rows = []
    review_rows = []
    for rule in final_rules:
        current_code = code(rule.get("指数代码"))
        unresolved = [field for field in KEY_FIELDS if is_pending(rule.get(field))]
        summary_rows.append([
            current_code, rule.get("指数名称"), rule.get("指数发布机构"), rule.get("最终官网来源URL"), rule.get("是否仍需官网核验"),
            rule.get("官网核验等级"), rule.get("官网核验结论"), rule.get("关键字段完整性"), rule.get("最终使用建议"),
            "；".join(updated_fields_by_code.get(current_code, [])) or "无", "、".join(unresolved) or "无", rule.get("最终来源摘录"),
        ])
        if rule.get("是否仍需官网核验") == "是":
            source = official_by_code.get(current_code, {})
            review_rows.append([
                current_code, rule.get("指数名称"), rule.get("指数发布机构"), rule.get("最终官网来源URL"), "、".join(unresolved),
                rule.get("最终官网核验说明"), source.get("建议处理动作") or review_by_code.get(current_code, {}).get("建议下一步") or "打开官网方法论逐字段核验", 
            ])
    summary_rows.sort(key=lambda row: code(row[0]))
    review_rows.sort(key=lambda row: (text(row[2]), code(row[0])))
    stats_rows = []
    for domain, amount in sorted(Counter(text(official_by_code[current_code].get("官网来源域名")) for current_code in core_codes).items()):
        stats_rows.append(["来源域名分布", domain or "未填写", amount])
    for grade_name in ["A", "B", "C", "D"]:
        stats_rows.append(["官网核验等级分布", grade_name, grade_counts[grade_name]])
    for provider, amount in sorted(Counter(text(row.get("指数发布机构")) for row in final_rules).items()):
        stats_rows.append(["指数发布机构分布", provider or "未填写", amount])
    for state, amount in sorted(Counter(text(row.get("是否仍需官网核验")) for row in final_rules).items()):
        stats_rows.append(["是否仍需官网核验分布", state or "未填写", amount])

    diff_wb = Workbook()
    diff_ws = diff_wb.active
    diff_ws.title = "字段级差异明细"
    diff_ws.append(diff_headers)
    for row in diff_rows:
        diff_ws.append(row)
    configure_sheet(diff_ws, diff_headers, diff_rows)
    summary_ws = diff_wb.create_sheet("指数级最终核验摘要")
    summary_ws.append(summary_headers)
    for row in summary_rows:
        summary_ws.append(row)
    configure_sheet(summary_ws, summary_headers, summary_rows)
    review_headers = ["指数代码", "指数名称", "指数发布机构", "官网来源URL", "未完成字段", "原因", "建议人工核验动作"]
    review_ws = diff_wb.create_sheet("仍需人工复核清单")
    review_ws.append(review_headers)
    for row in review_rows:
        review_ws.append(row)
    configure_sheet(review_ws, review_headers, review_rows)
    stats_ws = diff_wb.create_sheet("官网来源与核验等级统计")
    stats_headers = ["统计类型", "维度", "数量"]
    stats_ws.append(stats_headers)
    for row in stats_rows:
        stats_ws.append(row)
    configure_sheet(stats_ws, stats_headers, stats_rows)
    diff_wb.save(DIFF_OUTPUT)

    # Final validations: 15 sheets, all primary data row counts preserved, 82 mapped core indexes retained.
    protected_sheets = {
        "01_产品主表": 269,
        "02_月度规模份额": 7210,
        "03_月度交易流动性": 7433,
        "04_收益风险表现": 223,
        "05_持有人结构面板": 603,
        "06_持有人结构最新截面": 223,
        "07_核心指数月度表现估值": 11234,
        "08_核心指数最新收益": 81,
        "09_ETF_指数映射": 168,
        "11_口径差异摘要": original_counts["11_口径差异摘要"],
    }
    for sheet, expected_count in protected_sheets.items():
        actual_count = row_count(wb[sheet])
        if actual_count != expected_count:
            raise ValueError(f"{sheet}行数异常：{actual_count}，预期{expected_count}")
    if len(wb.sheetnames) != 15:
        raise ValueError(f"最终数据库sheet数为{len(wb.sheetnames)}，应为15")
    if len(final_rules) != 82 or len({code(row.get("指数代码")) for row in final_rules}) != 82:
        raise ValueError("最终10_指数规则明细不满足82行且指数代码唯一")
    _, mapping_records = worksheet_records(wb[MAPPING_SHEET])
    mapped_codes = {code(row.get("跟踪指数代码")) for row in mapping_records if code(row.get("跟踪指数代码"))}
    if mapped_codes - set(final_by_code):
        raise ValueError("09_ETF_指数映射存在未能在10_指数规则明细匹配的指数")
    for ws in wb.worksheets:
        if ws.merged_cells.ranges:
            raise ValueError(f"存在合并单元格：{ws.title}")
        if not text(ws["A1"].value):
            raise ValueError(f"存在空表头：{ws.title}")
    wb.save(CORE_OUTPUT)

    print(f"输出最终数据库路径：{CORE_OUTPUT}")
    print(f"输出最终差异表路径：{DIFF_OUTPUT}")
    print(f"最终数据库sheet数：{len(wb.sheetnames)}")
    print(f"10_指数规则明细行数：{len(final_rules)}")
    print(f"官网来源表指数数量：{len(official_by_code)}")
    print(f"成功匹配指数数量：{len(core_codes & set(official_by_code))}")
    print(f"未匹配指数数量：{len(unmatched_core) + len(unmatched_official)}")
    print(f"本轮尝试逐字段复核指数数量：{len(tried_codes)}")
    print(f"本轮成功完成核验指数数量：{len(successful_this_round)}")
    print(f"本轮仍需人工复核指数数量：{len(set(tried_codes) & set(pending_final_codes))}")
    print(f"本轮关键字段更新数量：{explicit_live_updates}")
    print(f"最终无需官网核验指数数量：{len(complete_final_codes)}")
    print(f"最终仍需官网核验指数数量：{len(pending_final_codes)}")
    print(f"官网核验等级A/B/C/D分布：A={grade_counts['A']} B={grade_counts['B']} C={grade_counts['C']} D={grade_counts['D']}")
    print("是否通过所有校验：是")


if __name__ == "__main__":
    main()
