from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT_PATH = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\全市场ETF基础信息_清洗与策略筛选.xlsx")
OUTPUT_PATH = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\全市场ETF基础信息_策略ETF池二次修正版.xlsx")

SOURCE_SHEET = "策略ETF_正式池_待复核"
FINAL_SHEET = "策略ETF_最终统计池"
WATCH_SHEET = "观察池_待人工确认"
NOTE_SHEET = "Step13_修正说明"

NEW_FIELDS = [
    "统计口径分类",
    "是否纳入核心策略ETF统计",
    "是否纳入广义策略ETF统计",
    "二次修正原因",
    "市场范围_二次修正",
]

WATCH_FIELDS = [
    "Wind代码",
    "证券简称",
    "基金简称",
    "跟踪指数名称",
    "命中关键词",
    "一级策略大类",
    "二级策略类别",
    "原是否纳入策略ETF统计",
    "统计口径分类",
    "二次修正原因",
    "备注",
]

CORE_PRIMARY_TYPES = {
    "红利",
    "自由现金流",
    "质量",
    "价值",
    "成长",
    "低波",
    "ESG",
    "基本面策略",
    "等权/另类加权",
}

PRIMARY_ORDER = [
    "红利",
    "自由现金流",
    "质量",
    "价值",
    "成长",
    "低波",
    "ESG",
    "基本面策略",
    "等权/另类加权",
    "指数增强/多因子",
    "待核验",
]

SECONDARY_ORDER = [
    "普通红利",
    "红利低波",
    "港股通低波红利",
    "红利质量",
    "红利价值",
    "红利成长",
    "央企红利",
    "国企红利",
    "港股通红利",
    "股东回报",
    "央企股东回报",
    "自由现金流",
    "质量",
    "价值",
    "成长",
    "低波",
    "ESG",
    "基本面策略",
    "等权/另类加权",
    "指数增强/多因子",
    "待核验",
]

CLASS_ORDER = {
    "核心策略指数ETF": 0,
    "广义策略ETF_指数增强": 1,
    "观察池": 2,
    "排除": 3,
}

WEAK_KEYWORDS = ["央企", "国企", "策略", "智选", "优选"]
STRONG_LOGIC = [
    "红利",
    "高股息",
    "股息",
    "低波",
    "质量",
    "价值",
    "成长",
    "自由现金流",
    "现金流",
    "基本面",
    "多因子",
    "增强",
    "ESG",
    "股东回报",
    "等权",
]

SPECIAL_OBSERVATION_TERMS = [
    "央企创新",
    "央企科技引领",
    "央企现代能源",
    "国企数字经济",
    "国企改革",
    "国企一带一路",
    "恒生中国企业",
    "中证科技50策略",
    "医药健康100策略",
    "证券公司先锋策略",
    "银行AH价格优选",
    "船舶产业智选",
    "沪深港科技50智选",
]

MARKET_TEXT_FIELDS = ["基金简称", "基金全称", "跟踪指数名称", "业绩比较基准"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=9, color="1F1F1F")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
GREEN_FILL = PatternFill("solid", fgColor="E2F0D9")
BLUE_FILL = PatternFill("solid", fgColor="D9EAF7")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN_GRAY = Side(style="thin", color="D9E1F2")
BODY_BORDER = Border(bottom=THIN_GRAY)

DATE_FIELDS = {
    "上市日期",
    "基金成立日",
    "托管费率_支持历史时间",
    "管理费率_支持历史时间",
}


def text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def joined_text(record: dict[str, Any], fields: list[str]) -> str:
    return " ".join(text(record.get(field)) for field in fields)


def contains_any(value: str, terms: list[str]) -> bool:
    return any(term in value for term in terms)


def classify_market(record: dict[str, Any]) -> str:
    combined = joined_text(record, MARKET_TEXT_FIELDS)
    combined_upper = combined.upper()
    index_code = text(record.get("跟踪指数代码")).upper()

    if "港股通" in combined:
        return "港股通"
    if contains_any(combined, ["恒生", "港股", "H股"]) and "港股通" not in combined:
        return "跨境/港股"
    if contains_any(combined, ["央企", "中央企业"]):
        return "央企"
    if contains_any(combined, ["国企", "国有企业"]):
        return "国企"
    if (
        contains_any(combined, ["中国A股", "中证", "国证", "上证", "深证"])
        or index_code.endswith((".CSI", ".CNI", ".SH", ".SZ"))
    ):
        return "A股"
    # Upper-case version is retained for robust Latin-text matching, even
    # though the current rule set is primarily Chinese.
    _ = combined_upper
    return "待核验"


def is_weak_only(record: dict[str, Any]) -> bool:
    evidence = " ".join(
        [
            text(record.get("命中关键词")),
            joined_text(record, MARKET_TEXT_FIELDS),
            text(record.get("证券简称")),
        ]
    )
    evidence_upper = evidence.upper()
    has_weak = contains_any(evidence, WEAK_KEYWORDS)
    has_strong = any(
        (term.upper() in evidence_upper if term == "ESG" else term in evidence)
        for term in STRONG_LOGIC
    )
    return has_weak and not has_strong


def has_weak_hit(record: dict[str, Any]) -> bool:
    hits = text(record.get("命中关键词"))
    return contains_any(hits, WEAK_KEYWORDS)


def refine_record(record: dict[str, Any]) -> dict[str, Any]:
    result = dict(record)
    wind_code = text(record.get("Wind代码")).upper()
    original_status = text(record.get("是否纳入策略ETF统计"))
    primary = text(record.get("一级策略大类"))
    combined = " ".join(
        [
            text(record.get("证券简称")),
            joined_text(record, MARKET_TEXT_FIELDS),
        ]
    )
    special_observation = contains_any(combined, SPECIAL_OBSERVATION_TERMS)
    weak_only = is_weak_only(record)

    if wind_code.endswith(".OF"):
        category = "观察池"
        core = "否"
        broad = "否"
        reason = "场外代码或待上市产品，暂不纳入上市 ETF 总量统计"
    elif original_status == "是" and primary in CORE_PRIMARY_TYPES:
        category = "核心策略指数ETF"
        core = "是"
        broad = "是"
        reason = "原已纳入策略ETF统计，且属于核心策略指数类别"
    elif original_status == "是" and primary == "指数增强/多因子":
        category = "广义策略ETF_指数增强"
        core = "否"
        broad = "是"
        reason = "指数增强产品纳入广义策略ETF统计，不纳入核心策略指数ETF统计"
    elif original_status == "待核验":
        category = "观察池"
        core = "否"
        broad = "否"
        reason = "原统计状态为待核验，转入观察池等待人工确认"
    elif original_status == "否":
        category = "排除"
        core = "否"
        broad = "否"
        reason = "原统计状态为否，暂不纳入策略ETF统计"
    else:
        category = "观察池"
        core = "否"
        broad = "否"
        reason = "原统计状态或策略分类无法直接判断，转入观察池"

    # These rules reinforce the observation treatment without overriding
    # original “否” records into the final statistics.
    if special_observation:
        core = "否"
        broad = "否"
        if category != "排除":
            category = "观察池"
        reason = "更偏主题、政策或行业策略，暂不纳入核心策略 ETF 统计"
    elif weak_only:
        core = "否"
        broad = "否"
        if category != "排除":
            category = "观察池"
        reason = "仅命中弱策略关键词，缺少明确强策略逻辑，进入观察池"

    result.update(
        {
            "统计口径分类": category,
            "是否纳入核心策略ETF统计": core,
            "是否纳入广义策略ETF统计": broad,
            "二次修正原因": reason,
            "市场范围_二次修正": classify_market(record),
        }
    )
    return result


def write_records(ws, headers: list[str], records: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for record in records:
        ws.append([record.get(header) for header in headers])


def style_table(ws, headers: list[str]) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BODY_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    sample_rows = min(ws.max_row, 300)
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, sample_rows + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 50))
        if header in {
            "基金全称",
            "业绩比较基准",
            "跟踪指数名称",
            "命中关键词",
            "分类依据",
            "二次修正原因",
            "备注",
        }:
            width = min(max(max_len * 1.2, 18), 40)
        else:
            width = min(max(max_len * 1.15, 11), 26)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for header in DATE_FIELDS:
        if header in headers:
            col = headers.index(header) + 1
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row_idx, col)
                if isinstance(cell.value, (date, datetime)):
                    cell.number_format = "yyyy-mm-dd"

    if "最新基金规模(亿)" in headers:
        col = headers.index("最新基金规模(亿)") + 1
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row_idx, col).number_format = "#,##0.0000"


def add_final_conditional_formatting(ws, headers: list[str]) -> None:
    if "统计口径分类" not in headers or ws.max_row < 2:
        return
    col = headers.index("统计口径分类") + 1
    letter = get_column_letter(col)
    target = f"{letter}2:{letter}{ws.max_row}"
    rules = [
        ("核心策略指数ETF", GREEN_FILL),
        ("广义策略ETF_指数增强", BLUE_FILL),
        ("观察池", YELLOW_FILL),
        ("排除", RED_FILL),
    ]
    for label, fill in rules:
        ws.conditional_formatting.add(
            target,
            FormulaRule(formula=[f'{letter}2="{label}"'], fill=fill),
        )


def build_note_rows(records: list[dict[str, Any]]) -> list[list[Any]]:
    core_records = [r for r in records if r.get("是否纳入核心策略ETF统计") == "是"]
    broad_records = [r for r in records if r.get("是否纳入广义策略ETF统计") == "是"]
    enhanced_records = [r for r in records if r.get("统计口径分类") == "广义策略ETF_指数增强"]
    watch_records = [r for r in records if r.get("统计口径分类") == "观察池"]
    excluded_records = [r for r in records if r.get("统计口径分类") == "排除"]

    def size_sum(items: list[dict[str, Any]]) -> float:
        return round(
            sum(
                float(r["最新基金规模(亿)"])
                for r in items
                if isinstance(r.get("最新基金规模(亿)"), (int, float))
            ),
            4,
        )

    rows: list[list[Any]] = []

    def add(section: str, metric: str, value: Any, note: str = "") -> None:
        rows.append([section, metric, value, note])

    add("总体统计", "原策略ETF正式池数量", len(records))
    add("总体统计", "核心策略指数ETF数量", len(core_records))
    add("总体统计", "广义策略ETF_指数增强数量", len(enhanced_records))
    add("总体统计", "广义策略ETF合计数量", len(broad_records), "核心策略指数ETF + 指数增强")
    add("总体统计", "观察池数量", len(watch_records))
    add("总体统计", "排除数量", len(excluded_records))
    add("规模统计", "核心策略指数ETF规模合计", size_sum(core_records), "单位：亿元；忽略规模缺失值")
    add("规模统计", "广义策略ETF规模合计", size_sum(broad_records), "单位：亿元；忽略规模缺失值")

    core_counts = Counter(text(r.get("一级策略大类")) or "待核验" for r in core_records)
    core_sizes: defaultdict[str, float] = defaultdict(float)
    broad_counts = Counter(text(r.get("一级策略大类")) or "待核验" for r in broad_records)
    market_counts = Counter(text(r.get("市场范围_二次修正")) or "待核验" for r in records)

    for record in core_records:
        size = record.get("最新基金规模(亿)")
        if isinstance(size, (int, float)):
            core_sizes[text(record.get("一级策略大类")) or "待核验"] += float(size)

    primary_rank = {value: idx for idx, value in enumerate(PRIMARY_ORDER)}
    for primary in sorted(core_counts, key=lambda x: (primary_rank.get(x, 999), x)):
        add("核心策略各一级策略数量", primary, core_counts[primary])
    for primary in sorted(core_sizes, key=lambda x: (primary_rank.get(x, 999), x)):
        add("核心策略各一级策略规模", primary, round(core_sizes[primary], 4), "单位：亿元")
    for primary in sorted(broad_counts, key=lambda x: (primary_rank.get(x, 999), x)):
        add("广义策略各一级策略数量", primary, broad_counts[primary])
    for market, count in sorted(market_counts.items(), key=lambda item: (-item[1], item[0])):
        add("市场范围_二次修正分布", market, count, "覆盖原策略ETF正式池全部记录")

    corrected_to_a = sum(
        text(r.get("市场范围")) == "待核验" and r.get("市场范围_二次修正") == "A股"
        for r in records
    )
    add("市场范围修正", "原市场范围为“待核验”但已修正为“A股”的数量", corrected_to_a)
    add(
        "口径说明",
        "核心与广义关系",
        "广义策略ETF包含核心策略指数ETF及指数增强产品",
    )
    add(
        "口径说明",
        "观察池用途",
        "保存待核验、场外/待上市及仅命中弱策略关键词的产品，不计入总量",
    )

    return rows


def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"输入文件不存在：{INPUT_PATH}")

    wb = load_workbook(INPUT_PATH)
    if SOURCE_SHEET not in wb.sheetnames:
        raise KeyError(f"输入工作簿缺少 sheet：{SOURCE_SHEET}")

    source_ws = wb[SOURCE_SHEET]
    original_headers = [cell.value for cell in source_ws[1]]
    required = {
        "Wind代码",
        "证券简称",
        "基金简称",
        "基金全称",
        "跟踪指数名称",
        "业绩比较基准",
        "跟踪指数代码",
        "最新基金规模(亿)",
        "命中关键词",
        "是否纳入策略ETF统计",
        "一级策略大类",
        "二级策略类别",
        "市场范围",
        "备注",
    }
    missing = required - set(original_headers)
    if missing:
        raise KeyError(f"正式池缺少必要字段：{sorted(missing)}")

    source_records: list[dict[str, Any]] = []
    for values in source_ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and text(value) != "" for value in values):
            continue
        source_records.append(dict(zip(original_headers, values)))

    refined_records = [refine_record(record) for record in source_records]

    primary_rank = {value: idx for idx, value in enumerate(PRIMARY_ORDER)}
    secondary_rank = {value: idx for idx, value in enumerate(SECONDARY_ORDER)}
    refined_records.sort(
        key=lambda r: (
            CLASS_ORDER.get(text(r.get("统计口径分类")), 999),
            primary_rank.get(text(r.get("一级策略大类")), 999),
            secondary_rank.get(text(r.get("二级策略类别")), 999),
            -(float(r["最新基金规模(亿)"]) if isinstance(r.get("最新基金规模(亿)"), (int, float)) else -1),
            text(r.get("Wind代码")),
        )
    )

    watch_records: list[dict[str, Any]] = []
    for record in refined_records:
        if record.get("统计口径分类") == "观察池" or (
            record.get("统计口径分类") == "排除" and has_weak_hit(record)
        ):
            watch = dict(record)
            watch["原是否纳入策略ETF统计"] = record.get("是否纳入策略ETF统计")
            watch_records.append(watch)

    for sheet_name in [FINAL_SHEET, WATCH_SHEET, NOTE_SHEET]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    final_ws = wb.create_sheet(FINAL_SHEET)
    watch_ws = wb.create_sheet(WATCH_SHEET)
    note_ws = wb.create_sheet(NOTE_SHEET)

    final_headers = original_headers + NEW_FIELDS
    write_records(final_ws, final_headers, refined_records)
    write_records(watch_ws, WATCH_FIELDS, watch_records)

    note_headers = ["统计类别", "指标", "数量/规模", "说明"]
    note_ws.append(note_headers)
    for row in build_note_rows(refined_records):
        note_ws.append(row)

    style_table(final_ws, final_headers)
    style_table(watch_ws, WATCH_FIELDS)
    style_table(note_ws, note_headers)
    add_final_conditional_formatting(final_ws, final_headers)

    previous_section = None
    for row_idx in range(2, note_ws.max_row + 1):
        section = note_ws.cell(row_idx, 1).value
        if section != previous_section:
            for col_idx in range(1, 5):
                note_ws.cell(row_idx, col_idx).fill = SECTION_FILL
                note_ws.cell(row_idx, col_idx).font = Font(
                    name="微软雅黑", size=9, bold=True, color="1F1F1F"
                )
        previous_section = section
        if isinstance(note_ws.cell(row_idx, 3).value, float):
            note_ws.cell(row_idx, 3).number_format = "#,##0.0000"
        elif isinstance(note_ws.cell(row_idx, 3).value, int):
            note_ws.cell(row_idx, 3).number_format = "#,##0"

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except AttributeError:
        pass

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)

    core_count = sum(r.get("是否纳入核心策略ETF统计") == "是" for r in refined_records)
    broad_count = sum(r.get("是否纳入广义策略ETF统计") == "是" for r in refined_records)
    watch_count = sum(r.get("统计口径分类") == "观察池" for r in refined_records)
    excluded_count = sum(r.get("统计口径分类") == "排除" for r in refined_records)

    print(f"核心策略指数ETF数量：{core_count}")
    print(f"广义策略ETF数量：{broad_count}")
    print(f"观察池数量：{watch_count}")
    print(f"排除数量：{excluded_count}")
    print(f"输出文件路径：{OUTPUT_PATH}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"处理失败：{exc}", file=sys.stderr)
        raise
