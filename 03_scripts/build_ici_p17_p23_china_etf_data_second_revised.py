from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT_PATH = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\原始数据\全市场ETF基础信息.csv.xlsx")
OUTPUT_DIR = Path(r"C:\Users\lvdon\Desktop\Fullgoal\ICI课题\P17-P23_中国ETF数据处理_二次修正版")

OUTPUT_FILES = {
    "00": "00_数据字段检查与描述性统计_二次修正版.xlsx",
    "P17": "P17_中国ETF年度新发与累计数量_二次修正版.xlsx",
    "P18": "P18_中国ETF资产类型与产品类型结构_二次修正版.xlsx",
    "P19": "P19_中国ETF管理人集中度与头部排名_二次修正版.xlsx",
    "P20": "P20_中国ETF规模区间分布与长尾分析_二次修正版.xlsx",
    "P21": "P21_中美对比所需中国ETF总览指标_二次修正版.xlsx",
    "P22P23": "P22-P23_结论页支撑数据摘要_二次修正版.xlsx",
    "ARCHIVE": "ICI课题_P17-P23_中国ETF作图数据汇总_二次修正版.xlsx",
    "README": "ICI课题_P17-P23_数据处理说明_二次修正版.md",
}

RAW_TO_STD = {
    "Wind代码": "Wind代码",
    "证券简称": "证券简称",
    "基金简称": "基金简称",
    "基金全称": "基金全称",
    "基金代码": "基金代码",
    "交易代码": "交易代码",
    "基金类型": "基金类型",
    "投资类型_二级分类": "投资类型",
    "业绩比较基准": "业绩比较基准",
    "最新基金规模(亿)": "基金规模_亿元",
    "基金管理人": "基金管理人",
    "ETF一级市场基金代码": "一级市场基金代码",
    "跟踪指数名称": "跟踪指数",
    "跟踪指数代码": "跟踪指数代码",
    "基金上市地点": "基金上市地点",
    "上市日期": "上市日期",
    "基金成立日": "成立日期",
    "基金托管人": "基金托管人",
    "管理费率_支持历史(%)": "管理费率",
    "托管费率_支持历史(%)": "托管费率",
}

BASE_COLUMNS = list(RAW_TO_STD.values())
CLASS_COLUMNS = [
    "资产类型",
    "股票ETF内部类型",
    "产品类型_汇总",
    "策略ETF_主分类",
    "策略增强标签",
    "是否增强策略",
    "是否跨境ETF",
    "是否债券ETF",
    "是否商品ETF",
    "是否货币ETF",
    "是否股票ETF",
    "是否宽基ETF",
    "是否行业主题ETF",
    "是否策略ETF",
    "是否多资产ETF",
    "命中关键词",
    "分类备注",
    "入池状态",
    "入池/观察原因",
]

EXCLUDE_TERMS = ["ETF联接", "ETF连接", "联接A", "联接C", "联接", "场外联接", "FOF", "LOF"]
MONEY_TERMS = ["货币", "现金", "保证金", "快线", "添益", "场内货币", "收益宝", "财富宝", "银华日利", "华宝添益", "交易货币", "易货币"]
MONEY_POOL_TERMS = ["银华日利", "华宝添益", "现金添益", "保证金", "场内货币", "收益宝", "财富宝", "货币ETF", "快线", "添富快线", "易货币", "交易货币"]
MONEY_REPRESENTATIVE_TERMS = ["银华日利", "华宝添益", "现金添益", "保证金", "场内货币", "收益宝", "财富宝"]
COMMODITY_TERMS = ["黄金ETF", "黄金基金", "豆粕ETF", "商品ETF", "能源化工", "原油ETF", "白银ETF"]
BOND_TERMS = ["国债", "政金债", "信用债", "公司债", "地方债", "城投债", "可转债", "短融", "债券", "债"]
CROSS_TERMS = [
    "恒生",
    "港股",
    "港股通",
    "H股",
    "纳斯达克",
    "纳指",
    "标普",
    "道琼斯",
    "日经",
    "德国",
    "法国",
    "海外",
    "中概",
    "东南亚",
    "QDII",
    "MSCI美国",
    "美国",
    "香港",
    "日韩",
    "越南",
    "沙特",
    "亚太",
    "全球",
]
BROAD_TERMS = [
    "沪深300",
    "中证500",
    "中证800",
    "中证1000",
    "中证2000",
    "中证A500",
    "中证A100",
    "中证A50",
    "中证A股",
    "上证50",
    "上证180",
    "上证380",
    "上证综指",
    "上证指数",
    "深证100",
    "深证50",
    "创业板指",
    "创业板50",
    "创业板综",
    "科创50",
    "科创100",
    "北证50",
    "MSCI中国A股",
    "MSCI中国A50互联互通",
    "MSCI A50",
    "富时中国A50",
    "国证A指",
    "万得全A",
    "中小板",
    "中小100",
]
CORE_BROAD_TERMS = BROAD_TERMS
INDUSTRY_ALL_SHARE_EXCLUDE = [
    "中证全指证券公司",
    "中证全指医药卫生",
    "中证全指医药",
    "中证全指半导体",
    "中证全指软件",
    "中证全指食品",
    "中证全指电力",
    "中证全指通信",
    "中证全指房地产",
    "中证全指银行",
    "中证全指建筑材料",
]
STRATEGY_TERMS = [
    "红利",
    "低波",
    "低波动",
    "红利低波",
    "质量",
    "价值",
    "成长",
    "质量成长",
    "现金流",
    "自由现金流",
    "股息",
    "高股息",
    "基本面",
    "等权",
    "等权重",
    "ESG",
    "动量",
    "回购",
    "分红",
    "龙头",
    "Smart Beta",
    "央企红利",
    "国企红利",
    "红利质量",
    "红利价值",
    "价值100",
    "成长100",
    "低估值",
    "高分红",
]
ENHANCED_TERMS = ["增强策略", "指数增强", "增强"]
INDUSTRY_TERMS = [
    "新能源",
    "光伏",
    "芯片",
    "半导体",
    "人工智能",
    "机器人",
    "军工",
    "医药",
    "消费",
    "金融",
    "证券公司",
    "证券行业",
    "银行",
    "地产",
    "传媒",
    "计算机",
    "通信",
    "汽车",
    "电池",
    "储能",
    "双碳",
    "绿色",
    "电力",
    "煤炭",
    "有色",
    "钢铁",
    "基建",
    "农业",
    "酒",
    "食品",
    "云计算",
    "软件",
    "游戏",
    "数字经济",
    "高端制造",
    "创新药",
    "医疗器械",
    "工业母机",
    "工程机械",
    "畜牧养殖",
    "稀有金属",
    "工业有色",
    "化工",
    "材料",
    "环保",
    "教育",
    "物流",
    "旅游",
    "黄金股",
    "矿业",
    "交运",
    "信息技术",
    "信息",
    "科技",
    "互联网",
    "生物",
    "医疗",
    "养老",
    "家电",
    "机械",
    "电子",
    "保险",
    "房地产",
]
CORE_BROAD_CHECK_TERMS = ["A500", "沪深300", "中证500", "中证1000", "创业板指", "科创50", "上证50"]
INDUSTRY_ALL_SHARE_CHECK_TERMS = ["全指证券", "全指医药", "全指食品", "全指软件", "全指电力", "全指银行", "全指通信"]
NOT_BROAD_CHECK_TERMS = [
    "A500红利低波",
    "中证A500红利低波",
    "沪深300质量",
    "沪深300红利",
    "中证500低波",
    "中证500质量成长",
    "中证500信息技术",
    "中证500医药",
    "中证500消费",
    "300价值",
    "300成长",
    "500质量",
    "500价值",
    "500成长",
]
SHOULD_BROAD_CHECK_TERMS = [
    "沪深300ETF",
    "中证500ETF",
    "中证1000ETF",
    "中证A500ETF",
    "创业板ETF",
    "科创50ETF",
    "上证50ETF",
    "中证1000增强策略ETF",
    "沪深300增强策略ETF",
    "创业板增强策略ETF",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "nat", "none"}:
        return ""
    return re.sub(r"\s+", " ", text)


def code_text(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text.zfill(6) if text.isdigit() and len(text) < 6 else text


def number_value(value: Any) -> float | None:
    text = clean_text(value).replace(",", "").replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def date_text(value: Any) -> str:
    if clean_text(value) == "":
        return ""
    dt = pd.to_datetime(value, errors="coerce")
    return "" if pd.isna(dt) else dt.strftime("%Y-%m-%d")


def contains_terms(text: str, terms: list[str]) -> list[str]:
    upper = text.upper()
    return [term for term in terms if term.upper() in upper]


def pct(value: float) -> str:
    return f"{value:.2%}"


def add_pct_display(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    pct_cols = [col for col in out.columns if "占比" in col or col in {"数量口径", "规模口径"}]
    for col in pct_cols:
        insert_at = out.columns.get_loc(col) + 1
        display_col = f"{col}_展示"
        out.insert(insert_at, display_col, out[col].map(lambda x: "" if pd.isna(x) else pct(float(x))))
    return out


def round4(value: Any) -> Any:
    if pd.isna(value):
        return ""
    if isinstance(value, (int, float)):
        return round(float(value), 4)
    return value


def manager_short_name(name: str) -> str:
    text = clean_text(name)
    for old in ["基金管理股份有限公司", "基金管理有限公司", "管理股份有限公司", "管理有限公司"]:
        text = text.replace(old, "")
    text = text.replace("(中国)", "")
    return text


def read_and_standardize() -> tuple[dict[str, pd.DataFrame], pd.DataFrame, pd.DataFrame, int]:
    raw_sheets = pd.read_excel(INPUT_PATH, sheet_name=None, dtype=object)
    raw = raw_sheets["Sheet1"].copy()
    raw_rows = len(raw)

    footer_mask = raw["Wind代码"].map(clean_text).str.startswith("数据来源")
    raw_no_footer = raw.loc[~footer_mask].copy()

    std = pd.DataFrame()
    for raw_col, std_col in RAW_TO_STD.items():
        if raw_col in raw_no_footer.columns:
            if std_col in {"Wind代码", "基金代码", "交易代码", "一级市场基金代码"}:
                std[std_col] = raw_no_footer[raw_col].map(code_text)
            elif std_col in {"上市日期", "成立日期"}:
                std[std_col] = raw_no_footer[raw_col].map(date_text)
            elif std_col in {"基金规模_亿元", "管理费率", "托管费率"}:
                std[std_col] = raw_no_footer[raw_col].map(number_value)
            else:
                std[std_col] = raw_no_footer[raw_col].map(clean_text)
        else:
            std[std_col] = ""
    std = std.reset_index(drop=True)

    field_mapping = pd.DataFrame(
        [
            {
                "标准字段名": std_col,
                "原始字段名": raw_col,
                "所在sheet": "Sheet1",
                "识别方式": "字段名精确映射",
                "备注": "原字段已为亿元口径，直接转数值" if std_col == "基金规模_亿元" else "",
            }
            for raw_col, std_col in RAW_TO_STD.items()
        ]
    )
    return raw_sheets, raw_no_footer, std, raw_rows


def pool_masks(std: pd.DataFrame) -> tuple[pd.Series, pd.Series, pd.DataFrame]:
    wind = std["Wind代码"].astype(str).str.upper()
    suffix = wind.str.extract(r"(\.[A-Z]+)$", expand=False).fillna("")
    name_text = (std["基金简称"].fillna("") + " " + std["基金全称"].fillna("")).astype(str)
    full_name = std["基金全称"].fillna("").astype(str)
    investment_type = std["投资类型"].fillna("").astype(str)

    listed_suffix = suffix.isin([".SH", ".SZ"])
    has_exchange = std["基金上市地点"].map(clean_text).ne("")
    has_listing_date = std["上市日期"].map(clean_text).ne("")
    has_start_date = std["成立日期"].map(clean_text).ne("")
    has_scale = pd.to_numeric(std["基金规模_亿元"], errors="coerce").gt(0)
    name_has_etf = name_text.str.contains("ETF", case=False, na=False)
    name_has_exchange_open = full_name.str.contains("交易型开放式", na=False)
    name_has_exchange_money = full_name.str.contains("交易型货币市场基金", na=False)
    is_money_type = investment_type.str.contains("货币市场型基金", na=False)
    name_has_money_pool_term = name_text.map(lambda x: bool(contains_terms(x, MONEY_POOL_TERMS)))
    included_product_form = (
        name_has_etf
        | name_has_exchange_open
        | name_has_exchange_money
        | (is_money_type & listed_suffix & has_exchange)
        | name_has_money_pool_term
    )
    exclude_hit = name_text.map(lambda x: bool(contains_terms(x, EXCLUDE_TERMS)))

    hard_conditions = listed_suffix & has_exchange & has_listing_date & has_start_date & has_scale
    main_mask = hard_conditions & included_product_form & ~exclude_hit
    key_missing = ~(has_exchange & has_listing_date & has_start_date & has_scale)
    observation_mask = (~main_mask) & (suffix.eq(".OF") | key_missing | name_has_etf | name_has_exchange_open | name_has_exchange_money)

    steps = [
        ("原始行数", len(std), "已剔除footer后的产品记录"),
        ("Wind代码后缀为.SH或.SZ", int(listed_suffix.sum()), "主分析池必须为场内上市交易代码"),
        ("基金上市地点非空", int((listed_suffix & has_exchange).sum()), "在.SH/.SZ基础上继续筛选"),
        ("上市日期非空", int((listed_suffix & has_exchange & has_listing_date).sum()), "在前一步基础上继续筛选"),
        ("成立日期非空", int((listed_suffix & has_exchange & has_listing_date & has_start_date).sum()), "在前一步基础上继续筛选"),
        ("规模非空且大于0", int((listed_suffix & has_exchange & has_listing_date & has_start_date & has_scale).sum()), "在前一步基础上继续筛选"),
        ("满足ETF/交易型开放式/场内货币纳入条件", int((hard_conditions & included_product_form).sum()), "硬条件后满足任一纳入条件"),
        ("排除联接/FOF/LOF等", int(main_mask.sum()), "最终上市ETF分析池"),
    ]
    screening_process = pd.DataFrame(steps, columns=["筛选步骤", "剩余/命中数量", "说明"])

    reasons = []
    for idx in std.index:
        item = []
        if suffix.loc[idx] == ".OF":
            item.append("Wind代码后缀为.OF")
        if suffix.loc[idx] not in [".SH", ".SZ", ".OF"]:
            item.append("Wind代码后缀非.SH/.SZ/.OF")
        if not has_exchange.loc[idx]:
            item.append("基金上市地点缺失")
        if not has_listing_date.loc[idx]:
            item.append("上市日期缺失")
        if not has_start_date.loc[idx]:
            item.append("成立日期缺失")
        if not has_scale.loc[idx]:
            item.append("规模缺失或小于等于0")
        if not included_product_form.loc[idx]:
            item.append("未命中ETF/交易型开放式/场内货币纳入条件")
        if exclude_hit.loc[idx]:
            item.append("命中联接/FOF/LOF等排除词")
        reasons.append("；".join(item) if item else "满足主分析池口径")
    std["入池状态"] = ["上市ETF分析池" if main_mask.loc[i] else ("观察池" if observation_mask.loc[i] else "非本次口径") for i in std.index]
    std["入池/观察原因"] = reasons
    return main_mask, observation_mask, screening_process


def classify_row(row: pd.Series) -> pd.Series:
    text = " ".join(
        clean_text(row.get(col))
        for col in ["证券简称", "基金简称", "基金全称", "跟踪指数", "业绩比较基准", "投资类型"]
    )
    investment_type = clean_text(row.get("投资类型"))
    hits: list[str] = []
    notes: list[str] = []

    money_hits = contains_terms(text, MONEY_TERMS)
    commodity_hits = contains_terms(text, COMMODITY_TERMS)
    bond_hits = contains_terms(text, BOND_TERMS)
    cross_hits = contains_terms(text, CROSS_TERMS)

    if "货币市场型基金" in investment_type or "交易型货币市场基金" in text or money_hits:
        asset_type = "货币ETF"
        hits += money_hits or ["投资类型=货币市场型基金"]
    elif "商品型基金" in investment_type or commodity_hits:
        asset_type = "商品ETF"
        hits += commodity_hits or ["投资类型=商品型基金"]
    elif "被动指数型债券基金" in investment_type or bond_hits:
        asset_type = "债券ETF"
        hits += bond_hits or ["投资类型=被动指数型债券基金"]
    elif "国际(QDII)股票型基金" in investment_type or cross_hits:
        asset_type = "跨境ETF"
        hits += cross_hits or ["投资类型=国际(QDII)股票型基金"]
    elif investment_type in {"被动指数型基金", "增强指数型基金"}:
        asset_type = "股票ETF"
    else:
        asset_type = "其他ETF"
        notes.append("投资类型和名称关键词不足，无法归入主要资产类型")

    strategy_hits = contains_terms(text, STRATEGY_TERMS)
    enhanced_hits = contains_terms(text, ENHANCED_TERMS)
    industry_exclude_hits = contains_terms(text, INDUSTRY_ALL_SHARE_EXCLUDE)
    broad_hits = [] if industry_exclude_hits else contains_terms(text, CORE_BROAD_TERMS)
    industry_hits = contains_terms(text, INDUSTRY_TERMS) + industry_exclude_hits
    is_enhanced = "增强指数型基金" in investment_type or bool(enhanced_hits)
    has_strategy_factor = bool(strategy_hits)
    is_core_broad = bool(broad_hits)

    if asset_type == "股票ETF":
        if industry_hits:
            stock_inner = "行业主题ETF"
            hits += industry_hits
        elif has_strategy_factor:
            stock_inner = "策略ETF"
            hits += strategy_hits
        elif is_enhanced and is_core_broad:
            stock_inner = "宽基ETF"
            hits += broad_hits + (enhanced_hits or ["增强指数型基金"])
            notes.append("宽基增强/指数增强，主分类按宽基处理")
        elif broad_hits:
            stock_inner = "宽基ETF"
            hits += broad_hits
        elif is_enhanced:
            stock_inner = "策略ETF"
            hits += enhanced_hits or ["增强指数型基金"]
        else:
            stock_inner = "其他股票ETF"
            notes.append("股票ETF但未命中宽基/策略/行业主题关键词")
    else:
        stock_inner = "不适用"
        if industry_exclude_hits:
            notes.append("非股票ETF，行业全指排除词不参与宽基分类")

    if asset_type == "股票ETF":
        product_summary = stock_inner
    elif asset_type in {"跨境ETF", "债券ETF", "商品ETF", "货币ETF"}:
        product_summary = asset_type
    else:
        product_summary = "其他ETF"

    strategy_main = stock_inner == "策略ETF"
    if asset_type == "股票ETF":
        if is_enhanced:
            strategy_label = "增强策略"
        elif has_strategy_factor:
            strategy_label = "Smart Beta / 策略因子"
        elif stock_inner == "行业主题ETF":
            strategy_label = "行业主题"
        elif stock_inner == "宽基ETF":
            strategy_label = "宽基"
        else:
            strategy_label = "无"
    else:
        strategy_label = "无"
    is_strategy_tag = strategy_main or strategy_label in {"增强策略", "Smart Beta / 策略因子"}

    flags = {
        "是否跨境ETF": asset_type == "跨境ETF",
        "是否债券ETF": asset_type == "债券ETF",
        "是否商品ETF": asset_type == "商品ETF",
        "是否货币ETF": asset_type == "货币ETF",
        "是否股票ETF": asset_type == "股票ETF",
        "是否宽基ETF": stock_inner == "宽基ETF",
        "是否行业主题ETF": stock_inner == "行业主题ETF",
        "是否策略ETF": is_strategy_tag,
        "是否多资产ETF": asset_type != "股票ETF",
        "是否增强策略": is_enhanced if asset_type == "股票ETF" else False,
    }
    return pd.Series(
        {
            "资产类型": asset_type,
            "股票ETF内部类型": stock_inner,
            "产品类型_汇总": product_summary,
            "策略ETF_主分类": "是" if strategy_main else "否",
            "策略增强标签": strategy_label,
            **{key: "是" if value else "否" for key, value in flags.items()},
            "命中关键词": "、".join(dict.fromkeys(hits)),
            "分类备注": "；".join(dict.fromkeys(notes)),
        }
    )


def finalize_pools(std: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    main_mask, observation_mask, screening_process = pool_masks(std)
    class_df = std.apply(classify_row, axis=1)
    full = pd.concat([std, class_df], axis=1)
    main = full.loc[main_mask, BASE_COLUMNS + CLASS_COLUMNS].copy().reset_index(drop=True)
    observation = full.loc[observation_mask, BASE_COLUMNS + CLASS_COLUMNS].copy().reset_index(drop=True)
    return main, observation, screening_process


def group_structure(df: pd.DataFrame, by_col: str) -> pd.DataFrame:
    total_count = len(df)
    total_scale = pd.to_numeric(df["基金规模_亿元"], errors="coerce").sum()
    rows = []
    for key, part in df.groupby(by_col, dropna=False):
        scale = pd.to_numeric(part["基金规模_亿元"], errors="coerce")
        rows.append(
            {
                by_col: key if clean_text(key) else "未分类",
                "ETF数量": len(part),
                "数量占比": len(part) / total_count if total_count else 0,
                "总规模_亿元": round4(scale.sum()),
                "规模占比": scale.sum() / total_scale if total_scale else 0,
                "平均规模_亿元": round4(scale.mean()),
                "规模中位数_亿元": round4(scale.median()),
            }
        )
    out = pd.DataFrame(rows).sort_values(["总规模_亿元", "ETF数量"], ascending=[False, False]).reset_index(drop=True)
    return add_pct_display(out)


def annual_table(df: pd.DataFrame, by_col: str | None = None) -> pd.DataFrame:
    temp = df.copy()
    temp["年份"] = pd.to_datetime(temp["成立日期"], errors="coerce").dt.year
    temp = temp.dropna(subset=["年份"]).copy()
    temp["年份"] = temp["年份"].astype(int)
    years = list(range(int(temp["年份"].min()), int(temp["年份"].max()) + 1)) if len(temp) else []
    if by_col is None:
        out = temp.groupby("年份").size().reindex(years, fill_value=0).reset_index(name="当年新成立ETF数量")
        out["累计ETF数量"] = out["当年新成立ETF数量"].cumsum()
        out["年度标记"] = out["年份"].map(lambda y: "YTD" if y == 2026 else "")
        return out
    keys = sorted(temp[by_col].dropna().unique().tolist())
    frame = pd.MultiIndex.from_product([years, keys], names=["年份", by_col]).to_frame(index=False)
    counts = temp.groupby(["年份", by_col]).size().reset_index(name="当年新成立ETF数量")
    out = frame.merge(counts, on=["年份", by_col], how="left").fillna({"当年新成立ETF数量": 0})
    out["当年新成立ETF数量"] = out["当年新成立ETF数量"].astype(int)
    out["累计ETF数量"] = out.groupby(by_col)["当年新成立ETF数量"].cumsum()
    return out


def manager_tables(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    total_count = len(df)
    total_scale = df["基金规模_亿元"].sum()
    rows = []
    for manager, part in df.groupby("基金管理人", dropna=False):
        scale = pd.to_numeric(part["基金规模_亿元"], errors="coerce")
        rows.append(
            {
                "基金管理人": manager,
                "管理人简称": manager_short_name(manager),
                "ETF数量": len(part),
                "ETF总规模_亿元": round4(scale.sum()),
                "规模占比": scale.sum() / total_scale if total_scale else 0,
                "平均单品规模_亿元": round4(scale.mean()),
                "规模中位数_亿元": round4(scale.median()),
                "最大单品规模_亿元": round4(scale.max()),
                "10亿以下产品数量": int((scale < 10).sum()),
                "10亿以下产品占比": int((scale < 10).sum()) / len(part) if len(part) else 0,
            }
        )
    ranking = pd.DataFrame(rows).sort_values(["ETF总规模_亿元", "ETF数量"], ascending=[False, False]).reset_index(drop=True)
    ranking.insert(0, "排名", range(1, len(ranking) + 1))
    ranking = add_pct_display(ranking)
    cr_rows = []
    for n in [3, 5, 10, 20]:
        top = ranking.head(n)
        cr_rows.append(
            {
                "指标": f"CR{n}",
                "数量口径": top["ETF数量"].sum() / total_count if total_count else 0,
                "规模口径": top["ETF总规模_亿元"].sum() / total_scale if total_scale else 0,
                "说明": f"按ETF总规模排名的前{n}大管理人合计占比",
            }
        )
    return ranking, add_pct_display(pd.DataFrame(cr_rows))


def size_tables(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    temp = df.copy()
    labels = ["<1亿", "1–2亿", "2–10亿", "10–50亿", "50–200亿", "200亿以上"]
    temp["规模区间"] = pd.cut(temp["基金规模_亿元"], [-math.inf, 1, 2, 10, 50, 200, math.inf], labels=labels, right=False)
    valid = temp.dropna(subset=["规模区间"]).copy()
    total_count = len(valid)
    total_scale = valid["基金规模_亿元"].sum()

    full_rows = []
    for label in labels:
        part = valid[valid["规模区间"].astype(str) == label]
        full_rows.append(
            {
                "规模区间": label,
                "ETF数量": len(part),
                "数量占比": len(part) / total_count if total_count else 0,
                "总规模_亿元": round4(part["基金规模_亿元"].sum()),
                "规模占比": part["基金规模_亿元"].sum() / total_scale if total_scale else 0,
            }
        )
    full = add_pct_display(pd.DataFrame(full_rows))

    by_type_rows = []
    for product_type, part0 in valid.groupby("产品类型_汇总"):
        type_count = len(part0)
        type_scale = part0["基金规模_亿元"].sum()
        for label in labels:
            part = part0[part0["规模区间"].astype(str) == label]
            by_type_rows.append(
                {
                    "产品类型_汇总": product_type,
                    "规模区间": label,
                    "ETF数量": len(part),
                    "数量占比": len(part) / type_count if type_count else 0,
                    "总规模_亿元": round4(part["基金规模_亿元"].sum()),
                    "规模占比": part["基金规模_亿元"].sum() / type_scale if type_scale else 0,
                }
            )
    by_type = add_pct_display(pd.DataFrame(by_type_rows))

    manager_rows = []
    for manager, part in temp.groupby("基金管理人", dropna=False):
        scale = part["基金规模_亿元"]
        count = len(part)
        manager_rows.append(
            {
                "基金管理人": manager,
                "管理人简称": manager_short_name(manager),
                "ETF总数量": count,
                "ETF总规模_亿元": round4(scale.sum()),
                "规模小于1亿数量": int((scale < 1).sum()),
                "规模小于2亿数量": int((scale < 2).sum()),
                "规模小于10亿数量": int((scale < 10).sum()),
                "规模小于1亿占比": int((scale < 1).sum()) / count if count else 0,
                "规模小于2亿占比": int((scale < 2).sum()) / count if count else 0,
                "规模小于10亿占比": int((scale < 10).sum()) / count if count else 0,
            }
        )
    manager = add_pct_display(pd.DataFrame(manager_rows).sort_values(["ETF总规模_亿元"], ascending=False).reset_index(drop=True))
    return full, by_type, manager


def quality_checks(main: pd.DataFrame, std: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, list[str]]:
    issues = []
    broad_count = int(main["是否宽基ETF"].eq("是").sum())
    stock_count = int(main["是否股票ETF"].eq("是").sum())
    of_count = int(main["Wind代码"].astype(str).str.upper().str.endswith(".OF").sum())
    missing_critical = main[
        main["基金上市地点"].map(clean_text).eq("")
        | main["上市日期"].map(clean_text).eq("")
        | main["成立日期"].map(clean_text).eq("")
        | main["基金规模_亿元"].isna()
        | main["基金规模_亿元"].le(0)
    ]
    if broad_count == 0:
        issues.append("宽基ETF数量为0，分类失败")
    if stock_count == 0:
        issues.append("股票ETF数量为0，分类失败")
    if of_count:
        issues.append(f"主分析池混入.OF产品{of_count}只")
    if len(missing_critical):
        issues.append(f"主分析池存在关键字段缺失产品{len(missing_critical)}只")

    std_text = std[["基金简称", "基金全称", "跟踪指数", "业绩比较基准"]].fillna("").agg(" ".join, axis=1)
    std_suffix = std["Wind代码"].astype(str).str.upper().str.extract(r"(\.[A-Z]+)$", expand=False).fillna("")
    money_rows = []
    for term in MONEY_REPRESENTATIVE_TERMS:
        candidates = std.loc[std_text.str.contains(term, regex=False, na=False) & std_suffix.isin([".SH", ".SZ"])].copy()
        in_main_codes = set(main["Wind代码"].astype(str))
        for _, row in candidates.iterrows():
            in_main = row["Wind代码"] in in_main_codes
            money_rows.append(
                {
                    "检查关键词": term,
                    "Wind代码": row["Wind代码"],
                    "基金简称": row["基金简称"],
                    "基金全称": row["基金全称"],
                    "是否进入主分析池": "是" if in_main else "否",
                    "入池/观察原因": row.get("入池/观察原因", ""),
                }
            )
    money_check = pd.DataFrame(money_rows)
    if not money_check.empty and money_check["是否进入主分析池"].eq("否").any():
        issues.append("场内货币代表产品存在未进入主分析池的.SH/.SZ产品")

    all_share_mask = main[["基金简称", "基金全称", "跟踪指数", "业绩比较基准"]].fillna("").agg(" ".join, axis=1).map(
        lambda x: bool(contains_terms(x, INDUSTRY_ALL_SHARE_CHECK_TERMS))
    )
    all_share = main.loc[all_share_mask, ["Wind代码", "基金简称", "跟踪指数", "股票ETF内部类型", "产品类型_汇总", "命中关键词", "分类备注"]].copy()
    all_share["是否分类异常"] = all_share["股票ETF内部类型"].eq("宽基ETF").map({True: "是", False: "否"})
    if all_share["是否分类异常"].eq("是").any():
        issues.append("行业全指类产品被误判为宽基ETF")

    text = main[["基金简称", "基金全称", "跟踪指数", "业绩比较基准"]].fillna("").agg(" ".join, axis=1)
    not_broad_rows = []
    for term in NOT_BROAD_CHECK_TERMS:
        sample = main.loc[text.str.contains(term, regex=False, na=False)].copy()
        for _, row in sample.iterrows():
            not_broad_rows.append(
                {
                    "检查关键词": term,
                    "Wind代码": row["Wind代码"],
                    "基金简称": row["基金简称"],
                    "跟踪指数": row["跟踪指数"],
                    "股票ETF内部类型": row["股票ETF内部类型"],
                    "是否分类异常": "是" if row["股票ETF内部类型"] == "宽基ETF" else "否",
                    "分类备注": row["分类备注"],
                }
            )
    not_broad_check = pd.DataFrame(not_broad_rows)
    if not_broad_check.empty:
        not_broad_check = pd.DataFrame(columns=["检查关键词", "Wind代码", "基金简称", "跟踪指数", "股票ETF内部类型", "是否分类异常", "分类备注"])
    if not_broad_check["是否分类异常"].eq("是").any():
        issues.append("策略/行业特征产品被误判为宽基ETF")

    should_broad_rows = []
    for term in SHOULD_BROAD_CHECK_TERMS:
        sample = main.loc[text.str.contains(term, regex=False, na=False)].copy()
        for _, row in sample.iterrows():
            ok = row["股票ETF内部类型"] == "宽基ETF"
            should_broad_rows.append(
                {
                    "检查关键词": term,
                    "Wind代码": row["Wind代码"],
                    "基金简称": row["基金简称"],
                    "跟踪指数": row["跟踪指数"],
                    "股票ETF内部类型": row["股票ETF内部类型"],
                    "是否分类异常": "否" if ok else "是",
                    "分类备注": row["分类备注"],
                }
            )
    should_broad_check = pd.DataFrame(should_broad_rows)
    if should_broad_check.empty:
        should_broad_check = pd.DataFrame(columns=["检查关键词", "Wind代码", "基金简称", "跟踪指数", "股票ETF内部类型", "是否分类异常", "分类备注"])
    if should_broad_check["是否分类异常"].eq("是").any():
        issues.append("应归为宽基或宽基增强的产品未归为宽基ETF")

    sample_rows = []
    for term in CORE_BROAD_CHECK_TERMS:
        sample = main.loc[text.str.contains(term, regex=False, na=False)].head(10).copy()
        for _, row in sample.iterrows():
            sample_rows.append(
                {
                    "命中关键词": term,
                    "Wind代码": row["Wind代码"],
                    "基金简称": row["基金简称"],
                    "跟踪指数": row["跟踪指数"],
                    "股票ETF内部类型": row["股票ETF内部类型"],
                    "产品类型_汇总": row["产品类型_汇总"],
                    "分类备注": row["分类备注"],
                }
            )
    broad_samples = pd.DataFrame(sample_rows)
    return all_share, broad_samples, money_check, not_broad_check, should_broad_check, issues


def note_df(lines: list[str]) -> pd.DataFrame:
    return pd.DataFrame({"序号": range(1, len(lines) + 1), "说明": lines})


def write_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    if path.exists():
        print(f"覆盖已有输出：{path}")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, data in sheets.items():
            data.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    style_workbook(path)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Microsoft YaHei", size=10)
    thin = Side(style="thin", color="D9E1F2")
    border = Border(bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
        for col_idx in range(1, ws.max_column + 1):
            values = [clean_text(ws.cell(row, col_idx).value) for row in range(1, min(ws.max_row, 250) + 1)]
            max_len = max((len(value) for value in values), default=8)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len * 1.15 + 2, 10), 42)
    wb.save(path)


def md_table(df: pd.DataFrame) -> str:
    if df.empty:
        return "（无记录）"
    cols = [str(c) for c in df.columns]
    rows = []
    for _, row in df.iterrows():
        rows.append([clean_text(row.get(c)).replace("|", "\\|") for c in df.columns])
    return "\n".join(
        [
            "| " + " | ".join(cols) + " |",
            "| " + " | ".join(["---"] * len(cols)) + " |",
            *["| " + " | ".join(row) + " |" for row in rows],
        ]
    )


def build_markdown(
    raw_rows: int,
    footer_count: int,
    raw_no_footer: pd.DataFrame,
    main: pd.DataFrame,
    observation: pd.DataFrame,
    field_mapping: pd.DataFrame,
    checks: pd.DataFrame,
    all_share_check: pd.DataFrame,
    money_check: pd.DataFrame,
    not_broad_check: pd.DataFrame,
    should_broad_check: pd.DataFrame,
    broad_samples: pd.DataFrame,
) -> str:
    total_scale = main["基金规模_亿元"].sum()
    asset = group_structure(main, "资产类型")
    product = group_structure(main, "产品类型_汇总")
    stock_inner = group_structure(main[main["资产类型"].eq("股票ETF")], "股票ETF内部类型")
    strategy_main_count = int(main["策略ETF_主分类"].eq("是").sum())
    strategy_tag_count = int(main["是否策略ETF"].eq("是").sum())
    money_count = int(main["资产类型"].eq("货币ETF").sum())
    money_scale = round4(main.loc[main["资产类型"].eq("货币ETF"), "基金规模_亿元"].sum())
    return f"""# ICI课题 P17-P23 中国ETF数据处理说明（二次修正版）

## 本次二次修正原因
- 修正版仍漏掉部分不含英文 ETF 字样、但属于交易所上市交易的场内货币基金。
- 修正版对 A500红利低波、沪深300质量、中证500低波、中证500信息技术等产品存在宽基过度识别风险。
- 本次将 `策略ETF_主分类` 与 `策略增强标签` 拆开，避免 P18 主图口径和 P21 补充标签口径混用。

## 路径
- 原始文件路径：`{INPUT_PATH}`
- 输出路径：`{OUTPUT_DIR}`

## 原始数据结构
- 原始 Sheet：Sheet1
- 原始行数：{raw_rows}
- 剔除 footer 行数：{footer_count}
- 剔除 footer 后产品记录：{len(raw_no_footer)}

## 剔除 footer 的逻辑
- 若 `Wind代码` 以“数据来源”开头，则视为数据来源说明行，不作为产品记录。

## 为什么主分析池只使用 .SH/.SZ
- P17-P23 需要反映全市场上市交易 ETF / 交易型开放式基金结构，`.SH` 和 `.SZ` 对应沪深交易所上市交易产品。
- `.OF` 后缀产品可能为场外、待上市或缺少交易信息产品，不能与上市交易 ETF 混算。

## 为什么场内货币基金要纳入
- 银华日利、华宝添益、现金添益、保证金、收益宝、财富宝等产品虽然不一定在简称中出现英文 ETF，但属于上市交易的交易型货币市场基金。
- 本次主池在满足 .SH/.SZ、上市地点、上市日期、成立日期、规模等硬条件后，将货币市场型基金和场内货币关键词产品纳入。

## .OF 产品如何处理
- `.OF` 后缀、关键字段缺失、或名称包含 ETF 但未进入主分析池的产品，统一进入“未上市或待上市观察池”。
- 观察池仅用于复核和说明，不进入 P17-P23 主图计算。

## 字段映射
{md_table(field_mapping)}

## 上市ETF分析池筛选口径
- Wind代码后缀为 `.SH` 或 `.SZ`；
- 基金上市地点、上市日期、成立日期均非空；
- 基金规模_亿元非空且大于 0；
- 基金简称/全称包含 ETF，或基金全称包含交易型开放式/交易型货币市场基金，或投资类型为货币市场型基金，或命中场内货币关键词；
- 排除 ETF联接、ETF连接、联接A、联接C、场外联接、FOF、LOF 等非交易型 ETF。

## 资产类型分类规则
- 互斥优先级：货币ETF > 商品ETF > 债券ETF > 跨境ETF > 股票ETF > 其他ETF。
- 港股通、恒生、QDII、美国、纳斯达克等归为跨境ETF。
- 有色金属、煤炭、钢铁、黄金股、矿业等股票行业主题不因名称含商品相关词而归为商品ETF，除非投资类型明确为商品型基金。

## 股票ETF内部类型分类规则
- 仅对资产类型为股票ETF的产品分类。
- 内部类型包括：宽基ETF、行业主题ETF、策略ETF、其他股票ETF。
- 分类优先级为行业主题ETF > 策略ETF > 宽基ETF > 其他股票ETF。
- 若“增强策略/指数增强”的底层是核心宽基，内部类型按宽基ETF，同时 `是否增强策略=是`、`策略增强标签=增强策略`。

## 策略ETF主分类与策略/增强标签的区别
- `策略ETF_主分类=是` 仅表示股票ETF内部类型为策略ETF，用于 P18 产品结构主图。
- `策略增强标签` 用于补充标记增强策略、Smart Beta/策略因子、行业主题、宽基等属性。
- `是否策略ETF=是` 表示策略主分类，或策略增强标签为增强策略 / Smart Beta / 策略因子，用于 P21 补充说明，不直接作为 P18 主图互斥分类。

## 宽基ETF识别规则
- 宽基ETF只识别纯宽基或宽基增强策略。
- 不将“中证全指”单独作为宽基关键词。
- 若命中行业主题或策略关键词，不归宽基，除非是增强策略/指数增强且底层为核心宽基。

## 行业主题ETF优先识别规则
- 信息技术、医药、消费、证券、银行、半导体、食品、电力、软件、通信、科技、互联网等行业主题词优先于宽基词。
- 中证500信息技术、中证500医药、中证500消费、中证全指证券公司、中证全指软件等不得归为宽基ETF。

## 策略ETF优先识别规则
- 红利、低波、质量、价值、成长、现金流、股息、等权、ESG、基本面等策略因子优先于宽基词。
- A500红利低波、沪深300质量、中证500低波、中证500质量成长、300价值、500成长等不得归为宽基ETF。

## 分类质量检查
{md_table(checks)}

## 场内货币纳入检查
{md_table(money_check.head(80))}

## 行业全指复核
{md_table(all_share_check.head(30))}

## 不得归宽基样本复核
{md_table(not_broad_check.head(50))}

## 应归宽基样本复核
{md_table(should_broad_check.head(50))}

## 核心宽基关键词样本
{md_table(broad_samples.head(50))}

## 每个输出文件对应 PPT 页
- P17：{OUTPUT_FILES["P17"]}
- P18：{OUTPUT_FILES["P18"]}
- P19：{OUTPUT_FILES["P19"]}
- P20：{OUTPUT_FILES["P20"]}
- P21：{OUTPUT_FILES["P21"]}
- P22-P23：{OUTPUT_FILES["P22P23"]}
- 汇总作图数据：{OUTPUT_FILES["ARCHIVE"]}

## P17-P23 初步描述性结论
- 上市ETF分析池共 {len(main)} 只，总规模约 {round4(total_scale)} 亿元。
- 货币ETF共 {money_count} 只，总规模约 {money_scale} 亿元，本次已纳入场内货币基金。
- 资产类型结构：{'; '.join(f"{r['资产类型']} {r['ETF数量']}只、{r['总规模_亿元']}亿元" for _, r in asset.iterrows())}。
- 产品类型汇总结构：{'; '.join(f"{r['产品类型_汇总']} {r['ETF数量']}只、{r['总规模_亿元']}亿元" for _, r in product.iterrows())}。
- 股票ETF内部结构：{'; '.join(f"{r['股票ETF内部类型']} {r['ETF数量']}只、{r['总规模_亿元']}亿元" for _, r in stock_inner.iterrows())}。
- 策略ETF主分类 {strategy_main_count} 只，策略/增强标签产品 {strategy_tag_count} 只，两者口径不同。
- 观察池共 {len(observation)} 只，不参与主图计算。
- 参考 PPT 主要是股票型 ETF 口径，且截止日期可能为 2025/10/31；本次 ICI 课题 P17-P23 主图数据以二次修正版的全市场上市 ETF 分析池为准。

## 分类复核建议
- 复核“增强策略 + 宽基指数”的双重标签产品。
- 复核行业全指、央企/国企主题、黄金股/矿业/有色金属等边界产品。
- 如后续需要与参考 PPT 对比，请保持“全市场上市ETF口径”和“股票型ETF口径”分开。

## 美国侧提示
- 美国数据来自 ICI 2026 Fact Book：ETF数量 4,495只，ETF总净资产 13.4万亿美元。
"""


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"输入文件不存在：{INPUT_PATH}")

    raw_sheets, raw_no_footer, std, raw_rows = read_and_standardize()
    footer_count = raw_rows - len(raw_no_footer)
    main_pool, observation_pool, screening_process = finalize_pools(std)

    # Enforce numeric types after filtering.
    for frame in [main_pool, observation_pool]:
        frame["基金规模_亿元"] = pd.to_numeric(frame["基金规模_亿元"], errors="coerce")
        frame["管理费率"] = pd.to_numeric(frame["管理费率"], errors="coerce")
        frame["托管费率"] = pd.to_numeric(frame["托管费率"], errors="coerce")

    all_share_check, broad_samples, money_check, not_broad_check, should_broad_check, issues = quality_checks(main_pool, std)
    quality_rows = [
        {"检查项": "宽基ETF数量是否为0", "结果": "通过" if main_pool["是否宽基ETF"].eq("是").sum() > 0 else "失败", "数值": int(main_pool["是否宽基ETF"].eq("是").sum()), "说明": ""},
        {"检查项": "股票ETF数量是否明显大于0", "结果": "通过" if main_pool["是否股票ETF"].eq("是").sum() > 0 else "失败", "数值": int(main_pool["是否股票ETF"].eq("是").sum()), "说明": ""},
        {"检查项": "是否有.OF混入主分析池", "结果": "通过" if not main_pool["Wind代码"].astype(str).str.upper().str.endswith(".OF").any() else "失败", "数值": int(main_pool["Wind代码"].astype(str).str.upper().str.endswith(".OF").sum()), "说明": ""},
        {"检查项": "是否有缺失日期/规模产品混入主分析池", "结果": "通过" if not issues or not any("关键字段缺失" in x for x in issues) else "失败", "数值": "", "说明": ""},
        {"检查项": "行业全指类是否误判宽基", "结果": "通过" if not all_share_check["是否分类异常"].eq("是").any() else "失败", "数值": int(all_share_check["是否分类异常"].eq("是").sum()) if len(all_share_check) else 0, "说明": ""},
        {"检查项": "场内货币代表产品是否入池", "结果": "通过" if money_check.empty or not money_check["是否进入主分析池"].eq("否").any() else "失败", "数值": int(money_check["是否进入主分析池"].eq("是").sum()) if len(money_check) else 0, "说明": "检查银华日利、华宝添益、现金添益、保证金、场内货币、收益宝、财富宝"},
        {"检查项": "不得归宽基样本是否误判", "结果": "通过" if not_broad_check.empty or not not_broad_check["是否分类异常"].eq("是").any() else "失败", "数值": int(not_broad_check["是否分类异常"].eq("是").sum()) if len(not_broad_check) else 0, "说明": ""},
        {"检查项": "应归宽基样本是否正确", "结果": "通过" if should_broad_check.empty or not should_broad_check["是否分类异常"].eq("是").any() else "失败", "数值": int(should_broad_check["是否分类异常"].eq("是").sum()) if len(should_broad_check) else 0, "说明": ""},
    ]
    checks = pd.DataFrame(quality_rows)
    if issues:
        raise RuntimeError("；".join(issues))

    missing = pd.DataFrame(
        [
            {
                "字段名": col,
                "缺失数量": int(main_pool[col].isna().sum() + main_pool[col].map(clean_text).eq("").sum()),
                "缺失比例": (main_pool[col].isna().sum() + main_pool[col].map(clean_text).eq("").sum()) / len(main_pool),
            }
            for col in BASE_COLUMNS
        ]
    )
    missing = add_pct_display(missing)
    duplicates = main_pool[main_pool.duplicated("Wind代码", keep=False)][["Wind代码", "基金简称", "基金管理人"]].copy()
    if duplicates.empty:
        duplicates = pd.DataFrame(columns=["Wind代码", "基金简称", "基金管理人", "重复次数", "备注"])
    else:
        duplicates["重复次数"] = duplicates["Wind代码"].map(main_pool["Wind代码"].value_counts())
        duplicates["备注"] = "Wind代码重复，需要复核"

    sheet_overview = pd.DataFrame(
        [{"sheet名称": name, "行数": len(df), "列数": len(df.columns), "是否读取成功": "是", "备注": ""} for name, df in raw_sheets.items()]
    )
    pool_check = pd.DataFrame(
        [
            {"检查项": "Wind代码后缀仅.SH/.SZ", "结果": "是", "数值": int(main_pool["Wind代码"].astype(str).str.contains(r"\\.(?:SH|SZ)$", regex=True).sum()), "说明": ""},
            {"检查项": "上市日期非空", "结果": "是", "数值": int(main_pool["上市日期"].map(clean_text).ne("").sum()), "说明": ""},
            {"检查项": "成立日期非空", "结果": "是", "数值": int(main_pool["成立日期"].map(clean_text).ne("").sum()), "说明": ""},
            {"检查项": "规模大于0", "结果": "是", "数值": int(main_pool["基金规模_亿元"].gt(0).sum()), "说明": ""},
            {"检查项": "观察池数量", "结果": "仅复核", "数值": len(observation_pool), "说明": ".OF、缺失关键交易信息或名称含ETF但未入主池"},
        ]
    )

    total_scale = main_pool["基金规模_亿元"].sum()
    desc = pd.DataFrame(
        [
            {"指标": "原始行数", "数值": raw_rows, "单位": "行", "说明": ""},
            {"指标": "剔除 footer 行数", "数值": footer_count, "单位": "行", "说明": "Wind数据来源说明行"},
            {"指标": "剔除 footer 后产品数", "数值": len(raw_no_footer), "单位": "只", "说明": ""},
            {"指标": "ETF候选产品数", "数值": int((std["基金简称"].fillna("") + std["基金全称"].fillna("")).str.contains("ETF|交易型开放式|交易型货币市场基金", case=False, regex=True, na=False).sum()), "单位": "只", "说明": "名称含ETF/交易型开放式/交易型货币市场基金"},
            {"指标": "上市ETF分析池数量", "数值": len(main_pool), "单位": "只", "说明": ""},
            {"指标": "观察池数量", "数值": len(observation_pool), "单位": "只", "说明": ""},
            {"指标": "ETF总规模_亿元", "数值": round4(total_scale), "单位": "亿元", "说明": ""},
            {"指标": "成立日期最早年份", "数值": pd.to_datetime(main_pool["成立日期"]).dt.year.min(), "单位": "年", "说明": ""},
            {"指标": "成立日期最新年份", "数值": pd.to_datetime(main_pool["成立日期"]).dt.year.max(), "单位": "年", "说明": ""},
            {"指标": "管理人数量", "数值": main_pool["基金管理人"].nunique(), "单位": "家", "说明": ""},
            {"指标": "资产类型数量分布", "数值": "；".join(f"{k}:{v}" for k, v in main_pool["资产类型"].value_counts().items()), "单位": "只", "说明": ""},
            {"指标": "资产类型规模分布", "数值": "；".join(f"{k}:{round4(v)}" for k, v in main_pool.groupby("资产类型")["基金规模_亿元"].sum().items()), "单位": "亿元", "说明": ""},
            {"指标": "股票ETF内部类型数量分布", "数值": "；".join(f"{k}:{v}" for k, v in main_pool.loc[main_pool["资产类型"].eq("股票ETF"), "股票ETF内部类型"].value_counts().items()), "单位": "只", "说明": ""},
            {"指标": "股票ETF内部类型规模分布", "数值": "；".join(f"{k}:{round4(v)}" for k, v in main_pool.loc[main_pool["资产类型"].eq("股票ETF")].groupby("股票ETF内部类型")["基金规模_亿元"].sum().items()), "单位": "亿元", "说明": ""},
            {"指标": "产品类型_汇总数量分布", "数值": "；".join(f"{k}:{v}" for k, v in main_pool["产品类型_汇总"].value_counts().items()), "单位": "只", "说明": ""},
            {"指标": "产品类型_汇总规模分布", "数值": "；".join(f"{k}:{round4(v)}" for k, v in main_pool.groupby("产品类型_汇总")["基金规模_亿元"].sum().items()), "单位": "亿元", "说明": ""},
            {"指标": "宽基ETF数量", "数值": int(main_pool["是否宽基ETF"].eq("是").sum()), "单位": "只", "说明": ""},
            {"指标": "行业主题ETF数量", "数值": int(main_pool["是否行业主题ETF"].eq("是").sum()), "单位": "只", "说明": ""},
            {"指标": "策略ETF主分类数量", "数值": int(main_pool["策略ETF_主分类"].eq("是").sum()), "单位": "只", "说明": ""},
            {"指标": "策略/增强标签产品数量", "数值": int(main_pool["是否策略ETF"].eq("是").sum()), "单位": "只", "说明": ""},
            {"指标": "货币ETF数量", "数值": int(main_pool["资产类型"].eq("货币ETF").sum()), "单位": "只", "说明": ""},
            {"指标": "货币ETF规模", "数值": round4(main_pool.loc[main_pool["资产类型"].eq("货币ETF"), "基金规模_亿元"].sum()), "单位": "亿元", "说明": ""},
        ]
    )

    p17_all = annual_table(main_pool)
    p17_asset = annual_table(main_pool, "资产类型")
    p17_product = annual_table(main_pool, "产品类型_汇总")
    p18_asset = group_structure(main_pool, "资产类型")
    p18_product = group_structure(main_pool, "产品类型_汇总")
    p18_stock = group_structure(main_pool[main_pool["资产类型"].eq("股票ETF")], "股票ETF内部类型")
    total_count = len(main_pool)
    total_scale = main_pool["基金规模_亿元"].sum()
    label_rows = []
    for label, part in main_pool.groupby("策略增强标签", dropna=False):
        label_rows.append(
            {
                "策略增强标签": label,
                "产品数量": len(part),
                "数量占比": len(part) / total_count if total_count else 0,
                "总规模_亿元": round4(part["基金规模_亿元"].sum()),
                "规模占比": part["基金规模_亿元"].sum() / total_scale if total_scale else 0,
            }
        )
    p18_strategy_label = add_pct_display(pd.DataFrame(label_rows).sort_values(["总规模_亿元", "产品数量"], ascending=[False, False]).reset_index(drop=True))
    sample_fields = ["Wind代码", "基金简称", "基金管理人", "资产类型", "股票ETF内部类型", "产品类型_汇总", "策略ETF_主分类", "策略增强标签", "是否增强策略", "是否策略ETF", "基金规模_亿元", "成立日期", "跟踪指数", "投资类型", "命中关键词", "分类备注"]
    sample_parts = []
    for _, part in main_pool.groupby("产品类型_汇总"):
        sample_parts.append(part.sample(min(len(part), 20), random_state=20260629))
    p18_sample = pd.concat(sample_parts, ignore_index=True)[sample_fields] if sample_parts else pd.DataFrame(columns=sample_fields)
    p19_rank, p19_cr = manager_tables(main_pool)
    p20_full, p20_type, p20_manager = size_tables(main_pool)

    cr_lookup = p19_cr.set_index("指标")
    metric_rows = [
        {"指标": "ETF总数量", "数值": len(main_pool), "单位": "只", "说明": "上市ETF分析池"},
        {"指标": "ETF总规模", "数值": round4(total_scale), "单位": "亿元", "说明": ""},
    ]
    for asset in ["股票ETF", "债券ETF", "商品ETF", "货币ETF", "跨境ETF"]:
        part = main_pool[main_pool["资产类型"].eq(asset)]
        metric_rows += [
            {"指标": f"{asset}数量", "数值": len(part), "单位": "只", "说明": ""},
            {"指标": f"{asset}规模", "数值": round4(part["基金规模_亿元"].sum()), "单位": "亿元", "说明": ""},
            {"指标": f"{asset}规模占比", "数值": part["基金规模_亿元"].sum() / total_scale if total_scale else 0, "数值展示": pct(part["基金规模_亿元"].sum() / total_scale if total_scale else 0), "单位": "%", "说明": ""},
        ]
    for product in ["宽基ETF", "行业主题ETF"]:
        part = main_pool[main_pool["股票ETF内部类型"].eq(product)]
        product_note = "股票ETF内部类型"
        metric_rows += [
            {"指标": f"{product}数量", "数值": len(part), "单位": "只", "说明": product_note},
            {"指标": f"{product}规模", "数值": round4(part["基金规模_亿元"].sum()), "单位": "亿元", "说明": product_note},
            {"指标": f"{product}规模占比", "数值": part["基金规模_亿元"].sum() / total_scale if total_scale else 0, "数值展示": pct(part["基金规模_亿元"].sum() / total_scale if total_scale else 0), "单位": "%", "说明": ""},
        ]
    strategy_main = main_pool[main_pool["策略ETF_主分类"].eq("是")]
    strategy_tag = main_pool[main_pool["是否策略ETF"].eq("是")]
    metric_rows += [
        {"指标": "策略ETF主分类数量", "数值": len(strategy_main), "单位": "只", "说明": "股票ETF内部类型=策略ETF"},
        {"指标": "策略ETF主分类规模", "数值": round4(strategy_main["基金规模_亿元"].sum()), "单位": "亿元", "说明": "股票ETF内部类型=策略ETF"},
        {"指标": "策略ETF主分类规模占比", "数值": strategy_main["基金规模_亿元"].sum() / total_scale if total_scale else 0, "数值展示": pct(strategy_main["基金规模_亿元"].sum() / total_scale if total_scale else 0), "单位": "%", "说明": ""},
        {"指标": "策略/增强标签产品数量", "数值": len(strategy_tag), "单位": "只", "说明": "策略主分类 + 增强策略 / Smart Beta标签"},
        {"指标": "策略/增强标签产品规模", "数值": round4(strategy_tag["基金规模_亿元"].sum()), "单位": "亿元", "说明": "策略主分类 + 增强策略 / Smart Beta标签"},
        {"指标": "策略/增强标签产品规模占比", "数值": strategy_tag["基金规模_亿元"].sum() / total_scale if total_scale else 0, "数值展示": pct(strategy_tag["基金规模_亿元"].sum() / total_scale if total_scale else 0), "单位": "%", "说明": ""},
    ]
    multi_asset = main_pool[~main_pool["资产类型"].eq("股票ETF")]
    metric_rows += [
        {"指标": "前五大管理人规模占比", "数值": cr_lookup.loc["CR5", "规模口径"], "数值展示": pct(cr_lookup.loc["CR5", "规模口径"]), "单位": "%", "说明": ""},
        {"指标": "前十大管理人规模占比", "数值": cr_lookup.loc["CR10", "规模口径"], "数值展示": pct(cr_lookup.loc["CR10", "规模口径"]), "单位": "%", "说明": ""},
        {"指标": "规模小于1亿产品占比", "数值": (main_pool["基金规模_亿元"] < 1).mean(), "数值展示": pct((main_pool["基金规模_亿元"] < 1).mean()), "单位": "%", "说明": ""},
        {"指标": "规模小于2亿产品占比", "数值": (main_pool["基金规模_亿元"] < 2).mean(), "数值展示": pct((main_pool["基金规模_亿元"] < 2).mean()), "单位": "%", "说明": ""},
        {"指标": "规模小于10亿产品占比", "数值": (main_pool["基金规模_亿元"] < 10).mean(), "数值展示": pct((main_pool["基金规模_亿元"] < 10).mean()), "单位": "%", "说明": ""},
        {"指标": "多资产ETF数量", "数值": len(multi_asset), "单位": "只", "说明": "非股票ETF"},
        {"指标": "多资产ETF规模", "数值": round4(multi_asset["基金规模_亿元"].sum()), "单位": "亿元", "说明": "非股票ETF"},
        {"指标": "多资产ETF规模占比", "数值": multi_asset["基金规模_亿元"].sum() / total_scale if total_scale else 0, "数值展示": pct(multi_asset["基金规模_亿元"].sum() / total_scale if total_scale else 0), "单位": "%", "说明": "非股票ETF"},
        {"指标": "平均管理费率", "数值": round4(main_pool["管理费率"].mean()), "单位": "%", "说明": "原始费率字段单位为%"},
        {"指标": "平均托管费率", "数值": round4(main_pool["托管费率"].mean()), "单位": "%", "说明": "原始费率字段单位为%"},
    ]
    p21 = pd.DataFrame(metric_rows)
    if "数值展示" not in p21.columns:
        p21["数值展示"] = ""

    p22 = pd.DataFrame(
        [
            {"转向": "发行导向 → 持营导向", "支撑指标": "规模小于10亿产品占比", "数值": (main_pool["基金规模_亿元"] < 10).mean(), "数值展示": pct((main_pool["基金规模_亿元"] < 10).mean()), "单位": "%", "解释": "长尾产品占比体现持营压力。"},
            {"转向": "指数覆盖 → 场景覆盖", "支撑指标": "行业主题ETF+策略ETF数量占比", "数值": main_pool["产品类型_汇总"].isin(["行业主题ETF", "策略ETF"]).mean(), "数值展示": pct(main_pool["产品类型_汇总"].isin(["行业主题ETF", "策略ETF"]).mean()), "单位": "%", "解释": "行业主题和策略产品体现从指数覆盖走向投资场景覆盖。"},
            {"转向": "费率竞争 → 生态竞争", "支撑指标": "平均管理费率", "数值": round4(main_pool["管理费率"].mean()), "数值展示": "", "单位": "%", "解释": "低费率只是基础，长尾与头部集中说明生态能力更关键。"},
            {"转向": "单品销售 → 组合解决方案", "支撑指标": "多资产ETF规模占比", "数值": multi_asset["基金规模_亿元"].sum() / total_scale if total_scale else 0, "数值展示": pct(multi_asset["基金规模_亿元"].sum() / total_scale if total_scale else 0), "单位": "%", "解释": "债券、商品、跨境等非股票ETF支持组合配置工具箱。"},
        ]
    )
    p23 = pd.DataFrame(
        [
            {"指标": "ETF总数量", "数值": len(main_pool), "单位": "只", "可用于哪一句结论": "用于说明中国上市ETF市场供给扩张"},
            {"指标": "ETF总规模", "数值": round4(total_scale), "单位": "亿元", "可用于哪一句结论": "用于说明中国ETF市场已有规模基础"},
            {"指标": "前十大管理人规模占比", "数值": cr_lookup.loc["CR10", "规模口径"], "数值展示": pct(cr_lookup.loc["CR10", "规模口径"]), "单位": "%", "可用于哪一句结论": "用于说明头部集中"},
            {"指标": "小规模产品占比", "数值": (main_pool["基金规模_亿元"] < 10).mean(), "数值展示": pct((main_pool["基金规模_亿元"] < 10).mean()), "单位": "%", "可用于哪一句结论": "用于说明持营压力"},
            {"指标": "股票ETF规模占比", "数值": main_pool.loc[main_pool["资产类型"].eq("股票ETF"), "基金规模_亿元"].sum() / total_scale, "数值展示": pct(main_pool.loc[main_pool["资产类型"].eq("股票ETF"), "基金规模_亿元"].sum() / total_scale), "单位": "%", "可用于哪一句结论": "用于说明权益属性较强"},
            {"指标": "多资产ETF规模占比", "数值": multi_asset["基金规模_亿元"].sum() / total_scale if total_scale else 0, "数值展示": pct(multi_asset["基金规模_亿元"].sum() / total_scale if total_scale else 0), "单位": "%", "可用于哪一句结论": "用于说明配置工具箱仍待完善"},
        ]
    )

    write_excel(
        OUTPUT_DIR / OUTPUT_FILES["00"],
        {
            "sheet概览": sheet_overview,
            "字段映射表": pd.DataFrame([{"标准字段名": v, "原始字段名": k, "所在sheet": "Sheet1", "识别方式": "字段名精确映射", "备注": ""} for k, v in RAW_TO_STD.items()]),
            "主分析池筛选过程": screening_process,
            "主分析池口径检查": pool_check,
            "ETF分析池_上市交易_二次修正": main_pool,
            "ETF观察池_OF或缺失交易信息_二次修正": observation_pool,
            "缺失值检查": missing,
            "重复值检查": duplicates,
            "分类结果检查": pd.concat([checks, pd.DataFrame([{"检查项": "行业全指复核样本数", "结果": "见分类异常复核", "数值": len(all_share_check), "说明": ""}, {"检查项": "核心宽基关键词样本数", "结果": "见分类异常复核", "数值": len(broad_samples), "说明": ""}])], ignore_index=True),
            "分类异常复核": pd.concat(
                [
                    all_share_check.assign(复核类型="行业全指不得误判宽基"),
                    not_broad_check.assign(复核类型="策略/行业特征不得归宽基"),
                    should_broad_check.assign(复核类型="应归宽基或宽基增强"),
                    broad_samples.assign(复核类型="核心宽基关键词样本"),
                ],
                ignore_index=True,
                sort=False,
            ),
            "场内货币纳入检查": money_check,
            "描述性统计": desc,
        },
    )
    write_excel(OUTPUT_DIR / OUTPUT_FILES["P17"], {"P17_年度新发累计_全市场": p17_all, "P17_年度新发累计_按资产类型": p17_asset, "P17_年度新发累计_按产品类型汇总": p17_product})
    write_excel(OUTPUT_DIR / OUTPUT_FILES["P18"], {"P18_资产类型结构": p18_asset, "P18_产品类型汇总结构": p18_product, "P18_股票ETF内部结构": p18_stock, "P18_策略标签补充统计": p18_strategy_label, "P18_分类复核样本": p18_sample})
    write_excel(OUTPUT_DIR / OUTPUT_FILES["P19"], {"P19_管理人规模排名": p19_rank, "P19_前十大管理人": p19_rank.head(10), "P19_CR集中度": p19_cr})
    write_excel(OUTPUT_DIR / OUTPUT_FILES["P20"], {"P20_规模区间分布_全市场": p20_full, "P20_规模区间分布_按产品类型汇总": p20_type, "P20_小规模产品占比_按管理人": p20_manager})
    write_excel(OUTPUT_DIR / OUTPUT_FILES["P21"], {"P21_中国ETF总览指标": p21, "说明": note_df(["美国数据来自 ICI 2026 Fact Book，ETF数量 4,495只，ETF总净资产 13.4万亿美元。"])})
    write_excel(OUTPUT_DIR / OUTPUT_FILES["P22P23"], {"P22_四个转向支撑数据": p22, "P23_最终升华可引用数据": p23})
    write_excel(
        OUTPUT_DIR / OUTPUT_FILES["ARCHIVE"],
        {
            "P17_年度新发累计": p17_all,
            "P17_按资产类型年度新发": p17_asset,
            "P18_资产类型结构": p18_asset,
            "P18_产品类型汇总结构": p18_product,
            "P18_股票ETF内部结构": p18_stock,
            "P18_策略标签补充统计": p18_strategy_label,
            "P19_管理人规模排名": p19_rank,
            "P19_CR集中度": p19_cr,
            "P20_规模区间分布": p20_full,
            "P20_小规模产品占比": p20_manager,
            "P21_中国ETF总览指标": p21,
            "P22_四个转向支撑数据": p22,
            "P23_最终升华引用数据": p23,
        },
    )

    readme = build_markdown(
        raw_rows,
        footer_count,
        raw_no_footer,
        main_pool,
        observation_pool,
        pd.DataFrame([{"标准字段名": v, "原始字段名": k, "所在sheet": "Sheet1", "识别方式": "字段名精确映射", "备注": ""} for k, v in RAW_TO_STD.items()]),
        checks,
        all_share_check,
        money_check,
        not_broad_check,
        should_broad_check,
        broad_samples,
    )
    readme_path = OUTPUT_DIR / OUTPUT_FILES["README"]
    if readme_path.exists():
        print(f"覆盖已有输出：{readme_path}")
    readme_path.write_text(readme, encoding="utf-8")

    generated = [OUTPUT_DIR / name for name in OUTPUT_FILES.values()]
    print("原始文件是否成功读取：是")
    print(f"原始行数：{raw_rows}")
    print(f"剔除 footer 后行数：{len(raw_no_footer)}")
    print(f"上市ETF分析池数量：{len(main_pool)}")
    print(f"观察池数量：{len(observation_pool)}")
    print(f"ETF总规模_亿元：{round4(total_scale)}")
    for label, mask in [
        ("货币ETF", main_pool["资产类型"].eq("货币ETF")),
        ("股票ETF", main_pool["资产类型"].eq("股票ETF")),
        ("宽基ETF", main_pool["是否宽基ETF"].eq("是")),
        ("行业主题ETF", main_pool["是否行业主题ETF"].eq("是")),
        ("策略ETF主分类", main_pool["策略ETF_主分类"].eq("是")),
        ("策略/增强标签产品", main_pool["是否策略ETF"].eq("是")),
    ]:
        part = main_pool.loc[mask]
        print(f"{label}数量与规模：{len(part)} / {round4(part['基金规模_亿元'].sum())} 亿元")
    print(f"是否有 .OF 混入主分析池：{'是' if main_pool['Wind代码'].astype(str).str.upper().str.endswith('.OF').any() else '否'}")
    print(f"是否有缺失日期/规模产品混入主分析池：{'是' if len(main_pool[(main_pool['上市日期'].eq('')) | (main_pool['成立日期'].eq('')) | (main_pool['基金规模_亿元'].isna()) | (main_pool['基金规模_亿元'].le(0))]) else '否'}")
    print(f"场内货币代表产品是否成功纳入：{'是' if money_check.empty or not money_check['是否进入主分析池'].eq('否').any() else '否'}")
    print(f"是否发现分类异常：{'是' if issues else '否'}")
    print("生成的文件清单：")
    for path in generated:
        print(f"- {path}")
    print("分类异常：未发现" if not issues else "分类异常：" + "；".join(issues))


if __name__ == "__main__":
    main()
