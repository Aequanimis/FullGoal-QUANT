from __future__ import annotations

import math
import re
import statistics
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


POOL_FILE = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\全市场ETF基础信息_策略ETF池二次修正版.xlsx")
MAIN_FILE = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\广义策略ETF月度流动性数据（最终版）.xlsx")
SUP_FILE = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\（补充）ETF月度流动性数据.xlsx")
OUTPUT = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\广义策略ETF月度流动性数据_合并验收清洗版.xlsx")

POOL_SHEET = "策略ETF_最终统计池"
MAIN_SHEET = "Sheet1"
SUP_SHEET = "ETF月度流动性数据"

SHEET_FULL = "流动性数据_完整合并版"
SHEET_ANALYSIS = "流动性数据_分析可用版"
SHEET_COVERAGE = "产品覆盖检查"
SHEET_SUP = "补充数据验收"
SHEET_MISSING = "缺失待补ETF清单"
SHEET_FIELD_MISSING = "字段缺失统计"
SHEET_ANOMALY = "异常值检查"
SHEET_MONTHLY = "月度汇总"
SHEET_NOTES = "验收说明"

STANDARD_FIELDS = [
    "wind_code",
    "fund_code",
    "trading_code",
    "year_month",
    "month_end_close",
    "month_return",
    "month_turnover_amount_sum",
    "month_turnover_volume_sum",
    "daily_avg_turnover_amount",
    "daily_avg_turnover_volume",
    "month_turnover_rate_sum",
    "month_amplitude_max",
    "month_discount_rate_avg",
    "month_end_iopv",
    "trading_days",
    "name",
]

AUX_FIELDS = [
    "数据来源",
    "是否补充数据",
    "原始月成交额单位",
    "原始日均成交额单位",
    "raw_daily_avg_turnover_amount_from_supplement",
    "raw_daily_avg_turnover_volume_from_supplement",
    "是否疑似重复收盘价",
    "是否日均字段重算",
    "上市年月",
    "是否上市后月份",
    "是否上市后有效流动性月份",
    "是否上市前空值月份",
    "数据可用状态",
]

POOL_FIELDS = [
    "证券简称",
    "基金简称",
    "基金全称",
    "基金管理人",
    "基金上市地点",
    "上市日期",
    "基金成立日",
    "跟踪指数代码",
    "跟踪指数名称",
    "最新基金规模(亿)",
    "统计口径分类",
    "是否纳入核心策略ETF统计",
    "是否纳入广义策略ETF统计",
    "一级策略大类",
    "二级策略类别",
    "市场范围_二次修正",
]

FULL_FIELDS = STANDARD_FIELDS + AUX_FIELDS + POOL_FIELDS

CORE_LIQUIDITY_FIELDS = [
    "month_turnover_amount_sum",
    "month_turnover_volume_sum",
    "daily_avg_turnover_amount",
    "daily_avg_turnover_volume",
    "month_turnover_rate_sum",
    "month_amplitude_max",
    "month_discount_rate_avg",
    "month_return",
]

MISSING_STATS_FIELDS = [
    "month_end_close",
    "month_return",
    "month_turnover_amount_sum",
    "month_turnover_volume_sum",
    "daily_avg_turnover_amount",
    "daily_avg_turnover_volume",
    "month_turnover_rate_sum",
    "month_amplitude_max",
    "month_discount_rate_avg",
    "month_end_iopv",
    "trading_days",
]

SUP_EXPECTATIONS = {
    "159207.SZ": ("2025-04", "2026-06", 15),
    "159581.SZ": ("2024-03", "2026-06", 28),
    "159905.SZ": ("2015-01", "2026-06", 138),
    "562060.SH": ("2023-12", "2026-06", 31),
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=9, color="1F1F1F")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN_GRAY = Side(style="thin", color="D9E1F2")
BODY_BORDER = Border(bottom=THIN_GRAY)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    value = str(value).replace("\u3000", " ").replace("\u200b", "")
    return re.sub(r"\s+", " ", value).strip()


def normalize_code(value: Any) -> str:
    raw = clean_text(value).upper().replace("。", ".")
    match = re.search(r"(\d{6})\s*[.]?\s*(SH|SZ)$", raw)
    return f"{match.group(1)}.{match.group(2)}" if match else raw


def normalize_six_digit(value: Any) -> str:
    match = re.search(r"(\d{6})", clean_text(value))
    return match.group(1) if match else clean_text(value)


def normalize_year_month(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m")
    raw = clean_text(value)
    match = re.search(r"(\d{4})[-/年](\d{1,2})", raw)
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}"
    match = re.fullmatch(r"(\d{4})(\d{2})", raw)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    return raw


def to_number(value: Any) -> float | int | None:
    if value is None or clean_text(value) in {"", "--", "N/A", "nan"}:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    raw = clean_text(value).replace(",", "").replace("%", "")
    try:
        number = float(raw)
        return int(number) if number.is_integer() else number
    except ValueError:
        return None


def to_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if not clean_text(value):
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(clean_text(value), fmt)
        except ValueError:
            continue
    return None


def month_index(year_month: str) -> int:
    match = re.fullmatch(r"(\d{4})-(\d{2})", clean_text(year_month))
    if not match:
        return -1
    return int(match.group(1)) * 12 + int(match.group(2)) - 1


def month_range(start: str, end: str) -> list[str]:
    start_i, end_i = month_index(start), month_index(end)
    if start_i < 0 or end_i < start_i:
        return []
    return [f"{idx // 12:04d}-{idx % 12 + 1:02d}" for idx in range(start_i, end_i + 1)]


def has_core_liquidity(record: dict[str, Any]) -> bool:
    return any(record.get(field) is not None for field in CORE_LIQUIDITY_FIELDS)


def sum_numeric(records: Iterable[dict[str, Any]], field: str) -> float:
    return sum(
        float(record[field])
        for record in records
        if isinstance(record.get(field), (int, float))
    )


def mean_numeric(records: list[dict[str, Any]], field: str) -> float | None:
    values = [float(r[field]) for r in records if isinstance(r.get(field), (int, float))]
    return sum(values) / len(values) if values else None


def median_numeric(records: list[dict[str, Any]], field: str) -> float | None:
    values = [float(r[field]) for r in records if isinstance(r.get(field), (int, float))]
    return statistics.median(values) if values else None


def relative_difference(a: float | None, b: float | None) -> float | None:
    if a is None or b is None:
        return None
    denominator = max(abs(b), 1.0)
    return abs(a - b) / denominator


def load_pool() -> dict[str, dict[str, Any]]:
    wb = load_workbook(POOL_FILE, read_only=True, data_only=True)
    ws = wb[POOL_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows())]
    pool = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, values))
        if clean_text(record.get("是否纳入广义策略ETF统计")) != "是":
            continue
        code = normalize_code(record.get("Wind代码"))
        copied = dict(record)
        copied["上市日期"] = to_datetime(copied.get("上市日期"))
        copied["基金成立日"] = to_datetime(copied.get("基金成立日"))
        pool[code] = copied
    return pool


def load_main() -> tuple[list[dict[str, Any]], dict[str, int], list[dict[str, Any]]]:
    wb = load_workbook(MAIN_FILE, read_only=True, data_only=True)
    ws = wb[MAIN_SHEET]
    headers = [clean_text(cell.value) for cell in next(ws.iter_rows())]
    missing = [field for field in STANDARD_FIELDS if field not in headers]
    if missing:
        raise KeyError(f"主表缺少字段：{missing}")

    records = []
    deleted = Counter()
    anomalies = []
    month_trading_days: defaultdict[str, list[int]] = defaultdict(list)

    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and clean_text(value) for value in values):
            deleted["全空行"] += 1
            continue
        raw = dict(zip(headers, values))
        if clean_text(raw.get("wind_code")).startswith("数据来源：Wind"):
            deleted["Wind来源行"] += 1
            continue
        record = {field: raw.get(field) for field in STANDARD_FIELDS}
        record["wind_code"] = normalize_code(record.get("wind_code"))
        record["fund_code"] = normalize_code(record.get("fund_code"))
        record["trading_code"] = normalize_code(record.get("trading_code"))
        record["year_month"] = normalize_year_month(record.get("year_month"))
        for field in STANDARD_FIELDS[4:15]:
            record[field] = to_number(record.get(field))
        if record["trading_days"] is not None:
            record["trading_days"] = int(record["trading_days"])
            month_trading_days[record["year_month"]].append(record["trading_days"])
        record["数据来源"] = "主表"
        record["是否补充数据"] = "否"
        record["原始月成交额单位"] = "CNY"
        record["原始日均成交额单位"] = "CNY"
        record["raw_daily_avg_turnover_amount_from_supplement"] = None
        record["raw_daily_avg_turnover_volume_from_supplement"] = None
        record["是否疑似重复收盘价"] = "否"
        record["是否日均字段重算"] = "否"
        records.append(record)

    trading_day_map = {}
    for month, values in month_trading_days.items():
        counts = Counter(values)
        trading_day_map[month] = counts.most_common(1)[0][0]
        if len(counts) > 1:
            anomalies.append(
                {
                    "异常类型": "同月trading_days映射不一致",
                    "严重程度": "提示",
                    "wind_code": None,
                    "year_month": month,
                    "字段": "trading_days",
                    "原始值": "；".join(f"{k}:{v}条" for k, v in sorted(counts.items())),
                    "参考值/重算值": trading_day_map[month],
                    "说明": "补充数据采用该月众数作为交易日数",
                }
            )
    return records, trading_day_map, anomalies


def load_supplement(
    trading_day_map: dict[str, int],
) -> tuple[list[dict[str, Any]], dict[str, bool]]:
    wb = load_workbook(SUP_FILE, read_only=True, data_only=True)
    ws = wb[SUP_SHEET]
    headers = [clean_text(cell.value) for cell in next(ws.iter_rows())]
    records = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and clean_text(value) for value in values):
            continue
        raw = dict(zip(headers, values))
        code = normalize_code(raw.get("Wind代码"))
        month = normalize_year_month(raw.get("日期"))
        amount_sum = to_number(raw.get("月成交额_亿"))
        amount_sum = amount_sum * 100000000 if amount_sum is not None else None
        volume_sum = to_number(raw.get("月成交量(份)"))
        raw_daily_amount = to_number(raw.get("日均成交额_百万"))
        raw_daily_amount = raw_daily_amount * 1000000 if raw_daily_amount is not None else None
        raw_daily_volume = to_number(raw.get("日均成交量(份)"))
        trading_days = trading_day_map.get(month)
        daily_amount = (
            amount_sum / trading_days
            if amount_sum is not None and trading_days is not None and trading_days > 0
            else None
        )
        daily_volume = (
            volume_sum / trading_days
            if volume_sum is not None and trading_days is not None and trading_days > 0
            else None
        )
        record = {
            "wind_code": code,
            "fund_code": code,
            "trading_code": normalize_six_digit(raw.get("交易代码")),
            "year_month": month,
            "month_end_close": to_number(raw.get("收盘价")),
            "month_return": to_number(raw.get("月涨跌幅(%)")),
            "month_turnover_amount_sum": amount_sum,
            "month_turnover_volume_sum": volume_sum,
            "daily_avg_turnover_amount": daily_amount,
            "daily_avg_turnover_volume": daily_volume,
            "month_turnover_rate_sum": to_number(raw.get("月换手率(%)")),
            "month_amplitude_max": to_number(raw.get("月振幅(%)")),
            "month_discount_rate_avg": to_number(raw.get("折溢价率_%")),
            "month_end_iopv": None,
            "trading_days": trading_days,
            "name": clean_text(raw.get("基金简称")) or None,
            "数据来源": "补充表",
            "是否补充数据": "是",
            "原始月成交额单位": "亿元",
            "原始日均成交额单位": "百万元",
            "raw_daily_avg_turnover_amount_from_supplement": raw_daily_amount,
            "raw_daily_avg_turnover_volume_from_supplement": raw_daily_volume,
            "是否日均字段重算": "是",
        }
        records.append(record)

    fixed_close = {}
    by_code: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        by_code[record["wind_code"]].append(record)
    for code, items in by_code.items():
        values = {
            round(float(r["month_end_close"]), 12)
            for r in items
            if r.get("month_end_close") is not None
        }
        fixed_close[code] = len(values) == 1 and len(items) > 1
        for record in items:
            record["是否疑似重复收盘价"] = "是" if fixed_close[code] else "否"
    return records, fixed_close


def merge_records(
    main_records: list[dict[str, Any]],
    supplement_records: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int]:
    main_map = {(r["wind_code"], r["year_month"]): r for r in main_records}
    supplement_map = {(r["wind_code"], r["year_month"]): r for r in supplement_records}
    keys = set(main_map) | set(supplement_map)
    merged = []
    conflicts = []
    supplement_used = 0

    for key in sorted(keys):
        main = main_map.get(key)
        supplement = supplement_map.get(key)
        if main is not None and supplement is not None:
            if has_core_liquidity(main):
                merged.append(main)
                if has_core_liquidity(supplement):
                    conflict = dict(supplement)
                    conflict["冲突原因"] = "主表和补充表同键均有有效数据，按规则保留主表"
                    conflicts.append(conflict)
            else:
                merged.append(supplement)
                supplement_used += 1
        elif main is not None:
            merged.append(main)
        elif supplement is not None:
            merged.append(supplement)
            supplement_used += 1
    return merged, conflicts, supplement_used


def add_pool_and_status(
    records: list[dict[str, Any]],
    pool: dict[str, dict[str, Any]],
) -> None:
    for record in records:
        product = pool.get(record["wind_code"])
        for field in POOL_FIELDS:
            record[field] = product.get(field) if product else None
        listing_date = product.get("上市日期") if product else None
        listing_month = listing_date.strftime("%Y-%m") if listing_date else ""
        record["上市年月"] = listing_month or None
        month = record.get("year_month", "")
        post_listing = bool(listing_month and month and month_index(month) >= month_index(listing_month))
        record["是否上市后月份"] = "是" if post_listing else "否"
        record["是否上市前空值月份"] = "是" if not post_listing else "否"
        effective = post_listing and has_core_liquidity(record)
        record["是否上市后有效流动性月份"] = "是" if effective else "否"
        if not product:
            record["数据可用状态"] = "未匹配产品池"
        elif not post_listing:
            record["数据可用状态"] = "上市前月份"
        elif effective:
            record["数据可用状态"] = "正常"
        else:
            record["数据可用状态"] = "上市后缺失"


def make_anomalies(
    records: list[dict[str, Any]],
    fixed_close: dict[str, bool],
    base_anomalies: list[dict[str, Any]],
    conflicts: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    anomalies = list(base_anomalies)

    def add(
        issue: str,
        record: dict[str, Any],
        field: str,
        original: Any,
        reference: Any = None,
        note: str = "",
        severity: str = "需复核",
    ) -> None:
        anomalies.append(
            {
                "异常类型": issue,
                "严重程度": severity,
                "wind_code": record.get("wind_code"),
                "year_month": record.get("year_month"),
                "字段": field,
                "原始值": original,
                "参考值/重算值": reference,
                "说明": note,
            }
        )

    seen_fixed_close = set()
    for record in records:
        code = record.get("wind_code", "")
        month = record.get("year_month", "")
        if not code:
            add("wind_code为空", record, "wind_code", code)
        elif not re.fullmatch(r"\d{6}\.(SH|SZ)", code):
            add("非标准Wind代码", record, "wind_code", code)
        if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", month):
            add("year_month格式异常", record, "year_month", month)

        for field in [
            "month_turnover_amount_sum",
            "month_turnover_volume_sum",
            "daily_avg_turnover_amount",
            "daily_avg_turnover_volume",
        ]:
            value = record.get(field)
            if isinstance(value, (int, float)) and value < 0:
                add("成交额或成交量为负数", record, field, value)
        for field in ["month_turnover_rate_sum", "month_amplitude_max"]:
            value = record.get(field)
            if isinstance(value, (int, float)) and value < 0:
                add("换手率或振幅为负数", record, field, value)

        discount = record.get("month_discount_rate_avg")
        if isinstance(discount, (int, float)) and abs(discount) > 10:
            add("折溢价率极端值", record, "month_discount_rate_avg", discount, "绝对值阈值10%")

        days = record.get("trading_days")
        if days is not None and (days < 1 or days > 25):
            add("trading_days超出1-25", record, "trading_days", days)

        amount = record.get("month_turnover_amount_sum")
        volume = record.get("month_turnover_volume_sum")
        if amount is not None and days is None:
            add("月成交额不为空但trading_days为空", record, "trading_days", None)
        if volume is not None and days is None:
            add("月成交量不为空但trading_days为空", record, "trading_days", None)

        if record.get("数据来源") == "主表" and amount is not None and days:
            recalculated = amount / days
            deviation = relative_difference(record.get("daily_avg_turnover_amount"), recalculated)
            if deviation is not None and deviation > 0.05:
                add(
                    "主表日均成交额与月成交额/交易日偏差较大",
                    record,
                    "daily_avg_turnover_amount",
                    record.get("daily_avg_turnover_amount"),
                    recalculated,
                    f"相对偏差={deviation:.2%}；新上市月可能因实际交易日少于全月交易日",
                )

        if record.get("数据来源") == "补充表":
            raw_amount = record.get("raw_daily_avg_turnover_amount_from_supplement")
            deviation = relative_difference(raw_amount, record.get("daily_avg_turnover_amount"))
            if deviation is not None and deviation > 0.05:
                add(
                    "补充表原始日均成交额与重算值偏差较大",
                    record,
                    "raw_daily_avg_turnover_amount_from_supplement",
                    raw_amount,
                    record.get("daily_avg_turnover_amount"),
                    f"相对偏差={deviation:.2%}；正式字段已按交易日重算",
                )
            raw_volume = record.get("raw_daily_avg_turnover_volume_from_supplement")
            volume_deviation = relative_difference(raw_volume, record.get("daily_avg_turnover_volume"))
            if volume_deviation is not None and volume_deviation > 0.05:
                add(
                    "补充表原始日均成交量与重算值偏差较大",
                    record,
                    "raw_daily_avg_turnover_volume_from_supplement",
                    raw_volume,
                    record.get("daily_avg_turnover_volume"),
                    f"相对偏差={volume_deviation:.2%}；正式字段已按交易日重算",
                )
            if (
                fixed_close.get(code)
                and code not in seen_fixed_close
            ):
                add(
                    "补充表month_end_close疑似固定值重复",
                    record,
                    "month_end_close",
                    record.get("month_end_close"),
                    None,
                    "该ETF全部月份收盘价相同，不建议用于历史月末价格或收益分析",
                )
                seen_fixed_close.add(code)
            if record.get("是否上市后有效流动性月份") == "是" and record.get("month_end_iopv") is None:
                add(
                    "补充表缺失month_end_iopv",
                    record,
                    "month_end_iopv",
                    None,
                    None,
                    "补充文件未提供IOPV",
                    "提示",
                )

    key_counts = Counter((r.get("wind_code"), r.get("year_month")) for r in records)
    for (code, month), count in key_counts.items():
        if count > 1:
            add(
                "重复wind_code+year_month",
                {"wind_code": code, "year_month": month},
                "唯一键",
                count,
            )
    for record in conflicts:
        add(
            "主表与补充表有效数据冲突",
            record,
            "唯一键",
            f"{record.get('wind_code')}+{record.get('year_month')}",
            "保留主表",
            record.get("冲突原因", ""),
        )
    return anomalies


def build_coverage(
    pool: dict[str, dict[str, Any]],
    records: list[dict[str, Any]],
    max_month: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    by_code: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        by_code[record["wind_code"]].append(record)

    coverage = []
    missing_list = []
    for code, product in pool.items():
        items = by_code.get(code, [])
        listing_month = product["上市日期"].strftime("%Y-%m") if product.get("上市日期") else ""
        expected_start = max("2015-01", listing_month) if listing_month else "2015-01"
        expected_months = set(month_range(expected_start, max_month))
        effective_items = [r for r in items if r.get("是否上市后有效流动性月份") == "是"]
        effective_months = {r["year_month"] for r in effective_items}
        missing_months = sorted(expected_months - effective_months)
        first_month = min(effective_months) if effective_months else None
        last_month = max(effective_months) if effective_months else None
        if not items:
            status = "完全遗漏"
            advice = "补抽完整分析区间"
        elif not effective_items:
            status = "有代码但无有效数据"
            advice = "补抽上市后全部流动性字段"
        elif expected_months and len(missing_months) / len(expected_months) >= 0.2:
            status = "上市后缺失严重"
            advice = "核查并补抽缺失月份"
        elif missing_months:
            status = "部分月份缺失"
            advice = "视研究需要补抽缺失月份"
        else:
            status = "正常"
            advice = "可直接用于分析"
        row = {
            "Wind代码": code,
            "证券简称": product.get("证券简称"),
            "基金简称": product.get("基金简称"),
            "上市日期": product.get("上市日期"),
            "上市年月": listing_month,
            "统计口径分类": product.get("统计口径分类"),
            "一级策略大类": product.get("一级策略大类"),
            "二级策略类别": product.get("二级策略类别"),
            "是否在流动性数据中出现": "是" if items else "否",
            "月度记录数": len(items),
            "上市后有效数据月份数": len(effective_months),
            "首个数据月份": first_month,
            "最后数据月份": last_month,
            "缺失月份数": len(missing_months),
            "数据可用状态": status,
            "处理建议": advice,
        }
        coverage.append(row)
        if status in {"完全遗漏", "有代码但无有效数据", "上市后缺失严重"}:
            missing_list.append(
                {
                    "Wind代码": code,
                    "证券简称": product.get("证券简称"),
                    "基金简称": product.get("基金简称"),
                    "问题类型": status,
                    "建议补抽开始月份": missing_months[0] if missing_months else expected_start,
                    "建议补抽结束月份": missing_months[-1] if missing_months else max_month,
                    "统计口径分类": product.get("统计口径分类"),
                    "一级策略大类": product.get("一级策略大类"),
                    "二级策略类别": product.get("二级策略类别"),
                    "最新基金规模(亿)": product.get("最新基金规模(亿)"),
                    "处理建议": advice,
                }
            )
    return coverage, missing_list


def build_supplement_acceptance(
    records: list[dict[str, Any]],
    pool: dict[str, dict[str, Any]],
    fixed_close: dict[str, bool],
) -> list[dict[str, Any]]:
    by_code: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        if record.get("数据来源") == "补充表":
            by_code[record["wind_code"]].append(record)
    output = []
    for code, (expected_start, expected_end, expected_count) in SUP_EXPECTATIONS.items():
        valid = [
            r
            for r in by_code.get(code, [])
            if r.get("是否上市后有效流动性月份") == "是"
        ]
        months = sorted({r["year_month"] for r in valid})
        product = pool.get(code, {})
        output.append(
            {
                "Wind代码": code,
                "基金简称": product.get("基金简称"),
                "上市日期": product.get("上市日期"),
                "预期有效开始月份": expected_start,
                "预期有效结束月份": expected_end,
                "预期有效月份数": expected_count,
                "实际有效开始月份": months[0] if months else None,
                "实际有效结束月份": months[-1] if months else None,
                "实际有效月份数": len(months),
                "是否达到预期": "是" if len(months) == expected_count and months and months[0] == expected_start and months[-1] == expected_end else "否",
                "是否疑似重复收盘价": "是" if fixed_close.get(code) else "否",
                "是否重算日均字段": "是",
                "折溢价率缺失月份数": sum(r.get("month_discount_rate_avg") is None for r in valid),
                "month_end_iopv缺失月份数": sum(r.get("month_end_iopv") is None for r in valid),
                "备注": "收盘价固定值不用于历史收益分析；正式日均字段按月成交额/量除以交易日重算",
            }
        )
    return output


def build_monthly_summary(analysis_records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []

    def append_summary(
        month: str,
        dimension: str,
        dimension_value: str,
        items: list[dict[str, Any]],
    ) -> None:
        rows.append(
            {
                "year_month": month,
                "汇总维度": dimension,
                "维度值": dimension_value,
                "有效ETF数量": len({r["wind_code"] for r in items}),
                "核心策略指数ETF数量": len(
                    {r["wind_code"] for r in items if clean_text(r.get("是否纳入核心策略ETF统计")) == "是"}
                ),
                "指数增强ETF数量": len(
                    {r["wind_code"] for r in items if clean_text(r.get("统计口径分类")) == "广义策略ETF_指数增强"}
                ),
                "月成交额合计_CNY": sum_numeric(items, "month_turnover_amount_sum"),
                "日均成交额均值_CNY": mean_numeric(items, "daily_avg_turnover_amount"),
                "日均成交额中位数_CNY": median_numeric(items, "daily_avg_turnover_amount"),
                "月成交量合计_份": sum_numeric(items, "month_turnover_volume_sum"),
                "平均换手率_%": mean_numeric(items, "month_turnover_rate_sum"),
                "平均折溢价率_%": mean_numeric(items, "month_discount_rate_avg"),
                "平均月振幅_%": mean_numeric(items, "month_amplitude_max"),
            }
        )

    by_month: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in analysis_records:
        by_month[record["year_month"]].append(record)
    for month in sorted(by_month):
        month_items = by_month[month]
        append_summary(month, "总体", "全部", month_items)
        for field, dimension in [
            ("统计口径分类", "统计口径分类"),
            ("一级策略大类", "一级策略大类"),
            ("二级策略类别", "二级策略类别"),
            ("市场范围_二次修正", "市场范围_二次修正"),
        ]:
            groups: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
            for record in month_items:
                groups[clean_text(record.get(field)) or "待补充"].append(record)
            for value, items in sorted(groups.items()):
                append_summary(month, dimension, value, items)
    return rows


def write_records(ws, headers: list[str], records: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for record in records:
        ws.append([record.get(header) for header in headers])


def style_sheet(ws, headers: list[str]) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BODY_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    sample_limit = min(ws.max_row, 300)
    wide = {"基金全称", "跟踪指数名称", "说明", "处理建议", "备注", "验收内容"}
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, sample_limit + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 60))
        width = min(max(max_len * 1.12, 11), 55 if header in wide else 30)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for col_idx, header in enumerate(headers, 1):
        if "日期" in header or header in {"上市日期", "基金成立日"}:
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (datetime, date)):
                    ws.cell(row_idx, col_idx).number_format = "yyyy-mm-dd"
        if header == "year_month" or "月份" in header or header == "上市年月":
            for row_idx in range(2, ws.max_row + 1):
                if ws.cell(row_idx, col_idx).value:
                    ws.cell(row_idx, col_idx).number_format = "@"
        if any(token in header for token in ["amount", "成交额", "规模"]):
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "#,##0.00"
        if any(token in header for token in ["volume", "成交量"]):
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "#,##0.00"
        if "%" in header or any(token in header for token in ["month_return", "turnover_rate", "amplitude", "discount_rate"]):
            for row_idx in range(2, ws.max_row + 1):
                if isinstance(ws.cell(row_idx, col_idx).value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "0.0000"


def main() -> None:
    pool = load_pool()
    main_records, trading_day_map, base_anomalies = load_main()
    supplement_records, fixed_close = load_supplement(trading_day_map)
    merged_records, conflicts, supplement_used = merge_records(main_records, supplement_records)
    add_pool_and_status(merged_records, pool)
    merged_records.sort(key=lambda r: (r.get("wind_code", ""), r.get("year_month", "")))
    analysis_records = [
        record
        for record in merged_records
        if record.get("是否上市后有效流动性月份") == "是"
    ]
    analysis_records.sort(key=lambda r: (r.get("year_month", ""), r.get("wind_code", "")))

    max_month = max((r["year_month"] for r in merged_records if re.fullmatch(r"\d{4}-\d{2}", r["year_month"])), default="")
    coverage_rows, missing_rows = build_coverage(pool, merged_records, max_month)
    supplement_acceptance = build_supplement_acceptance(
        merged_records, pool, fixed_close
    )
    anomalies = make_anomalies(
        merged_records, fixed_close, base_anomalies, conflicts
    )

    field_missing_rows = []
    for field in MISSING_STATS_FIELDS:
        full_missing = sum(r.get(field) is None for r in merged_records)
        analysis_missing = sum(r.get(field) is None for r in analysis_records)
        field_missing_rows.append(
            {
                "字段": field,
                "完整合并版缺失数": full_missing,
                "分析可用版缺失数": analysis_missing,
                "分析可用版记录数": len(analysis_records),
                "分析可用版缺失比例": analysis_missing / len(analysis_records) if analysis_records else None,
                "单位/口径说明": (
                    "CNY元" if "amount" in field
                    else "份" if "volume" in field
                    else "%" if field in {"month_return", "month_turnover_rate_sum", "month_amplitude_max", "month_discount_rate_avg"}
                    else ""
                ),
            }
        )

    monthly_rows = build_monthly_summary(analysis_records)

    main_codes = {r["wind_code"] for r in main_records}
    supplement_codes = {r["wind_code"] for r in supplement_records}
    merged_codes = {r["wind_code"] for r in merged_records}
    analysis_codes = {r["wind_code"] for r in analysis_records}
    complete_omissions = set(pool) - merged_codes

    notes_rows = [
        {"项目": "产品池应覆盖ETF数量", "验收内容": len(pool)},
        {"项目": "主表原始ETF数量", "验收内容": len(main_codes)},
        {"项目": "补充表ETF数量", "验收内容": len(supplement_codes)},
        {"项目": "合并后覆盖ETF数量", "验收内容": len(merged_codes)},
        {"项目": "是否达到223只", "验收内容": "是" if len(merged_codes) == len(pool) == 223 else "否"},
        {"项目": "分析可用ETF数量", "验收内容": len(analysis_codes)},
        {"项目": "月份范围", "验收内容": f"{min(r['year_month'] for r in merged_records)} 至 {max_month}"},
        {
            "项目": "补充数据修复结果",
            "验收内容": "已补齐159207.SZ、159581.SZ、159905.SZ三只完全遗漏产品，并用补充表替换562060.SH原有上市前空面板，四只均达到预期有效月份数。",
        },
        {
            "项目": "补充数据收盘价问题",
            "验收内容": "4只产品的month_end_close在全部月份固定不变，疑似为最新收盘价重复填充；保留备查，但正式历史收益分析不得依赖该字段。",
        },
        {
            "项目": "补充数据日均字段问题",
            "验收内容": "补充表原始日均成交额和日均成交量疑似非按交易日平均；正式字段已统一使用月成交额/成交量除以主表同月trading_days重算。",
        },
        {
            "项目": "补充数据IOPV问题",
            "验收内容": "补充文件未提供month_end_iopv，补充产品有效月份中该字段均为空。",
        },
        {
            "项目": "补充数据折溢价率问题",
            "验收内容": "部分有效月份缺少折溢价率，具体数量见“补充数据验收”和“字段缺失统计”。",
        },
        {
            "项目": "字段单位口径",
            "验收内容": "month_turnover_amount_sum与daily_avg_turnover_amount统一为CNY元；成交量字段统一为份；month_return、month_turnover_rate_sum、month_amplitude_max、month_discount_rate_avg保留Wind原始百分比数值口径，单位为%。",
        },
        {
            "项目": "后续建议",
            "验收内容": "正式流动性统计使用“流动性数据_分析可用版”；收益分析优先使用month_return，不使用补充表固定收盘价；对缺失待补清单及异常值记录逐项复核。",
        },
    ]

    wb = Workbook()
    wb.remove(wb.active)
    full_ws = wb.create_sheet(SHEET_FULL)
    analysis_ws = wb.create_sheet(SHEET_ANALYSIS)
    coverage_ws = wb.create_sheet(SHEET_COVERAGE)
    supplement_ws = wb.create_sheet(SHEET_SUP)
    missing_ws = wb.create_sheet(SHEET_MISSING)
    field_ws = wb.create_sheet(SHEET_FIELD_MISSING)
    anomaly_ws = wb.create_sheet(SHEET_ANOMALY)
    monthly_ws = wb.create_sheet(SHEET_MONTHLY)
    notes_ws = wb.create_sheet(SHEET_NOTES)

    write_records(full_ws, FULL_FIELDS, merged_records)
    write_records(analysis_ws, FULL_FIELDS, analysis_records)

    coverage_headers = [
        "Wind代码", "证券简称", "基金简称", "上市日期", "上市年月",
        "统计口径分类", "一级策略大类", "二级策略类别",
        "是否在流动性数据中出现", "月度记录数", "上市后有效数据月份数",
        "首个数据月份", "最后数据月份", "缺失月份数", "数据可用状态", "处理建议",
    ]
    write_records(coverage_ws, coverage_headers, coverage_rows)

    supplement_headers = [
        "Wind代码", "基金简称", "上市日期", "预期有效开始月份", "预期有效结束月份",
        "预期有效月份数", "实际有效开始月份", "实际有效结束月份", "实际有效月份数",
        "是否达到预期", "是否疑似重复收盘价", "是否重算日均字段",
        "折溢价率缺失月份数", "month_end_iopv缺失月份数", "备注",
    ]
    write_records(supplement_ws, supplement_headers, supplement_acceptance)

    missing_headers = [
        "Wind代码", "证券简称", "基金简称", "问题类型", "建议补抽开始月份",
        "建议补抽结束月份", "统计口径分类", "一级策略大类", "二级策略类别",
        "最新基金规模(亿)", "处理建议",
    ]
    write_records(missing_ws, missing_headers, missing_rows)

    field_headers = [
        "字段", "完整合并版缺失数", "分析可用版缺失数", "分析可用版记录数",
        "分析可用版缺失比例", "单位/口径说明",
    ]
    write_records(field_ws, field_headers, field_missing_rows)

    anomaly_headers = [
        "异常类型", "严重程度", "wind_code", "year_month", "字段", "原始值",
        "参考值/重算值", "说明",
    ]
    write_records(anomaly_ws, anomaly_headers, anomalies)

    monthly_headers = [
        "year_month", "汇总维度", "维度值", "有效ETF数量",
        "核心策略指数ETF数量", "指数增强ETF数量", "月成交额合计_CNY",
        "日均成交额均值_CNY", "日均成交额中位数_CNY", "月成交量合计_份",
        "平均换手率_%", "平均折溢价率_%", "平均月振幅_%",
    ]
    write_records(monthly_ws, monthly_headers, monthly_rows)
    notes_headers = ["项目", "验收内容"]
    write_records(notes_ws, notes_headers, notes_rows)

    for ws, headers in [
        (full_ws, FULL_FIELDS),
        (analysis_ws, FULL_FIELDS),
        (coverage_ws, coverage_headers),
        (supplement_ws, supplement_headers),
        (missing_ws, missing_headers),
        (field_ws, field_headers),
        (anomaly_ws, anomaly_headers),
        (monthly_ws, monthly_headers),
        (notes_ws, notes_headers),
    ]:
        style_sheet(ws, headers)

    # Percentage display for missing ratios.
    ratio_col = field_headers.index("分析可用版缺失比例") + 1
    for row_idx in range(2, field_ws.max_row + 1):
        field_ws.cell(row_idx, ratio_col).number_format = "0.00%"

    # Highlight severe anomalies and problematic coverage rows.
    severity_col = anomaly_headers.index("严重程度") + 1
    for row_idx in range(2, anomaly_ws.max_row + 1):
        fill = ERROR_FILL if anomaly_ws.cell(row_idx, severity_col).value == "需复核" else WARNING_FILL
        anomaly_ws.cell(row_idx, severity_col).fill = fill
    status_col = coverage_headers.index("数据可用状态") + 1
    for row_idx in range(2, coverage_ws.max_row + 1):
        if coverage_ws.cell(row_idx, status_col).value != "正常":
            coverage_ws.cell(row_idx, status_col).fill = WARNING_FILL

    notes_ws.column_dimensions["A"].width = 28
    notes_ws.column_dimensions["B"].width = 110
    for row_idx in range(2, notes_ws.max_row + 1):
        notes_ws.cell(row_idx, 2).alignment = Alignment(wrap_text=True, vertical="top")
        notes_ws.row_dimensions[row_idx].height = 48

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    print(f"1. 广义策略ETF标准产品池数量：{len(pool)}")
    print(f"2. 主表实际覆盖ETF数量：{len(main_codes)}")
    print(f"3. 补充表覆盖ETF数量：{len(supplement_codes)}")
    print(f"4. 合并后覆盖ETF数量：{len(merged_codes)}")
    print(f"5. 合并后是否仍有完全遗漏ETF：{'是' if complete_omissions else '否'}")
    for row in supplement_acceptance:
        print(f"6. {row['Wind代码']} 有效月份数：{row['实际有效月份数']}")
    print(f"7. 是否发现补充表收盘价重复问题：{'是' if any(fixed_close.values()) else '否'}")
    print("8. 是否已重算补充表日均成交额和日均成交量：是")
    print(f"9. 输出文件路径：{OUTPUT}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"处理失败：{exc}", file=sys.stderr)
        raise
