from __future__ import annotations

import shutil
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\广义策略ETF月度流动性数据_合并验收清洗版.xlsx")
OUTPUT = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\广义策略ETF月度流动性数据_合并验收清洗版_v2.xlsx")

CORE_SHEETS = [
    "流动性数据_完整合并版",
    "流动性数据_分析可用版",
    "产品覆盖检查",
    "补充数据验收",
    "月度汇总",
]

NON_BLOCKING = {
    "补充表原始日均成交额与重算值偏差较大",
    "补充表原始日均成交量与重算值偏差较大",
    "补充表缺失month_end_iopv",
    "补充表month_end_close疑似固定值重复",
    "主表日均成交额与月成交额/交易日偏差较大",
    "同月trading_days映射不一致",
}

REVIEW_NON_BLOCKING = {"折溢价率极端值"}

BLOCKING_PATTERNS = [
    "完全遗漏",
    "重复wind_code+year_month",
    "重复唯一键",
    "非标准Wind代码",
    "wind_code为空",
    "成交额或成交量为负数",
    "负数成交额",
    "负数成交量",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=9, color="1F1F1F")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN_GRAY = Side(style="thin", color="D9E1F2")
BODY_BORDER = Border(bottom=THIN_GRAY)


def classify_blocking(issue: str) -> str:
    if issue in NON_BLOCKING:
        return "否"
    if issue in REVIEW_NON_BLOCKING:
        return "否，但建议人工复核"
    if any(pattern in issue for pattern in BLOCKING_PATTERNS):
        return "是"
    return "否，但建议人工复核"


def snapshot_sheet(ws) -> tuple[int, int, list[tuple]]:
    values = [tuple(row) for row in ws.iter_rows(values_only=True)]
    return ws.max_row, ws.max_column, values


def style_notes_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A1:B{ws.max_row}"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 115
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, 1).font = BODY_FONT
        ws.cell(row_idx, 2).font = BODY_FONT
        ws.cell(row_idx, 1).border = BODY_BORDER
        ws.cell(row_idx, 2).border = BODY_BORDER
        ws.cell(row_idx, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_idx].height = 48


def style_usage_sheet(ws) -> None:
    style_notes_sheet(ws)


def main() -> None:
    if not INPUT.exists():
        raise FileNotFoundError(f"输入文件不存在：{INPUT}")

    # Copy first so the original workbook is never overwritten.
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(INPUT, OUTPUT)
    wb = load_workbook(OUTPUT)

    missing_sheets = [
        name
        for name in [*CORE_SHEETS, "异常值检查", "验收说明"]
        if name not in wb.sheetnames
    ]
    if missing_sheets:
        raise KeyError(f"输入工作簿缺少sheet：{missing_sheets}")

    snapshots = {name: snapshot_sheet(wb[name]) for name in CORE_SHEETS}

    # Add blocking classification to anomaly sheet without changing existing cells.
    anomaly_ws = wb["异常值检查"]
    headers = [cell.value for cell in anomaly_ws[1]]
    issue_col = headers.index("异常类型") + 1
    if "是否阻断分析" in headers:
        blocking_col = headers.index("是否阻断分析") + 1
    else:
        blocking_col = anomaly_ws.max_column + 1
        anomaly_ws.cell(1, blocking_col, "是否阻断分析")

    anomaly_count = anomaly_ws.max_row - 1
    blocking_count = 0
    for row_idx in range(2, anomaly_ws.max_row + 1):
        issue = str(anomaly_ws.cell(row_idx, issue_col).value or "")
        status = classify_blocking(issue)
        anomaly_ws.cell(row_idx, blocking_col, status)
        if status == "是":
            blocking_count += 1

    header_cell = anomaly_ws.cell(1, blocking_col)
    if anomaly_ws.max_column > 1:
        source_header = anomaly_ws.cell(1, blocking_col - 1)
        header_cell._style = copy(source_header._style)
    else:
        header_cell.fill = HEADER_FILL
        header_cell.font = HEADER_FONT
    header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    anomaly_ws.column_dimensions[get_column_letter(blocking_col)].width = 22
    anomaly_ws.freeze_panes = "A2"
    anomaly_ws.auto_filter.ref = f"A1:{get_column_letter(anomaly_ws.max_column)}{anomaly_ws.max_row}"
    for row_idx in range(2, anomaly_ws.max_row + 1):
        anomaly_ws.cell(row_idx, blocking_col).font = BODY_FONT
        anomaly_ws.cell(row_idx, blocking_col).border = BODY_BORDER

    # Revise the acceptance notes.
    notes_ws = wb["验收说明"]
    notes_ws.delete_rows(1, notes_ws.max_row)
    notes_rows = [
        ("产品池与覆盖", "本文件覆盖223只广义策略ETF，合并后无完全遗漏ETF。"),
        ("数据规模", "完整面板30,774行；分析可用数据7,433行；时间范围为2015-01至2026-06。"),
        ("异常记录总数", f"“异常值检查”共{anomaly_count}条记录。"),
        ("阻断性异常", f"阻断性异常为{blocking_count}条。当前异常主要是质量提示和复核标记，并不代表数据不可用。"),
        (
            "主要异常构成",
            "主要包括：补充表原始日均成交额/成交量与重算值偏差、补充表IOPV缺失、补充表收盘价固定重复、主表日均成交额与月成交额/交易日偏差提示，以及少量折溢价率极端值。",
        ),
        (
            "流动性分析可用性",
            "本文件可以用于月度流动性分析，正式分析建议使用“流动性数据_分析可用版”。",
        ),
        (
            "收益风险分析限制",
            "本文件不建议用于收益风险分析。收益风险应以后续单独导出的净值、复权净值或收益率数据为准；补充表month_end_close为固定重复值，不应作为历史价格。",
        ),
        (
            "日均字段处理",
            "补充表正式日均成交额和日均成交量已按月成交额/成交量除以trading_days重算；原始日均字段仅用于备查。",
        ),
        (
            "字段单位",
            "成交额统一为CNY元，成交量统一为份；换手率、折溢价率、月振幅和月收益率保留Wind原始百分比数值口径，单位为%。",
        ),
        (
            "使用建议",
            "月度流动性统计、分组比较和流动性趋势分析使用“流动性数据_分析可用版”；异常值检查中的非阻断提示可在敏感性分析或个案研究时进一步复核。",
        ),
    ]
    notes_ws.append(["项目", "验收内容"])
    for row in notes_rows:
        notes_ws.append(row)
    style_notes_sheet(notes_ws)

    # Create final usage guide.
    usage_name = "最终使用说明"
    if usage_name in wb.sheetnames:
        del wb[usage_name]
    usage_ws = wb.create_sheet(usage_name)
    usage_ws.append(["项目", "使用说明"])
    usage_rows = [
        ("覆盖范围", "本文件覆盖223只广义策略ETF。"),
        ("完整面板", "30,774行。"),
        ("分析可用数据", "7,433行。"),
        ("时间范围", "2015-01至2026-06。"),
        ("完全遗漏ETF", "无。"),
        ("重复唯一键", "无重复wind_code + year_month。"),
        ("缺失待补ETF", "“缺失待补ETF清单”为空。"),
        ("正式分析sheet", "使用“流动性数据_分析可用版”。"),
        (
            "可分析字段",
            "月成交额、日均成交额、月成交量、日均成交量、换手率、折溢价率、月振幅。",
        ),
        (
            "历史价格限制",
            "不建议使用补充表的month_end_close进行历史价格或收益分析，该字段在4只补充ETF中疑似固定值重复。",
        ),
        (
            "收益风险分析",
            "后续收益风险分析应使用单独导出的收益风险表现表、净值、复权净值或收益率数据。",
        ),
        (
            "异常解释",
            f"异常值检查共{anomaly_count}条，阻断性异常{blocking_count}条；其余为质量提示或建议人工复核项，不影响月度流动性分析主流程。",
        ),
    ]
    for row in usage_rows:
        usage_ws.append(row)
    style_usage_sheet(usage_ws)

    # Freeze all sheets as requested, without changing their values.
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_column > 0:
            for cell in ws[1]:
                cell.font = copy(cell.font)
                cell.font = Font(
                    name=cell.font.name or "微软雅黑",
                    size=cell.font.sz or 10,
                    bold=True,
                    color=cell.font.color,
                )

    # Verify core sheets remain byte-for-byte equivalent at the cell-value level.
    for name, before in snapshots.items():
        after = snapshot_sheet(wb[name])
        if before != after:
            raise RuntimeError(f"核心sheet数据被意外修改：{name}")

    wb.save(OUTPUT)

    print(f"输出文件路径：{OUTPUT}")
    print("是否新增“最终使用说明”：是")
    print("是否新增“是否阻断分析”字段：是")
    print(f"异常值检查总记录数：{anomaly_count}")
    print(f"阻断性异常数量：{blocking_count}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"处理失败：{exc}", file=sys.stderr)
        raise
