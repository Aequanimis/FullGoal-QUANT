from pathlib import Path

from openpyxl import load_workbook


BASE = Path.home() / "Desktop" / "Fullgoal" / "课题研究" / "处理后数据"

paths = [
    BASE / "全市场ETF基础信息_策略ETF池二次修正版.xlsx",
    BASE / "wind代码池" / "sheet7" / "ETF与指数映射关系表.xlsx",
    BASE / "wind代码池" / "sheet7" / "指数编制规则详细信息表.xlsx",
    BASE / "wind代码池" / "sheet7" / "完整数据合并表.xlsx",
]

for path in paths:
    print(f"\nFILE {path.name}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    print("Sheets:", workbook.sheetnames)
    for sheet_name in workbook.sheetnames[:4]:
        sheet = workbook[sheet_name]
        print(" Sheet:", sheet_name, "max", sheet.max_row, sheet.max_column)
        for row in sheet.iter_rows(max_row=3, values_only=True):
            print("  ", tuple(row[:25]))
