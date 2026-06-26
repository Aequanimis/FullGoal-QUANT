from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path.home() / "Desktop" / "Fullgoal" / "课题研究" / "处理后数据" / "wind代码池" / "sheet7"

ORIGINAL_FILE = BASE / "核心策略ETF_跟踪指数规则_验收修正版.xlsx"
SUPPLEMENT_FILE = BASE / "33只核心策略指数编制规则信息补充表.xlsx"
OUTPUT_FILE = BASE / "核心策略ETF_跟踪指数规则_补充整合终版.xlsx"

CORE_FIELDS = ["样本空间", "成分股数量", "选样指标", "选样方法", "加权方式", "调样频率"]
RULE_FIELDS = [
    "发布机构",
    "指数基日",
    "指数发布日期",
    "指数类型",
    "样本空间",
    "成分股数量",
    "选样指标",
    "选样方法",
    "加权方式",
    "单只成分权重上限",
    "行业权重限制",
    "调样频率",
    "指数简介",
    "指数编制方案链接",
]

SHEET4_COLS = [
    "Index_Code",
    "Index_Name",
    *RULE_FIELDS,
    "规则补充来源",
    "规则完整性状态",
    "仍缺失字段说明",
    "是否仍需官网核验",
]

SHEET5_COLS = [
    "ETF_Code",
    "ETF_Name",
    "基金简称",
    "基金全称",
    "基金管理人",
    "一级策略大类",
    "二级策略类别",
    "市场范围_二次修正",
    "最新基金规模(亿)",
    "上市日期",
    "基金成立日",
    "Index_Code",
    "Index_Name",
    *RULE_FIELDS,
    "规则补充来源",
    "规则完整性状态",
    "仍缺失字段说明",
    "是否仍需官网核验",
]


def norm_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).replace("\u3000", " ").replace("\xa0", " ").strip()
    return " ".join(text.split())


def norm_code(value: Any) -> str:
    return norm_text(value).upper()


def is_source_row(row: pd.Series) -> bool:
    return any("数据来源：Wind" in norm_text(v) or "数据来源:Wind" in norm_text(v) for v in row)


def clean_input_df(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    df = df.copy()
    df.columns = [norm_text(c) for c in df.columns]
    if df.empty:
        return df, 0
    source_mask = df.apply(is_source_row, axis=1)
    source_count = int(source_mask.sum())
    df = df.loc[~source_mask].dropna(how="all").copy()
    return df, source_count


def invalid_value(value: Any) -> bool:
    text = norm_text(value)
    if text == "":
        return True
    return text.lower() in {"nan", "none", "nat", "n/a"} or text in {
        "-",
        "--",
        "Wind暂无",
        "待补充",
        "待官网补充",
        "暂无",
        "暂无数据",
    }


def supplement_value(value: Any) -> str:
    return "待官网补充" if invalid_value(value) else norm_text(value)


def output_value(value: Any) -> str:
    return "待官网补充" if invalid_value(value) else norm_text(value)


def effective(value: Any) -> bool:
    return not invalid_value(value)


def missing_fields(row: pd.Series, fields: list[str]) -> list[str]:
    return [field for field in fields if not effective(row.get(field))]


def safe_float(value: Any) -> float:
    try:
        if invalid_value(value):
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def suggest_source(code: str, index_name: str, publisher: str) -> str:
    code_u = norm_code(code)
    name = norm_text(index_name)
    pub = norm_text(publisher)
    if code_u.endswith(".CSI") or "中证" in pub:
        return "中证指数有限公司官网"
    if code_u.endswith(".CNI") or "国证" in pub:
        return "国证指数有限公司官网"
    if code_u.endswith(".SH") or "上证" in name:
        return "中证指数有限公司官网或上证指数页面"
    if code_u.endswith(".SZ") or "深证" in name:
        return "国证指数有限公司官网或深证指数页面"
    if code_u.endswith(".HI") or "恒生" in pub:
        return "恒生指数有限公司官网"
    if code_u.endswith(".SPI") or "标普" in pub or "标普" in name:
        return "S&P Dow Jones Indices 官网"
    if code_u.endswith(".MI") or "MSCI" in pub.upper() or "MSCI" in name.upper():
        return "MSCI 官网"
    return "待确认"


def has_generic_source(value: Any) -> bool:
    text = norm_text(value)
    if not text or text == "待官网补充":
        return False
    if "Wind/官网" in text or text.startswith("Wind/") or text == "Wind":
        return True
    if text in {"官网", "Wind及官网", "Wind、官网"}:
        return True
    return False


def apply_formatting(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row and sheet.max_column:
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill

        headers = {sheet.cell(row=1, column=col).value: col for col in range(1, sheet.max_column + 1)}

        for col_idx in range(1, sheet.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in sheet[letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)

        for header, col_idx in headers.items():
            if header in {"上市日期", "基金成立日", "指数基日", "指数发布日期", "归档时间", "运行时间"}:
                for cell in sheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for item in cell:
                        item.number_format = "yyyy-mm-dd"
            if header in {"最新基金规模(亿)", "对应ETF规模合计", "涉及ETF规模合计", "受仍缺失规则影响的 ETF 规模合计"}:
                for cell in sheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for item in cell:
                        item.number_format = "0.0000"

    workbook.save(path)


def main() -> None:
    # Sheet 1 and 2 are standard outputs from the previous corrected workbook.
    sheet1 = pd.read_excel(ORIGINAL_FILE, sheet_name="核心ETF_指数映射_标准版", dtype=object)
    sheet2 = pd.read_excel(ORIGINAL_FILE, sheet_name="核心指数清单_去重版", dtype=object)
    old_rules = pd.read_excel(ORIGINAL_FILE, sheet_name="指数规则明细_标准版", dtype=object)
    sheet1.columns = [norm_text(c) for c in sheet1.columns]
    sheet2.columns = [norm_text(c) for c in sheet2.columns]
    old_rules.columns = [norm_text(c) for c in old_rules.columns]

    sheet1["ETF_Code"] = sheet1["ETF_Code"].map(norm_code)
    sheet1["Index_Code"] = sheet1["Index_Code"].map(norm_code)
    sheet2["Index_Code"] = sheet2["Index_Code"].map(norm_code)
    old_rules["Index_Code"] = old_rules["Index_Code"].map(norm_code)

    # Supplement clean-up.
    supplement_raw = pd.read_excel(SUPPLEMENT_FILE, dtype=object)
    raw_source_rows = int(supplement_raw.apply(is_source_row, axis=1).sum()) if not supplement_raw.empty else 0
    supplement, source_rows_removed = clean_input_df(supplement_raw)
    supplement = supplement.rename(
        columns={
            "指数发布机构": "发布机构",
            "指数编制方案链接或来源": "指数编制方案链接",
        }
    )
    for col in ["指数代码", "指数名称", *RULE_FIELDS]:
        if col not in supplement.columns:
            supplement[col] = ""

    supplement["指数代码"] = supplement["指数代码"].map(norm_code)
    supplement["指数名称"] = supplement["指数名称"].map(norm_text)
    supplement = supplement[supplement["指数代码"] != ""].copy()
    supplement = supplement.dropna(how="all").copy()

    for col in RULE_FIELDS:
        supplement[col] = supplement[col].map(supplement_value)

    supplement["补充表字段完整性状态"] = supplement.apply(
        lambda row: "核心字段完整" if not missing_fields(row, CORE_FIELDS) else "核心字段不完整",
        axis=1,
    )
    supplement["补充表缺失字段说明"] = supplement.apply(
        lambda row: "无" if not missing_fields(row, CORE_FIELDS) else "缺失：" + "、".join(missing_fields(row, CORE_FIELDS)),
        axis=1,
    )
    supplement_clean = supplement[
        ["指数代码", "指数名称", *RULE_FIELDS, "补充表字段完整性状态", "补充表缺失字段说明"]
    ].sort_values("指数代码")

    supplement_duplicate_count = int(supplement_clean["指数代码"].duplicated().sum())
    supplement_unique = (
        supplement_clean.assign(_complete=(supplement_clean["补充表字段完整性状态"] == "核心字段完整").astype(int))
        .sort_values(["指数代码", "_complete"], ascending=[True, False], kind="stable")
        .drop_duplicates("指数代码", keep="first")
        .drop(columns=["_complete"])
        .reset_index(drop=True)
    )
    supplement_map = supplement_unique.set_index("指数代码", drop=False)

    # Original missing list before integration.
    if "是否规则已补全" in old_rules.columns:
        old_complete_count = int((old_rules["是否规则已补全"].map(norm_text) == "是").sum())
    else:
        old_complete_count = int(old_rules.apply(lambda row: not missing_fields(row, CORE_FIELDS), axis=1).sum())
    old_index_count = old_rules["Index_Code"].nunique()
    old_pending_codes = set(old_rules.loc[old_rules.get("是否规则已补全", "").map(norm_text) != "是", "Index_Code"])

    # Build integrated rules.
    records: list[dict[str, Any]] = []
    supplemented_codes: set[str] = set()
    for _, old_row in old_rules.iterrows():
        code = norm_code(old_row.get("Index_Code"))
        record: dict[str, Any] = {
            "Index_Code": code,
            "Index_Name": output_value(old_row.get("Index_Name")),
        }
        sources = ["原修正版"]
        fields_filled_from_supplement: list[str] = []
        supp_row = supplement_map.loc[code] if code in supplement_map.index else None
        for field in RULE_FIELDS:
            old_value = old_row.get(field)
            supp_value = supp_row.get(field) if supp_row is not None else None
            if effective(old_value):
                record[field] = norm_text(old_value)
            elif supp_row is not None and effective(supp_value):
                record[field] = norm_text(supp_value)
                fields_filled_from_supplement.append(field)
            else:
                record[field] = "待官网补充"

        if fields_filled_from_supplement:
            supplemented_codes.add(code)
            sources.append("Wind补充表")

        core_missing = missing_fields(pd.Series(record), CORE_FIELDS)
        if not core_missing:
            integrity = "核心规则已补全"
        elif len(core_missing) < len(CORE_FIELDS):
            integrity = "部分补全"
        else:
            # If only non-core fields are available but all six core fields missing, treat as pending.
            integrity = "待补充"

        all_missing = missing_fields(pd.Series(record), RULE_FIELDS)
        record["规则补充来源"] = " + ".join(sources)
        record["规则完整性状态"] = integrity
        record["仍缺失字段说明"] = "无" if not all_missing else "缺失：" + "、".join(all_missing)
        if all_missing:
            record["是否仍需官网核验"] = "是"
        elif has_generic_source(record.get("指数编制方案链接")):
            record["是否仍需官网核验"] = "建议核验"
        else:
            record["是否仍需官网核验"] = "否"
        records.append(record)

    integrated_rules = pd.DataFrame(records)[SHEET4_COLS]

    # Preserve only 82 core indices, no extras from supplement.
    core_index_codes = set(sheet2["Index_Code"])
    integrated_rules = integrated_rules[integrated_rules["Index_Code"].isin(core_index_codes)].copy()

    # ETF-rule merged table.
    sheet5 = sheet1.merge(integrated_rules, on=["Index_Code", "Index_Name"], how="left")
    for col in RULE_FIELDS + ["规则补充来源", "规则完整性状态", "仍缺失字段说明", "是否仍需官网核验"]:
        if col not in sheet5.columns:
            sheet5[col] = "待官网补充"
        sheet5[col] = sheet5[col].map(output_value)
    sheet5 = sheet5[SHEET5_COLS].sort_values(
        ["一级策略大类", "二级策略类别", "最新基金规模(亿)"],
        ascending=[True, True, False],
        kind="stable",
    )

    # Remaining verification list.
    index_meta = sheet2.rename(
        columns={
            "对应ETF数量": "涉及ETF数量",
            "对应ETF规模合计": "涉及ETF规模合计",
            "涉及一级策略大类": "涉及一级策略大类",
            "涉及二级策略类别": "涉及二级策略类别",
        }
    )
    remaining = integrated_rules[
        (integrated_rules["规则完整性状态"] != "核心规则已补全")
        | (integrated_rules["是否仍需官网核验"].isin(["是", "建议核验"]))
    ].copy()
    sheet6 = remaining.merge(
        index_meta[["Index_Code", "涉及ETF数量", "涉及ETF规模合计", "涉及一级策略大类", "涉及二级策略类别"]],
        on="Index_Code",
        how="left",
    )
    sheet6["建议补充来源"] = sheet6.apply(
        lambda row: suggest_source(row["Index_Code"], row["Index_Name"], row.get("发布机构", "")),
        axis=1,
    )
    sheet6 = sheet6[
        [
            "Index_Code",
            "Index_Name",
            "发布机构",
            "涉及ETF数量",
            "涉及ETF规模合计",
            "涉及一级策略大类",
            "涉及二级策略类别",
            "规则完整性状态",
            "仍缺失字段说明",
            "建议补充来源",
        ]
    ].sort_values(["涉及一级策略大类", "涉及二级策略类别", "涉及ETF规模合计"], ascending=[True, True, False])

    # Comparison sheet.
    supplement_codes = set(supplement_unique["指数代码"])
    original_missing_codes = old_pending_codes
    covered_missing_codes = supplement_codes & original_missing_codes
    unmatched_to_missing_codes = supplement_codes - original_missing_codes
    core_complete = int((integrated_rules["规则完整性状态"] == "核心规则已补全").sum())
    partial_complete = int((integrated_rules["规则完整性状态"] == "部分补全").sum())
    still_pending = int((integrated_rules["规则完整性状态"] == "待补充").sum())
    still_verify = int(integrated_rules["是否仍需官网核验"].isin(["是", "建议核验"]).sum())
    affected_indices = set(sheet6["Index_Code"])
    affected_etfs = sheet5[sheet5["Index_Code"].isin(affected_indices)].copy()
    affected_etf_count = affected_etfs["ETF_Code"].nunique()
    affected_scale = affected_etfs["最新基金规模(亿)"].map(safe_float).sum()

    comparison_stats = pd.DataFrame(
        [
            ["原唯一指数数量", old_index_count],
            ["原规则已补全指数数量", old_complete_count],
            ["原规则待补充指数数量", old_index_count - old_complete_count],
            ["补充表有效指数数量", supplement_unique["指数代码"].nunique()],
            ["补充表覆盖原缺失指数数量", len(covered_missing_codes)],
            ["补充表未匹配到原缺失清单的指数数量", len(unmatched_to_missing_codes)],
            ["整合后核心规则已补全指数数量", core_complete],
            ["整合后部分补全指数数量", partial_complete],
            ["整合后仍待补充指数数量", still_pending],
            ["整合后仍需官网核验指数数量", still_verify],
            ["受仍缺失规则影响的 ETF 数量", affected_etf_count],
            ["受仍缺失规则影响的 ETF 规模合计", affected_scale],
        ],
        columns=["统计项", "数值"],
    )
    by_strategy = (
        affected_etfs.groupby("一级策略大类", as_index=False)
        .agg(
            剩余需核验指数数量=("Index_Code", "nunique"),
            涉及ETF数量=("ETF_Code", "nunique"),
            涉及ETF规模合计=("最新基金规模(亿)", lambda s: sum(safe_float(v) for v in s)),
        )
        .sort_values(["剩余需核验指数数量", "涉及ETF规模合计"], ascending=[False, False])
    )
    sheet7 = pd.concat(
        [
            comparison_stats,
            pd.DataFrame([["", ""]], columns=["统计项", "数值"]),
            pd.DataFrame([["按一级策略大类的剩余缺失情况", ""]], columns=["统计项", "数值"]),
            by_strategy.rename(columns={"一级策略大类": "统计项", "剩余需核验指数数量": "数值"})[
                ["统计项", "数值", "涉及ETF数量", "涉及ETF规模合计"]
            ],
        ],
        ignore_index=True,
    )

    # Data quality checks.
    sheet1_dup = int(sheet1["ETF_Code"].duplicated().sum())
    sheet2_unique = sheet2["Index_Code"].nunique()
    has_source_in_supp_clean = bool(supplement_clean.apply(is_source_row, axis=1).any())
    all_supp_match_old_missing = supplement_codes.issubset(original_missing_codes)
    noncore_mixed = bool(set(integrated_rules["Index_Code"]) - core_index_codes)
    lost_core = bool(core_index_codes - set(integrated_rules["Index_Code"]))

    standard_map = sheet1[["ETF_Code", "Index_Code"]].rename(columns={"Index_Code": "Index_Code_标准"})
    merged_check = sheet5[["ETF_Code", "Index_Code"]].merge(standard_map, on="ETF_Code", how="left")
    etf_index_mismatch = int((merged_check["Index_Code"].map(norm_code) != merged_check["Index_Code_标准"].map(norm_code)).sum())
    index_code_same = "是" if etf_index_mismatch == 0 and len(sheet5) == len(sheet1) else "否"
    core_incomplete = len(integrated_rules) - core_complete

    quality = pd.DataFrame(
        [
            ["Sheet 1 ETF 行数", len(sheet1), ""],
            ["Sheet 1 ETF_Code 重复数量", sheet1_dup, ""],
            ["Sheet 2 唯一 Index_Code 数量", sheet2_unique, ""],
            ["Sheet 3 补充表有效指数数量", supplement_unique["指数代码"].nunique(), ""],
            ["Sheet 3 补充表重复指数代码数量", supplement_duplicate_count, ""],
            ["Sheet 3 是否存在“数据来源：Wind”行", "是" if has_source_in_supp_clean else "否", f"原补充表删除来源行 {source_rows_removed} 行；原始检测 {raw_source_rows} 行"],
            ["补充表 33 个指数是否全部能匹配原缺失规则清单", "是" if all_supp_match_old_missing else "否", f"未匹配数量 {len(unmatched_to_missing_codes)}"],
            ["整合后指数规则明细行数", len(integrated_rules), ""],
            ["整合后是否混入非核心指数", "是" if noncore_mixed else "否", ""],
            ["整合后是否丢失核心指数", "是" if lost_core else "否", ""],
            ["ETF_指数规则合并表行数", len(sheet5), ""],
            ["ETF_指数规则合并表 Index_Code 是否与 Sheet 1 完全一致", index_code_same, ""],
            ["本次整合后 ETF-指数错配数量", etf_index_mismatch, ""],
            ["核心 6 字段完整指数数量", core_complete, ""],
            ["核心 6 字段不完整指数数量", core_incomplete, ""],
            ["仍需官网核验指数数量", still_verify, ""],
            ["脚本运行时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""],
        ],
        columns=["检查项", "结果", "说明"],
    )

    # Replace display blanks in text-heavy outputs without damaging numeric/date fields.
    for df in [supplement_clean, integrated_rules, sheet5, sheet6, sheet7, quality]:
        object_cols = df.select_dtypes(include=["object"]).columns
        for col in object_cols:
            df[col] = df[col].map(lambda v: "待官网补充" if norm_text(v) == "" else v)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        sheet1.to_excel(writer, sheet_name="核心ETF_指数映射_标准版", index=False)
        sheet2.to_excel(writer, sheet_name="核心指数清单_去重版", index=False)
        supplement_clean.to_excel(writer, sheet_name="补充表_清洗版", index=False)
        integrated_rules.to_excel(writer, sheet_name="指数规则明细_整合版", index=False)
        sheet5.to_excel(writer, sheet_name="ETF_指数规则合并表_整合版", index=False)
        sheet6.to_excel(writer, sheet_name="剩余待核验字段清单", index=False)
        sheet7.to_excel(writer, sheet_name="补充前后对比", index=False)
        quality.to_excel(writer, sheet_name="数据质量检查", index=False)

    apply_formatting(OUTPUT_FILE)

    print(f"补充表有效指数数量: {supplement_unique['指数代码'].nunique()}")
    print(f"成功匹配并补充的指数数量: {len(supplemented_codes)}")
    print(f"整合后核心规则已补全指数数量: {core_complete}")
    print(f"整合后部分补全指数数量: {partial_complete}")
    print(f"整合后仍需官网核验指数数量: {still_verify}")
    print(f"ETF-指数错配数量: {etf_index_mismatch}")
    print(f"输出文件路径: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
