from pathlib import Path

from openpyxl import load_workbook


path = (
    Path.home()
    / "Desktop"
    / "Fullgoal"
    / "课题研究"
    / "处理后数据"
    / "wind代码池"
    / "sheet7"
    / "核心策略ETF_跟踪指数规则_补充整合终版.xlsx"
)

wb = load_workbook(path, read_only=True, data_only=True)
print("FILE_SIZE", path.stat().st_size)
print("SHEETS", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(name, ws.max_row, ws.max_column)

ws1 = wb["核心ETF_指数映射_标准版"]
h1 = [c.value for c in next(ws1.iter_rows(max_row=1))]
r1 = list(ws1.iter_rows(min_row=2, values_only=True))
print("SHEET1_ROWS", len(r1))
print("ETF_UNIQUE", len({row[h1.index('ETF_Code')] for row in r1}))
print("INDEX_UNIQUE_IN_SHEET1", len({row[h1.index('Index_Code')] for row in r1}))

ws2 = wb["核心指数清单_去重版"]
h2 = [c.value for c in next(ws2.iter_rows(max_row=1))]
r2 = list(ws2.iter_rows(min_row=2, values_only=True))
print("SHEET2_INDEX_UNIQUE", len({row[h2.index('Index_Code')] for row in r2}))

ws3 = wb["补充表_清洗版"]
h3 = [c.value for c in next(ws3.iter_rows(max_row=1))]
r3 = list(ws3.iter_rows(min_row=2, values_only=True))
print("SHEET3_ROWS", len(r3), "SUPP_INDEX_UNIQUE", len({row[h3.index('指数代码')] for row in r3}))
print("SHEET3_DUP", len(r3) - len({row[h3.index('指数代码')] for row in r3}))

ws4 = wb["指数规则明细_整合版"]
h4 = [c.value for c in next(ws4.iter_rows(max_row=1))]
r4 = list(ws4.iter_rows(min_row=2, values_only=True))
status_idx = h4.index("规则完整性状态")
verify_idx = h4.index("是否仍需官网核验")
print("SHEET4_ROWS", len(r4), "INDEX_UNIQUE", len({row[h4.index('Index_Code')] for row in r4}))
print("CORE_COMPLETE", sum(1 for row in r4 if row[status_idx] == "核心规则已补全"))
print("PARTIAL", sum(1 for row in r4 if row[status_idx] == "部分补全"))
print("PENDING", sum(1 for row in r4 if row[status_idx] == "待补充"))
print("VERIFY", sum(1 for row in r4 if row[verify_idx] in ("是", "建议核验")))

ws5 = wb["ETF_指数规则合并表_整合版"]
h5 = [c.value for c in next(ws5.iter_rows(max_row=1))]
r5 = list(ws5.iter_rows(min_row=2, values_only=True))
print("SHEET5_ROWS", len(r5))
map1 = {row[h1.index("ETF_Code")]: row[h1.index("Index_Code")] for row in r1}
mismatch = 0
for row in r5:
    etf = row[h5.index("ETF_Code")]
    idx = row[h5.index("Index_Code")]
    if map1.get(etf) != idx:
        mismatch += 1
print("ETF_INDEX_MISMATCH", mismatch)

has_wind_source = False
for ws in wb.worksheets:
    for row in ws.iter_rows(values_only=True):
        # Ignore the quality-check label itself; detect actual source rows in tabular data.
        nonempty = [str(v) for v in row if v is not None and str(v).strip()]
        if len(nonempty) == 1 and ("数据来源：Wind" in nonempty[0] or "数据来源:Wind" in nonempty[0]):
            has_wind_source = True
print("HAS_ACTUAL_WIND_SOURCE_ROW", has_wind_source)
