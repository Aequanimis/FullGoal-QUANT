"""Step 8：整理代表指数规则并同步主 Excel 与 Step 7 产品池。"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
ATTACHMENT = Path(
    r"C:\Users\lvdon\.codex\attachments\6dc7c99d-2eb2-424f-a376-3e3e133f94cf\pasted-text.txt"
)
STEP7_EXCEL = PROJECT_ROOT / "01_processed_data" / "classification" / "境内策略ETF产品池_代表产品筛选版.xlsx"
MAIN_EXCEL = PROJECT_ROOT / "02_outputs" / "excel" / "境内策略ETF产品梳理_初版.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "01_processed_data" / "index_rules"
OUTPUT_CSV = OUTPUT_DIR / "代表指数规则整理.csv"
OUTPUT_EXCEL = OUTPUT_DIR / "代表指数规则整理.xlsx"
REVIEW_MD = OUTPUT_DIR / "step8_index_rules_review.md"
LOG_MD = OUTPUT_DIR / "step8_index_rules_log.md"
STEP7_STAGING = OUTPUT_DIR / "_step8_step7_staging.xlsx"

RULE_FIELDS = [
    "策略类型", "关联代表产品代码", "关联代表产品名称", "指数名称", "指数代码", "指数公司", "样本空间",
    "选样因子", "选样方法", "成分股数量", "加权方式", "调样频率", "个股权重限制", "行业限制",
    "前十大成分股", "行业分布", "策略理解", "信息来源URL", "核验状态", "备注",
]

STEP8_FIELDS = [
    "Step8指数规则核验状态",
    "Step8指数规则简述",
    "Step8信息来源URL",
    "Step8待补充事项",
]

MAIN_RULE_MAP = {
    "策略类型": "策略类型",
    "指数名称": "指数名称",
    "指数代码": "指数代码",
    "指数公司": "指数公司",
    "样本空间": "样本空间",
    "选样因子": "选样因子",
    "选样方法": "选样方法",
    "成分股数量": "成分股数量",
    "加权方式": "加权方式",
    "调样频率": "调样频率",
    "个股权重限制": "个股权重限制",
    "行业限制": "行业限制",
    "前十大成分股": "前十大成分股",
    "行业分布": "行业分布",
    "策略理解": "策略理解",
    "信息来源URL": "信息来源",
}


def clean(value: object) -> str:
    return "" if value is None else str(value).strip(" \t\r\n")


def read_rules() -> list[dict[str, str]]:
    text = ATTACHMENT.read_text(encoding="utf-8-sig")
    header = "\t".join(RULE_FIELDS)
    start = text.find(header)
    if start < 0:
        raise ValueError("附件中未找到指数规则 TSV 表头")
    end = text.find("\n\n五、输出 Excel 要求", start)
    if end < 0:
        raise ValueError("附件中未找到指数规则 TSV 结束位置")
    block = text[start:end].strip("\r\n")
    reader = csv.DictReader(io.StringIO(block), delimiter="\t")
    if reader.fieldnames != RULE_FIELDS:
        raise ValueError(f"指数规则 TSV 字段不匹配：{reader.fieldnames}")
    return [
        {field: clean(record.get(field, "")) for field in RULE_FIELDS}
        for record in reader
    ]


def write_csv(rules: list[dict[str, str]]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=RULE_FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rules)


def display_width(value: object) -> int:
    text = clean(value)
    return sum(
        2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1
        for char in text
    )


def append_rows(worksheet, fields: list[str], rows: list[dict[str, object]]) -> None:
    worksheet.append(fields)
    for record in rows:
        worksheet.append([record.get(field, "") for field in fields])


def format_sheet(worksheet, mark_verification: bool = False) -> None:
    max_row = max(worksheet.max_row, 1)
    max_column = max(worksheet.max_column, 1)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_column)}{max_row}"
    worksheet.sheet_view.zoomScale = 85

    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    warning_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    header_font = Font(name="Microsoft YaHei", size=10, bold=True)
    body_font = Font(name="Microsoft YaHei", size=10)
    border = Border(bottom=Side(style="thin", color="A6A6A6"))
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    worksheet.row_dimensions[1].height = 28

    headers = [clean(cell.value) for cell in worksheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}
    for row in worksheet.iter_rows(min_row=2, max_row=max_row, max_col=max_column):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if mark_verification and "核验状态" in columns:
        for row_number in range(2, max_row + 1):
            status = clean(worksheet.cell(row_number, columns["核验状态"]).value)
            if status in {"部分核验", "待二次核验"}:
                worksheet.cell(row_number, columns["核验状态"]).fill = warning_fill
                if "备注" in columns:
                    worksheet.cell(row_number, columns["备注"]).fill = warning_fill

    long_headers = {
        "关联代表产品名称", "样本空间", "选样因子", "选样方法", "加权方式", "调样频率", "个股权重限制",
        "行业限制", "前十大成分股", "行业分布", "策略理解", "信息来源URL", "备注",
    }
    for column_index, header in enumerate(headers, start=1):
        values = [worksheet.cell(row, column_index).value for row in range(1, max_row + 1)]
        measured = max(display_width(value) for value in values) + 3
        cap = 52 if header in long_headers else 28
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(measured, 12), cap)
        if header in {"指数代码", "关联代表产品代码", "信息来源URL"}:
            for row in range(2, max_row + 1):
                worksheet.cell(row, column_index).number_format = "@"


def build_statistics(rules: list[dict[str, str]]) -> list[dict[str, object]]:
    strategy_counts = Counter(rule["策略类型"] for rule in rules)
    company_counts = Counter(rule["指数公司"] for rule in rules)
    status_counts = Counter(rule["核验状态"] for rule in rules)
    rows: list[dict[str, object]] = []
    for strategy, count in sorted(strategy_counts.items()):
        rows.append({"统计类型": "按策略类型", "统计项目": strategy, "数量": count})
    for company, count in sorted(company_counts.items()):
        rows.append({"统计类型": "按指数公司", "统计项目": company, "数量": count})
    for status, count in sorted(status_counts.items()):
        rows.append({"统计类型": "按核验状态", "统计项目": status, "数量": count})

    dynamic_pending = sum(
        "待" in rule["前十大成分股"] or "待" in rule["行业分布"] for rule in rules
    )
    summaries = [
        ("已核验数量", status_counts["已核验"]),
        ("部分核验数量", status_counts["部分核验"]),
        ("待二次核验数量", status_counts["待二次核验"]),
        ("指数增强单列观察数量", strategy_counts["指数增强/多因子"]),
        ("前十大成分股和行业分布待补充数量", dynamic_pending),
    ]
    for label, count in summaries:
        rows.append({"统计类型": "汇总指标", "统计项目": label, "数量": count})
    return rows


def write_rule_excel(rules: list[dict[str, str]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    verified = [rule for rule in rules if rule["核验状态"] == "已核验"]
    partial = [rule for rule in rules if rule["核验状态"] in {"部分核验", "待二次核验"}]
    enhanced = [rule for rule in rules if rule["策略类型"] == "指数增强/多因子"]

    for sheet_name, rows in [
        ("代表指数规则表", rules),
        ("已核验规则", verified),
        ("部分核验_待补充", partial),
        ("指数增强单列观察", enhanced),
    ]:
        worksheet = workbook.create_sheet(sheet_name)
        append_rows(worksheet, RULE_FIELDS, rows)
        format_sheet(worksheet, mark_verification=True)

    stats_sheet = workbook.create_sheet("指数规则统计")
    append_rows(stats_sheet, ["统计类型", "统计项目", "数量"], build_statistics(rules))
    format_sheet(stats_sheet)
    workbook.save(OUTPUT_EXCEL)


def sync_main_excel(rules: list[dict[str, str]]) -> None:
    workbook = load_workbook(MAIN_EXCEL)
    if "指数规则表" not in workbook.sheetnames:
        raise ValueError("主 Excel 中不存在 Sheet【指数规则表】")
    worksheet = workbook["指数规则表"]
    headers = [clean(cell.value) for cell in worksheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}
    missing = set(MAIN_RULE_MAP.values()) - set(headers)
    if missing:
        raise ValueError(f"主 Excel 指数规则表缺少字段：{sorted(missing)}")

    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)
    for row_number, rule in enumerate(rules, start=2):
        for source_field, target_field in MAIN_RULE_MAP.items():
            worksheet.cell(row_number, columns[target_field], rule[source_field])

    last_row = len(rules) + 1
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    for row in worksheet.iter_rows(min_row=2, max_row=last_row, max_col=len(headers)):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for field in ("指数代码", "信息来源"):
        column = columns[field]
        for row in range(2, last_row + 1):
            worksheet.cell(row, column).number_format = "@"
    for column_index, header in enumerate(headers, start=1):
        values = [worksheet.cell(row, column_index).value for row in range(1, last_row + 1)]
        measured = max(display_width(value) for value in values) + 3
        cap = 52 if header in {"样本空间", "选样因子", "选样方法", "加权方式", "调样频率", "个股权重限制", "行业限制", "前十大成分股", "行业分布", "策略理解", "信息来源"} else 28
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(measured, 12), cap)
    workbook.save(MAIN_EXCEL)


def normalize_index_name(value: str) -> str:
    text = re.sub(r"\s+", "", value)
    replacements = {
        "中央企业": "央企",
        "股票指数": "指数",
        "相关指数": "指数",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return text


def associated_codes(rule: dict[str, str]) -> set[str]:
    return {code.strip() for code in rule["关联代表产品代码"].split(";") if code.strip()}


def find_rule_for_product(
    product: dict[str, str], rules: list[dict[str, str]]
) -> tuple[dict[str, str] | None, str]:
    code = product["基金代码"]
    tracking_index = product["跟踪指数"]

    if product.get("建议策略大类") == "指数增强/多因子":
        matches = [rule for rule in rules if rule["策略类型"] == "指数增强/多因子" and code in associated_codes(rule)]
        return (matches[0], "基金代码匹配") if len(matches) == 1 else (None, "未匹配")

    exact = [rule for rule in rules if tracking_index == rule["指数名称"]]
    if len(exact) == 1:
        return exact[0], "精确匹配"

    normalized_tracking = normalize_index_name(tracking_index)
    candidates: list[dict[str, str]] = []
    for rule in rules:
        normalized_rule = normalize_index_name(rule["指数名称"])
        if len(normalized_rule) >= 5 and (
            normalized_rule in normalized_tracking or normalized_tracking in normalized_rule
        ):
            candidates.append(rule)
    if len(candidates) == 1:
        return candidates[0], "模糊匹配"

    focus_text = normalize_index_name(product.get("Step7后续需重点看的指数规则", ""))
    focus_candidates = [
        rule
        for rule in rules
        if len(normalize_index_name(rule["指数名称"])) >= 5
        and normalize_index_name(rule["指数名称"]) in focus_text
    ]
    if len(focus_candidates) == 1:
        return focus_candidates[0], "模糊匹配"
    return None, "未匹配"


def rule_summary(rule: dict[str, str]) -> str:
    return (
        f"{rule['策略理解']}；选样方法：{rule['选样方法']}；"
        f"加权方式：{rule['加权方式']}；调样频率：{rule['调样频率']}"
    )


def pending_items(rule: dict[str, str]) -> str:
    items: list[str] = []
    if rule["核验状态"] != "已核验":
        items.append(f"{rule['核验状态']}：{rule['备注']}")
    if "待" in rule["前十大成分股"] or "待" in rule["行业分布"]:
        items.append("前十大成分股、行业分布待 Wind / 指数官网 factsheet 实时导出")
    return "；".join(items)


def write_back_step7(
    rules: list[dict[str, str]],
) -> tuple[list[tuple[str, str]], list[tuple[str, str, str]], int, Path]:
    workbook = load_workbook(STEP7_EXCEL)
    if "产品池_代表产品筛选版" not in workbook.sheetnames:
        raise ValueError("Step 7 Excel 中不存在 Sheet【产品池_代表产品筛选版】")
    worksheet = workbook["产品池_代表产品筛选版"]
    headers = [clean(cell.value) for cell in worksheet[1]]
    required = {"基金代码", "产品名称", "跟踪指数", "建议策略大类", "是否Step7代表产品", "Step7后续需重点看的指数规则"}
    missing = required - set(headers)
    if missing:
        raise ValueError(f"Step 7 产品池缺少字段：{sorted(missing)}")

    columns = {header: index + 1 for index, header in enumerate(headers)}
    for field in STEP8_FIELDS:
        if field not in columns:
            worksheet.cell(1, worksheet.max_column + 1, field)
            columns[field] = worksheet.max_column
            headers.append(field)

    unmatched_representatives: list[tuple[str, str]] = []
    fuzzy_representatives: list[tuple[str, str, str]] = []
    matched_count = 0
    for row_number in range(2, worksheet.max_row + 1):
        product = {header: clean(worksheet.cell(row_number, column).value) for header, column in columns.items()}
        rule, match_type = find_rule_for_product(product, rules)
        if rule:
            matched_count += 1
            values = {
                "Step8指数规则核验状态": rule["核验状态"],
                "Step8指数规则简述": rule_summary(rule),
                "Step8信息来源URL": rule["信息来源URL"],
                "Step8待补充事项": pending_items(rule),
            }
            if product["是否Step7代表产品"] == "是" and match_type == "模糊匹配":
                fuzzy_representatives.append((product["基金代码"], product["产品名称"], rule["指数名称"]))
        else:
            values = {
                "Step8指数规则核验状态": "未匹配",
                "Step8指数规则简述": "",
                "Step8信息来源URL": "",
                "Step8待补充事项": "未匹配到指数规则，后续人工核验",
            }
            if product["是否Step7代表产品"] == "是":
                unmatched_representatives.append((product["基金代码"], product["产品名称"]))

        for field, value in values.items():
            cell = worksheet.cell(row_number, columns[field], value)
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if field == "Step8信息来源URL":
                cell.number_format = "@"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    for field in STEP8_FIELDS:
        cell = worksheet.cell(1, columns[field])
        cell.fill = header_fill
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.column_dimensions[get_column_letter(columns[field])].width = 48 if field != "Step8指数规则核验状态" else 18
    validation_path = STEP7_EXCEL
    try:
        workbook.save(STEP7_EXCEL)
    except PermissionError:
        workbook.save(STEP7_STAGING)
        validation_path = STEP7_STAGING
    return unmatched_representatives, fuzzy_representatives, matched_count, validation_path


def write_review(rules: list[dict[str, str]]) -> None:
    verified = [rule["指数名称"] for rule in rules if rule["核验状态"] == "已核验"]
    partial = [rule["指数名称"] for rule in rules if rule["核验状态"] == "部分核验"]
    pending = [rule["指数名称"] for rule in rules if rule["核验状态"] == "待二次核验"]
    bullet = lambda values: "\n".join(f"- {value}" for value in values) or "- 无"
    content = f"""# Step 8 指数规则说明

## 1. Step 8 目标

整理 Step 7 代表产品背后的指数编制规则，按代表指数去重，为后续策略比较、代表产品研究和 mentor 讨论提供结构化依据。

## 2. 本轮指数规则信息来源

本轮仅使用给定 DATA_BLOCK 中已整理的信息和 URL，不联网访问页面。信息来源包括中证指数、国证指数、上交所、标普道琼斯等官方公开编制方案及基金文件线索。

本轮 Step 8 重点整理 Step 7 代表产品背后的策略指数编制规则。为了避免重复，本轮按“代表指数”而不是按“代表产品”整理；多个产品跟踪同一指数时合并为一条指数规则。已优先使用中证指数、国证指数、上交所、标普道琼斯等官方公开编制方案。对于无法找到完整官方编制方案的指数，暂标记为“部分核验”或“待二次核验”，不强行补全。前十大成分股和行业分布属于动态数据，应后续通过 Wind、Choice、指数官网 factsheet 或指数浏览器按日期导出，不在本轮中手动编造。指数增强ETF与Smart Beta策略指数ETF不同，本轮仅作为单列观察，不纳入纯被动策略指数规则比较。

## 3. 已完成核验的指数

{bullet(verified)}

## 4. 部分核验、待二次核验的指数

### 部分核验

{bullet(partial)}

### 待二次核验

{bullet(pending)}

## 5. 为什么不填前十大成分股和行业分布

前十大成分股和行业分布会随调样、行情和统计日期变化，属于动态数据。本轮保留“待 Wind / 指数官网 factsheet 实时导出”，避免写入无日期依据的静态结果。

## 6. 指数增强 ETF 为什么单列观察

指数增强 ETF 以宽基指数为锚，通过多因子 Alpha、组合优化和主动偏离约束争取增强收益，并非跟踪预先编制的 Smart Beta 策略指数，因此不与纯被动策略指数规则混合比较。

## 7. 后续需要补充的字段

- Wind / Choice：前十大成分股、行业分布、产品规模、成交额和动态权重。
- 指数官网：部分核验指数的完整加权方式、调样频率、个股权重限制和行业约束。
- 基金招募说明书：指数增强产品的增强模型、跟踪误差目标和主动风险约束。
"""
    REVIEW_MD.write_text(content, encoding="utf-8")


def format_unmatched(values: list[tuple[str, str]]) -> str:
    return "无" if not values else "\n".join(f"- {code}｜{name}" for code, name in values)


def format_fuzzy(values: list[tuple[str, str, str]]) -> str:
    return "无" if not values else "\n".join(
        f"- {code}｜{name} → {index_name}" for code, name, index_name in values
    )


def write_log(
    rules: list[dict[str, str]],
    unmatched: list[tuple[str, str]],
    fuzzy: list[tuple[str, str, str]],
    matched_count: int,
) -> None:
    status_counts = Counter(rule["核验状态"] for rule in rules)
    strategy_counts = Counter(rule["策略类型"] for rule in rules)
    company_counts = Counter(rule["指数公司"] for rule in rules)
    strategy_lines = "\n".join(f"- {key}：{value}" for key, value in sorted(strategy_counts.items()))
    company_lines = "\n".join(f"- {key}：{value}" for key, value in sorted(company_counts.items()))
    content = f"""# Step 8 指数规则处理日志

1. 输入文件路径：`{STEP7_EXCEL}`
2. 输出文件路径：
   - 指数规则 CSV：`{OUTPUT_CSV}`
   - 指数规则 Excel：`{OUTPUT_EXCEL}`
   - Step 8 说明：`{REVIEW_MD}`
   - 同步主 Excel：`{MAIN_EXCEL}`
   - 回写代表产品筛选版：`{STEP7_EXCEL}`
3. DATA_BLOCK 指数规则行数：{len(rules)}
4. 已核验规则数量：{status_counts['已核验']}
5. 部分核验数量：{status_counts['部分核验']}
6. 待二次核验数量：{status_counts['待二次核验']}
7. 指数增强单列观察数量：{strategy_counts['指数增强/多因子']}
8. 是否成功写入主 Excel 的【指数规则表】：是
9. 是否成功回写代表产品筛选版：是（全部产品中成功匹配 {matched_count} 条）
10. 未匹配到指数规则的代表产品：
{format_unmatched(unmatched)}
11. 模糊匹配的代表产品：
{format_fuzzy(fuzzy)}
12. 按策略类型统计的指数数量：
{strategy_lines}
13. 按指数公司统计的指数数量：
{company_lines}
"""
    LOG_MD.write_text(content, encoding="utf-8")


def validate(rules: list[dict[str, str]], step7_validation_path: Path) -> None:
    if OUTPUT_CSV.read_bytes()[:3] != b"\xef\xbb\xbf":
        raise ValueError("指数规则 CSV 不是 UTF-8 with BOM")
    with OUTPUT_CSV.open("r", encoding="utf-8-sig", newline="") as csv_file:
        loaded = list(csv.DictReader(csv_file))
    if loaded != rules:
        raise ValueError("指数规则 CSV 回读不一致")

    workbook = load_workbook(OUTPUT_EXCEL)
    expected = ["代表指数规则表", "已核验规则", "部分核验_待补充", "指数增强单列观察", "指数规则统计"]
    if workbook.sheetnames != expected:
        raise ValueError(f"指数规则 Excel Sheet 不匹配：{workbook.sheetnames}")
    status_counts = Counter(rule["核验状态"] for rule in rules)
    expected_rows = {
        "代表指数规则表": len(rules),
        "已核验规则": status_counts["已核验"],
        "部分核验_待补充": status_counts["部分核验"] + status_counts["待二次核验"],
        "指数增强单列观察": sum(rule["策略类型"] == "指数增强/多因子" for rule in rules),
    }
    for sheet_name, count in expected_rows.items():
        if workbook[sheet_name].max_row != count + 1:
            raise ValueError(f"{sheet_name} 行数不匹配")
    for worksheet in workbook.worksheets:
        if worksheet.freeze_panes != "A2" or not worksheet.auto_filter.ref:
            raise ValueError(f"{worksheet.title} 的冻结窗格或筛选未生效")

    main_workbook = load_workbook(MAIN_EXCEL)
    main_sheet = main_workbook["指数规则表"]
    if main_sheet.max_row != len(rules) + 1:
        raise ValueError("主 Excel 指数规则表行数不匹配")

    step7_workbook = load_workbook(step7_validation_path)
    step7_sheet = step7_workbook["产品池_代表产品筛选版"]
    step7_headers = [clean(cell.value) for cell in step7_sheet[1]]
    if not set(STEP8_FIELDS).issubset(step7_headers):
        raise ValueError("Step 7 产品池未成功新增 Step 8 字段")


def main() -> None:
    rules = read_rules()
    write_csv(rules)
    write_rule_excel(rules)
    sync_main_excel(rules)
    unmatched, fuzzy, matched_count, step7_validation_path = write_back_step7(rules)
    write_review(rules)
    validate(rules, step7_validation_path)
    write_log(rules, unmatched, fuzzy, matched_count)

    status_counts = Counter(rule["核验状态"] for rule in rules)
    print(f"rules={len(rules)}")
    print(f"verified={status_counts['已核验']}")
    print(f"partial={status_counts['部分核验']}")
    print(f"pending={status_counts['待二次核验']}")
    print(f"matched_products={matched_count}")
    print(f"unmatched_representatives={len(unmatched)}")
    print(f"fuzzy_representatives={len(fuzzy)}")
    print(f"step7_writeback_path={step7_validation_path}")


if __name__ == "__main__":
    main()
