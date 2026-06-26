from __future__ import annotations

import hashlib
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


SOURCE = Path(
    r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\wind代码池\sheet6"
    r"\核心策略指数ETF跟踪指数_验收清洗版.xlsx"
)
ARCHIVE_DIR = Path(
    r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据"
    r"\阶段性归档_境内广义策略ETF数据底座"
)
TARGET = ARCHIVE_DIR / "07_核心策略指数ETF_跟踪指数_验收清洗版.xlsx"
INDEX = ARCHIVE_DIR / "阶段性归档索引表.xlsx"
README = ARCHIVE_DIR / "阶段性归档说明.md"


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def headers(ws) -> dict[str, int]:
    return {str(c.value).strip(): c.column for c in ws[1] if c.value is not None}


def auto_width(ws) -> None:
    for col in range(1, ws.max_column + 1):
        values = [str(ws.cell(row, col).value or "") for row in range(1, ws.max_row + 1)]
        ws.column_dimensions[get_column_letter(col)].width = min(
            max(max((len(v) for v in values), default=8) + 2, 10), 70
        )


def update_index(now_text: str, file_size: int, digest: str, backup: Path | None) -> None:
    wb = load_workbook(INDEX)
    ws = wb["归档文件索引"]
    h = headers(ws)
    row = None
    for r in range(2, ws.max_row + 1):
        values = (
            str(ws.cell(r, h.get("模块名称", 1)).value or ""),
            str(ws.cell(r, h.get("归档后文件名", 1)).value or ""),
        )
        if any("跟踪指数" in value and ("表现估值" in value or "验收清洗" in value) for value in values):
            row = r
            break
    if row is None:
        row = ws.max_row + 1

    values = {
        "序号": "07",
        "模块名称": "核心策略指数跟踪数据",
        "归档后文件名": TARGET.name,
        "归档后文件路径": str(TARGET),
        "原始文件路径": str(SOURCE),
        "文件定位": "用于核心策略指数ETF跟踪指数的月度表现、估值、最新期间收益及对照基准分析",
        "正式分析使用sheet": "核心策略指数月度表现估值_清洗版；核心策略指数最新期间收益_清洗版",
        "是否核心产出": "是",
        "注意事项": "82个核心指数全部覆盖；历史估值缺失值未进行人为填补",
        "源文件是否存在": "是",
        "是否成功复制": "是",
        "归档文件大小_字节": file_size,
        "SHA256": digest,
        "校验备注": f"旧版已备份：{backup}" if backup else "源文件与归档文件SHA256一致",
        "文件类型": "Excel",
        "文件用途": "分析核心策略指数的月度收益、估值水平、期间收益及与基准指数的对照表现",
        "数据范围": "168只核心策略指数ETF，对应82个唯一跟踪指数",
        "数据状态": "已完成验收、补充合并、清洗",
        "覆盖情况": "原始文件覆盖75个，补充文件覆盖7个，合并后82个全部覆盖，缺失0",
        "归档时间": now_text,
    }
    for name, value in values.items():
        if name in h:
            ws.cell(row, h[name], value)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    auto_width(ws)

    status_ws = wb["当前数据完成情况"]
    sh = headers(status_ws)
    status_row = None
    for r in range(2, status_ws.max_row + 1):
        if "跟踪指数表现估值" in str(status_ws.cell(r, sh["数据模块"]).value or ""):
            status_row = r
            break
    if status_row is None:
        status_row = status_ws.max_row + 1
    status_values = {
        "数据模块": "跟踪指数表现估值",
        "当前状态": "已完成",
        "对应文件": TARGET.name,
        "主要用途": "核心策略指数月度收益、估值、期间收益及基准对照分析",
        "后续是否还需补充": "需定期更新",
        "备注": "82个核心指数全部覆盖；SPCADMCP.SPI缺少最新期间收益，月度历史数据可用",
    }
    for name, value in status_values.items():
        if name in sh:
            status_ws.cell(status_row, sh[name], value)
    status_ws.freeze_panes = "A2"
    status_ws.auto_filter.ref = f"A1:{get_column_letter(status_ws.max_column)}{status_ws.max_row}"
    auto_width(status_ws)

    pending_ws = wb["后续数据收集清单"]
    ph = headers(pending_ws)
    for r in range(pending_ws.max_row, 1, -1):
        if "跟踪指数表现估值" in str(pending_ws.cell(r, ph["后续数据模块"]).value or ""):
            pending_ws.delete_rows(r)
    pending_ws.freeze_panes = "A2"
    pending_ws.auto_filter.ref = f"A1:{get_column_letter(pending_ws.max_column)}{pending_ws.max_row}"
    auto_width(pending_ws)

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = copy(cell.font)
            cell.font = Font(
                name=cell.font.name,
                size=cell.font.size,
                bold=True,
                italic=cell.font.italic,
                color=cell.font.color,
            )
    wb.save(INDEX)


def update_readme() -> None:
    text = README.read_text(encoding="utf-8-sig")
    completed = "- 核心策略指数跟踪数据：已完成，覆盖168只核心策略指数ETF及82个唯一跟踪指数。"
    if completed not in text:
        marker = "## 7. 后续待补充数据模块"
        position = text.find(marker)
        if position >= 0:
            text = text[:position].rstrip() + "\n" + completed + "\n\n" + text[position:]
    text = text.replace("- 跟踪指数表现估值表；\n", "")
    text = text.replace("- 跟踪指数表现估值表\n", "")

    recommendation = (
        '- 核心策略指数表现估值分析以“07_核心策略指数ETF_跟踪指数_验收清洗版.xlsx”'
        '中的“核心策略指数月度表现估值_清洗版”和“核心策略指数最新期间收益_清洗版”为准；'
    )
    if recommendation not in text:
        marker = "## 6. 已完成数据模块"
        position = text.find(marker)
        if position >= 0:
            text = text[:position].rstrip() + "\n" + recommendation + "\n\n" + text[position:]

    heading = "## 核心策略指数跟踪数据归档"
    paragraph = (
        "07_核心策略指数ETF_跟踪指数_验收清洗版.xlsx 已完成归档。"
        "本文件覆盖168只核心策略指数ETF对应的82个唯一跟踪指数。"
        "原始文件覆盖75个指数，补充文件覆盖7个指数，合并后82个指数全部覆盖，缺失指数0。"
        "文件已完成指数代码标准化、补充数据优先合并、月度主键去重、收益率单位统一、"
        "估值0值缺失化和额外指数隔离，可用于核心策略指数历史表现、估值及期间收益分析。"
    )
    if heading in text:
        start = text.index(heading)
        next_heading = text.find("\n## ", start + len(heading))
        text = text[:start].rstrip() if next_heading < 0 else text[:start].rstrip() + "\n\n" + text[next_heading + 1:].lstrip()
    text = text.rstrip() + f"\n\n{heading}\n\n{paragraph}\n"
    README.write_text(text, encoding="utf-8-sig")


def main() -> None:
    if not SOURCE.exists():
        raise FileNotFoundError(SOURCE)
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    source_hash = sha256(SOURCE)
    now = datetime.now()
    backup = None
    if TARGET.exists():
        backup = TARGET.with_name(
            f"{TARGET.stem}_旧版备份_{now.strftime('%Y%m%d_%H%M%S')}{TARGET.suffix}"
        )
        shutil.copy2(TARGET, backup)
    shutil.copy2(SOURCE, TARGET)
    target_hash = sha256(TARGET)
    if target_hash != source_hash:
        raise RuntimeError("归档文件与源文件哈希不一致")
    update_index(now.strftime("%Y-%m-%d %H:%M:%S"), TARGET.stat().st_size, target_hash, backup)
    update_readme()
    print(f"源文件存在：是")
    print(f"归档复制成功：是")
    print(f"旧版备份：{backup if backup else '未生成（原目标不存在）'}")
    print(f"归档文件路径：{TARGET}")
    print(f"文件大小：{TARGET.stat().st_size} 字节")
    print("阶段性归档索引表：已更新")
    print("阶段性归档说明：已更新")


if __name__ == "__main__":
    main()
