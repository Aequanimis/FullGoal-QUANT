from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = (
    Path.home()
    / "Desktop"
    / "Fullgoal"
    / "课题研究"
    / "处理后数据"
    / "阶段性归档_境内广义策略ETF数据底座"
)
OUTPUT = ROOT / "境内广义策略ETF_核心数据库.xlsx"

FILES = {
    "pool": "01_广义策略ETF_产品池与分类口径_主文件.xlsx",
    "difference": "02_Wind口径_vs_自建策略ETF口径_差异分析.xlsx",
    "size": "03_广义策略ETF_月度规模份额_上市后分析版.xlsx",
    "liquidity": "04_广义策略ETF_月度交易流动性_合并验收清洗版_v2.xlsx",
    "risk": "05_广义策略ETF_收益风险表现_合并验收清洗版.xlsx",
    "holder": "06_广义策略ETF_持有人结构表_半年报年报清洗归档版.xlsx",
    "index": "07_核心策略指数ETF_跟踪指数_验收清洗版.xlsx",
    "rules": "08_核心策略ETF_跟踪指数规则_官网补充核验最终版.xlsx",
}

SHEETS = {
    "pool": "策略ETF_最终统计池",
    "size": "月度规模份额_上市后分析版",
    "liquidity": "流动性数据_分析可用版",
    "risk": "收益风险_合并清洗版",
    "holder_panel": "持有人结构_面板清洗版",
    "holder_latest": "2025年年报_最新截面",
    "index_monthly": "核心策略指数月度表现估值_清洗版",
    "index_latest": "核心策略指数最新期间收益_清洗版",
    "mapping": "核心ETF_指数映射_标准版",
    "rules": "指数规则明细_官网补充版",
    "difference_summary": "口径对比汇总",
    "scope_advice": "研究范围建议",
}

CLASS_FIELDS = [
    "统计口径分类",
    "是否纳入核心策略ETF统计",
    "是否纳入广义策略ETF统计",
    "一级策略大类",
    "二级策略类别",
    "市场范围_二次修正",
    "基金管理人",
]

PRODUCT_COLUMNS = [
    "Wind代码",
    "基金代码",
    "交易代码",
    "证券简称",
    "基金简称",
    "基金全称",
    "基金管理人",
    "基金托管人",
    "上市日期",
    "基金成立日",
    "投资类型_二级分类",
    "跟踪指数代码",
    "跟踪指数名称",
    "最新基金规模(亿)",
    "最新基金份额(份)",
    "一级策略大类",
    "二级策略类别",
    "市场范围_二次修正",
    "统计口径分类",
    "是否纳入核心策略ETF统计",
    "是否纳入广义策略ETF统计",
    "是否红利+策略",
    "分类依据",
    "备注",
]


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="top")
THIN_GREY = Side(style="thin", color="D9E2F3")
HEADER_BORDER = Border(bottom=Side(style="medium", color="1F4E78"))


def normalize_code(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\u3000", " ").upper()


def is_missing(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def load_records(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"未找到 sheet：{path.name} / {sheet_name}")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        return [], []
    headers = [str(x).strip() if x is not None else "" for x in header_row]
    records: list[dict[str, Any]] = []
    for row in rows:
        if all(is_missing(value) for value in row):
            continue
        records.append({headers[index]: row[index] if index < len(row) else None for index in range(len(headers))})
    return headers, records


def select_record(record: dict[str, Any], headers: list[str]) -> list[Any]:
    return [record.get(header) for header in headers]


def update_width(widths: list[int], values: list[Any]) -> None:
    for index, value in enumerate(values):
        text = "" if value is None else str(value)
        widths[index] = max(widths[index], min(max(len(text) + 2, 10), 40))


def header_number_format(header: str) -> str | None:
    if header in {"日期", "上市日期", "基金成立日", "统计日期", "指数基日", "指数发布日期", "成立以来收益率起始时间", "成立以来收益率截止时间", "基金成立日_Wind"}:
        return "yyyy-mm-dd"
    if "日期" in header and "数据" not in header:
        return "yyyy-mm-dd"
    if any(key in header for key in ["规模", "份额", "成交额", "成交量", "净值", "点位", "收益率", "波动率", "回撤", "Sharpe", "SHARPE", "跟踪误差", "信息比率", "超越", "换手", "折溢", "振幅", "PE_", "PB_"]):
        return "0.0000"
    if any(key in header for key in ["数量", "行数", "记录数", "覆盖", "重复"]):
        return "0"
    return None


def write_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    add_filter: bool = True,
) -> None:
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    widths = [min(max(len(str(header)) + 2, 10), 40) for header in headers]

    header_cells = []
    for header in headers:
        cell = WriteOnlyCell(ws, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER
        header_cells.append(cell)
    ws.append(header_cells)

    formats = [header_number_format(header) for header in headers]
    for values in rows:
        output_cells = []
        for index, value in enumerate(values):
            if formats[index] and value is not None:
                cell = WriteOnlyCell(ws, value=value)
                cell.number_format = formats[index]
                cell.alignment = BODY_ALIGNMENT
                output_cells.append(cell)
            else:
                output_cells.append(value)
        ws.append(output_cells)
        update_width(widths, values)

    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.row_dimensions[1].height = 30
    if add_filter and headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"


def add_classification(
    headers: list[str],
    rows: list[dict[str, Any]],
    code_header: str,
    product_by_code: dict[str, dict[str, Any]],
    rename: dict[str, str] | None = None,
) -> tuple[list[str], list[list[Any]], set[str]]:
    rename = rename or {}
    output_headers = [rename.get(header, header) for header in headers]
    existing_by_output = {output: source for source, output in zip(headers, output_headers)}
    for field in CLASS_FIELDS:
        if field not in output_headers:
            output_headers.append(field)

    unmatched: set[str] = set()
    output_rows: list[list[Any]] = []
    for row in rows:
        code = normalize_code(row.get(code_header))
        product = product_by_code.get(code)
        if code and product is None:
            unmatched.add(code)
        values: list[Any] = []
        for output_header in output_headers:
            source_header = existing_by_output.get(output_header)
            current = row.get(source_header) if source_header else None
            if output_header in CLASS_FIELDS and is_missing(current) and product is not None:
                current = product.get(output_header)
            values.append(current)
        output_rows.append(values)
    return output_headers, output_rows, unmatched


def date_bounds(rows: list[dict[str, Any]], header: str) -> tuple[Any, Any]:
    values = [row.get(header) for row in rows if row.get(header) is not None]
    if not values:
        return None, None
    try:
        return min(values), max(values)
    except TypeError:
        values = [str(value) for value in values]
        return min(values), max(values)


def coverage_rate(rows: list[dict[str, Any]], field: str) -> str:
    if not rows:
        return "不适用"
    count = sum(not is_missing(row.get(field)) for row in rows)
    return f"{count}/{len(rows)} ({count / len(rows):.1%})"


def source_prefix(file_key: str) -> Path:
    return ROOT / FILES[file_key]


def rule_verified(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    if text == "否":
        return "是"
    if text == "是":
        return "否"
    if text:
        return text
    return "待官网核验"


def main() -> None:
    for source in FILES.values():
        if not (ROOT / source).exists():
            raise FileNotFoundError(f"缺少归档源文件：{ROOT / source}")

    # 01 产品主表与全库 ETF 分类映射
    pool_headers, pool_records = load_records(source_prefix("pool"), SHEETS["pool"])
    product_by_code = {normalize_code(row.get("Wind代码")): row for row in pool_records if normalize_code(row.get("Wind代码"))}
    product_rows = [
        select_record(row, PRODUCT_COLUMNS)
        for row in sorted(
            pool_records,
            key=lambda row: (
                0 if row.get("是否纳入广义策略ETF统计") == "是" else 1,
                str(row.get("统计口径分类") or ""),
                str(row.get("一级策略大类") or ""),
                normalize_code(row.get("Wind代码")),
            ),
        )
    ]
    broad_codes = {
        normalize_code(row.get("Wind代码"))
        for row in pool_records
        if row.get("是否纳入广义策略ETF统计") == "是" and normalize_code(row.get("Wind代码"))
    }
    core_codes = {
        normalize_code(row.get("Wind代码"))
        for row in pool_records
        if row.get("是否纳入核心策略ETF统计") == "是" and normalize_code(row.get("Wind代码"))
    }
    broad_count = len(broad_codes)
    core_count = len(core_codes)
    enhanced_count = sum(row.get("统计口径分类") == "广义策略ETF_指数增强" for row in pool_records)
    observation_count = sum(row.get("统计口径分类") == "观察池" for row in pool_records)
    excluded_count = sum(row.get("统计口径分类") == "排除" for row in pool_records)

    # 02 月度规模份额
    size_headers, size_records = load_records(source_prefix("size"), SHEETS["size"])
    size_out_headers, size_out_rows, size_unmatched = add_classification(
        size_headers, size_records, "Wind代码", product_by_code
    )
    size_codes = {normalize_code(row.get("Wind代码")) for row in size_records if normalize_code(row.get("Wind代码"))}
    size_start, size_end = date_bounds(size_records, "日期")
    size_not_post = sum(row.get("是否上市后月份") != "是" for row in size_records)

    # 03 月度交易流动性（标准化主键和时间字段名称）
    liquidity_headers, liquidity_records = load_records(source_prefix("liquidity"), SHEETS["liquidity"])
    liquidity_out_headers, liquidity_out_rows, liquidity_unmatched = add_classification(
        liquidity_headers,
        liquidity_records,
        "wind_code",
        product_by_code,
        rename={"wind_code": "Wind代码", "year_month": "日期"},
    )
    liquidity_codes = {normalize_code(row.get("wind_code")) for row in liquidity_records if normalize_code(row.get("wind_code"))}
    liquidity_start, liquidity_end = date_bounds(liquidity_records, "year_month")
    liquidity_amount_missing = sum(is_missing(row.get("month_turnover_amount_sum")) for row in liquidity_records)
    liquidity_discount_missing = sum(is_missing(row.get("month_discount_rate_avg")) for row in liquidity_records)

    # 04 收益风险
    risk_headers, risk_records = load_records(source_prefix("risk"), SHEETS["risk"])
    risk_out_headers, risk_out_rows, risk_unmatched = add_classification(
        risk_headers, risk_records, "Wind代码", product_by_code
    )
    risk_codes_list = [normalize_code(row.get("Wind代码")) for row in risk_records if normalize_code(row.get("Wind代码"))]
    risk_codes = set(risk_codes_list)
    risk_duplicates = len(risk_codes_list) - len(risk_codes)

    # 05 / 06 持有人结构
    holder_panel_headers, holder_panel_records = load_records(source_prefix("holder"), SHEETS["holder_panel"])
    holder_panel_out_headers, holder_panel_out_rows, holder_panel_unmatched = add_classification(
        holder_panel_headers, holder_panel_records, "Wind代码", product_by_code
    )
    holder_panel_codes = {normalize_code(row.get("Wind代码")) for row in holder_panel_records if normalize_code(row.get("Wind代码"))}

    holder_latest_headers, holder_latest_records = load_records(source_prefix("holder"), SHEETS["holder_latest"])
    holder_latest_out_headers, holder_latest_out_rows, holder_latest_unmatched = add_classification(
        holder_latest_headers, holder_latest_records, "Wind代码", product_by_code
    )
    latest_covered = sum(row.get("是否出现在2025年年报") == "是" for row in holder_latest_records)
    latest_uncovered = len(holder_latest_records) - latest_covered
    missing_reason_counter = Counter(
        str(row.get("2025年年报缺失原因"))
        for row in holder_latest_records
        if row.get("是否出现在2025年年报") != "是" and not is_missing(row.get("2025年年报缺失原因"))
    )
    missing_reason_text = "；".join(f"{reason}：{count}只" for reason, count in missing_reason_counter.items()) or "无"

    # 07 / 08 核心指数数据
    index_monthly_headers, index_monthly_records = load_records(source_prefix("index"), SHEETS["index_monthly"])
    index_monthly_rows = [select_record(row, index_monthly_headers) for row in index_monthly_records]
    index_monthly_codes = {normalize_code(row.get("指数代码")) for row in index_monthly_records if normalize_code(row.get("指数代码"))}

    index_latest_headers, index_latest_records = load_records(source_prefix("index"), SHEETS["index_latest"])
    index_latest_rows = [select_record(row, index_latest_headers) for row in index_latest_records]
    index_latest_codes = {normalize_code(row.get("指数代码")) for row in index_latest_records if normalize_code(row.get("指数代码"))}

    # 09 ETF-指数映射：统一为数据库级字段名称，并从产品池补齐统计口径。
    mapping_headers, mapping_records = load_records(source_prefix("rules"), SHEETS["mapping"])
    mapping_out_headers = [
        "Wind代码", "证券简称", "基金简称", "基金全称", "基金管理人",
        "一级策略大类", "二级策略类别", "市场范围_二次修正", "统计口径分类",
        "跟踪指数代码", "跟踪指数名称", "最新基金规模(亿)", "上市日期", "基金成立日",
    ]
    mapping_rows: list[list[Any]] = []
    mapping_unmatched: set[str] = set()
    mapping_index_mismatch = 0
    for row in mapping_records:
        code = normalize_code(row.get("ETF_Code"))
        product = product_by_code.get(code)
        if not product:
            mapping_unmatched.add(code)
        expected_index = normalize_code(product.get("跟踪指数代码")) if product else ""
        actual_index = normalize_code(row.get("Index_Code"))
        if product and expected_index != actual_index:
            mapping_index_mismatch += 1
        mapping_rows.append([
            row.get("ETF_Code"), row.get("ETF_Name"), row.get("基金简称"), row.get("基金全称"), row.get("基金管理人"),
            row.get("一级策略大类"), row.get("二级策略类别"), row.get("市场范围_二次修正"),
            product.get("统计口径分类") if product else None,
            row.get("Index_Code"), row.get("Index_Name"), row.get("最新基金规模(亿)"), row.get("上市日期"), row.get("基金成立日"),
        ])
    mapping_rows.sort(key=lambda r: (str(r[8] or ""), str(r[5] or ""), str(r[6] or ""), str(r[0] or "")))

    # 10 指数规则明细：字段名称统一，保留待官网核验信息。
    rule_headers, rule_records = load_records(source_prefix("rules"), SHEETS["rules"])
    rule_out_headers = [
        "指数代码", "指数名称", "指数发布机构", "指数基日", "指数发布日期", "指数类型",
        "样本空间", "成分股数量", "选样指标", "选样方法", "加权方式", "单只成分权重上限",
        "行业权重限制", "调样频率", "指数简介", "指数编制方案链接", "信息来源",
        "规则完整性状态", "仍缺失字段说明", "是否已核验", "是否仍需官网核验",
    ]
    rule_rows: list[list[Any]] = []
    for row in rule_records:
        rule_rows.append([
            row.get("Index_Code"), row.get("Index_Name"), row.get("发布机构"), row.get("指数基日"), row.get("指数发布日期"), row.get("指数类型"),
            row.get("样本空间"), row.get("成分股数量"), row.get("选样指标"), row.get("选样方法"), row.get("加权方式"), row.get("单只成分权重上限"),
            row.get("行业权重限制"), row.get("调样频率"), row.get("指数简介"), row.get("指数编制方案链接"), row.get("规则补充来源"),
            row.get("规则完整性状态"), row.get("仍缺失字段说明"), rule_verified(row.get("是否仍需官网核验")), row.get("是否仍需官网核验"),
        ])
    rule_rows.sort(key=lambda row: normalize_code(row[0]))
    remaining_rule_review = sum(str(row.get("是否仍需官网核验") or "").strip() != "否" for row in rule_records)

    # 11 口径差异摘要
    diff_headers, diff_records = load_records(source_prefix("difference"), SHEETS["difference_summary"])
    advice_headers, advice_records = load_records(source_prefix("difference"), SHEETS["scope_advice"])
    difference_out_headers = ["摘要类型", "指标", "数值/内容", "说明"]
    difference_rows: list[list[Any]] = []
    for row in diff_records:
        difference_rows.append([
            "口径对比汇总", row.get("指标"), row.get("数量/规模"),
            "；".join(str(row.get(key)) for key in ["统计类别", "说明"] if not is_missing(row.get(key))),
        ])
    for row in advice_records:
        difference_rows.append(["研究范围建议", row.get("项目"), row.get("建议内容"), ""])

    # 12 数据质量总览：将每个正式模块的可用范围与匹配状态集中展示。
    quality_headers = ["模块", "指标", "数值", "状态/说明"]
    quality_rows: list[list[Any]] = []
    quality_rows.extend([
        ["产品主表", "产品主表行数", len(product_rows), "含广义策略ETF、观察池和排除池产品"],
        ["产品主表", "广义策略ETF数量", broad_count, "是否纳入广义策略ETF统计=是"],
        ["产品主表", "核心策略指数ETF数量", core_count, "是否纳入核心策略ETF统计=是"],
        ["产品主表", "指数增强/多因子ETF数量", enhanced_count, "统计口径分类=广义策略ETF_指数增强"],
        ["产品主表", "观察池数量", observation_count, "统计口径分类=观察池"],
        ["产品主表", "排除数量", excluded_count, "统计口径分类=排除"],
        ["月度规模份额", "数据行数", len(size_records), "仅保留上市后有效记录"],
        ["月度规模份额", "覆盖Wind代码数量", len(size_codes), ""],
        ["月度规模份额", "缺失Wind代码数量", len(broad_codes - size_codes), "相对于223只广义策略ETF"],
        ["月度规模份额", "时间起点", size_start, ""],
        ["月度规模份额", "时间终点", size_end, ""],
        ["月度规模份额", "是否仅上市后数据", "是" if size_not_post == 0 else "否", f"非上市后记录：{size_not_post}行"],
        ["月度交易流动性", "数据行数", len(liquidity_records), "正式分析可用版"],
        ["月度交易流动性", "覆盖Wind代码数量", len(liquidity_codes), ""],
        ["月度交易流动性", "缺失Wind代码数量", len(broad_codes - liquidity_codes), "相对于223只广义策略ETF"],
        ["月度交易流动性", "时间起点", liquidity_start, ""],
        ["月度交易流动性", "时间终点", liquidity_end, ""],
        ["月度交易流动性", "月成交额缺失数", liquidity_amount_missing, "month_turnover_amount_sum"],
        ["月度交易流动性", "折溢价率缺失数", liquidity_discount_missing, "month_discount_rate_avg"],
        ["收益风险表现", "覆盖Wind代码数量", len(risk_codes), ""],
        ["收益风险表现", "缺失Wind代码数量", len(broad_codes - risk_codes), "相对于223只广义策略ETF"],
        ["收益风险表现", "重复Wind代码数量", risk_duplicates, ""],
        ["收益风险表现", "年化波动率覆盖率", coverage_rate(risk_records, "年化波动率(%)"), ""],
        ["收益风险表现", "最大回撤覆盖率", coverage_rate(risk_records, "最大回撤(%)"), ""],
        ["收益风险表现", "Sharpe覆盖率", coverage_rate(risk_records, "SHARPE"), ""],
        ["收益风险表现", "跟踪误差覆盖率", coverage_rate(risk_records, "跟踪误差(%)"), ""],
        ["持有人结构", "面板行数", len(holder_panel_records), "半年报/年报口径"],
        ["持有人结构", "面板覆盖Wind代码数量", len(holder_panel_codes), ""],
        ["持有人结构", "最新截面覆盖数量", latest_covered, "2025年年报已覆盖"],
        ["持有人结构", "最新截面未覆盖数量", latest_uncovered, missing_reason_text],
        ["核心指数数据", "核心ETF数量", core_count, ""],
        ["核心指数数据", "唯一跟踪指数数量", len(index_monthly_codes), "按月度指数数据中的指数代码"],
        ["核心指数数据", "月度指数数据覆盖数量", len(index_monthly_codes), ""],
        ["核心指数数据", "最新期间收益覆盖数量", len(index_latest_codes), ""],
        ["核心指数数据", "缺失指数数量", max(0, 82 - len(index_monthly_codes)), "相对于82个核心跟踪指数"],
        ["指数规则", "ETF-指数映射数量", len(mapping_rows), ""],
        ["指数规则", "唯一指数规则数量", len({normalize_code(row.get("Index_Code")) for row in rule_records if normalize_code(row.get("Index_Code"))}), ""],
        ["指数规则", "ETF-指数错配数量", mapping_index_mismatch, "以产品主表的跟踪指数代码为标准"],
        ["指数规则", "待官网核验字段或指数数量", remaining_rule_review, "是否仍需官网核验不等于否"],
    ])

    unmatched_by_module = {
        "月度规模份额": size_unmatched,
        "月度交易流动性": liquidity_unmatched,
        "收益风险表现": risk_unmatched,
        "持有人结构面板": holder_panel_unmatched,
        "持有人结构最新截面": holder_latest_unmatched,
        "ETF-指数映射": mapping_unmatched,
    }
    for module, codes in unmatched_by_module.items():
        if codes:
            for code in sorted(code for code in codes if code):
                quality_rows.append([module, "未匹配Wind代码", code, "未能在01_产品主表按Wind代码匹配"])
        else:
            quality_rows.append([module, "未匹配Wind代码", 0, "全部可与01_产品主表按Wind代码匹配"])

    description_rows = [
        ["数据库名称", "境内广义策略ETF核心数据库"],
        ["产品池口径", "广义策略ETF：223只；核心策略指数ETF：168只；指数增强/多因子ETF：55只"],
        ["主分析建议", "主分析使用核心策略指数ETF口径；广义策略ETF口径用于补充观察指数增强/多因子ETF"],
        ["ETF层面主键", "统一使用Wind代码关联"],
        ["指数层面主键", "统一使用指数代码关联"],
        ["产品分类", "以01_产品主表为准"],
        ["规模份额", "以02_月度规模份额为准；仅保留上市后有效记录"],
        ["交易流动性", "以03_月度交易流动性为准；正式分析使用原归档文件的流动性数据_分析可用版"],
        ["收益风险", "以04_收益风险表现为准"],
        ["持有人结构", "以05_持有人结构面板和06_持有人结构最新截面为准"],
        ["核心指数表现估值", "以07_核心指数月度表现估值和08_核心指数最新收益为准"],
        ["指数规则", "以09_ETF_指数映射和10_指数规则明细为准；待官网核验字段原样保留"],
        ["数据库定位", "本工作簿仅保留正式分析主表、长表、映射表和简化说明；不包含原始备份、过程性验收和异常检查sheet"],
    ]

    # 创建核心数据库工作簿；所有数据均为原归档文件读取后的独立副本。
    output_wb = Workbook(write_only=True)
    write_sheet(output_wb, "00_数据库说明", ["项目", "内容"], description_rows)
    write_sheet(output_wb, "01_产品主表", PRODUCT_COLUMNS, product_rows)
    write_sheet(output_wb, "02_月度规模份额", size_out_headers, size_out_rows)
    write_sheet(output_wb, "03_月度交易流动性", liquidity_out_headers, liquidity_out_rows)
    write_sheet(output_wb, "04_收益风险表现", risk_out_headers, risk_out_rows)
    write_sheet(output_wb, "05_持有人结构面板", holder_panel_out_headers, holder_panel_out_rows)
    write_sheet(output_wb, "06_持有人结构最新截面", holder_latest_out_headers, holder_latest_out_rows)
    write_sheet(output_wb, "07_核心指数月度表现估值", index_monthly_headers, index_monthly_rows)
    write_sheet(output_wb, "08_核心指数最新收益", index_latest_headers, index_latest_rows)
    write_sheet(output_wb, "09_ETF_指数映射", mapping_out_headers, mapping_rows)
    write_sheet(output_wb, "10_指数规则明细", rule_out_headers, rule_rows)
    write_sheet(output_wb, "11_口径差异摘要", difference_out_headers, difference_rows)
    write_sheet(output_wb, "12_数据质量总览", quality_headers, quality_rows)
    output_wb.save(OUTPUT)

    all_unmatched = set().union(*unmatched_by_module.values())
    print(f"输出文件路径：{OUTPUT}")
    print(f"01_产品主表行数：{len(product_rows)}")
    print(f"广义策略ETF数量：{broad_count}")
    print(f"核心策略指数ETF数量：{core_count}")
    print(f"指数增强/多因子ETF数量：{enhanced_count}")
    print(f"02_月度规模份额覆盖Wind代码数量：{len(size_codes)}")
    print(f"03_月度交易流动性覆盖Wind代码数量：{len(liquidity_codes)}")
    print(f"04_收益风险表现覆盖Wind代码数量：{len(risk_codes)}")
    print(f"05_持有人结构面板覆盖Wind代码数量：{len(holder_panel_codes)}")
    print(f"核心指数唯一数量：{len(index_monthly_codes)}")
    print(f"所有模块未匹配Wind代码数量：{len(all_unmatched)}")
    print("12_数据质量总览：已成功生成")


if __name__ == "__main__":
    main()
