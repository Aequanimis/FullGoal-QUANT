from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\lvdon\Desktop\Fullgoal\课题研究\处理后数据\阶段性归档_境内广义策略ETF数据底座")
INDEX = ROOT / "阶段性归档索引表.xlsx"
README = ROOT / "阶段性归档说明.md"

OLD_SUBDIRS = [
    "01_产品池与分类口径",
    "02_Wind口径_vs_自建口径差异",
    "03_月度规模份额",
    "04_月度交易流动性",
    "99_归档说明",
]


def main() -> None:
    wb = load_workbook(INDEX)
    ws = wb["归档文件索引"]
    headers = [cell.value for cell in ws[1]]
    path_col = headers.index("归档后文件路径") + 1
    file_col = headers.index("归档后文件名") + 1
    for row in range(2, ws.max_row + 1):
        filename = ws.cell(row, file_col).value
        ws.cell(row, path_col, str(ROOT / filename))
    wb.save(INDEX)

    content = README.read_text(encoding="utf-8-sig")
    for subdir in OLD_SUBDIRS:
        content = content.replace(f"{ROOT}\\{subdir}\\", f"{ROOT}\\")
    README.write_text(content, encoding="utf-8-sig", newline="\n")

    print(f"索引路径已更新：{INDEX}")
    print(f"说明路径已更新：{README}")


if __name__ == "__main__":
    main()
