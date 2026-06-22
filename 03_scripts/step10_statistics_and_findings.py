"""Step 10：基于本地工作簿生成策略 ETF 统计分析和初步发现。"""

from __future__ import annotations

import csv
import unicodedata
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PRODUCT_EXCEL = PROJECT_ROOT / "01_processed_data" / "classification" / "境内策略ETF产品池_代表产品筛选版.xlsx"
RULES_EXCEL = PROJECT_ROOT / "01_processed_data" / "index_rules" / "代表指数规则整理.xlsx"
CLASSIFICATION_EXCEL = PROJECT_ROOT / "01_processed_data" / "classification" / "策略分类表.xlsx"
MAIN_EXCEL = PROJECT_ROOT / "02_outputs" / "excel" / "境内策略ETF产品梳理_初版.xlsx"

STATISTICS_DIR = PROJECT_ROOT / "01_processed_data" / "statistics"
SUMMARY_DIR = PROJECT_ROOT / "02_outputs" / "summary"
MENTOR_DIR = PROJECT_ROOT / "02_outputs" / "mentor_feedback"
STATISTICS_EXCEL = STATISTICS_DIR / "策略ETF统计分析表.xlsx"
PRODUCT_CSV = STATISTICS_DIR / "产品池统计.csv"
STRATEGY_CSV = STATISTICS_DIR / "策略类型统计.csv"
REPRESENTATIVE_CSV = STATISTICS_DIR / "代表产品统计.csv"
RULES_CSV = STATISTICS_DIR / "指数规则统计.csv"
FINDINGS_MD = SUMMARY_DIR / "境内策略ETF初步发现.md"
MENTOR_MD = MENTOR_DIR / "Step10_mentor讨论要点.md"
LOG_MD = STATISTICS_DIR / "step10_statistics_log.md"

LEVELS = ["主线", "补充观察", "单列观察", "跨境补充", "仅作线索"]
PENDING_KEYWORDS = [
    "规模", "上市日期", "基金公司", "管理人", "指数全称", "跟踪指数", "费率", "成交额", "Wind", "Choice", "校验"
]
MAINLINE_STRATEGIES = {
    "红利", "红利低波", "自由现金流", "红利质量", "质量", "价值", "成长", "低波", "基本面策略"
}


def clean(value: object) -> str:
    return "" if value is None else str(value).strip(" \t\r\n")


def ensure_directories() -> None:
    for directory in (STATISTICS_DIR, SUMMARY_DIR, MENTOR_DIR):
        directory.mkdir(parents=True, exist_ok=True)


def verify_inputs() -> None:
    required = [PRODUCT_EXCEL, RULES_EXCEL, CLASSIFICATION_EXCEL, MAIN_EXCEL]
    missing = [path for path in required if not path.exists()]
    if missing:
        ensure_directories()
        LOG_MD.write_text(
            "# Step 10 统计日志\n\n缺少输入文件：\n"
            + "\n".join(f"- `{path}`" for path in missing)
            + "\n\nStep 10 已停止，未继续执行。\n",
            encoding="utf-8",
        )
        raise FileNotFoundError("缺少输入文件：" + "；".join(str(path) for path in missing))


def read_records(path: Path, sheet_name: str) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"{path.name} 中不存在 Sheet【{sheet_name}】")
    worksheet = workbook[sheet_name]
    headers = [clean(cell.value) for cell in worksheet[1]]
    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        record = {header: clean(value) for header, value in zip(headers, values)}
        if any(record.values()):
            rows.append(record)
    return rows


def require_fields(rows: list[dict[str, str]], fields: set[str], label: str) -> None:
    available = set(rows[0]) if rows else set()
    missing = fields - available
    if missing:
        raise ValueError(f"{label} 缺少字段：{sorted(missing)}")


def ratio(count: int, total: int) -> float:
    return count / total if total else 0.0


def calculate_statistics(
    products: list[dict[str, str]],
    rules: list[dict[str, str]],
    classifications: list[dict[str, str]],
) -> dict[str, object]:
    total_products = len(products)
    level_counts = Counter(product.get("建议纳入级别") or "其他或待判断" for product in products)
    known_level_total = sum(level_counts[level] for level in LEVELS)
    level_counts["其他或待判断"] = total_products - known_level_total
    pending_products = [product for product in products if product.get("是否待校验") == "是"]
    representatives = [product for product in products if product.get("是否Step7代表产品") == "是"]
    non_representative_count = total_products - len(representatives)

    overview = [
        {"统计项目": "当前产品池产品总数", "数量": total_products, "占比": 1.0, "备注": "当前核心产品池"},
        {"统计项目": "主线产品数量", "数量": level_counts["主线"], "占比": ratio(level_counts["主线"], total_products), "备注": "建议纳入级别=主线"},
        {"统计项目": "补充观察产品数量", "数量": level_counts["补充观察"], "占比": ratio(level_counts["补充观察"], total_products), "备注": "建议纳入级别=补充观察"},
        {"统计项目": "单列观察产品数量", "数量": level_counts["单列观察"], "占比": ratio(level_counts["单列观察"], total_products), "备注": "建议纳入级别=单列观察"},
        {"统计项目": "跨境补充产品数量", "数量": level_counts["跨境补充"], "占比": ratio(level_counts["跨境补充"], total_products), "备注": "建议纳入级别=跨境补充"},
        {"统计项目": "仅作线索产品数量", "数量": level_counts["仅作线索"], "占比": ratio(level_counts["仅作线索"], total_products), "备注": "建议纳入级别=仅作线索"},
        {"统计项目": "待校验产品数量", "数量": len(pending_products), "占比": ratio(len(pending_products), total_products), "备注": "是否待校验=是"},
        {"统计项目": "Step7代表产品数量", "数量": len(representatives), "占比": ratio(len(representatives), total_products), "备注": "是否Step7代表产品=是"},
        {"统计项目": "非代表产品数量", "数量": non_representative_count, "占比": ratio(non_representative_count, total_products), "备注": "未入选Step7代表产品"},
    ]

    strategy_counts = Counter(product.get("建议策略大类") or product.get("策略大类") or "未分类" for product in products)
    strategy_rows = []
    for strategy, count in sorted(strategy_counts.items(), key=lambda item: (-item[1], item[0])):
        strategy_products = [product for product in products if (product.get("建议策略大类") or product.get("策略大类") or "未分类") == strategy]
        strategy_rows.append(
            {
                "策略大类": strategy,
                "产品数量": count,
                "产品占比": ratio(count, total_products),
                "主线产品数量": sum(product.get("建议纳入级别") == "主线" for product in strategy_products),
                "补充观察产品数量": sum(product.get("建议纳入级别") == "补充观察" for product in strategy_products),
                "待校验产品数量": sum(product.get("是否待校验") == "是" for product in strategy_products),
                "代表产品数量": sum(product.get("是否Step7代表产品") == "是" for product in strategy_products),
            }
        )

    level_rows = [
        {
            "建议纳入级别": level,
            "产品数量": level_counts[level],
            "产品占比": ratio(level_counts[level], total_products),
        }
        for level in [*LEVELS, "其他或待判断"]
    ]

    cross_rows = []
    for strategy, _ in sorted(strategy_counts.items(), key=lambda item: (-item[1], item[0])):
        strategy_products = [product for product in products if (product.get("建议策略大类") or product.get("策略大类") or "未分类") == strategy]
        row: dict[str, object] = {"策略大类": strategy}
        for level in [*LEVELS, "其他或待判断"]:
            row[level] = sum((product.get("建议纳入级别") or "其他或待判断") == level for product in strategy_products)
        row["合计"] = len(strategy_products)
        cross_rows.append(row)

    representative_strategy_counts = Counter(product.get("建议策略大类") or product.get("策略大类") for product in representatives)
    representative_level_counts = Counter(product.get("建议纳入级别") for product in representatives)
    representative_grade_counts = Counter(product.get("Step7代表产品级别") or "未填写" for product in representatives)
    representative_rows = [
        {"统计类型": "汇总", "统计项目": "Step7代表产品总数", "数量": len(representatives), "占代表产品比": 1.0}
    ]
    representative_rows.extend(
        {"统计类型": "按策略类型", "统计项目": strategy, "数量": count, "占代表产品比": ratio(count, len(representatives))}
        for strategy, count in sorted(representative_strategy_counts.items(), key=lambda item: (-item[1], item[0]))
    )
    representative_rows.extend(
        {"统计类型": "按建议纳入级别", "统计项目": level, "数量": representative_level_counts[level], "占代表产品比": ratio(representative_level_counts[level], len(representatives))}
        for level in LEVELS
    )
    representative_rows.extend(
        {"统计类型": "按Step7代表产品级别", "统计项目": grade, "数量": count, "占代表产品比": ratio(count, len(representatives))}
        for grade, count in sorted(representative_grade_counts.items(), key=lambda item: (-item[1], item[0]))
    )

    pending_by_strategy = Counter(
        product.get("建议策略大类") or product.get("策略大类") or "未分类" for product in pending_products
    )
    keyword_counts = Counter()
    for product in pending_products:
        text = product.get("待校验事项", "")
        for keyword in PENDING_KEYWORDS:
            if keyword in text:
                keyword_counts[keyword] += 1
    pending_rows = [
        {"统计类型": "汇总", "统计项目": "待校验产品数量", "数量": len(pending_products), "占比": ratio(len(pending_products), total_products)},
        {"统计类型": "汇总", "统计项目": "待校验产品占比", "数量": len(pending_products), "占比": ratio(len(pending_products), total_products)},
    ]
    pending_rows.extend(
        {
            "统计类型": "按策略大类",
            "统计项目": strategy,
            "数量": count,
            "占比": ratio(count, strategy_counts[strategy]),
        }
        for strategy, count in sorted(pending_by_strategy.items(), key=lambda item: (-item[1], item[0]))
    )
    pending_rows.extend(
        {"统计类型": "待校验事项关键词", "统计项目": keyword, "数量": keyword_counts[keyword], "占比": ratio(keyword_counts[keyword], len(pending_products))}
        for keyword in PENDING_KEYWORDS
    )

    rule_status_counts = Counter(rule.get("核验状态") for rule in rules)
    rule_strategy_counts = Counter(rule.get("策略类型") for rule in rules)
    rule_company_counts = Counter(rule.get("指数公司") for rule in rules)
    top10_pending = sum("待" in rule.get("前十大成分股", "") for rule in rules)
    industry_pending = sum("待" in rule.get("行业分布", "") for rule in rules)
    rule_rows = [
        {"统计类型": "汇总", "统计项目": "指数规则总数", "数量": len(rules)},
        {"统计类型": "汇总", "统计项目": "已核验数量", "数量": rule_status_counts["已核验"]},
        {"统计类型": "汇总", "统计项目": "部分核验数量", "数量": rule_status_counts["部分核验"]},
        {"统计类型": "汇总", "统计项目": "待二次核验数量", "数量": rule_status_counts["待二次核验"]},
        {"统计类型": "汇总", "统计项目": "指数增强单列观察数量", "数量": rule_strategy_counts["指数增强/多因子"]},
        {"统计类型": "汇总", "统计项目": "前十大成分股待补充数量", "数量": top10_pending},
        {"统计类型": "汇总", "统计项目": "行业分布待补充数量", "数量": industry_pending},
    ]
    rule_rows.extend(
        {"统计类型": "按策略类型", "统计项目": strategy, "数量": count}
        for strategy, count in sorted(rule_strategy_counts.items(), key=lambda item: (-item[1], item[0]))
    )
    rule_rows.extend(
        {"统计类型": "按指数公司", "统计项目": company, "数量": count}
        for company, count in sorted(rule_company_counts.items(), key=lambda item: (-item[1], item[0]))
    )

    maturity_values = [row.get("境内发展成熟度", "") for row in classifications]
    classification_rows = [
        {"统计项目": "策略类型总数", "数量": len(classifications)},
        {"统计项目": "主线策略数量", "数量": sum(row.get("策略类型") in MAINLINE_STRATEGIES for row in classifications)},
        {"统计项目": "补充观察/单列观察策略数量", "数量": sum(row.get("策略类型") not in MAINLINE_STRATEGIES and "待补充" not in row.get("境内发展成熟度", "") for row in classifications)},
        {"统计项目": "待补充策略数量", "数量": sum("待补充" in value for value in maturity_values)},
        {"统计项目": "成熟度高的策略数量", "数量": sum("成熟度高" in value for value in maturity_values)},
        {"统计项目": "快速发展的策略数量", "数量": sum("快速发展" in value for value in maturity_values)},
        {"统计项目": "包含待补充的策略数量", "数量": sum("待补充" in value for value in maturity_values)},
    ]

    return {
        "overview": overview,
        "strategy_rows": strategy_rows,
        "level_rows": level_rows,
        "cross_rows": cross_rows,
        "representative_rows": representative_rows,
        "pending_rows": pending_rows,
        "rule_rows": rule_rows,
        "classification_rows": classification_rows,
        "level_counts": level_counts,
        "strategy_counts": strategy_counts,
        "representatives": representatives,
        "pending_products": pending_products,
        "rule_status_counts": rule_status_counts,
    }


def write_csv(path: Path, fields: list[str], rows: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fields, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def display_width(value: object) -> int:
    text = clean(value)
    return sum(
        2 if unicodedata.east_asian_width(character) in {"W", "F", "A"} else 1
        for character in text
    )


def append_rows(worksheet, fields: list[str], rows: list[dict[str, object]]) -> None:
    worksheet.append(fields)
    for row in rows:
        worksheet.append([row.get(field, "") for field in fields])


def format_sheet(worksheet) -> None:
    max_row = max(worksheet.max_row, 1)
    max_column = max(worksheet.max_column, 1)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_column)}{max_row}"
    worksheet.sheet_view.zoomScale = 90
    fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    header_font = Font(name="Microsoft YaHei", size=10, bold=True)
    body_font = Font(name="Microsoft YaHei", size=10)
    border = Border(bottom=Side(style="thin", color="A6A6A6"))
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    worksheet.row_dimensions[1].height = 28

    headers = [clean(cell.value) for cell in worksheet[1]]
    for row in worksheet.iter_rows(min_row=2, max_row=max_row, max_col=max_column):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_index, header in enumerate(headers, start=1):
        values = [worksheet.cell(row, column_index).value for row in range(1, max_row + 1)]
        measured = max(display_width(value) for value in values) + 3
        cap = 68 if header in {"内容", "备注"} else 30
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(measured, 12), cap)
        if "占比" in header:
            for row in range(2, max_row + 1):
                worksheet.cell(row, column_index).number_format = "0.00%"
        if header in {"数量", "产品数量", "主线产品数量", "补充观察产品数量", "待校验产品数量", "代表产品数量", "合计"}:
            for row in range(2, max_row + 1):
                worksheet.cell(row, column_index).number_format = "0"


def create_findings_sections(stats: dict[str, object]) -> list[dict[str, str]]:
    overview = {row["统计项目"]: row["数量"] for row in stats["overview"]}  # type: ignore[index]
    strategy_rows: list[dict[str, object]] = stats["strategy_rows"]  # type: ignore[assignment]
    top_text = "、".join(f"{row['策略大类']}（{row['产品数量']}只）" for row in strategy_rows[:5])
    small_text = "、".join(f"{row['策略大类']}（{row['产品数量']}只）" for row in strategy_rows if int(row["产品数量"]) <= 3)
    rule_status: Counter = stats["rule_status_counts"]  # type: ignore[assignment]
    rule_total = sum(rule_status.values())
    return [
        {
            "标题": "1. 当前产品池覆盖情况",
            "内容": (
                f"当前产品池共{overview['当前产品池产品总数']}只，其中主线{overview['主线产品数量']}只、"
                f"补充观察{overview['补充观察产品数量']}只、单列观察{overview['单列观察产品数量']}只、"
                f"跨境补充{overview['跨境补充产品数量']}只。当前不是全量117只数据库，而是核心产品池初版。"
            ),
        },
        {
            "标题": "2. 策略类型分布",
            "内容": f"当前产品数量较多的策略为{top_text}。红利、红利低波和自由现金流构成优先研究主线；当前样本较少的类别包括{small_text}。",
        },
        {
            "标题": "3. 代表产品情况",
            "内容": f"Step7共筛选{overview['Step7代表产品数量']}只代表产品，其中主线代表产品20只，覆盖红利、红利低波、自由现金流、红利质量、价值、成长、低波和基本面策略等方向，后续用于指数规则拆解。",
        },
        {
            "标题": "4. 指数规则整理情况",
            "内容": f"已整理{rule_total}条代表指数规则，其中已核验{rule_status['已核验']}条、部分核验{rule_status['部分核验']}条、待二次核验{rule_status['待二次核验']}条。前十大成分股和行业分布需后续通过Wind、Choice或指数官网factsheet补齐。",
        },
        {
            "标题": "5. 初步判断",
            "内容": "境内策略ETF已由单一红利扩展至红利低波、自由现金流、红利质量、价值、成长、低波和基本面策略。当前核心产品池中，红利、红利低波、自由现金流最值得优先展开；红利质量可作为红利策略升级方向。价值、成长、低波和基本面策略已有布局，但仍需结合规模和流动性判断成熟度。指数增强/多因子、央国企红利/股东回报、ESG、港股通红利/低波暂作为补充观察；纯质量、等权/另类加权和动量仍需补充。",
        },
        {
            "标题": "6. 后续待补充事项",
            "内容": "待Wind/Choice到位后补充规模、成交额、费率、成立日期和上市日期；补齐全量117只策略ETF；补充纯质量、等权和动量等缺口；校验所有待校验产品；导出前十大成分股和行业分布。",
        },
    ]


def write_statistics_excel(stats: dict[str, object], findings_sections: list[dict[str, str]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = [
        ("产品总览", ["统计项目", "数量", "占比", "备注"], stats["overview"]),
        ("策略大类统计", ["策略大类", "产品数量", "产品占比", "主线产品数量", "补充观察产品数量", "待校验产品数量", "代表产品数量"], stats["strategy_rows"]),
        ("纳入级别统计", ["建议纳入级别", "产品数量", "产品占比"], stats["level_rows"]),
        ("策略大类_纳入级别交叉", ["策略大类", *LEVELS, "其他或待判断", "合计"], stats["cross_rows"]),
        ("代表产品统计", ["统计类型", "统计项目", "数量", "占代表产品比"], stats["representative_rows"]),
        ("待校验统计", ["统计类型", "统计项目", "数量", "占比"], stats["pending_rows"]),
        ("指数规则统计", ["统计类型", "统计项目", "数量"], stats["rule_rows"]),
        ("策略分类统计", ["统计项目", "数量"], stats["classification_rows"]),
        ("初步发现", ["标题", "内容"], findings_sections),
    ]
    for sheet_name, fields, rows in sheets:
        worksheet = workbook.create_sheet(sheet_name)
        append_rows(worksheet, fields, rows)  # type: ignore[arg-type]
        format_sheet(worksheet)

    strategy_sheet = workbook["策略大类统计"]
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "策略大类产品数量"
    chart.y_axis.title = "策略大类"
    chart.x_axis.title = "产品数量"
    data = Reference(strategy_sheet, min_col=2, min_row=1, max_row=strategy_sheet.max_row)
    categories = Reference(strategy_sheet, min_col=1, min_row=2, max_row=strategy_sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 9
    chart.width = 16
    strategy_sheet.add_chart(chart, "I2")
    workbook.save(STATISTICS_EXCEL)


def build_findings_markdown(stats: dict[str, object]) -> str:
    overview = {row["统计项目"]: row["数量"] for row in stats["overview"]}  # type: ignore[index]
    strategy_rows: list[dict[str, object]] = stats["strategy_rows"]  # type: ignore[assignment]
    top_text = "、".join(f"{row['策略大类']}（{row['产品数量']}只）" for row in strategy_rows[:5])
    small_text = "、".join(f"{row['策略大类']}（{row['产品数量']}只）" for row in strategy_rows if int(row["产品数量"]) <= 3)
    representative_rows: list[dict[str, object]] = stats["representative_rows"]  # type: ignore[assignment]
    rep_total = next(int(row["数量"]) for row in representative_rows if row["统计项目"] == "Step7代表产品总数")
    rep_mainline = next(int(row["数量"]) for row in representative_rows if row["统计类型"] == "按建议纳入级别" and row["统计项目"] == "主线")
    rule_status: Counter = stats["rule_status_counts"]  # type: ignore[assignment]
    rule_total = sum(rule_status.values())
    return f"""# 境内策略 ETF 初步统计与发现

## 1. 当前产品池覆盖情况

- 当前产品池产品总数：{overview['当前产品池产品总数']}只；
- 主线产品：{overview['主线产品数量']}只；
- 补充观察：{overview['补充观察产品数量']}只；单列观察：{overview['单列观察产品数量']}只；跨境补充：{overview['跨境补充产品数量']}只；
- 当前产品池不是全量117只数据库，而是核心产品池初版。

### 产品池覆盖范围说明

当前产品池为公开搜索和代表性整理形成的核心产品池，不等于全市场完整 117 只策略 ETF 数据库。根据前期研究口径，2026 年一季度末境内策略类 ETF 约 117 只，跟踪 59 个标的指数；当前产品池主要覆盖红利、红利低波、自由现金流、红利质量、价值、成长、低波、基本面策略、指数增强/多因子、央国企红利/股东回报、ESG 和港股通红利/低波等方向，但对纯质量、等权/另类加权、动量、更多价值/成长细分产品、更多同指数多管理人产品仍未完全覆盖。后续需在 Wind / Choice 到位后补齐全量产品池。

## 2. 策略类型分布

当前产品数量较多的策略类型为{top_text}。从核心产品池的产品数量和主线定位看，红利、红利低波、自由现金流是当前优先研究的三条主线。样本相对较少的类别包括{small_text}，现阶段不宜仅依据产品数量判断其市场成熟度。

## 3. 代表产品情况

Step7共筛选{rep_total}只代表产品，其中主线代表产品{rep_mainline}只。代表产品主要覆盖红利、红利低波、自由现金流、红利质量、价值、成长、低波、基本面策略，以及若干补充观察方向，后续将用于指数规则拆解和同类比较。

## 4. 指数规则整理情况

当前共整理{rule_total}条代表指数规则，其中已核验{rule_status['已核验']}条、部分核验{rule_status['部分核验']}条、待二次核验{rule_status['待二次核验']}条。前十大成分股和行业分布属于动态数据，后续需通过 Wind、Choice、指数官网 factsheet 或指数浏览器按日期补齐。

## 5. 初步判断

- 境内策略 ETF 已经从单一红利扩展到红利低波、自由现金流、红利质量、价值、成长、低波、基本面策略等方向；
- 从当前核心产品池看，红利、红利低波、自由现金流是最值得优先展开的三条主线；
- 红利质量可以作为红利策略升级方向；
- 价值、成长、低波、基本面策略已有布局，但整体需要结合规模和流动性进一步判断成熟度；
- 指数增强/多因子、央国企红利/股东回报、ESG、港股通红利/低波暂作为补充观察；
- 纯质量、等权/另类加权、动量等方向后续需要补充。

## 6. 后续待补充事项

- 等 Wind / Choice 到位后补规模；
- 补成交额；
- 补费率；
- 补成立日期和上市日期；
- 补全量 117 只策略 ETF；
- 补纯质量、等权、动量等缺口；
- 校验所有待校验产品；
- 导出前十大成分股和行业分布。
"""


def write_mentor_markdown() -> None:
    content = """# Step 10 Mentor 讨论要点

1. 当前研究是否先以核心产品池为主，还是需要尽快补齐全量 117 只策略 ETF？
2. ESG、央国企红利/股东回报、港股通红利是否纳入主线，还是作为补充观察？
3. 指数增强 ETF 是否纳入策略 ETF 分析范围，还是单列对比？
4. 后续海外案例是否优先对应红利低波、自由现金流、质量、价值、低波、护城河等成熟策略？
5. 最终成果更偏产品发展建议，还是偏指数策略设计建议？

## 可发给 mentor 的简短汇报

好的老师，我这边先用公开信息整理了一版境内策略 ETF 的核心产品池，并按照红利、红利低波、自由现金流、红利质量、价值、成长、低波、基本面策略等方向做了分类。当前这版不是全量 117 只策略 ETF 数据库，而是先围绕代表性产品和代表指数搭建分析框架。下一步我会等 Wind / Choice 权限到位后补齐规模、成交额、费率、上市日期以及纯质量、等权、动量等缺口方向，同时继续完善代表指数的编制规则。
"""
    MENTOR_MD.write_text(content, encoding="utf-8")


def sync_main_excel(stats: dict[str, object], findings_sections: list[dict[str, str]]) -> None:
    workbook = load_workbook(MAIN_EXCEL)
    for sheet_name in ("统计分析", "初步发现"):
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])

    statistics_sheet = workbook.create_sheet("统计分析")
    main_rows: list[dict[str, object]] = []
    for row in stats["overview"]:  # type: ignore[union-attr]
        main_rows.append({"统计模块": "产品总览", "统计项目": row["统计项目"], "数量": row["数量"], "占比": row["占比"], "备注": row["备注"]})
    for row in stats["strategy_rows"]:  # type: ignore[union-attr]
        main_rows.append({"统计模块": "策略大类", "统计项目": row["策略大类"], "数量": row["产品数量"], "占比": row["产品占比"], "备注": f"主线{row['主线产品数量']}；待校验{row['待校验产品数量']}；代表产品{row['代表产品数量']}"})
    for row in stats["level_rows"]:  # type: ignore[union-attr]
        main_rows.append({"统计模块": "纳入级别", "统计项目": row["建议纳入级别"], "数量": row["产品数量"], "占比": row["产品占比"], "备注": ""})
    for row in stats["representative_rows"]:  # type: ignore[union-attr]
        if row["统计类型"] in {"汇总", "按建议纳入级别"}:
            main_rows.append({"统计模块": "代表产品", "统计项目": row["统计项目"], "数量": row["数量"], "占比": row["占代表产品比"], "备注": row["统计类型"]})
    for row in stats["rule_rows"]:  # type: ignore[union-attr]
        if row["统计类型"] == "汇总":
            main_rows.append({"统计模块": "指数规则", "统计项目": row["统计项目"], "数量": row["数量"], "占比": "", "备注": ""})
    append_rows(statistics_sheet, ["统计模块", "统计项目", "数量", "占比", "备注"], main_rows)
    format_sheet(statistics_sheet)

    findings_sheet = workbook.create_sheet("初步发现")
    append_rows(findings_sheet, ["标题", "内容"], findings_sections)
    format_sheet(findings_sheet)
    workbook.save(MAIN_EXCEL)


def write_log(
    products: list[dict[str, str]], rules: list[dict[str, str]], classifications: list[dict[str, str]], stats: dict[str, object]
) -> None:
    overview = {row["统计项目"]: row["数量"] for row in stats["overview"]}  # type: ignore[index]
    rule_status: Counter = stats["rule_status_counts"]  # type: ignore[assignment]
    content = f"""# Step 10 统计分析日志

1. 输入文件路径：
   - 产品池：`{PRODUCT_EXCEL}`
   - 指数规则：`{RULES_EXCEL}`
   - 策略分类：`{CLASSIFICATION_EXCEL}`
   - 主 Excel：`{MAIN_EXCEL}`
2. 输出文件路径：
   - 统计分析 Excel：`{STATISTICS_EXCEL}`
   - 产品池统计 CSV：`{PRODUCT_CSV}`
   - 策略类型统计 CSV：`{STRATEGY_CSV}`
   - 代表产品统计 CSV：`{REPRESENTATIVE_CSV}`
   - 指数规则统计 CSV：`{RULES_CSV}`
   - 初步发现：`{FINDINGS_MD}`
   - Mentor 讨论要点：`{MENTOR_MD}`
3. 产品池读取行数：{len(products)}
4. 指数规则读取行数：{len(rules)}
5. 策略分类读取行数：{len(classifications)}
6. 当前产品池产品总数：{overview['当前产品池产品总数']}
7. 主线产品数量：{overview['主线产品数量']}
8. 补充观察产品数量：{overview['补充观察产品数量']}
9. 单列观察产品数量：{overview['单列观察产品数量']}
10. 跨境补充产品数量：{overview['跨境补充产品数量']}
11. 代表产品数量：{overview['Step7代表产品数量']}
12. 已核验指数规则数量：{rule_status['已核验']}
13. 部分核验指数规则数量：{rule_status['部分核验']}
14. 待二次核验指数规则数量：{rule_status['待二次核验']}
15. 是否成功生成统计分析 Excel：是
16. 是否成功生成初步发现 Markdown：是
17. 是否成功生成 mentor 讨论要点：是
18. 是否成功同步更新主 Excel：是
"""
    LOG_MD.write_text(content, encoding="utf-8")


def validate(
    products: list[dict[str, str]], rules: list[dict[str, str]], classifications: list[dict[str, str]]
) -> None:
    for path in (PRODUCT_CSV, STRATEGY_CSV, REPRESENTATIVE_CSV, RULES_CSV):
        if path.read_bytes()[:3] != b"\xef\xbb\xbf":
            raise ValueError(f"CSV 不是 UTF-8 with BOM：{path}")
    workbook = load_workbook(STATISTICS_EXCEL)
    expected_sheets = [
        "产品总览", "策略大类统计", "纳入级别统计", "策略大类_纳入级别交叉", "代表产品统计",
        "待校验统计", "指数规则统计", "策略分类统计", "初步发现",
    ]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(f"统计分析 Excel Sheet 不匹配：{workbook.sheetnames}")
    for worksheet in workbook.worksheets:
        if worksheet.freeze_panes != "A2" or not worksheet.auto_filter.ref:
            raise ValueError(f"{worksheet.title} 的冻结窗格或筛选未生效")
    if workbook["产品总览"].max_row != 10:
        raise ValueError("产品总览统计行数异常")

    main_workbook = load_workbook(MAIN_EXCEL)
    if not {"统计分析", "初步发现"}.issubset(main_workbook.sheetnames):
        raise ValueError("主 Excel 未成功新增统计分析或初步发现 Sheet")
    if len(products) != 58 or len(rules) != 26 or len(classifications) != 15:
        raise ValueError("输入数据行数与当前项目阶段预期不一致")


def main() -> None:
    ensure_directories()
    verify_inputs()
    products = read_records(PRODUCT_EXCEL, "产品池_代表产品筛选版")
    rules = read_records(RULES_EXCEL, "代表指数规则表")
    classifications = read_records(CLASSIFICATION_EXCEL, "策略分类表")
    require_fields(products, {"建议纳入级别", "是否Step7代表产品", "是否待校验", "建议策略大类", "待校验事项"}, "产品池")
    require_fields(rules, {"策略类型", "指数公司", "核验状态", "前十大成分股", "行业分布"}, "指数规则")
    require_fields(classifications, {"策略类型", "境内发展成熟度"}, "策略分类")

    stats = calculate_statistics(products, rules, classifications)
    findings_sections = create_findings_sections(stats)
    write_csv(PRODUCT_CSV, ["统计项目", "数量", "占比", "备注"], stats["overview"])  # type: ignore[arg-type]
    write_csv(STRATEGY_CSV, ["策略大类", "产品数量", "产品占比", "主线产品数量", "补充观察产品数量", "待校验产品数量", "代表产品数量"], stats["strategy_rows"])  # type: ignore[arg-type]
    write_csv(REPRESENTATIVE_CSV, ["统计类型", "统计项目", "数量", "占代表产品比"], stats["representative_rows"])  # type: ignore[arg-type]
    write_csv(RULES_CSV, ["统计类型", "统计项目", "数量"], stats["rule_rows"])  # type: ignore[arg-type]
    write_statistics_excel(stats, findings_sections)
    FINDINGS_MD.write_text(build_findings_markdown(stats), encoding="utf-8")
    write_mentor_markdown()
    sync_main_excel(stats, findings_sections)
    validate(products, rules, classifications)
    write_log(products, rules, classifications, stats)

    overview = {row["统计项目"]: row["数量"] for row in stats["overview"]}  # type: ignore[index]
    rule_status: Counter = stats["rule_status_counts"]  # type: ignore[assignment]
    print(f"products={overview['当前产品池产品总数']}")
    print(f"mainline={overview['主线产品数量']}")
    print(f"supplemental={overview['补充观察产品数量']}")
    print(f"representatives={overview['Step7代表产品数量']}")
    print(f"rules={sum(rule_status.values())}")
    print(f"verified={rule_status['已核验']}")
    print(f"partial={rule_status['部分核验']}")
    print(f"pending={rule_status['待二次核验']}")


if __name__ == "__main__":
    main()
