from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path.home() / "Desktop" / "Fullgoal" / "课题研究" / "处理后数据" / "wind代码池" / "sheet7"
INPUT_FILE = BASE / "核心策略ETF_跟踪指数规则_补充整合终版.xlsx"
OUTPUT_FILE = BASE / "核心策略ETF_跟踪指数规则_官网补充核验版.xlsx"

RULE_FIELDS = [
    "发布机构",
    "指数基日",
    "指数发布日期",
    "指数类型",
    "样本空间",
    "成分股数量",
    "选样指标",
    "选样方法",
    "加权方式",
    "单只成分权重上限",
    "行业权重限制",
    "调样频率",
    "指数简介",
    "指数编制方案链接",
]

CORE_FIELDS = ["样本空间", "成分股数量", "选样指标", "选样方法", "加权方式", "调样频率"]

RULE_OUTPUT_COLS = [
    "Index_Code",
    "Index_Name",
    *RULE_FIELDS,
    "规则补充来源",
    "规则完整性状态",
    "仍缺失字段说明",
    "是否仍需官网核验",
]

ETF_RULE_COLS = [
    "ETF_Code",
    "ETF_Name",
    "基金简称",
    "基金全称",
    "基金管理人",
    "一级策略大类",
    "二级策略类别",
    "市场范围_二次修正",
    "最新基金规模(亿)",
    "上市日期",
    "基金成立日",
    "Index_Code",
    "Index_Name",
    *RULE_FIELDS,
    "规则补充来源",
    "规则完整性状态",
    "仍缺失字段说明",
    "是否仍需官网核验",
]


def norm_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return " ".join(str(value).replace("\u3000", " ").replace("\xa0", " ").strip().split())


def norm_code(value: Any) -> str:
    return norm_text(value).upper()


def is_missing(value: Any) -> bool:
    text = norm_text(value)
    if text == "":
        return True
    if text.lower() in {"nan", "none", "nat", "n/a"}:
        return True
    return text in {"-", "--", "Wind暂无", "待补充", "待官网补充", "暂无", "暂无数据"}


def is_effective(value: Any) -> bool:
    text = norm_text(value)
    if is_missing(text):
        return False
    return text != "待官网核验"


def output_missing(value: Any) -> str:
    return "待官网核验" if is_missing(value) else norm_text(value)


def safe_float(value: Any) -> float:
    try:
        if is_missing(value):
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def suggest_source(code: str, index_name: str, publisher: str) -> str:
    code_u = norm_code(code)
    name = norm_text(index_name)
    pub = norm_text(publisher)
    if code_u.endswith(".CSI") or "中证" in pub:
        return "中证指数有限公司官网"
    if code_u.endswith(".CNI") or "国证" in pub:
        return "国证指数有限公司官网"
    if code_u.endswith(".SH") or "上证" in name:
        return "中证指数有限公司官网或上证指数页面"
    if code_u.endswith(".SZ") or "深证" in name:
        return "国证指数有限公司官网或深证指数页面"
    if code_u.endswith(".HI") or "恒生" in pub:
        return "恒生指数有限公司官网"
    if code_u.endswith(".SPI") or "标普" in pub or "标普" in name:
        return "S&P Dow Jones Indices 官网"
    if code_u.endswith(".MI") or "MSCI" in pub.upper() or "MSCI" in name.upper():
        return "MSCI 官网"
    return "待确认"


def update_payloads() -> dict[str, dict[str, Any]]:
    return {
        "CSPSADRP.CI": {
            "Index_Name": "标普A股红利",
            "发布机构": "S&P Dow Jones Indices / 标普道琼斯指数",
            "样本空间": "S&P China A Domestic BMI，剔除 ST / *ST 等特殊处理股票",
            "成分股数量": "100",
            "选样指标": "高股息、盈利能力、股息增长、市值、流动性等",
            "选样方法": "从符合条件的中国 A 股股票中，按照股息收益率及相关质量、流动性条件筛选，选取高股息特征突出的股票构成指数",
            "加权方式": "股息率/股息相关加权，具体权重约束待官网核验",
            "单只成分权重上限": "待官网核验",
            "行业权重限制": "待官网核验",
            "调样频率": "待官网核验",
            "指数简介": "该指数旨在衡量中国 A 股市场中高股息股票的表现，强调分红能力、盈利能力和流动性条件",
            "指数编制方案链接": "S&P Dow Jones Indices 官网 / S&P China A Share Dividend Opportunities Index 页面及相关方法论",
            "规则补充来源": "GPT整理 + S&P Dow Jones Indices 官网线索",
            "规则完整性状态": "核心规则基本补全，仍建议官网核验",
            "是否仍需官网核验": "建议核验",
            "仍缺失字段说明": "单只成分权重上限、行业权重限制、调样频率、具体 A 股版本加权约束仍待官网核验",
        },
        "399667.SZ": {
            "Index_Name": "创业成长",
            "发布机构": "国证指数 / 深圳证券信息有限公司",
            "样本空间": "创业板市场",
            "成分股数量": "50",
            "选样指标": "业绩增长率、预期盈利增长率等成长指标",
            "选样方法": "计算综合成长得分，选取得分最高的 50 只证券作为样本",
            "加权方式": "综合得分倾斜因子 × 自由流通市值加权",
            "单只成分权重上限": "10%；前五大权重上限40%",
            "调样频率": "季度调整",
            "指数简介": "反映创业板市场中成长能力较强上市公司的整体表现",
            "指数编制方案链接": "国证指数有限公司官网，创业成长指数编制方案",
            "规则补充来源": "GPT整理 + 国证指数有限公司官网",
            "规则完整性状态": "核心规则已补全",
            "是否仍需官网核验": "否",
            "仍缺失字段说明": "无",
        },
        "399326.SZ": {
            "Index_Name": "成长40",
            "发布机构": "国证指数 / 深圳证券信息有限公司",
            "样本空间": "深交所 A 股",
            "成分股数量": "40",
            "选样指标": "成长因子、动量因子",
            "选样方法": "剔除低流动性、低市值、扣非净利润 TTM 为负等股票后，按成长因子筛选成长特征突出的股票",
            "加权方式": "成长/动量因子得分倾斜后的自由流通市值加权",
            "单只成分权重上限": "10%",
            "调样频率": "季度调整",
            "指数简介": "反映深市 A 股中成长特征较强股票的整体表现",
            "指数编制方案链接": "国证指数有限公司官网，成长40指数编制方案",
            "规则补充来源": "GPT整理 + 国证指数有限公司官网",
            "规则完整性状态": "核心规则已补全",
            "是否仍需官网核验": "否",
            "仍缺失字段说明": "无",
        },
        "399372.SZ": {
            "Index_Name": "大盘成长",
            "发布机构": "国证指数 / 深圳证券信息有限公司",
            "样本空间": "巨潮大盘指数样本空间，或国证大盘类样本空间，具体以国证官网编制方案为准",
            "成分股数量": "待官网核验",
            "选样指标": "成长类指标，通常包括主营业务收入增长、净利润增长、内部增长率等成长因子，具体以国证官网编制方案为准",
            "选样方法": "在大盘样本中计算成长得分，选取成长得分较高的股票构成指数，具体数量与规则待官网核验",
            "加权方式": "待官网核验",
            "单只成分权重上限": "待官网核验",
            "行业权重限制": "待官网核验",
            "调样频率": "待官网核验",
            "指数简介": "反映大盘股票中成长特征较强股票的表现",
            "指数编制方案链接": "国证指数有限公司官网，需人工核验 399372.SZ 编制方案",
            "规则补充来源": "GPT整理，仍需国证官网核验",
            "规则完整性状态": "部分补全，待官网核验",
            "是否仍需官网核验": "是",
            "仍缺失字段说明": "成分股数量、加权方式、单只成分权重上限、行业权重限制、调样频率需国证官网核验",
        },
        "399373.SZ": {
            "Index_Name": "大盘价值",
            "发布机构": "国证指数 / 深圳证券信息有限公司",
            "样本空间": "巨潮大盘指数",
            "成分股数量": "66",
            "选样指标": "E/P、CF/P、B/P、股息收益率",
            "选样方法": "在大盘样本中计算价值得分，选取价值得分较高的股票构成指数",
            "加权方式": "调整市值加权",
            "单只成分权重上限": "15%；前五大权重上限60%",
            "调样频率": "半年调整",
            "指数简介": "反映大盘股票中价值特征较强股票的整体表现",
            "指数编制方案链接": "国证指数有限公司官网，大盘价值指数编制方案",
            "规则补充来源": "GPT整理 + 国证指数有限公司官网",
            "规则完整性状态": "核心规则已补全",
            "是否仍需官网核验": "否",
            "仍缺失字段说明": "无",
        },
        "000029.SH": {
            "Index_Name": "180价值",
            "发布机构": "中证指数有限公司 / 上海证券交易所",
            "样本空间": "上证 180 指数样本",
            "成分股数量": "60",
            "选样指标": "股息收益率、账面市值比、现金流市值比、盈利市值比",
            "选样方法": "计算价值因子得分，选取价值得分最高的 60 只证券作为指数样本",
            "加权方式": "调整市值加权",
            "单只成分权重上限": "10%",
            "调样频率": "半年调整",
            "指数简介": "反映上证 180 样本中价值特征较强股票的表现",
            "指数编制方案链接": "中证指数有限公司官网，180价值指数编制方案",
            "规则补充来源": "GPT整理 + 中证指数有限公司官网",
            "规则完整性状态": "核心规则已补全",
            "是否仍需官网核验": "否",
            "仍缺失字段说明": "无",
        },
        "399348.SZ": {
            "Index_Name": "深证价值",
            "发布机构": "国证指数 / 深圳证券信息有限公司",
            "样本空间": "深证 300 指数样本",
            "成分股数量": "100",
            "选样指标": "E/P、CF/P、B/P、股息收益率",
            "选样方法": "计算价值因子得分，选取价值得分较高的股票构成指数",
            "加权方式": "派氏加权法",
            "单只成分权重上限": "待官网核验",
            "行业权重限制": "待官网核验",
            "调样频率": "半年调整",
            "指数简介": "反映深市 A 股中价值特征较强股票的表现",
            "指数编制方案链接": "国证指数有限公司官网，深证价值指数编制方案",
            "规则补充来源": "GPT整理 + 国证指数有限公司官网",
            "规则完整性状态": "核心规则基本补全，仍建议官网核验",
            "是否仍需官网核验": "建议核验",
            "仍缺失字段说明": "单只成分权重上限、行业权重限制需官网核验",
        },
        "H30356.CSI": {
            "Index_Name": "中证800价值",
            "发布机构": "中证指数有限公司",
            "样本空间": "中证 800 指数样本",
            "成分股数量": "250",
            "选样指标": "股息收益率、账面市值比、现金流市值比、盈利市值比",
            "选样方法": "计算价值因子得分，选取价值得分最高的 250 只证券作为指数样本",
            "加权方式": "调整市值加权",
            "单只成分权重上限": "待官网核验",
            "行业权重限制": "待官网核验",
            "调样频率": "半年调整",
            "指数简介": "反映中证 800 样本中价值特征较强股票的表现",
            "指数编制方案链接": "中证指数有限公司官网，中证800价值指数编制方案",
            "规则补充来源": "GPT整理 + 中证指数有限公司官网线索",
            "规则完整性状态": "核心规则基本补全，仍建议官网核验",
            "是否仍需官网核验": "建议核验",
            "仍缺失字段说明": "单只成分权重上限、行业权重限制需官网核验",
        },
        "H30352.CSI": {
            "Index_Name": "中证500价值",
            "发布机构": "中证指数有限公司",
            "样本空间": "中证 500 指数样本",
            "成分股数量": "150",
            "选样指标": "股息收益率、账面市值比、现金流市值比、盈利市值比",
            "选样方法": "计算价值因子得分，选取价值得分最高的 150 只证券作为指数样本",
            "加权方式": "调整市值加权",
            "单只成分权重上限": "待官网核验",
            "行业权重限制": "待官网核验",
            "调样频率": "半年调整",
            "指数简介": "反映中证 500 样本中价值特征较强股票的表现",
            "指数编制方案链接": "中证指数有限公司官网，中证500价值指数编制方案",
            "规则补充来源": "GPT整理 + 中证指数有限公司官网线索",
            "规则完整性状态": "核心规则基本补全，仍建议官网核验",
            "是否仍需官网核验": "建议核验",
            "仍缺失字段说明": "单只成分权重上限、行业权重限制需官网核验",
        },
        "000021.SH": {
            "Index_Name": "180治理",
            "发布机构": "中证指数有限公司 / 上海证券交易所",
            "样本空间": "上证 180 指数和上证公司治理指数样本交集",
            "成分股数量": "100",
            "选样指标": "公司治理水平、流动性、市值等综合排名",
            "选样方法": "从上证 180 和上证公司治理指数样本交集中选取综合排名前 100 只证券",
            "加权方式": "自由流通市值加权",
            "单只成分权重上限": "待官网核验",
            "行业权重限制": "待官网核验",
            "调样频率": "待官网核验",
            "指数简介": "反映上证 180 样本中公司治理水平较好的上市公司证券表现",
            "指数编制方案链接": "中证指数有限公司官网，180治理指数编制方案",
            "规则补充来源": "GPT整理 + 中证指数有限公司官网线索",
            "规则完整性状态": "核心规则基本补全，仍建议官网核验",
            "是否仍需官网核验": "建议核验",
            "仍缺失字段说明": "调样频率、单只成分权重上限、行业权重限制需官网核验",
        },
        "931088.CSI": {
            "Index_Name": "180ESG",
            "发布机构": "中证指数有限公司",
            "样本空间": "上证 180 指数样本",
            "成分股数量": "待官网核验",
            "选样指标": "中证 ESG 评价结果",
            "选样方法": "剔除中证 ESG 评价在各行业内排名后 20% 的证券",
            "加权方式": "ESG 综合评级得分调整后的自由流通市值加权",
            "行业权重限制": "行业权重与上证 180 保持一致",
            "单只成分权重上限": "10%；前五大权重上限40%",
            "调样频率": "半年调整",
            "指数简介": "在上证 180 样本基础上剔除 ESG 评价较低证券，反映 ESG 表现相对较好的上市公司证券表现",
            "指数编制方案链接": "中证指数有限公司官网，180ESG指数编制方案",
            "规则补充来源": "GPT整理 + 中证指数有限公司官网线索",
            "规则完整性状态": "核心规则基本补全，仍建议官网核验",
            "是否仍需官网核验": "建议核验",
            "仍缺失字段说明": "成分股数量需官网核验",
        },
        "399378.SZ": {
            "Index_Name": "ESG 300",
            "发布机构": "国证指数 / 深圳证券信息有限公司",
            "样本空间": "国证 1000 指数样本",
            "成分股数量": "300",
            "选样指标": "ESG 评分、ESG 风险评价",
            "选样方法": "剔除 ESG 风险较高证券后，按 ESG 评分筛选 300 只证券",
            "加权方式": "派氏加权法",
            "单只成分权重上限": "10%",
            "行业权重限制": "待官网核验",
            "调样频率": "半年调整",
            "指数简介": "反映国证 1000 样本中 ESG 表现较好的上市公司证券表现",
            "指数编制方案链接": "国证指数有限公司官网，ESG 300 指数编制方案",
            "规则补充来源": "GPT整理 + 国证指数有限公司官网",
            "规则完整性状态": "核心规则基本补全，仍建议官网核验",
            "是否仍需官网核验": "建议核验",
            "仍缺失字段说明": "行业权重限制需官网核验",
        },
        "931463.CSI": {
            "Index_Name": "300 ESG",
            "发布机构": "中证指数有限公司",
            "指数基日": "2017-06-30",
            "指数发布日期": "2020-04-30",
            "样本空间": "沪深 300 指数样本",
            "成分股数量": "待官网核验",
            "选样指标": "中证 ESG 分数 / ESG 评价",
            "选样方法": "剔除中证一级行业内 ESG 分数最低 20% 的上市公司证券",
            "加权方式": "待官网核验",
            "单只成分权重上限": "待官网核验",
            "行业权重限制": "待官网核验",
            "调样频率": "待官网核验",
            "指数简介": "在沪深 300 样本基础上剔除 ESG 评分较低证券，反映 ESG 表现较好的大盘上市公司证券表现",
            "指数编制方案链接": "中证指数有限公司官网待核验；辅助来源包括交易所基金招募说明书",
            "规则补充来源": "GPT整理，仍需中证官网核验",
            "规则完整性状态": "部分补全，待官网核验",
            "是否仍需官网核验": "是",
            "仍缺失字段说明": "成分股数量、加权方式、单只成分权重上限、行业权重限制、调样频率需中证官网核验",
        },
        "931243.CSI": {
            "Index_Name": "诚通央企ESG",
            "发布机构": "中证指数有限公司",
            "样本空间": "央企上市公司",
            "成分股数量": "50",
            "选样指标": "中证 ESG 评价、央企属性",
            "选样方法": "基于中证 ESG 评价，选取权重调整后 ESG 综合得分较高的 50 只央企上市公司证券",
            "加权方式": "待官网核验",
            "单只成分权重上限": "待官网核验",
            "行业权重限制": "待官网核验",
            "调样频率": "待官网核验",
            "指数简介": "反映央企上市公司中 ESG 综合表现较优证券的整体表现",
            "指数编制方案链接": "国资委官方新闻可确认策略逻辑和成分数量；中证指数官网编制方案仍需核验",
            "规则补充来源": "GPT整理，仍需中证官网核验；国资委新闻作辅助确认",
            "规则完整性状态": "部分补全，待官网核验",
            "是否仍需官网核验": "是",
            "仍缺失字段说明": "加权方式、单只成分权重上限、行业权重限制、调样频率需中证官网核验",
        },
        "721638.MI": {
            "Index_Name": "MSCI CHINA A RMB ESG UNIVERSAL",
            "发布机构": "MSCI",
            "样本空间": "MSCI China A RMB Index，覆盖沪深交易所上市的大中盘 A 股",
            "成分股数量": "待官网核验",
            "选样指标": "ESG 评分、ESG 趋势、自由流通市值",
            "选样方法": "在母指数基础上，根据 ESG 评分和 ESG 趋势调整成分权重，提高 ESG profile 和 ESG trend 暴露",
            "加权方式": "自由流通市值权重基础上的 ESG 再加权",
            "单只成分权重上限": "待官网核验",
            "行业权重限制": "待官网核验",
            "调样频率": "待官网核验",
            "指数简介": "在 MSCI China A RMB 母指数基础上，提高 ESG 评分和 ESG 趋势暴露，同时保持较广市场覆盖",
            "指数编制方案链接": "MSCI 官网，MSCI China A RMB ESG Universal Index 页面及 ESG Universal 方法论",
            "规则补充来源": "GPT整理 + MSCI 官网线索",
            "规则完整性状态": "核心规则基本补全，仍建议官网核验",
            "是否仍需官网核验": "建议核验",
            "仍缺失字段说明": "成分股数量、单只成分权重上限、行业权重限制、调样频率需 MSCI 官网核验",
        },
        "707717L.MI": {
            "Index_Name": "MSCI中国A股国际质量指数",
            "发布机构": "MSCI",
            "样本空间": "MSCI China A International Index",
            "成分股数量": "待官网核验",
            "选样指标": "ROE、盈利增长稳定性、低财务杠杆",
            "选样方法": "基于质量因子评分选取并加权成分股，强调高 ROE、盈利稳定、低杠杆",
            "加权方式": "质量得分 × 市值加权",
            "单只成分权重上限": "待官网核验",
            "行业权重限制": "待官网核验",
            "调样频率": "半年度调整",
            "指数简介": "反映 MSCI China A International 母指数中质量因子暴露较高证券的表现",
            "指数编制方案链接": "MSCI 官网，MSCI China A International Quality Index 页面及 MSCI Quality Indexes 方法论",
            "规则补充来源": "GPT整理 + MSCI 官网线索",
            "规则完整性状态": "核心规则基本补全，仍建议官网核验",
            "是否仍需官网核验": "建议核验",
            "仍缺失字段说明": "成分股数量、单只成分权重上限、行业权重限制需 MSCI 官网核验",
        },
    }


def apply_formatting(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)
        for header, col_idx in headers.items():
            if header in {"上市日期", "基金成立日", "指数基日", "指数发布日期", "运行时间"}:
                for col in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for cell in col:
                        cell.number_format = "yyyy-mm-dd"
            if header in {"最新基金规模(亿)", "涉及ETF规模合计", "受仍缺失字段影响的 ETF 规模合计"}:
                for col in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for cell in col:
                        cell.number_format = "0.0000"
    wb.save(path)


def main() -> None:
    sheet1 = pd.read_excel(INPUT_FILE, sheet_name="核心ETF_指数映射_标准版", dtype=object)
    sheet2 = pd.read_excel(INPUT_FILE, sheet_name="核心指数清单_去重版", dtype=object)
    rules = pd.read_excel(INPUT_FILE, sheet_name="指数规则明细_整合版", dtype=object)

    for df in [sheet1, sheet2, rules]:
        df.columns = [norm_text(c) for c in df.columns]

    sheet1["ETF_Code"] = sheet1["ETF_Code"].map(norm_code)
    sheet1["Index_Code"] = sheet1["Index_Code"].map(norm_code)
    sheet2["Index_Code"] = sheet2["Index_Code"].map(norm_code)
    rules["Index_Code"] = rules["Index_Code"].map(norm_code)

    for col in RULE_OUTPUT_COLS:
        if col not in rules.columns:
            rules[col] = "待官网核验"

    original_core_count = len(rules)
    before_complete = int((rules["规则完整性状态"].map(norm_text) == "核心规则已补全").sum())

    updates = update_payloads()
    updated_codes: list[str] = []
    rules = rules.copy()
    rules = rules[RULE_OUTPUT_COLS].copy()

    for idx, row in rules.iterrows():
        code = norm_code(row["Index_Code"])
        if code in updates:
            payload = updates[code]
            updated_codes.append(code)
            # Keep original index name if payload lacks it; update explicit official/GPT fields.
            for field, value in payload.items():
                if field in rules.columns:
                    rules.at[idx, field] = output_missing(value)

        # Global missing token harmonization for every row.
        for field in RULE_FIELDS + ["规则补充来源", "规则完整性状态", "仍缺失字段说明", "是否仍需官网核验"]:
            rules.at[idx, field] = output_missing(rules.at[idx, field])

    # For rows where missing statements still contain the old token, normalize wording.
    rules["仍缺失字段说明"] = rules["仍缺失字段说明"].map(
        lambda v: norm_text(v).replace("待官网补充", "待官网核验").replace("待补充", "待官网核验").replace("Wind暂无", "待官网核验")
    )

    # Ensure no blank output in the rule table.
    for col in RULE_OUTPUT_COLS:
        rules[col] = rules[col].map(lambda v: "待官网核验" if norm_text(v) == "" else v)

    core_codes = set(sheet2["Index_Code"])
    rules = rules[rules["Index_Code"].isin(core_codes)].copy()

    # Rebuild ETF-rule merged table; use sheet1 mapping as the master.
    merged = sheet1.merge(rules, on=["Index_Code", "Index_Name"], how="left")
    if len(merged) != len(sheet1) or merged["规则完整性状态"].isna().any():
        # Fall back to Index_Code-only merge if a name was touched.
        merged = sheet1.merge(
            rules.drop(columns=["Index_Name"]).copy(),
            on="Index_Code",
            how="left",
        )
    for col in ETF_RULE_COLS:
        if col not in merged.columns:
            merged[col] = "待官网核验"
    for col in merged.columns:
        if merged[col].dtype == "object":
            merged[col] = merged[col].map(lambda v: "待官网核验" if norm_text(v) == "" else output_missing(v))
    merged = merged[ETF_RULE_COLS].sort_values(
        ["一级策略大类", "二级策略类别", "最新基金规模(亿)"],
        ascending=[True, True, False],
        kind="stable",
    )

    # Remaining official verification list.
    def needs_verify(row: pd.Series) -> bool:
        any_wait = any("待官网核验" in norm_text(row.get(col)) for col in RULE_FIELDS + ["仍缺失字段说明"])
        return any_wait or norm_text(row.get("是否仍需官网核验")) in {"是", "建议核验"} or norm_text(row.get("规则完整性状态")) != "核心规则已补全"

    remaining = rules[rules.apply(needs_verify, axis=1)].copy()
    index_meta = sheet2.rename(
        columns={
            "对应ETF数量": "涉及ETF数量",
            "对应ETF规模合计": "涉及ETF规模合计",
        }
    )
    remaining_list = remaining.merge(
        index_meta[["Index_Code", "涉及ETF数量", "涉及ETF规模合计", "涉及一级策略大类", "涉及二级策略类别"]],
        on="Index_Code",
        how="left",
    )
    remaining_list["建议补充来源"] = remaining_list.apply(
        lambda row: suggest_source(row["Index_Code"], row["Index_Name"], row.get("发布机构", "")),
        axis=1,
    )
    remaining_list = remaining_list[
        [
            "Index_Code",
            "Index_Name",
            "发布机构",
            "涉及ETF数量",
            "涉及ETF规模合计",
            "涉及一级策略大类",
            "涉及二级策略类别",
            "规则完整性状态",
            "仍缺失字段说明",
            "建议补充来源",
        ]
    ].sort_values(["涉及一级策略大类", "涉及二级策略类别", "涉及ETF规模合计"], ascending=[True, True, False])

    complete_after = int((rules["规则完整性状态"].map(norm_text) == "核心规则已补全").sum())
    basic_after = int(rules["规则完整性状态"].map(norm_text).str.contains("基本补全", na=False).sum())
    partial_after = int(rules["规则完整性状态"].map(norm_text).str.contains("部分补全", na=False).sum())
    verify_after = int(rules["是否仍需官网核验"].map(norm_text).isin(["是", "建议核验"]).sum())
    affected_codes = set(remaining_list["Index_Code"])
    affected_etfs = merged[merged["Index_Code"].isin(affected_codes)].copy()
    affected_etf_count = affected_etfs["ETF_Code"].nunique()
    affected_scale = sum(safe_float(v) for v in affected_etfs["最新基金规模(亿)"])

    stats = pd.DataFrame(
        [
            ["原核心指数数量", original_core_count, "", ""],
            ["更新后核心指数数量", len(rules), "", ""],
            ["更新前核心规则已补全指数数量", before_complete, "", ""],
            ["更新后核心规则已补全指数数量", complete_after, "", ""],
            ["更新后核心规则基本补全但建议核验数量", basic_after, "", ""],
            ["更新后部分补全待官网核验数量", partial_after, "", ""],
            ["更新后仍需官网核验指数数量", verify_after, "", ""],
            ["受仍缺失字段影响的 ETF 数量", affected_etf_count, "", ""],
            ["受仍缺失字段影响的 ETF 规模合计", affected_scale, "", ""],
            ["", "", "", ""],
            ["按一级策略大类统计剩余待核验指数数量、涉及 ETF 数量、涉及规模", "", "", ""],
        ],
        columns=["统计项", "数值", "涉及ETF数量", "涉及ETF规模合计"],
    )
    by_strategy = (
        affected_etfs.groupby("一级策略大类", as_index=False)
        .agg(
            数值=("Index_Code", "nunique"),
            涉及ETF数量=("ETF_Code", "nunique"),
            涉及ETF规模合计=("最新基金规模(亿)", lambda s: sum(safe_float(v) for v in s)),
        )
        .rename(columns={"一级策略大类": "统计项"})
        .sort_values(["数值", "涉及ETF规模合计"], ascending=[False, False])
    )
    comparison = pd.concat([stats, by_strategy[["统计项", "数值", "涉及ETF数量", "涉及ETF规模合计"]]], ignore_index=True)

    standard_map = sheet1[["ETF_Code", "Index_Code"]].rename(columns={"Index_Code": "Index_Code_标准"})
    check_merge = merged[["ETF_Code", "Index_Code"]].merge(standard_map, on="ETF_Code", how="left")
    mismatch_count = int((check_merge["Index_Code"].map(norm_code) != check_merge["Index_Code_标准"].map(norm_code)).sum())

    noncore_added = bool(set(rules["Index_Code"]) - core_codes)
    core_lost = bool(core_codes - set(rules["Index_Code"]))
    updated_all = set(updates.keys()).issubset(set(rules["Index_Code"]))
    # Check blanks only in the newly created rule and ETF-rule sheets, because Sheet 1/2 are kept as original artifacts.
    key_outputs = pd.concat(
        [rules.astype(object), merged.astype(object)],
        ignore_index=True,
        sort=False,
    )
    has_blank = bool(key_outputs.apply(lambda col: col.map(lambda v: norm_text(v) == "" if pd.notna(v) else True)).any().any())
    has_wind_missing = bool(key_outputs.apply(lambda col: col.map(lambda v: "Wind暂无" in norm_text(v))).any().any())
    has_old_pending = bool(key_outputs.apply(lambda col: col.map(lambda v: norm_text(v) in {"待补充", "待官网补充"})).any().any())
    unified_missing = "是" if (not has_wind_missing and not has_old_pending) else "否"

    quality = pd.DataFrame(
        [
            ["Sheet 1 ETF 行数是否为 168", "是" if len(sheet1) == 168 else "否", len(sheet1)],
            ["Sheet 1 ETF_Code 是否重复", "否" if not sheet1["ETF_Code"].duplicated().any() else "是", int(sheet1["ETF_Code"].duplicated().sum())],
            ["Sheet 3 指数数量是否为 82", "是" if len(rules) == 82 else "否", len(rules)],
            ["是否新增非核心指数", "是" if noncore_added else "否", ""],
            ["是否丢失核心指数", "是" if core_lost else "否", ""],
            ["Sheet 4 行数是否为 168", "是" if len(merged) == 168 else "否", len(merged)],
            ["Sheet 4 ETF-Index 错配数量", mismatch_count, ""],
            ["16 个待处理指数是否全部完成字段更新", "是" if updated_all and len(set(updated_codes)) == 16 else "否", len(set(updated_codes))],
            ["是否仍存在空值", "是" if has_blank else "否", "检查范围：Sheet 3 与 Sheet 4"],
            ["是否仍存在“Wind暂无”", "是" if has_wind_missing else "否", ""],
            ["是否仍存在“待补充”", "是" if has_old_pending else "否", ""],
            ["是否已统一改为“待官网核验”", unified_missing, ""],
            ["输出文件路径", str(OUTPUT_FILE), ""],
            ["运行时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""],
        ],
        columns=["检查项", "结果", "说明"],
    )

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        sheet1.to_excel(writer, sheet_name="核心ETF_指数映射_标准版", index=False)
        sheet2.to_excel(writer, sheet_name="核心指数清单_去重版", index=False)
        rules.to_excel(writer, sheet_name="指数规则明细_官网补充版", index=False)
        merged.to_excel(writer, sheet_name="ETF_指数规则合并表_官网补充版", index=False)
        remaining_list.to_excel(writer, sheet_name="剩余待官网核验清单", index=False)
        comparison.to_excel(writer, sheet_name="官网补充前后对比", index=False)
        quality.to_excel(writer, sheet_name="数据质量检查", index=False)

    apply_formatting(OUTPUT_FILE)

    print(f"核心 ETF 数量: {len(sheet1)}")
    print(f"核心指数数量: {len(rules)}")
    print(f"本次更新的指数数量: {len(set(updated_codes))}")
    print(f"更新后核心规则已补全指数数量: {complete_after}")
    print(f"更新后仍需官网核验指数数量: {verify_after}")
    print(f"ETF-指数错配数量: {mismatch_count}")
    print(f"输出文件路径: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
