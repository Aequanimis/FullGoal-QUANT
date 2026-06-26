from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path.home() / "Desktop" / "Fullgoal" / "课题研究" / "处理后数据" / "wind代码池" / "sheet7"
INPUT_FILE = BASE / "核心策略ETF_跟踪指数规则_官网补充核验版.xlsx"
OUTPUT_FILE = BASE / "核心策略ETF_跟踪指数规则_官网补充核验最终版.xlsx"

SHEET1 = "核心ETF_指数映射_标准版"
SHEET3 = "指数规则明细_官网补充版"
SHEET4 = "ETF_指数规则合并表_官网补充版"


def normalize(value):
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").replace("\xa0", " ").strip()


def fill_blank_cells(ws):
    filled = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if normalize(cell.value) == "":
                cell.value = "待官网核验"
                filled += 1
    return filled


def header_index(ws, header):
    headers = [cell.value for cell in ws[1]]
    return headers.index(header) + 1


def apply_basic_formatting(wb):
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)


def main():
    wb = load_workbook(INPUT_FILE)
    for required in [SHEET1, SHEET3, SHEET4]:
        if required not in wb.sheetnames:
            raise KeyError(f"缺少必要 sheet：{required}")

    ws1 = wb[SHEET1]
    ws3 = wb[SHEET3]
    ws4 = wb[SHEET4]

    filled3 = fill_blank_cells(ws3)
    filled4 = fill_blank_cells(ws4)

    apply_basic_formatting(wb)
    wb.save(OUTPUT_FILE)

    # Re-open output for verification.
    check_wb = load_workbook(OUTPUT_FILE, read_only=True, data_only=True)
    c_ws1 = check_wb[SHEET1]
    c_ws3 = check_wb[SHEET3]
    c_ws4 = check_wb[SHEET4]

    s1_etf_col = header_index(c_ws1, "ETF_Code")
    s1_idx_col = header_index(c_ws1, "Index_Code")
    s3_idx_col = header_index(c_ws3, "Index_Code")
    s4_etf_col = header_index(c_ws4, "ETF_Code")
    s4_idx_col = header_index(c_ws4, "Index_Code")

    sheet1_rows = list(c_ws1.iter_rows(min_row=2, values_only=True))
    sheet3_rows = list(c_ws3.iter_rows(min_row=2, values_only=True))
    sheet4_rows = list(c_ws4.iter_rows(min_row=2, values_only=True))

    etf_count = len(sheet1_rows)
    unique_etf_count = len({row[s1_etf_col - 1] for row in sheet1_rows})
    index_count = len({row[s3_idx_col - 1] for row in sheet3_rows})

    standard_map = {row[s1_etf_col - 1]: row[s1_idx_col - 1] for row in sheet1_rows}
    mismatch_count = 0
    for row in sheet4_rows:
        etf_code = row[s4_etf_col - 1]
        index_code = row[s4_idx_col - 1]
        if standard_map.get(etf_code) != index_code:
            mismatch_count += 1

    blank_sheet3 = sum(
        1
        for row in c_ws3.iter_rows(min_row=2, values_only=True)
        for value in row
        if normalize(value) == ""
    )
    blank_sheet4 = sum(
        1
        for row in c_ws4.iter_rows(min_row=2, values_only=True)
        for value in row
        if normalize(value) == ""
    )

    print(f"Sheet3 填补空值数量: {filled3}")
    print(f"Sheet4 填补空值数量: {filled4}")
    print(f"ETF 数量: {etf_count}")
    print(f"ETF_Code 唯一数量: {unique_etf_count}")
    print(f"核心指数数量: {index_count}")
    print(f"ETF-指数错配数量: {mismatch_count}")
    print(f"Sheet3 剩余空值数量: {blank_sheet3}")
    print(f"Sheet4 剩余空值数量: {blank_sheet4}")
    print(f"输出文件路径: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
