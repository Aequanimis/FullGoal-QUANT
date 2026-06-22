"""Step 7：按给定 DATA_BLOCK 选择各策略代表产品。"""

from __future__ import annotations

import csv
import io
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
ATTACHMENT = Path(
    r"C:\Users\lvdon\.codex\attachments\8bb1cd8d-1f51-4940-a6d5-6a3150b0a24e\pasted-text.txt"
)
INPUT_EXCEL = PROJECT_ROOT / "01_processed_data" / "classification" / "境内策略ETF产品池_分类修正版.xlsx"
MAIN_EXCEL = PROJECT_ROOT / "02_outputs" / "excel" / "境内策略ETF产品梳理_初版.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "01_processed_data" / "classification"
OUTPUT_CSV = OUTPUT_DIR / "境内策略ETF代表产品筛选表.csv"
OUTPUT_EXCEL = OUTPUT_DIR / "境内策略ETF产品池_代表产品筛选版.xlsx"
REVIEW_MD = OUTPUT_DIR / "step7_representative_products.md"
LOG_MD = OUTPUT_DIR / "step7_representative_products_log.md"

STEP6_FIELDS = [
    "策略大类", "策略细分", "基金代码", "产品名称", "基金公司", "上市交易所", "跟踪指数", "指数公司",
    "纳入级别", "是否主线", "是否补充观察", "是否待校验", "待校验事项", "是否代表产品初筛",
    "代表产品初筛理由", "信息来源备注", "清洗备注", "原策略大类", "原策略细分", "原纳入级别",
    "原指数公司", "建议策略大类", "建议策略细分", "建议纳入级别", "建议指数公司", "是否需要修改",
    "修改类型", "修改原因", "后续是否需要查指数规则", "分类修正备注",
]

STEP7_FIELDS = [
    "是否Step7代表产品",
    "Step7代表产品级别",
    "Step7选择理由",
    "Step7后续需重点看的指数规则",
    "Step7备注",
]
OUTPUT_FIELDS = [*STEP6_FIELDS, *STEP7_FIELDS]

DATA_FIELDS = [
    "策略类型",
    "基金代码",
    "代表产品",
    "基金公司",
    "跟踪指数",
    "Step7代表产品级别",
    "Step7选择理由",
    "Step7后续需重点看的指数规则",
    "Step7备注",
]


def clean(value: object) -> str:
    return "" if value is None else str(value).strip(" \t\r\n")


def product_key(record: dict[str, str]) -> tuple[str, str]:
    return record["基金代码"], record["产品名称"]


def data_key(record: dict[str, str]) -> tuple[str, str]:
    return record["基金代码"], record["代表产品"]


def read_step6_pool() -> list[dict[str, str]]:
    if not INPUT_EXCEL.exists():
        raise FileNotFoundError(f"缺少输入文件：{INPUT_EXCEL}")
    workbook = load_workbook(INPUT_EXCEL, data_only=False)
    if "产品池_分类修正版" not in workbook.sheetnames:
        raise ValueError("输入 Excel 中不存在 Sheet【产品池_分类修正版】")
    worksheet = workbook["产品池_分类修正版"]
    headers = [clean(cell.value) for cell in worksheet[1]]
    if headers != STEP6_FIELDS:
        raise ValueError(f"Step 6 字段不匹配：{headers}")

    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        record = {field: clean(value) for field, value in zip(STEP6_FIELDS, values)}
        if any(record.values()):
            rows.append(record)
    return rows


def read_data_block() -> list[dict[str, str]]:
    text = ATTACHMENT.read_text(encoding="utf-8-sig")
    header = "\t".join(DATA_FIELDS)
    start = text.find(header)
    if start < 0:
        raise ValueError("附件中未找到 Step 7 TSV 表头")
    end = text.find("\n\n五、代表产品筛选版 Excel 要求", start)
    if end < 0:
        raise ValueError("附件中未找到 Step 7 TSV 结束位置")
    block = text[start:end].strip("\r\n")
    reader = csv.DictReader(io.StringIO(block), delimiter="\t")
    if reader.fieldnames != DATA_FIELDS:
        raise ValueError(f"Step 7 TSV 字段不匹配：{reader.fieldnames}")
    return [
        {field: clean(record.get(field, "")) for field in DATA_FIELDS}
        for record in reader
    ]


def select_representatives(
    products: list[dict[str, str]], selections: list[dict[str, str]]
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[tuple[str, str]], list[tuple[str, str]]]:
    pool_counts = Counter(product_key(record) for record in products)
    selection_by_key = {data_key(record): record for record in selections}
    selection_keys = set(selection_by_key)
    unmatched = [key for key in selection_keys if pool_counts[key] == 0]
    ambiguous = [key for key in selection_keys if pool_counts[key] > 1]
    usable = selection_keys - set(unmatched) - set(ambiguous)

    all_rows: list[dict[str, str]] = []
    representative_by_key: dict[tuple[str, str], dict[str, str]] = {}
    for product in products:
        key = product_key(product)
        selection = selection_by_key.get(key) if key in usable else None
        if selection:
            step7 = {
                "是否Step7代表产品": "是",
                "Step7代表产品级别": selection["Step7代表产品级别"],
                "Step7选择理由": selection["Step7选择理由"],
                "Step7后续需重点看的指数规则": selection["Step7后续需重点看的指数规则"],
                "Step7备注": selection["Step7备注"],
            }
        else:
            step7 = {field: "否" if field == "是否Step7代表产品" else "" for field in STEP7_FIELDS}
        result = {**product, **step7}
        all_rows.append(result)
        if selection:
            representative_by_key[key] = result

    representatives = [
        representative_by_key[data_key(selection)]
        for selection in selections
        if data_key(selection) in representative_by_key
    ]
    return all_rows, representatives, unmatched, ambiguous


def write_csv(rows: list[dict[str, str]]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=OUTPUT_FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def display_width(value: object) -> int:
    text = clean(value)
    return sum(
        2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1
        for char in text
    )


def append_rows(worksheet, fields: list[str], rows: list[dict[str, object]]) -> None:
    worksheet.append(fields)
    for record in rows:
        worksheet.append([record.get(field, "") for field in fields])


def format_sheet(worksheet) -> None:
    max_row = max(worksheet.max_row, 1)
    max_column = max(worksheet.max_column, 1)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_column)}{max_row}"
    worksheet.sheet_view.zoomScale = 85

    fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    header_font = Font(name="Microsoft YaHei", size=10, bold=True)
    body_font = Font(name="Microsoft YaHei", size=10)
    border = Border(bottom=Side(style="thin", color="A6A6A6"))
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    worksheet.row_dimensions[1].height = 28

    headers = [cell.value for cell in worksheet[1]]
    for row in worksheet.iter_rows(min_row=2, max_row=max_row, max_col=max_column):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    long_headers = {
        "跟踪指数", "待校验事项", "信息来源备注", "清洗备注", "修改原因", "分类修正备注",
        "代表产品初筛理由", "Step7选择理由", "Step7后续需重点看的指数规则", "Step7备注",
    }
    for column_index, header in enumerate(headers, start=1):
        values = [worksheet.cell(row, column_index).value for row in range(1, max_row + 1)]
        measured = max(display_width(value) for value in values) + 3
        cap = 48 if header in long_headers else 28
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(measured, 12), cap)
        if header == "基金代码":
            for row in range(2, max_row + 1):
                worksheet.cell(row, column_index).number_format = "@"


def build_statistics(
    products: list[dict[str, str]], representatives: list[dict[str, str]]
) -> list[dict[str, object]]:
    strategy_counts = Counter(row["建议策略大类"] for row in representatives)
    representative_level_counts = Counter(row["Step7代表产品级别"] for row in representatives)
    inclusion_counts = Counter(row["建议纳入级别"] for row in representatives)
    rows: list[dict[str, object]] = []

    for strategy, count in sorted(strategy_counts.items()):
        rows.append({"统计类型": "按策略类型", "策略类型": strategy, "代表产品级别": "", "产品数量": count})
    for level, count in sorted(representative_level_counts.items()):
        rows.append({"统计类型": "按代表产品级别", "策略类型": "", "代表产品级别": level, "产品数量": count})

    summaries = [
        ("主线代表产品", inclusion_counts["主线"]),
        ("补充观察代表产品", inclusion_counts["补充观察"]),
        ("单列观察代表产品", inclusion_counts["单列观察"]),
        ("跨境补充代表产品", inclusion_counts["跨境补充"]),
        ("非代表产品", len(products) - len(representatives)),
    ]
    for label, count in summaries:
        rows.append({"统计类型": "汇总指标", "策略类型": label, "代表产品级别": "", "产品数量": count})
    return rows


def write_excel(products: list[dict[str, str]], representatives: list[dict[str, str]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    mainline = [row for row in representatives if row["建议纳入级别"] == "主线"]
    supplemental = [
        row
        for row in representatives
        if row["建议纳入级别"] in {"补充观察", "跨境补充", "单列观察", "仅作线索"}
    ]
    sheets = [
        ("产品池_代表产品筛选版", products),
        ("Step7代表产品清单", representatives),
        ("主线代表产品", mainline),
        ("补充观察代表产品", supplemental),
    ]
    for name, rows in sheets:
        worksheet = workbook.create_sheet(name)
        append_rows(worksheet, OUTPUT_FIELDS, rows)
        format_sheet(worksheet)

    stats_sheet = workbook.create_sheet("代表产品统计")
    stats_fields = ["统计类型", "策略类型", "代表产品级别", "产品数量"]
    append_rows(stats_sheet, stats_fields, build_statistics(products, representatives))
    format_sheet(stats_sheet)
    workbook.save(OUTPUT_EXCEL)


def sync_main_excel(
    products: list[dict[str, str]], representatives: list[dict[str, str]]
) -> tuple[list[tuple[str, str]], list[tuple[str, str]]]:
    workbook = load_workbook(MAIN_EXCEL)
    if "产品总表" not in workbook.sheetnames:
        raise ValueError("主 Excel 中不存在 Sheet【产品总表】")
    worksheet = workbook["产品总表"]
    headers = [clean(cell.value) for cell in worksheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}
    required = {"基金代码", "产品名称", "是否代表产品", "代表性理由", "备注"}
    missing_headers = required - set(headers)
    if missing_headers:
        raise ValueError(f"主 Excel 缺少字段：{sorted(missing_headers)}")

    main_positions: dict[tuple[str, str], list[int]] = defaultdict(list)
    for row in range(2, worksheet.max_row + 1):
        key = (
            clean(worksheet.cell(row, columns["基金代码"]).value),
            clean(worksheet.cell(row, columns["产品名称"]).value),
        )
        main_positions[key].append(row)

    representative_keys = {product_key(row) for row in representatives}
    for product in products:
        key = product_key(product)
        positions = main_positions.get(key, [])
        if len(positions) == 1 and key not in representative_keys:
            worksheet.cell(positions[0], columns["是否代表产品"], "否")

    missing: list[tuple[str, str]] = []
    ambiguous: list[tuple[str, str]] = []
    for record in representatives:
        key = product_key(record)
        positions = main_positions.get(key, [])
        if not positions:
            missing.append(key)
            continue
        if len(positions) > 1:
            ambiguous.append(key)
            continue
        row = positions[0]
        worksheet.cell(row, columns["是否代表产品"], "是")
        worksheet.cell(row, columns["代表性理由"], record["Step7选择理由"])

        note_text = (
            f"Step7代表产品级别：{record['Step7代表产品级别']}；"
            f"后续需看指数规则：{record['Step7后续需重点看的指数规则']}"
        )
        note_cell = worksheet.cell(row, columns["备注"])
        existing = clean(note_cell.value)
        if note_text not in existing:
            note_cell.value = existing + ("\n" if existing else "") + note_text
        note_cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{worksheet.max_row}"
    workbook.save(MAIN_EXCEL)
    return missing, ambiguous


def format_key_list(keys: Iterable[tuple[str, str]]) -> str:
    values = list(keys)
    return "无" if not values else "\n".join(f"- {code}｜{name}" for code, name in values)


def write_review(representatives: list[dict[str, str]]) -> None:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in representatives:
        grouped[row["建议策略大类"]].append(row)

    category_sections = []
    for strategy, rows in sorted(grouped.items()):
        lines = [f"### {strategy}", ""]
        lines.extend(
            f"- {row['基金代码']} {row['产品名称']}（{row['Step7代表产品级别']}）：{row['Step7选择理由']}"
            for row in rows
        )
        category_sections.append("\n".join(lines))

    mainline = [row for row in representatives if row["建议纳入级别"] == "主线"]
    supplemental = [row for row in representatives if row["建议纳入级别"] != "主线"]
    mainline_lines = "\n".join(f"- {row['基金代码']} {row['产品名称']}" for row in mainline)
    supplemental_lines = "\n".join(
        f"- {row['基金代码']} {row['产品名称']}（{row['建议纳入级别']}）" for row in supplemental
    )
    rules_lines = "\n".join(
        f"- {row['建议策略大类']}｜{row['产品名称']}：{row['Step7后续需重点看的指数规则']}"
        for row in representatives
    )

    content = f"""# Step 7 代表产品说明

## 1. Step 7 目标

从 Step 6 分类修正版产品池中，为每类策略选择少量代表产品，作为后续指数规则拆解和 mentor 讨论的重点样本。

## 2. 代表产品选择标准

- 策略逻辑具有典型性；
- 跟踪指数具有代表性；
- 基金公司及产品具有一定代表性；
- 产品具有一定市场认知度；
- 有助于同类策略、不同指数体系及境内外策略 ETF 的后续比较。

本轮 Step 7 不是对所有 ETF 做深度研究，而是从 Step 6 的分类修正版产品池中，为每类策略选择 1–3 个代表产品，用于后续指数规则拆解和 mentor 讨论。选择标准包括：策略逻辑是否典型、指数是否有代表性、基金公司是否有代表性、产品是否具有一定市场认知度、是否有助于后续和海外策略 ETF 对比。由于当前尚未接入 Wind / Choice，本轮结果为代表产品初筛，后续仍需补充规模、成交额、费率、成立日期和上市日期等数据。

## 3. 每类代表产品清单

{chr(10).join(category_sections)}

## 4. 主线代表产品

{mainline_lines}

## 5. 补充观察代表产品

{supplemental_lines}

## 6. 后续需重点查指数规则的方向

{rules_lines}

## 7. 数据校验说明

本轮为公开信息和策略代表性初筛，后续需用 Wind / Choice 校验规模、成交额、费率、上市日期。
"""
    REVIEW_MD.write_text(content, encoding="utf-8")


def write_log(
    products: list[dict[str, str]],
    selections: list[dict[str, str]],
    representatives: list[dict[str, str]],
    unmatched: list[tuple[str, str]],
    ambiguous: list[tuple[str, str]],
    main_missing: list[tuple[str, str]],
    main_ambiguous: list[tuple[str, str]],
) -> None:
    strategy_counts = Counter(row["建议策略大类"] for row in representatives)
    inclusion_counts = Counter(row["建议纳入级别"] for row in representatives)
    strategy_lines = "\n".join(
        f"- {strategy}：{count}" for strategy, count in sorted(strategy_counts.items())
    )
    content = f"""# Step 7 代表产品筛选日志

1. 输入文件路径：`{INPUT_EXCEL}`
2. 输出文件路径：
   - 代表产品筛选表 CSV：`{OUTPUT_CSV}`
   - 代表产品筛选版 Excel：`{OUTPUT_EXCEL}`
   - Step 7 说明：`{REVIEW_MD}`
   - 同步主 Excel：`{MAIN_EXCEL}`
3. 原始产品数量：{len(products)}
4. DATA_BLOCK 代表产品数量：{len(selections)}
5. 成功匹配代表产品数量：{len(representatives)}
6. 未匹配到的 DATA_BLOCK 产品：
{format_key_list(unmatched)}
   - 无法唯一匹配的 DATA_BLOCK 产品：
{format_key_list(ambiguous)}
7. 是否成功生成代表产品筛选版 Excel：是
8. 是否成功同步更新主 Excel：是
   - 主 Excel 中未找到的代表产品：
{format_key_list(main_missing)}
   - 主 Excel 中无法唯一匹配的代表产品：
{format_key_list(main_ambiguous)}
9. 按策略类型统计代表产品数量：
{strategy_lines}
10. 主线代表产品数量：{inclusion_counts['主线']}
11. 补充观察代表产品数量：{inclusion_counts['补充观察']}
12. 单列观察代表产品数量：{inclusion_counts['单列观察']}
13. 跨境补充代表产品数量：{inclusion_counts['跨境补充']}
"""
    LOG_MD.write_text(content, encoding="utf-8")


def validate(
    products: list[dict[str, str]], representatives: list[dict[str, str]]
) -> None:
    if OUTPUT_CSV.read_bytes()[:3] != b"\xef\xbb\xbf":
        raise ValueError("代表产品 CSV 不是 UTF-8 with BOM")
    with OUTPUT_CSV.open("r", encoding="utf-8-sig", newline="") as csv_file:
        loaded = list(csv.DictReader(csv_file))
    if loaded != products:
        raise ValueError("代表产品 CSV 回读不一致")

    workbook = load_workbook(OUTPUT_EXCEL)
    expected_sheets = [
        "产品池_代表产品筛选版", "Step7代表产品清单", "主线代表产品", "补充观察代表产品", "代表产品统计"
    ]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(f"代表产品 Excel Sheet 不匹配：{workbook.sheetnames}")
    expected_rows = {
        "产品池_代表产品筛选版": len(products),
        "Step7代表产品清单": len(representatives),
        "主线代表产品": sum(row["建议纳入级别"] == "主线" for row in representatives),
        "补充观察代表产品": sum(row["建议纳入级别"] != "主线" for row in representatives),
    }
    for sheet_name, count in expected_rows.items():
        if workbook[sheet_name].max_row != count + 1:
            raise ValueError(f"{sheet_name} 行数不匹配")
    for worksheet in workbook.worksheets:
        if worksheet.freeze_panes != "A2" or not worksheet.auto_filter.ref:
            raise ValueError(f"{worksheet.title} 的冻结窗格或筛选未生效")

    main_workbook = load_workbook(MAIN_EXCEL)
    worksheet = main_workbook["产品总表"]
    headers = [clean(cell.value) for cell in worksheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}
    main_map = {
        (
            clean(worksheet.cell(row, columns["基金代码"]).value),
            clean(worksheet.cell(row, columns["产品名称"]).value),
        ): row
        for row in range(2, worksheet.max_row + 1)
    }
    for record in representatives:
        key = product_key(record)
        if key in main_map and clean(worksheet.cell(main_map[key], columns["是否代表产品"]).value) != "是":
            raise ValueError(f"主 Excel 代表产品同步失败：{key}")


def main() -> None:
    products = read_step6_pool()
    selections = read_data_block()
    all_rows, representatives, unmatched, ambiguous = select_representatives(products, selections)
    write_csv(all_rows)
    write_excel(all_rows, representatives)
    write_review(representatives)
    main_missing, main_ambiguous = sync_main_excel(all_rows, representatives)
    validate(all_rows, representatives)
    write_log(
        all_rows, selections, representatives, unmatched, ambiguous, main_missing, main_ambiguous
    )

    print(f"products={len(all_rows)}")
    print(f"selections={len(selections)}")
    print(f"matched={len(representatives)}")
    print(f"unmatched={len(unmatched)}")
    print(f"ambiguous={len(ambiguous)}")
    print(f"main_missing={len(main_missing)}")
    print("strategy_counts=" + repr(dict(sorted(Counter(r['建议策略大类'] for r in representatives).items()))))


if __name__ == "__main__":
    main()
