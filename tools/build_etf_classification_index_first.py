from __future__ import annotations

import math
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_PATH = Path(r"C:\Users\lvdon\Desktop\Fullgoal\ICI课题\P17-P23_中国ETF数据处理_二次修正版\00_数据字段检查与描述性统计_二次修正版.xlsx")
INPUT_SHEET = "ETF分析池_上市交易_二次修正"
OUTPUT_DIR = Path(r"C:\Users\lvdon\Desktop\Fullgoal\ICI课题\ETF分类主表_指数优先版")
OUTPUT_PATH = OUTPUT_DIR / "ETF分类主表_指数优先版.xlsx"

EXPECTED_COUNT = 1574
EXPECTED_SCALE = 46949.9638


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def contains_any(s: str, keywords: list[str]) -> bool:
    return any(k and k.lower() in s.lower() for k in keywords)


def match_keywords(s: str, keywords: list[str]) -> list[str]:
    out: list[str] = []
    lower = s.lower()
    for k in keywords:
        if k and k.lower() in lower and k not in out:
            out.append(k)
    return out


def add_tag(tags: list[str], tag: str) -> None:
    if tag and tag not in tags:
        tags.append(tag)


def fmt_date(value) -> str:
    if pd.isna(value):
        return ""
    try:
        return pd.to_datetime(value).strftime("%Y-%m-%d")
    except Exception:
        return text(value)


def normalize_index(index_name: str) -> str:
    s = text(index_name)
    if not s:
        return ""
    replacements = {
        "500SNLV": "中证500低波",
        "证券公司": "中证全指证券公司",
        "电力指数": "中证全指电力公用事业",
        "科创50": "上证科创板50",
        "科创价格": "上证科创板综合价格",
        "创业200": "创业板200",
        "创业300": "创业板300",
        "创业板指": "创业板指",
    }
    if s in replacements:
        return replacements[s]
    if re.fullmatch(r"1000", s):
        return "中证1000"
    if re.fullmatch(r"A500", s, flags=re.I):
        return "中证A500"
    if re.fullmatch(r"300", s):
        return "沪深300"
    if re.fullmatch(r"500", s):
        return "中证500"
    if s.startswith("300质量"):
        return "沪深300质量"
    if s.startswith("300红利低波"):
        return "沪深300红利低波"
    if s.startswith("500质量成长"):
        return "中证500质量成长"
    return s


MONEY_KEYWORDS = ["货币", "现金", "保证金", "快线", "添益", "场内货币", "收益宝", "财富宝", "银华日利", "华宝添益", "交易货币"]
COMMODITY_KEYWORDS = ["黄金ETF", "黄金基金", "豆粕ETF", "商品ETF", "能源化工", "原油ETF", "白银ETF"]
BOND_KEYWORDS = ["国债", "政金债", "信用债", "公司债", "地方债", "城投债", "可转债", "短融", "债券", "债"]
CROSS_KEYWORDS = ["恒生", "港股", "港股通", "H股", "纳斯达克", "标普", "道琼斯", "日经", "德国", "法国", "海外", "中概", "东南亚", "QDII", "MSCI美国", "美国", "香港", "日韩", "越南", "沙特", "亚太", "全球"]

ENHANCE_KEYWORDS = ["增强策略", "指数增强", "增强"]
STRATEGY_KEYWORDS = [
    "红利", "低波", "低波动", "红利低波", "质量", "价值", "成长", "质量成长", "现金流", "自由现金流",
    "股息", "高股息", "基本面", "等权", "等权重", "ESG", "动量", "回购", "分红", "央企红利",
    "国企红利", "红利质量", "红利价值", "价值100", "成长100", "低估值", "高分红", "Smart Beta",
    "价值稳健", "成长创新", "质量低波", "红利低波动", "股东回报", "龙头红利", "高质量", "盈利",
    "盈利质量", "央企股东回报", "国企股东回报", "央企分红", "国企分红", "央企价值", "国企价值",
]
THEME_KEYWORDS = [
    "新能源", "光伏", "芯片", "半导体", "人工智能", "机器人", "军工", "医药", "消费", "金融",
    "证券公司", "全指证券公司", "非银", "银行", "地产", "房地产", "传媒", "计算机", "通信", "汽车",
    "电池", "储能", "双碳", "绿色", "电力", "煤炭", "有色", "钢铁", "基建", "农业", "酒", "食品",
    "云计算", "软件", "游戏", "数字经济", "高端制造", "创新药", "医疗器械", "工业母机", "机床",
    "工程机械", "畜牧养殖", "稀有金属", "稀土", "稀土产业", "工业有色", "化工", "材料", "环保",
    "教育", "物流", "旅游", "黄金股", "矿业", "交运", "信息技术", "科技", "互联网", "生物", "医疗",
    "养老", "家电", "机械", "电子", "保险", "石油天然气", "油气产业", "卫星产业", "航天", "航空",
    "通用航空", "低空经济", "碳中和", "低碳经济", "虚拟现实", "元宇宙", "粮食产业", "中药",
    "央企创新", "央企创新驱动", "国企改革", "国新央企", "长江保护", "长江经济带", "长三角", "粤港澳",
    "一带一路", "央企科技", "央企现代能源", "央企结构调整", "央企共赢", "专精特新", "新材料", "新经济",
    "新消费", "新能车", "智能车", "智能驾驶", "物联网", "大数据", "信创",
]
BROAD_KEYWORDS = [
    "沪深300", "中证500", "中证800", "中证1000", "中证2000", "中证A500", "中证A100", "中证A50",
    "中证A股", "上证50", "上证180", "上证380", "上证综指", "上证指数", "深证100", "深证50",
    "创业板指", "创业板50", "创业板200", "创业板综", "科创50", "科创100", "科创200", "科创板50",
    "科创板100", "科创板200", "科创板综合", "科创综指", "科创创业50", "双创50", "双创", "北证50",
    "MSCI中国A股", "MSCI中国A50互联互通", "MSCI A50", "富时中国A50", "国证A指", "万得全A", "中小板", "中小100",
    "创业200", "创业300", "科创价格",
]
GENERIC_KEYWORDS = ["证券", "核心", "创新", "优势", "产业", "主题", "信息", "龙头", "成长", "精选", "优选"]
COMMODITY_STOCK_CONFLICT = ["黄金", "黄金股", "有色", "稀有金属", "稀土", "油气", "煤炭", "钢铁", "石油天然气", "矿业"]


def asset_type(row: pd.Series) -> tuple[str, list[str], str, str]:
    short = text(row.get("基金简称"))
    full = text(row.get("基金全称"))
    index_name = text(row.get("跟踪指数"))
    benchmark = text(row.get("业绩比较基准"))
    investment = text(row.get("投资类型"))
    classification_text = " ".join([short, index_name, benchmark])

    money_text = classification_text.replace("自由现金流", "").replace("现金流", "")
    if "货币市场型基金" in investment or "交易型货币市场基金" in full or contains_any(money_text, MONEY_KEYWORDS):
        return "货币ETF", match_keywords(money_text, MONEY_KEYWORDS), "命中货币ETF规则", "货币;场内货币"
    if "商品型基金" in investment or contains_any(classification_text, COMMODITY_KEYWORDS):
        return "商品ETF", match_keywords(classification_text, COMMODITY_KEYWORDS), "命中商品ETF规则", "商品"
    if "被动指数型债券基金" in investment or contains_any(classification_text, BOND_KEYWORDS):
        return "债券ETF", match_keywords(classification_text, BOND_KEYWORDS), "命中债券ETF规则", "债券"
    if "国际(QDII)股票型基金" in investment or contains_any(classification_text, CROSS_KEYWORDS):
        hits = match_keywords(classification_text, CROSS_KEYWORDS)
        tags = ["跨境"]
        for tag in ["港股", "港股通", "恒生", "美国", "纳斯达克", "香港"]:
            if tag in hits:
                tags.append("港股" if tag in ["港股", "港股通", "恒生", "香港"] else tag)
        return "跨境ETF", hits, "命中跨境ETF规则", ";".join(dict.fromkeys(tags))
    if not index_name and not investment:
        return "其他ETF", [], "投资类型和跟踪指数缺失，无法判断资产类型", "待复核"
    return "股票ETF", [], "未命中非股票资产类型规则，按境内股票ETF处理", ""


def classify_stock_index(index_name: str, benchmark: str, short_names: list[str]) -> dict:
    idx = text(index_name)
    norm = normalize_index(idx)
    short_text = " ".join(text(x) for x in short_names[:5])
    # Index priority: index first, benchmark second, fund short name third. Full fund name is intentionally excluded.
    idx_text = " ".join([idx, norm])
    aux_text = " ".join([benchmark, short_text])
    all_text = " ".join([idx_text, aux_text])

    tags: list[str] = []
    reasons: list[str] = []
    confidence = "高"

    enhance_hits = match_keywords(all_text, ENHANCE_KEYWORDS)
    strategy_hits = match_keywords(all_text, STRATEGY_KEYWORDS)
    theme_hits = match_keywords(all_text, THEME_KEYWORDS)
    broad_hits = match_keywords(idx_text, BROAD_KEYWORDS)
    generic_hits = match_keywords(all_text, GENERIC_KEYWORDS)
    keyword_hits = list(dict.fromkeys(enhance_hits + strategy_hits + theme_hits + broad_hits + generic_hits))

    # Avoid treating "证券" as industry unless the index/short name explicitly points to securities company ETFs.
    securities_company = ("证券公司" in all_text) or ("全指证券公司" in all_text) or ("证券公司ETF" in short_text) or bool(re.search(r"(^|[^投资])证券ETF", short_text))
    if securities_company and "证券公司" not in theme_hits:
        theme_hits.append("证券公司")
        keyword_hits.append("证券公司")

    if not idx:
        return {
            "分类": "其他ETF", "标签": "待复核", "纯宽基": "否", "策略": "否", "行业主题": "否",
            "置信度": "低", "依据": "跟踪指数缺失", "命中关键词": "",
            "需复核": "是", "复核原因": "跟踪指数缺失", "规范化": norm,
        }

    # Pure broad means broad hit from index text, with no strategy/theme modifiers.
    if broad_hits and enhance_hits and not strategy_hits and not theme_hits:
        add_tag(tags, "宽基")
        add_tag(tags, "指数增强")
        confidence = "中"
        classification = "宽基ETF"
        basis = "宽基增强/指数增强，主分类按宽基处理"
        pure_broad = "是"
        is_strategy = "是"
        is_theme = "否"
    elif strategy_hits:
        add_tag(tags, "Smart Beta")
        for h in strategy_hits[:2]:
            add_tag(tags, h)
        classification = "策略ETF"
        basis = "跟踪指数/业绩基准/基金简称命中明确策略因子"
        pure_broad = "否"
        is_strategy = "是"
        is_theme = "否"
        if broad_hits or theme_hits:
            confidence = "中"
    elif theme_hits:
        add_tag(tags, "行业主题")
        for h in theme_hits[:2]:
            add_tag(tags, h)
        classification = "行业主题ETF"
        basis = "跟踪指数/业绩基准/基金简称命中明确行业主题关键词"
        pure_broad = "否"
        is_strategy = "否"
        is_theme = "是"
        if broad_hits:
            confidence = "中"
    elif broad_hits:
        add_tag(tags, "宽基")
        classification = "宽基ETF"
        basis = "跟踪指数命中纯宽基或板块宽基关键词，且无策略/行业主题修饰"
        pure_broad = "是"
        is_strategy = "否"
        is_theme = "否"
    else:
        add_tag(tags, "待复核")
        classification = "其他ETF"
        basis = "未命中明确宽基、策略或行业主题指数规则"
        pure_broad = "否"
        is_strategy = "否"
        is_theme = "否"
        confidence = "低"
        reasons.append("指数分类为其他ETF")

    hit_kinds = sum(bool(x) for x in [broad_hits, strategy_hits, theme_hits])
    if hit_kinds >= 3:
        confidence = "低"
        reasons.append("同时命中宽基、策略、行业主题关键词")
    if generic_hits and not broad_hits and not strategy_hits and not theme_hits:
        confidence = "低"
        reasons.append("只命中泛词，缺少明确分类依据")
    if contains_any(all_text, ["央企", "国企"]) and not contains_any(all_text, [
        "央企红利", "国企红利", "央企股东回报", "国企股东回报", "央企分红", "国企分红", "央企价值", "国企价值",
        "央企创新", "央企创新驱动", "央企科技", "央企现代能源", "央企结构调整", "国企改革", "国新央企", "央企共赢", "央企主题",
    ]):
        confidence = "低"
        reasons.append("央企/国企含义不清")
    if contains_any(all_text, COMMODITY_STOCK_CONFLICT):
        reasons.append("股票指数含商品/资源相关行业词，建议人工确认")
    if classification == "其他ETF" or confidence == "低":
        need_review = "是"
    else:
        need_review = "否"
    if need_review == "是" and not reasons:
        reasons.append("低置信度分类")

    return {
        "分类": classification,
        "标签": ";".join(tags),
        "纯宽基": pure_broad,
        "策略": is_strategy,
        "行业主题": is_theme,
        "置信度": confidence,
        "依据": basis,
        "命中关键词": "、".join(keyword_hits),
        "需复核": need_review,
        "复核原因": "；".join(dict.fromkeys(reasons)),
        "规范化": norm,
    }


def grouped_summary(df: pd.DataFrame, field: str, label: str) -> pd.DataFrame:
    rows = []
    total_count = len(df)
    total_scale = df["基金规模_亿元"].sum()
    for key, g in df.groupby(field, dropna=False):
        scale = g["基金规模_亿元"].sum()
        rows.append({
            label: key,
            "ETF数量": len(g),
            "数量占比": round(len(g) / total_count, 6) if total_count else 0,
            "总规模_亿元": round(scale, 4),
            "规模占比": round(scale / total_scale, 6) if total_scale else 0,
            "平均规模_亿元": round(g["基金规模_亿元"].mean(), 4) if len(g) else 0,
            "规模中位数_亿元": round(g["基金规模_亿元"].median(), 4) if len(g) else 0,
        })
    return pd.DataFrame(rows).sort_values("总规模_亿元", ascending=False)


def make_label_stats(df: pd.DataFrame) -> pd.DataFrame:
    buckets: dict[str, list[int]] = {}
    for i, raw in enumerate(df["ETF标签"].fillna("")):
        for label in [x.strip() for x in str(raw).split(";") if x.strip()]:
            buckets.setdefault(label, []).append(i)
    rows = []
    for label, idxs in buckets.items():
        g = df.iloc[idxs]
        rows.append({
            "标签": label,
            "产品数量": len(g),
            "总规模_亿元": round(g["基金规模_亿元"].sum(), 4),
            "代表产品示例": "；".join(g["基金简称"].head(5).astype(str)),
        })
    return pd.DataFrame(rows).sort_values("总规模_亿元", ascending=False)


def is_pure_broad_candidate(row: pd.Series, needle: str) -> bool:
    idx = " ".join([text(row.get("跟踪指数")), normalize_index(text(row.get("跟踪指数")))])
    short = text(row.get("基金简称"))
    search = f"{idx} {short}"
    broad_hits = match_keywords(idx, BROAD_KEYWORDS)
    if not broad_hits:
        return False
    variants = [needle]
    if needle == "科创50":
        variants.append("科创板50")
    if needle == "创业板":
        variants.extend(["创业板指", "创业板50", "创业板200", "创业板300"])
    if not any(re.search(re.escape(v), idx, flags=re.I) for v in variants):
        return False
    if contains_any(search, STRATEGY_KEYWORDS + THEME_KEYWORDS):
        return False
    return True


def check_keywords(df: pd.DataFrame, keywords: list[str], expected: str, mode: str) -> pd.DataFrame:
    rows = []
    searchable = (
        df["基金简称"].fillna("").astype(str) + " " +
        df["跟踪指数"].fillna("").astype(str) + " " +
        df["业绩比较基准"].fillna("").astype(str)
    )
    for kw in keywords:
        needle = kw.replace(" ETF", "").replace("ETF", "").strip()
        if mode == "broad":
            matches = df[df.apply(lambda r: is_pure_broad_candidate(r, needle), axis=1)]
        elif mode == "strategy":
            matches = df[(df["ETF资产类型"].eq("股票ETF")) & searchable.str.contains(re.escape(needle), case=False, na=False)]
        elif mode == "not_broad":
            matches = df[(df["ETF资产类型"].eq("股票ETF")) & searchable.str.contains(re.escape(needle), case=False, na=False)]
        else:
            matches = df[searchable.str.contains(re.escape(needle), case=False, na=False)]
        if matches.empty:
            rows.append({"检查关键词": kw, "Wind代码": "", "基金简称": "", "跟踪指数": "", "最终ETF分类": "", "是否通过": "未找到", "问题说明": "主分析池中未匹配到关键词"})
            continue
        for _, r in matches.iterrows():
            final = text(r["最终ETF分类"])
            ok = final == expected if mode != "not_broad" else final != "宽基ETF"
            rows.append({
                "检查关键词": kw,
                "Wind代码": r["Wind代码"],
                "基金简称": r["基金简称"],
                "跟踪指数": r["跟踪指数"],
                "最终ETF分类": final,
                "是否通过": "通过" if ok else "失败",
                "问题说明": "" if ok else (f"应为{expected}" if mode != "not_broad" else "不应归为宽基ETF"),
            })
    return pd.DataFrame(rows)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="366092")
    header_font = Font(bold=True, color="FFFFFF")
    text_cols = {"Wind代码", "基金代码", "交易代码", "跟踪指数代码"}
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        headers = [text(c.value) for c in ws[1]]
        for idx, h in enumerate(headers, 1):
            if h in text_cols:
                for cell in ws[get_column_letter(idx)]:
                    cell.number_format = "@"
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col[:200]:
                max_len = max(max_len, len(text(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 38)
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError("未找到二次修正版主分析池：ETF分析池_上市交易_二次修正。请检查输入文件路径。")

    try:
        df = pd.read_excel(INPUT_PATH, sheet_name=INPUT_SHEET, dtype={"Wind代码": str, "基金代码": str, "交易代码": str, "跟踪指数代码": str})
    except ValueError as exc:
        raise RuntimeError("未找到二次修正版主分析池：ETF分析池_上市交易_二次修正。请检查输入文件路径。") from exc

    for col in ["成立日期", "上市日期"]:
        if col in df.columns:
            df[col] = df[col].apply(fmt_date)
    df["基金规模_亿元"] = pd.to_numeric(df["基金规模_亿元"], errors="coerce")
    if "管理人简称" not in df.columns:
        df["管理人简称"] = ""

    # Asset type first, based on investment type and index/benchmark/short name. Fund full name is only used for money-market ETF cue.
    assets = df.apply(asset_type, axis=1)
    df["ETF资产类型"] = [x[0] for x in assets]
    df["_资产命中关键词"] = ["、".join(x[1]) for x in assets]
    df["_资产分类依据"] = [x[2] for x in assets]
    df["_资产标签"] = [x[3] for x in assets]

    total_count = len(df)
    total_scale = round(float(df["基金规模_亿元"].sum()), 4)
    of_count = int(df["Wind代码"].astype(str).str.endswith(".OF").sum())
    missing_start = int(df["成立日期"].isna().sum() + (df["成立日期"].astype(str).str.strip() == "").sum())
    missing_list = int(df["上市日期"].isna().sum() + (df["上市日期"].astype(str).str.strip() == "").sum())
    bad_scale = int(df["基金规模_亿元"].isna().sum() + (df["基金规模_亿元"] <= 0).sum())
    fatal = []
    if total_count != EXPECTED_COUNT:
        fatal.append(f"主分析池数量={total_count}")
    if abs(total_scale - EXPECTED_SCALE) > 0.1:
        fatal.append(f"ETF总规模_亿元={total_scale}")
    if of_count:
        fatal.append(f".OF产品数量={of_count}")
    if missing_start:
        fatal.append(f"成立日期缺失数量={missing_start}")
    if missing_list:
        fatal.append(f"上市日期缺失数量={missing_list}")
    if bad_scale:
        fatal.append(f"基金规模缺失或小于等于0数量={bad_scale}")
    if fatal:
        raise RuntimeError("质量检查失败，已停止输出：" + "; ".join(fatal))

    index_rows = []
    index_map: dict[str, dict] = {}
    for index_name, g in df.groupby("跟踪指数", dropna=False):
        idx = text(index_name)
        assets_for_index = sorted(set(g["ETF资产类型"].astype(str)))
        if len(assets_for_index) == 1:
            index_asset = assets_for_index[0]
        else:
            index_asset = "混合/需复核"
        representative = text(g.sort_values("基金规模_亿元", ascending=False).iloc[0]["基金简称"])
        benchmark = " ".join(g["业绩比较基准"].dropna().astype(str).unique()[:3])
        short_names = list(g["基金简称"].dropna().astype(str).unique())
        if index_asset == "股票ETF":
            c = classify_stock_index(idx, benchmark, short_names)
            index_class = c["分类"]
            tags = c["标签"]
            pure_broad = c["纯宽基"]
            is_strategy = c["策略"]
            is_theme = c["行业主题"]
            confidence = c["置信度"]
            basis = c["依据"]
            keywords = c["命中关键词"]
            need_review = c["需复核"]
            review_reason = c["复核原因"]
            norm = c["规范化"]
        elif index_asset in ["货币ETF", "商品ETF", "债券ETF", "跨境ETF"]:
            index_class = index_asset
            tags = {"货币ETF": "货币;场内货币", "商品ETF": "商品", "债券ETF": "债券", "跨境ETF": "跨境"}.get(index_asset, "")
            pure_broad = is_strategy = is_theme = "否"
            confidence = "高"
            basis = f"关联ETF资产类型为{index_asset}"
            keywords = "、".join([x for x in g["_资产命中关键词"].dropna().astype(str).unique() if x])
            need_review = "否"
            review_reason = ""
            norm = normalize_index(idx)
        else:
            index_class = "其他ETF"
            tags = "待复核"
            pure_broad = is_strategy = is_theme = "否"
            confidence = "低"
            basis = "同一跟踪指数关联多种资产类型或资产类型无法判断"
            keywords = ""
            need_review = "是"
            review_reason = "指数资产类型为混合/需复核"
            norm = normalize_index(idx)
        row = {
            "跟踪指数": idx,
            "跟踪指数_规范化": norm,
            "关联ETF数量": len(g),
            "关联ETF规模_亿元": round(float(g["基金规模_亿元"].sum()), 4),
            "代表ETF简称": representative,
            "指数资产类型": index_asset,
            "指数分类_自动": index_class,
            "指数标签": tags,
            "是否纯宽基指数": pure_broad,
            "是否策略指数": is_strategy,
            "是否行业主题指数": is_theme,
            "分类置信度": confidence,
            "分类依据": basis,
            "命中关键词": keywords,
            "是否需要人工复核": need_review,
            "复核原因": review_reason,
            "人工修正指数分类": "",
            "最终指数分类": index_class,
            "最终指数分类来源": "自动分类",
        }
        index_rows.append(row)
        index_map[idx] = row

    index_df = pd.DataFrame(index_rows).sort_values(["指数资产类型", "关联ETF规模_亿元"], ascending=[True, False])

    mapped = df["跟踪指数"].apply(lambda x: index_map.get(text(x), {}))
    df["跟踪指数_规范化"] = mapped.apply(lambda x: x.get("跟踪指数_规范化", ""))
    df["指数分类_自动"] = mapped.apply(lambda x: x.get("指数分类_自动", "其他ETF"))
    df["最终指数分类"] = mapped.apply(lambda x: x.get("最终指数分类", "其他ETF"))
    df["ETF分类_自动"] = df.apply(lambda r: r["ETF资产类型"] if r["ETF资产类型"] in ["货币ETF", "商品ETF", "债券ETF", "跨境ETF"] else (r["最终指数分类"] or "其他ETF"), axis=1)
    df["是否增强策略"] = mapped.apply(lambda x: "是" if "指数增强" in text(x.get("指数标签", "")) else "否")
    df["是否SmartBeta策略"] = mapped.apply(lambda x: "是" if "Smart Beta" in text(x.get("指数标签", "")) else "否")
    df["是否宽基ETF"] = df["ETF分类_自动"].apply(lambda x: "是" if x == "宽基ETF" else "否")
    df["是否行业主题ETF"] = df["ETF分类_自动"].apply(lambda x: "是" if x == "行业主题ETF" else "否")
    df["是否策略ETF"] = df.apply(lambda r: "是" if r["ETF分类_自动"] == "策略ETF" or r["是否增强策略"] == "是" or r["是否SmartBeta策略"] == "是" else "否", axis=1)
    df["是否跨境ETF"] = df["ETF资产类型"].apply(lambda x: "是" if x == "跨境ETF" else "否")
    df["是否债券ETF"] = df["ETF资产类型"].apply(lambda x: "是" if x == "债券ETF" else "否")
    df["是否商品ETF"] = df["ETF资产类型"].apply(lambda x: "是" if x == "商品ETF" else "否")
    df["是否货币ETF"] = df["ETF资产类型"].apply(lambda x: "是" if x == "货币ETF" else "否")
    df["是否多资产ETF"] = (df["基金简称"].fillna("").astype(str) + " " + df["跟踪指数"].fillna("").astype(str) + " " + df["业绩比较基准"].fillna("").astype(str)).apply(lambda s: "是" if contains_any(s, ["多资产", "资产配置"]) else "否")
    df["分类置信度"] = mapped.apply(lambda x: x.get("分类置信度", "低"))
    df["分类依据"] = mapped.apply(lambda x: x.get("分类依据", ""))
    df["命中关键词"] = mapped.apply(lambda x: x.get("命中关键词", ""))
    df["ETF标签"] = df.apply(lambda r: r["_资产标签"] if r["ETF资产类型"] != "股票ETF" else text(index_map.get(text(r["跟踪指数"]), {}).get("指数标签", "")), axis=1)
    df["是否需要人工复核"] = mapped.apply(lambda x: x.get("是否需要人工复核", "否"))
    df["复核原因"] = mapped.apply(lambda x: x.get("复核原因", ""))
    df.loc[df["ETF分类_自动"].eq("其他ETF"), ["是否需要人工复核", "分类置信度"]] = ["是", "低"]
    df.loc[df["跟踪指数"].fillna("").astype(str).str.strip().eq(""), ["是否需要人工复核", "分类置信度", "复核原因"]] = ["是", "低", "跟踪指数缺失"]
    df["人工修正分类"] = ""
    df["最终ETF分类"] = df["ETF分类_自动"]
    df["最终分类来源"] = "自动分类"

    original_fields = list(df.columns)
    remove_internal = ["_资产命中关键词", "_资产分类依据", "_资产标签"]
    original_fields = [c for c in original_fields if c not in remove_internal and c not in [
        "ETF资产类型", "跟踪指数_规范化", "指数分类_自动", "最终指数分类", "ETF分类_自动", "ETF标签",
        "是否增强策略", "是否SmartBeta策略", "是否宽基ETF", "是否行业主题ETF", "是否策略ETF", "是否跨境ETF",
        "是否债券ETF", "是否商品ETF", "是否货币ETF", "是否多资产ETF", "分类置信度", "分类依据", "命中关键词",
        "是否需要人工复核", "复核原因", "人工修正分类", "最终ETF分类", "最终分类来源",
    ]]
    required_original = [
        "Wind代码", "证券简称", "基金简称", "基金全称", "基金代码", "交易代码", "基金管理人", "管理人简称", "基金类型",
        "投资类型", "业绩比较基准", "基金规模_亿元", "成立日期", "上市日期", "基金上市地点", "跟踪指数", "跟踪指数代码",
        "基金托管人", "管理费率", "托管费率",
    ]
    for col in required_original:
        if col not in df.columns:
            df[col] = ""
            original_fields.append(col)
    # Put required fields first, then any additional original columns from the source, then new fields.
    extra_original = [c for c in original_fields if c not in required_original]
    new_fields = [
        "ETF资产类型", "跟踪指数_规范化", "指数分类_自动", "最终指数分类", "ETF分类_自动", "ETF标签",
        "是否增强策略", "是否SmartBeta策略", "是否宽基ETF", "是否行业主题ETF", "是否策略ETF", "是否跨境ETF",
        "是否债券ETF", "是否商品ETF", "是否货币ETF", "是否多资产ETF", "分类置信度", "分类依据", "命中关键词",
        "是否需要人工复核", "复核原因", "人工修正分类", "最终ETF分类", "最终分类来源",
    ]
    master_df = df[required_original + extra_original + new_fields].copy()

    review_df = master_df[master_df["是否需要人工复核"].eq("是")][[
        "Wind代码", "基金简称", "基金全称", "基金管理人", "基金规模_亿元", "跟踪指数", "跟踪指数_规范化",
        "ETF资产类型", "最终指数分类", "ETF分类_自动", "ETF标签", "分类置信度", "分类依据", "命中关键词",
        "复核原因", "人工修正分类", "最终ETF分类",
    ]].copy()

    final_summary = grouped_summary(master_df, "最终ETF分类", "最终ETF分类")
    asset_summary = grouped_summary(master_df, "ETF资产类型", "ETF资产类型")
    stats_df = pd.concat([
        final_summary.assign(汇总维度="最终ETF分类"),
        asset_summary.assign(汇总维度="ETF资产类型"),
    ], ignore_index=True)
    stats_df = stats_df[["汇总维度", "最终ETF分类", "ETF资产类型", "ETF数量", "数量占比", "总规模_亿元", "规模占比", "平均规模_亿元", "规模中位数_亿元"]]

    index_stats = index_df.groupby("最终指数分类", dropna=False).agg(
        跟踪指数数量=("跟踪指数", "count"),
        关联ETF数量=("关联ETF数量", "sum"),
        关联ETF规模_亿元=("关联ETF规模_亿元", "sum"),
    ).reset_index().sort_values("关联ETF规模_亿元", ascending=False)
    index_stats["关联ETF规模_亿元"] = index_stats["关联ETF规模_亿元"].round(4)

    label_stats = make_label_stats(master_df)
    broad_sample = master_df[master_df["最终ETF分类"].eq("宽基ETF")].sort_values("基金规模_亿元", ascending=False).head(100)
    strategy_sample = master_df[master_df["最终ETF分类"].eq("策略ETF")].sort_values("基金规模_亿元", ascending=False).head(100)
    theme_sample = master_df[master_df["最终ETF分类"].eq("行业主题ETF")].sort_values("基金规模_亿元", ascending=False).head(100)

    rules = pd.DataFrame([
        ("本次输入文件", str(INPUT_PATH)),
        ("本次输入sheet", INPUT_SHEET),
        ("本次输出文件", str(OUTPUT_PATH)),
        ("本次是否改变主分析池", "否，保留二次修正版1574只上市ETF主分析池"),
        ("主分析池数量", total_count),
        ("主分析池总规模", total_scale),
        ("为什么采用指数优先分类", "前几版对产品名称做关键词分类时，基金全称中的“证券投资基金”会误触发证券行业关键词；指数优先可直接按跟踪指数经济含义分类。"),
        ("为什么基金全称不得用于行业主题关键词判断", "基金全称通常包含法律形式和“证券投资基金”等通用表述，不代表跟踪指数行业属性；本次只将基金全称用于ETF/货币基金辅助判断。"),
        ("ETF资产类型规则", "按货币、商品、债券、跨境、股票、其他的优先级互斥判断。"),
        ("跟踪指数分类规则", "对去重跟踪指数分类；股票指数按纯宽基/板块宽基、策略指数、行业主题指数、其他ETF判断。"),
        ("ETF分类映射规则", "非股票ETF由资产类型直接映射；股票ETF由最终指数分类映射为ETF分类。"),
        ("策略ETF判断规则", "识别红利、低波、质量、价值、成长、现金流、股息、等权、ESG、股东回报等明确策略因子。"),
        ("行业主题ETF判断规则", "识别行业、产业、区域、央企/国企主题、科技制造、能源材料等明确指数关键词；证券仅在证券公司相关语境下触发。"),
        ("宽基ETF判断规则", "仅当跟踪指数本身为纯宽基或板块宽基且无策略/行业主题修饰时归为宽基ETF。"),
        ("央企/国企类判断规则", "央企/国企红利、价值、股东回报归策略；央企创新、科技、现代能源、国企改革等归行业主题；仅央企/国企需复核。"),
        ("人工复核规则", "其他ETF、低置信度、多类关键词冲突、只命中泛词、央企国企含义不清、跟踪指数缺失、商品/资源相关股票主题均标记复核。"),
        ("后续人工修正使用方式", "可先在人工修正指数分类或人工修正分类中填入人工结论，后续用人工字段覆盖最终分类。"),
    ], columns=["项目", "说明"])

    quality_rows = []
    def add_check(item: str, ok: bool, value, note: str) -> None:
        quality_rows.append({"检查项": item, "结果": "通过" if ok else "失败", "数值": value, "说明": note})
    empty_index_class = int(index_df["指数分类_自动"].fillna("").astype(str).str.strip().eq("").sum())
    empty_auto = int(master_df["ETF分类_自动"].fillna("").astype(str).str.strip().eq("").sum())
    empty_final = int(master_df["最终ETF分类"].fillna("").astype(str).str.strip().eq("").sum())
    other_count = int(master_df["最终ETF分类"].eq("其他ETF").sum())
    review_count = len(review_df)
    add_check("主分析池数量是否为1574", total_count == EXPECTED_COUNT, total_count, "必须为1574")
    add_check("总规模是否约为46949.9638亿元", abs(total_scale - EXPECTED_SCALE) <= 0.1, total_scale, "误差不超过0.1亿元")
    add_check("是否存在.OF产品", of_count == 0, of_count, "必须为0")
    add_check("成立日期是否有缺失", missing_start == 0, missing_start, "必须为0")
    add_check("上市日期是否有缺失", missing_list == 0, missing_list, "必须为0")
    add_check("基金规模是否有缺失或小于等于0", bad_scale == 0, bad_scale, "必须为0")
    add_check("跟踪指数分类是否有空值", empty_index_class == 0, empty_index_class, "必须为0")
    add_check("ETF分类_自动是否有空值", empty_auto == 0, empty_auto, "必须为0")
    add_check("最终ETF分类是否有空值", empty_final == 0, empty_final, "必须为0")
    add_check("待人工复核数量", True, review_count, "提示项")
    add_check("其他ETF数量", True, other_count, "提示项")
    for cat in ["宽基ETF", "行业主题ETF", "策略ETF", "跨境ETF", "债券ETF", "商品ETF", "货币ETF"]:
        add_check(f"{cat}数量", True, int(master_df["最终ETF分类"].eq(cat).sum()), "提示项")
    quality_df = pd.DataFrame(quality_rows)

    broad_check = check_keywords(master_df, ["沪深300ETF", "中证500ETF", "中证1000ETF", "中证2000ETF", "中证A500ETF", "上证50ETF", "创业板ETF", "创业板50ETF", "科创50ETF", "科创100ETF", "科创200ETF", "科创板综合ETF", "科创创业50ETF", "双创50ETF", "北证50ETF"], "宽基ETF", "broad")
    theme_check = check_keywords(master_df, ["中证全指证券公司ETF", "中证全指医药ETF", "中证全指半导体ETF", "中证全指食品ETF", "中证全指电力ETF", "中证全指银行ETF", "中证全指通信ETF", "中证全指软件ETF", "中证500信息技术ETF", "中证500医药ETF", "中证500消费ETF"], "行业主题ETF", "not_broad")
    strategy_check = check_keywords(master_df, ["A500红利低波ETF", "中证A500红利低波动ETF", "沪深300质量ETF", "沪深300红利ETF", "中证500低波ETF", "中证500质量成长ETF", "300价值ETF", "300成长ETF", "500质量ETF", "500价值ETF", "500成长ETF", "红利质量ETF", "现金流ETF", "自由现金流ETF", "股息ETF", "等权ETF", "ESG ETF"], "策略ETF", "strategy")

    if (quality_df["结果"] == "失败").any():
        failed = "; ".join(quality_df.loc[quality_df["结果"].eq("失败"), "检查项"].astype(str))
        raise RuntimeError("质量检查失败，已停止输出：" + failed)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        index_df.to_excel(writer, sheet_name="跟踪指数分类表", index=False)
        master_df.to_excel(writer, sheet_name="ETF分类主表", index=False)
        review_df.to_excel(writer, sheet_name="待人工复核ETF", index=False)
        stats_df.to_excel(writer, sheet_name="分类统计_数量规模", index=False)
        index_stats.to_excel(writer, sheet_name="跟踪指数统计", index=False)
        label_stats.to_excel(writer, sheet_name="标签统计", index=False)
        broad_sample.to_excel(writer, sheet_name="宽基ETF样本", index=False)
        strategy_sample.to_excel(writer, sheet_name="策略ETF样本", index=False)
        theme_sample.to_excel(writer, sheet_name="行业主题ETF样本", index=False)
        rules.to_excel(writer, sheet_name="分类规则说明", index=False)
        quality_df.to_excel(writer, sheet_name="质量检查", index=False)
        broad_check.to_excel(writer, sheet_name="宽基误判检查", index=False)
        theme_check.to_excel(writer, sheet_name="行业主题误判检查", index=False)
        strategy_check.to_excel(writer, sheet_name="策略误判检查", index=False)

    style_workbook(OUTPUT_PATH)

    final_dist = master_df.groupby("最终ETF分类").agg(数量=("Wind代码", "count"), 规模=("基金规模_亿元", "sum")).sort_values("规模", ascending=False)
    asset_dist = master_df.groupby("ETF资产类型").agg(数量=("Wind代码", "count"), 规模=("基金规模_亿元", "sum")).sort_values("规模", ascending=False)
    index_dist = index_df.groupby("最终指数分类").agg(跟踪指数数量=("跟踪指数", "count"), 关联ETF数量=("关联ETF数量", "sum"), 规模=("关联ETF规模_亿元", "sum")).sort_values("规模", ascending=False)
    def cat_stats(cat: str) -> tuple[int, float]:
        g = master_df[master_df["最终ETF分类"].eq(cat)]
        return len(g), round(float(g["基金规模_亿元"].sum()), 4)

    print(f"输入文件路径：{INPUT_PATH}")
    print(f"输入 sheet 名称：{INPUT_SHEET}")
    print(f"输出文件路径：{OUTPUT_PATH}")
    print(f"主分析池数量：{total_count}")
    print(f"ETF总规模_亿元：{total_scale}")
    print("最终ETF分类 分布：")
    for k, r in final_dist.iterrows():
        print(f"  {k}: {int(r['数量'])}只, {round(float(r['规模']), 4)}亿元")
    print("ETF资产类型 分布：")
    for k, r in asset_dist.iterrows():
        print(f"  {k}: {int(r['数量'])}只, {round(float(r['规模']), 4)}亿元")
    print("跟踪指数分类分布：")
    for k, r in index_dist.iterrows():
        print(f"  {k}: {int(r['跟踪指数数量'])}个指数, {int(r['关联ETF数量'])}只ETF, {round(float(r['规模']), 4)}亿元")
    print(f"待人工复核数量：{review_count}")
    print(f"其他ETF数量：{other_count}")
    for cat in ["宽基ETF", "行业主题ETF", "策略ETF"]:
        cnt, scale = cat_stats(cat)
        print(f"{cat}数量和规模：{cnt}只, {scale}亿元")
    print("是否存在质量检查失败项：否")
    print("处理完成：ETF分类主表_指数优先版.xlsx 已生成")


if __name__ == "__main__":
    main()
