from __future__ import annotations

import math
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_PATH = Path(r"C:\Users\lvdon\Desktop\Fullgoal\ICI课题\ETF分类主表_指数优先版\ETF分类主表_指数优先版.xlsx")
OUTPUT_DIR = Path(r"C:\Users\lvdon\Desktop\Fullgoal\ICI课题\ETF分类主表_指数优先版_v2")
OUTPUT_PATH = OUTPUT_DIR / "ETF分类主表_指数优先版_v2.xlsx"

EXPECTED_COUNT = 1574
EXPECTED_SCALE = 46949.9638


def text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def num(value) -> float:
    try:
        if pd.isna(value):
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def contains_any(s: str, keywords: list[str]) -> bool:
    lower = s.lower()
    return any(k.lower() in lower for k in keywords if k)


def match_keywords(s: str, keywords: list[str]) -> list[str]:
    lower = s.lower()
    out: list[str] = []
    for k in keywords:
        if k and k.lower() in lower and k not in out:
            out.append(k)
    return out


def split_tags(raw) -> list[str]:
    tags = []
    for part in text(raw).split(";"):
        part = part.strip()
        if part and part not in tags:
            tags.append(part)
    return tags


def join_tags(tags: list[str]) -> str:
    out = []
    for tag in tags:
        tag = text(tag)
        if tag and tag not in out:
            out.append(tag)
    return ";".join(out)


def add_tag(tags: list[str], tag: str) -> None:
    if tag and tag not in tags:
        tags.append(tag)


def remove_tag(tags: list[str], tag: str) -> list[str]:
    return [x for x in tags if x != tag]


def fmt_date(value) -> str:
    if text(value) == "":
        return ""
    try:
        return pd.to_datetime(value).strftime("%Y-%m-%d")
    except Exception:
        return text(value)


def safe_sheet_read(path: Path, sheet: str) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name=sheet, engine="openpyxl", dtype={"Wind代码": str, "基金代码": str, "交易代码": str, "跟踪指数代码": str})
    except ValueError as exc:
        if sheet == "ETF分类主表":
            raise RuntimeError("未找到 ETF分类主表 sheet，请检查输入文件。") from exc
        return pd.DataFrame()


ENHANCE_KEYWORDS = ["增强策略", "指数增强", "增强指数", "增强型", "增强指数型", "被动指数增强", "主动增强"]

THEME_PATCH_KEYWORDS = [
    "电网设备", "电网", "证券龙头", "证券公司", "国防", "船舶", "船舶产业", "石化", "石化产业",
    "油气", "油气资源", "油气产业", "石油天然气", "能源", "能源指数", "家用电器", "家电",
    "集成电路", "农牧渔", "农牧", "畜牧", "5G", "5G产业", "电信", "电信主题", "智能制造",
    "高端制造", "工业母机", "机床", "中药", "稀土", "稀土产业", "碳中和", "低碳经济",
    "卫星产业", "卫星", "通用航空", "低空经济", "粮食产业", "长江保护", "央企创新",
    "央企创新驱动", "央企科技", "央企现代能源", "央企结构调整", "央企共赢", "国企改革",
    "国新央企", "专精特新", "新材料", "新经济", "新消费", "智能车", "智能驾驶", "物联网",
    "大数据", "信创", "环保", "旅游", "物流", "教育", "医疗服务", "医疗创新", "医疗保健",
    "创新药", "医药卫生", "医疗器械", "消费电子", "电子", "机械", "工程机械", "建材",
    "建筑材料", "房地产", "地产", "煤炭", "钢铁", "有色金属", "工业有色", "黄金股", "矿业", "资源",
]
BROAD_PATCH_KEYWORDS = [
    "深成指", "深证成指", "深证成份", "深成", "国证2000", "创业板大盘", "上证中盘", "上证中小",
    "中证小盘", "中小盘", "国证1000", "国证300", "上证150", "上证380", "科创200", "科创板200",
    "科创板综合", "科创综指", "科创创业50", "双创50", "创业板50", "创业板200", "深证100",
    "深证50", "北证50",
]
STRATEGY_PATCH_KEYWORDS = [
    "自由现金流", "现金流", "股息", "红利", "低波", "质量", "价值", "成长", "等权", "ESG",
    "股东回报", "高股息", "高分红", "基本面", "盈利质量", "回购", "红利质量", "红利低波",
    "质量低波", "价值稳健", "成长创新",
]
SHS_KEYWORDS = [
    "沪港深", "沪深港", "SHS", "中证沪港深", "沪港深500", "沪港深科技", "沪港深云计算",
    "沪港深消费", "沪港深黄金", "沪深港黄金", "沪港深创新药",
]

COMMON_BROAD_SAMPLES = [
    "南方中证500ETF", "易方达中证A500ETF", "中金中证A500ETF", "南方中证1000ETF", "易方达创业板ETF",
    "华泰柏瑞沪深300ETF", "易方达沪深300ETF", "华夏上证50ETF", "富国上证综指ETF", "嘉实沪深300ETF",
    "南方创业板ETF", "华夏科创50ETF",
]
REAL_ENHANCED_SAMPLES = ["中证1000增强策略ETF", "沪深300增强策略ETF", "创业板增强策略ETF", "中证A500增强策略ETF", "中证500增强策略ETF", "中证2000增强策略ETF"]
THEME_SAMPLE_NAMES = [
    "华夏中证电网设备主题ETF", "鹏华国证证券龙头ETF", "富国中证智选船舶产业ETF", "鹏华中证国防ETF",
    "易方达中证石化产业ETF", "汇添富中证油气资源ETF", "国泰中证全指家用电器ETF", "国泰中证全指集成电路ETF",
    "博时中证5G产业50ETF", "易方达中证电信主题ETF",
]
BROAD_SAMPLE_NAMES = ["南方深成ETF", "万家国证2000ETF", "西部利得创业板大盘ETF", "易方达上证中盘ETF"]


def priority_text(row: pd.Series) -> str:
    return " ".join(text(row.get(c)) for c in ["跟踪指数", "业绩比较基准", "基金简称", "证券简称"])


def enhance_text(row: pd.Series) -> str:
    return " ".join(text(row.get(c)) for c in ["基金简称", "证券简称", "投资类型", "跟踪指数", "业绩比较基准", "基金类型"])


def has_real_enhance(row: pd.Series) -> bool:
    return contains_any(enhance_text(row), ENHANCE_KEYWORDS)


def set_class_flags(df: pd.DataFrame) -> None:
    df["是否宽基ETF"] = df["最终ETF分类"].apply(lambda x: "是" if text(x) == "宽基ETF" else "否")
    df["是否行业主题ETF"] = df["最终ETF分类"].apply(lambda x: "是" if text(x) == "行业主题ETF" else "否")
    df["是否跨境ETF"] = df["ETF资产类型"].apply(lambda x: "是" if text(x) == "跨境ETF" else "否")
    df["是否债券ETF"] = df["ETF资产类型"].apply(lambda x: "是" if text(x) == "债券ETF" else "否")
    df["是否商品ETF"] = df["ETF资产类型"].apply(lambda x: "是" if text(x) == "商品ETF" else "否")
    df["是否货币ETF"] = df["ETF资产类型"].apply(lambda x: "是" if text(x) == "货币ETF" else "否")
    df["是否策略ETF"] = df.apply(
        lambda r: "是" if text(r.get("最终ETF分类")) == "策略ETF" or text(r.get("是否增强策略")) == "是" or text(r.get("是否SmartBeta策略")) == "是" else "否",
        axis=1,
    )


def summary_by(df: pd.DataFrame, field: str, label: str) -> pd.DataFrame:
    total_count = len(df)
    total_scale = df["基金规模_亿元"].sum()
    rows = []
    for key, g in df.groupby(field, dropna=False):
        scale = g["基金规模_亿元"].sum()
        rows.append({
            label: key,
            "ETF数量": len(g),
            "数量占比": round(len(g) / total_count, 6) if total_count else 0,
            "总规模_亿元": round(scale, 4),
            "规模占比": round(scale / total_scale, 6) if total_scale else 0,
            "平均规模_亿元": round(g["基金规模_亿元"].mean(), 4) if len(g) else 0,
            "规模中位数_亿元": round(g["基金规模_亿元"].median(), 4) if len(g) else 0,
        })
    return pd.DataFrame(rows).sort_values("总规模_亿元", ascending=False)


def label_stats(df: pd.DataFrame) -> pd.DataFrame:
    buckets: dict[str, list[int]] = {}
    for i, raw in enumerate(df["ETF标签"].fillna("")):
        for tag in [x.strip() for x in str(raw).split(";") if x.strip()]:
            buckets.setdefault(tag, []).append(i)
    rows = []
    for tag, idxs in buckets.items():
        g = df.iloc[idxs]
        rows.append({
            "标签": tag,
            "产品数量": len(g),
            "总规模_亿元": round(g["基金规模_亿元"].sum(), 4),
            "代表产品示例": "；".join(g["基金简称"].head(5).astype(str)),
        })
    return pd.DataFrame(rows).sort_values("总规模_亿元", ascending=False) if rows else pd.DataFrame(columns=["标签", "产品数量", "总规模_亿元", "代表产品示例"])


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="366092")
    font = Font(bold=True, color="FFFFFF")
    text_cols = {"Wind代码", "基金代码", "交易代码", "跟踪指数代码"}
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for c in ws[1]:
            c.fill = fill
            c.font = font
        headers = [text(c.value) for c in ws[1]]
        for idx, header in enumerate(headers, 1):
            if header in text_cols:
                for cell in ws[get_column_letter(idx)]:
                    cell.number_format = "@"
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col_cells[:200]:
                max_len = max(max_len, len(text(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def find_rows(df: pd.DataFrame, sample: str) -> pd.DataFrame:
    needle = sample.replace(" ETF", "").replace("ETF", "")
    search = df["基金简称"].fillna("").astype(str) + " " + df["证券简称"].fillna("").astype(str) + " " + df["跟踪指数"].fillna("").astype(str)
    return df[search.str.contains(re.escape(needle), case=False, na=False)]


def find_exact_fund(df: pd.DataFrame, sample: str) -> pd.DataFrame:
    compact_sample = sample.replace(" ", "")
    short = df["基金简称"].fillna("").astype(str).str.replace(" ", "", regex=False)
    sec = df["证券简称"].fillna("").astype(str).str.replace(" ", "", regex=False)
    return df[(short == compact_sample) | (sec == compact_sample)]


def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"输入文件不存在：{INPUT_PATH}")
    df = safe_sheet_read(INPUT_PATH, "ETF分类主表")
    _index = safe_sheet_read(INPUT_PATH, "跟踪指数分类表")
    _review = safe_sheet_read(INPUT_PATH, "待人工复核ETF")
    _stats = safe_sheet_read(INPUT_PATH, "分类统计_数量规模")

    for col in ["成立日期", "上市日期"]:
        if col in df.columns:
            df[col] = df[col].apply(fmt_date)
    df["基金规模_亿元"] = pd.to_numeric(df["基金规模_亿元"], errors="coerce")
    for col in [
        "ETF资产类型", "ETF分类_自动", "ETF标签", "是否增强策略", "是否SmartBeta策略", "是否宽基ETF",
        "是否行业主题ETF", "是否策略ETF", "是否跨境ETF", "是否债券ETF", "是否商品ETF", "是否货币ETF",
        "是否多资产ETF", "分类置信度", "分类依据", "命中关键词", "是否需要人工复核", "复核原因",
        "人工修正分类", "最终ETF分类", "最终分类来源",
    ]:
        if col not in df.columns:
            df[col] = ""

    count = len(df)
    scale = round(float(df["基金规模_亿元"].sum()), 4)
    of_count = int(df["Wind代码"].astype(str).str.endswith(".OF").sum())
    missing_start = int(df["成立日期"].isna().sum() + df["成立日期"].astype(str).str.strip().eq("").sum())
    missing_list = int(df["上市日期"].isna().sum() + df["上市日期"].astype(str).str.strip().eq("").sum())
    bad_scale = int(df["基金规模_亿元"].isna().sum() + (df["基金规模_亿元"] <= 0).sum())
    fatal = []
    if count != EXPECTED_COUNT:
        fatal.append(f"行数={count}")
    if abs(scale - EXPECTED_SCALE) > 0.1:
        fatal.append(f"总规模={scale}")
    if of_count:
        fatal.append(f".OF数量={of_count}")
    if missing_start:
        fatal.append(f"成立日期缺失={missing_start}")
    if missing_list:
        fatal.append(f"上市日期缺失={missing_list}")
    if bad_scale:
        fatal.append(f"规模缺失或小于等于0={bad_scale}")
    if fatal:
        raise RuntimeError("质量检查失败，已停止输出：" + "; ".join(fatal))

    original_label = df["ETF标签"].copy()
    original_enhance = df["是否增强策略"].copy()
    original_strategy = df["是否策略ETF"].copy()
    original_asset = df["ETF资产类型"].copy()
    original_auto = df["ETF分类_自动"].copy()
    original_final = df["最终ETF分类"].copy()

    patched_codes: set[str] = set()

    enhance_rows = []
    for idx, row in df.iterrows():
        old_label = text(row["ETF标签"])
        old_enhance = text(row["是否增强策略"])
        old_strategy = text(row["是否策略ETF"])
        tags = split_tags(old_label)
        real = has_real_enhance(row)
        changed = False
        reason = ""
        if real:
            if text(row["是否增强策略"]) != "是":
                df.at[idx, "是否增强策略"] = "是"
                changed = True
            if "指数增强" not in tags:
                add_tag(tags, "指数增强")
                changed = True
            if text(row["最终ETF分类"]) == "宽基ETF":
                df.at[idx, "是否策略ETF"] = "是"
            reason = "明确命中增强关键词，保留/补充指数增强标签"
        else:
            if "指数增强" in tags or old_enhance == "是":
                tags = remove_tag(tags, "指数增强")
                df.at[idx, "是否增强策略"] = "否"
                if "宽基增强/指数增强" in text(row["分类依据"]):
                    df.at[idx, "分类依据"] = "纯宽基指数"
                if text(row["最终ETF分类"]) != "策略ETF" and text(row["是否SmartBeta策略"]) != "是":
                    df.at[idx, "是否策略ETF"] = "否"
                changed = True
                reason = "未命中明确增强关键词，删除误标指数增强"
        new_label = join_tags(tags) if tags else old_label
        df.at[idx, "ETF标签"] = new_label
        if changed:
            patched_codes.add(text(row["Wind代码"]))
            enhance_rows.append({
                "Wind代码": row["Wind代码"],
                "基金简称": row["基金简称"],
                "跟踪指数": row["跟踪指数"],
                "原ETF标签": old_label,
                "新ETF标签": new_label,
                "原是否增强策略": old_enhance,
                "新是否增强策略": df.at[idx, "是否增强策略"],
                "原是否策略ETF": old_strategy,
                "新是否策略ETF": df.at[idx, "是否策略ETF"],
                "修正原因": reason,
            })

    other_patch_rows = []
    other_mask = df["最终ETF分类"].astype(str).eq("其他ETF")
    for idx, row in df[other_mask].iterrows():
        ptext = priority_text(row)
        strategy_hits = match_keywords(ptext, STRATEGY_PATCH_KEYWORDS)
        theme_hits = match_keywords(ptext, THEME_PATCH_KEYWORDS)
        broad_hits = match_keywords(ptext, BROAD_PATCH_KEYWORDS)
        new_class = ""
        hits: list[str] = []
        reason = ""
        if strategy_hits:
            new_class = "策略ETF"
            hits = strategy_hits
            reason = "其他ETF命中策略补充关键词"
        elif theme_hits:
            new_class = "行业主题ETF"
            hits = theme_hits
            reason = "其他ETF命中行业主题补充关键词"
        elif broad_hits:
            new_class = "宽基ETF"
            hits = broad_hits
            reason = "其他ETF命中宽基/板块宽基补充关键词"
        if new_class:
            old_final = text(row["最终ETF分类"])
            old_label = text(row["ETF标签"])
            tags = split_tags(old_label)
            if new_class == "策略ETF":
                add_tag(tags, "Smart Beta")
                add_tag(tags, hits[0])
                df.at[idx, "是否SmartBeta策略"] = "是"
            elif new_class == "行业主题ETF":
                add_tag(tags, "行业主题")
                add_tag(tags, hits[0])
            elif new_class == "宽基ETF":
                add_tag(tags, "宽基")
            new_label = join_tags([t for t in tags if t != "待复核"])
            df.at[idx, "ETF资产类型"] = "股票ETF"
            df.at[idx, "ETF分类_自动"] = new_class
            df.at[idx, "最终ETF分类"] = new_class
            df.at[idx, "最终分类来源"] = "补丁修正"
            df.at[idx, "ETF标签"] = new_label
            df.at[idx, "分类置信度"] = "中"
            df.at[idx, "分类依据"] = reason
            df.at[idx, "命中关键词"] = "、".join(hits)
            df.at[idx, "是否需要人工复核"] = "否"
            df.at[idx, "复核原因"] = ""
            patched_codes.add(text(row["Wind代码"]))
            other_patch_rows.append({
                "Wind代码": row["Wind代码"],
                "基金简称": row["基金简称"],
                "跟踪指数": row["跟踪指数"],
                "原最终ETF分类": old_final,
                "新最终ETF分类": new_class,
                "原ETF标签": old_label,
                "新ETF标签": new_label,
                "命中关键词": "、".join(hits),
                "修正原因": reason,
                "是否仍需人工复核": "否",
            })

    shs_rows = []
    for idx, row in df.iterrows():
        ptext = priority_text(row)
        hits = match_keywords(ptext, SHS_KEYWORDS)
        if not hits:
            continue
        old_asset = text(row["ETF资产类型"])
        old_final = text(row["最终ETF分类"])
        old_label = text(row["ETF标签"])
        tags = split_tags(old_label)
        add_tag(tags, "跨市场")
        add_tag(tags, "沪港深")
        new_label = join_tags([t for t in tags if t != "待复核"])
        if old_asset == "跨境ETF" and old_final == "跨境ETF":
            reason = "已为跨境ETF，补充跨市场/沪港深标签"
        else:
            reason = "沪港深/沪深港/SHS跨市场指数，归入跨境ETF口径"
        df.at[idx, "ETF资产类型"] = "跨境ETF"
        df.at[idx, "ETF分类_自动"] = "跨境ETF"
        df.at[idx, "最终ETF分类"] = "跨境ETF"
        df.at[idx, "是否跨境ETF"] = "是"
        df.at[idx, "是否多资产ETF"] = "是"
        df.at[idx, "ETF标签"] = new_label
        df.at[idx, "分类依据"] = "沪港深/沪深港/SHS跨市场指数，归入跨境ETF口径。"
        df.at[idx, "分类置信度"] = "中"
        df.at[idx, "是否需要人工复核"] = "否"
        df.at[idx, "复核原因"] = ""
        df.at[idx, "最终分类来源"] = "补丁修正"
        patched_codes.add(text(row["Wind代码"]))
        shs_rows.append({
            "Wind代码": row["Wind代码"],
            "基金简称": row["基金简称"],
            "跟踪指数": row["跟踪指数"],
            "原ETF资产类型": old_asset,
            "新ETF资产类型": "跨境ETF",
            "原最终ETF分类": old_final,
            "新最终ETF分类": "跨境ETF",
            "原ETF标签": old_label,
            "新ETF标签": new_label,
            "修正原因": reason,
        })

    # Manual correction has priority if present.
    manual = df["人工修正分类"].fillna("").astype(str).str.strip()
    manual_mask = manual.ne("")
    df.loc[manual_mask, "最终ETF分类"] = manual[manual_mask]
    df.loc[manual_mask, "最终分类来源"] = "人工修正"
    df.loc[~manual_mask & df["Wind代码"].astype(str).isin(patched_codes), "最终分类来源"] = "补丁修正"
    df.loc[~manual_mask & ~df["Wind代码"].astype(str).isin(patched_codes) & df["最终分类来源"].fillna("").astype(str).str.strip().eq(""), "最终分类来源"] = "自动分类"

    set_class_flags(df)
    df.loc[df["最终ETF分类"].eq("策略ETF"), "是否SmartBeta策略"] = df.loc[df["最终ETF分类"].eq("策略ETF"), "是否SmartBeta策略"].replace("", "是")
    df.loc[df["最终ETF分类"].eq("策略ETF"), "是否策略ETF"] = "是"
    df.loc[df["最终ETF分类"].eq("其他ETF"), ["是否需要人工复核", "分类置信度"]] = ["是", "低"]

    empty_final = int(df["最终ETF分类"].fillna("").astype(str).str.strip().eq("").sum())
    empty_asset = int(df["ETF资产类型"].fillna("").astype(str).str.strip().eq("").sum())
    review_df = df[df["是否需要人工复核"].astype(str).eq("是")].copy()
    other_count = int(df["最终ETF分类"].eq("其他ETF").sum())

    ordinary_broad_wrong = 0
    for sample in COMMON_BROAD_SAMPLES:
        m = find_exact_fund(df, sample)
        if not m.empty:
            ordinary_broad_wrong += int(((m["最终ETF分类"] == "宽基ETF") & ((m["是否增强策略"] == "是") | m["ETF标签"].fillna("").astype(str).str.contains("指数增强"))).sum())

    obvious_theme_still_other = 0
    for sample in THEME_SAMPLE_NAMES:
        m = find_rows(df, sample)
        if not m.empty:
            obvious_theme_still_other += int(m["最终ETF分类"].eq("其他ETF").sum())
    obvious_broad_still_other = 0
    for sample in BROAD_SAMPLE_NAMES:
        m = find_rows(df, sample)
        if not m.empty:
            obvious_broad_still_other += int(m["最终ETF分类"].eq("其他ETF").sum())
    shs_unfixed = int(df[df.apply(lambda r: bool(match_keywords(priority_text(r), SHS_KEYWORDS)), axis=1)]["最终ETF分类"].ne("跨境ETF").sum())

    final_summary = summary_by(df, "最终ETF分类", "最终ETF分类")
    asset_summary = summary_by(df, "ETF资产类型", "ETF资产类型")
    stats_df = pd.concat([
        final_summary.assign(汇总维度="最终ETF分类"),
        asset_summary.assign(汇总维度="ETF资产类型"),
    ], ignore_index=True)
    stats_df = stats_df[["汇总维度", "最终ETF分类", "ETF资产类型", "ETF数量", "数量占比", "总规模_亿元", "规模占比", "平均规模_亿元", "规模中位数_亿元"]]

    tag_df = label_stats(df)
    index_stats = df.groupby(["最终ETF分类", "跟踪指数"], dropna=False).agg(
        关联ETF数量=("Wind代码", "count"),
        关联ETF规模_亿元=("基金规模_亿元", "sum"),
        代表ETF简称=("基金简称", lambda x: "；".join(x.astype(str).head(3))),
    ).reset_index()
    index_stats["关联ETF规模_亿元"] = index_stats["关联ETF规模_亿元"].round(4)
    index_stats = index_stats.sort_values(["最终ETF分类", "关联ETF规模_亿元"], ascending=[True, False])

    sample_cols = [c for c in df.columns]
    broad_sample = df[df["最终ETF分类"].eq("宽基ETF")].sort_values("基金规模_亿元", ascending=False).head(100)[sample_cols]
    strategy_sample = df[df["最终ETF分类"].eq("策略ETF")].sort_values("基金规模_亿元", ascending=False).head(100)[sample_cols]
    theme_sample = df[df["最终ETF分类"].eq("行业主题ETF")].sort_values("基金规模_亿元", ascending=False).head(100)[sample_cols]

    quality_rows = []
    def add_check(item: str, ok: bool, value, note: str) -> None:
        quality_rows.append({"检查项": item, "结果": "通过" if ok else "失败", "数值": value, "说明": note})
    add_check("主分析池数量是否仍为1574", len(df) == EXPECTED_COUNT, len(df), "必须为1574")
    add_check("总规模是否仍约为46949.9638亿元", abs(round(float(df["基金规模_亿元"].sum()), 4) - EXPECTED_SCALE) <= 0.1, round(float(df["基金规模_亿元"].sum()), 4), "误差不超过0.1亿元")
    add_check("是否存在.OF产品", of_count == 0, of_count, "必须为0")
    add_check("成立日期是否有缺失", missing_start == 0, missing_start, "必须为0")
    add_check("上市日期是否有缺失", missing_list == 0, missing_list, "必须为0")
    add_check("基金规模是否有缺失或小于等于0", bad_scale == 0, bad_scale, "必须为0")
    add_check("最终ETF分类是否有空值", empty_final == 0, empty_final, "必须为0")
    add_check("ETF资产类型是否有空值", empty_asset == 0, empty_asset, "必须为0")
    add_check("普通宽基ETF是否仍被误标指数增强", ordinary_broad_wrong == 0, ordinary_broad_wrong, "必须为0")
    add_check("是否仍有明显行业主题产品留在其他ETF", obvious_theme_still_other == 0, obvious_theme_still_other, "关键样本不应留在其他ETF")
    add_check("是否仍有明显宽基产品留在其他ETF", obvious_broad_still_other == 0, obvious_broad_still_other, "关键样本不应留在其他ETF")
    add_check("沪港深/SHS产品是否已统一处理", shs_unfixed == 0, shs_unfixed, "命中沪港深/SHS应归跨境ETF")
    add_check("待人工复核数量", True, len(review_df), "提示项")
    add_check("其他ETF数量", True, other_count, "提示项")
    for cat in ["宽基ETF", "行业主题ETF", "策略ETF", "跨境ETF", "债券ETF", "商品ETF", "货币ETF"]:
        g = df[df["最终ETF分类"].eq(cat)]
        add_check(f"{cat}数量和规模", True, f"{len(g)}只 / {round(float(g['基金规模_亿元'].sum()), 4)}亿元", "提示项")

    # Key samples as additional quality rows.
    for sample in COMMON_BROAD_SAMPLES:
        m = find_exact_fund(df, sample)
        wrong = int(((m["最终ETF分类"] == "宽基ETF") & ((m["是否增强策略"] == "是") | m["ETF标签"].fillna("").astype(str).str.contains("指数增强"))).sum()) if not m.empty else 0
        add_check(f"普通宽基样本不应增强：{sample}", wrong == 0, wrong if not m.empty else "未找到", "应为宽基ETF且不标记指数增强")
    for sample in REAL_ENHANCED_SAMPLES:
        m = find_rows(df, sample)
        ok = bool((not m.empty) and (((m["是否增强策略"] == "是") | m["ETF标签"].fillna("").astype(str).str.contains("指数增强")).any()))
        add_check(f"真实增强样本保留增强：{sample}", ok or m.empty, "未找到" if m.empty else ("通过" if ok else "未保留"), "若存在应保留增强标签")
    for sample in THEME_SAMPLE_NAMES:
        m = find_rows(df, sample)
        ok = bool(m.empty or not m["最终ETF分类"].eq("其他ETF").any())
        add_check(f"行业主题样本不留其他ETF：{sample}", ok, "未找到" if m.empty else int(m["最终ETF分类"].eq("其他ETF").sum()), "若存在不应留在其他ETF")
    for sample in BROAD_SAMPLE_NAMES:
        m = find_rows(df, sample)
        ok = bool(m.empty or not m["最终ETF分类"].eq("其他ETF").any())
        add_check(f"宽基样本不留其他ETF：{sample}", ok, "未找到" if m.empty else int(m["最终ETF分类"].eq("其他ETF").sum()), "若存在不应留在其他ETF")
    quality_df = pd.DataFrame(quality_rows)

    if (quality_df["结果"] == "失败").any():
        failed = "；".join(quality_df.loc[quality_df["结果"].eq("失败"), "检查项"].astype(str))
        raise RuntimeError("质量检查失败，已停止输出：" + failed)

    rule_rows = pd.DataFrame([
        ("本次输入文件", str(INPUT_PATH)),
        ("本次输出文件", str(OUTPUT_PATH)),
        ("本次是否改变主分析池", "否。仅基于现有ETF分类主表做补丁修正，不新增、不删除ETF。"),
        ("主分析池数量和规模", f"{len(df)}只，{round(float(df['基金规模_亿元'].sum()), 4)}亿元"),
        ("本次为什么只做补丁修正", "指数优先版产品池和主体分类已稳定，本次仅修正增强误标、其他ETF中明确样本、沪港深/SHS跨市场口径。"),
        ("增强标签修正逻辑", "仅当基金简称、证券简称、投资类型、跟踪指数、业绩比较基准、基金类型明确包含增强策略、指数增强、增强指数、增强型、增强指数型、被动指数增强、主动增强时保留增强标签；不使用基金全称判断增强。"),
        ("其他ETF补丁分类逻辑", "只对最终ETF分类=其他ETF的产品，按跟踪指数、业绩比较基准、基金简称、证券简称匹配策略、行业主题和宽基补充关键词；无法判断则保留其他ETF。"),
        ("沪港深/SHS跨市场口径", "命中沪港深、沪深港、SHS等跨市场关键词的产品统一归为跨境ETF，标签增加跨市场;沪港深，并标记是否多资产ETF=是。"),
        ("仍需人工复核事项", f"v2仍有{len(review_df)}只产品标记待人工复核，其他ETF数量为{other_count}只。"),
        ("后续如何使用最终ETF分类", "P17-P23后续统计应以ETF分类主表_v2中的最终ETF分类字段为准；若人工修正分类有值，应优先覆盖最终ETF分类。"),
    ], columns=["项目", "说明"])

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="ETF分类主表_v2", index=False)
        stats_df.to_excel(writer, sheet_name="分类统计_数量规模_v2", index=False)
        tag_df.to_excel(writer, sheet_name="标签统计_v2", index=False)
        index_stats.to_excel(writer, sheet_name="跟踪指数统计_v2", index=False)
        review_df.to_excel(writer, sheet_name="待人工复核ETF_v2", index=False)
        pd.DataFrame(enhance_rows).to_excel(writer, sheet_name="增强标签修正明细", index=False)
        pd.DataFrame(other_patch_rows).to_excel(writer, sheet_name="其他ETF补丁修正明细", index=False)
        pd.DataFrame(shs_rows).to_excel(writer, sheet_name="沪港深口径修正明细", index=False)
        quality_df.to_excel(writer, sheet_name="质量检查_v2", index=False)
        broad_sample.to_excel(writer, sheet_name="宽基ETF样本_v2", index=False)
        strategy_sample.to_excel(writer, sheet_name="策略ETF样本_v2", index=False)
        theme_sample.to_excel(writer, sheet_name="行业主题ETF样本_v2", index=False)
        rule_rows.to_excel(writer, sheet_name="分类规则说明_v2", index=False)

    style_workbook(OUTPUT_PATH)

    final_dist = df.groupby("最终ETF分类").agg(数量=("Wind代码", "count"), 规模=("基金规模_亿元", "sum")).sort_values("规模", ascending=False)
    asset_dist = df.groupby("ETF资产类型").agg(数量=("Wind代码", "count"), 规模=("基金规模_亿元", "sum")).sort_values("规模", ascending=False)
    print(f"输入文件路径：{INPUT_PATH}")
    print(f"输出文件路径：{OUTPUT_PATH}")
    print(f"主分析池数量：{len(df)}")
    print(f"ETF总规模_亿元：{round(float(df['基金规模_亿元'].sum()), 4)}")
    print("最终ETF分类分布：")
    for k, r in final_dist.iterrows():
        print(f"  {k}: {int(r['数量'])}只, {round(float(r['规模']), 4)}亿元")
    print("ETF资产类型分布：")
    for k, r in asset_dist.iterrows():
        print(f"  {k}: {int(r['数量'])}只, {round(float(r['规模']), 4)}亿元")
    print(f"增强标签修正数量：{len(enhance_rows)}")
    print(f"其他ETF补丁修正数量：{len(other_patch_rows)}")
    print(f"沪港深口径修正数量：{len(shs_rows)}")
    print(f"待人工复核数量：{len(review_df)}")
    print(f"其他ETF数量：{other_count}")
    print(f"普通宽基误标增强数量是否为0：{'是' if ordinary_broad_wrong == 0 else '否'}")
    print("质量检查是否通过：是")
    print("输出文件清单：")
    print(f"  {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
