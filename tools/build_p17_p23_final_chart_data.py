from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_PATH = Path(r"C:\Users\lvdon\Desktop\Fullgoal\ICI课题\ETF分类主表_指数优先版_v3\ETF分类主表_指数优先版_v3.xlsx")
INPUT_SHEET = "ETF分类主表_v3"
OUTPUT_DIR = Path(r"C:\Users\lvdon\Desktop\Fullgoal\ICI课题\P17-P23_作图数据_最终版")
OUTPUT_XLSX = OUTPUT_DIR / "ICI课题_P17-P23_作图数据_最终版.xlsx"
OUTPUT_MD = OUTPUT_DIR / "ICI课题_P17-P23_作图数据说明_最终版.md"

EXPECTED_COUNT = 1574
EXPECTED_SCALE = 46949.9638
ALLOWED_CLASSES = ["宽基ETF", "行业主题ETF", "策略ETF", "跨境ETF", "债券ETF", "商品ETF", "货币ETF"]
CLASS_ORDER = {name: i + 1 for i, name in enumerate(ALLOWED_CLASSES)}
ASSET_ORDER_LIST = ["股票ETF", "跨境ETF", "债券ETF", "商品ETF", "货币ETF"]
ASSET_ORDER = {name: i + 1 for i, name in enumerate(ASSET_ORDER_LIST)}
BUCKETS = [
    ("<1亿", 0, 1),
    ("1–2亿", 1, 2),
    ("2–10亿", 2, 10),
    ("10–50亿", 10, 50),
    ("50–200亿", 50, 200),
    ("200亿以上", 200, None),
]


def text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def fmt_date(value) -> str:
    if text(value) == "":
        return ""
    try:
        return pd.to_datetime(value).strftime("%Y-%m-%d")
    except Exception:
        return text(value)


def manager_short(name: str) -> str:
    s = text(name)
    replacements = [
        "基金管理股份有限公司",
        "基金管理有限公司",
        "资产管理股份有限公司",
        "资产管理有限公司",
        "基金有限公司",
        "管理有限公司",
        "有限公司",
    ]
    for suffix in replacements:
        s = s.replace(suffix, "")
    return s.strip()


def summary_by(df: pd.DataFrame, field: str, label: str, order_map: dict[str, int] | None = None) -> pd.DataFrame:
    total_count = len(df)
    total_scale = df["基金规模_亿元"].sum()
    rows = []
    for key, g in df.groupby(field, dropna=False):
        scale = g["基金规模_亿元"].sum()
        rows.append({
            label: key,
            "ETF数量": len(g),
            "数量占比": round(len(g) / total_count, 6) if total_count else 0,
            "总规模_亿元": round(scale, 4),
            "规模占比": round(scale / total_scale, 6) if total_scale else 0,
            "平均规模_亿元": round(g["基金规模_亿元"].mean(), 4) if len(g) else 0,
            "规模中位数_亿元": round(g["基金规模_亿元"].median(), 4) if len(g) else 0,
            "作图排序": order_map.get(key, 999) if order_map else 999,
        })
    return pd.DataFrame(rows).sort_values("作图排序")


def scale_bucket(value: float) -> str:
    v = float(value)
    for name, low, high in BUCKETS:
        if high is None and v >= low:
            return name
        if high is not None and low <= v < high:
            return name
    return "<1亿"


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="366092")
    font = Font(bold=True, color="FFFFFF")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col_cells[:200]:
                max_len = max(max_len, len(text(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 38)
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def add_metric(rows: list[dict], name: str, value, unit: str, note: str = "") -> None:
    rows.append({"指标": name, "数值": value, "单位": unit, "说明": note})


def cat_count_scale(df: pd.DataFrame, category: str) -> tuple[int, float]:
    g = df[df["最终ETF分类"].eq(category)]
    return len(g), round(float(g["基金规模_亿元"].sum()), 4)


def main() -> None:
    try:
        df = pd.read_excel(
            INPUT_PATH,
            sheet_name=INPUT_SHEET,
            engine="openpyxl",
            dtype={"Wind代码": str, "基金代码": str, "交易代码": str, "跟踪指数代码": str},
        )
    except ValueError as exc:
        raise RuntimeError("未找到 ETF分类主表_v3，请检查输入文件。") from exc

    for col in ["成立日期", "上市日期"]:
        if col in df.columns:
            df[col] = df[col].apply(fmt_date)
    df["成立日期_dt"] = pd.to_datetime(df["成立日期"], errors="coerce")
    df["上市日期_dt"] = pd.to_datetime(df["上市日期"], errors="coerce")
    df["基金规模_亿元"] = pd.to_numeric(df["基金规模_亿元"], errors="coerce")
    for col in ["管理费率", "托管费率"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        else:
            df[col] = pd.NA
    if "管理人简称" not in df.columns:
        df["管理人简称"] = ""
    df["管理人简称"] = df.apply(lambda r: text(r["管理人简称"]) or manager_short(r["基金管理人"]), axis=1)

    # Input validation.
    count = len(df)
    scale = round(float(df["基金规模_亿元"].sum()), 4)
    of_count = int(df["Wind代码"].astype(str).str.endswith(".OF").sum())
    missing_start = int(df["成立日期_dt"].isna().sum())
    missing_list = int(df["上市日期_dt"].isna().sum())
    bad_scale = int(df["基金规模_亿元"].isna().sum() + (df["基金规模_亿元"] <= 0).sum())
    empty_final = int(df["最终ETF分类"].fillna("").astype(str).str.strip().eq("").sum())
    categories = sorted(df["最终ETF分类"].dropna().astype(str).unique())
    invalid_categories = sorted(set(categories) - set(ALLOWED_CLASSES))
    fatal = []
    if count != EXPECTED_COUNT:
        fatal.append(f"主分析池数量={count}")
    if abs(scale - EXPECTED_SCALE) > 0.1:
        fatal.append(f"ETF总规模_亿元={scale}")
    if of_count:
        fatal.append(f".OF产品数量={of_count}")
    if missing_start:
        fatal.append(f"成立日期缺失数量={missing_start}")
    if missing_list:
        fatal.append(f"上市日期缺失数量={missing_list}")
    if bad_scale:
        fatal.append(f"规模缺失或小于等于0数量={bad_scale}")
    if empty_final:
        fatal.append(f"最终ETF分类空值数量={empty_final}")
    if invalid_categories:
        fatal.append("最终ETF分类存在非法类别=" + "、".join(invalid_categories))
    if fatal:
        raise RuntimeError("质量检查失败，已停止输出：" + "；".join(fatal))

    quality_rows = [
        {"检查项": "主分析池数量", "结果": "通过", "数值": count, "说明": "必须为1574"},
        {"检查项": "ETF总规模_亿元", "结果": "通过", "数值": scale, "说明": "与46949.9638亿元误差不超过0.1"},
        {"检查项": ".OF产品数量", "结果": "通过", "数值": of_count, "说明": "必须为0"},
        {"检查项": "成立日期缺失数量", "结果": "通过", "数值": missing_start, "说明": "必须为0"},
        {"检查项": "上市日期缺失数量", "结果": "通过", "数值": missing_list, "说明": "必须为0"},
        {"检查项": "规模缺失或小于等于0数量", "结果": "通过", "数值": bad_scale, "说明": "必须为0"},
        {"检查项": "最终ETF分类空值数量", "结果": "通过", "数值": empty_final, "说明": "必须为0"},
        {"检查项": "最终ETF分类类别", "结果": "通过", "数值": "、".join(categories), "说明": "仅包含验收后的7类"},
    ]
    for cat in ALLOWED_CLASSES:
        c, s = cat_count_scale(df, cat)
        quality_rows.append({"检查项": f"{cat}数量和规模", "结果": "通过", "数值": f"{c}只 / {s}亿元", "说明": "按最终ETF分类"})
    quality_df = pd.DataFrame(quality_rows)

    # P17.
    df["年份"] = df["成立日期_dt"].dt.year.astype(int)
    year_counts = df.groupby("年份").size().reset_index(name="当年新成立ETF数量").sort_values("年份")
    year_counts["累计ETF数量"] = year_counts["当年新成立ETF数量"].cumsum()
    year_counts["年度标记"] = year_counts["年份"].apply(lambda y: "YTD" if y == 2026 else "完整年度")
    p17_total = year_counts[["年份", "当年新成立ETF数量", "累计ETF数量", "年度标记"]]

    years = sorted(df["年份"].unique())
    grid = pd.MultiIndex.from_product([years, ALLOWED_CLASSES], names=["年份", "最终ETF分类"]).to_frame(index=False)
    ycat = df.groupby(["年份", "最终ETF分类"]).size().reset_index(name="当年新成立ETF数量")
    p17_by_cat = grid.merge(ycat, on=["年份", "最终ETF分类"], how="left").fillna({"当年新成立ETF数量": 0})
    p17_by_cat["当年新成立ETF数量"] = p17_by_cat["当年新成立ETF数量"].astype(int)
    p17_by_cat["累计ETF数量"] = p17_by_cat.groupby("最终ETF分类")["当年新成立ETF数量"].cumsum()
    p17_by_cat["年度标记"] = p17_by_cat["年份"].apply(lambda y: "YTD" if y == 2026 else "完整年度")
    p17_by_cat["作图排序"] = p17_by_cat["最终ETF分类"].map(CLASS_ORDER)
    p17_by_cat = p17_by_cat.sort_values(["年份", "作图排序"])[["年份", "最终ETF分类", "当年新成立ETF数量", "累计ETF数量", "年度标记"]]

    # P18.
    p18_class = summary_by(df, "最终ETF分类", "最终ETF分类", CLASS_ORDER)
    p18_asset = summary_by(df, "ETF资产类型", "ETF资产类型", ASSET_ORDER)
    stock_df = df[df["最终ETF分类"].isin(["宽基ETF", "行业主题ETF", "策略ETF"])].copy()
    stock_total_count = len(stock_df)
    stock_total_scale = stock_df["基金规模_亿元"].sum()
    stock_rows = []
    for cat in ["宽基ETF", "行业主题ETF", "策略ETF"]:
        g = stock_df[stock_df["最终ETF分类"].eq(cat)]
        scale_g = g["基金规模_亿元"].sum()
        stock_rows.append({
            "股票ETF内部分类": cat,
            "ETF数量": len(g),
            "数量占比_股票ETF内部": round(len(g) / stock_total_count, 6) if stock_total_count else 0,
            "总规模_亿元": round(scale_g, 4),
            "规模占比_股票ETF内部": round(scale_g / stock_total_scale, 6) if stock_total_scale else 0,
            "平均规模_亿元": round(g["基金规模_亿元"].mean(), 4) if len(g) else 0,
            "规模中位数_亿元": round(g["基金规模_亿元"].median(), 4) if len(g) else 0,
            "作图排序": CLASS_ORDER[cat],
        })
    p18_stock = pd.DataFrame(stock_rows)

    # P19.
    manager = df.groupby("基金管理人", dropna=False).agg(
        ETF数量=("Wind代码", "count"),
        ETF总规模_亿元=("基金规模_亿元", "sum"),
        平均单品规模_亿元=("基金规模_亿元", "mean"),
        规模中位数_亿元=("基金规模_亿元", "median"),
        最大单品规模_亿元=("基金规模_亿元", "max"),
        **{"10亿以下产品数量": ("基金规模_亿元", lambda x: int((x < 10).sum()))},
    ).reset_index()
    manager["ETF总规模_亿元"] = manager["ETF总规模_亿元"].round(4)
    manager["规模占比"] = (manager["ETF总规模_亿元"] / scale).round(6)
    manager["平均单品规模_亿元"] = manager["平均单品规模_亿元"].round(4)
    manager["规模中位数_亿元"] = manager["规模中位数_亿元"].round(4)
    manager["最大单品规模_亿元"] = manager["最大单品规模_亿元"].round(4)
    manager["10亿以下产品占比"] = (manager["10亿以下产品数量"] / manager["ETF数量"]).round(6)
    short_map = df.groupby("基金管理人")["管理人简称"].agg(lambda x: next((text(v) for v in x if text(v)), "")).to_dict()
    manager["管理人简称"] = manager["基金管理人"].map(short_map)
    manager["管理人简称"] = manager.apply(lambda r: text(r["管理人简称"]) or manager_short(r["基金管理人"]), axis=1)
    manager = manager.sort_values("ETF总规模_亿元", ascending=False).reset_index(drop=True)
    manager.insert(0, "排名", range(1, len(manager) + 1))
    p19_manager = manager[["排名", "基金管理人", "管理人简称", "ETF数量", "ETF总规模_亿元", "规模占比", "平均单品规模_亿元", "规模中位数_亿元", "最大单品规模_亿元", "10亿以下产品数量", "10亿以下产品占比"]]

    cr_rows = []
    for n in [3, 5, 10, 20]:
        top = p19_manager.head(n)
        cr_rows.append({
            "指标": f"CR{n}",
            "数量口径": round(top["ETF数量"].sum() / count, 6),
            "规模口径": round(top["ETF总规模_亿元"].sum() / scale, 6),
            "说明": f"前{n}大管理人ETF数量/规模占全市场比例",
        })
    p19_cr = pd.DataFrame(cr_rows)
    cr5_scale = float(p19_cr.loc[p19_cr["指标"].eq("CR5"), "规模口径"].iloc[0])
    cr10_scale = float(p19_cr.loc[p19_cr["指标"].eq("CR10"), "规模口径"].iloc[0])

    # P20.
    df["规模区间"] = df["基金规模_亿元"].apply(scale_bucket)
    bucket_order = {name: i + 1 for i, (name, _, _) in enumerate(BUCKETS)}
    bucket_rows = []
    for name, _, _ in BUCKETS:
        g = df[df["规模区间"].eq(name)]
        bucket_rows.append({
            "规模区间": name,
            "ETF数量": len(g),
            "数量占比": round(len(g) / count, 6),
            "总规模_亿元": round(g["基金规模_亿元"].sum(), 4),
            "规模占比": round(g["基金规模_亿元"].sum() / scale, 6),
            "作图排序": bucket_order[name],
        })
    p20_bucket = pd.DataFrame(bucket_rows)
    p20_by_class_rows = []
    for cat in ALLOWED_CLASSES:
        cg = df[df["最终ETF分类"].eq(cat)]
        c_count = len(cg)
        c_scale = cg["基金规模_亿元"].sum()
        for name, _, _ in BUCKETS:
            g = cg[cg["规模区间"].eq(name)]
            p20_by_class_rows.append({
                "最终ETF分类": cat,
                "规模区间": name,
                "ETF数量": len(g),
                "数量占比_分类内": round(len(g) / c_count, 6) if c_count else 0,
                "总规模_亿元": round(g["基金规模_亿元"].sum(), 4),
                "规模占比_分类内": round(g["基金规模_亿元"].sum() / c_scale, 6) if c_scale else 0,
                "作图排序": bucket_order[name],
            })
    p20_by_class = pd.DataFrame(p20_by_class_rows)
    small_mgr_rows = []
    for mgr, g in df.groupby("基金管理人", dropna=False):
        if len(g) < 5:
            continue
        small_mgr_rows.append({
            "基金管理人": mgr,
            "管理人简称": text(g["管理人简称"].iloc[0]) or manager_short(mgr),
            "ETF总数量": len(g),
            "ETF总规模_亿元": round(g["基金规模_亿元"].sum(), 4),
            "规模小于1亿数量": int((g["基金规模_亿元"] < 1).sum()),
            "规模小于2亿数量": int((g["基金规模_亿元"] < 2).sum()),
            "规模小于10亿数量": int((g["基金规模_亿元"] < 10).sum()),
            "规模小于1亿占比": round((g["基金规模_亿元"] < 1).sum() / len(g), 6),
            "规模小于2亿占比": round((g["基金规模_亿元"] < 2).sum() / len(g), 6),
            "规模小于10亿占比": round((g["基金规模_亿元"] < 10).sum() / len(g), 6),
        })
    p20_small_mgr = pd.DataFrame(small_mgr_rows).sort_values("ETF总规模_亿元", ascending=False)
    lt1_share = round((df["基金规模_亿元"] < 1).sum() / count, 6)
    lt2_share = round((df["基金规模_亿元"] < 2).sum() / count, 6)
    lt10_share = round((df["基金规模_亿元"] < 10).sum() / count, 6)

    # P21.
    metrics = []
    add_metric(metrics, "ETF总数量", count, "只", "ETF分类主表v3，全市场上市ETF口径")
    add_metric(metrics, "ETF总规模", scale, "亿元", "ETF分类主表v3，全市场上市ETF口径")
    for cat in ALLOWED_CLASSES:
        c, s = cat_count_scale(df, cat)
        add_metric(metrics, f"{cat}数量", c, "只", "按最终ETF分类")
        add_metric(metrics, f"{cat}规模", s, "亿元", "按最终ETF分类")
        add_metric(metrics, f"{cat}规模占比", round(s / scale, 6), "比例", "按最终ETF分类")
    stock_count = len(stock_df)
    stock_scale = round(float(stock_df["基金规模_亿元"].sum()), 4)
    multi_df = df[df["最终ETF分类"].isin(["跨境ETF", "债券ETF", "商品ETF", "货币ETF"])]
    multi_count = len(multi_df)
    multi_scale = round(float(multi_df["基金规模_亿元"].sum()), 4)
    add_metric(metrics, "股票ETF数量", stock_count, "只", "宽基ETF+行业主题ETF+策略ETF")
    add_metric(metrics, "股票ETF规模", stock_scale, "亿元", "宽基ETF+行业主题ETF+策略ETF")
    add_metric(metrics, "股票ETF规模占比", round(stock_scale / scale, 6), "比例", "宽基ETF+行业主题ETF+策略ETF")
    add_metric(metrics, "多资产ETF数量", multi_count, "只", "跨境ETF+债券ETF+商品ETF+货币ETF")
    add_metric(metrics, "多资产ETF规模", multi_scale, "亿元", "跨境ETF+债券ETF+商品ETF+货币ETF")
    add_metric(metrics, "多资产ETF规模占比", round(multi_scale / scale, 6), "比例", "跨境ETF+债券ETF+商品ETF+货币ETF")
    add_metric(metrics, "前五大管理人规模占比", cr5_scale, "比例", "P19_CR集中度 CR5规模口径")
    add_metric(metrics, "前十大管理人规模占比", cr10_scale, "比例", "P19_CR集中度 CR10规模口径")
    add_metric(metrics, "规模小于1亿产品占比", lt1_share, "比例", "按基金规模_亿元")
    add_metric(metrics, "规模小于2亿产品占比", lt2_share, "比例", "按基金规模_亿元")
    add_metric(metrics, "规模小于10亿产品占比", lt10_share, "比例", "按基金规模_亿元")
    add_metric(metrics, "平均管理费率", round(df["管理费率"].dropna().mean(), 6), "%", "简单平均")
    add_metric(metrics, "平均托管费率", round(df["托管费率"].dropna().mean(), 6), "%", "简单平均")
    p21_overview = pd.DataFrame(metrics)
    p21_card = pd.DataFrame([
        {"市场": "美国", "指标": "ETF数量", "数值": 4495, "单位": "只", "说明": "ICI 2026 Fact Book"},
        {"市场": "美国", "指标": "ETF总净资产", "数值": 13.4, "单位": "万亿美元", "说明": "ICI 2026 Fact Book"},
        {"市场": "中国", "指标": "ETF数量", "数值": count, "单位": "只", "说明": "ETF分类主表v3，全市场上市ETF口径"},
        {"市场": "中国", "指标": "ETF总规模", "数值": scale, "单位": "亿元", "说明": "ETF分类主表v3，全市场上市ETF口径"},
        {"市场": "中国", "指标": "前十大管理人规模占比", "数值": cr10_scale, "单位": "比例", "说明": "ETF分类主表v3"},
        {"市场": "中国", "指标": "规模小于10亿产品占比", "数值": lt10_share, "单位": "比例", "说明": "ETF分类主表v3"},
    ])

    # P22 / P23.
    theme_count, _ = cat_count_scale(df, "行业主题ETF")
    strategy_count, _ = cat_count_scale(df, "策略ETF")
    theme_strategy_share = round((theme_count + strategy_count) / count, 6)
    avg_mgmt_fee = round(df["管理费率"].dropna().mean(), 6)
    p22 = pd.DataFrame([
        {"转向": "发行导向 → 持营导向", "支撑指标": "规模小于10亿产品占比", "数值": lt10_share, "单位": "比例", "解释": "大量小规模 ETF 说明市场竞争重点不能停留在发行端，需要转向持续运营、流动性建设和客户使用场景。"},
        {"转向": "指数覆盖 → 场景覆盖", "支撑指标": "行业主题ETF + 策略ETF 数量占比", "数值": theme_strategy_share, "单位": "比例", "解释": "主题和策略产品数量较多，说明 ETF 供给已经从宽基覆盖延伸到行业、主题、策略等具体配置场景。"},
        {"转向": "费率竞争 → 生态竞争", "支撑指标": "平均管理费率", "数值": avg_mgmt_fee, "单位": "%", "解释": "费率只是基础竞争要素，长期竞争更依赖规模、流动性、渠道、投教和账户嵌入。"},
        {"转向": "单品销售 → 组合解决方案", "支撑指标": "多资产ETF规模占比", "数值": round(multi_scale / scale, 6), "单位": "比例", "解释": "跨境、债券、商品、货币等多资产 ETF 已具备一定规模，ETF 应从单只产品销售进入组合配置解决方案。"},
    ])
    broad_count, broad_scale = cat_count_scale(df, "宽基ETF")
    p23 = pd.DataFrame([
        {"指标": "ETF总数量", "数值": count, "单位": "只", "可用于哪一句结论": "用于说明中国 ETF 市场已经形成较完整供给基础。"},
        {"指标": "ETF总规模", "数值": scale, "单位": "亿元", "可用于哪一句结论": "用于说明 ETF 已经成为资管行业重要工具。"},
        {"指标": "宽基ETF规模", "数值": broad_scale, "单位": "亿元", "可用于哪一句结论": "用于说明宽基仍是资产配置和长期资金承接的重要底座。"},
        {"指标": "行业主题ETF数量", "数值": theme_count, "单位": "只", "可用于哪一句结论": "用于说明行业主题供给丰富，配置场景持续细分。"},
        {"指标": "策略ETF数量", "数值": strategy_count, "单位": "只", "可用于哪一句结论": "用于说明策略和Smart Beta产品正在形成补充工具箱。"},
        {"指标": "前十大管理人规模占比", "数值": cr10_scale, "单位": "比例", "可用于哪一句结论": "用于说明管理人格局呈现头部集中。"},
        {"指标": "规模小于10亿产品占比", "数值": lt10_share, "单位": "比例", "可用于哪一句结论": "用于说明产品长尾和持营压力。"},
        {"指标": "多资产ETF规模占比", "数值": round(multi_scale / scale, 6), "单位": "比例", "可用于哪一句结论": "用于说明 ETF 工具箱正从权益扩展到多资产配置。"},
    ])
    readme = pd.DataFrame([
        {"页码": "P17", "图表名称": "年度新发与累计数量", "对应sheet": "P17_年度新发累计", "建议图形": "柱状图+折线图", "关键强调数据": "近年新发提速，2026为YTD", "备注": "使用成立日期"},
        {"页码": "P18", "图表名称": "ETF分类结构", "对应sheet": "P18_分类结构_数量规模", "建议图形": "横向条形图/环形图", "关键强调数据": "宽基、行业主题、跨境、债券等结构", "备注": "使用最终ETF分类"},
        {"页码": "P18", "图表名称": "资产类型结构", "对应sheet": "P18_资产类型结构", "建议图形": "环形图/堆积条", "关键强调数据": "股票ETF与多资产ETF结构", "备注": "使用ETF资产类型"},
        {"页码": "P19", "图表名称": "管理人规模排名", "对应sheet": "P19_管理人规模排名", "建议图形": "柱状图", "关键强调数据": "前十大管理人规模排名", "备注": "按规模降序"},
        {"页码": "P19", "图表名称": "管理人集中度", "对应sheet": "P19_CR集中度", "建议图形": "指标卡/柱状图", "关键强调数据": "CR5/CR10/CR20", "备注": "强调头部集中"},
        {"页码": "P20", "图表名称": "规模区间分布", "对应sheet": "P20_规模区间分布", "建议图形": "柱状图+折线图", "关键强调数据": "10亿以下产品占比、200亿以上规模贡献", "备注": "展示长尾"},
        {"页码": "P21", "图表名称": "中美对比数据卡", "对应sheet": "P21_中美对比数据卡", "建议图形": "数据卡片+矩阵表", "关键强调数据": "美国4495只/13.4万亿美元，中国1574只/4.69万亿元", "备注": "不做汇率换算"},
        {"页码": "P22", "图表名称": "四个转向", "对应sheet": "P22_四个转向支撑数据", "建议图形": "四卡片", "关键强调数据": "持营、场景、生态、组合", "备注": "数据支撑结论"},
        {"页码": "P23", "图表名称": "最终升华", "对应sheet": "P23_最终升华引用数据", "建议图形": "金字塔+数据卡片", "关键强调数据": "ETF数量、规模、CR10、小规模占比", "备注": "结尾页引用"},
    ])

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    sheets = {
        "00_口径校验": quality_df,
        "P17_年度新发累计": p17_total,
        "P17_年度新发_按分类": p17_by_cat,
        "P18_分类结构_数量规模": p18_class,
        "P18_资产类型结构": p18_asset,
        "P18_股票ETF内部结构": p18_stock,
        "P19_管理人规模排名": p19_manager,
        "P19_CR集中度": p19_cr,
        "P20_规模区间分布": p20_bucket,
        "P20_规模区间_按分类": p20_by_class,
        "P20_小规模占比_管理人": p20_small_mgr,
        "P21_中国ETF总览指标": p21_overview,
        "P21_中美对比数据卡": p21_card,
        "P22_四个转向支撑数据": p22,
        "P23_最终升华引用数据": p23,
        "README_图表清单": readme,
    }
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        for sheet_name, sheet_df in sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
    style_workbook(OUTPUT_XLSX)

    class_lines = "\n".join(
        f"- {row['最终ETF分类']}：{int(row['ETF数量'])}只，{row['总规模_亿元']}亿元，规模占比{row['规模占比']:.4f}"
        for _, row in p18_class.iterrows()
    )
    md = f"""# ICI课题 P17-P23 作图数据说明（最终版）

## 本次任务目的

基于已验收的 `ETF分类主表_指数优先版_v3.xlsx`，生成 P17-P23 所需作图数据底表和汇报引用指标。

## 输入文件路径

{INPUT_PATH}

## 输出文件路径

{OUTPUT_XLSX}

## 产品池口径

中国 ETF 数据为全市场上市交易 ETF / 交易型开放式基金口径，共 {count} 只，总规模 {scale} 亿元；不包含 `.OF`、未上市、待上市、联接基金和关键交易信息缺失产品。

## 为什么使用 `最终ETF分类`

`最终ETF分类` 是已经过指数优先分类、补丁修正和人工映射验收后的分类字段，后续 P17-P23 统计均以该字段为准。

## 为什么不再重新分类

本次任务只做作图数据汇总。重新自动分类会引入口径漂移风险，因此不再改动产品池、基础字段或分类字段。

## P17-P23 数据对应关系

- P17：`P17_年度新发累计`、`P17_年度新发_按分类`
- P18：`P18_分类结构_数量规模`、`P18_资产类型结构`、`P18_股票ETF内部结构`
- P19：`P19_管理人规模排名`、`P19_CR集中度`
- P20：`P20_规模区间分布`、`P20_规模区间_按分类`、`P20_小规模占比_管理人`
- P21：`P21_中国ETF总览指标`、`P21_中美对比数据卡`
- P22：`P22_四个转向支撑数据`
- P23：`P23_最终升华引用数据`

## 关键统计结果

- ETF总数量：{count}只
- ETF总规模：{scale}亿元
- 前十大管理人规模占比：{cr10_scale}
- 规模小于10亿产品占比：{lt10_share}
- 多资产ETF规模占比：{round(multi_scale / scale, 6)}

### 最终ETF分类结构

{class_lines}

## 后续 PPT 作图建议

P17 使用柱状图+折线图展示年度新发和累计数量；P18 用条形图或环形图展示结构；P19 用排名柱状图和集中度指标卡；P20 用规模区间柱状图展示长尾；P21-P23 使用数据卡片支撑中美对比和启示结论。

## 注意事项

- 中国 ETF 数据为全市场上市交易 ETF 口径。
- 美国数据来自 ICI 2026 Fact Book。
- 中美规模币种不同，不直接换算。
- 2026 年为 YTD，不作为完整年度同比。
"""
    OUTPUT_MD.write_text(md, encoding="utf-8")

    print(f"输入文件路径：{INPUT_PATH}")
    print(f"输出文件夹路径：{OUTPUT_DIR}")
    print(f"输出总文件路径：{OUTPUT_XLSX}")
    print(f"主分析池数量：{count}")
    print(f"ETF总规模_亿元：{scale}")
    print("最终ETF分类分布：")
    for _, row in p18_class.iterrows():
        print(f"  {row['最终ETF分类']}: {int(row['ETF数量'])}只, {row['总规模_亿元']}亿元")
    print(f"P17年度范围：{min(years)}-{max(years)}")
    print("P18分类结构：")
    for _, row in p18_class.iterrows():
        print(f"  {row['最终ETF分类']}: 数量占比 {row['数量占比']}, 规模占比 {row['规模占比']}")
    print(f"P19 CR10规模口径：{cr10_scale}")
    print(f"P20 10亿以下产品占比：{lt10_share}")
    print(f"P21 中国ETF总数量和总规模：{count}只, {scale}亿元")
    print("生成的sheet清单：")
    for name in sheets:
        print(f"  {name}")
    print("质量检查是否通过：是")


if __name__ == "__main__":
    main()
