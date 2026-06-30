from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_PATH = Path(r"C:\Users\lvdon\Desktop\Fullgoal\ICI课题\ETF分类主表_指数优先版_v2\ETF分类主表_指数优先版_v2.xlsx")
INPUT_SHEET = "ETF分类主表_v2"
OUTPUT_DIR = Path(r"C:\Users\lvdon\Desktop\Fullgoal\ICI课题\ETF分类主表_指数优先版_v3")
OUTPUT_PATH = OUTPUT_DIR / "ETF分类主表_指数优先版_v3.xlsx"

EXPECTED_COUNT = 1574
EXPECTED_SCALE = 46949.9638

MAPPING = {
    "510010.SH": "策略ETF",
    "159623.SZ": "行业主题ETF",
    "159909.SZ": "行业主题ETF",
    "560650.SH": "策略ETF",
    "159301.SZ": "行业主题ETF",
    "159965.SZ": "策略ETF",
    "510160.SH": "行业主题ETF",
    "560190.SH": "行业主题ETF",
    "530530.SH": "宽基ETF",
    "562920.SH": "行业主题ETF",
    "510090.SH": "策略ETF",
    "561320.SH": "行业主题ETF",
    "510270.SH": "行业主题ETF",
    "561130.SH": "行业主题ETF",
    "515090.SH": "策略ETF",
    "515760.SH": "行业主题ETF",
    "563060.SH": "行业主题ETF",
    "159855.SZ": "行业主题ETF",
    "159973.SZ": "行业主题ETF",
    "512870.SH": "行业主题ETF",
    "159872.SZ": "行业主题ETF",
    "510810.SH": "行业主题ETF",
    "517180.SH": "行业主题ETF",
    "516620.SH": "行业主题ETF",
    "515200.SH": "行业主题ETF",
    "159912.SZ": "宽基ETF",
    "517090.SH": "行业主题ETF",
    "560620.SH": "行业主题ETF",
    "530100.SH": "宽基ETF",
    "159613.SZ": "行业主题ETF",
    "561500.SH": "策略ETF",
    "159578.SZ": "宽基ETF",
    "159743.SZ": "行业主题ETF",
    "159804.SZ": "宽基ETF",
    "562910.SH": "行业主题ETF",
    "562050.SH": "行业主题ETF",
    "159106.SZ": "行业主题ETF",
    "159666.SZ": "行业主题ETF",
    "510060.SH": "行业主题ETF",
    "510770.SH": "行业主题ETF",
    "159719.SZ": "行业主题ETF",
    "512770.SH": "行业主题ETF",
    "512190.SH": "行业主题ETF",
    "159159.SZ": "行业主题ETF",
    "159918.SZ": "宽基ETF",
    "510020.SH": "宽基ETF",
    "159638.SZ": "行业主题ETF",
    "516320.SH": "行业主题ETF",
}


def text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def fmt_date(value) -> str:
    if text(value) == "":
        return ""
    try:
        return pd.to_datetime(value).strftime("%Y-%m-%d")
    except Exception:
        return text(value)


def class_basis(category: str) -> str:
    if category == "宽基ETF":
        return "人工复核：跟踪指数属于宽基/板块宽基指数"
    if category == "行业主题ETF":
        return "人工复核：跟踪指数或基金简称属于行业/区域/主题指数"
    if category == "策略ETF":
        return "人工复核：跟踪指数属于策略/治理/ESG/质量类指数"
    return "人工复核：按人工映射表完成分类"


def class_tag(category: str) -> str:
    return {
        "宽基ETF": "宽基",
        "行业主题ETF": "行业主题",
        "策略ETF": "Smart Beta / 策略因子",
    }.get(category, "")


def update_flags(df: pd.DataFrame) -> None:
    df["是否宽基ETF"] = df["最终ETF分类"].apply(lambda x: "是" if text(x) == "宽基ETF" else "否")
    df["是否行业主题ETF"] = df["最终ETF分类"].apply(lambda x: "是" if text(x) == "行业主题ETF" else "否")
    df["是否策略ETF"] = df["最终ETF分类"].apply(lambda x: "是" if text(x) == "策略ETF" else "否")
    df["是否跨境ETF"] = df["ETF资产类型"].apply(lambda x: "是" if text(x) == "跨境ETF" else "否")
    df["是否债券ETF"] = df["ETF资产类型"].apply(lambda x: "是" if text(x) == "债券ETF" else "否")
    df["是否商品ETF"] = df["ETF资产类型"].apply(lambda x: "是" if text(x) == "商品ETF" else "否")
    df["是否货币ETF"] = df["ETF资产类型"].apply(lambda x: "是" if text(x) == "货币ETF" else "否")


def summary(df: pd.DataFrame, field: str, label: str) -> pd.DataFrame:
    total_count = len(df)
    total_scale = df["基金规模_亿元"].sum()
    rows = []
    for key, g in df.groupby(field, dropna=False):
        scale = g["基金规模_亿元"].sum()
        rows.append({
            label: key,
            "ETF数量": len(g),
            "数量占比": round(len(g) / total_count, 6) if total_count else 0,
            "总规模_亿元": round(scale, 4),
            "规模占比": round(scale / total_scale, 6) if total_scale else 0,
            "平均规模_亿元": round(g["基金规模_亿元"].mean(), 4) if len(g) else 0,
            "规模中位数_亿元": round(g["基金规模_亿元"].median(), 4) if len(g) else 0,
        })
    return pd.DataFrame(rows).sort_values("总规模_亿元", ascending=False)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="366092")
    font = Font(bold=True, color="FFFFFF")
    text_cols = {"Wind代码", "基金代码", "交易代码", "跟踪指数代码"}
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        headers = [text(c.value) for c in ws[1]]
        for idx, header in enumerate(headers, 1):
            if header in text_cols:
                for cell in ws[get_column_letter(idx)]:
                    cell.number_format = "@"
        for col_idx, cells in enumerate(ws.columns, 1):
            max_len = 0
            for cell in cells[:200]:
                max_len = max(max_len, len(text(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main() -> None:
    try:
        df = pd.read_excel(
            INPUT_PATH,
            sheet_name=INPUT_SHEET,
            engine="openpyxl",
            dtype={"Wind代码": str, "基金代码": str, "交易代码": str, "跟踪指数代码": str},
        )
    except ValueError as exc:
        raise RuntimeError("未找到 ETF分类主表_v2，请检查输入文件。") from exc

    for col in ["成立日期", "上市日期"]:
        if col in df.columns:
            df[col] = df[col].apply(fmt_date)
    df["基金规模_亿元"] = pd.to_numeric(df["基金规模_亿元"], errors="coerce")
    class_text_cols = [
        "ETF资产类型", "ETF分类_自动", "ETF标签", "是否增强策略", "是否SmartBeta策略", "是否宽基ETF",
        "是否行业主题ETF", "是否策略ETF", "是否跨境ETF", "是否债券ETF", "是否商品ETF", "是否货币ETF",
        "是否多资产ETF", "分类置信度", "分类依据", "命中关键词", "是否需要人工复核", "复核原因",
        "人工修正分类", "最终ETF分类", "最终分类来源",
    ]
    for col in class_text_cols:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].apply(text).astype("object")

    count = len(df)
    scale = round(float(df["基金规模_亿元"].sum()), 4)
    of_count = int(df["Wind代码"].astype(str).str.endswith(".OF").sum())
    missing_start = int(df["成立日期"].isna().sum() + df["成立日期"].astype(str).str.strip().eq("").sum())
    missing_list = int(df["上市日期"].isna().sum() + df["上市日期"].astype(str).str.strip().eq("").sum())
    bad_scale = int(df["基金规模_亿元"].isna().sum() + (df["基金规模_亿元"] <= 0).sum())
    fatal = []
    if count != EXPECTED_COUNT:
        fatal.append(f"行数={count}")
    if abs(scale - EXPECTED_SCALE) > 0.1:
        fatal.append(f"总规模={scale}")
    if of_count:
        fatal.append(f".OF数量={of_count}")
    if missing_start:
        fatal.append(f"成立日期缺失={missing_start}")
    if missing_list:
        fatal.append(f"上市日期缺失={missing_list}")
    if bad_scale:
        fatal.append(f"规模缺失或小于等于0={bad_scale}")
    if fatal:
        raise RuntimeError("质量检查失败，已停止输出：" + "；".join(fatal))

    before = df.set_index("Wind代码")[["ETF分类_自动", "最终ETF分类", "ETF资产类型"]].copy()
    hit_codes = set(df["Wind代码"].astype(str)).intersection(MAPPING.keys())
    missing_codes = sorted(set(MAPPING) - hit_codes)

    detail_rows = []
    for code, category in MAPPING.items():
        mask = df["Wind代码"].astype(str).eq(code)
        if not mask.any():
            continue
        idx = df.index[mask][0]
        original_final = text(df.at[idx, "最终ETF分类"])
        df.at[idx, "ETF资产类型"] = "股票ETF"
        df.at[idx, "最终ETF分类"] = category
        df.at[idx, "ETF分类_自动"] = category
        df.at[idx, "人工修正分类"] = category
        df.at[idx, "最终分类来源"] = "人工修正"
        df.at[idx, "是否需要人工复核"] = "否"
        df.at[idx, "复核原因"] = "已按人工映射表完成分类"
        df.at[idx, "分类置信度"] = "高"
        df.at[idx, "分类依据"] = class_basis(category)
        df.at[idx, "ETF标签"] = class_tag(category)
        df.at[idx, "是否增强策略"] = "否"
        df.at[idx, "是否SmartBeta策略"] = "是" if category == "策略ETF" else "否"
        df.at[idx, "是否多资产ETF"] = "否"
        detail_rows.append({
            "Wind代码": code,
            "基金简称": df.at[idx, "基金简称"],
            "跟踪指数": df.at[idx, "跟踪指数"],
            "原最终ETF分类": original_final,
            "新最终ETF分类": category,
            "人工修正分类": category,
            "基金规模_亿元": df.at[idx, "基金规模_亿元"],
            "修正原因": class_basis(category),
        })

    update_flags(df)

    # Check non-mapped products were not reclassified.
    after = df.set_index("Wind代码")[["ETF分类_自动", "最终ETF分类", "ETF资产类型"]].copy()
    unchanged_codes = sorted(set(before.index) - set(MAPPING))
    non_mapped_changed = int((before.loc[unchanged_codes] != after.loc[unchanged_codes]).any(axis=1).sum()) if unchanged_codes else 0

    mapped_df = df[df["Wind代码"].astype(str).isin(MAPPING.keys())].copy()
    mapped_not_other = int(mapped_df["最终ETF分类"].ne("其他ETF").sum())
    mapped_all_hit = len(hit_codes) == len(MAPPING)
    mapped_all_not_other = mapped_not_other == len(MAPPING)
    other_count = int(df["最终ETF分类"].eq("其他ETF").sum())
    empty_final = int(df["最终ETF分类"].fillna("").astype(str).str.strip().eq("").sum())
    empty_asset = int(df["ETF资产类型"].fillna("").astype(str).str.strip().eq("").sum())
    review_df = df[df["是否需要人工复核"].astype(str).eq("是")].copy()

    class_stats = summary(df, "最终ETF分类", "最终ETF分类")
    asset_stats = summary(df, "ETF资产类型", "ETF资产类型")
    broad_sample = df[df["最终ETF分类"].eq("宽基ETF")].sort_values("基金规模_亿元", ascending=False).head(100)
    strategy_sample = df[df["最终ETF分类"].eq("策略ETF")].sort_values("基金规模_亿元", ascending=False).head(100)
    theme_sample = df[df["最终ETF分类"].eq("行业主题ETF")].sort_values("基金规模_亿元", ascending=False).head(100)

    quality_rows = []
    def add_check(item: str, ok: bool, value, note: str) -> None:
        quality_rows.append({"检查项": item, "结果": "通过" if ok else "失败", "数值": value, "说明": note})

    add_check("主分析池数量是否仍为1574", len(df) == EXPECTED_COUNT, len(df), "必须为1574")
    add_check("总规模是否仍约为46949.9638亿元", abs(round(float(df["基金规模_亿元"].sum()), 4) - EXPECTED_SCALE) <= 0.1, round(float(df["基金规模_亿元"].sum()), 4), "误差不超过0.1亿元")
    add_check("是否存在.OF产品", of_count == 0, of_count, "必须为0")
    add_check("成立日期是否有缺失", missing_start == 0, missing_start, "必须为0")
    add_check("上市日期是否有缺失", missing_list == 0, missing_list, "必须为0")
    add_check("基金规模是否有缺失或小于等于0", bad_scale == 0, bad_scale, "必须为0")
    add_check("最终ETF分类是否有空值", empty_final == 0, empty_final, "必须为0")
    add_check("ETF资产类型是否有空值", empty_asset == 0, empty_asset, "必须为0")
    add_check("其他ETF数量", other_count == 0, other_count, "必须为0")
    add_check("待人工复核数量", True, len(review_df), "提示项")
    add_check("48只人工映射是否全部命中", mapped_all_hit, len(hit_codes), f"缺失：{'、'.join(missing_codes) if missing_codes else '无'}")
    add_check("48只人工映射是否全部不再属于其他ETF", mapped_all_not_other, mapped_not_other, "必须全部不为其他ETF")
    add_check("未列入映射表产品分类是否保持不变", non_mapped_changed == 0, non_mapped_changed, "必须为0")
    for category in ["宽基ETF", "行业主题ETF", "策略ETF", "跨境ETF", "债券ETF", "商品ETF", "货币ETF"]:
        g = df[df["最终ETF分类"].eq(category)]
        add_check(f"{category}数量和规模", True, f"{len(g)}只 / {round(float(g['基金规模_亿元'].sum()), 4)}亿元", "提示项")
    quality_df = pd.DataFrame(quality_rows)

    if (quality_df["结果"] == "失败").any():
        failed = "；".join(quality_df.loc[quality_df["结果"].eq("失败"), "检查项"].astype(str))
        raise RuntimeError("质量检查失败，已停止输出：" + failed)

    rule_df = pd.DataFrame([
        ("本次输入文件", str(INPUT_PATH)),
        ("本次输出文件", str(OUTPUT_PATH)),
        ("本次是否改变主分析池", "否"),
        ("本次是否新增或删除产品", "否"),
        ("本次修正对象", "v2中剩余48只其他ETF"),
        ("本次修正方式", "按Wind代码人工映射覆盖最终ETF分类、ETF分类_自动、人工修正分类、标签和布尔字段"),
        ("后续P17-P23汇总字段", "应使用最终ETF分类字段"),
        ("后续分类原则", "不应再重新自动分类；如有变更，应继续使用人工修正分类字段覆盖"),
    ], columns=["项目", "说明"])

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="ETF分类主表_v3", index=False)
        pd.DataFrame(detail_rows).to_excel(writer, sheet_name="人工修正明细_48只", index=False)
        class_stats.to_excel(writer, sheet_name="分类统计_数量规模_v3", index=False)
        asset_stats.to_excel(writer, sheet_name="资产类型统计_v3", index=False)
        review_df.to_excel(writer, sheet_name="待人工复核ETF_v3", index=False)
        quality_df.to_excel(writer, sheet_name="质量检查_v3", index=False)
        broad_sample.to_excel(writer, sheet_name="宽基ETF样本_v3", index=False)
        strategy_sample.to_excel(writer, sheet_name="策略ETF样本_v3", index=False)
        theme_sample.to_excel(writer, sheet_name="行业主题ETF样本_v3", index=False)
        rule_df.to_excel(writer, sheet_name="分类规则说明_v3", index=False)
    style_workbook(OUTPUT_PATH)

    final_dist = df.groupby("最终ETF分类").agg(数量=("Wind代码", "count"), 规模=("基金规模_亿元", "sum")).sort_values("规模", ascending=False)
    print(f"输入文件路径：{INPUT_PATH}")
    print(f"输出文件路径：{OUTPUT_PATH}")
    print(f"主分析池数量：{len(df)}")
    print(f"ETF总规模_亿元：{round(float(df['基金规模_亿元'].sum()), 4)}")
    print(f"人工映射命中数量：{len(hit_codes)}")
    print(f"人工修正后其他ETF数量：{other_count}")
    print(f"待人工复核数量：{len(review_df)}")
    print("最终ETF分类分布：")
    for category, row in final_dist.iterrows():
        print(f"  {category}: {int(row['数量'])}只, {round(float(row['规模']), 4)}亿元")
    print("质量检查是否通过：是")
    print("输出文件清单：")
    print(f"  {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
